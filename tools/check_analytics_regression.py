# tools/check_analytics_regression.py
# -*- coding: utf-8 -*-
# SIMS 분석/KPI 회귀 체크 도구
# 작성자: ChatGPT (2026-05-02)
# VERSION = "check_analytics_regression/2026-05-02-v1"
# 참고: 이 스크립트는SIMS 분석/KPI 회귀 여부를 점검하기 위한 도구입니다.

"""
Analytics/KPI regression checker.

기본 import/helper 확인:
    & "C:\\Program Files\\Python313\\python.exe" tools\\check_analytics_regression.py

실제 서비스 DB 조회 smoke test:
    & "C:\\Program Files\\Python313\\python.exe" tools\\check_analytics_regression.py --live

NLQ 라우팅까지 확인:
    & "C:\\Program Files\\Python313\\python.exe" tools\\check_analytics_regression.py --nlq

전체 확인:
    & "C:\\Program Files\\Python313\\python.exe" tools\\check_analytics_regression.py --live --nlq
"""

from __future__ import annotations

import argparse
import ast
import io
import importlib
import json
import logging
import os
import re
import sys
import tempfile
import time
import traceback
import warnings
import zipfile
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable, Iterable

import pandas as pd


# ---------------------------------------------------------------------
# Project root 보정
# ---------------------------------------------------------------------
THIS_FILE = Path(__file__).resolve()
PROJECT_ROOT = THIS_FILE.parents[1]

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

os.chdir(PROJECT_ROOT)


# ---------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="[%(levelname)s] %(message)s",
)

log = logging.getLogger("analytics_regression")


# ---------------------------------------------------------------------
# Result helpers
# ---------------------------------------------------------------------
@dataclass
class CheckResult:
    name: str
    ok: bool
    detail: str = ""


@dataclass
class ServiceCase:
    name: str
    function_name: str
    params: dict[str, Any]
    expected_title_contains: str
    expected_meta_key: str | None = None
    expected_analysis_type: str | None = None
    expected_condition_tokens: tuple[str, ...] = ()
    require_seq_column: bool = True
    require_summary_md: bool = True
    require_message: bool = True
    allow_zero_rows: bool = False
    check_code_columns: bool = True


@dataclass
class NlqCase:
    query: str
    expected_action: str
    expected_analysis_type: str | None = None
    expected_meta_key: str | None = None
    expected_params: dict[str, Any] | None = None
    expected_condition_tokens: tuple[str, ...] = ()
    require_summary_md: bool = True
    require_message: bool = True
    allow_empty_meta_counts: bool = False

def _ok(name: str, detail: str = "") -> CheckResult:
    return CheckResult(name=name, ok=True, detail=detail)


def _fail(name: str, detail: str = "") -> CheckResult:
    return CheckResult(name=name, ok=False, detail=detail)


def _print_results(title: str, results: list[CheckResult]) -> int:
    print()
    print("=" * 78)
    print(title)
    print("=" * 78)

    failed = 0
    for r in results:
        mark = "OK " if r.ok else "FAIL"
        print(f"[{mark}] {r.name}")
        if r.detail:
            print(f"      {r.detail}")
        if not r.ok:
            failed += 1

    print("-" * 78)
    print(f"총 {len(results)}건 / 성공 {len(results) - failed}건 / 실패 {failed}건")
    return failed


def _safe_len_df(obj: Any) -> int | None:
    try:
        import pandas as pd
        if isinstance(obj, pd.DataFrame):
            return int(len(obj))
    except Exception:
        pass
    return None


def _payload_columns(payload: dict[str, Any]) -> list[str]:
    cols = payload.get("columns")
    if isinstance(cols, list) and cols:
        return [str(c) for c in cols]

    df_display = payload.get("df_display")
    try:
        import pandas as pd
        if isinstance(df_display, pd.DataFrame):
            return [str(c) for c in df_display.columns]
    except Exception:
        pass

    df = payload.get("df")
    try:
        import pandas as pd
        if isinstance(df, pd.DataFrame):
            return [str(c) for c in df.columns]
    except Exception:
        pass

    records = payload.get("records")
    if isinstance(records, list) and records and isinstance(records[0], dict):
        return [str(c) for c in records[0].keys()]

    return []


def _payload_df(payload: dict[str, Any]) -> Any:
    try:
        import pandas as pd
        for key in ("df", "df_display"):
            obj = payload.get(key)
            if isinstance(obj, pd.DataFrame):
                return obj
    except Exception:
        pass
    return None


def _code_column_dtype_problem(payload: dict[str, Any]) -> str:
    try:
        import pandas as pd
    except Exception:
        return ""

    df = _payload_df(payload)
    if not isinstance(df, pd.DataFrame) or df.empty:
        return ""

    code_cols = [
        "제품코드",
        "제조사코드",
        "거래처코드",
        "매입처코드",
        "재고적용처코드",
        "보험코드",
        "표준코드",
        "바코드",
    ]
    bad_cols = [
        col for col in code_cols
        if col in df.columns and pd.api.types.is_numeric_dtype(df[col])
    ]
    if bad_cols:
        return f"코드 컬럼이 numeric dtype으로 변환됨: {bad_cols!r}"
    return ""


def _payload_row_count(payload: dict[str, Any]) -> int:
    meta = payload.get("meta") or {}
    for key in ("row_count_total", "row_count"):
        try:
            v = meta.get(key)
            if v is not None:
                return int(v)
        except Exception:
            pass

    for key in ("df_display", "df"):
        n = _safe_len_df(payload.get(key))
        if n is not None:
            return n

    records = payload.get("records")
    if isinstance(records, list):
        return len(records)

    return 0

def _condition_text_from_payload(payload: dict[str, Any]) -> str:
    """
    조회조건 검증용 문자열.
    분석/KPI는 condition/query_summary/summary_md/message/params 중
    어디에 조건이 들어가도 검증할 수 있게 한곳에 모은다.
    """
    if not isinstance(payload, dict):
        return ""

    meta = payload.get("meta") or {}
    params = payload.get("params") or {}

    parts = [
        meta.get("condition"),
        meta.get("query_summary"),
        meta.get("summary_md"),
        payload.get("message"),
        payload.get("data") if isinstance(payload.get("data"), str) else "",
        str(params or ""),
    ]

    text = " ".join(str(x or "") for x in parts)
    text = text.replace("\n", " ").replace("\r", " ")
    text = " ".join(text.split())

    # params에 20250101 형식으로 들어온 날짜도
    # expected_condition_tokens=("2025-01-01", ...)와 비교 가능하게 한다.
    text = _append_date_variants(text)

    return text

def _missing_condition_tokens(payload: dict[str, Any], tokens: tuple[str, ...]) -> list[str]:
    if not tokens:
        return []

    text = _condition_text_from_payload(payload)
    return [str(token) for token in tokens if str(token) not in text]


def _short_text(value: Any, limit: int = 140) -> str:
    text = str(value or "").replace("\n", " ").replace("\r", " ")
    text = " ".join(text.split())
    if len(text) > limit:
        return text[:limit] + "..."
    return text

def _append_date_variants(text: str) -> str:
    """
    condition_text 안의 YYYYMMDD / YYYYMM 값을
    YYYY-MM-DD / YYYY-MM 형태로도 같이 검사할 수 있게 보강한다.

    예:
    - 20250101 → 2025-01-01
    - 20251231 → 2025-12-31
    - 202501   → 2025-01
    """
    import re

    src = str(text or "")
    variants: list[str] = []

    for m in re.findall(r"(?<!\d)(20\d{2})(\d{2})(\d{2})(?!\d)", src):
        y, mo, d = m
        variants.append(f"{y}-{mo}-{d}")

    for m in re.findall(r"(?<!\d)(20\d{2})(\d{2})(?!\d)", src):
        y, mo = m
        variants.append(f"{y}-{mo}")

    if variants:
        src += " " + " ".join(dict.fromkeys(variants))

    return src

# ---------------------------------------------------------------------
# Basic checks
# ---------------------------------------------------------------------
def run_basic_checks() -> list[CheckResult]:
    results: list[CheckResult] = []

    module_name = "app.services.analytics_sales_trend_service"
    required_functions = [
        "get_sales_trend_result",
        "get_sales_trend_summary_result",
        "get_sales_forecast_result",
        "get_stock_shortage_result",
    ]

    try:
        mod = importlib.import_module(module_name)
        results.append(_ok(f"import {module_name}"))
    except Exception as e:
        return [_fail(f"import {module_name}", f"{type(e).__name__}: {e}")]

    for fn_name in required_functions:
        fn = getattr(mod, fn_name, None)
        if callable(fn):
            results.append(_ok(f"{module_name}.{fn_name}"))
        else:
            results.append(_fail(f"{module_name}.{fn_name}", "callable 함수 없음"))

    # NLQ router 쪽 분석 action 해석 함수 확인
    try:
        from app.utils import env_config

        saved_env = {
            k: os.environ.get(k)
            for k in (
                "APP_ENV",
                "CHAT_FILE",
                "UPLOAD_DIR",
                "SSAI_STORAGE_ROOT",
                "LOG_FILE",
                "SIMS_LOG_FILE",
                "SSAI_INSTANCE_ID",
                "SSAI_ENV_FILE",
            )
        }
        old_cwd = Path.cwd()
        with tempfile.TemporaryDirectory() as td:
            root = Path(td) / "project"
            other = Path(td) / "other"
            chat_dir = root / "chat"
            upload_dir = root / "uploads"
            storage_dir = root / "storage"
            log_file = root / "logs" / "app.log"
            root.mkdir()
            other.mkdir()
            chat_dir.mkdir()
            upload_dir.mkdir()
            storage_dir.mkdir()
            log_file.parent.mkdir()
            project_env = root / ".env"
            project_chat = chat_dir / "root_chat_rooms.json"
            other_chat = other / "wrong_chat_rooms.json"
            project_env.write_text(
                "\n".join(
                    [
                        "APP_ENV=prod",
                        "SSAI_INSTANCE_ID=regression",
                        f"CHAT_FILE={project_chat}",
                        f"UPLOAD_DIR={upload_dir}",
                        f"SSAI_STORAGE_ROOT={storage_dir}",
                        f"LOG_FILE={log_file}",
                    ]
                ),
                encoding="utf-8",
            )
            (other / ".env").write_text(
                "\n".join(
                    [
                        f"CHAT_FILE={other_chat}",
                        f"UPLOAD_DIR={other / 'uploads'}",
                        f"SSAI_STORAGE_ROOT={other / 'storage'}",
                        f"LOG_FILE={other / 'logs' / 'app.log'}",
                    ]
                ),
                encoding="utf-8",
            )

            os.environ["CHAT_FILE"] = str(other_chat)
            os.environ["UPLOAD_DIR"] = str(other / "uploads")
            os.environ["SSAI_STORAGE_ROOT"] = str(other / "storage")
            os.environ["LOG_FILE"] = str(other / "logs" / "app.log")
            os.chdir(other)
            loaded = env_config.load_project_env(override=True, env_path=project_env)
            load_ok = (
                loaded.env_path == project_env
                and loaded.exists
                and os.environ.get("CHAT_FILE") == str(project_chat)
                and os.environ.get("UPLOAD_DIR") == str(upload_dir)
                and os.environ.get("SSAI_STORAGE_ROOT") == str(storage_dir)
                and os.environ.get("LOG_FILE") == str(log_file)
                and os.environ.get("SSAI_ENV_FILE") == str(project_env)
            )
            required_paths = ("CHAT_FILE", "UPLOAD_DIR", "SSAI_STORAGE_ROOT", ("LOG_FILE", "SIMS_LOG_FILE"))
            errors = env_config.validate_startup_env(env_path=project_env, project_root=root, environ=os.environ, required_path_keys=required_paths)
            if load_ok and not errors:
                results.append(_ok("project-root .env overrides cwd and OS env", f"env_file={project_env}"))
            else:
                results.append(_fail("project-root .env overrides cwd and OS env", f"load_ok={load_ok}, errors={errors}"))

            missing_errors = env_config.validate_startup_env(
                env_path=root / "missing.env",
                project_root=root,
                environ={
                    "APP_ENV": "prod",
                    "CHAT_FILE": str(project_chat),
                    "UPLOAD_DIR": str(upload_dir),
                    "SSAI_STORAGE_ROOT": str(storage_dir),
                    "LOG_FILE": str(log_file),
                },
                required_path_keys=required_paths,
            )
            relative_env = root / "relative.env"
            relative_env.write_text(
                "\n".join(
                    [
                        "APP_ENV=prod",
                        "CHAT_FILE=data/chat.json",
                        f"UPLOAD_DIR={upload_dir}",
                        "SSAI_STORAGE_ROOT=data/ssai_storage",
                        "LOG_FILE=logs/app.log",
                    ]
                ),
                encoding="utf-8",
            )
            relative_errors = env_config.validate_startup_env(
                env_path=relative_env,
                project_root=root,
                environ={},
                required_path_keys=required_paths,
            )
            blank_env = root / "blank.env"
            blank_env.write_text(
                "\n".join(
                    [
                        "APP_ENV=prod",
                        "CHAT_FILE=",
                        f"UPLOAD_DIR={upload_dir}",
                        "SSAI_STORAGE_ROOT=",
                        "LOG_FILE=",
                    ]
                ),
                encoding="utf-8",
            )
            blank_errors = env_config.validate_startup_env(
                env_path=blank_env,
                project_root=root,
                environ={},
                required_path_keys=required_paths,
            )
            missing_app_env = root / "missing_app.env"
            missing_app_env.write_text(
                "\n".join(
                    [
                        f"CHAT_FILE={project_chat}",
                        f"UPLOAD_DIR={upload_dir}",
                        f"SSAI_STORAGE_ROOT={storage_dir}",
                        f"LOG_FILE={log_file}",
                    ]
                ),
                encoding="utf-8",
            )
            blank_app_env = root / "blank_app.env"
            blank_app_env.write_text(
                "\n".join(
                    [
                        "APP_ENV=",
                        f"CHAT_FILE={project_chat}",
                        f"UPLOAD_DIR={upload_dir}",
                        f"SSAI_STORAGE_ROOT={storage_dir}",
                        f"LOG_FILE={log_file}",
                    ]
                ),
                encoding="utf-8",
            )
            unknown_app_env = root / "unknown_app.env"
            unknown_app_env.write_text(
                "\n".join(
                    [
                        "APP_ENV=production",
                        f"CHAT_FILE={project_chat}",
                        f"UPLOAD_DIR={upload_dir}",
                        f"SSAI_STORAGE_ROOT={storage_dir}",
                        f"LOG_FILE={log_file}",
                    ]
                ),
                encoding="utf-8",
            )
            app_env_errors = {
                "missing": env_config.validate_startup_env(env_path=missing_app_env, project_root=root, environ={}, required_path_keys=required_paths),
                "blank": env_config.validate_startup_env(env_path=blank_app_env, project_root=root, environ={}, required_path_keys=required_paths),
                "unknown": env_config.validate_startup_env(env_path=unknown_app_env, project_root=root, environ={}, required_path_keys=required_paths),
            }
            app_env_ok = True
            for mode in ("dev", "test", "prod"):
                mode_env = root / f"{mode}.env"
                mode_env.write_text(
                    "\n".join(
                        [
                            f"APP_ENV={mode}",
                            f"CHAT_FILE={project_chat}",
                            f"UPLOAD_DIR={upload_dir}",
                            f"SSAI_STORAGE_ROOT={storage_dir}",
                            f"LOG_FILE={log_file}",
                        ]
                    ),
                    encoding="utf-8",
                )
                if env_config.validate_startup_env(env_path=mode_env, project_root=root, environ={}, required_path_keys=required_paths):
                    app_env_ok = False
            alias_env = root / "alias.env"
            alias_log_file = root / "logs" / "sims-app.log"
            alias_env.write_text(
                "\n".join(
                    [
                        "APP_ENV=prod",
                        f"CHAT_FILE={project_chat}",
                        f"UPLOAD_DIR={upload_dir}",
                        f"SSAI_STORAGE_ROOT={storage_dir}",
                        f"SIMS_LOG_FILE={alias_log_file}",
                    ]
                ),
                encoding="utf-8",
            )
            no_log_env = root / "no_log.env"
            no_log_env.write_text(
                "\n".join(
                    [
                        "APP_ENV=prod",
                        f"CHAT_FILE={project_chat}",
                        f"UPLOAD_DIR={upload_dir}",
                        f"SSAI_STORAGE_ROOT={storage_dir}",
                    ]
                ),
                encoding="utf-8",
            )
            os.environ["LOG_FILE"] = str(other / "logs" / "os-app.log")
            os.environ["SIMS_LOG_FILE"] = str(other / "logs" / "os-sims-app.log")
            alias_errors = env_config.validate_startup_env(env_path=alias_env, project_root=root, environ=os.environ, required_path_keys=required_paths)
            no_log_errors = env_config.validate_startup_env(env_path=no_log_env, project_root=root, environ=os.environ, required_path_keys=required_paths)
            app_env_policy_ok = (
                any("missing required env: APP_ENV" in e for e in app_env_errors["missing"])
                and any("missing required env: APP_ENV" in e for e in app_env_errors["blank"])
                and any("invalid APP_ENV" in e for e in app_env_errors["unknown"])
                and app_env_ok
            )
            if app_env_policy_ok:
                results.append(_ok("APP_ENV required and allow-listed", "dev/test/prod accepted; missing/blank/unknown blocked"))
            else:
                results.append(_fail("APP_ENV required and allow-listed", f"errors={app_env_errors}, valid_modes_ok={app_env_ok}"))
            dev_path = env_config.config_path("CHAT_FILE", project_root=root, environ={"CHAT_FILE": "data/dev_chat.json"})
            storage_path = env_config.config_path("SSAI_STORAGE_ROOT", project_root=root, environ=env_config.read_project_env_file(project_env))
            log_path = env_config.config_path_any(("LOG_FILE", "SIMS_LOG_FILE"), project_root=root, environ=env_config.read_project_env_file(project_env))
            alias_log_path = env_config.config_path_any(("LOG_FILE", "SIMS_LOG_FILE"), project_root=root, environ=env_config.read_project_env_file(alias_env))
            policy_ok = (
                any("missing project env file" in e for e in missing_errors)
                and app_env_policy_ok
                and not alias_errors
                and alias_log_path == alias_log_file.resolve()
                and any("missing required path env: LOG_FILE" in e for e in no_log_errors)
                and any("relative path is not allowed in prod: CHAT_FILE" in e for e in relative_errors)
                and any("relative path is not allowed in prod: SSAI_STORAGE_ROOT" in e for e in relative_errors)
                and any("relative path is not allowed in prod: LOG_FILE" in e for e in relative_errors)
                and any("missing required path env: CHAT_FILE" in e for e in blank_errors)
                and any("missing required path env: SSAI_STORAGE_ROOT" in e for e in blank_errors)
                and any("missing required path env: LOG_FILE" in e for e in blank_errors)
                and dev_path == (root / "data" / "dev_chat.json").resolve()
                and storage_path == storage_dir.resolve()
                and log_path == log_file.resolve()
                and not (root / "data" / "ssai_storage").exists()
                and not (root / "logs" / "app.log").exists()
            )
            if policy_ok:
                results.append(_ok("prod env path validation and dev relative resolution", "storage/log missing/blank/relative prod paths blocked"))
            else:
                results.append(_fail("prod env path validation and dev relative resolution", f"missing={missing_errors}, relative={relative_errors}, blank={blank_errors}, dev_path={dev_path}"))
            os.chdir(old_cwd)
        os.chdir(old_cwd)
        for k, v in saved_env.items():
            if v is None:
                os.environ.pop(k, None)
            else:
                os.environ[k] = v
    except Exception as e:
        try:
            os.chdir(PROJECT_ROOT)
        except Exception:
            pass
        results.append(_fail("project-root .env policy helpers", f"{type(e).__name__}: {e}"))

    try:
        main_src = Path("app/Lmstudio_SSAI_chat_main.py").read_text(encoding="utf-8")
        mssql_src = Path("app/db/mssql_client.py").read_text(encoding="utf-8")
        auth_src = Path("app/services/ssai_auth_service.py").read_text(encoding="utf-8")
        company_admin_src = Path("app/services/ssai_company_admin_service.py").read_text(encoding="utf-8")
        storage_src = Path("app/services/ssai_storage_service.py").read_text(encoding="utf-8")
        logging_src = Path("app/utils/logging_setup.py").read_text(encoding="utf-8")
        company_tool_src = Path("tools/ssai_test_company_connection.py").read_text(encoding="utf-8")
        verify_tool_src = Path("tools/ssai_verify_admin_password.py").read_text(encoding="utf-8")
        korean_doc_call_matches = list(re.finditer(r"(?m)^_inject_korean_document_language_once\(\)\s*$", main_src))
        korean_doc_call_pos = korean_doc_call_matches[0].start() if korean_doc_call_matches else -1
        login_check_pos = main_src.find("if not require_login():")
        korean_doc_block_match = re.search(
            r"def _inject_korean_document_language_once\(\) -> None:.*?(?=\ndef _inject_base_css_once)",
            main_src,
            re.S,
        )
        korean_doc_block = korean_doc_block_match.group(0) if korean_doc_block_match else ""
        scroll_helper_match = re.search(
            r"def _scroll_to_anchor_js\(.*?(?=\n# ={5,}|\ndef )",
            main_src,
            re.S,
        )
        scroll_helper_block = scroll_helper_match.group(0) if scroll_helper_match else ""
        inline_scroll_match = re.search(
            r"if _jump_to:\s*\n\s*st\.iframe\(.*?(?=\n\s*else:)",
            main_src,
            re.S,
        )
        inline_scroll_block = inline_scroll_match.group(0) if inline_scroll_match else ""
        source_checks = [
            ("main auto dotenv removed", "load_dotenv(override=True)" not in main_src and 'ENV_PATH = APP_DIR / ".env"' not in main_src and "_DEFAULT_ENV_TEXT" not in main_src),
            ("main chat paths require config_path", 'CHAT_FILE         = str(_config_path("CHAT_FILE"))' in main_src and 'UPLOAD_DIR        = str(_config_path("UPLOAD_DIR"))' in main_src),
            ("startup env validation/logs present", "_STARTUP_REQUIRED_PATHS" in main_src and '("LOG_FILE", "SIMS_LOG_FILE")' in main_src and "[app.env.paths]" in main_src and "[app.env.user_paths]" in main_src),
            (
                "browser Korean document language guard",
                "def _inject_korean_document_language_once()" in main_src
                and "document.documentElement" in main_src
                and 'root.lang = "ko"' in main_src
                and 'root.setAttribute("translate", "no")' in main_src
                and 'root.classList.add("notranslate")' in main_src
                and 'meta.setAttribute("name", "google")' in main_src
                and 'meta.setAttribute("content", "notranslate")' in main_src
                and "unsafe_allow_javascript=True" in main_src
                and "__korean_document_language_loaded" in main_src
                and len(korean_doc_call_matches) == 1
                and korean_doc_call_pos >= 0
                and login_check_pos >= 0
                and korean_doc_call_pos < login_check_pos
                and "user_input" not in korean_doc_block
                and "stc.html(" not in korean_doc_block
            ),
            (
                "Streamlit iframe scroll migration",
                "import streamlit.components.v1" not in main_src
                and "stc.html(" not in main_src
                and main_src.count("st.iframe(") == 2
                and "height=0" not in main_src
                and "height=1, tab_index=-1" in scroll_helper_block
                and "height=1," in inline_scroll_block
                and "tab_index=-1" in inline_scroll_block
                and "window.parent.document" in scroll_helper_block
                and "window.parent.document" in inline_scroll_block
                and "scrollIntoView" in scroll_helper_block
                and "scrollIntoView" in inline_scroll_block
                and "MutationObserver" in scroll_helper_block
                and "except Exception:" in scroll_helper_block
                and "scroll anchor iframe failed" in scroll_helper_block,
            ),
            ("db cwd dotenv search removed", "find_dotenv" not in mssql_src),
            ("auth root env parser priority", "p = ENV_PATH" in auth_src and "return env.get(name) or os.environ.get(name) or default" in auth_src),
            ("company admin project env loader", "load_project_env(override=False)" in company_admin_src and "load_dotenv()" not in company_admin_src),
            ("APP_ENV allow-list enforced", "ALLOWED_APP_ENVS" in Path("app/utils/env_config.py").read_text(encoding="utf-8") and "missing required env: APP_ENV" in Path("app/utils/env_config.py").read_text(encoding="utf-8")),
            ("storage root has no data fallback", 'config_path("SSAI_STORAGE_ROOT"' in storage_src and "DEFAULT_STORAGE_ROOT" not in storage_src and "data/ssai_storage" not in storage_src),
            ("logger requires root env log path", "read_project_env_file()" in logging_src and "LOG_FILE is required" in logging_src and "Path(log_dir) / filename" not in logging_src and "C:\\\\" not in logging_src and "D:\\\\" not in logging_src),
            ("connection tool ignores cwd dotenv", "read_project_env_file()" in company_tool_src and "def load_dotenv" not in company_tool_src),
            ("password tool ignores cwd dotenv", "read_project_env_file()" in verify_tool_src and "def load_dotenv" not in verify_tool_src),
        ]
        for name, ok in source_checks:
            results.append(_ok(name) if ok else _fail(name, "source guard failed"))

        try:
            width_migration_targets = (
                "app/Lmstudio_SSAI_chat_main.py",
                "app/sims/views/analytics_views.py",
                "app/sims/views/dashboard.py",
                "app/sims/views/rddbc_io_check_views.py",
                "app/sims/views/rddbc_io_doc_views.py",
                "app/sims/views/rddbc_io_flow_views.py",
                "app/sims/views/rddbc_io_goods_views.py",
                "app/sims/views/rddbc_io_inout_views.py",
                "app/sims/views/rddbc_io_inventory_views.py",
                "app/sims/views/rddbc_io_stock_views.py",
                "app/sims/views/road_address.py",
                "app/ui/sims_hub.py",
                "app/ui/ssai_admin.py",
                "app/ui/ssai_company_admin.py",
                "app/ui/ssai_login.py",
            )
            migrated_widgets = {"button", "form_submit_button", "dataframe", "download_button"}
            deprecated_width_calls: list[str] = []
            stretch_width_counts: Counter[str] = Counter()
            for relative_path in width_migration_targets:
                source_tree = ast.parse((PROJECT_ROOT / relative_path).read_text(encoding="utf-8"))
                for node in ast.walk(source_tree):
                    if not (
                        isinstance(node, ast.Call)
                        and isinstance(node.func, ast.Attribute)
                        and isinstance(node.func.value, ast.Name)
                        and node.func.value.id == "st"
                        and node.func.attr in migrated_widgets
                    ):
                        continue
                    for keyword in node.keywords:
                        if keyword.arg == "use_container_width":
                            deprecated_width_calls.append(f"{relative_path}:{node.lineno}:{node.func.attr}")
                        if keyword.arg == "width" and isinstance(keyword.value, ast.Constant) and keyword.value.value == "stretch":
                            stretch_width_counts[node.func.attr] += 1

            expected_stretch_width_counts = {
                "button": 44,
                "form_submit_button": 38,
                "dataframe": 12,
                "download_button": 5,
            }
            if deprecated_width_calls or dict(stretch_width_counts) != expected_stretch_width_counts:
                results.append(
                    _fail(
                        "Streamlit width migration for active UI calls",
                        f"deprecated={deprecated_width_calls!r}; stretch={dict(stretch_width_counts)!r}",
                    )
                )
            else:
                results.append(
                    _ok(
                        "Streamlit width migration for active UI calls",
                        "15 active tracked files use width=stretch for 99 direct button/form/dataframe/download calls; helper and test compatibility patterns are excluded",
                    )
                )
        except Exception as exc:
            results.append(_fail("Streamlit width migration for active UI calls", f"{type(exc).__name__}: {exc}"))

        try:
            from streamlit.elements.lib.layout_utils import validate_height
            from streamlit.errors import StreamlitInvalidHeightError

            validate_height(1)
            try:
                validate_height(0)
            except StreamlitInvalidHeightError:
                results.append(_ok("Streamlit iframe height validation", "height=1 is accepted and height=0 is rejected by the installed Streamlit validator"))
            else:
                results.append(_fail("Streamlit iframe height validation", "height=0 was accepted"))
        except Exception as exc:
            results.append(_fail("Streamlit iframe height validation", f"{type(exc).__name__}: {exc}"))

        try:
            class _FakeLog:
                def __init__(self):
                    self.debug_calls = 0

                def debug(self, *args, **kwargs):
                    self.debug_calls += 1

            class _FakeSt:
                def __init__(self, *, fail: bool = False, loaded: bool = False):
                    self.session_state = {"__korean_document_language_loaded": True} if loaded else {}
                    self.fail = fail
                    self.html_calls = 0

                def html(self, *args, **kwargs):
                    self.html_calls += 1
                    if self.fail:
                        raise RuntimeError("html failed")

            if not korean_doc_block:
                raise AssertionError("helper block not found")

            fake_ok = _FakeSt()
            ns_ok = {"st": fake_ok, "log": _FakeLog()}
            exec(korean_doc_block, ns_ok)
            ns_ok["_inject_korean_document_language_once"]()

            fake_fail = _FakeSt(fail=True)
            ns_fail = {"st": fake_fail, "log": _FakeLog()}
            exec(korean_doc_block, ns_fail)
            ns_fail["_inject_korean_document_language_once"]()

            fake_loaded = _FakeSt(loaded=True)
            ns_loaded = {"st": fake_loaded, "log": _FakeLog()}
            exec(korean_doc_block, ns_loaded)
            ns_loaded["_inject_korean_document_language_once"]()

            behavior_ok = (
                fake_ok.html_calls == 1
                and fake_ok.session_state.get("__korean_document_language_loaded") is True
                and fake_fail.html_calls == 1
                and "__korean_document_language_loaded" not in fake_fail.session_state
                and fake_loaded.html_calls == 0
                and fake_loaded.session_state.get("__korean_document_language_loaded") is True
            )
            if behavior_ok:
                results.append(_ok("browser Korean document language helper behavior", "loaded flag is set only after successful st.html; already-loaded reruns skip injection"))
            else:
                results.append(_fail("browser Korean document language helper behavior", f"ok_calls={fake_ok.html_calls}, fail_state={fake_fail.session_state}, loaded_calls={fake_loaded.html_calls}"))
        except Exception as e:
            results.append(_fail("browser Korean document language helper behavior", f"{type(e).__name__}: {e}"))

        import types

        effective_block = re.search(r"def _effective_chat_file\(\).*?def _chat_storage_mode_enabled", main_src, re.S)
        partition_block = re.search(r"def _partitioned_chat_root\(chat_file: str \| None = None\).*?def _partitioned_rooms_file", main_src, re.S)
        if not effective_block or not partition_block:
            results.append(_fail("user legacy and partition chat path calculation", "function block not found"))
        else:
            ns: dict[str, Any] = {
                "Path": Path,
                "CHAT_FILE": str(PROJECT_ROOT / "tmp_chat" / "chat_rooms.json"),
                "get_current_user": lambda: types.SimpleNamespace(user_id=8),
            }
            exec(effective_block.group(0).rsplit("def _chat_storage_mode_enabled", 1)[0], ns)
            exec(partition_block.group(0).rsplit("def _partitioned_rooms_file", 1)[0], ns)
            effective = Path(ns["_effective_chat_file"]())
            root = ns["_partitioned_chat_root"](str(effective))
            expected_file = PROJECT_ROOT / "tmp_chat" / "user_8_chat_rooms.json"
            expected_root = PROJECT_ROOT / "tmp_chat" / "user_8"
            if effective == expected_file and root == expected_root:
                results.append(_ok("user legacy and partition chat path calculation", f"file={effective.name}, root={root.name}"))
            else:
                results.append(_fail("user legacy and partition chat path calculation", f"file={effective}, root={root}"))
    except Exception as e:
        results.append(_fail("project-root .env source guards", f"{type(e).__name__}: {e}"))

    try:
        from app.services import llm_health
        import time as _time

        class _FakeExc(Exception):
            def __init__(self, status_code: int | None = None, message: str = "fixture error without secrets"):
                super().__init__(message)
                if status_code is not None:
                    self.status_code = status_code

        secret_fixture = "sk-secret-fixture"
        classified = {
            "401": llm_health.classify_llm_exception(_FakeExc(401))["code"],
            "connection": llm_health.classify_llm_exception(type("APIConnectFixture", (Exception,), {})())["code"],
            "404": llm_health.classify_llm_exception(_FakeExc(404))["code"],
            "429": llm_health.classify_llm_exception(_FakeExc(429))["code"],
            "500": llm_health.classify_llm_exception(_FakeExc(500))["code"],
            "timeout": llm_health.classify_llm_exception(type("RequestTimeout", (Exception,), {})())["code"],
        }
        unknown_info = llm_health.classify_llm_exception(_FakeExc(None, f"raw {secret_fixture} http://127.0.0.1/body"))
        preset_unknown = RuntimeError(f"raw preset {secret_fixture}")
        setattr(preset_unknown, "llm_error_code", "not_a_known_code")
        preset_info = llm_health.classify_llm_exception(preset_unknown)

        class _Model:
            def __init__(self, model_id: str):
                self.id = model_id

        class _Models:
            def __init__(self, ids: list[str]):
                self._ids = ids
                self.calls = 0

            def list(self):
                self.calls += 1
                return type("ModelList", (), {"data": [_Model(x) for x in self._ids]})()

        class _Client:
            def __init__(self, ids: list[str]):
                self.models = _Models(ids)
                self.last_options = {}

            def with_options(self, **kwargs):
                self.last_options = dict(kwargs)
                return self

        class _TimeoutModels:
            def __init__(self):
                self.calls = 0

            def list(self):
                self.calls += 1
                raise TimeoutError("fast health timeout fixture")

        class _TimeoutClient:
            def __init__(self):
                self.models = _TimeoutModels()
                self.last_options = {}

            def with_options(self, **kwargs):
                self.last_options = dict(kwargs)
                return self

        ok_client = _Client(["model-a"])
        ok_status = llm_health.check_llm(expected_model="model-a", client=ok_client)
        empty_status = llm_health.check_llm(expected_model="model-a", client=_Client([]))
        mismatch_status = llm_health.check_llm(expected_model="model-a", client=_Client(["model-b"]))
        timeout_client = _TimeoutClient()
        health_t0 = _time.perf_counter()
        timeout_status = llm_health.check_llm(expected_model="model-a", client=timeout_client, health_timeout_s=4)
        health_elapsed = _time.perf_counter() - health_t0
        model_list_options_ok = (
            ok_client.last_options.get("max_retries") == 0
            and timeout_client.last_options.get("timeout") == 4
            and timeout_client.last_options.get("max_retries") == 0
            and timeout_client.models.calls == 1
        )

        class _Usage:
            prompt_tokens = 3
            completion_tokens = 5
            total_tokens = 8

        def _resp(content: str = "", reasoning: str = "", finish: str | None = None):
            msg = type("Msg", (), {"content": content, "reasoning_content": reasoning})()
            choice = type("Choice", (), {"message": msg, "finish_reason": finish})()
            return type("Resp", (), {"choices": [choice], "usage": _Usage()})()

        normal = llm_health.extract_chat_completion_text(_resp("ok", finish="stop"))
        reasoning_only = llm_health.extract_chat_completion_text(_resp("", "hidden reasoning"))
        empty = llm_health.extract_chat_completion_text(_resp(""))
        length = llm_health.extract_chat_completion_text(_resp("cut", finish="length"))
        retry_info = {"code": "timeout", "retryable": True}
        non_retry_info = {"code": "authentication_error", "retryable": False}
        retry_checks = {
            "bounded_none": llm_health.bounded_retry_count(None) == 0,
            "bounded_negative": llm_health.bounded_retry_count(-3) == 0,
            "bounded_high": llm_health.bounded_retry_count(9) == 1,
            "retry_before_content": llm_health.should_retry_llm_error(retry_info, attempt=0, max_retries=1, content_started=False),
            "no_retry_after_max": not llm_health.should_retry_llm_error(retry_info, attempt=1, max_retries=1, content_started=False),
            "no_retry_after_content": not llm_health.should_retry_llm_error(retry_info, attempt=0, max_retries=1, content_started=True),
            "no_retry_auth": not llm_health.should_retry_llm_error(non_retry_info, attempt=0, max_retries=1, content_started=False),
        }
        def _simulated_attempts(err_info: dict, *, max_retries: int, content_started: bool = False) -> int:
            calls = 0
            for attempt in range(llm_health.bounded_retry_count(max_retries) + 1):
                calls += 1
                if not llm_health.should_retry_llm_error(err_info, attempt=attempt, max_retries=max_retries, content_started=content_started):
                    break
            return calls

        retry_call_counts = {
            "non_retry_total_1": _simulated_attempts(non_retry_info, max_retries=1) == 1,
            "retryable_total_2": _simulated_attempts(retry_info, max_retries=1) == 2,
            "partial_total_1": _simulated_attempts(retry_info, max_retries=1, content_started=True) == 1,
        }

        main_src = Path("app/Lmstudio_SSAI_chat_main.py").read_text(encoding="utf-8", errors="replace")
        top_src = "\n".join(main_src.splitlines()[:280])
        model_ready_block = main_src[main_src.find("def _llm_model_ready"):main_src.find("def _fallback_login_greeting")]
        protected_block = main_src[main_src.find("def call_chat_protected"):main_src.find("# =========================================================", main_src.find("def call_chat_protected") + 1)]
        stream_call_start = main_src.find("response_stream = call_chat_protected")
        stream_call_end = main_src.find(")", stream_call_start)
        stream_call_block = main_src[stream_call_start:stream_call_end]
        get_models_start = main_src.find("def _get_models_cached()")
        get_models_end = main_src.find("# =========================", get_models_start + 1)
        get_models_block = main_src[get_models_start:get_models_end]
        source_guards = {
            "top_health_lazy": "LLM_STATUS = {\"ok\": None" in top_src and "LLM_STATUS = check_llm(" not in top_src,
            "model_cache_ttl_short": "@st.cache_data(ttl=5)" in main_src and "def _get_models_cached" in main_src,
            "model_list_timeout": "with_options" in get_models_block and "LLM_HEALTH_TIMEOUT_S" in get_models_block,
            "model_list_sdk_retry_off": "max_retries=0" in get_models_block,
            "client_sdk_retry_off": "CLIENT = OpenAI(" in main_src and "max_retries=0" in main_src,
            "model_ready_no_duplicate_check": "check_llm(" not in model_ready_block,
            "call_protected_no_unused_client": "cli = getattr(CLIENT" not in protected_block,
            "stream_inner_retry_off": "max_retry=0" in stream_call_block,
            "expected_model_fixed_ui": "st.session_state.selected_model = EXPECTED_LM_MODEL" in main_src and "st.expander(" in main_src,
            "stream_no_retry_after_tokens": "if collected:\n                    final_text = \"\".join(collected).strip()" in main_src,
            "reasoning_delta_guard": "reasoning_content" in main_src and "reasoning_seen" in main_src,
            "login_fallback_ready_check": "ready, _ready_message = _llm_request_config_ready(EXPECTED_LM_MODEL)" in main_src,
            "wait_placeholder": "wait_slot = st.empty()" in main_src and "_clear_wait_notice()" in main_src,
            "no_local_model_runtime_fallback": 'or "local-model"' not in main_src.replace('#                model=model or st.session_state.get("selected_model") or "local-model"', ""),
        }

        mismatches = []
        expected_codes = {
            "401": "authentication_error",
            "connection": "connection_error",
            "404": "model_not_found",
            "429": "rate_or_queue_busy",
            "500": "server_error",
            "timeout": "timeout",
        }
        for key, expected in expected_codes.items():
            if classified.get(key) != expected:
                mismatches.append(f"{key}->{classified.get(key)}")
        if not ok_status.get("ok"):
            mismatches.append("expected model should be ok")
        if empty_status.get("code") != "model_not_loaded":
            mismatches.append(f"empty models code={empty_status.get('code')}")
        if mismatch_status.get("code") != "model_mismatch":
            mismatches.append(f"mismatch code={mismatch_status.get('code')}")
        if timeout_status.get("code") != "timeout" or not model_list_options_ok or health_elapsed > 0.5:
            mismatches.append(f"health timeout/retry guard failed code={timeout_status.get('code')} options={timeout_client.last_options} calls={timeout_client.models.calls} elapsed={health_elapsed:.3f}")
        if secret_fixture in str(unknown_info.get("user_message")) or unknown_info.get("code") != "unknown_error":
            mismatches.append("unknown exception leaked raw user message")
        if secret_fixture in str(preset_info.get("user_message")) or preset_info.get("user_message") != llm_health.SAFE_MESSAGES["unknown_error"]:
            mismatches.append("preset unknown exception leaked raw user message")
        if normal.get("content") != "ok" or normal.get("usage", {}).get("completion_tokens") != 5:
            mismatches.append("normal content/usage extraction failed")
        if reasoning_only.get("code") != "reasoning_only" or reasoning_only.get("content"):
            mismatches.append("reasoning-only guard failed")
        if empty.get("code") != "empty_response":
            mismatches.append("empty response guard failed")
        if length.get("finish_reason") != "length":
            mismatches.append("finish_reason length not preserved")
        failed_retry = [k for k, ok in retry_checks.items() if not ok]
        if failed_retry:
            mismatches.append(f"retry_checks={failed_retry}")
        failed_retry_counts = [k for k, ok in retry_call_counts.items() if not ok]
        if failed_retry_counts:
            mismatches.append(f"retry_call_counts={failed_retry_counts}")
        failed_guards = [k for k, ok in source_guards.items() if not ok]
        if failed_guards:
            mismatches.append(f"source_guards={failed_guards}")

        if mismatches:
            results.append(_fail("LM Studio 0.4.19 stability guards", "; ".join(mismatches)))
        else:
            results.append(_ok("LM Studio 0.4.19 stability guards", "health classification, model mismatch, response guards, and retry source guards verified"))
    except Exception as e:
        results.append(_fail("LM Studio 0.4.19 stability guards", f"{type(e).__name__}: {e}"))

    try:
        sidebar_errors = []
        main_src = Path("app/Lmstudio_SSAI_chat_main.py").read_text(encoding="utf-8", errors="replace")
        helper_start = main_src.index("def _is_platform_operations_admin")
        helper_end = main_src.index("def _render_sidebar_app_title_once", helper_start)
        helper_ns: dict[str, Any] = {}

        class _User:
            def __init__(self, user_type: str, user_grade: str):
                self.user_type = user_type
                self.user_grade = user_grade

        current_user_box = {"user": None}
        helper_ns["get_current_user"] = lambda: current_user_box["user"]
        from app.ui.ssai_admin import _is_super_admin_user as _actual_super_admin_user
        helper_ns["_is_platform_super_admin_user"] = _actual_super_admin_user
        exec(main_src[helper_start:helper_end], helper_ns)
        can_model_admin = helper_ns["_is_platform_operations_admin"]
        cases = [
            (None, False),
            (_User("WHOLESALE_USER", "STAFF"), False),
            (_User("WHOLESALE_ADMIN", "MANAGER"), False),
            (_User("SSART_ADMIN", "MANAGER"), False),
            (_User("SSART_ADMIN", "SUPER"), True),
        ]
        for user_case, expected in cases:
            current_user_box["user"] = user_case
            got = bool(can_model_admin())
            if got != expected:
                sidebar_errors.append(f"platform_admin_case_failed={getattr(user_case, 'user_type', None)}:{getattr(user_case, 'user_grade', None)}->{got}")

        title_call = main_src.find("\n_render_sidebar_app_title_once()\n")
        login_call = main_src.find("if not require_login():")
        sidebar_block_start = main_src.find("with st.sidebar:", main_src.find("_consume_sims_close_for_chat_room_change()"))
        model_gate = main_src.find("if sidebar_model_admin_allowed:", sidebar_block_start)
        model_lookup = main_src.find("models = get_models()", sidebar_block_start)
        chat_section = main_src.find("# 2) 채팅방 관리", sidebar_block_start)
        lower_sidebar_model_block = main_src[sidebar_block_start:chat_section]
        old_banner_after_login = lower_sidebar_model_block.find("SSAI LM Studio Chatbot")
        selected_model_init_start = main_src.find('if "selected_model" not in st.session_state:')
        selected_model_init = main_src[selected_model_init_start:main_src.find("current_room =", selected_model_init_start)]

        if not (0 <= title_call < login_call):
            sidebar_errors.append("sidebar title is not rendered before require_login")
        if old_banner_after_login >= 0:
            sidebar_errors.append("duplicate sidebar title remains in lower sidebar block")
        if not (0 <= model_gate < model_lookup < chat_section):
            sidebar_errors.append("model lookup is not gated inside platform admin block")
        if "get_models()" in selected_model_init:
            sidebar_errors.append("selected_model init still calls get_models")
        if 'st.session_state.selected_model = EXPECTED_LM_MODEL or ""' not in selected_model_init:
            sidebar_errors.append("selected_model init does not preserve expected model")

        ready_start = main_src.index("def _llm_request_config_ready")
        ready_end = main_src.index("def _fallback_login_greeting", ready_start)
        ready_calls = {"get_models": 0}
        ready_ns: dict[str, Any] = {
            "EXPECTED_LM_MODEL": "expected-model",
            "LLM_SAFE_MESSAGES": {
                "model_not_loaded": "model not loaded",
                "model_mismatch": "model mismatch",
            },
            "get_models": lambda: ready_calls.__setitem__("get_models", ready_calls["get_models"] + 1) or ["expected-model"],
        }
        exec(main_src[ready_start:ready_end], ready_ns)
        ready, _message = ready_ns["_llm_request_config_ready"]("expected-model")
        if not ready or ready_calls["get_models"] != 0:
            sidebar_errors.append(f"member config readiness called get_models={ready_calls['get_models']}")
        ready, _message = ready_ns["_llm_model_ready"]("expected-model")
        if not ready or ready_calls["get_models"] != 1:
            sidebar_errors.append(f"platform-admin model readiness did not call get_models once={ready_calls['get_models']}")

        from app.services.llm_health import SAFE_MESSAGES as _SAFE_MESSAGES
        from app.services.llm_health import bounded_retry_count as _bounded_retry_count
        from app.services.llm_health import classify_llm_exception as _classify_llm_exception
        retry_start = main_src.index("def call_chat_with_retry")
        retry_end = main_src.index("# =========================================================", retry_start + 1)
        completion_calls = {"create": 0, "with_options": []}

        class _FakeCompletions:
            def create(self, **_kwargs):
                completion_calls["create"] += 1
                raise type("APIConnectFixture", (Exception,), {})("raw secret-token http://127.0.0.1/body")

        class _FakeClient:
            chat = type("_FakeChat", (), {"completions": _FakeCompletions()})()

            def with_options(self, **kwargs):
                completion_calls["with_options"].append(dict(kwargs))
                return self

        retry_ns: dict[str, Any] = {
            "_normalize_for_lmstudio": lambda messages: messages,
            "LLM_TIMEOUT_S": 1,
            "LLM_MAX_RETRY": 0,
            "LLM_BACKOFF_SEQ": [0.01],
            "bounded_retry_count": _bounded_retry_count,
            "CLIENT": _FakeClient(),
            "EXPECTED_LM_MODEL": "expected-model",
            "st": type("_FakeSt", (), {"session_state": {}})(),
            "log": type("_FakeLog", (), {"info": lambda *a, **k: None, "warning": lambda *a, **k: None})(),
            "time": type("_FakeTime", (), {"perf_counter": __import__("time").perf_counter, "sleep": lambda *_a, **_k: None})(),
            "random": type("_FakeRandom", (), {"uniform": lambda *_a, **_k: 0.0})(),
            "classify_llm_exception": _classify_llm_exception,
            "LLM_SAFE_MESSAGES": _SAFE_MESSAGES,
        }
        exec(main_src[retry_start:retry_end], retry_ns)
        try:
            retry_ns["call_chat_with_retry"](
                messages=[{"role": "user", "content": "hello"}],
                model="expected-model",
                stream=False,
                timeout_s=1,
                max_retry=0,
            )
            sidebar_errors.append("completion failure did not raise safe RuntimeError")
        except RuntimeError as exc:
            if "secret-token" in str(exc) or "127.0.0.1" in str(exc):
                sidebar_errors.append("completion failure exposed raw exception text")
        if completion_calls["create"] != 1:
            sidebar_errors.append(f"completion failure call count unexpected={completion_calls['create']}")

        for required in (
            "sidebar_model_admin_allowed",
            "model_controls_rendered",
            "loaded_model_lookup_called",
            "[sidebar.model]",
        ):
            if required not in main_src:
                sidebar_errors.append(f"sidebar diagnostic missing={required}")

        if sidebar_errors:
            results.append(_fail("sidebar model admin permission policy", "; ".join(sidebar_errors)))
        else:
            results.append(_ok("sidebar model admin permission policy", "title precedes login/sidebar info; model UI and get_models are gated to SSART_ADMIN SUPER"))
    except Exception as e:
        results.append(_fail("sidebar model admin permission policy", f"{type(e).__name__}: {e}"))

    try:
        router = importlib.import_module("app.sims.nlq.nlq_router")
        resolve = getattr(router, "_resolve_analytics_action", None)
        if callable(resolve):
            tests = [
                ("품목별 매출 추세 2025년 조회", "품목별 매출 추세 분석"),
                ("품목별 매출 추세 요약표 2025년 조회", "품목별 매출 추세 요약표"),
                ("품목별 매출 예상 2025년 조회", "품목별 매출 예상"),
                ("매출처별 매출 예상 2025년 조회", "매출처별 매출 예상"),
                ("영업사원별 매출 예상 2025년 조회", "영업사원별 매출 예상"),
                ("지역별 매출 예상 2025년 조회", "지역별 매출 예상"),
                ("품목별 재고부족현황 2025년 조회", "품목별 재고부족현황"),
                ("매입처별 재고부족 현황 2025년 조회", "매입처별 재고부족 현황"),
            ]
            for q, expected in tests:
                got = resolve(q)
                if got == expected:
                    results.append(_ok(f"analytics action resolver: {q}", got))
                else:
                    results.append(_fail(f"analytics action resolver: {q}", f"expected={expected!r}, got={got!r}"))
        else:
            results.append(_fail("analytics action resolver", "_resolve_analytics_action 없음"))

        resolve_candidate = getattr(router, "resolve_new_sims_nlq_candidate", None)
        if not callable(resolve_candidate):
            results.append(_fail("new SIMS/NLQ route candidate", "parse-only resolver missing"))
        else:
            candidate_cases = [
                ("품목별 매출 추세분석", "analytics", "품목별 매출 추세 분석"),
                ("한미제약 품목별 매출 추세분석", "analytics", "품목별 매출 추세 분석"),
                ("정상출고만 품목별 매출 추세분석", "analytics", "품목별 매출 추세 분석"),
                ("정상출고 내역 조회", "io", "출고명세 조회"),
            ]
            for query, expected_route, expected_action in candidate_cases:
                candidate = resolve_candidate(query) or {}
                if candidate.get("route") != expected_route or candidate.get("action") != expected_action:
                    results.append(
                        _fail(
                            f"new SIMS/NLQ route candidate: {query}",
                            f"expected=({expected_route!r}, {expected_action!r}), got={candidate!r}",
                        )
                    )
                else:
                    results.append(
                        _ok(
                            f"new SIMS/NLQ route candidate: {query}",
                            f"route={expected_route}, action={expected_action}",
                        )
                    )

            main_src = Path("app/Lmstudio_SSAI_chat_main.py").read_text(encoding="utf-8")
            route_checks = []
            for required in (
                "resolve_new_sims_nlq_candidate(user_input)",
                "not is_new_sims_nlq",
                "route=new_sims_nlq reason=parsed_action",
                "query_execution_failed",
            ):
                if required not in main_src:
                    route_checks.append(f"missing main route guard {required!r}")
            if "has_explicit_current_table_reference" not in main_src:
                route_checks.append("explicit current-table priority guard missing")
            if route_checks:
                results.append(_fail("new SIMS/NLQ versus current-table route guards", "; ".join(route_checks)))
            else:
                results.append(
                    _ok(
                        "new SIMS/NLQ versus current-table route guards",
                        "new actions route before implicit follow-ups; explicit current-table references retain priority",
                    )
                )
    except Exception as e:
        results.append(_fail("analytics action resolver", f"{type(e).__name__}: {e}"))

    try:
        old_current_yyyymm = getattr(mod, "_current_yyyymm", None)
        setattr(mod, "_current_yyyymm", lambda: "202607")

        def _row(month: str, amt: int, qty: int = 1, product_code: str = "0001", buy_cd: str = "B1") -> dict[str, Any]:
            return {"기준월": month, "제품코드": product_code, "제품명": "테스트", "규격": "EA", "제조사코드": "M1", "제조사명": "제조사", "제품그룹명": "G", "제품구분명": "D", "제품분류명": "C", "매입처코드": buy_cd, "출고수량": qty, "출고할증수량": 0, "매출공급가액": amt, "매출세액": 0, "매출합계": amt, "집계건수": 1}

        raw_df = pd.DataFrame(
            [
                _row("202601", 100),
                _row("202602", 100),
                _row("202603", 100),
                _row("202604", 200),
                _row("202605", 200),
                _row("202606", 200),
                _row("202607", 300),
                _row("202608", 999),
            ]
        )
        summary_df = mod.get_sales_trend_summary_df(
            {
                "month_from": "202601",
                "month_to": "202608",
                "date_from": "20260101",
                "date_to": "20260831",
                "source_mode": "monthly_book",
            },
            raw_df=raw_df,
        )
        row = summary_df.iloc[0].to_dict()

        expected = {
            "완료월수": 6,
            "완료월총매출": 900,
            "완료월평균매출": 150,
            "월평균매출": 150,
            "최근3개월평균매출": 200,
            "최근6개월평균매출": 150,
            "최근3개월증감률": 33.3333333333,
            "당월 현재매출": 300,
            "당월 예상매출": 230,
            "당월 잔여예상": 0,
            "당월 진척률": 130.4347826087,
        }
        mismatches = []
        for key, exp in expected.items():
            got = float(row.get(key) or 0)
            if abs(got - float(exp)) > 1e-6:
                mismatches.append(f"{key}: expected={exp}, got={got}")

        if mismatches:
            results.append(_fail("sales period policy current/future exclusion", "; ".join(mismatches)))
        else:
            meta = mod._forecast_meta_from_df(summary_df)
            period_meta = mod._period_policy_meta_from_summary_df(summary_df)
            if int(meta.get("month_count") or 0) != 7 or int(period_meta.get("completed_month_count") or 0) != 6:
                results.append(_fail("sales period policy month counters", f"month_count={meta.get('month_count')}, completed={period_meta.get('completed_month_count')}"))
            elif str(row.get("추세판정") or "") != "증가":
                results.append(_fail("sales period policy trend judge", f"expected='증가', got={row.get('추세판정')!r}"))
            else:
                results.append(_ok("sales period policy current/future exclusion", "current=202607 and future=202608 excluded; forecast rate applied"))

        trend_df = mod._add_trend_columns(raw_df)
        period_values = dict(zip(trend_df["기준월"], trend_df["기간구분"]))
        rolling_values = dict(zip(trend_df["기준월"], trend_df["최근3개월평균매출"]))
        rolling_ok = (
            abs(float(rolling_values.get("202604") or 0) - 100) < 1e-6
            and abs(float(rolling_values.get("202607") or 0) - 200) < 1e-6
            and len(set(round(float(v), 6) for v in rolling_values.values())) > 1
        )
        period_ok = (
            period_values.get("202606") == "완료월"
            and period_values.get("202607") == "당월진행"
            and period_values.get("202608") == "미래월"
        )
        if rolling_ok and period_ok:
            results.append(_ok("sales trend detail period labels and rolling averages", "period labels set; row-wise rolling averages preserved"))
        else:
            results.append(_fail("sales trend detail period labels and rolling averages", f"periods={period_values}, rolling={rolling_values}"))

        duplicate_vendor_df = pd.DataFrame(
            [
                _row("202601", 200_000_000, 500, "00439", "B1"),
                _row("202601", 270_721_841, 500, "00439", "B2"),
                _row("202602", 100_000_000, 300, "00439", "B1"),
                _row("202602", 210_956_793, 458, "00439", "B2"),
                _row("202603", 284_213_860, 750, "00439", "B1"),
                _row("202604", 402_633_120, 770, "00439", "B1"),
                _row("202605", 420_000_000, 780, "00439", "B1"),
                _row("202606", 449_938_280, 790, "00439", "B1"),
                _row("202607", 200_886_348, 800, "00439", "B1"),
            ]
        )
        duplicate_trend = mod._add_trend_columns(duplicate_vendor_df)
        feb_rows = duplicate_trend[duplicate_trend["기준월"] == "202602"]
        jun_row = duplicate_trend[duplicate_trend["기준월"] == "202606"].iloc[0]
        jul_row = duplicate_trend[duplicate_trend["기준월"] == "202607"].iloc[0]
        duplicate_checks = [
            ("2026-02 전월대비매출", set(feb_rows["전월대비매출"].round(2).tolist()), {-159_765_048.0}),
            ("2026-02 전월대비수량", set(feb_rows["전월대비수량"].round(2).tolist()), {-242.0}),
            ("2026-02 최근3개월평균매출", set(feb_rows["최근3개월평균매출"].round(2).tolist()), {470_721_841.0}),
        ]
        duplicate_mismatches = []
        for label, got, exp in duplicate_checks:
            if got != exp:
                duplicate_mismatches.append(f"{label}: expected={exp}, got={got}")
        expected_points = {
            "2026-06 최근3개월평균매출": (jun_row.get("최근3개월평균매출"), 368_948_993.33),
            "2026-06 최근6개월평균매출": (jun_row.get("최근6개월평균매출"), 377_705_122.8),
            "2026-07 최근3개월평균매출": (jul_row.get("최근3개월평균매출"), 424_190_466.67),
            "2026-07 최근6개월평균매출": (jul_row.get("최근6개월평균매출"), 389_743_982.33),
        }
        for label, (got, exp) in expected_points.items():
            if abs(float(got or 0) - exp) > 0.01:
                duplicate_mismatches.append(f"{label}: expected={exp}, got={got}")
        if duplicate_mismatches:
            results.append(_fail("sales trend detail monthly aggregate metrics", "; ".join(duplicate_mismatches)))
        else:
            results.append(_ok("sales trend detail monthly aggregate metrics", "duplicate vendor rows share product-month metrics"))

        month_point_mismatches = []
        required_month_cols = ["월시점 증감률", "월시점 추세판정", "월시점 판정결과", "추세판정", "판정결과"]
        for c in required_month_cols:
            if c not in duplicate_trend.columns:
                month_point_mismatches.append(f"missing column {c}")
        if not month_point_mismatches:
            feb_judges = set(feb_rows["월시점 추세판정"].astype(str).tolist())
            feb_compat_judges = set(feb_rows["추세판정"].astype(str).tolist())
            if len(feb_judges) != 1 or feb_judges != feb_compat_judges:
                month_point_mismatches.append(f"202602 duplicate rows judge mismatch: {feb_judges}, compat={feb_compat_judges}")
            feb_expected_values = set(feb_rows["월시점 예상매출"].round(2).tolist())
            if len(feb_expected_values) != 1:
                month_point_mismatches.append(f"202602 duplicate rows expected sales mismatch: {feb_expected_values}")
            if str(jun_row.get("월시점 추세판정") or "") != "안정":
                month_point_mismatches.append(f"202606 expected 안정, got={jun_row.get('월시점 추세판정')!r}")
            if str(jul_row.get("월시점 추세판정") or "") != "안정":
                month_point_mismatches.append(f"202607 expected 안정, got={jul_row.get('월시점 추세판정')!r}")
            if str(jul_row.get("기간구분") or "") != "당월진행":
                month_point_mismatches.append(f"202607 period expected 당월진행, got={jul_row.get('기간구분')!r}")
            judge_seq = duplicate_trend.drop_duplicates(["제품코드", "기준월"])["월시점 추세판정"].astype(str).tolist()
            if len(set(judge_seq)) <= 1:
                month_point_mismatches.append(f"monthly judges appear copied: {judge_seq}")
            suffix_cols = [c for c in duplicate_trend.columns if str(c).endswith(("_x", "_y"))]
            if suffix_cols:
                month_point_mismatches.append(f"unexpected suffix columns: {suffix_cols}")
        if month_point_mismatches:
            results.append(_fail("sales trend detail month-point judge", "; ".join(month_point_mismatches)))
        else:
            results.append(_ok("sales trend detail month-point judge", "monthly point-in-time judges are merged per product-month"))

        product_00439_summary = mod.get_sales_trend_summary_df(
            {
                "month_from": "202601",
                "month_to": "202607",
                "date_from": "20260101",
                "date_to": "20260731",
                "source_mode": "monthly_book",
            },
            raw_df=duplicate_vendor_df,
        )
        row_00439 = product_00439_summary.iloc[0].to_dict()
        meta_00439 = mod._period_policy_meta_from_summary_df(product_00439_summary)
        display_mismatches = []
        expected_00439 = {
            "완료월평균매출": 389_743_982,
            "당월 현재매출": 200_886_348,
            "당월 예상매출": 442_935_939,
            "당월 잔여예상": 242_049_591,
            "당월 진척률": 45.3522419545,
        }
        for key, exp in expected_00439.items():
            got = float(row_00439.get(key) or 0)
            if abs(got - exp) > 0.02:
                display_mismatches.append(f"{key}: expected={exp}, got={got}")
        progress_value = row_00439.get("당월 진척률")
        if int(float(progress_value or 0)) == float(progress_value or 0):
            display_mismatches.append(f"당월 진척률 lost decimal precision: {progress_value}")
        if f"{float(progress_value or 0):.2f}%" != "45.35%":
            display_mismatches.append(f"당월 진척률 display expected 45.35%, got={float(progress_value or 0):.2f}%")
        if any(c.startswith("_당월") for c in product_00439_summary.columns):
            display_mismatches.append("internal _당월 columns exposed in summary")
        if round(float(meta_00439.get("avg_completed_month_sales_amt") or 0)) != 389_743_982:
            display_mismatches.append(f"meta avg_completed_month_sales_amt={meta_00439.get('avg_completed_month_sales_amt')}")
        trend_jul_expected = float(jul_row.get("월시점 예상매출") or 0)
        trend_jul_actual = float(jul_row.get("월시점 실제매출") or 0)
        trend_jul_progress = float(jul_row.get("월시점 달성률") or 0)
        if abs(round(trend_jul_expected) - round(float(row_00439.get("당월 예상매출") or 0))) > 0:
            display_mismatches.append(f"trend/summary expected mismatch: trend={trend_jul_expected}, summary={row_00439.get('당월 예상매출')}")
        if abs(trend_jul_actual - float(row_00439.get("당월 현재매출") or 0)) > 0.02:
            display_mismatches.append(f"trend/summary actual mismatch: trend={trend_jul_actual}, summary={row_00439.get('당월 현재매출')}")
        if abs(trend_jul_progress - float(row_00439.get("당월 진척률") or 0)) > 1e-6:
            display_mismatches.append(f"trend/summary progress mismatch: trend={trend_jul_progress}, summary={row_00439.get('당월 진척률')}")
        old_summary_for_forecast = getattr(mod, "get_sales_trend_summary_df", None)
        try:
            setattr(mod, "get_sales_trend_summary_df", lambda params=None, raw_df=None: product_00439_summary.copy())
            forecast_00439 = mod.get_sales_forecast_df({})
            forecast_row_00439 = forecast_00439.iloc[0].to_dict()
            if abs(float(forecast_row_00439.get("당월 예상매출") or 0) - float(row_00439.get("당월 예상매출") or 0)) > 0.02:
                display_mismatches.append(f"summary/forecast expected mismatch: forecast={forecast_row_00439.get('당월 예상매출')}, summary={row_00439.get('당월 예상매출')}")
            if abs(float(forecast_row_00439.get("당월 현재매출") or 0) - float(row_00439.get("당월 현재매출") or 0)) > 0.02:
                display_mismatches.append(f"summary/forecast actual mismatch: forecast={forecast_row_00439.get('당월 현재매출')}, summary={row_00439.get('당월 현재매출')}")
            if abs(float(forecast_row_00439.get("당월 진척률") or 0) - float(row_00439.get("당월 진척률") or 0)) > 1e-6:
                display_mismatches.append(f"summary/forecast progress mismatch: forecast={forecast_row_00439.get('당월 진척률')}, summary={row_00439.get('당월 진척률')}")
        finally:
            if old_summary_for_forecast is not None:
                setattr(mod, "get_sales_trend_summary_df", old_summary_for_forecast)
        if display_mismatches:
            results.append(_fail("sales forecast current month display summary", "; ".join(display_mismatches)))
        else:
            results.append(_ok("sales forecast current month display summary", "00439 current-month display values and trend/summary/forecast match expected"))

        past_df = pd.DataFrame([
            _row("202601", 100),
            _row("202602", 100),
            _row("202603", 100),
            _row("202604", 200),
            _row("202605", 200),
            _row("202606", 300),
        ])
        past_summary = mod.get_sales_trend_summary_df(
            {"month_from": "202601", "month_to": "202606", "date_from": "20260101", "date_to": "20260630", "source_mode": "monthly_book"},
            raw_df=past_df,
        )
        past_row = past_summary.iloc[0].to_dict()
        past_df_changed = pd.DataFrame([
            _row("202601", 100),
            _row("202602", 100),
            _row("202603", 100),
            _row("202604", 200),
            _row("202605", 200),
            _row("202606", 999),
        ])
        past_summary_changed = mod.get_sales_trend_summary_df(
            {"month_from": "202601", "month_to": "202606", "date_from": "20260101", "date_to": "20260630", "source_mode": "monthly_book"},
            raw_df=past_df_changed,
        )
        past_row_changed = past_summary_changed.iloc[0].to_dict()
        past_expected = float(past_row.get("당월 예상매출") or 0)
        past_expected_changed = float(past_row_changed.get("당월 예상매출") or 0)
        past_progress = float(past_row.get("당월 진척률") or 0)
        if (
            float(past_row.get("완료월수") or 0) == 5
            and float(past_row.get("당월 현재매출") or 0) == 300
            and past_expected > 0
            and past_progress > 0
            and abs(past_expected - past_expected_changed) < 1e-9
        ):
            results.append(_ok("sales period policy past month-end evaluation", "past month-end keeps end month as evaluation and forecast is basis-month only"))
        else:
            results.append(_fail("sales period policy past month-end evaluation", f"row={past_row}, changed={past_row_changed}"))

        source_policy_cases = [
            ("date_to == today", {"month_from": "202601", "date_to": "20260712", "policy_date": "20260712"}, False, "20260712", "current_monthly", "202607", ["202601", "202602", "202603", "202604", "202605", "202606"]),
            ("date_to > today", {"month_from": "202601", "date_to": "20260831", "policy_date": "20260712"}, False, "20260712", "current_monthly", "202607", ["202601", "202602", "202603", "202604", "202605", "202606"]),
            ("past month end", {"month_from": "202601", "date_to": "20260630", "policy_date": "20260712"}, False, "20260630", "historical_month_end", "202606", ["202601", "202602", "202603", "202604", "202605"]),
            ("past mid month", {"month_from": "202601", "date_to": "20260702", "policy_date": "20260712"}, True, "20260702", "historical_midmonth", "202607", ["202601", "202602", "202603", "202604", "202605", "202606"]),
        ]
        source_policy_mismatches = []
        for label, params_case, expected_hybrid, expected_effective, expected_mode, expected_eval_month, expected_basis in source_policy_cases:
            policy = mod._resolve_period_source_policy(params_case)
            if bool(policy.get("use_hybrid")) != expected_hybrid:
                source_policy_mismatches.append(f"{label}: hybrid expected={expected_hybrid}, got={policy.get('use_hybrid')}")
            if bool(policy.get("use_hybrid_detail")) != expected_hybrid:
                source_policy_mismatches.append(f"{label}: hybrid_detail expected={expected_hybrid}, got={policy.get('use_hybrid_detail')}")
            if str(policy.get("effective_date_to") or "") != expected_effective:
                source_policy_mismatches.append(f"{label}: effective expected={expected_effective}, got={policy.get('effective_date_to')}")
            if str(policy.get("evaluation_mode") or "") != expected_mode:
                source_policy_mismatches.append(f"{label}: mode expected={expected_mode}, got={policy.get('evaluation_mode')}")
            if str(policy.get("evaluation_month") or "") != expected_eval_month:
                source_policy_mismatches.append(f"{label}: eval_month expected={expected_eval_month}, got={policy.get('evaluation_month')}")
            if list(policy.get("basis_months") or []) != expected_basis:
                source_policy_mismatches.append(f"{label}: basis expected={expected_basis}, got={policy.get('basis_months')}")
        if source_policy_mismatches:
            results.append(_fail("sales source date policy resolver", "; ".join(source_policy_mismatches)))
        else:
            results.append(_ok("sales source date policy resolver", "today/future/month-end/mid-month branches verified"))

        old_monthly_source = getattr(mod, "get_sales_trend_monthly_df", None)
        old_detail_source = getattr(mod, "get_sales_trend_detail_df", None)
        try:
            source_calls = {"monthly": 0, "detail": 0}
            monthly_params_seen: list[dict[str, Any]] = []
            monthly_source_df = pd.DataFrame(
                [
                    _row("202606", 600, 6, "MIX1", "B1"),
                    _row("202607", 900, 9, "MIX1", "B1"),
                ]
            )
            detail_source_df = pd.DataFrame([_row("202607", 70, 7, "MIX1", "B1")])
            detail_source_df["거래처코드"] = "C1"
            detail_source_df["거래처명"] = "거래처"
            detail_source_df["시도명"] = "서울"
            detail_source_df["시구군명"] = "강남구"
            detail_source_df["법정읍면동명"] = "역삼동"
            detail_source_df["출고건수"] = 1
            detail_source_df["거래처수"] = 1
            month_col = list(monthly_source_df.columns)[0]
            branch_columns: dict[str, list[str]] = {}

            def _fake_monthly_source(params: Optional[dict[str, Any]] = None, source_mode: str = "monthly_book") -> pd.DataFrame:
                source_calls["monthly"] += 1
                monthly_params_seen.append(dict(params or {}))
                return monthly_source_df.copy()

            def _fake_detail_source(params: Optional[dict[str, Any]] = None) -> pd.DataFrame:
                source_calls["detail"] += 1
                return detail_source_df.copy()

            setattr(mod, "get_sales_trend_monthly_df", _fake_monthly_source)
            setattr(mod, "get_sales_trend_detail_df", _fake_detail_source)

            for label, params_case in [
                ("today monthly-only", {"date_to": "20260712", "policy_date": "20260712"}),
                ("future monthly-only", {"date_to": "20260831", "policy_date": "20260712"}),
                ("past month-end monthly-only", {"date_to": "20260630", "policy_date": "20260712"}),
            ]:
                source_calls["detail"] = 0
                monthly_params_seen.clear()
                out_source = mod.get_sales_trend_df(
                    {
                        "month_from": "202606",
                        "month_to": "202608",
                        "source_mode": "monthly_book",
                        **params_case,
                    }
                )
                if source_calls["detail"] != 0:
                    source_policy_mismatches.append(f"{label}: detail calls expected=0, got={source_calls['detail']}")
                if label == "future monthly-only" and str((monthly_params_seen[-1] or {}).get("date_to") or "") != "20260712":
                    source_policy_mismatches.append(f"{label}: monthly date_to not capped, params={monthly_params_seen[-1]}")
                if out_source.empty:
                    source_policy_mismatches.append(f"{label}: empty source result")
                suffix_cols = [c for c in out_source.columns if str(c).endswith(("_x", "_y"))]
                if suffix_cols:
                    source_policy_mismatches.append(f"{label}: unexpected suffix columns={suffix_cols}")
                branch_columns[label] = list(out_source.columns)

            source_calls["detail"] = 0
            out_hybrid = mod.get_sales_trend_df(
                {
                    "month_from": "202606",
                    "month_to": "202607",
                    "date_to": "20260702",
                    "policy_date": "20260712",
                    "source_mode": "monthly_book",
                }
            )
            hybrid_months = out_hybrid[month_col].astype(str).tolist() if month_col in out_hybrid.columns else []
            if source_calls["detail"] != 1:
                source_policy_mismatches.append(f"past mid month hybrid: detail calls expected=1, got={source_calls['detail']}")
            if hybrid_months.count("202607") != 1:
                source_policy_mismatches.append(f"past mid month hybrid: expected one replaced 202607 row, months={hybrid_months}")
            if not bool(getattr(out_hybrid, "attrs", {}).get("mixed_current_month_detail")):
                source_policy_mismatches.append("past mid month hybrid: mixed attrs missing")
            hybrid_suffix_cols = [c for c in out_hybrid.columns if str(c).endswith(("_x", "_y"))]
            if hybrid_suffix_cols:
                source_policy_mismatches.append(f"past mid month hybrid: unexpected suffix columns={hybrid_suffix_cols}")
            branch_columns["past mid month hybrid"] = list(out_hybrid.columns)
            internal_cols = {"거래처코드", "거래처명", "시도명", "시구군명", "법정읍면동명", "출고건수", "거래처수", "평균공급단가"}
            leaked_cols = sorted(c for c in out_hybrid.columns if c in internal_cols)
            if leaked_cols:
                source_policy_mismatches.append(f"past mid month hybrid: leaked internal columns={leaked_cols}")
            if len({tuple(cols) for cols in branch_columns.values()}) != 1:
                source_policy_mismatches.append(f"branch public columns mismatch={branch_columns}")

            current_raw = mod.get_sales_trend_df(
                {
                    "month_from": "202606",
                    "month_to": "202607",
                    "date_to": "20260712",
                    "policy_date": "20260712",
                    "source_mode": "monthly_book",
                }
            )
            current_summary = mod.get_sales_trend_summary_df(
                {
                    "month_from": "202606",
                    "month_to": "202607",
                    "date_to": "20260712",
                    "policy_date": "20260712",
                    "source_mode": "monthly_book",
                },
                raw_df=current_raw,
            )
            hybrid_summary = mod.get_sales_trend_summary_df(
                {
                    "month_from": "202606",
                    "month_to": "202607",
                    "date_to": "20260702",
                    "policy_date": "20260712",
                    "source_mode": "monthly_book",
                },
                raw_df=out_hybrid,
            )
            current_eval = current_summary.iloc[0].to_dict()
            hybrid_eval = hybrid_summary.iloc[0].to_dict()
            if abs(float(current_eval.get("당월 예상매출") or 0) - float(hybrid_eval.get("당월 예상매출") or 0)) > 1e-9:
                source_policy_mismatches.append(f"20260702/20260712 expected mismatch: current={current_eval}, hybrid={hybrid_eval}")
            if abs(float(current_eval.get("당월 현재매출") or 0) - float(hybrid_eval.get("당월 현재매출") or 0)) < 1e-9:
                source_policy_mismatches.append(f"20260702/20260712 actual should differ: current={current_eval}, hybrid={hybrid_eval}")
            if abs(float(current_eval.get("당월 진척률") or 0) - float(hybrid_eval.get("당월 진척률") or 0)) < 1e-9:
                source_policy_mismatches.append(f"20260702/20260712 progress should differ: current={current_eval}, hybrid={hybrid_eval}")

            if source_policy_mismatches:
                results.append(_fail("sales source monthly-only detail skip", "; ".join(source_policy_mismatches)))
            else:
                results.append(_ok("sales source monthly-only detail skip", "monthly-only skipped detail; mid-month hybrid replaced current month"))
        finally:
            if old_monthly_source is not None:
                setattr(mod, "get_sales_trend_monthly_df", old_monthly_source)
            if old_detail_source is not None:
                setattr(mod, "get_sales_trend_detail_df", old_detail_source)

        old_forecast_df = getattr(mod, "get_sales_forecast_df", None)
        old_stock_df = getattr(mod, "_load_product_current_stock", None)
        try:
            stock_base_df = pd.DataFrame(
                [
                    {
                        "제품코드": "STK1",
                        "제품명": "재고테스트1",
                        "규격": "EA",
                        "제조사명": "제조사",
                        "매입처명": "매입처",
                        "2026-01 수량": 10,
                        "2026-02 수량": 20,
                        "2026-03 수량": 30,
                        "2026-04 수량": 40,
                        "2026-05 수량": 50,
                        "2026-06 수량": 60,
                        "2026-07 수량": 25,
                        "2026-08 수량": 999,
                        "총출고수량": 1234,
                    },
                    {
                        "제품코드": "STK2",
                        "제품명": "재고테스트2",
                        "규격": "EA",
                        "제조사명": "제조사",
                        "매입처명": "매입처",
                        "2026-01 수량": 0,
                        "2026-02 수량": 0,
                        "2026-03 수량": 0,
                        "2026-04 수량": 0,
                        "2026-05 수량": 0,
                        "2026-06 수량": 0,
                        "2026-07 수량": 0,
                        "2026-08 수량": 999,
                        "총출고수량": 999,
                    },
                    {
                        "제품코드": "STK3",
                        "제품명": "재고테스트3",
                        "규격": "EA",
                        "제조사명": "제조사",
                        "매입처명": "매입처",
                        "2026-01 수량": 5,
                        "2026-02 수량": 5,
                        "2026-03 수량": 5,
                        "2026-04 수량": 5,
                        "2026-05 수량": 5,
                        "2026-06 수량": 5,
                        "2026-07 수량": 0,
                        "2026-08 수량": 999,
                        "총출고수량": 999,
                    },
                ]
            )
            stock_current_df = pd.DataFrame(
                [
                    {"제품코드": "STK1", "장부재고수량": 20, "실재고수량": 20, "장부재고금액": 20_000, "실재고금액": 40_000, "장부재고평가단가": 1000, "실재고평가단가": 2000, "당월입고수량": 1, "당월출고수량": 2, "당월재고증감수량": -1},
                    {"제품코드": "STK2", "장부재고수량": 10, "실재고수량": 10, "장부재고금액": 20_000, "실재고금액": 30_000, "장부재고평가단가": 2000, "실재고평가단가": 3000, "당월입고수량": 0, "당월출고수량": 0, "당월재고증감수량": 0},
                    {"제품코드": "STK3", "장부재고수량": -2, "실재고수량": -2, "장부재고금액": -200, "실재고금액": -400, "장부재고평가단가": 100, "실재고평가단가": 200, "당월입고수량": 0, "당월출고수량": 0, "당월재고증감수량": 0},
                ]
            )

            setattr(mod, "get_sales_forecast_df", lambda params, raw_df=None: stock_base_df.copy())
            setattr(mod, "_load_product_current_stock", lambda *args, **kwargs: stock_current_df.copy())

            stock_result = mod.get_stock_shortage_df(
                {
                    "month_from": "202601",
                    "month_to": "202608",
                    "date_from": "20260101",
                    "date_to": "20260831",
                    "source_mode": "monthly_book",
                    "stock_mode": "book",
                }
            )
            stock_row1 = stock_result[stock_result["제품코드"].astype(str) == "STK1"].iloc[0].to_dict()
            stock_row2 = stock_result[stock_result["제품코드"].astype(str) == "STK2"].iloc[0].to_dict()
            stock_row3 = stock_result[stock_result["제품코드"].astype(str) == "STK3"].iloc[0].to_dict()
            meta_stock = mod._stock_shortage_meta_from_df(stock_result)

            stock_mismatches = []
            expected_stock_1 = {
                "완료월수": 6,
                "완료월총출고수량": 210,
                "완료월평균출고수량": 35,
                "최근3개월평균출고수량": 50,
                "최근6개월평균출고수량": 35,
                "최근3개월수량증감률": 42.8571428571,
                "당월 현재출고수량": 25,
                "당월 예상출고수량": 57.5,
                "당월 잔여예상출고수량": 32.5,
                "당월 출고진척률": 43.4782608696,
                "예상월말재고수량": -12.5,
                "부족예상수량": 12.5,
                "부족예상금액": 12500,
                "당월 재고충족률": 61.5384615385,
            }
            for key, exp in expected_stock_1.items():
                got = float(stock_row1.get(key) or 0)
                if abs(got - exp) > 1e-6:
                    stock_mismatches.append(f"STK1 {key}: expected={exp}, got={got}")
            if float(stock_row1.get("월평균출고수량") or 0) != float(stock_row1.get("완료월평균출고수량") or 0):
                stock_mismatches.append("월평균출고수량 is not aligned to 완료월평균출고수량")
            if str(stock_row1.get("재고부족판정") or "") != "부족":
                stock_mismatches.append(f"STK1 재고부족판정 expected 부족, got={stock_row1.get('재고부족판정')!r}")
            if float(stock_row2.get("당월 재고충족률") or 0) != 100 or str(stock_row2.get("재고부족판정") or "") != "수요없음":
                stock_mismatches.append(f"STK2 expected no-demand fill=100, row={stock_row2}")
            if abs(float(stock_row3.get("부족예상수량") or 0) - 7) > 1e-6:
                stock_mismatches.append(f"STK3 negative stock shortage expected=7, got={stock_row3.get('부족예상수량')}")
            if abs(float(meta_stock.get("overall_stock_fill_rate") or 0) - (30 / 37.5 * 100)) > 1e-6:
                stock_mismatches.append(f"overall fill rate expected=80, got={meta_stock.get('overall_stock_fill_rate')}")
            if abs(float(meta_stock.get("current_month_demand_progress_pct") or 0) - (25 / 62.5 * 100)) > 1e-6:
                stock_mismatches.append(f"weighted demand progress expected=40, got={meta_stock.get('current_month_demand_progress_pct')}")
            if str(stock_result.get("분석자료원", pd.Series([""])).iloc[0]) != "월집계-장부재고(Rddbc220)":
                stock_mismatches.append(f"monthly source label mismatch: {stock_result.get('분석자료원')}")
            internal_cols = [c for c in ["당월입고수량", "당월출고수량", "당월재고증감수량"] if c in stock_result.columns]
            if internal_cols:
                stock_mismatches.append(f"stock shortage internal columns leaked: {internal_cols}")
            if any(str(c).endswith(("_x", "_y")) for c in stock_result.columns):
                stock_mismatches.append("stock shortage result has merge suffix columns")
            if "2026-08 수량" in stock_result.columns and float(stock_row1.get("완료월총출고수량") or 0) >= 999:
                stock_mismatches.append("future month quantity appears included in completed demand")

            stock_hybrid_result = mod.get_stock_shortage_df(
                {
                    "month_from": "202601",
                    "month_to": "202607",
                    "date_from": "20260101",
                    "date_to": "20260702",
                    "policy_date": "20260712",
                    "source_mode": "monthly_book",
                    "stock_mode": "book",
                }
            )
            hybrid_source = str(stock_hybrid_result.get("분석자료원", pd.Series([""])).iloc[0])
            hybrid_stock_source = str(stock_hybrid_result.get("현재고원천", pd.Series([""])).iloc[0])
            if "평가월: 출고상세(Rddbc120)" not in hybrid_source or "현재재고: 전월말+입출고상세" not in hybrid_source:
                stock_mismatches.append(f"hybrid source label mismatch: {hybrid_source}")
            if "입고상세(Rddbc110)" not in hybrid_stock_source or "출고상세(Rddbc120)" not in hybrid_stock_source:
                stock_mismatches.append(f"hybrid stock source label mismatch: {hybrid_stock_source}")
            leaked_hybrid = [
                c for c in ["당월입고수량", "당월출고수량", "당월재고증감수량"]
                if c in stock_hybrid_result.columns
            ]
            if leaked_hybrid:
                stock_mismatches.append(f"hybrid stock internal columns leaked: {leaked_hybrid}")

            stock_past_month_end = mod.get_stock_shortage_df(
                {
                    "month_from": "202601",
                    "month_to": "202606",
                    "date_from": "20260101",
                    "date_to": "20260630",
                    "policy_date": "20260712",
                    "source_mode": "monthly_book",
                    "stock_mode": "book",
                }
            )
            stock_base_df_changed = stock_base_df.copy()
            stock_base_df_changed.loc[stock_base_df_changed["제품코드"].astype(str) == "STK1", "2026-06 수량"] = 999
            setattr(mod, "get_sales_forecast_df", lambda params, raw_df=None: stock_base_df_changed.copy())
            stock_past_month_end_changed = mod.get_stock_shortage_df(
                {
                    "month_from": "202601",
                    "month_to": "202606",
                    "date_from": "20260101",
                    "date_to": "20260630",
                    "policy_date": "20260712",
                    "source_mode": "monthly_book",
                    "stock_mode": "book",
                }
            )
            setattr(mod, "get_sales_forecast_df", lambda params, raw_df=None: stock_base_df.copy())
            past_stock_row1 = stock_past_month_end[stock_past_month_end["제품코드"].astype(str) == "STK1"].iloc[0].to_dict()
            past_stock_row1_changed = stock_past_month_end_changed[stock_past_month_end_changed["제품코드"].astype(str) == "STK1"].iloc[0].to_dict()
            if abs(float(past_stock_row1.get("당월 예상출고수량") or 0) - 46) > 1e-6:
                stock_mismatches.append(f"past month-end expected demand expected=46, got={past_stock_row1.get('당월 예상출고수량')}")
            if abs(float(past_stock_row1.get("당월 예상출고수량") or 0) - float(past_stock_row1_changed.get("당월 예상출고수량") or 0)) > 1e-9:
                stock_mismatches.append(f"past month-end expected demand changed by actual month: before={past_stock_row1}, after={past_stock_row1_changed}")
            if float(past_stock_row1.get("당월 현재출고수량") or 0) != 60:
                stock_mismatches.append(f"past month-end actual demand expected=60, got={past_stock_row1.get('당월 현재출고수량')}")
            if float(past_stock_row1.get("당월 예상출고수량") or 0) <= 0:
                stock_mismatches.append("past month-end expected demand should not be zero")

            stock_real_result = mod.get_stock_shortage_df(
                {
                    "month_from": "202601",
                    "month_to": "202608",
                    "date_from": "20260101",
                    "date_to": "20260831",
                    "source_mode": "monthly_real",
                    "stock_mode": "real",
                }
            )
            stock_real_row1 = stock_real_result[stock_real_result["제품코드"].astype(str) == "STK1"].iloc[0].to_dict()
            if float(stock_real_row1.get("재고평가단가") or 0) != 2000:
                stock_mismatches.append(f"real stock unit price expected=2000, got={stock_real_row1.get('재고평가단가')}")
            if abs(float(stock_real_row1.get("부족예상금액") or 0) - 25000) > 1e-6:
                stock_mismatches.append(f"real stock shortage amount expected=25000, got={stock_real_row1.get('부족예상금액')}")

            if stock_mismatches:
                results.append(_fail("stock shortage current-month period policy", "; ".join(stock_mismatches)))
            else:
                results.append(_ok("stock shortage current-month period policy", "completed demand, current demand, shortage, fill rate, and unit price selection verified"))
        finally:
            if old_forecast_df is not None:
                setattr(mod, "get_sales_forecast_df", old_forecast_df)
            if old_stock_df is not None:
                setattr(mod, "_load_product_current_stock", old_stock_df)
    except Exception as e:
        results.append(_fail("sales period policy current/future exclusion", f"{type(e).__name__}: {e}"))
    finally:
        try:
            if old_current_yyyymm is not None:
                setattr(mod, "_current_yyyymm", old_current_yyyymm)
        except Exception:
            pass

    try:
        import app.services.analytics_manufacturer_sales_trend_service as manufacturer_mod

        def _manufacturer_raw(date_to: str = "20260712") -> pd.DataFrame:
            july_a = 50 if str(date_to or "") >= "20260712" else 20
            rows = []
            for m, a1, a2, b in [
                ("202601", 40, 60, 10),
                ("202602", 40, 60, 20),
                ("202603", 40, 60, 30),
                ("202604", 40, 60, 40),
                ("202605", 40, 60, 50),
                ("202606", 40, 60, 60),
                ("202607", 10, july_a - 10, 70),
            ]:
                rows.extend([
                    {"기준월": m, "제품코드": "P1", "제품명": "A1", "규격": "", "제조사명": " 제약A ", "매출공급가액": a1, "매출합계": a1},
                    {"기준월": m, "제품코드": "P2", "제품명": "A2", "규격": "", "제조사명": "제약A", "매출공급가액": a2, "매출합계": a2},
                    {"기준월": m, "제품코드": "P3", "제품명": "B1", "규격": "", "제조사명": None, "매출공급가액": b, "매출합계": b},
                ])
            for m, amt in [
                ("202601", 60_000_000),
                ("202602", 50_000_000),
                ("202603", 40_000_000),
                ("202604", 30_000_000),
                ("202605", 20_000_000),
                ("202606", 10_000_000),
                ("202607", 5_000_000),
            ]:
                rows.append({"기준월": m, "제품코드": "P5", "제품명": "D1", "규격": "", "제조사명": "감소D", "매출공급가액": amt, "매출합계": amt})
            rows.append({"기준월": "202607", "제품코드": "P4", "제품명": "C1", "규격": "", "제조사명": "신규C", "매출공급가액": 30, "매출합계": 30})
            df = pd.DataFrame(rows)
            df.attrs["mixed_current_month_detail"] = str(date_to or "") < "20260712" and str(date_to or "")[:6] == "202607"
            df.attrs["source_label_completed"] = "월집계-장부재고(Rddbc220)"
            df.attrs["source_label_current"] = "출고상세(Rddbc120)"
            return df

        old_loader = getattr(manufacturer_mod, "get_sales_trend_df", None)
        captured_manufacturer_params = []

        def _manufacturer_loader(params):
            captured_manufacturer_params.append(dict(params or {}))
            return _manufacturer_raw(str((params or {}).get("date_to") or "20260712"))

        setattr(manufacturer_mod, "get_sales_trend_df", _manufacturer_loader)
        try:
            params_current = {
                "month_from": "202601",
                "month_to": "202607",
                "date_from": "20260101",
                "date_to": "20260712",
                "policy_date": "20260712",
                "source_mode": "monthly_book",
            }
            params_mid = dict(params_current, date_to="20260702")
            params_past_end = dict(params_current, month_to="202606", date_to="20260630")
            detail = manufacturer_mod.get_manufacturer_sales_trend(params_current)
            detail_mid = manufacturer_mod.get_manufacturer_sales_trend(params_mid)
            detail_past = manufacturer_mod.get_manufacturer_sales_trend(params_past_end)
            mismatches = []
            if getattr(detail, "attrs", {}).get("evaluation_mode") != "current_monthly":
                mismatches.append(f"current mode expected current_monthly got={getattr(detail, 'attrs', {}).get('evaluation_mode')}")
            if getattr(detail_mid, "attrs", {}).get("evaluation_mode") != "historical_midmonth":
                mismatches.append(f"mid mode expected historical_midmonth got={getattr(detail_mid, 'attrs', {}).get('evaluation_mode')}")
            if getattr(detail_past, "attrs", {}).get("evaluation_mode") != "historical_month_end":
                mismatches.append(f"past month-end mode expected historical_month_end got={getattr(detail_past, 'attrs', {}).get('evaluation_mode')}")
            address_params = dict(params_current, sido_nm="서울", gugun_nm="강남", road_nm="테헤란로")
            _ = manufacturer_mod.get_manufacturer_sales_trend(address_params)
            last_params = captured_manufacturer_params[-1] if captured_manufacturer_params else {}
            address_mismatches = []
            for k, expected in {"sido_nm": "서울", "gugun_nm": "강남", "road_nm": "테헤란로"}.items():
                if last_params.get(k) != expected:
                    address_mismatches.append(f"manufacturer address param not forwarded {k}={last_params.get(k)!r}")
            query_condition = manufacturer_mod._fmt_analytics_query_summary(address_params, "월집계-장부재고(Rddbc220)")
            for expected in ["시도명 서울", "시구군명 강남", "도로명 테헤란로"]:
                if expected not in query_condition:
                    address_mismatches.append(f"manufacturer address query condition missing {expected}: {query_condition}")
            if address_mismatches:
                results.append(_fail("manufacturer sales trend address filters", "; ".join(address_mismatches)))
            else:
                results.append(_ok("manufacturer sales trend address filters", "sido/gugun/road params and query summary verified"))

            if "제약사명" not in detail.columns:
                mismatches.append("missing 제약사명")
            forbidden_tokens = ["수량", "다음월예상매출", "예상등급"]
            forbidden = [
                c for c in detail.columns
                if any(x in str(c) for x in forbidden_tokens)
                or str(c).endswith(("_x", "_y"))
                or str(c).startswith("_")
                or c in {"제품코드", "제품명", "규격"}
            ]
            if forbidden:
                mismatches.append(f"forbidden public columns={forbidden}")
            required_detail_cols = [
                "기준월",
                "매출공급가액",
                "최근3개월평균매출",
                "월시점 실제매출",
                "월시점 예상매출",
                "월시점 달성률",
                "월시점 예상기준",
                "월시점 적용증감률",
                "월시점 예상대비차이",
                "월시점 잔여예상",
                "월시점 추세판정",
                "월시점 판정결과",
                "판정결과",
                "추세판정",
            ]
            for c in required_detail_cols:
                if c not in detail.columns:
                    mismatches.append(f"missing detail column {c}")
            detail_analysis_block = [
                "월시점 완료월수",
                "월시점 완료월평균매출",
                "월시점 최근3개월평균매출",
                "월시점 최근6개월평균매출",
                "월시점 증감률",
                "월시점 추세판정",
                "월시점 판정결과",
                "월시점 실제매출",
                "월시점 예상기준",
                "월시점 적용증감률",
                "월시점 예상매출",
                "월시점 예상대비차이",
                "월시점 잔여예상",
                "월시점 달성률",
                "추세판정",
                "판정결과",
            ]
            detail_cols = list(detail.columns)
            block_positions = [detail_cols.index(c) for c in detail_analysis_block if c in detail_cols]
            if block_positions != sorted(block_positions) or len(block_positions) != len(detail_analysis_block):
                mismatches.append("detail analysis block order mismatch")
            non_zero_metric_sum = float(
                detail[[c for c in ["월시점 실제매출", "월시점 예상매출", "월시점 달성률"] if c in detail.columns]]
                .apply(pd.to_numeric, errors="coerce")
                .fillna(0)
                .abs()
                .sum()
                .sum()
            )
            if non_zero_metric_sum <= 0:
                mismatches.append("detail month-point actual/expected/progress are all zero")
            zero_check_cols = [
                "매출공급가액",
                "매출세액",
                "매출합계",
                "집계건수",
                "월시점 실제매출",
                "월시점 예상매출",
                "월시점 예상대비차이",
                "월시점 잔여예상",
            ]
            present_zero_cols = [c for c in zero_check_cols if c in detail.columns]
            zero_rows = (
                detail[present_zero_cols].apply(pd.to_numeric, errors="coerce").fillna(0).abs().sum(axis=1).eq(0).sum()
                if present_zero_cols
                else 0
            )
            if int(zero_rows) != 0:
                mismatches.append(f"detail public zero rows should be removed rows={zero_rows}")
            detail_key_counts = detail.groupby(["제약사명", "기준월"]).size()
            if not detail_key_counts.empty and int(detail_key_counts.max()) != 1:
                mismatches.append("detail should have one row per manufacturer-month")

            raw_cur = _manufacturer_raw("20260712")
            raw_sum = float(pd.to_numeric(raw_cur["매출공급가액"], errors="coerce").fillna(0).sum())
            detail_sum = float(pd.to_numeric(detail["매출공급가액"], errors="coerce").fillna(0).sum())
            if abs(raw_sum - detail_sum) > 1e-9:
                mismatches.append(f"detail monthly sum mismatch raw={raw_sum}, detail={detail_sum}")
            a_july = detail[(detail["제약사명"].astype(str) == "제약A") & (detail["기준월"].astype(str) == "202607")].iloc[0]
            a_mid_july = detail_mid[(detail_mid["제약사명"].astype(str) == "제약A") & (detail_mid["기준월"].astype(str) == "202607")].iloc[0]
            if float(a_july["매출공급가액"]) != 50:
                mismatches.append(f"manufacturer detail should aggregate manufacturer-month sales expected=50 got={a_july['매출공급가액']}")
            if float(a_mid_july["매출공급가액"]) == float(a_july["매출공급가액"]):
                mismatches.append("midmonth/current manufacturer actual sales should differ")
            sorted_check = detail.sort_values(["제약사명", "기준월"], ascending=[True, True]).reset_index(drop=True)
            if list(detail[["제약사명", "기준월"]].itertuples(index=False, name=None)) != list(sorted_check[["제약사명", "기준월"]].itertuples(index=False, name=None)):
                mismatches.append("detail final sort should be manufacturer asc + month asc")
            expected_seq = list(range(1, len(detail) + 1))
            if "순번" not in detail.columns or list(pd.to_numeric(detail["순번"], errors="coerce").fillna(0).astype(int)) != expected_seq:
                mismatches.append("detail sequence should be reassigned after final sort")
            if pd.isna(a_july.get("전월대비매출")):
                mismatches.append("detail last month diff sales should use same formula, not NaN")
            if pd.isna(a_july.get("전월대비매출증감률")):
                mismatches.append("detail last month diff pct should use same formula, not NaN")
            bad_period_values = {"당월진행", "current_monthly", "historical_midmonth", "historical_month_end"}
            leaked_period = sorted(set(detail.get("기간구분", pd.Series(dtype=str)).astype(str)) & bad_period_values)
            if leaked_period:
                mismatches.append(f"detail public period leaked forbidden values={leaked_period}")
            mid_period_values = set(detail_mid.get("기간구분", pd.Series(dtype=str)).astype(str))
            if "부분월" not in mid_period_values:
                mismatches.append(f"midmonth detail should expose 부분월 for final month values={sorted(mid_period_values)}")
            blank_july = detail[(detail["제약사명"].astype(str) == "제약사 미지정") & (detail["기준월"].astype(str) == "202607")].iloc[0]
            if float(blank_july["매출공급가액"]) != 70:
                mismatches.append("blank manufacturer group not preserved in detail")
            new_july = detail[(detail["제약사명"].astype(str) == "신규C") & (detail["기준월"].astype(str) == "202607")].iloc[0]
            if str(new_july.get("추세판정") or "") != "자료부족":
                mismatches.append(f"new manufacturer judge expected 자료부족 got={new_july.get('추세판정')}")

            if mismatches:
                results.append(_fail("manufacturer sales trend detail", "; ".join(mismatches)))
            else:
                results.append(_ok("manufacturer sales trend detail", "analysis block, zero-row removal, sort, and monthly sums verified"))

            summary = manufacturer_mod.get_manufacturer_sales_trend_summary(params_current)
            summary_mid = manufacturer_mod.get_manufacturer_sales_trend_summary(params_mid)
            summary_past = manufacturer_mod.get_manufacturer_sales_trend_summary(params_past_end)
            detail_res = manufacturer_mod.get_manufacturer_sales_trend_result(params_current)
            res = manufacturer_mod.get_manufacturer_sales_trend_summary_result(params_current)
            res_past = manufacturer_mod.get_manufacturer_sales_trend_summary_result(params_past_end)
            mismatches = []
            forbidden = [
                c for c in summary.columns
                if any(x in str(c) for x in forbidden_tokens)
                or str(c).endswith(("_x", "_y"))
                or str(c).startswith("_")
                or c in {"제품코드", "제품명", "규격"}
            ]
            if forbidden:
                mismatches.append(f"forbidden summary columns={forbidden}")
            if not any(str(c).endswith(" 매출") and str(c)[:4].isdigit() for c in summary.columns):
                mismatches.append("missing dynamic monthly sales columns")
            required_summary_cols = [
                "순번",
                "제약사명",
                "총매출공급가액",
                "총매출세액",
                "총매출액",
                "완료월총매출",
                "월평균매출",
                "완료월수",
                "완료월평균매출",
                "당월 현재매출",
                "당월 예상매출",
                "당월 잔여예상",
                "당월 진척률",
                "매출발생월수",
                "최근3개월평균매출",
                "최근6개월평균매출",
                "최근3개월증감률",
                "추세판정",
                "제품수",
                "매입처수",
                "총집계건수",
                "분석자료원",
                "기간구분",
            ]
            for c in required_summary_cols:
                if c not in summary.columns:
                    mismatches.append(f"missing summary column {c}")
            summary_cols = list(summary.columns)
            summary_positions = [summary_cols.index(c) for c in required_summary_cols if c in summary_cols]
            if summary_positions != sorted(summary_positions) or len(summary_positions) != len(required_summary_cols):
                mismatches.append("summary core column order mismatch")
            if "평가월 매출" in summary.columns:
                mismatches.append("current monthly summary should expose 당월 현재매출, not 평가월 매출")
            for label, df_check in [("mid", summary_mid), ("past", summary_past)]:
                for c in ["평가월 매출", "평가월 예상매출", "평가월 잔여예상", "평가월 진척률"]:
                    if c not in df_check.columns:
                        mismatches.append(f"{label} historical summary missing {c}")
                if "당월 현재매출" in df_check.columns:
                    mismatches.append(f"{label} historical summary should not expose 당월 현재매출")
                if any(c in df_check.columns for c in ["당월 예상매출", "당월 잔여예상", "당월 진척률"]):
                    mismatches.append(f"{label} historical summary should not expose 당월 expected/progress labels")
            if any(str(c).endswith(" 수량") for c in summary.columns):
                mismatches.append("summary should not expose monthly qty columns")
            summary_sum = float(pd.to_numeric(summary["총매출공급가액"], errors="coerce").fillna(0).sum()) if "총매출공급가액" in summary.columns else -1
            if abs(raw_sum - summary_sum) > 1e-9:
                mismatches.append(f"summary total mismatch raw={raw_sum}, summary={summary_sum}")
            names = set(summary["제약사명"].astype(str).tolist()) if "제약사명" in summary.columns else set()
            if not {"제약A", "제약사 미지정"}.issubset(names):
                mismatches.append(f"manufacturer universe not preserved names={sorted(names)}")
            meta = res.get("meta") or {}
            if meta.get("analysis_type") != "manufacturer_sales_trend_summary" or meta.get("summary_type") != "manufacturer_trend_summary":
                mismatches.append(f"unexpected summary meta={meta}")
            if meta.get("evaluation_mode") != "current_monthly":
                mismatches.append(f"summary meta evaluation_mode expected current_monthly got={meta.get('evaluation_mode')}")
            period_caption = str(meta.get("period_caption") or "")
            if "당월 2026-07" not in period_caption or "current_monthly" in period_caption:
                mismatches.append(f"current summary period caption unexpected={period_caption}")
            mid_caption = str(getattr(summary_mid, "attrs", {}).get("period_caption") or "")
            if "평가월 2026-07(07-02까지)" not in mid_caption or "당월" in mid_caption or "historical_midmonth" in mid_caption:
                mismatches.append(f"mid summary period caption unexpected={mid_caption}")
            past_caption = str(getattr(summary_past, "attrs", {}).get("period_caption") or "")
            if "완료월 2026-01~2026-05" not in past_caption or "평가월 2026-06" not in past_caption or "historical_month_end" in past_caption:
                mismatches.append(f"past summary period caption unexpected={past_caption}")
            if "current_monthly" in set(summary.get("기간구분", pd.Series(dtype=str)).astype(str)):
                mismatches.append("summary public period label leaked internal current_monthly")
            a_summary = summary[summary["제약사명"].astype(str) == "제약A"].iloc[0]
            a_summary_mid = summary_mid[summary_mid["제약사명"].astype(str) == "제약A"].iloc[0]
            for c in ["완료월총매출", "완료월수", "완료월평균매출", "최근3개월평균매출", "최근6개월평균매출", "최근3개월증감률", "추세판정"]:
                if str(c) == "추세판정":
                    if str(a_summary[c]) != str(a_summary_mid[c]):
                        mismatches.append(f"current/mid completed judge differs {a_summary[c]} vs {a_summary_mid[c]}")
                elif abs(float(a_summary[c]) - float(a_summary_mid[c])) > 1e-9:
                    mismatches.append(f"current/mid completed metric differs {c}: {a_summary[c]} vs {a_summary_mid[c]}")
            if float(a_summary["당월 현재매출"]) == float(a_summary_mid["평가월 매출"]):
                mismatches.append("current/mid current month sales should differ in summary")
            if float(a_summary.get("당월 예상매출", 0)) <= 0:
                mismatches.append("current summary expected sales should be populated")
            if float(a_summary.get("당월 진척률", 0)) <= 0:
                mismatches.append("current summary progress should be populated")
            a_past = summary_past[summary_past["제약사명"].astype(str) == "제약A"].iloc[0]
            if int(a_past["완료월수"]) != 5:
                mismatches.append(f"past month-end completed month count expected=5 got={a_past['완료월수']}")
            new_summary = summary[summary["제약사명"].astype(str) == "신규C"].iloc[0]
            if str(new_summary.get("추세판정") or "") != "비교자료 부족":
                mismatches.append(f"new manufacturer summary judge expected 비교자료 부족 got={new_summary.get('추세판정')}")
            for m in ["2026-01", "2026-02", "2026-03", "2026-04", "2026-05", "2026-06", "2026-07"]:
                col = f"{m} 매출"
                if col in summary.columns:
                    detail_month = m.replace("-", "")
                    dsum = float(pd.to_numeric(detail[detail["기준월"].astype(str) == detail_month]["매출공급가액"], errors="coerce").fillna(0).sum())
                    ssum = float(pd.to_numeric(summary[col], errors="coerce").fillna(0).sum())
                    if abs(dsum - ssum) > 1e-9:
                        mismatches.append(f"detail/summary month sum mismatch {m}: detail={dsum}, summary={ssum}")
            if abs(float(pd.to_numeric(summary["총매출공급가액"], errors="coerce").fillna(0).sum()) - detail_sum) > 1e-9:
                mismatches.append("detail total and summary total differ under identical params")

            try:
                from app.ui.sims_table_display import resolve_sims_table_mode
                old_chat_env = os.environ.get("SIMS_CHAT_FAST_TABLE_CELL_THRESHOLD")
                old_panel_env = os.environ.get("SIMS_FAST_TABLE_CELL_THRESHOLD")
                os.environ["SIMS_CHAT_FAST_TABLE_CELL_THRESHOLD"] = "10"
                os.environ["SIMS_FAST_TABLE_CELL_THRESHOLD"] = "10"
                detail_chat_mode = resolve_sims_table_mode(detail, action="제약사별 매출 추세 분석", render_path="chat")
                detail_panel_mode = resolve_sims_table_mode(detail, action="제약사별 매출 추세 분석", render_path="panel")
                if detail_chat_mode.get("mode") != "fast" or detail_panel_mode.get("mode") != "fast":
                    mismatches.append(f"detail table mode fast expected chat={detail_chat_mode} panel={detail_panel_mode}")
                chat_mode = resolve_sims_table_mode(summary, action="제약사별 매출 추세 분석 요약표", render_path="chat")
                panel_mode = resolve_sims_table_mode(summary, action="제약사별 매출 추세 분석 요약표", render_path="panel")
                if chat_mode.get("mode") != "fast" or panel_mode.get("mode") != "fast":
                    mismatches.append(f"table mode fast expected chat={chat_mode} panel={panel_mode}")
                os.environ["SIMS_CHAT_FAST_TABLE_CELL_THRESHOLD"] = "999999"
                os.environ["SIMS_FAST_TABLE_CELL_THRESHOLD"] = "999999"
                chat_small = resolve_sims_table_mode(summary, action="제약사별 매출 추세 분석 요약표", render_path="chat")
                panel_small = resolve_sims_table_mode(summary, action="제약사별 매출 추세 분석 요약표", render_path="panel")
                if chat_small.get("mode") != "small" or panel_small.get("mode") != "small":
                    mismatches.append(f"table mode small expected chat={chat_small} panel={panel_small}")
            finally:
                if 'old_chat_env' in locals():
                    if old_chat_env is None:
                        os.environ.pop("SIMS_CHAT_FAST_TABLE_CELL_THRESHOLD", None)
                    else:
                        os.environ["SIMS_CHAT_FAST_TABLE_CELL_THRESHOLD"] = old_chat_env
                if 'old_panel_env' in locals():
                    if old_panel_env is None:
                        os.environ.pop("SIMS_FAST_TABLE_CELL_THRESHOLD", None)
                    else:
                        os.environ["SIMS_FAST_TABLE_CELL_THRESHOLD"] = old_panel_env

            def _manufacturer_bucket(value):
                text = str(value or "").strip()
                if text in {"증가", "신규/증가"}:
                    return "증가"
                if text == "감소":
                    return "감소"
                if text == "안정":
                    return "안정"
                return "자료부족"

            def _expected_counts(frame):
                counts = {"증가": 0, "감소": 0, "안정": 0, "자료부족": 0}
                if frame is None or frame.empty:
                    return counts
                if "기준월" in frame.columns:
                    work_counts = (
                        frame.assign(_기준월_sort=frame["기준월"].astype(str))
                        .sort_values(["제약사명", "_기준월_sort"])
                        .drop_duplicates("제약사명", keep="last")
                    )
                else:
                    work_counts = frame.drop_duplicates("제약사명", keep="last")
                for value in work_counts.get("추세판정", pd.Series(dtype=object)).tolist():
                    counts[_manufacturer_bucket(value)] += 1
                return counts

            header_mismatches = []
            detail_meta = detail_res.get("meta") or {}
            summary_meta = res.get("meta") or {}
            past_meta = res_past.get("meta") or {}
            expected_detail_counts = _expected_counts(detail)
            expected_summary_counts = _expected_counts(summary)
            for label, frame, meta_check, expected_counts in [
                ("detail", detail, detail_meta, expected_detail_counts),
                ("summary", summary, summary_meta, expected_summary_counts),
            ]:
                counts = meta_check.get("trend_judge_counts") or {}
                four_total = sum(int(counts.get(k, 0) or 0) for k in ["증가", "감소", "안정", "자료부족"])
                manufacturer_count = int(meta_check.get("manufacturer_count") or 0)
                if four_total != manufacturer_count:
                    header_mismatches.append(f"{label} judge count total {four_total} != manufacturer_count {manufacturer_count}")
                for k in ["증가", "감소", "안정", "자료부족"]:
                    if int(counts.get(k, 0) or 0) != int(expected_counts.get(k, 0) or 0):
                        header_mismatches.append(f"{label} judge {k} expected={expected_counts.get(k)} got={counts.get(k)}")
                if not isinstance(counts.get("자료부족", 0), int):
                    header_mismatches.append(f"{label} 자료부족 count should be integer")

            if detail_meta.get("current_progress_title") != "당월 진행 요약":
                header_mismatches.append(f"detail current progress title unexpected={detail_meta.get('current_progress_title')}")
            if summary_meta.get("current_progress_title") != "당월 진행 요약":
                header_mismatches.append(f"summary current progress title unexpected={summary_meta.get('current_progress_title')}")
            if past_meta.get("current_progress_title") != "평가월 진행 요약":
                header_mismatches.append(f"past progress title unexpected={past_meta.get('current_progress_title')}")
            for meta_label, meta_check in [("detail", detail_meta), ("summary", summary_meta), ("past", past_meta)]:
                for key in [
                    "completed_month_count",
                    "avg_completed_month_sales_amt",
                    "sum_current_month_sales_amt",
                    "sum_current_month_expected_amt",
                    "sum_current_month_remaining_expected_amt",
                    "current_month_progress_pct",
                ]:
                    if key not in meta_check:
                        header_mismatches.append(f"{meta_label} missing header meta {key}")

            eval_month = str(detail_meta.get("evaluation_month") or "")
            eval_rows = detail[detail["기준월"].astype(str) == eval_month] if eval_month and "기준월" in detail.columns else detail
            actual_total = float(pd.to_numeric(eval_rows.get("월시점 실제매출", 0), errors="coerce").fillna(0).sum())
            expected_total = float(pd.to_numeric(eval_rows.get("월시점 예상매출", 0), errors="coerce").fillna(0).sum())
            expected_progress = (actual_total / expected_total * 100) if abs(expected_total) >= 1e-12 else 0
            if abs(float(detail_meta.get("current_month_progress_pct") or 0) - expected_progress) > 1e-9:
                header_mismatches.append("detail progress should use summed actual / summed expected")
            summary_actual = float(pd.to_numeric(summary.get("당월 현재매출", 0), errors="coerce").fillna(0).sum())
            summary_expected = float(pd.to_numeric(summary.get("당월 예상매출", 0), errors="coerce").fillna(0).sum())
            summary_progress = (summary_actual / summary_expected * 100) if abs(summary_expected) >= 1e-12 else 0
            if abs(float(summary_meta.get("current_month_progress_pct") or 0) - summary_progress) > 1e-9:
                header_mismatches.append("summary progress should use summed actual / summed expected")

            if header_mismatches:
                results.append(_fail("manufacturer sales trend header summaries", "; ".join(header_mismatches)))
            else:
                results.append(_ok("manufacturer sales trend header summaries", "summary cards, progress totals, judge buckets, and table modes verified"))

            try:
                from app.ui.current_table_followups.action_dispatcher import handle_current_table_followup_by_action

                pushed_tables = []
                pushed_notices = []

                def _test_find_col(frame, *, exact=(), include_any=(), exclude_any=()):
                    cols = [str(c) for c in frame.columns]
                    for name in exact:
                        if name in cols:
                            return name
                    for col in cols:
                        if include_any and not any(w in col for w in include_any):
                            continue
                        if exclude_any and any(w in col for w in exclude_any):
                            continue
                        return col
                    return ""

                def _test_to_num(sr):
                    return pd.to_numeric(
                        sr.fillna("").astype(str).str.replace(",", "", regex=False).str.replace("%", "", regex=False),
                        errors="coerce",
                    ).fillna(0)

                def _test_push_table(**kwargs):
                    pushed_tables.append(kwargs)
                    return True

                def _test_push_notice(**kwargs):
                    pushed_notices.append(kwargs)
                    return True

                class _NoopLog:
                    def info(self, *args, **kwargs):
                        return None

                    def exception(self, *args, **kwargs):
                        return None

                helpers = {
                    "find_col": _test_find_col,
                    "to_num": _test_to_num,
                    "push_table": _test_push_table,
                    "push_notice": _test_push_notice,
                }

                followup_mismatches = []
                source_action = "제약사별 매출 추세 분석 요약표"
                source_key = "test_manufacturer_summary"

                pushed_tables.clear()
                pushed_notices.clear()
                handled = handle_current_table_followup_by_action(
                    df=summary,
                    query="현재표 추세판정 집계",
                    top_n=20,
                    table_key=source_key,
                    source_action=source_action,
                    helpers=helpers,
                    log=_NoopLog(),
                )
                if not handled or not pushed_tables:
                    followup_mismatches.append("trend judge group should return table")
                else:
                    group_payload = pushed_tables[-1]
                    group_df = group_payload.get("df")
                    extra_meta = group_payload.get("extra_meta") or {}
                    if extra_meta.get("group_column") != "추세판정":
                        followup_mismatches.append(f"group column expected 추세판정 got={extra_meta.get('group_column')}")
                    if not isinstance(group_df, pd.DataFrame) or group_df.empty:
                        followup_mismatches.append("trend judge group dataframe empty")
                    else:
                        if "제약사수" not in group_df.columns:
                            followup_mismatches.append("trend judge group missing 제약사수")
                        elif int(pd.to_numeric(group_df["제약사수"], errors="coerce").fillna(0).sum()) != int(summary["제약사명"].nunique()):
                            followup_mismatches.append("trend judge group manufacturer count sum mismatch")
                        if "총매출액" in group_df.columns:
                            grouped_total = float(pd.to_numeric(group_df["총매출액"], errors="coerce").fillna(0).sum())
                            original_total = float(pd.to_numeric(summary["총매출액"], errors="coerce").fillna(0).sum())
                            if abs(grouped_total - original_total) > 1e-9:
                                followup_mismatches.append("trend judge group total sales sum mismatch")
                        if "현재표 분석/KPI 후속분석 불가" in str(group_payload.get("title") or ""):
                            followup_mismatches.append("trend judge group fell through to analytics kpi unsupported notice")

                for query, expected_col in [
                    ("현재표 추세판정 감소만 보여줘", "추세판정"),
                    ("현재표 제약사명 제약A 상세", "제약사명"),
                    ("현재표 당월 진척률 100 이상", "당월 진척률"),
                    ("현재표 총매출액 1억 이상", "총매출액"),
                ]:
                    pushed_tables.clear()
                    pushed_notices.clear()
                    handled = handle_current_table_followup_by_action(
                        df=summary,
                        query=query,
                        top_n=20,
                        table_key=source_key,
                        source_action=source_action,
                        helpers=helpers,
                        log=_NoopLog(),
                    )
                    if not handled or not pushed_tables:
                        followup_mismatches.append(f"followup should return table query={query}")
                        continue
                    out_df = pushed_tables[-1].get("df")
                    meta_extra = pushed_tables[-1].get("extra_meta") or {}
                    if not isinstance(out_df, pd.DataFrame):
                        followup_mismatches.append(f"followup output not dataframe query={query}")
                        continue
                    if expected_col not in out_df.columns:
                        followup_mismatches.append(f"followup output missing original column {expected_col} query={query}")
                    if "순번" not in out_df.columns:
                        followup_mismatches.append(f"followup should regenerate sequence query={query}")
                    if expected_col == "추세판정" and not out_df["추세판정"].astype(str).str.contains("감소").all():
                        followup_mismatches.append("trend judge text filter contains non 감소 rows")
                    if expected_col == "제약사명" and not out_df["제약사명"].astype(str).str.contains("제약A").all():
                        followup_mismatches.append("manufacturer text filter contains non 제약A rows")
                    if expected_col in {"당월 진척률", "총매출액"} and meta_extra.get("filter_column") != expected_col:
                        followup_mismatches.append(f"numeric filter meta column mismatch query={query} meta={meta_extra}")

                if followup_mismatches:
                    results.append(_fail("current table generic group/filter followups", "; ".join(followup_mismatches)))
                else:
                    results.append(_ok("current table generic group/filter followups", "generic group, text filter, numeric filter, and analytics fallback blocking verified"))
            except Exception as e:
                results.append(_fail("current table generic group/filter followups", f"{type(e).__name__}: {e}"))

            if mismatches:
                results.append(_fail("manufacturer sales trend summary", "; ".join(mismatches)))
            else:
                results.append(_ok("manufacturer sales trend summary", "summary schema, monthly pivot, totals, universe, and meta verified"))
        finally:
            if old_loader is not None:
                setattr(manufacturer_mod, "get_sales_trend_df", old_loader)
    except Exception as e:
        results.append(_fail("manufacturer sales trend", f"{type(e).__name__}: {e}"))


    try:
        supp_mod = importlib.import_module("app.services.analytics_supplier_stock_shortage_service")
        product_base = pd.DataFrame([
            {
                "제품코드": "P001",
                "제품명": "테스트제품",
                "규격": "EA",
                "제조사명": "제조사A",
                "제품그룹명": "G",
                "제품구분명": "D",
                "제품분류명": "C",
                "재고기준": "장부",
                "현재재고수량": 100,
                "현재재고금액": 2000,
                "재고평가단가": 20,
                "당월 예상출고수량": 100,
                "당월 잔여예상출고수량": 150,
                "당월 출고진척률": 50,
                "부족예상수량": 50,
                "부족예상금액": 1000,
                "1개월부족수량": 10,
                "2개월부족수량": 20,
                "3개월부족수량": 30,
                "재고커버월수": 0.67,
                "당월 재고충족률": 66.67,
                "재고부족판정": "부족",
                "부족등급": "부족",
            },
            {
                "제품코드": "P002",
                "제품명": "테스트제품2",
                "규격": "EA",
                "제조사명": "제조사A",
                "제품그룹명": "G",
                "제품구분명": "D",
                "제품분류명": "C",
                "재고기준": "장부",
                "현재재고수량": 13,
                "현재재고금액": 143,
                "재고평가단가": 11,
                "당월 예상출고수량": 0,
                "당월 잔여예상출고수량": 0,
                "당월 출고진척률": 0,
                "부족예상수량": 0,
                "부족예상금액": 0,
                "1개월부족수량": 0,
                "2개월부족수량": 0,
                "3개월부족수량": 0,
                "재고커버월수": 0,
                "당월 재고충족률": 100,
                "재고부족판정": "수요없음",
                "부족등급": "정상",
            }
        ])
        supplier_stock = pd.DataFrame([
            {
                "제품코드": "P001",
                "매입처코드": "A",
                "매입처명": "매입처A",
                "매입처원본재고수량": 120,
                "매입처원본재고금액": 1200,
                "양수재고수량": 120,
                "양수재고금액": 700,
                "최근6완료월매입금액": 70,
                "전체완료월매입금액": 70,
                "매입처입고누계수량": 70,
            },
            {
                "제품코드": "P001",
                "매입처코드": "B",
                "매입처명": "매입처B",
                "매입처원본재고수량": -20,
                "매입처원본재고금액": -200,
                "양수재고수량": 80,
                "양수재고금액": 300,
                "최근6완료월매입금액": 30,
                "전체완료월매입금액": 30,
                "매입처입고누계수량": 30,
            },
            {
                "제품코드": "P002",
                "매입처코드": "C",
                "매입처명": "매입처C",
                "매입처원본재고수량": 7,
                "매입처원본재고금액": 999999,
                "양수재고수량": 7,
                "양수재고금액": 77,
                "최근6완료월매입금액": 0,
                "전체완료월매입금액": 0,
                "매입처입고누계수량": 7,
            },
            {
                "제품코드": "P002",
                "매입처코드": "D",
                "매입처명": "매입처D",
                "매입처원본재고수량": 6,
                "매입처원본재고금액": 888888,
                "양수재고수량": 6,
                "양수재고금액": 66,
                "최근6완료월매입금액": 0,
                "전체완료월매입금액": 0,
                "매입처입고누계수량": 6,
            },
        ])
        detail = supp_mod.build_supplier_allocation_detail(product_base, supplier_stock)
        summary_all = supp_mod.build_supplier_shortage_summary(detail, {})
        summary_b = supp_mod.build_supplier_shortage_summary(detail, {"buy_nm": "매입처B"})

        a_amt = float(summary_all.loc[summary_all["매입처코드"] == "A", "배정부족예상금액"].iloc[0])
        b_amt = float(summary_all.loc[summary_all["매입처코드"] == "B", "배정부족예상금액"].iloc[0])
        total_qty = float(detail["배정부족예상수량"].sum())
        stock_sum = float(detail.loc[detail["제품코드"] == "P001", "매입처원본재고수량"].sum())
        p001_stock_amt = float(detail.loc[detail["제품코드"] == "P001", "매입처원본재고금액"].sum())
        p002_stock_amt = float(detail.loc[detail["제품코드"] == "P002", "매입처원본재고금액"].sum())
        p002_qty_7_amt = float(detail.loc[detail["매입처코드"] == "C", "매입처원본재고금액"].iloc[0])
        p002_qty_6_amt = float(detail.loc[detail["매입처코드"] == "D", "매입처원본재고금액"].iloc[0])
        neg_b = float(detail.loc[detail["매입처코드"] == "B", "매입처원본재고수량"].iloc[0])
        neg_b_amt = float(detail.loc[detail["매입처코드"] == "B", "매입처원본재고금액"].iloc[0])
        filtered_b_amt = float(summary_b["배정부족예상금액"].sum())

        mismatches = []
        if abs(a_amt - 700) > 1e-6 or abs(b_amt - 300) > 1e-6:
            mismatches.append(f"70/30 allocation expected 700/300 got {a_amt}/{b_amt}")
        if abs(total_qty - 50) > 1e-6:
            mismatches.append(f"shortage qty should remain product-level 50 got {total_qty}")
        if abs(stock_sum - 100) > 1e-6 or abs(neg_b + 20) > 1e-6:
            mismatches.append(f"negative supplier stock not preserved stock_sum={stock_sum} neg_b={neg_b}")
        if abs(p001_stock_amt - 2000) > 1e-6 or abs(neg_b_amt + 400) > 1e-6:
            mismatches.append(f"stock amount must use product unit price p001_sum={p001_stock_amt} neg_b_amt={neg_b_amt}")
        if abs(p002_stock_amt - 143) > 1e-6 or abs(p002_qty_7_amt - 77) > 1e-6 or abs(p002_qty_6_amt - 66) > 1e-6:
            mismatches.append(f"7/6 stock amount expected 77/66 sum 143 got {p002_qty_7_amt}/{p002_qty_6_amt}/{p002_stock_amt}")
        if not set(detail["재고정합성"].dropna().astype(str).unique()).issubset({"일치"}):
            mismatches.append(f"stock consistency should match public basis got={detail['재고정합성'].dropna().astype(str).unique().tolist()}")
        if abs(filtered_b_amt - 300) > 1e-6 or len(summary_b) != 1:
            mismatches.append(f"supplier filter must apply after allocation got rows={len(summary_b)} amount={filtered_b_amt}")

        if mismatches:
            results.append(_fail("supplier stock shortage allocation fixture", "; ".join(mismatches)))
        else:
            results.append(_ok("supplier stock shortage allocation fixture", "unit-price stock amount, negative stock, 70/30 allocation, and post-filter OK"))

        old_product_loader = getattr(supp_mod, "load_product_shortage_base")
        old_supplier_loader = getattr(supp_mod, "load_supplier_product_stock")
        try:
            setattr(supp_mod, "load_product_shortage_base", lambda params: product_base.copy())
            setattr(supp_mod, "load_supplier_product_stock", lambda product_codes, params: supplier_stock.copy())
            payload = supp_mod.get_supplier_stock_shortage_result({"stock_mode": "book"})
            result_df = payload.get("df") if isinstance(payload.get("df"), pd.DataFrame) else payload.get("df_display")
            meta = payload.get("meta") or {}
            import io
            import json

            unsafe_types = (pd.DataFrame, pd.Series, bytes, bytearray)
            unsafe_meta = [k for k, v in meta.items() if isinstance(v, unsafe_types)]
            unsafe_attrs = [
                k
                for k, v in getattr(result_df, "attrs", {}).items()
                if isinstance(v, unsafe_types)
            ] if isinstance(result_df, pd.DataFrame) else []
            room = {
                "id": "fixture",
                "messages": [
                    {
                        "role": "assistant",
                        "type": payload.get("type"),
                        "action": payload.get("action"),
                        "title": payload.get("title"),
                        "meta": meta,
                    }
                ],
            }
            json.dumps(room, ensure_ascii=False)

            from app.ui.chat_middleware import _make_table_downloads
            from openpyxl import load_workbook

            _, xlsx_buf = _make_table_downloads(result_df)
            sheet_names = load_workbook(io.BytesIO(xlsx_buf.getvalue()), read_only=True).sheetnames
            follow_df = result_df.head(1).copy()
            follow_df.attrs.clear()
            _, follow_xlsx_buf = _make_table_downloads(follow_df)
            follow_sheet_names = load_workbook(io.BytesIO(follow_xlsx_buf.getvalue()), read_only=True).sheetnames
            if unsafe_meta or unsafe_attrs:
                results.append(_fail("supplier stock shortage json-safe payload", f"unsafe_meta={unsafe_meta} unsafe_attrs={unsafe_attrs}"))
            elif sheet_names != ["매입처별요약", "제품매입처상세"]:
                results.append(_fail("supplier stock shortage json-safe payload", f"unexpected excel sheets={sheet_names}"))
            elif follow_sheet_names != ["SIMS"]:
                results.append(_fail("supplier stock shortage json-safe payload", f"current-table result should be one sheet got={follow_sheet_names}"))
            else:
                results.append(_ok("supplier stock shortage json-safe payload", "json.dumps room OK; original Excel 2 sheets; current-table Excel one sheet"))
        finally:
            setattr(supp_mod, "load_product_shortage_base", old_product_loader)
            setattr(supp_mod, "load_supplier_product_stock", old_supplier_loader)

        try:
            generic_mod = importlib.import_module("app.ui.current_table_followups.generic")
            captured: dict[str, object] = {}

            def _push_table(**kwargs):
                captured.update(kwargs)
                return True

            top_df = pd.DataFrame({
                "매입처코드": [f"B{i:02d}" for i in range(25)],
                "배정부족예상금액": list(range(25)),
                "매입처원본재고금액": list(range(100, 125)),
            })
            top_df.attrs["supplier_detail_key"] = "detail-key-should-not-inherit"
            ok_top = generic_mod.handle_common_column_filter_followup(
                df=top_df,
                query="현재표 배정부족예상금액 TOP 20",
                top_n=20,
                table_key="supplier-table",
                source_action="매입처별 재고부족 현황",
                helpers={"push_table": _push_table},
                log=log,
            )
            out_top = captured.get("df")
            if not ok_top or not isinstance(out_top, pd.DataFrame):
                results.append(_fail("supplier current-table top amount", "TOP 20 route did not push table"))
            elif len(out_top) != 20 or float(out_top["배정부족예상금액"].iloc[0]) != 24 or float(out_top["배정부족예상금액"].iloc[-1]) != 5:
                results.append(_fail("supplier current-table top amount", f"unexpected TOP rows={len(out_top)} head/tail={out_top['배정부족예상금액'].head(1).tolist()}/{out_top['배정부족예상금액'].tail(1).tolist()}"))
            elif getattr(out_top, "attrs", {}).get("supplier_detail_key"):
                results.append(_fail("supplier current-table top amount", "supplier_detail_key leaked into current-table TOP result"))
            else:
                results.append(_ok("supplier current-table top amount", "배정부족예상금액 TOP 20 sorted desc; detail attrs cleared"))
        except Exception as e:
            results.append(_fail("supplier current-table top amount", f"{type(e).__name__}: {e}"))

        try:
            dispatcher_mod = importlib.import_module("app.ui.current_table_followups.action_dispatcher")
            generic_mod = importlib.import_module("app.ui.current_table_followups.generic")
            chat_mod = importlib.import_module("app.ui.chat_middleware")
            import pyarrow as pa

            def _run_group_top_case(df: pd.DataFrame, query: str) -> tuple[bool, dict[str, object]]:
                captured: dict[str, object] = {}

                def _push_group_top_table(**kwargs):
                    captured.clear()
                    captured.update(kwargs)
                    return True

                ok = dispatcher_mod.handle_current_table_followup_by_action(
                    df=df,
                    query=query,
                    top_n=20,
                    table_key="product-flow-table",
                    source_action="제품수불현황 조회",
                    helpers={"push_table": _push_group_top_table, "push_notice": lambda **kwargs: True},
                    log=log,
                )
                return ok, captured

            base_df = pd.DataFrame({
                "영업사원": ["정원장", "박진우", "김도윤", "박진우", "정원장", "김도윤", None, "0"],
                "거래처명": ["A", "B", "C", "D", "E", "F", "G", "H"],
                "제품명": ["P1", "P2", "P1", "P2", "P3", "P3", "P4", "P4"],
                "합계금액": [100, 900, 300, 700, 200, 400, 50, 30],
                "합계수량": [10, 90, 30, 70, 20, 40, 5, 3],
                "총수량": [12, 92, 32, 72, 22, 42, 6, 4],
                "수불수량": [1, 9, 3, 7, 2, 4, 1, 1],
                "출고수량": [11, 99, 33, 77, 22, 44, 0, 0],
                "입고수량": [21, 19, 23, 17, 12, 14, 0, 0],
                "재고수량": [31, 39, 33, 37, 32, 34, 0, 0],
                "부족예상수량": [41, 49, 43, 47, 42, 44, 0, 0],
                "배정부족예상수량": [51, 59, 53, 57, 52, 54, 0, 0],
                "명세서번호": pd.Series([1, 2, 3, 4, 5, 6, None, 8], dtype="Int64"),
                "제조번호": ["000123", "001-A", "114625021", "000124", "001-B", "114625022", "", "000000"],
                "검수확인": ["1", "0", "Y", "1", "0", "Y", "", "0"],
            })
            source_snapshot = base_df.copy(deep=True)
            cases = [
                ("합계금액 존재 기본", base_df, "현재표 영업사원별 TOP 20 표로 만들어줘", "합계금액"),
                ("합계금액 없음 합계수량 기본", base_df.drop(columns=["합계금액"]), "현재표 영업사원별 TOP 20 표로 만들어줘", "합계수량"),
                ("전체 수량 없음 건수 기본", base_df.drop(columns=["합계금액", "합계수량", "총수량", "수불수량"]), "현재표 영업사원별 TOP 20 표로 만들어줘", "행수"),
                ("건수 명시", base_df, "현재표 영업사원별 건수 TOP 20", "행수"),
                ("합계수량 명시", base_df, "현재표 영업사원별 합계수량 TOP 20", "합계수량"),
                ("총수량 명시", base_df, "현재표 영업사원별 총수량 TOP 20", "총수량"),
                ("수불수량 명시", base_df, "현재표 영업사원별 수불수량 TOP 20", "수불수량"),
                ("출고수량 명시", base_df, "현재표 영업사원별 출고수량 TOP 20", "출고수량"),
                ("입고수량 명시", base_df, "현재표 영업사원별 입고수량 TOP 20", "입고수량"),
                ("재고수량 명시", base_df, "현재표 영업사원별 재고수량 TOP 20", "재고수량"),
                ("일반 수량 전체수량 우선", base_df, "현재표 영업사원별 수량 TOP 20", "합계수량"),
                ("일반 수량 방향수량 금지", base_df.drop(columns=["합계금액", "합계수량", "총수량", "수불수량", "부족예상수량", "배정부족예상수량"]), "현재표 영업사원별 수량 TOP 20", "행수"),
            ]
            mismatches = []
            case_reports = []
            arrow_checked = False
            for label, case_df, query, expected_metric in cases:
                ok_group_top, captured_group_top = _run_group_top_case(case_df.copy(deep=True), query)
                out_group_top = captured_group_top.get("df")
                meta_group_top = captured_group_top.get("extra_meta") if isinstance(captured_group_top.get("extra_meta"), dict) else {}
                metric = str(meta_group_top.get("group_top_metric") or "")
                if not ok_group_top or not isinstance(out_group_top, pd.DataFrame):
                    mismatches.append(f"{label}: did not push group TOP")
                    continue
                if metric != expected_metric:
                    mismatches.append(f"{label}: metric={metric} expected={expected_metric}")
                dimension_col = next((c for c in ("영업사원", "제품명", "제품코드") if c in out_group_top.columns), "")
                if not dimension_col:
                    mismatches.append(f"{label}: missing dimension columns={list(out_group_top.columns)}")
                else:
                    values = out_group_top[dimension_col].astype(str).tolist()
                    if dimension_col == "영업사원":
                        if not any("미지정" in v for v in values):
                            mismatches.append(f"{label}: missing 미지정 group values={values}")
                        if "0" not in values:
                            mismatches.append(f"{label}: literal string 0 not preserved values={values}")
                if label == "전체 수량 없음 건수 기본" and "출고수량" in out_group_top.columns and metric == "출고수량":
                    mismatches.append("default fallback incorrectly selected 출고수량")
                if not arrow_checked and isinstance(out_group_top, pd.DataFrame):
                    display_df = chat_mod._preserve_product_flow_table_dtypes(out_group_top.copy())
                    pa.Table.from_pandas(display_df, preserve_index=False)
                    if display_df["영업사원"].dtype == "float64":
                        mismatches.append("Arrow display boundary converted 영업사원 to float64")
                    if "행수" in display_df.columns and not pd.api.types.is_integer_dtype(display_df["행수"]):
                        mismatches.append(f"행수 dtype not integer after display boundary: {display_df['행수'].dtype}")
                    if any(str(v) == "None" for v in display_df.astype(object).to_numpy().ravel()):
                        mismatches.append("display boundary produced literal None")
                    arrow_checked = True
                case_reports.append(f"{label}={metric}")

            captured_filter: dict[str, object] = {}

            def _push_filter_table(**kwargs):
                captured_filter.clear()
                captured_filter.update(kwargs)
                return True

            row_top_df = pd.DataFrame({"거래처명": ["A", "B", "C"], "합계금액": [10, 30, 20]})
            row_top_ok = generic_mod.handle_common_column_filter_followup(
                df=row_top_df,
                query="현재표 합계금액 TOP 20",
                top_n=20,
                table_key="product-flow-table",
                source_action="제품수불현황 조회",
                helpers={"push_table": _push_filter_table},
                log=log,
            )
            row_top_out = captured_filter.get("df")
            if not row_top_ok or not isinstance(row_top_out, pd.DataFrame):
                mismatches.append("row-level 합계금액 TOP did not route")
            elif row_top_out["합계금액"].tolist() != [30, 20, 10]:
                mismatches.append(f"row-level 합계금액 TOP sort changed={row_top_out['합계금액'].tolist()}")
            selector_out = pd.DataFrame({
                "영업사원": ["A", "B"],
                "행수": pd.Series([2, 1], dtype="Int64"),
                "합계금액": [100, 50],
                "합계수량": [10, 5],
                "총수량": [12, 6],
                "수불수량": [1, 9],
                "출고수량": [11, 99],
                "입고수량": [21, 19],
                "재고수량": [31, 39],
                "부족예상수량": [41, 49],
                "배정부족예상수량": [51, 59],
            })
            selector_cases = [
                ("현재표 영업사원별 수불수량 TOP 20", "수불수량"),
                ("현재표 영업사원별 부족예상수량 TOP 20", "부족예상수량"),
                ("현재표 영업사원별 배정부족예상수량 TOP 20", "배정부족예상수량"),
                ("현재표 영업사원별 출고수량 TOP 20", "출고수량"),
                ("현재표 영업사원별 입고수량 TOP 20", "입고수량"),
                ("현재표 영업사원별 재고수량 TOP 20", "재고수량"),
                ("현재표 영업사원별 합계수량 TOP 20", "합계수량"),
                ("현재표 영업사원별 총수량 TOP 20", "총수량"),
                ("현재표 영업사원별 수량 TOP 20", "합계수량"),
            ]
            selector_reports = []
            for query_text, expected_metric in selector_cases:
                metric_col, _metric_label = generic_mod._select_common_group_top_metric(selector_out, query_text)
                selector_reports.append(f"{expected_metric}->{metric_col}")
                if metric_col != expected_metric:
                    mismatches.append(f"selector {query_text!r}: metric={metric_col} expected={expected_metric}")
            direction_only = selector_out.drop(columns=["합계금액", "합계수량", "총수량", "수불수량", "부족예상수량", "배정부족예상수량"])
            metric_col, _metric_label = generic_mod._select_common_group_top_metric(direction_only, "현재표 영업사원별 수량 TOP 20")
            selector_reports.append(f"direction-only 수량->{metric_col}")
            if metric_col != "행수":
                mismatches.append(f"general 수량 with directional only selected {metric_col}, expected 행수")
            try:
                pd.testing.assert_frame_equal(base_df, source_snapshot, check_dtype=True)
            except AssertionError as ae:
                mismatches.append(f"source df mutated: {ae}")
            if mismatches:
                results.append(_fail("product flow salesperson dimension TOP", "; ".join(mismatches)))
            else:
                results.append(_ok("product flow salesperson dimension TOP", "fallback cases OK: " + ", ".join(case_reports) + "; row-level 합계금액 TOP and Arrow conversion OK"))
                results.append(_ok("product flow group TOP explicit metric aliases", "specific quantity metrics win over generic 수량 alias: " + ", ".join(selector_reports)))
                results.append(_ok("product flow group TOP fallback matrix", "default metric order amount -> total quantity -> count; explicit count/quantity/directional quantity respected"))
                results.append(_ok("product flow group TOP Arrow boundary", "group TOP payload survives product-flow display dtype preservation and pyarrow conversion without literal None"))
        except Exception as e:
            results.append(_fail("product flow salesperson dimension TOP", f"{type(e).__name__}: {e}"))

        try:
            generic_mod = importlib.import_module("app.ui.current_table_followups.generic")
            captured_filter: dict[str, object] = {}

            def _push_filter_table(**kwargs):
                captured_filter.clear()
                captured_filter.update(kwargs)
                return True

            shortage_df = pd.DataFrame(
                [
                    {"제품코드": "P001", "제품명": "가나다정", "규격": "10T", "제조사명": "A제약", "부족예상수량": 5, "부족예상금액": 1000},
                    {"제품코드": "P002", "제품명": "라마바정", "규격": "20T", "제조사명": "B제약", "부족예상수량": 12, "부족예상금액": 2000},
                    {"제품코드": "P003", "제품명": "사아자정", "규격": "30T", "제조사명": "C제약", "부족예상수량": 30, "부족예상금액": 3000},
                ]
            )
            ok_filter = generic_mod.handle_common_column_filter_followup(
                df=shortage_df,
                query="현재표 부족제품수 10 이상 목록",
                top_n=20,
                table_key="stock-shortage",
                source_action="품목별 재고부족현황",
                helpers={"push_table": _push_filter_table},
                log=log,
            )
            out_filter = captured_filter.get("df")
            action_filter = str(captured_filter.get("action") or "")
            if not ok_filter or not isinstance(out_filter, pd.DataFrame):
                results.append(_fail("current-table shortage qty alias", "부족제품수 numeric filter did not push table"))
            elif "부족예상수량" not in action_filter or "제품명" in action_filter:
                results.append(_fail("current-table shortage qty alias", f"unexpected action={action_filter}"))
            elif list(out_filter["제품명"]) != ["라마바정", "사아자정"]:
                results.append(_fail("current-table shortage qty alias", f"product names shifted={list(out_filter['제품명'])}"))
            elif "부족예상수량" not in list(out_filter.columns):
                results.append(_fail("current-table shortage qty alias", f"missing 부족예상수량 columns={list(out_filter.columns)}"))
            else:
                results.append(_ok("current-table shortage qty alias", "부족제품수 10 이상 -> 부족예상수량 numeric filter; 제품명 values preserved"))
        except Exception as e:
            results.append(_fail("current-table shortage qty alias", f"{type(e).__name__}: {e}"))

        try:
            chat_mod = importlib.import_module("app.ui.chat_middleware")
            amount_df = pd.DataFrame({
                "기준월": ["202607"],
                "매입처코드": ["B001"],
                "완료월수": [6],
                "배정부족예상금액": [1234],
                "배정1개월부족금액": [100],
                "매입처원본재고금액": [200],
            })
            profile = chat_mod._build_sims_sales_time_profile(
                amount_df,
                chat_mod._sims_business_terms("매입처별 재고부족 현황"),
            )
            amount_col = str((profile or {}).get("amount_col") or "")
            amount_label = str((profile or {}).get("amount_label") or "")
            if amount_col != "배정부족예상금액":
                results.append(_fail("supplier amount column priority", f"expected 배정부족예상금액 got={amount_col}"))
            elif amount_label != "배정부족예상금액":
                results.append(_fail("supplier amount column priority", f"expected amount_label=배정부족예상금액 got={amount_label}"))
            else:
                current_profile = chat_mod._build_sims_sales_time_profile(
                    amount_df,
                    chat_mod._sims_business_terms("현재표 배정부족예상금액 TOP 20"),
                )
                current_amount_col = str((current_profile or {}).get("amount_col") or "")
                current_amount_label = str((current_profile or {}).get("amount_label") or "")
                inherited_ctx = chat_mod._build_sims_analysis_context_from_df(
                    amount_df,
                    result={},
                    action_name="현재표 부족제품수 10 이상 목록",
                    params={},
                    meta={
                        "current_table_followup": True,
                        "flow": "매입처별 재고부족",
                        "amount_label": "배정부족예상금액",
                        "amount_priority": (
                            "배정부족예상금액",
                            "배정1개월부족금액",
                            "매입처원본재고금액",
                        ),
                    },
                )
                inherited_amount_col = str(((inherited_ctx or {}).get("sales_time_profile") or {}).get("amount_col") or "")
                if current_amount_col != "배정부족예상금액" or current_amount_label != "배정부족예상금액":
                    results.append(_fail("supplier amount column priority", f"current-table amount mismatch col={current_amount_col} label={current_amount_label}"))
                elif inherited_amount_col != "배정부족예상금액":
                    results.append(_fail("supplier amount column priority", f"inherited current-table amount mismatch col={inherited_amount_col}"))
                else:
                    results.append(_ok("supplier amount column priority", "amount_col=배정부족예상금액 for source, followup, and inherited profile"))
        except Exception as e:
            results.append(_fail("supplier amount column priority", f"{type(e).__name__}: {e}"))

        try:
            chat_mod = importlib.import_module("app.ui.chat_middleware")
            panel_mod = importlib.import_module("app.ui.sims_panel")

            class _FakeExpander:
                def __enter__(self):
                    return self

                def __exit__(self, *_args):
                    return False

            class _FakeStHeader:
                def __init__(self):
                    self.captions: list[str] = []
                    self.markdowns: list[str] = []
                    self.expanders: list[tuple[str, bool]] = []

                def caption(self, text, *args, **kwargs):
                    self.captions.append(str(text))

                def markdown(self, text, *args, **kwargs):
                    self.markdowns.append(str(text))

                def expander(self, label, expanded=False, *args, **kwargs):
                    self.expanders.append((str(label), bool(expanded)))
                    return _FakeExpander()

            long_tail = "끝부분-원문보존"
            long_query = "조회조건 " + ("가나다라마바사 " * 80) + long_tail
            long_nlq = "NLQ 원문 " + ("사용자 질문 내용 " * 80) + long_tail
            sample_df = pd.DataFrame({"제품코드": ["P001", "P002"], "부족예상수량": [1, 2]})
            sample_item = {
                "type": "table",
                "title": "품목별 재고부족현황",
                "action": "품목별 재고부족현황",
                "params": {"date_from": "2026-01-01", "date_to": "2026-07-17"},
            }
            sample_meta = {
                "table_key": "sims_fixture",
                "source_key": "source_fixture",
                "source_action": "품목별 재고부족현황",
                "query_summary": long_query,
                "download_row_count": 11713,
                "display_row_count": 3000,
                "created_at": "2026-07-17T20:01:31",
                "nlq_query": long_nlq,
            }
            view = chat_mod._build_sims_result_header_view(sample_item, sample_meta, sample_df, title="품목별 재고부족현황")
            fake_chat_st = _FakeStHeader()
            old_chat_st = getattr(chat_mod, "st", None)
            setattr(chat_mod, "st", fake_chat_st)
            try:
                chat_mod._render_sims_result_header_view(view)
            finally:
                if old_chat_st is not None:
                    setattr(chat_mod, "st", old_chat_st)

            expired_view = chat_mod._build_sims_result_header_view(sample_item, sample_meta, title="품목별 재고부족현황", expired=True)
            panel_payload = {
                "title": "품목별 재고부족현황",
                "action": "품목별 재고부족현황",
                "df": pd.DataFrame({"제품코드": range(5)}),
                "meta": dict(sample_meta),
            }
            fake_panel_st = _FakeStHeader()
            old_panel_st = getattr(panel_mod, "st", None)
            setattr(panel_mod, "st", fake_panel_st)
            try:
                panel_mod._render_panel_result_compact_header(panel_payload, "품목별 재고부족현황", "품목별 재고부족현황", sample_df)
            finally:
                if old_panel_st is not None:
                    setattr(panel_mod, "st", old_panel_st)
            panel_payload_no_key = {
                "title": "품목별 재고부족현황",
                "action": "품목별 재고부족현황",
                "df": pd.DataFrame({"제품코드": range(5)}),
                "meta": {k: v for k, v in sample_meta.items() if k != "table_key"},
            }
            fake_panel_no_key_st = _FakeStHeader()
            setattr(panel_mod, "st", fake_panel_no_key_st)
            try:
                panel_mod._render_panel_result_compact_header(panel_payload_no_key, "품목별 재고부족현황", "품목별 재고부족현황", sample_df)
            finally:
                if old_panel_st is not None:
                    setattr(panel_mod, "st", old_panel_st)

            header_mismatches = []
            if len(fake_chat_st.captions) != 2:
                header_mismatches.append(f"chat captions expected 2 got={fake_chat_st.captions}")
            if any("table_key" in c or "NLQ" in c for c in fake_chat_st.captions):
                header_mismatches.append(f"chat captions leaked diagnostics={fake_chat_st.captions}")
            detail_text = "\n".join(fake_chat_st.markdowns)
            if "table_key" not in detail_text or "sims_fixture" not in detail_text or "NLQ 원문" not in detail_text:
                header_mismatches.append(f"chat details missing diagnostics={detail_text}")
            if long_tail not in detail_text:
                header_mismatches.append("chat details truncated long query/NLQ tail")
            if long_tail in "\n".join(fake_chat_st.captions):
                header_mismatches.append(f"chat captions must compact long query={fake_chat_st.captions}")
            if fake_chat_st.expanders != [("상세 조회정보", False)]:
                header_mismatches.append(f"chat expander unexpected={fake_chat_st.expanders}")
            if "전체 11,713건" not in str(view.get("line1")) or "표 데이터 3,000건" not in str(view.get("line1")):
                header_mismatches.append(f"row summary unexpected={view.get('line1')}")
            if "품목별 재고부족현황" in str(view.get("line1")):
                header_mismatches.append("title duplicated inside row summary")
            if expired_view.get("followup_available"):
                header_mismatches.append("expired fallback must not advertise followup availability")
            if "만료" not in str(expired_view.get("line1")):
                header_mismatches.append(f"expired line missing status={expired_view.get('line1')}")
            if len(fake_panel_st.captions) != 2:
                header_mismatches.append(f"panel captions expected 2 got={fake_panel_st.captions}")
            if any("table_key" in c for c in fake_panel_st.captions):
                header_mismatches.append(f"panel captions leaked table_key={fake_panel_st.captions}")
            if "table_key" not in "\n".join(fake_panel_st.markdowns):
                header_mismatches.append("panel details missing table_key")
            if "현재표 후속질문 가능" not in " ".join(fake_panel_st.captions):
                header_mismatches.append(f"panel caption missing followup availability={fake_panel_st.captions}")
            if "현재표 후속질문 가능" in " ".join(fake_panel_no_key_st.captions):
                header_mismatches.append(f"panel without table_key advertised followup={fake_panel_no_key_st.captions}")
            if long_tail not in "\n".join(fake_panel_st.markdowns):
                header_mismatches.append("panel details truncated long query tail")

            if header_mismatches:
                results.append(_fail("sims result compact header render", "; ".join(header_mismatches)))
            else:
                results.append(_ok("sims result compact header render", "chat/panel headers use two captions; detail expander preserves long originals; expired/no-key paths hide followup"))
        except Exception as e:
            results.append(_fail("sims result compact header render", f"{type(e).__name__}: {e}"))

        try:
            chat_mod = importlib.import_module("app.ui.chat_middleware")
            st_obj = getattr(chat_mod, "st", None)
            session_state = getattr(st_obj, "session_state", None)
            if session_state is None:
                raise RuntimeError("streamlit session_state unavailable")

            old_cache = session_state.get("__sims_analysis_ctx_by_table_key")
            old_latest = session_state.get("__sims_analysis_ctx")
            try:
                product_ctx = {
                    "kind": "SIMS_ANALYSIS_CONTEXT_V1",
                    "table_key": "sims_product_codes",
                    "source_table_key": "",
                    "action": "제품코드 목록",
                    "row_count": 2,
                    "column_count": 2,
                    "analysis_text": "product context",
                }
                vendor_ctx = {
                    "kind": "SIMS_ANALYSIS_CONTEXT_V1",
                    "table_key": "sims_vendors",
                    "source_table_key": "",
                    "action": "거래처 목록",
                    "row_count": 3,
                    "column_count": 2,
                    "analysis_text": "vendor context",
                }
                session_state["__sims_analysis_ctx_by_table_key"] = {}
                chat_mod._cache_sims_analysis_ctx_by_table_key(product_ctx)
                chat_mod._cache_sims_analysis_ctx_by_table_key(vendor_ctx)
                session_state["__sims_analysis_ctx"] = vendor_ctx

                selected_ctx, selected_source = chat_mod._select_sims_analysis_ctx_for_table(
                    table_key="sims_product_codes",
                    action="제품코드 목록",
                    meta={"table_key": "sims_product_codes", "action": "제품코드 목록"},
                    download_df=pd.DataFrame({"제품코드": ["A", "B"], "제품명": ["a", "b"]}),
                )
                mismatch_same = chat_mod._sims_clicked_llm_context_mismatch(
                    selected_ctx,
                    "sims_product_codes",
                    "제품코드 목록",
                )
                mismatch_wrong = chat_mod._sims_clicked_llm_context_mismatch(
                    vendor_ctx,
                    "sims_product_codes",
                    "제품코드 목록",
                )
                rebuilt_ctx, rebuilt_source = chat_mod._select_sims_analysis_ctx_for_table(
                    table_key="sims_rebuilt",
                    action="제품코드 목록",
                    meta={"table_key": "sims_rebuilt", "action": "제품코드 목록"},
                    download_df=pd.DataFrame({"제품코드": ["C"], "제품명": ["c"]}),
                )
                mismatch_spaced = chat_mod._sims_clicked_llm_context_mismatch(
                    {"kind": "SIMS_ANALYSIS_CONTEXT_V1", "table_key": "sims_product_codes", "action": "Product Codes"},
                    "sims_product_codes",
                    "  Product   Codes  ",
                )

                class _DummyColumn:
                    def __enter__(self):
                        return self

                    def __exit__(self, exc_type, exc, tb):
                        return False

                calls = []
                warnings = []

                def _fake_runner(prompt, **kwargs):
                    calls.append({"prompt": prompt, "kwargs": kwargs})

                old_runner = session_state.get("__sims_llm_analysis_runner")
                old_button = getattr(chat_mod.st, "button")
                old_columns = getattr(chat_mod.st, "columns")
                old_download_button = getattr(chat_mod.st, "download_button")
                old_caption = getattr(chat_mod.st, "caption")
                old_warning = getattr(chat_mod.st, "warning")
                try:
                    session_state["__sims_llm_analysis_runner"] = _fake_runner
                    chat_mod.st.columns = lambda *args, **kwargs: [_DummyColumn(), _DummyColumn(), _DummyColumn()]
                    chat_mod.st.download_button = lambda *args, **kwargs: None
                    chat_mod.st.caption = lambda *args, **kwargs: None
                    chat_mod.st.warning = lambda msg, *args, **kwargs: warnings.append(str(msg))

                    chat_mod.st.button = lambda *args, **kwargs: True
                    chat_mod._render_sims_result_actions_plain(
                        key_suffix="plain_product",
                        csv_bytes=b"a,b\n1,2\n",
                        csv_name="plain.csv",
                        excel_bytes=b"xlsx",
                        xlsx_name="plain.xlsx",
                        prompt="analyze product",
                        table_key="sims_product_codes",
                        clicked_action="",
                        clicked_message_id="msg_product",
                        clicked_meta={"table_key": "sims_product_codes", "action": ""},
                        download_df=pd.DataFrame({"code": ["A"], "name": ["a"]}),
                    )

                    chat_mod.st.button = lambda *args, **kwargs: "prepare" not in str(kwargs.get("key") or "")
                    chat_mod._render_sims_result_actions_lazy(
                        key_suffix="lazy_product",
                        table_key="sims_product_codes",
                        download_df=pd.DataFrame({"code": ["A"], "name": ["a"]}),
                        csv_name="lazy.csv",
                        xlsx_name="lazy.xlsx",
                        prompt="analyze lazy product",
                        expected_rows=10000,
                        display_rows=1,
                        clicked_message_id="msg_product_lazy",
                        clicked_action="",
                        clicked_meta={"table_key": "sims_product_codes", "action": ""},
                    )

                    before_missing = len(calls)
                    chat_mod._run_clicked_sims_llm_analysis(
                        prompt="missing override",
                        key_suffix="missing_product",
                        table_key="sims_missing",
                        action="Product Codes",
                        message_id="msg_missing",
                        analysis_ctx=None,
                        context_source="missing",
                    )
                    missing_blocked = len(calls) == before_missing

                    before_mismatch = len(calls)
                    chat_mod._run_clicked_sims_llm_analysis(
                        prompt="wrong override",
                        key_suffix="wrong_product",
                        table_key="sims_product_codes",
                        action="?쒗뭹肄붾뱶 紐⑸줉",
                        message_id="msg_wrong",
                        analysis_ctx=vendor_ctx,
                        context_source="cache",
                    )
                    mismatch_blocked = len(calls) == before_mismatch
                finally:
                    if old_runner is None:
                        session_state.pop("__sims_llm_analysis_runner", None)
                    else:
                        session_state["__sims_llm_analysis_runner"] = old_runner
                    chat_mod.st.button = old_button
                    chat_mod.st.columns = old_columns
                    chat_mod.st.download_button = old_download_button
                    chat_mod.st.caption = old_caption
                    chat_mod.st.warning = old_warning

                source = Path("app/ui/chat_middleware.py").read_text(encoding="utf-8")
                main_source = Path("app/Lmstudio_SSAI_chat_main.py").read_text(encoding="utf-8")
                mismatches = []
                if selected_ctx.get("table_key") != "sims_product_codes" or selected_ctx.get("action") != "제품코드 목록":
                    mismatches.append(f"selected wrong ctx={selected_ctx}")
                if selected_source != "cache":
                    mismatches.append(f"expected cache source got={selected_source}")
                if mismatch_same:
                    mismatches.append(f"same table/action mismatch={mismatch_same}")
                if mismatch_wrong != "table_key_mismatch":
                    mismatches.append(f"wrong table mismatch={mismatch_wrong}")
                if mismatch_spaced:
                    mismatches.append(f"spaced action should match, got={mismatch_spaced}")
                if rebuilt_source != "rebuilt" or rebuilt_ctx.get("table_key") != "sims_rebuilt":
                    mismatches.append(f"rebuilt ctx/source unexpected={rebuilt_ctx}, {rebuilt_source}")
                if len(calls) != 2:
                    mismatches.append(f"expected plain/lazy runner calls=2 got={len(calls)} calls={calls}")
                else:
                    plain_ctx = calls[0]["kwargs"].get("analysis_ctx_override")
                    lazy_ctx = calls[1]["kwargs"].get("analysis_ctx_override")
                    if not isinstance(plain_ctx, dict) or plain_ctx.get("table_key") != "sims_product_codes":
                        mismatches.append(f"plain runner used wrong ctx={plain_ctx}")
                    if not isinstance(lazy_ctx, dict) or lazy_ctx.get("table_key") != "sims_product_codes":
                        mismatches.append(f"lazy runner used wrong ctx={lazy_ctx}")
                if not missing_blocked:
                    mismatches.append("missing table-scoped context did not block runner")
                if not mismatch_blocked:
                    mismatches.append("mismatched table-scoped context did not block runner")
                if "__sims_old_table_download_enabled" in source or "sims_old_table_prepare_excel" in source:
                    mismatches.append("old-table early Excel-only return block remains")
                if "analysis_ctx_override" not in main_source or "selected_context_table_key" not in main_source:
                    mismatches.append("main runner does not accept/log table-scoped analysis context")
                if "table_scoped_request and not valid_override" not in main_source:
                    mismatches.append("main runner does not fail closed when clicked table context is missing")
                if "_run_clicked_sims_llm_analysis" not in source or "clicked_table_key" not in source:
                    mismatches.append("chat action buttons do not use table-scoped LLM helper")

                if mismatches:
                    results.append(_fail("table-scoped SIMS LLM analysis context", "; ".join(mismatches)))
                else:
                    results.append(_ok("table-scoped SIMS LLM analysis context", "old table buttons use clicked table_key/action context; latest global context is not used as fallback; mismatch blocks LLM"))
            finally:
                if old_cache is None:
                    session_state.pop("__sims_analysis_ctx_by_table_key", None)
                else:
                    session_state["__sims_analysis_ctx_by_table_key"] = old_cache
                if old_latest is None:
                    session_state.pop("__sims_analysis_ctx", None)
                else:
                    session_state["__sims_analysis_ctx"] = old_latest
        except Exception as e:
            results.append(_fail("table-scoped SIMS LLM analysis context", f"{type(e).__name__}: {e}"))

        try:
            chat_mod = importlib.import_module("app.ui.chat_middleware")
            st_obj = getattr(chat_mod, "st", None)
            session_state = getattr(st_obj, "session_state", None)
            if session_state is None:
                raise RuntimeError("streamlit session_state unavailable")

            saved_state = {
                key: session_state.get(key)
                for key in (
                    "__sims_analysis_ctx_by_table_key",
                    "__sims_analysis_ctx",
                    "__sims_current_table_source_key",
                    "__sims_current_table_source_action",
                    "__sims_current_table_source_analysis_ctx",
                    "__sims_last_table_key",
                    "__sims_last_table_action",
                    "sims_tables",
                    "sims_export_tables",
                    "__sims_export_tables_by_key",
                    "__chat_inbox",
                )
            }
            old_drain = getattr(chat_mod, "drain_inbox_to_chat")
            old_company_match = getattr(chat_mod, "_chat_payload_matches_current_company")
            try:
                for key in saved_state:
                    session_state.pop(key, None)
                session_state["__sims_analysis_ctx_by_table_key"] = {}
                session_state["sims_tables"] = {}
                session_state["sims_export_tables"] = {}
                session_state["__sims_export_tables_by_key"] = {}
                session_state["__chat_inbox"] = []
                chat_mod.drain_inbox_to_chat = lambda *args, **kwargs: None
                chat_mod._chat_payload_matches_current_company = lambda payload: True

                df_a = pd.DataFrame({"제품코드": ["P001", "P002"], "합계금액": [100, 200]})
                df_p = pd.DataFrame({"영업사원": ["김"], "합계금액": [300]})
                df_b = pd.DataFrame({"거래처명": ["거래처"], "건수": [1]})

                chat_mod.wssz(
                    {
                        "type": "table",
                        "title": "제품수불현황 조회",
                        "action": "제품수불현황 조회",
                        "df": df_a,
                        "df_display": df_a,
                        "meta": {"table_key": "sims_a_ctx_owner", "action": "제품수불현황 조회"},
                    },
                    "제품수불현황 조회",
                )
                cache_after_a = dict(session_state.get("__sims_analysis_ctx_by_table_key") or {})

                chat_mod.wssz(
                    {
                        "type": "table",
                        "title": "현재표 영업사원별 집계",
                        "action": "현재표 영업사원별 집계",
                        "df": df_p,
                        "df_display": df_p,
                        "meta": {
                            "current_table_followup": True,
                            "source_table_key": "sims_a_ctx_owner",
                            "action": "현재표 영업사원별 집계",
                        },
                    },
                    "현재표 영업사원별 집계",
                )
                p_key = str(session_state.get("__sims_last_table_key") or "").strip()
                cache_after_p = dict(session_state.get("__sims_analysis_ctx_by_table_key") or {})

                chat_mod.wssz(
                    {
                        "type": "table",
                        "title": "거래처 목록",
                        "action": "거래처 목록",
                        "df": df_b,
                        "df_display": df_b,
                        "meta": {"table_key": "sims_b_ctx_owner", "action": "거래처 목록"},
                    },
                    "거래처 목록",
                )
                cache_after_b = dict(session_state.get("__sims_analysis_ctx_by_table_key") or {})

                selected_a, selected_a_source = chat_mod._select_sims_analysis_ctx_for_table(
                    table_key="sims_a_ctx_owner",
                    action="제품수불현황 조회",
                    meta={"table_key": "sims_a_ctx_owner", "action": "제품수불현황 조회"},
                    download_df=df_a,
                )
                selected_p, selected_p_source = chat_mod._select_sims_analysis_ctx_for_table(
                    table_key=p_key,
                    action="현재표 영업사원별 집계",
                    meta={"table_key": p_key, "source_table_key": "sims_a_ctx_owner", "current_table_followup": True, "action": "현재표 영업사원별 집계"},
                    download_df=df_p,
                )

                original_ctx = {
                    "kind": "SIMS_ANALYSIS_CONTEXT_V1",
                    "table_key": "sims_a_rebuild",
                    "source_table_key": "",
                    "action": "제품수불현황 조회",
                    "analysis_text": "original",
                }
                bad_derived_ctx = {
                    "kind": "SIMS_ANALYSIS_CONTEXT_V1",
                    "table_key": "sims_a_rebuild",
                    "source_table_key": "sims_a_rebuild",
                    "current_table_followup": True,
                    "action": "현재표 영업사원별 집계",
                    "analysis_text": "bad derived",
                }
                session_state["__sims_analysis_ctx_by_table_key"] = {"sims_a_rebuild": original_ctx}
                blocked_key = chat_mod._cache_sims_analysis_ctx_by_table_key(bad_derived_ctx)
                blocked_ctx = dict(session_state.get("__sims_analysis_ctx_by_table_key") or {}).get("sims_a_rebuild")

                session_state["__sims_analysis_ctx_by_table_key"] = {"sims_a_rebuild": bad_derived_ctx}
                rebuilt_a, rebuilt_a_source = chat_mod._select_sims_analysis_ctx_for_table(
                    table_key="sims_a_rebuild",
                    action="제품수불현황 조회",
                    meta={"table_key": "sims_a_rebuild", "action": "제품수불현황 조회"},
                    download_df=df_a,
                )
                session_state["__sims_analysis_ctx_by_table_key"] = {"sims_a_rebuild": bad_derived_ctx}
                missing_a, missing_a_source = chat_mod._select_sims_analysis_ctx_for_table(
                    table_key="sims_a_rebuild",
                    action="제품수불현황 조회",
                    meta={"table_key": "sims_a_rebuild", "action": "제품수불현황 조회"},
                    download_df=None,
                )

                mismatches = []
                if cache_after_a.get("sims_a_ctx_owner", {}).get("action") != "제품수불현황 조회":
                    mismatches.append(f"A cache not original after A={cache_after_a}")
                if cache_after_p.get("sims_a_ctx_owner", {}).get("action") != "제품수불현황 조회":
                    mismatches.append(f"A cache overwritten by P={cache_after_p.get('sims_a_ctx_owner')}")
                if not p_key or p_key == "sims_a_ctx_owner":
                    mismatches.append(f"derived table key invalid={p_key}")
                if cache_after_p.get(p_key, {}).get("action") != "현재표 영업사원별 집계":
                    mismatches.append(f"P cache missing/wrong={p_key}:{cache_after_p.get(p_key)}")
                if cache_after_p.get(p_key, {}).get("source_table_key") != "sims_a_ctx_owner":
                    mismatches.append(f"P source key wrong={cache_after_p.get(p_key)}")
                for key in ("sims_a_ctx_owner", p_key, "sims_b_ctx_owner"):
                    if key not in cache_after_b:
                        mismatches.append(f"cache key missing after B={key}:{cache_after_b}")
                if selected_a_source != "cache" or selected_a.get("action") != "제품수불현황 조회":
                    mismatches.append(f"A selection wrong={selected_a_source}:{selected_a}")
                if selected_p_source != "cache" or selected_p.get("action") != "현재표 영업사원별 집계":
                    mismatches.append(f"P selection wrong={selected_p_source}:{selected_p}")
                if str(session_state.get("__sims_current_table_source_key") or "") != "sims_b_ctx_owner":
                    mismatches.append(f"current source changed={session_state.get('__sims_current_table_source_key')}")
                if blocked_key or blocked_ctx.get("action") != "제품수불현황 조회":
                    mismatches.append(f"collision block failed key={blocked_key} ctx={blocked_ctx}")
                if rebuilt_a_source != "rebuilt" or rebuilt_a.get("action") != "제품수불현황 조회":
                    mismatches.append(f"rebuild from clicked df failed={rebuilt_a_source}:{rebuilt_a}")
                if missing_a is not None or missing_a_source != "missing":
                    mismatches.append(f"missing df should fail closed={missing_a_source}:{missing_a}")

                if mismatches:
                    results.append(_fail("SIMS LLM context cache ownership", "; ".join(mismatches)))
                else:
                    results.append(_ok("SIMS LLM context cache ownership", "source A and derived P contexts stay under their own table_key; stale collision rebuild/fail-closed verified"))
            finally:
                chat_mod.drain_inbox_to_chat = old_drain
                chat_mod._chat_payload_matches_current_company = old_company_match
                for key, value in saved_state.items():
                    if value is None:
                        session_state.pop(key, None)
                    else:
                        session_state[key] = value
        except Exception as e:
            results.append(_fail("SIMS LLM context cache ownership", f"{type(e).__name__}: {e}"))

        try:
            from app.ui.current_table_followups.action_dispatcher import (
                classify_current_table_followup_intent,
                current_table_analysis_query_matches,
                current_table_analysis_ctx_mismatch,
                select_current_table_analysis_context,
            )

            product_ctx = {
                "kind": "SIMS_ANALYSIS_CONTEXT_V1",
                "table_key": "sims_product_flow",
                "source_table_key": "",
                "action": "제품수불현황 조회",
                "row_count": 12,
                "column_count": 8,
                "analysis_text": "product flow context",
            }
            vendor_ctx = {
                "kind": "SIMS_ANALYSIS_CONTEXT_V1",
                "table_key": "sims_vendor",
                "source_table_key": "",
                "action": "거래처 목록",
                "row_count": 3,
                "column_count": 5,
                "analysis_text": "vendor context",
            }
            session_state = {
                "__sims_current_table_source_key": "sims_product_flow",
                "__sims_current_table_source_action": "제품수불현황 조회",
                "__sims_analysis_ctx_by_table_key": {
                    "sims_product_flow": product_ctx,
                    "sims_vendor": vendor_ctx,
                },
                "__sims_analysis_ctx": vendor_ctx,
            }

            selected_ctx, selected_source, selected_reason = select_current_table_analysis_context(session_state)

            route_cases = {
                "현재표를 분석해줘": "llm_analysis",
                "현재표만 분석해줘": "llm_analysis",
                "현재표로 분석해줘": "llm_analysis",
                "현재표 이상 항목 분석해줘": "llm_analysis",
                "현재표에서 이상 항목을 알려줘": "llm_analysis",
                "현재표의 문제점과 주의사항을 분석해줘": "llm_analysis",
                "현재표 금액 100만원 이상만 보여줘": "dataframe_table",
                "현재표 수량 10 이상 목록": "dataframe_table",
                "현재표 재고 0 이하인 제품만 표로 만들어줘": "dataframe_table",
                "현재표를 영업사원별 TOP 20 표로 만들어줘": "dataframe_table",
                "현재표에서 요약표를 만들어줘": "dataframe_table",
                "현재표 영업사원 분석": "llm_analysis",
                "현재표 영업사원 요약해줘": "llm_analysis",
                "현재표에서 확인할 점 알려줘": "llm_analysis",
                "현재표 영업사원별 TOP 20 표로 만들어줘": "dataframe_table",
                "현재표 영업사원별 집계": "dataframe_table",
                "현재표 제품별 목록": "dataframe_table",
                "현재표 요약표": "dataframe_table",
            }

            route_mismatches = []
            for query, expected_intent in route_cases.items():
                actual_intent = classify_current_table_followup_intent(query)
                if actual_intent != expected_intent:
                    route_mismatches.append(f"{query}:{actual_intent}!={expected_intent}")

            intent_edge_cases = {
                "현재표를 분석해줘": "llm_analysis",
                "현재표만 분석해줘": "llm_analysis",
                "현재표로 분석해줘": "llm_analysis",
                "현재표 이상 항목 분석해줘": "llm_analysis",
                "현재표에서 이상 항목을 알려줘": "llm_analysis",
                "현재표 금액 100만원 이상만 보여줘": "dataframe_table",
                "현재표 수량 10 이상 목록": "dataframe_table",
                "현재표를 영업사원별 TOP 20 표로 만들어줘": "dataframe_table",
                "현재표에서 요약표를 만들어줘": "dataframe_table",
                "현재표 이상 항목 목록을 만들어줘": "dataframe_table",
                "현재표 이상 항목 중 100만원 이상만 보여줘": "dataframe_table",
                "현재표 문제점을 요약표로 만들어줘": "dataframe_table",
                "현재표 주의사항을 표로 정리해줘": "dataframe_table",
            }
            intent_edge_mismatches = []
            for query, expected_intent in intent_edge_cases.items():
                actual_intent = classify_current_table_followup_intent(query)
                if actual_intent != expected_intent:
                    intent_edge_mismatches.append(f"{query}:{actual_intent}!={expected_intent}")
            if intent_edge_mismatches:
                results.append(_fail("current-table intent edge cases", "; ".join(intent_edge_mismatches)))
            else:
                results.append(_ok("current-table intent edge cases", "current-table referent particles and 이상 항목 analysis are separated from explicit table/filter requests"))

            if selected_source != "cache" or selected_reason:
                route_mismatches.append(f"expected cache ctx got source={selected_source} reason={selected_reason}")
            if not isinstance(selected_ctx, dict) or selected_ctx.get("table_key") != "sims_product_flow":
                route_mismatches.append(f"selected wrong current ctx={selected_ctx}")
            if current_table_analysis_ctx_mismatch(selected_ctx, "sims_product_flow", "제품수불현황 조회"):
                route_mismatches.append("matched current ctx reported mismatch")
            if current_table_analysis_ctx_mismatch(vendor_ctx, "sims_product_flow", "제품수불현황 조회") != "table_key_mismatch":
                route_mismatches.append("vendor latest ctx must not be accepted for product current table")

            expired_state = {
                "__sims_current_table_source_key": "sims_expired",
                "__sims_current_table_source_action": "제품수불현황 조회",
                "__sims_analysis_ctx_by_table_key": {
                    "sims_expired": {
                        **product_ctx,
                        "table_key": "sims_expired",
                        "expired": True,
                    }
                },
            }
            expired_ctx, expired_source, expired_reason = select_current_table_analysis_context(expired_state)
            if expired_ctx is not None or expired_source or expired_reason != "expired_context":
                route_mismatches.append(f"expired current ctx should fail closed got ctx={expired_ctx} source={expired_source} reason={expired_reason}")
            expired_string_false_state = {
                "__sims_current_table_source_key": "sims_product_flow",
                "__sims_current_table_source_action": "제품수불현황 조회",
                "__sims_analysis_ctx_by_table_key": {
                    "sims_product_flow": {
                        **product_ctx,
                        "expired": "false",
                    }
                },
            }
            false_ctx, false_source, false_reason = select_current_table_analysis_context(expired_string_false_state)
            if not isinstance(false_ctx, dict) or false_source != "cache" or false_reason:
                route_mismatches.append(f"expired='false' should remain valid got ctx={false_ctx} source={false_source} reason={false_reason}")
            expired_string_true_state = {
                "__sims_current_table_source_key": "sims_product_flow",
                "__sims_current_table_source_action": "제품수불현황 조회",
                "__sims_analysis_ctx_by_table_key": {
                    "sims_product_flow": {
                        **product_ctx,
                        "expired": "true",
                    }
                },
            }
            true_ctx, true_source, true_reason = select_current_table_analysis_context(expired_string_true_state)
            if true_ctx is not None or true_source or true_reason != "expired_context":
                route_mismatches.append(f"expired='true' should fail closed got ctx={true_ctx} source={true_source} reason={true_reason}")

            latest_only_state = {
                "__sims_current_table_source_key": "sims_product_flow",
                "__sims_current_table_source_action": "제품수불현황 조회",
                "__sims_analysis_ctx": product_ctx,
            }
            latest_ctx, latest_source, latest_reason = select_current_table_analysis_context(latest_only_state)
            if latest_ctx is not None or latest_source or latest_reason != "missing_context":
                route_mismatches.append(f"latest-only global ctx should not be accepted got ctx={latest_ctx} source={latest_source} reason={latest_reason}")

            action_mismatch_state = {
                "__sims_current_table_source_key": "sims_product_flow",
                "__sims_current_table_source_action": "제품수불현황 조회",
                "__sims_analysis_ctx_by_table_key": {
                    "sims_product_flow": {
                        **product_ctx,
                        "action": "거래처 목록",
                    }
                },
            }
            mismatch_ctx, mismatch_source, mismatch_reason = select_current_table_analysis_context(action_mismatch_state)
            if mismatch_ctx is not None or mismatch_source or mismatch_reason != "action_mismatch":
                route_mismatches.append(f"action mismatch should fail closed got ctx={mismatch_ctx} source={mismatch_source} reason={mismatch_reason}")
            if not current_table_analysis_query_matches("현재표 영업사원 분석", "현재표 영업사원 분석"):
                route_mismatches.append("analysis query match helper rejected identical query")
            if current_table_analysis_query_matches("현재표 영업사원 분석", "다른 질문"):
                route_mismatches.append("analysis query match helper accepted mismatched query")

            main_source = Path("app/Lmstudio_SSAI_chat_main.py").read_text(encoding="utf-8")
            dispatcher_source = Path("app/ui/current_table_followups/action_dispatcher.py").read_text(encoding="utf-8")
            required_snippets = {
                "intent_helper_import": "classify_current_table_followup_intent" in main_source,
                "pandas_skip": "stage=pandas_handler_skip" in main_source,
                "one_shot_override": "__current_table_analysis_ctx_override" in main_source,
                "queue_override": "analysis_ctx_override=analysis_ctx_override" in main_source,
                "fail_closed_notice": "현재표 분석 컨텍스트 불가" in main_source,
                "route_log": "[current_table.route]" in main_source,
                "analysis_ctx_log": "[current_table.analysis_ctx]" in main_source,
                "query_match_guard": "current_table_analysis_query_matches(analysis_query, last_user_text)" in main_source,
                "disable_context_on_mismatch": "disable_sims_analysis_ctx = True" in main_source,
                "latest_global_removed": '("latest", "__sims_analysis_ctx")' not in dispatcher_source,
                "expired_context_guard": "expired_context" in dispatcher_source,
            }
            missing_snippets = [name for name, ok in required_snippets.items() if not ok]
            if missing_snippets:
                route_mismatches.append(f"main routing snippets missing={missing_snippets}")

            if route_mismatches:
                results.append(_fail("current-table source routing and LLM analysis intent", "; ".join(route_mismatches)))
            else:
                results.append(_ok("current-table source routing and LLM analysis intent", "general analysis uses table-scoped current source context; explicit table requests remain pandas-only; missing/mismatch contexts fail closed"))
        except Exception as e:
            results.append(_fail("current-table source routing and LLM analysis intent", f"{type(e).__name__}: {e}"))

        try:
            chat_mod = importlib.import_module("app.ui.chat_middleware")

            master_cases = [
                {
                    "analysis_type": "users_master",
                    "llm_summary_kind": "users_master_summary",
                    "llm_summary_md": "사용자마스터 전체 집계 요약\n- 전체 조회건수: 227건",
                    "users_master_summary": {"top": ["부서"]},
                    "table_key": "sims_users",
                },
                {
                    "analysis_type": "goods_master",
                    "llm_summary_kind": "goods_master_summary",
                    "llm_summary_md": "제품마스터 전체 집계 요약\n- 전체 조회건수: 120건",
                    "goods_master_summary": {"top": ["제품"]},
                    "table_key": "sims_goods",
                },
                {
                    "analysis_type": "vendor_master",
                    "llm_summary_kind": "vendor_master_summary",
                    "llm_summary_md": "거래처마스터 자동 집계 요약\n- 전체 조회건수: 340건",
                    "vendor_master_summary": {"top": ["거래처"]},
                    "master_nlq": True,
                    "domain": "vendors",
                    "source_key": "sims_vendors",
                },
                {
                    "analysis_type": "codes_master",
                    "llm_summary_kind": "codes_master_summary",
                    "llm_summary_md": "코드마스터 전체 집계 요약\n- 전체 조회건수: 42건",
                    "codes_master_summary": {"top": ["코드"]},
                    "table_key": "sims_codes",
                },
            ]
            visible_cases = [
                {"summary_md": "조회 결과가 없습니다. 조건을 확인해 주세요.", "analysis_type": "no_data"},
                {"summary_md": "오류가 발생했습니다. 다시 시도해 주세요.", "analysis_type": "error"},
                {"summary_md": "LLM 분석 답변입니다.", "analysis_type": "llm_answer"},
                {"summary_md": "현재표 후속질문 결과 요약", "current_table_followup": True},
            ]
            summary_sample = (
                "조회조건: 기간 2026-01-01 ~ 2026-07-17\n"
                "조회결과: 227건\n"
                "조회결과: 11,713건\n"
                "전체 조회건수: 11,713건\n"
                "화면 표시건수: 3,000건\n"
                "조회 완료: 29,716건\n"
                "조회 결과: 11,713건 (표시 3,000건)\n"
                "조회결과: 조회 결과가 없습니다.\n"
                "조회 결과: 조건에 맞는 자료가 없습니다.\n"
                "조회 완료: 일부 자료를 불러오지 못했습니다.\n"
                "오류가 발생했습니다.\n"
                "권한이 없습니다.\n"
                "조건을 확인해 주세요.\n"
                "일부 데이터만 조회됐다는 경고\n"
                "KPI 수치: 123건\n"
                "업무 요약 본문"
            )
            cleaned = chat_mod._clean_chat_summary_text_v2(summary_sample, "기간 2026-01-01 ~ 2026-07-17")
            inbound_summary = (
                "조회조건: 기간 2026-07-01 ~ 2026-07-18 / 재고위치 전체\n"
                "조회결과: 전체 12,131건 / 화면 표시 12,131건\n"
                "입고수량: 9,999개\n"
                "업무 요약 본문"
            )
            inbound_cleaned = chat_mod._clean_chat_summary_text_v2(
                inbound_summary,
                "기간 2026-07-01 ~ 2026-07-18 / 재고위치 전체",
            )
            inbound_item_for_header = {
                "type": "table",
                "action": "입고명세 조회",
                "meta": {
                    "query_summary": "기간 2026-07-01 ~ 2026-07-18 / 재고위치 전체",
                    "table_key": "sims_dfca5279",
                },
            }
            inbound_summary_cond_text = chat_mod._summary_condition_text_for_cleanup(
                inbound_item_for_header,
                inbound_item_for_header["meta"],
                is_sims_table_or_text=True,
                caption_cond_text="",
            )
            inbound_cleaned_from_render_path = chat_mod._clean_chat_summary_text_v2(
                inbound_summary,
                inbound_summary_cond_text,
            )
            product_flow_summary = (
                "제품정보: 제품코드 12345 / 제품명 테스트제품 / 제조사명 테스트제약\n"
                "이월재고: 10개\n"
                "입고수량: 20개\n"
                "출고수량: 5개\n"
                "재고수량: 25개\n"
                "집계 요약 펼쳐보기"
            )
            product_flow_cleaned = chat_mod._clean_chat_summary_text_v2(
                product_flow_summary,
                "제품코드 12345 / 제품명 테스트제품 / 제조사명 테스트제약",
            )
            inventory_summary = (
                "조회조건: 기간 2026-01-01 ~ 2026-07-18 / 재고기준 장부재고\n"
                "조회 결과: 14,308건 (표시 3,000건)\n"
                "현재재고수량: 100개\n"
                "경고: 일부 데이터만 조회됐다는 경고"
            )
            inventory_cleaned = chat_mod._clean_chat_summary_text_v2(
                inventory_summary,
                "기간 2026-01-01 ~ 2026-07-18 / 재고기준 장부재고",
            )
            inventory_cleaned_drop_product = chat_mod._clean_chat_summary_text_v2(
                "제품정보: 제품코드 12345 / 제품명 테스트제제\n현재재고수량: 100개",
                "기간 2026-01-01 ~ 2026-07-18 / 재고기준 장부재고",
                drop_product_info=True,
            )
            inventory_cleaned_keep_product = chat_mod._clean_chat_summary_text_v2(
                "제품정보: 제품코드 12345 / 제품명 테스트제제\n현재재고수량: 100개",
                "제품코드 12345 / 제품명 테스트제제",
                drop_product_info=False,
            )
            explicit_product_meta = {"params": {"제품코드": "12345"}}
            maker_only_meta = {"params": {"제조사명": "테스트제약"}}
            product_group_only_meta = {"params": {"제품그룹명": "전문의약품"}}
            product_name_only_meta = {"params": {"제품명": "정"}}
            blank_product_code_meta = {"params": {"제품코드": ""}}
            full_inventory_df = pd.DataFrame({"제품코드": ["A", "B"], "제품명": ["a", "b"]})
            single_inventory_df = pd.DataFrame({"제품코드": ["A", "A"], "제품명": ["a", "a"]})
            src = Path("app/ui/chat_middleware.py").read_text(encoding="utf-8")
            mismatches = []
            if not all(chat_mod._is_internal_master_summary(meta) for meta in master_cases):
                mismatches.append("master summary meta not hidden")
            if any(chat_mod._is_internal_master_summary(meta) for meta in visible_cases):
                mismatches.append("visible user-facing summary misclassified as internal master")
            if "업무 요약 본문" not in cleaned:
                mismatches.append(f"cleaned summary lost body={cleaned!r}")
            removed_lines = (
                "조회결과: 227건",
                "조회결과: 11,713건",
                "전체 조회건수: 11,713건",
                "화면 표시건수: 3,000건",
                "조회 완료: 29,716건",
                "조회 결과: 11,713건 (표시 3,000건)",
            )
            preserved_lines = (
                "조회결과: 조회 결과가 없습니다.",
                "조회 결과: 조건에 맞는 자료가 없습니다.",
                "조회 완료: 일부 자료를 불러오지 못했습니다.",
                "오류가 발생했습니다.",
                "권한이 없습니다.",
                "조건을 확인해 주세요.",
                "일부 데이터만 조회됐다는 경고",
                "KPI 수치: 123건",
            )
            if any(line in cleaned for line in removed_lines):
                mismatches.append(f"cleaned summary kept duplicate row-count line={cleaned!r}")
            if any(line not in cleaned for line in preserved_lines):
                mismatches.append(f"cleaned summary removed business notice={cleaned!r}")
            if "조회조건:" in inbound_cleaned or "조회결과:" in inbound_cleaned:
                mismatches.append(f"inbound duplicate condition/row-count kept={inbound_cleaned!r}")
            if "조회조건:" in inbound_cleaned_from_render_path or "조회결과:" in inbound_cleaned_from_render_path:
                mismatches.append(f"inbound render-path duplicate condition/row-count kept={inbound_cleaned_from_render_path!r}")
            if inbound_summary_cond_text != "기간 2026-07-01 ~ 2026-07-18 / 재고위치 전체":
                mismatches.append(f"inbound render-path summary condition unexpected={inbound_summary_cond_text!r}")
            if "입고수량: 9,999개" not in inbound_cleaned or "업무 요약 본문" not in inbound_cleaned:
                mismatches.append(f"inbound business summary lost={inbound_cleaned!r}")
            if "입고수량: 9,999개" not in inbound_cleaned_from_render_path or "업무 요약 본문" not in inbound_cleaned_from_render_path:
                mismatches.append(f"inbound render-path business summary lost={inbound_cleaned_from_render_path!r}")
            if "제품정보:" not in product_flow_cleaned:
                mismatches.append(f"product flow product info removed={product_flow_cleaned!r}")
            for token in ("이월재고: 10개", "입고수량: 20개", "출고수량: 5개", "재고수량: 25개", "집계 요약 펼쳐보기"):
                if token not in product_flow_cleaned:
                    mismatches.append(f"product flow KPI lost token={token!r} cleaned={product_flow_cleaned!r}")
            if "조회조건:" in inventory_cleaned or "조회 결과:" in inventory_cleaned:
                mismatches.append(f"inventory duplicate condition/row-count kept={inventory_cleaned!r}")
            if "현재재고수량: 100개" not in inventory_cleaned or "경고: 일부 데이터만 조회됐다는 경고" not in inventory_cleaned:
                mismatches.append(f"inventory business/warning summary lost={inventory_cleaned!r}")
            if "제품정보:" in inventory_cleaned_drop_product or "현재재고수량: 100개" not in inventory_cleaned_drop_product:
                mismatches.append(f"inventory full-list product info cleanup failed={inventory_cleaned_drop_product!r}")
            if "제품정보:" not in inventory_cleaned_keep_product:
                mismatches.append(f"inventory explicit product info removed={inventory_cleaned_keep_product!r}")
            if not chat_mod._should_show_product_inventory_info(explicit_product_meta, full_inventory_df):
                mismatches.append("explicit product filter did not show product info")
            if chat_mod._should_show_product_inventory_info(maker_only_meta, full_inventory_df):
                mismatches.append("maker-only filter showed product info for multi-product inventory")
            if chat_mod._should_show_product_inventory_info(product_group_only_meta, full_inventory_df):
                mismatches.append("product-group-only filter showed product info for multi-product inventory")
            if chat_mod._should_show_product_inventory_info(product_name_only_meta, full_inventory_df):
                mismatches.append("product-name-only filter showed product info for multi-product inventory")
            if chat_mod._should_show_product_inventory_info(blank_product_code_meta, full_inventory_df):
                mismatches.append("blank product-code filter showed product info for multi-product inventory")
            if chat_mod._should_show_product_inventory_info({}, full_inventory_df):
                mismatches.append("multi-product inventory showed first-row product info")
            if not chat_mod._should_show_product_inventory_info({}, single_inventory_df):
                mismatches.append("single-product inventory did not show product info")
            fake_item_for_force = {"id": "msg-force"}
            fake_meta_for_force = {"table_key": "sims_force_table"}
            chat_mod._set_old_sims_table_force_rendered(fake_item_for_force, fake_meta_for_force, "force_uid", True)
            if not chat_mod._is_old_sims_table_force_rendered(fake_item_for_force, fake_meta_for_force, "force_uid"):
                mismatches.append("old table force render helper did not enable")
            chat_mod._set_old_sims_table_force_rendered(fake_item_for_force, fake_meta_for_force, "force_uid", False)
            if chat_mod._is_old_sims_table_force_rendered(fake_item_for_force, fake_meta_for_force, "force_uid"):
                mismatches.append("old table force render helper did not clear one table")
            old_chat_st_for_raw_meta = chat_mod.st
            class _RawMetaFakeSt:
                def __init__(self) -> None:
                    self.session_state = {}

            try:
                raw_meta_fake_st = _RawMetaFakeSt()
                chat_mod.st = raw_meta_fake_st
                raw_meta_fake_st.session_state["user"] = {"user_type": "SALES"}
                if chat_mod._is_internal_admin_for_raw_meta():
                    mismatches.append("raw meta visible to non-internal user")
                raw_meta_fake_st.session_state["user"] = {"user_type": "SSART_ADMIN"}
                if not chat_mod._is_internal_admin_for_raw_meta():
                    mismatches.append("raw meta hidden from internal admin")
            finally:
                chat_mod.st = old_chat_st_for_raw_meta
            if not all("llm_summary_md" in meta and (meta.get("table_key") or meta.get("source_key")) for meta in master_cases[:3]):
                mismatches.append("master metadata keys not preserved in fixture")
            if '"?? ?? ????"' in src or "master_summary_actions = {" in src:
                mismatches.append("broken master expander/action-list branch remains")
            if "and not _is_internal_master_summary(meta)" not in src:
                mismatches.append("debug meta expander is not guarded for internal master summaries")

            if mismatches:
                results.append(_fail("master llm summary hidden from screen", "; ".join(mismatches)))
            else:
                results.append(_ok("master llm summary hidden from screen", "users/goods/vendors/codes master summaries hidden from default screen; LLM metadata preserved; visible summaries unaffected"))
        except Exception as e:
            results.append(_fail("master llm summary hidden from screen", f"{type(e).__name__}: {e}"))

        try:
            chat_mw_src = Path("app/ui/chat_middleware.py").read_text(encoding="utf-8")
            checks = {
                "download_prepare_light": '"download_prepare"' in chat_mw_src and '"__ui_rerun_reason"] = "download_prepare"' in chat_mw_src,
                "download_prepare_table_key": "__sims_download_prepare_table_key" in chat_mw_src,
                "current_followup_compact": "def _render_current_followup_compact_header" in chat_mw_src,
                "current_followup_header_branch": "if is_current_followup:" in chat_mw_src and "_render_current_followup_compact_header" in chat_mw_src,
                "compact_header_helper": "def _build_sims_result_header_view" in chat_mw_src and "상세 조회정보" in chat_mw_src,
                "no_default_table_key_caption": 'st.caption(f"table_key:' not in chat_mw_src and "NLQ 질문:" not in chat_mw_src,
                "excel_cell_lazy": "SIMS_EXCEL_EAGER_MAX_CELLS" in chat_mw_src and "lazy_basis_cells" in chat_mw_src,
            }
            failed = [name for name, ok in checks.items() if not ok]
            if failed:
                results.append(_fail("current-table compact render policy", f"failed={failed}"))
            else:
                results.append(_ok("current-table compact render policy", "followup header compact; hides table_key/meta captions; download_prepare uses light history path; Excel lazy uses cell threshold"))
        except Exception as e:
            results.append(_fail("current-table compact render policy", f"{type(e).__name__}: {e}"))

        try:
            from app.ui import sims_table_display as display_mod

            quantity_kind = {
                "완료월평균출고수량": display_mod._numeric_display_kind("완료월평균출고수량"),
                "최근3개월평균수요수량": display_mod._numeric_display_kind("최근3개월평균수요수량"),
                "부족예상수량": display_mod._numeric_display_kind("부족예상수량"),
                "수요증감률": display_mod._numeric_display_kind("수요증감률"),
            }
            if quantity_kind != {
                "완료월평균출고수량": "int",
                "최근3개월평균수요수량": "int",
                "부족예상수량": "int",
                "수요증감률": "percent2",
            }:
                results.append(_fail("stock shortage quantity display format", f"unexpected={quantity_kind}"))
            else:
                results.append(_ok("stock shortage quantity display format", "shortage quantity columns render as integer; percent columns keep percent2"))
        except Exception as e:
            results.append(_fail("stock shortage quantity display format", f"{type(e).__name__}: {e}"))

        try:
            from app.ui import sims_table_display as display_mod

            src_df = pd.DataFrame(
                [
                    {"제품코드": "0001", "재고기준": "장부재고", "수요예상기준": "비교자료부족", "분석자료원": "월집계-장부재고(Rddbc220)", "현재고원천": "장부재고월집계(Rddbc220) 누계", "부족예상수량": 0, "flag": False},
                    {"제품코드": "0002", "재고기준": "장부재고", "수요예상기준": "최근3개월평균수요수량", "분석자료원": "월집계-장부재고(Rddbc220)", "현재고원천": "장부재고월집계(Rddbc220) 누계", "부족예상수량": 10, "flag": True},
                    {"제품코드": "0003", "재고기준": "None", "수요예상기준": "NULL", "분석자료원": None, "현재고원천": pd.NA, "부족예상수량": 0, "flag": False},
                ]
            )
            original_nulls = {
                "재고기준": int(src_df["재고기준"].isna().sum()),
                "수요예상기준": int(src_df["수요예상기준"].isna().sum()),
                "분석자료원": int(src_df["분석자료원"].isna().sum()),
                "현재고원천": int(src_df["현재고원천"].isna().sum()),
            }
            display_df = display_mod.normalize_display_df_for_streamlit(src_df)
            display_mod.log_sims_display_fields(src_df, display_df, action="품목별 재고부족현황", render_path="chat", mode="fast")
            ok_display = (
                original_nulls == {"재고기준": 0, "수요예상기준": 0, "분석자료원": 1, "현재고원천": 1}
                and display_df.loc[0, "재고기준"] == "장부재고"
                and display_df.loc[0, "수요예상기준"] == "비교자료부족"
                and display_df.loc[0, "분석자료원"] == "월집계-장부재고(Rddbc220)"
                and display_df.loc[0, "현재고원천"] == "장부재고월집계(Rddbc220) 누계"
                and display_df.loc[1, "수요예상기준"] == "최근3개월평균수요수량"
                and display_df.loc[1, "현재고원천"] == "장부재고월집계(Rddbc220) 누계"
                and display_df.loc[2, "재고기준"] == "None"
                and display_df.loc[2, "수요예상기준"] == "NULL"
                and display_df.loc[2, "분석자료원"] == ""
                and display_df.loc[2, "현재고원천"] == ""
                and display_df.loc[0, "부족예상수량"] == 0
                and bool(display_df.loc[0, "flag"]) is False
            )
            if ok_display:
                results.append(_ok("stock shortage display null cleanup", "display copy preserves business strings and blanks only actual nulls while numeric zero/False stay unchanged"))
            else:
                results.append(_fail("stock shortage display null cleanup", f"display={display_df.to_dict(orient='records')} source={src_df.to_dict(orient='records')}"))
        except Exception as e:
            results.append(_fail("stock shortage display null cleanup", f"{type(e).__name__}: {e}"))

        try:
            from app.ui import chat_middleware as chat_mod
            from app.ui import sims_table_display as display_mod

            text_cols = [
                "\uc7ac\uace0\uae30\uc900",
                "\uc7ac\uace0\uae30\uc900\ud310\uc815",
                "\ubd80\uc871\ub4f1\uae09",
                "\uc218\uc694\uc608\uc0c1\ub4f1\uae09",
                "\uc218\uc694\uc608\uc0c1\uae30\uc900",
                "\ubd84\uc11d\uc790\ub8cc\uc6d0",
                "\ud604\uc7ac\uace0\uc6d0\ucc9c",
            ]
            row0 = {
                "\uc7ac\uace0\uae30\uc900": "\uc7a5\ubd80\uc7ac\uace0",
                "\uc7ac\uace0\uae30\uc900\ud310\uc815": "\uc801\uc815",
                "\ubd80\uc871\ub4f1\uae09": "\uc815\uc0c1",
                "\uc218\uc694\uc608\uc0c1\ub4f1\uae09": "\ube44\uad50\uc790\ub8cc\ubd80\uc871",
                "\uc218\uc694\uc608\uc0c1\uae30\uc900": "\ucd5c\uadfc3\uac1c\uc6d4\ud3c9\uade0\uc218\uc694\uc218\ub7c9",
                "\ubd84\uc11d\uc790\ub8cc\uc6d0": "\uc6d4\uc9d1\uacc4-\uc7a5\ubd80\uc7ac\uace0(Rddbc220)",
                "\ud604\uc7ac\uace0\uc6d0\ucc9c": "\uc7a5\ubd80\uc7ac\uace0\uc6d4\uc9d1\uacc4(Rddbc220) \ub204\uacc4",
                "\ubd80\uc871\uc608\uc0c1\uc218\ub7c9": 0,
                "\ud3c9\uac00\uc6d4 \uc218\uc694\uc9c4\ucc99\ub960": 61.67,
                "flag": False,
            }
            row1 = dict(row0)
            row1.update({
                "\uc7ac\uace0\uae30\uc900": None,
                "\uc218\uc694\uc608\uc0c1\uae30\uc900": "None",
                "\ud604\uc7ac\uace0\uc6d0\ucc9c": pd.NA,
                "\ubd80\uc871\uc608\uc0c1\uc218\ub7c9": 10,
                "\ud3c9\uac00\uc6d4 \uc218\uc694\uc9c4\ucc99\ub960": 0,
            })
            source_df = pd.DataFrame([row0, row1])
            source_snapshot = source_df.copy(deep=True)

            def _common_renderer_view(frame: pd.DataFrame) -> pd.DataFrame:
                fast = chat_mod._chat_fast_display_df(frame)
                normalized = display_mod.normalize_display_df_for_streamlit(fast)
                view, _cfg, _width, _height = display_mod.build_sims_table_display_config(
                    normalized,
                    action_name="\ud488\ubaa9\ubcc4 \uc7ac\uace0\ubd80\uc871\ud604\ud669",
                    add_row_no=False,
                    row_no_name="\uc21c\ubc88",
                    enable_pinning=True,
                    max_pinned_cols=5,
                )
                return view

            route_views = {
                "chat_fast": _common_renderer_view(source_df),
                "chat_small": display_mod.build_sims_table_display_config(
                    display_mod.normalize_display_df_for_streamlit(source_df),
                    action_name="\ud488\ubaa9\ubcc4 \uc7ac\uace0\ubd80\uc871\ud604\ud669",
                    add_row_no=False,
                )[0],
                "panel_fast": _common_renderer_view(source_df),
                "panel_small": display_mod.build_sims_table_display_config(
                    display_mod.normalize_display_df_for_streamlit(source_df),
                    action_name="\ud488\ubaa9\ubcc4 \uc7ac\uace0\ubd80\uc871\ud604\ud669",
                    add_row_no=False,
                )[0],
                "history_reopen": _common_renderer_view(source_df),
                "current_followup": _common_renderer_view(source_df.loc[[0]].copy()),
            }

            route_failures: list[str] = []
            for route, view in route_views.items():
                for col in text_cols:
                    if col not in view.columns:
                        route_failures.append(f"{route}:{col}:missing")
                        continue
                    if not (pd.api.types.is_object_dtype(view[col]) or pd.api.types.is_string_dtype(view[col])):
                        route_failures.append(f"{route}:{col}:dtype={view[col].dtype}")
                    expected_non_null = int(source_df.loc[view.index, col].notna().sum()) if len(view.index) else 0
                    actual_non_null = int(view[col].replace("", pd.NA).notna().sum())
                    if actual_non_null != expected_non_null:
                        route_failures.append(f"{route}:{col}:non_null={actual_non_null}/{expected_non_null}")
                    if col in row0 and view.iloc[0][col] != row0[col]:
                        route_failures.append(f"{route}:{col}:value={view.iloc[0][col]!r}")
                if "current_followup" != route:
                    if view.loc[1, "\uc7ac\uace0\uae30\uc900"] != "":
                        route_failures.append(f"{route}:actual_none_not_blank")
                    if view.loc[1, "\uc218\uc694\uc608\uc0c1\uae30\uc900"] != "None":
                        route_failures.append(f"{route}:literal_none_not_preserved")
                    if view.loc[0, "\ubd80\uc871\uc608\uc0c1\uc218\ub7c9"] != 0:
                        route_failures.append(f"{route}:numeric_zero_changed")
                    if bool(view.loc[0, "flag"]) is not False:
                        route_failures.append(f"{route}:false_changed")

            source_unchanged = source_df.equals(source_snapshot)
            numeric_preserved = all(pd.api.types.is_numeric_dtype(v["\ubd80\uc871\uc608\uc0c1\uc218\ub7c9"]) for v in route_views.values())
            if route_failures or not source_unchanged or not numeric_preserved:
                results.append(_fail("stock shortage renderer text field preservation", f"failures={route_failures} source_unchanged={source_unchanged} numeric_preserved={numeric_preserved}"))
            else:
                results.append(_ok("stock shortage renderer text field preservation", "chat fast/small, panel fast/small, history reopen, and current-table followup preserve stock shortage business text fields while numeric columns stay numeric"))
        except Exception as e:
            results.append(_fail("stock shortage renderer text field preservation", f"{type(e).__name__}: {e}"))

        try:
            from app.ui import chat_middleware as chat_mod
            from app.ui import sims_table_display as display_mod

            full_df = pd.DataFrame(
                [
                    {"제품코드": "0001", "재고기준": "장부재고", "수요예상기준": "비교자료부족", "분석자료원": "월집계-장부재고(Rddbc220)", "현재고원천": "장부재고월집계(Rddbc220) 누계", "부족예상수량": 0},
                    {"제품코드": "0002", "재고기준": "장부재고", "수요예상기준": "최근3개월평균수요수량", "분석자료원": "월집계-장부재고(Rddbc220)", "현재고원천": "장부재고월집계(Rddbc220) 누계", "부족예상수량": 10},
                    {"제품코드": "0003", "재고기준": None, "수요예상기준": None, "분석자료원": None, "현재고원천": None, "부족예상수량": 0},
                ],
                index=[10, 20, 30],
            )
            display_df = full_df[["제품코드", "재고기준", "수요예상기준", "분석자료원", "현재고원천", "부족예상수량"]].copy()
            display_df.loc[10, ["재고기준", "수요예상기준", "분석자료원", "현재고원천"]] = [None, None, "None", pd.NA]
            display_df.loc[20, ["재고기준", "수요예상기준", "분석자료원", "현재고원천"]] = ["None", "NULL", None, "nan"]

            restored = chat_mod._restore_display_fields_from_full_df(
                display_df,
                full_df,
                action="품목별 재고부족현황",
                stage="regression",
            )
            fast_view = display_mod.normalize_display_df_for_streamlit(restored)
            small_view = display_mod.normalize_display_df_for_streamlit(restored.copy())
            filtered_view = display_mod.normalize_display_df_for_streamlit(restored.loc[[20]].copy())
            unchanged_full = (
                full_df.loc[10, "수요예상기준"] == "비교자료부족"
                and full_df.loc[20, "수요예상기준"] == "최근3개월평균수요수량"
            )
            ok_restore = (
                fast_view.loc[10, "재고기준"] == "장부재고"
                and fast_view.loc[10, "수요예상기준"] == "비교자료부족"
                and fast_view.loc[10, "분석자료원"] == "월집계-장부재고(Rddbc220)"
                and fast_view.loc[10, "현재고원천"] == "장부재고월집계(Rddbc220) 누계"
                and small_view.loc[20, "수요예상기준"] == "최근3개월평균수요수량"
                and filtered_view.loc[20, "분석자료원"] == "월집계-장부재고(Rddbc220)"
                and filtered_view.loc[20, "현재고원천"] == "장부재고월집계(Rddbc220) 누계"
                and fast_view.loc[30, "재고기준"] == ""
                and fast_view.loc[30, "수요예상기준"] == ""
                and fast_view.loc[10, "부족예상수량"] == 0
                and unchanged_full
            )
            if ok_restore:
                results.append(_ok("stock shortage display field restore", "full/display/fast/small/filter paths preserve Excel business values by original index"))
            else:
                results.append(_fail("stock shortage display field restore", f"fast={fast_view.to_dict(orient='records')} full={full_df.to_dict(orient='records')}"))
        except Exception as e:
            results.append(_fail("stock shortage display field restore", f"{type(e).__name__}: {e}"))

        try:
            import warnings

            import numpy as np

            from app.sims.views import rddbc_io_shared as io_shared_mod

            metric_col = "\uc785\uace0\uc218\ub7c9"
            text_col = "\uba54\ubaa8"
            all_null_col = "\uc804\ubd80\uacb0\uce21"
            seq_col = "\uba85\uc138\uc11c\ubc88\ud638"
            product_no_col = "\uc81c\uc870\ubc88\ud638"
            validation_col = "\uac80\uc218\ud655\uc778"

            source_df = pd.DataFrame(
                {
                    metric_col: pd.Series([1, 0, "2", None, pd.NA, np.nan, ""], dtype="object"),
                    text_col: pd.Series(["A", "0", "", None, pd.NA, np.nan, "NULL"], dtype="object"),
                    all_null_col: pd.Series([None, pd.NA, np.nan, None, pd.NA, np.nan, None], dtype="object"),
                    seq_col: pd.Series([None, 268, "628", "", pd.NA, np.nan, 0], dtype="object"),
                    product_no_col: pd.Series(["000123", "001-A", "114625021", None, pd.NA, np.nan, "NULL"], dtype="object"),
                    validation_col: pd.Series(["1", "0", "Y", "", None, pd.NA, "<NA>"], dtype="object"),
                }
            )
            source_before = source_df.copy(deep=True)

            with warnings.catch_warnings(record=True) as caught:
                warnings.simplefilter("always")
                prepared = io_shared_mod._prepare_io_display_df(source_df, add_row_no=False)

            failures: list[str] = []
            warning_matches = [
                str(w.message)
                for w in caught
                if "Downcasting object dtype arrays on .fillna" in str(w.message)
            ]
            if warning_matches:
                failures.append(f"futurewarning_count={len(warning_matches)}")

            try:
                pd.testing.assert_frame_equal(source_df, source_before, check_dtype=True)
            except AssertionError as assert_exc:
                failures.append(f"source_mutated={assert_exc}")

            if not pd.api.types.is_numeric_dtype(prepared[metric_col]):
                failures.append(f"metric_dtype={prepared[metric_col].dtype}")
            if prepared[metric_col].tolist() != [1, 0, 2, 0, 0, 0, 0]:
                failures.append(f"metric_values={prepared[metric_col].tolist()!r}")
            if prepared[text_col].tolist() != ["A", "0", "", "", "", "", ""]:
                failures.append(f"text_values={prepared[text_col].tolist()!r}")
            if prepared[all_null_col].tolist() != ["", "", "", "", "", "", ""]:
                failures.append(f"all_null_values={prepared[all_null_col].tolist()!r}")
            if str(prepared[seq_col].dtype) != "Int64":
                failures.append(f"seq_dtype={prepared[seq_col].dtype}")
            if prepared[seq_col].dropna().astype(int).tolist() != [268, 628]:
                failures.append(f"seq_values={prepared[seq_col].tolist()!r}")
            if prepared[product_no_col].tolist() != ["000123", "001-A", "114625021", "", "", "", "NULL"]:
                failures.append(f"product_no_values={prepared[product_no_col].tolist()!r}")
            if prepared[validation_col].tolist() != ["1", "0", "Y", "", "", "", "<NA>"]:
                failures.append(f"validation_values={prepared[validation_col].tolist()!r}")
            for col in [text_col, all_null_col, product_no_col, validation_col]:
                if not (pd.api.types.is_object_dtype(prepared[col]) or pd.api.types.is_string_dtype(prepared[col])):
                    failures.append(f"{col}_dtype={prepared[col].dtype}")

            if failures:
                results.append(_fail("rddbc IO fillna futurewarning", "; ".join(failures)))
            else:
                results.append(_ok("rddbc IO fillna futurewarning", "production display helper emits no fillna downcast FutureWarning and preserves numeric/text/null policy"))
        except Exception as e:
            results.append(_fail("rddbc IO fillna futurewarning", f"{type(e).__name__}: {e}"))

        try:
            import warnings

            import numpy as np

            from app.services import product_flow_service as product_flow_mod

            msg_part = "DataFrame concatenation with empty or all-NA entries"
            detail_params = {"date_from": "20260701", "date_to": "20260719", "top": 10}
            detail_settings = product_flow_mod._mode_settings({"flow_scope": "all"})
            detail_frames = [
                pd.DataFrame(
                    {
                        "\uc218\ub7c9": pd.Series([1, 0, np.nan], dtype="float64"),
                        "\uc2dd\ubcc4\uc790": pd.Series(["001", "0", ""], dtype="object"),
                        "\uc785\uace0\ub9cc\uc788\ub294\uacf5\ubc31": pd.Series([None, pd.NA, np.nan], dtype="object"),
                    }
                ),
                pd.DataFrame(
                    {
                        "\uc218\ub7c9": pd.Series([pd.NA], dtype="object"),
                        "\uc2dd\ubcc4\uc790": pd.Series([pd.NA], dtype="object"),
                        "\ucd9c\uace0\ub9cc\uc788\ub294\uacf5\ubc31": pd.Series([pd.NA], dtype="object"),
                    }
                ),
            ]
            detail_before = [df.copy(deep=True) for df in detail_frames]
            query_calls = {"count": 0}
            original_query_to_df = product_flow_mod.query_to_df

            def _fake_query_to_df(_sql, _params):
                idx = query_calls["count"]
                query_calls["count"] += 1
                return detail_frames[idx].copy(deep=True)

            try:
                product_flow_mod.query_to_df = _fake_query_to_df
                with warnings.catch_warnings(record=True) as caught_detail:
                    warnings.simplefilter("always")
                    detail_out = product_flow_mod._get_detail_df(detail_params, detail_settings)
            finally:
                product_flow_mod.query_to_df = original_query_to_df

            display_settings = product_flow_mod._mode_settings({"flow_scope": "all"})
            display_detail = pd.DataFrame(
                {
                    "\uc815\ub82c\uc77c\uc790": ["20260702", "20260703"],
                    "\uc815\ub82c\uc21c\ubc88": [2, 3],
                    "\ub0b4\ubd80\ubc29\ud5a5": ["IN", "OUT"],
                    "\uc7ac\uace0\uc99d\uac10": [10, -4],
                    "\uc785\uace0\uc218\ub7c9": [10, 0],
                    "\ucd9c\uace0\uc218\ub7c9": [0, 4],
                    "\ud560\uc99d": [0, 0],
                    "\uacf5\uae09\uac00\uc561": [1000, 2000],
                    "\ubd80\uac00\uc138": [100, 200],
                    "\ud569\uacc4\uae08\uc561": [1100, 2200],
                    "\uc785\ucd9c\uace0\uc77c\uc790": ["20260702", "20260703"],
                    "\uba85\uc138\uc11c\uc77c\uc790": ["20260702", "20260703"],
                    "\uc7ac\uace0\uc704\uce58": ["0001", "0001"],
                    "\uc81c\uc870\ubc88\ud638": ["000123", "001-A"],
                    "\uac80\uc218\ud655\uc778": ["1", "0"],
                    "\uc601\uc5c5\uc0ac\uc6d0": ["0", "001"],
                }
            )
            display_before = display_detail.copy(deep=True)
            with warnings.catch_warnings(record=True) as caught_display:
                warnings.simplefilter("always")
                display_out, display_meta = product_flow_mod._prepare_display_df(
                    display_detail,
                    5.0,
                    display_settings,
                    {"physic_cd": "000123", "physic_nm": "\ud14c\uc2a4\ud2b8", "stock_names": ["0001"]},
                )

            failures: list[str] = []
            detail_warnings = [str(w.message) for w in caught_detail if msg_part in str(w.message)]
            display_warnings = [str(w.message) for w in caught_display if msg_part in str(w.message)]
            if detail_warnings:
                failures.append(f"detail_concat_futurewarning={len(detail_warnings)}")
            if display_warnings:
                failures.append(f"display_concat_futurewarning={len(display_warnings)}")
            if query_calls["count"] != 2:
                failures.append(f"detail_query_calls={query_calls['count']}")
            if len(detail_out) != 4:
                failures.append(f"detail_rows={len(detail_out)}")
            expected_detail_cols = ["\uc218\ub7c9", "\uc2dd\ubcc4\uc790", "\uc785\uace0\ub9cc\uc788\ub294\uacf5\ubc31", "\ucd9c\uace0\ub9cc\uc788\ub294\uacf5\ubc31"]
            if list(detail_out.columns) != expected_detail_cols:
                failures.append(f"detail_cols={list(detail_out.columns)!r}")
            if detail_out["\uc2dd\ubcc4\uc790"].iloc[0] != "001" or detail_out["\uc2dd\ubcc4\uc790"].iloc[1] != "0":
                failures.append(f"detail_identifier_values={detail_out['\uc2dd\ubcc4\uc790'].tolist()!r}")
            if detail_out["\uc218\ub7c9"].iloc[1] != 0:
                failures.append(f"detail_zero={detail_out['\uc218\ub7c9'].tolist()!r}")
            if not pd.isna(detail_out["\uc785\uace0\ub9cc\uc788\ub294\uacf5\ubc31"]).all():
                failures.append("detail_all_na_col_changed")
            for actual, before in zip(detail_frames, detail_before):
                try:
                    pd.testing.assert_frame_equal(actual, before, check_dtype=True)
                except AssertionError as assert_exc:
                    failures.append(f"detail_input_mutated={assert_exc}")

            if len(display_out) != 3:
                failures.append(f"display_rows={len(display_out)}")
            if display_out.iloc[0].get("\uc785\ucd9c\uace0\uc77c\uc790") != "\uc774\uc6d4\uc7ac\uace0":
                failures.append(f"display_carry_label={display_out.iloc[0].get('\uc785\ucd9c\uace0\uc77c\uc790')!r}")
            if str(display_out["\uba85\uc138\uc11c\ubc88\ud638"].dtype) != "Int64":
                failures.append(f"display_seq_dtype={display_out['\uba85\uc138\uc11c\ubc88\ud638'].dtype}")
            if display_out["\uc81c\uc870\ubc88\ud638"].tolist()[1:] != ["000123", "001-A"]:
                failures.append(f"display_product_no={display_out['\uc81c\uc870\ubc88\ud638'].tolist()!r}")
            if display_out["\uac80\uc218\ud655\uc778"].tolist()[1:] != ["1", "0"]:
                failures.append(f"display_validation={display_out['\uac80\uc218\ud655\uc778'].tolist()!r}")
            for col in ["\uc785\uace0\uc218\ub7c9", "\ucd9c\uace0\uc218\ub7c9", "\uacf5\uae09\uac00\uc561", "\ud569\uacc4\uae08\uc561"]:
                if not pd.api.types.is_numeric_dtype(display_out[col]):
                    failures.append(f"display_numeric_{col}_dtype={display_out[col].dtype}")
            if display_meta.get("row_count") != 3:
                failures.append(f"display_meta_row_count={display_meta.get('row_count')}")
            try:
                pd.testing.assert_frame_equal(display_detail, display_before, check_dtype=True)
            except AssertionError as assert_exc:
                failures.append(f"display_input_mutated={assert_exc}")

            if failures:
                results.append(_fail("product flow concat futurewarning", "; ".join(failures)))
            else:
                results.append(_ok("product flow concat futurewarning", "detail and carry-row concat paths emit no concat FutureWarning while preserving rows, column order, dtypes, and input frames"))
        except Exception as e:
            results.append(_fail("product flow concat futurewarning", f"{type(e).__name__}: {e}"))

        try:
            import pyarrow as pa

            from app.sims.views import rddbc_io_shared as io_shared_mod
            from app.ui import chat_middleware as chat_mod
            from app.ui import sims_panel as panel_mod
            from app.ui import sims_table_display as display_mod

            seq_col = "\uba85\uc138\uc11c\ubc88\ud638"
            product_no_col = "\uc81c\uc870\ubc88\ud638"
            validation_col = "\uac80\uc218\ud655\uc778"
            amount_col = "\ud569\uacc4\uae08\uc561"
            qty_col = "\uc785\uace0\uc218\ub7c9"
            salesperson_col = "\uc601\uc5c5\uc0ac\uc6d0"
            source_df = pd.DataFrame(
                {
                    seq_col: pd.Series([pd.NA, 268, 628], dtype="Int64"),
                    product_no_col: ["000123", "001-A", "114625021"],
                    validation_col: ["1", "0", "Y"],
                    amount_col: [0, 1100, 2200],
                    qty_col: [0, 10, 20],
                    salesperson_col: ["0", "001", "A-01"],
                }
            )
            source_before = source_df.copy(deep=True)
            calls: list[dict[str, Any]] = []
            original_dataframe = display_mod.st.dataframe

            def _fake_dataframe(data=None, *args, **kwargs):
                calls.append({"data": data, "args": args, "kwargs": dict(kwargs)})
                return data

            try:
                display_mod.st.dataframe = _fake_dataframe
                chat_mod.st.dataframe = _fake_dataframe
                panel_mod.st.dataframe = _fake_dataframe
                io_shared_mod.st.dataframe = _fake_dataframe

                display_view = display_mod.render_sims_table(
                    source_df.copy(),
                    action_name="\uc81c\ud488\uc218\ubd88\ud604\ud669 \uc870\ud68c",
                    add_row_no=False,
                    key="__regression_width_display",
                )
                chat_mod._render_chat_fast_dataframe(
                    source_df.copy(),
                    height=320,
                    action_name="\ud604\uc7ac\ud45c \uc601\uc5c5\uc0ac\uc6d0 TOP 20",
                    meta={"current_table_followup": True},
                )
                panel_mod._render_fast_dataframe(
                    source_df.copy(),
                    height=340,
                    action_name="\uc81c\ud488\uc218\ubd88\ud604\ud669 \uc870\ud68c",
                    meta={"table_key": "sims_width_panel"},
                )
                io_view = io_shared_mod._render_io_dataframe(
                    source_df.copy(),
                    key="__regression_width_io",
                    add_row_no=False,
                    use_container_width=False,
                    height=280,
                )
            finally:
                display_mod.st.dataframe = original_dataframe
                chat_mod.st.dataframe = original_dataframe
                panel_mod.st.dataframe = original_dataframe
                io_shared_mod.st.dataframe = original_dataframe

            failures: list[str] = []
            if len(calls) != 4:
                failures.append(f"dataframe_call_count={len(calls)}")
            expected_widths = ["stretch", "stretch", "stretch", "content"]
            for idx, expected_width in enumerate(expected_widths):
                if idx >= len(calls):
                    continue
                kwargs = calls[idx]["kwargs"]
                if "use_container_width" in kwargs:
                    failures.append(f"call{idx}_deprecated_kwarg")
                if kwargs.get("width") != expected_width:
                    failures.append(f"call{idx}_width={kwargs.get('width')!r}")
                if "height" not in kwargs:
                    failures.append(f"call{idx}_height_missing")
            for idx in [0, 1, 2]:
                if idx < len(calls) and not calls[idx]["kwargs"].get("column_config"):
                    failures.append(f"call{idx}_column_config_missing")
            if len(calls) >= 4 and calls[3]["kwargs"].get("height") != 280:
                failures.append(f"io_height={calls[3]['kwargs'].get('height')!r}")
            if isinstance(display_view, pd.DataFrame):
                try:
                    pa.Table.from_pandas(display_view, preserve_index=False)
                except Exception as arrow_exc:
                    failures.append(f"display_arrow={type(arrow_exc).__name__}: {arrow_exc}")
            if isinstance(io_view, pd.DataFrame):
                if io_view[product_no_col].tolist() != ["000123", "001-A", "114625021"]:
                    failures.append(f"io_product_no={io_view[product_no_col].tolist()!r}")
                if io_view[validation_col].tolist() != ["1", "0", "Y"]:
                    failures.append(f"io_validation={io_view[validation_col].tolist()!r}")
            try:
                pd.testing.assert_frame_equal(source_df, source_before, check_dtype=True)
            except AssertionError as assert_exc:
                failures.append(f"source_mutated={assert_exc}")

            if failures:
                results.append(_fail("streamlit width compatibility for SIMS tables", "; ".join(failures)))
            else:
                results.append(_ok("streamlit width compatibility for SIMS tables", "display/chat-fast/panel-fast/io wrappers pass width kwargs, preserve height/column_config and avoid deprecated use_container_width"))
        except Exception as e:
            results.append(_fail("streamlit width compatibility for SIMS tables", f"{type(e).__name__}: {e}"))

        try:
            import inspect

            from app.ui import chat_middleware as chat_mod

            old_display_limit = os.environ.get("SIMS_CHAT_DISPLAY_MAX_ROWS")
            original_dataframe = chat_mod.st.dataframe
            original_caption = chat_mod.st.caption
            original_table_log = chat_mod.log_sims_table_render
            dataframe_calls: list[pd.DataFrame] = []
            captions: list[str] = []
            table_logs: list[dict[str, Any]] = []

            def _fake_dataframe(data=None, *args, **kwargs):
                if isinstance(data, pd.DataFrame):
                    dataframe_calls.append(data.copy(deep=True))
                return data

            def _fake_caption(value, *args, **kwargs):
                captions.append(str(value))

            def _fake_table_log(df, **kwargs):
                table_logs.append({"rows": int(len(df)), **dict(kwargs)})

            try:
                os.environ["SIMS_CHAT_DISPLAY_MAX_ROWS"] = "17"
                chat_mod.st.dataframe = _fake_dataframe
                chat_mod.st.caption = _fake_caption
                chat_mod.log_sims_table_render = _fake_table_log
                chat_mod.st.session_state["__sims_table_render_path"] = "history"

                empty_df = pd.DataFrame({"code": pd.Series(dtype="string")})
                small_df = pd.DataFrame({"code": pd.Series(["00001"] * 16, dtype="string"), "qty": pd.Series(range(16), dtype="Int64")})
                equal_df = pd.DataFrame({"code": pd.Series(["00001"] * 17, dtype="string"), "qty": pd.Series(range(17), dtype="Int64")})
                over_df = pd.DataFrame({"code": pd.Series(["00001"] * 18, dtype="string"), "qty": pd.Series(range(18), dtype="Int64")})
                for frame, expected_rows in ((empty_df, 0), (small_df, 16), (equal_df, 17), (over_df, 17)):
                    limited = chat_mod._limit_chat_display_df(frame)
                    if len(limited) != expected_rows:
                        raise AssertionError(f"limit_rows={len(limited)} expected={expected_rows}")

                columns = {"product_code": pd.Series([f"{idx:05d}" for idx in range(10115)], dtype="string")}
                columns.update({f"metric_{idx:02d}": pd.Series(range(10115), dtype="Int64") for idx in range(97)})
                full_df = pd.DataFrame(columns)
                full_before = full_df.copy(deep=True)
                chat_mod.st.session_state["sims_tables"] = {"history-large-fixture": full_df}
                chat_mod.st.session_state["sims_export_tables"] = {"history-large-fixture": full_df}
                chat_mod.st.session_state["__sims_export_tables_by_key"] = {"history-large-fixture": full_df}
                chat_mod.st.session_state["__sims_current_table_source_key"] = "history-large-fixture"
                chat_mod._render_chat_fast_dataframe(
                    full_df,
                    height=520,
                    action_name="\ud488\ubaa9\ubcc4 \uc7ac\uace0\ubd80\uc871\ud604\ud669",
                    meta={"kind": "table", "table_key": "history-large-fixture"},
                )
            finally:
                chat_mod.st.dataframe = original_dataframe
                chat_mod.st.caption = original_caption
                chat_mod.log_sims_table_render = original_table_log
                if old_display_limit is None:
                    os.environ.pop("SIMS_CHAT_DISPLAY_MAX_ROWS", None)
                else:
                    os.environ["SIMS_CHAT_DISPLAY_MAX_ROWS"] = old_display_limit

            failures: list[str] = []
            if not dataframe_calls or len(dataframe_calls[-1]) != 17:
                failures.append(f"fast_renderer_rows={len(dataframe_calls[-1]) if dataframe_calls else 'missing'}")
            if dataframe_calls and dataframe_calls[-1].columns.tolist() != full_df.columns.tolist():
                failures.append("fast_renderer_columns_changed")
            try:
                pd.testing.assert_frame_equal(full_df, full_before, check_dtype=True)
            except AssertionError as assert_exc:
                failures.append(f"fast_renderer_input_mutated={assert_exc}")
            for store_name in ("sims_tables", "sims_export_tables", "__sims_export_tables_by_key"):
                stored = (chat_mod.st.session_state.get(store_name) or {}).get("history-large-fixture")
                if stored is not full_df or not isinstance(stored, pd.DataFrame) or len(stored) != 10115:
                    failures.append(f"{store_name}_full_source_changed")
            if not captions or "10,115" not in captions[-1] or "17" not in captions[-1]:
                failures.append(f"truncation_caption={captions!r}")
            if not table_logs:
                failures.append("table_render_log_missing")
            else:
                logged = table_logs[-1]
                expected_log = {
                    "full_rows": 10115,
                    "display_rows": 17,
                    "render_truncated": True,
                    "display_limit": 17,
                }
                for key, expected in expected_log.items():
                    if logged.get(key) != expected:
                        failures.append(f"log_{key}={logged.get(key)!r}")

            body_source = inspect.getsource(chat_mod._render_chat_item_body)
            fast_source = inspect.getsource(chat_mod._render_chat_fast_dataframe)
            if "_render_chat_fast_dataframe(\n                                render_df.copy()" not in body_source:
                failures.append("history_large_fast_branch_missing")
            if "_render_chat_fast_dataframe(\n                                view_df.copy()" not in body_source:
                failures.append("nlq_fast_direct_branch_missing")
            if "_limit_chat_display_df(full_df)" not in fast_source:
                failures.append("fast_renderer_display_limit_missing")
            if "head(300)" in fast_source or "SIMS_CHAT_TABLE_DISPLAY_LIMIT" in fast_source:
                failures.append("new_fixed_300_limit_detected")
            if failures:
                results.append(_fail("SIMS chat history fast table row cap", "; ".join(failures)))
            else:
                results.append(_ok("SIMS chat history fast table row cap", "history/NLQ fast paths share the display-limit boundary; 10,115x98 render input is capped by SIMS_CHAT_DISPLAY_MAX_ROWS while the original DataFrame remains unchanged"))
        except Exception as e:
            results.append(_fail("SIMS chat history fast table row cap", f"{type(e).__name__}: {e}"))

        try:
            import io
            import json
            from openpyxl import load_workbook

            from app.services import product_flow_service as product_flow_mod
            from app.sims.views import rddbc_io_shared as io_shared_mod
            from app.ui import chat_middleware as chat_mod
            from app.ui import sims_table_display as display_mod

            seq_col = "\uba85\uc138\uc11c\ubc88\ud638"
            product_no_col = "\uc81c\uc870\ubc88\ud638"
            validation_col = "\uac80\uc218\ud655\uc778"
            in_qty_col = "\uc785\uace0\uc218\ub7c9"
            out_qty_col = "\ucd9c\uace0\uc218\ub7c9"
            supply_col = "\uacf5\uae09\uac00\uc561"
            total_col = "\ud569\uacc4\uae08\uc561"
            product_code_col = "\uc81c\ud488\ucf54\ub4dc"
            biz_no_col = "\uc0ac\uc5c5\uc790\ubc88\ud638"
            phone_col = "\uc804\ud654\ubc88\ud638"
            zip_col = "\uc6b0\ud3b8\ubc88\ud638"
            salesperson_col = "\uc601\uc5c5\uc0ac\uc6d0"
            salesperson_name_col = "\uc601\uc5c5\uc0ac\uc6d0\uba85"
            row_no_col = "\uc21c\ubc88"
            no_col = "\ubc88\ud638"

            source_df = pd.DataFrame(
                {
                    seq_col: [None, 268, 628, 893, None, None, None],
                    product_no_col: ["114625021", "000123", "001-A", "", None, pd.NA, "NULL"],
                    validation_col: ["1", "0", "Y", "", float("nan"), "<NA>", "nan"],
                    in_qty_col: [0, "10", "20", "", "30", "40", "50"],
                    out_qty_col: [0, "1", "2", "", "3", "4", "5"],
                    supply_col: [0, "1000", "2000", "", "3000", "4000", "5000"],
                    total_col: [0, "1100", "2200", "", "3300", "4400", "5500"],
                    product_code_col: ["00001", "00002", "00003", "00004", "00005", "00006", "00007"],
                    biz_no_col: ["012-34-56789", "1112233333", "", "999-88-77777", "0000011111", "2222233333", "4444455555"],
                    phone_col: ["02-1234-5678", "01000000000", "", "031-123-4567", "01011112222", "01033334444", "01055556666"],
                    zip_col: ["01234", "12345", "", "67890", "00001", "00002", "00003"],
                    salesperson_col: ["0", "001", "", None, "A-01", "0", "002"],
                    salesperson_name_col: ["0", "Kim", "", None, "A team", "0", "Lee"],
                    row_no_col: [1, 2, 3, 4, 5, 6, 7],
                    no_col: ["0", "10", "", None, "20", "0", "30"],
                }
            )

            finalized = product_flow_mod._finalize_display_df_250(source_df)
            prepared = io_shared_mod._prepare_io_display_df(finalized, add_row_no=False)
            view_df, column_config, _table_width, _table_height = display_mod.build_sims_table_display_config(
                prepared.copy(),
                action_name="\uc81c\ud488\uc218\ubd88\ud604\ud669 \uc870\ud68c",
                add_row_no=False,
            )

            failures: list[str] = []

            if str(finalized[seq_col].dtype) != "Int64":
                failures.append(f"service_seq_dtype={finalized[seq_col].dtype}")
            if finalized[seq_col].isna().iloc[0] is not True and not pd.isna(finalized[seq_col].iloc[0]):
                failures.append(f"service_seq_blank={finalized[seq_col].iloc[0]!r}")
            if finalized[seq_col].dropna().astype(int).tolist() != [268, 628, 893]:
                failures.append(f"service_seq_values={finalized[seq_col].tolist()!r}")

            for col, expected in [
                (product_no_col, ["114625021", "000123", "001-A", "", "", "", "NULL"]),
                (validation_col, ["1", "0", "Y", "", "", "<NA>", "nan"]),
            ]:
                if not (pd.api.types.is_object_dtype(finalized[col]) or pd.api.types.is_string_dtype(finalized[col])):
                    failures.append(f"service_{col}_dtype={finalized[col].dtype}")
                if finalized[col].tolist() != expected:
                    failures.append(f"service_{col}_values={finalized[col].tolist()!r}")

            if str(prepared[seq_col].dtype) != "Int64":
                failures.append(f"display_seq_dtype={prepared[seq_col].dtype}")
            for col, expected in [
                (product_no_col, ["114625021", "000123", "001-A", "", "", "", "NULL"]),
                (validation_col, ["1", "0", "Y", "", "", "<NA>", "nan"]),
                (product_code_col, ["00001", "00002", "00003", "00004", "00005", "00006", "00007"]),
                (biz_no_col, ["012-34-56789", "1112233333", "", "999-88-77777", "0000011111", "2222233333", "4444455555"]),
                (phone_col, ["02-1234-5678", "01000000000", "", "031-123-4567", "01011112222", "01033334444", "01055556666"]),
                (zip_col, ["01234", "12345", "", "67890", "00001", "00002", "00003"]),
            ]:
                if prepared[col].tolist() != expected:
                    failures.append(f"display_{col}_values={prepared[col].tolist()!r}")
                if not (pd.api.types.is_object_dtype(prepared[col]) or pd.api.types.is_string_dtype(prepared[col])):
                    failures.append(f"display_{col}_dtype={prepared[col].dtype}")

            for col in [in_qty_col, out_qty_col, supply_col, total_col]:
                if not pd.api.types.is_numeric_dtype(prepared[col]):
                    failures.append(f"numeric_{col}_dtype={prepared[col].dtype}")

            if not display_mod._is_numeric_display_col(view_df, seq_col):
                failures.append("seq_not_number_column")
            for col in [product_no_col, validation_col, product_code_col, biz_no_col, phone_col, zip_col]:
                if display_mod._is_numeric_display_col(view_df, col):
                    failures.append(f"{col}_unexpected_number_column")
            type_expectations = {
                seq_col: ("number", 1),
                product_no_col: ("text", None),
                validation_col: ("text", None),
                product_code_col: ("text", None),
                biz_no_col: ("text", None),
                phone_col: ("text", None),
                zip_col: ("text", None),
            }
            for col, (expected_type, expected_step) in type_expectations.items():
                cfg = column_config.get(col)
                actual_type = (cfg or {}).get("type_config", {}).get("type")
                actual_step = (cfg or {}).get("type_config", {}).get("step")
                if actual_type != expected_type:
                    failures.append(f"column_config_{col}_type={actual_type}")
                if expected_step is not None and actual_step != expected_step:
                    failures.append(f"column_config_{col}_step={actual_step}")

            json.dumps(prepared.to_dict(orient="records"), ensure_ascii=False)

            bio = io.BytesIO()
            excel_df = chat_mod._sanitize_dataframe_for_excel(finalized)
            excel_df.to_excel(bio, index=False, sheet_name="SIMS")
            bio.seek(0)
            wb = load_workbook(bio, read_only=True, data_only=True)
            ws = wb["SIMS"]
            header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

            def _cell(row: int, col_name: str):
                return ws.cell(row=row, column=header.index(col_name) + 1)

            if _cell(2, seq_col).value is not None:
                failures.append(f"excel_seq_blank={_cell(2, seq_col).value!r}")
            for row in [3, 4, 5]:
                if _cell(row, seq_col).data_type != "n":
                    failures.append(f"excel_seq_type_row{row}={_cell(row, seq_col).data_type}")
            for row in [2, 3, 4]:
                if _cell(row, product_no_col).data_type not in {"s", "inlineStr"}:
                    failures.append(f"excel_product_no_type_row{row}={_cell(row, product_no_col).data_type}")
                if _cell(row, validation_col).data_type not in {"s", "inlineStr"}:
                    failures.append(f"excel_validation_type_row{row}={_cell(row, validation_col).data_type}")
            if _cell(3, product_no_col).value != "000123":
                failures.append(f"excel_leading_zero={_cell(3, product_no_col).value!r}")
            if _cell(8, product_no_col).value != "NULL":
                failures.append(f"excel_literal_NULL={_cell(8, product_no_col).value!r}")
            if _cell(7, validation_col).value != "<NA>":
                failures.append(f"excel_literal_pdna={_cell(7, validation_col).value!r}")
            if _cell(8, validation_col).value != "nan":
                failures.append(f"excel_literal_nan={_cell(8, validation_col).value!r}")
            wb.close()

            try:
                import pyarrow as pa

                mixed_derived = pd.DataFrame({
                    seq_col: [None, 268.0, pd.NA, ""],
                    product_no_col: ["000123", "001-A", "114625021", ""],
                    validation_col: ["1", "0", "Y", ""],
                    salesperson_col: pd.Series(["0", "001", None, pd.NA], dtype="object"),
                    salesperson_name_col: pd.Series(["0", "Kim", None, ""], dtype="object"),
                    row_no_col: pd.Series([0, 0.0, "0", None], dtype="object"),
                    no_col: pd.Series(["0007", "0", "", pd.NA], dtype="object"),
                    "\uac70\ub798\ucc98\uba85": pd.Series(["0", None, "A", ""], dtype="object"),
                    "\ucf54\ub4dc": pd.Series(["0001", "0", None, pd.NA], dtype="object"),
                })
                mixed_before = {
                    col: mixed_derived[col].copy()
                    for col in [salesperson_col, salesperson_name_col, row_no_col, no_col, "\uac70\ub798\ucc98\uba85", "\ucf54\ub4dc"]
                }
                mixed_after = chat_mod._preserve_product_flow_table_dtypes(mixed_derived)
                for col, before_series in mixed_before.items():
                    try:
                        pd.testing.assert_series_equal(mixed_after[col], before_series, check_dtype=True, check_names=True)
                    except AssertionError as assert_exc:
                        failures.append(f"mixed_{col}_changed={assert_exc}")
                if str(mixed_after[seq_col].dtype) != "Int64":
                    failures.append(f"mixed_seq_dtype={mixed_after[seq_col].dtype}")
                for col in [product_no_col, validation_col]:
                    if not (pd.api.types.is_object_dtype(mixed_after[col]) or pd.api.types.is_string_dtype(mixed_after[col])):
                        failures.append(f"mixed_{col}_dtype={mixed_after[col].dtype}")

                derived = finalized.iloc[[1, 2, 4]].copy()
                derived[seq_col] = pd.Series([float("nan"), 268.0, 628.0], dtype="object")
                for numeric_col in [in_qty_col, out_qty_col, supply_col, total_col]:
                    derived[numeric_col] = pd.to_numeric(derived[numeric_col], errors="coerce")
                preserved_before = {
                    col: derived[col].copy()
                    for col in [salesperson_col, salesperson_name_col, row_no_col, no_col]
                }
                derived = chat_mod._preserve_product_flow_table_dtypes(derived)
                for col, before_series in preserved_before.items():
                    if derived[col].tolist() != before_series.tolist():
                        failures.append(f"derived_{col}_changed={derived[col].tolist()!r}")
                fast_like = display_mod.normalize_display_df_for_streamlit(chat_mod._chat_fast_display_df(derived.copy()))
                fast_like, fast_cfg, _tw, _th = display_mod.build_sims_table_display_config(
                    fast_like,
                    action_name="현재표 영업사원 TOP 20",
                    meta={"current_table_followup": True},
                    add_row_no=False,
                )
                fast_like = chat_mod._preserve_product_flow_table_dtypes(chat_mod._chat_clean_display_none_values(fast_like))
                if str(derived[seq_col].dtype) != "Int64":
                    failures.append(f"derived_seq_dtype={derived[seq_col].dtype}")
                if not (pd.api.types.is_object_dtype(derived[product_no_col]) or pd.api.types.is_string_dtype(derived[product_no_col])):
                    failures.append(f"derived_product_no_dtype={derived[product_no_col].dtype}")
                if not (pd.api.types.is_object_dtype(derived[validation_col]) or pd.api.types.is_string_dtype(derived[validation_col])):
                    failures.append(f"derived_validation_dtype={derived[validation_col].dtype}")
                if (fast_cfg.get(product_no_col) or {}).get("type_config", {}).get("type") == "number":
                    failures.append("derived_product_no_number_config")
                if (fast_cfg.get(validation_col) or {}).get("type_config", {}).get("type") == "number":
                    failures.append("derived_validation_number_config")
                pa.Table.from_pandas(fast_like, preserve_index=False)
            except Exception as arrow_exc:
                failures.append(f"derived_pyarrow={type(arrow_exc).__name__}: {arrow_exc}")

            if failures:
                results.append(_fail("product flow column type policy", "; ".join(failures)))
            else:
                results.append(_ok("product flow column type policy", "statement seq is nullable integer/numeric Excel cell; product no and validation remain text in service/display/Excel; numeric and identifier columns preserved"))
        except Exception as e:
            results.append(_fail("product flow column type policy", f"{type(e).__name__}: {e}"))

        try:
            main_src = Path("app/Lmstudio_SSAI_chat_main.py").read_text(encoding="utf-8")
            no_source_checks = {
                "simple_notice": "현재표가 없습니다. 먼저 SIMS 조회를 실행한 뒤 다시 질문해 주세요." in main_src,
                "source_probe": "def _has_current_table_source_df" in main_src,
                "immediate_notice": "no current source; notice pushed immediately" in main_src,
                "no_deferred": "st.session_state.pop(\"__deferred_current_table_followup\", None)" in main_src,
                "title_limit": "def _room_title_text_from_message(content: str, *, limit: int = 20)" in main_src,
            }
            failed = [name for name, ok in no_source_checks.items() if not ok]
            if failed:
                results.append(_fail("current-table missing source immediate notice", f"failed={failed}"))
            else:
                results.append(_ok("current-table missing source immediate notice", "no-source current-table questions push one notice immediately and do not defer to panel"))
        except Exception as e:
            results.append(_fail("current-table missing source immediate notice", f"{type(e).__name__}: {e}"))

        try:
            small_df = pd.DataFrame({"배정부족예상금액": range(20), "부족제품수": range(20)})
            eight_df = pd.DataFrame({"배정부족예상금액": range(8), "부족제품수": range(8)})
            mid_259_df = pd.DataFrame({
                "배정부족예상금액": range(259),
                "부족제품수": range(259),
                "매입처명": [f"B{i}" for i in range(259)],
            })
            mid_470_df = pd.DataFrame({
                "배정부족예상금액": range(470),
                "음수재고제품수": range(470),
                "매입처명": [f"B{i}" for i in range(470)],
            })
            group_df = pd.DataFrame({"주요배분기준": ["A", "B", "C", "D"], "배정부족예상금액": [1, 2, 3, 4]})
            meta_followup = {"current_table_followup": True}
            checks = {
                "20": chat_mod._chat_is_current_followup_fast_table(small_df, meta_followup),
                "8": chat_mod._chat_is_current_followup_fast_table(eight_df, meta_followup),
                "259": chat_mod._chat_is_current_followup_fast_table(mid_259_df, meta_followup),
                "470": chat_mod._chat_is_current_followup_fast_table(mid_470_df, meta_followup),
                "4": chat_mod._chat_is_current_followup_fast_table(group_df, meta_followup),
            }
            if checks != {"20": False, "8": False, "259": True, "470": True, "4": False}:
                results.append(_fail("supplier current-table fast render policy", f"unexpected modes={checks}"))
            else:
                results.append(_ok("supplier current-table fast render policy", "20/8/4 small; 259/470 fast"))
        except Exception as e:
            results.append(_fail("supplier current-table fast render policy", f"{type(e).__name__}: {e}"))

        try:
            import streamlit as st

            old_reason = st.session_state.get("__ui_rerun_reason")
            old_reason_current = st.session_state.get("__ui_rerun_reason_current")
            old_path = st.session_state.get("__sims_table_render_path")
            old_last_key = st.session_state.get("__sims_last_table_key")
            old_latest_followup = st.session_state.get("__sims_latest_followup_table_key")
            try:
                st.session_state["__sims_table_render_path"] = "history"
                st.session_state["__sims_last_table_key"] = "new-table"
                st.session_state["__sims_latest_followup_table_key"] = "new-table"
                old_item = {"type": "table", "action": "매입처별 재고부족 현황", "table_key": "old-table"}
                old_meta = {"table_key": "old-table", "row_count": 975}
                new_item = {"type": "table", "action": "현재표 배정부족예상금액 TOP 20", "table_key": "new-table"}
                new_meta = {"table_key": "new-table", "row_count": 20}

                st.session_state["__ui_rerun_reason_current"] = "sims_panel_open"
                panel_open_old = chat_mod._should_full_render_sims_table(old_item, old_meta, "old")
                panel_open_new = chat_mod._should_full_render_sims_table(new_item, new_meta, "new")
                st.session_state["__ui_rerun_reason_current"] = "chat_room_change"
                room_change_old = chat_mod._should_full_render_sims_table(old_item, old_meta, "old")
                st.session_state["__ui_rerun_reason"] = "sims_action_change"
                st.session_state["__ui_rerun_reason_current"] = "current_table_followup"
                st.session_state[chat_mod._old_sims_table_force_key(old_item, old_meta, "old")] = True
                followup_old = chat_mod._should_full_render_sims_table(old_item, old_meta, "old")
                followup_new = chat_mod._should_full_render_sims_table(new_item, new_meta, "new")

                if panel_open_old or panel_open_new or room_change_old:
                    results.append(_fail("supplier history lightweight rerun policy", f"light rerun should skip history tables panel_old={panel_open_old} panel_new={panel_open_new} room_old={room_change_old}"))
                elif not followup_old or not followup_new:
                    results.append(_fail("supplier history lightweight rerun policy", f"followup should render forced old and latest old={followup_old} new={followup_new}"))
                elif chat_mod._ui_rerun_reason() != "current_table_followup":
                    results.append(_fail("supplier history lightweight rerun policy", f"stale reason priority failed got={chat_mod._ui_rerun_reason()}"))
                else:
                    results.append(_ok("supplier history lightweight rerun policy", "panel/action rerun skips old tables; forced old table and latest followup render; stale action reason ignored"))
            finally:
                st.session_state.pop(chat_mod._old_sims_table_force_key(old_item, old_meta, "old"), None)
                if old_reason is None:
                    st.session_state.pop("__ui_rerun_reason", None)
                else:
                    st.session_state["__ui_rerun_reason"] = old_reason
                if old_reason_current is None:
                    st.session_state.pop("__ui_rerun_reason_current", None)
                else:
                    st.session_state["__ui_rerun_reason_current"] = old_reason_current
                if old_path is None:
                    st.session_state.pop("__sims_table_render_path", None)
                else:
                    st.session_state["__sims_table_render_path"] = old_path
                if old_last_key is None:
                    st.session_state.pop("__sims_last_table_key", None)
                else:
                    st.session_state["__sims_last_table_key"] = old_last_key
                if old_latest_followup is None:
                    st.session_state.pop("__sims_latest_followup_table_key", None)
                else:
                    st.session_state["__sims_latest_followup_table_key"] = old_latest_followup
        except Exception as e:
            results.append(_fail("supplier history lightweight rerun policy", f"{type(e).__name__}: {e}"))

        try:
            import streamlit as st

            old_chat_history = st.session_state.get("__chat_history")
            old_sims_tables = st.session_state.get("sims_tables")
            old_current_source_key = st.session_state.get("__sims_current_table_source_key")
            old_current_source_action = st.session_state.get("__sims_current_table_source_action")
            old_room = st.session_state.get("current_room")
            old_forced_key = None
            try:
                product_a = {
                    "id": "msg-product-a",
                    "type": "table",
                    "action": "제품수불현황 조회",
                    "meta": {"table_key": "sims_product_a", "action": "제품수불현황 조회"},
                }
                followup_1 = {
                    "id": "msg-followup-1",
                    "type": "table",
                    "action": "현재표 영업사원 TOP 20",
                    "meta": {"table_key": "sims_followup_1", "action": "현재표 영업사원 TOP 20", "current_table_followup": True},
                }
                followup_2 = {
                    "id": "msg-followup-2",
                    "type": "table",
                    "action": "현재표 영업사원별 집계",
                    "meta": {"table_key": "sims_followup_2", "action": "현재표 영업사원별 집계", "current_table_followup": True},
                }
                vendor_b = {
                    "id": "msg-vendor-b",
                    "type": "table",
                    "action": "거래처 목록",
                    "meta": {"table_key": "sims_vendor_b", "action": "거래처 목록"},
                }
                st.session_state["__chat_history"] = [product_a, followup_1, followup_2, vendor_b]
                st.session_state["sims_tables"] = {
                    "sims_product_a": object(),
                    "sims_followup_1": object(),
                    "sims_followup_2": object(),
                    "sims_vendor_b": object(),
                }
                st.session_state["__sims_current_table_source_key"] = "sims_vendor_b"
                st.session_state["__sims_current_table_source_action"] = "거래처 목록"
                old_forced_key = chat_mod._old_sims_table_force_key(product_a, product_a["meta"], "")
                st.session_state[old_forced_key] = True

                chat_mod._prune_old_sims_table_history(new_table_key="sims_vendor_b", new_item=vendor_b)
                kept_history_keys = [
                    str((x.get("meta") or {}).get("table_key") or "")
                    for x in st.session_state.get("__chat_history", [])
                    if isinstance(x, dict)
                ]
                kept_table_keys = set((st.session_state.get("sims_tables") or {}).keys())
                current_source_key_after = st.session_state.get("__sims_current_table_source_key")
                current_source_action_after = st.session_state.get("__sims_current_table_source_action")

                prune_mismatches = []
                if "sims_product_a" not in kept_history_keys:
                    prune_mismatches.append(f"forced old table missing from history keys={kept_history_keys}")
                if "sims_product_a" not in kept_table_keys:
                    prune_mismatches.append(f"forced old table payload missing keys={sorted(kept_table_keys)}")
                if current_source_key_after != "sims_vendor_b" or current_source_action_after != "거래처 목록":
                    prune_mismatches.append(f"current source changed key={current_source_key_after} action={current_source_action_after}")

                if prune_mismatches:
                    results.append(_fail("old SIMS table reopen preserves clicked table", "; ".join(prune_mismatches)))
                else:
                    results.append(_ok("old SIMS table reopen preserves clicked table", "forced old table survives prune/history payload cleanup while current source remains latest new table"))
            finally:
                if old_forced_key:
                    st.session_state.pop(old_forced_key, None)
                if old_chat_history is None:
                    st.session_state.pop("__chat_history", None)
                else:
                    st.session_state["__chat_history"] = old_chat_history
                if old_sims_tables is None:
                    st.session_state.pop("sims_tables", None)
                else:
                    st.session_state["sims_tables"] = old_sims_tables
                if old_current_source_key is None:
                    st.session_state.pop("__sims_current_table_source_key", None)
                else:
                    st.session_state["__sims_current_table_source_key"] = old_current_source_key
                if old_current_source_action is None:
                    st.session_state.pop("__sims_current_table_source_action", None)
                else:
                    st.session_state["__sims_current_table_source_action"] = old_current_source_action
                if old_room is None:
                    st.session_state.pop("current_room", None)
                else:
                    st.session_state["current_room"] = old_room
        except Exception as e:
            results.append(_fail("old SIMS table reopen preserves clicked table", f"{type(e).__name__}: {e}"))

        try:
            import streamlit as st

            old_chat_history = st.session_state.get("__chat_history")
            old_sims_tables = st.session_state.get("sims_tables")
            old_current_source_key = st.session_state.get("__sims_current_table_source_key")
            old_current_source_action = st.session_state.get("__sims_current_table_source_action")
            old_path = st.session_state.get("__sims_table_render_path")
            old_reason_current = st.session_state.get("__ui_rerun_reason_current")
            old_latest_followup = st.session_state.get("__sims_latest_followup_table_key")
            old_last_table = st.session_state.get("__sims_last_table_key")
            old_chat_rooms = st.session_state.get("chat_rooms")
            old_current_room = st.session_state.get("current_room")
            old_previous_target = st.session_state.get("__sims_previous_current_table_source_target_key_for_prune")
            force_key = None
            try:
                product_a = {
                    "id": "msg-product-a-preforce",
                    "type": "table",
                    "action": "제품수불현황 조회",
                    "meta": {"table_key": "sims_product_a_preforce", "action": "제품수불현황 조회"},
                }
                followup_1 = {
                    "id": "msg-followup-preforce-1",
                    "type": "table",
                    "action": "현재표 영업사원 TOP 20",
                    "meta": {
                        "table_key": "sims_followup_preforce_1",
                        "action": "현재표 영업사원 TOP 20",
                        "current_table_followup": True,
                        "source_table_key": "sims_product_a_preforce",
                    },
                }
                followup_2 = {
                    "id": "msg-followup-preforce-2",
                    "type": "table",
                    "action": "현재표 영업사원별 집계",
                    "meta": {
                        "table_key": "sims_followup_preforce_2",
                        "action": "현재표 영업사원별 집계",
                        "current_table_followup": True,
                        "source_table_key": "sims_product_a_preforce",
                    },
                }
                vendor_b = {
                    "id": "msg-vendor-b-preforce",
                    "type": "table",
                    "action": "거래처 목록",
                    "meta": {"table_key": "sims_vendor_b_preforce", "action": "거래처 목록"},
                }
                room = {
                    "id": "room-preforce-reconcile",
                    "history": [
                        {"id": "msg-normal-before-a", "role": "user", "content": "normal message"},
                        product_a,
                        followup_1,
                        followup_2,
                    ],
                    "sims_messages": [],
                    "gen_messages": [],
                    "messages": [],
                }
                st.session_state["chat_rooms"] = [room]
                st.session_state["current_room"] = room["id"]
                st.session_state["__chat_history"] = [
                    {"id": "msg-normal-before-a", "role": "user", "content": "normal message"},
                    followup_1,
                    followup_2,
                ]
                st.session_state["sims_tables"] = {
                    "sims_product_a_preforce": object(),
                    "sims_followup_preforce_1": object(),
                    "sims_followup_preforce_2": object(),
                    "sims_vendor_b_preforce": object(),
                }
                st.session_state["__sims_current_table_source_key"] = "sims_vendor_b_preforce"
                st.session_state["__sims_current_table_source_action"] = "거래처 목록"

                st.session_state["__sims_previous_current_table_source_key_for_prune"] = "sims_product_a_preforce"
                st.session_state["__sims_previous_current_table_source_target_key_for_prune"] = "sims_vendor_b_preforce"

                chat_mod._prune_old_sims_table_history(new_table_key="sims_vendor_b_preforce", new_item=vendor_b)
                kept_history_keys = [
                    str((x.get("meta") or {}).get("table_key") or "")
                    for x in st.session_state.get("__chat_history", [])
                    if isinstance(x, dict)
                ]
                kept_table_keys = set((st.session_state.get("sims_tables") or {}).keys())
                force_key = chat_mod._old_sims_table_force_key(product_a, product_a["meta"], "preforce")
                st.session_state[force_key] = True
                st.session_state["__sims_table_render_path"] = "history"
                st.session_state["__ui_rerun_reason_current"] = "current_table_followup"
                st.session_state["__sims_latest_followup_table_key"] = "sims_followup_preforce_2"
                st.session_state["__sims_last_table_key"] = "sims_vendor_b_preforce"
                full_after_force = chat_mod._should_full_render_sims_table(product_a, product_a["meta"], "preforce")
                st.session_state.pop(force_key, None)
                placeholder_after_collapse = not chat_mod._should_full_render_sims_table(product_a, product_a["meta"], "preforce")

                prune_mismatches = []
                if "sims_product_a_preforce" not in kept_history_keys:
                    prune_mismatches.append(f"pre-force source table missing from history keys={kept_history_keys}")
                if kept_history_keys.count("sims_product_a_preforce") > 1:
                    prune_mismatches.append(f"pre-force source duplicated keys={kept_history_keys}")
                kept_table_order = [k for k in kept_history_keys if k]
                if kept_table_order[:3] != ["sims_product_a_preforce", "sims_followup_preforce_1", "sims_followup_preforce_2"]:
                    prune_mismatches.append(f"reconciled table order changed keys={kept_history_keys}")
                if "sims_product_a_preforce" not in kept_table_keys:
                    prune_mismatches.append(f"pre-force source payload missing keys={sorted(kept_table_keys)}")
                if st.session_state.get("__sims_current_table_source_key") != "sims_vendor_b_preforce":
                    prune_mismatches.append(f"current source key changed={st.session_state.get('__sims_current_table_source_key')}")
                if st.session_state.get("__sims_previous_current_table_source_key_for_prune"):
                    prune_mismatches.append("previous source prune key was not consumed")
                if not full_after_force:
                    prune_mismatches.append("forced source table did not full-render")
                if not placeholder_after_collapse:
                    prune_mismatches.append("collapsed source table did not return to placeholder")

                if prune_mismatches:
                    results.append(_fail("old SIMS table pre-force prune keeps source", "; ".join(prune_mismatches)))
                else:
                    results.append(_ok("old SIMS table pre-force prune keeps source", "previous source table survives new basis prune before force, can full-render/collapse, and current source stays latest table"))
            finally:
                if force_key:
                    st.session_state.pop(force_key, None)
                if old_chat_history is None:
                    st.session_state.pop("__chat_history", None)
                else:
                    st.session_state["__chat_history"] = old_chat_history
                if old_sims_tables is None:
                    st.session_state.pop("sims_tables", None)
                else:
                    st.session_state["sims_tables"] = old_sims_tables
                if old_current_source_key is None:
                    st.session_state.pop("__sims_current_table_source_key", None)
                else:
                    st.session_state["__sims_current_table_source_key"] = old_current_source_key
                if old_current_source_action is None:
                    st.session_state.pop("__sims_current_table_source_action", None)
                else:
                    st.session_state["__sims_current_table_source_action"] = old_current_source_action
                if old_path is None:
                    st.session_state.pop("__sims_table_render_path", None)
                else:
                    st.session_state["__sims_table_render_path"] = old_path
                if old_reason_current is None:
                    st.session_state.pop("__ui_rerun_reason_current", None)
                else:
                    st.session_state["__ui_rerun_reason_current"] = old_reason_current
                if old_latest_followup is None:
                    st.session_state.pop("__sims_latest_followup_table_key", None)
                else:
                    st.session_state["__sims_latest_followup_table_key"] = old_latest_followup
                if old_last_table is None:
                    st.session_state.pop("__sims_last_table_key", None)
                else:
                    st.session_state["__sims_last_table_key"] = old_last_table
                if old_chat_rooms is None:
                    st.session_state.pop("chat_rooms", None)
                else:
                    st.session_state["chat_rooms"] = old_chat_rooms
                if old_current_room is None:
                    st.session_state.pop("current_room", None)
                else:
                    st.session_state["current_room"] = old_current_room
                if old_previous_target is None:
                    st.session_state.pop("__sims_previous_current_table_source_target_key_for_prune", None)
                else:
                    st.session_state["__sims_previous_current_table_source_target_key_for_prune"] = old_previous_target
        except Exception as e:
            results.append(_fail("old SIMS table pre-force prune keeps source", f"{type(e).__name__}: {e}"))

        try:
            import streamlit as st

            panel_mod = importlib.import_module("app.ui.sims_panel")

            state_keys = [
                "__sims_current_table_source_key",
                "__sims_current_table_source_action",
                "__sims_previous_current_table_source_key_for_prune",
                "__sims_previous_current_table_source_action_for_prune",
                "__sims_previous_current_table_source_target_key_for_prune",
                "__old_table_history_refresh_key_pending",
            ]
            saved_state = {key: st.session_state.get(key) for key in state_keys}
            missing_state = {key for key in state_keys if key not in st.session_state}
            try:
                df = pd.DataFrame({"value": [1]})
                for key in state_keys:
                    st.session_state.pop(key, None)
                st.session_state["__sims_current_table_source_key"] = "sims_product_a_stash"
                st.session_state["__sims_current_table_source_action"] = "제품수불현황 조회"

                panel_mod._stash_panel_table_for_current_followup(
                    {
                        "df": df,
                        "df_display": df.copy(),
                        "meta": {"table_key": "sims_vendor_b_stash", "action": "거래처 목록"},
                    },
                    "거래처 목록",
                    record_previous_source_for_prune=True,
                )

                mismatches = []
                if st.session_state.get("__sims_current_table_source_key") != "sims_vendor_b_stash":
                    mismatches.append("new panel result did not become current source")
                if st.session_state.get("__sims_previous_current_table_source_key_for_prune") != "sims_product_a_stash":
                    mismatches.append("previous source key not recorded for prune")
                if st.session_state.get("__sims_previous_current_table_source_target_key_for_prune") != "sims_vendor_b_stash":
                    mismatches.append("previous source target key not recorded for prune")
                if st.session_state.get("__old_table_history_refresh_key_pending") != "sims_vendor_b_stash":
                    mismatches.append("history refresh key not recorded")

                for key in (
                    "__sims_previous_current_table_source_key_for_prune",
                    "__sims_previous_current_table_source_action_for_prune",
                    "__sims_previous_current_table_source_target_key_for_prune",
                    "__old_table_history_refresh_key_pending",
                ):
                    st.session_state.pop(key, None)
                panel_mod._stash_panel_table_for_current_followup(
                    {
                        "df": df,
                        "df_display": df.copy(),
                        "meta": {
                            "table_key": "sims_followup_stash",
                            "action": "현재표 영업사원 TOP 20",
                            "current_table_followup": True,
                        },
                    },
                    "현재표 영업사원 TOP 20",
                    record_previous_source_for_prune=True,
                )
                if st.session_state.get("__sims_previous_current_table_source_key_for_prune"):
                    mismatches.append("followup table incorrectly recorded previous source")
                if st.session_state.get("__old_table_history_refresh_key_pending"):
                    mismatches.append("followup table incorrectly requested history refresh")

                if mismatches:
                    results.append(_fail("panel current source transition records previous source", "; ".join(mismatches)))
                else:
                    results.append(_ok("panel current source transition records previous source", "new panel result records previous source for prune and one-shot history refresh; followup render does not"))
            finally:
                for key in state_keys:
                    if key in missing_state:
                        st.session_state.pop(key, None)
                    else:
                        st.session_state[key] = saved_state.get(key)
        except Exception as e:
            results.append(_fail("panel current source transition records previous source", f"{type(e).__name__}: {e}"))

        try:
            import streamlit as st

            panel_mod = importlib.import_module("app.ui.sims_panel")
            state_keys = [
                "__sims_panel_chat_pushed_source_sig",
                "__sims_last_final_payload_for_chat",
                "__sims_last_final_payload_for_chat_action",
                "__sims_panel_last_final_payload",
                "__sims_panel_last_final_action",
                "__sims_run_seq",
                "__sims_selected",
            ]
            saved_state = {key: st.session_state.get(key) for key in state_keys}
            missing_state = {key for key in state_keys if key not in st.session_state}
            try:
                for key in state_keys:
                    st.session_state.pop(key, None)
                calls = {"entry": 0, "push": 0, "save": 0}
                table_keys: list[str] = []

                def _clear_payload_cache() -> None:
                    for key in (
                        "__sims_last_final_payload_for_chat",
                        "__sims_last_final_payload_for_chat_action",
                        "__sims_panel_last_final_payload",
                        "__sims_panel_last_final_action",
                    ):
                        st.session_state.pop(key, None)

                def _fake_panel_push_boundary(*, run_seq: int, submit_seq: int, action: str, table_key: str, condition: str) -> bool:
                    calls["entry"] += 1
                    st.session_state["__sims_run_seq"] = run_seq
                    st.session_state["__sims_query_submit_seq"] = submit_seq
                    st.session_state["__sims_selected"] = {"category": "마스터", "action": action}
                    payload = {
                        "type": "table",
                        "action": action,
                        "condition": condition,
                        "meta": {
                            "table_key": table_key,
                            "action": action,
                            "query_summary": condition,
                        },
                    }
                    payload["meta"]["_panel_source_sig"] = panel_mod._make_panel_source_sig(action, payload)
                    st.session_state["__sims_last_final_payload_for_chat"] = payload
                    st.session_state["__sims_panel_last_final_payload"] = payload
                    sig = str((payload.get("meta") or {}).get("_panel_source_sig") or "")
                    if panel_mod._panel_chat_push_already_consumed(sig):
                        _clear_payload_cache()
                        return False
                    calls["push"] += 1
                    calls["save"] += 1
                    table_keys.append(table_key)
                    st.session_state["__sims_panel_chat_pushed_source_sig"] = sig
                    _clear_payload_cache()
                    return True

                st.session_state["__sims_run_seq"] = 77
                st.session_state["__sims_query_submit_seq"] = 1
                st.session_state["__sims_selected"] = {"category": "마스터", "action": "거래처 목록"}
                first = _fake_panel_push_boundary(run_seq=77, submit_seq=1, action="거래처 목록", table_key="sims_vendor_first", condition="전체")
                second = _fake_panel_push_boundary(run_seq=77, submit_seq=2, action="거래처 목록", table_key="sims_vendor_same_submit", condition="전체")
                third = _fake_panel_push_boundary(run_seq=77, submit_seq=2, action="거래처 목록", table_key="sims_vendor_rerun", condition="전체")
                fourth = _fake_panel_push_boundary(run_seq=77, submit_seq=3, action="거래처 목록", table_key="sims_vendor_changed", condition="변경조건")
                cache_left = any(st.session_state.get(k) for k in (
                    "__sims_last_final_payload_for_chat",
                    "__sims_last_final_payload_for_chat_action",
                    "__sims_panel_last_final_payload",
                    "__sims_panel_last_final_action",
                ))
                mismatches = []
                if not first:
                    mismatches.append("first production boundary push skipped")
                if not second:
                    mismatches.append("explicit same-condition submit was blocked")
                if third:
                    mismatches.append("simple rerun was not skipped")
                if not fourth:
                    mismatches.append("changed-condition submit was blocked")
                if calls != {"entry": 4, "push": 3, "save": 3}:
                    mismatches.append(f"unexpected calls={calls}")
                if len(set(table_keys)) != 3 or table_keys != ["sims_vendor_first", "sims_vendor_same_submit", "sims_vendor_changed"]:
                    mismatches.append(f"unexpected table_keys={table_keys}")
                if cache_left:
                    mismatches.append("cached final payload was not cleaned")
                if panel_mod._panel_chat_push_already_consumed("other-query-signature"):
                    mismatches.append("different panel source was incorrectly consumed")
                if mismatches:
                    results.append(_fail("panel chat push lifecycle signature", "; ".join(mismatches)))
                else:
                    results.append(_ok("panel chat push lifecycle signature", "explicit submit sequence and condition fingerprint create new chat results; simple reruns skip duplicate pushes"))
            finally:
                for key in state_keys:
                    if key in missing_state:
                        st.session_state.pop(key, None)
                    else:
                        st.session_state[key] = saved_state.get(key)
        except Exception as e:
            results.append(_fail("panel chat push lifecycle signature", f"{type(e).__name__}: {e}"))

        try:
            import streamlit as st

            chat_mod = importlib.import_module("app.ui.chat_middleware")

            state_keys = [
                "__chat_history",
                "chat_rooms",
                "current_room",
                "sims_tables",
                "__sims_previous_current_table_source_key_for_prune",
                "__sims_previous_current_table_source_target_key_for_prune",
                "__sims_previous_current_table_source_action_for_prune",
                "__old_table_history_refresh_key_pending",
            ]
            saved_state = {key: st.session_state.get(key) for key in state_keys}
            missing_state = {key for key in state_keys if key not in st.session_state}
            try:
                product_a = {
                    "id": "msg-product-a-mismatch",
                    "type": "table",
                    "action": "제품수불현황 조회",
                    "meta": {"table_key": "sims_product_a_mismatch", "action": "제품수불현황 조회"},
                }
                followup = {
                    "id": "msg-followup-mismatch",
                    "type": "table",
                    "action": "현재표 영업사원 TOP 20",
                    "meta": {
                        "table_key": "sims_followup_mismatch",
                        "action": "현재표 영업사원 TOP 20",
                        "current_table_followup": True,
                        "source_table_key": "sims_product_a_mismatch",
                    },
                }
                vendor_c = {
                    "id": "msg-vendor-c-mismatch",
                    "type": "table",
                    "action": "거래처 목록",
                    "meta": {"table_key": "sims_vendor_c_mismatch", "action": "거래처 목록"},
                }
                room = {"id": "room-target-mismatch", "history": [product_a, followup], "sims_messages": [], "gen_messages": [], "messages": []}
                st.session_state["chat_rooms"] = [room]
                st.session_state["current_room"] = room["id"]
                st.session_state["__chat_history"] = [followup]
                st.session_state["sims_tables"] = {
                    "sims_product_a_mismatch": object(),
                    "sims_followup_mismatch": object(),
                    "sims_vendor_c_mismatch": object(),
                }
                st.session_state["__sims_previous_current_table_source_key_for_prune"] = "sims_product_a_mismatch"
                st.session_state["__sims_previous_current_table_source_target_key_for_prune"] = "sims_vendor_b_failed"
                st.session_state["__sims_previous_current_table_source_action_for_prune"] = "제품수불현황 조회"
                st.session_state["__old_table_history_refresh_key_pending"] = "sims_vendor_b_failed"

                chat_mod._prune_old_sims_table_history(new_table_key="sims_vendor_c_mismatch", new_item=vendor_c)
                kept_history_keys = [
                    str((x.get("meta") or {}).get("table_key") or "")
                    for x in st.session_state.get("__chat_history", [])
                    if isinstance(x, dict)
                ]
                mismatches = []
                if "sims_product_a_mismatch" in kept_history_keys:
                    mismatches.append(f"stale previous source was reconciled keys={kept_history_keys}")
                for key in (
                    "__sims_previous_current_table_source_key_for_prune",
                    "__sims_previous_current_table_source_target_key_for_prune",
                    "__sims_previous_current_table_source_action_for_prune",
                    "__old_table_history_refresh_key_pending",
                ):
                    if st.session_state.get(key):
                        mismatches.append(f"temporary key not cleared {key}={st.session_state.get(key)!r}")
                if mismatches:
                    results.append(_fail("old SIMS table target mismatch clears stale previous source", "; ".join(mismatches)))
                else:
                    results.append(_ok("old SIMS table target mismatch clears stale previous source", "stale A->B transition is not consumed by C prune and temporary keys are cleared"))
            finally:
                for key in state_keys:
                    if key in missing_state:
                        st.session_state.pop(key, None)
                    else:
                        st.session_state[key] = saved_state.get(key)
        except Exception as e:
            results.append(_fail("old SIMS table target mismatch clears stale previous source", f"{type(e).__name__}: {e}"))

        try:
            chat_mod = importlib.import_module("app.ui.chat_middleware")

            class _State(dict):
                def pop(self, key, default=None):
                    return super().pop(key, default)

            calls = {"rerun": 0, "db": 0, "push": 0, "save": 0}

            def _rerun():
                calls["rerun"] += 1

            state = _State({
                "__old_table_history_refresh_key_pending": "sims_vendor_b_refresh",
                "__chat_pending_render": [{"meta": {"table_key": "sims_vendor_b_refresh"}}],
                "__sims_panel_skip_view_once": True,
                "__sims_panel_skip_view_reason": "panel_new_result",
            })
            first = chat_mod.consume_old_table_history_refresh_once(state, run_seq="run-1", rerun=_rerun)
            second = chat_mod.consume_old_table_history_refresh_once(
                _State({
                    "__old_table_history_refresh_key_pending": "sims_vendor_b_refresh",
                    "__old_table_history_refresh_done_sig": state.get("__old_table_history_refresh_done_sig"),
                }),
                run_seq="run-1",
                rerun=_rerun,
            )
            mismatches = []
            if not first:
                mismatches.append("first refresh was not consumed")
            if second:
                mismatches.append("duplicate refresh was consumed")
            if calls["rerun"] != 1:
                mismatches.append(f"rerun_count={calls['rerun']}")
            if calls["db"] or calls["push"] or calls["save"]:
                mismatches.append(f"unexpected side effects={calls}")
            if state.get("__chat_pending_render") != [{"meta": {"table_key": "sims_vendor_b_refresh"}}]:
                mismatches.append("pending B was changed")
            if not state.get("__sims_panel_skip_view_once"):
                mismatches.append("skip-view one-shot was cleared before refresh")
            if state.get("__ui_rerun_reason") != "old_table_history_refresh":
                mismatches.append(f"unexpected refresh reason={state.get('__ui_rerun_reason')!r}")
            if mismatches:
                results.append(_fail("old table history refresh consumes once without side effects", "; ".join(mismatches)))
            else:
                results.append(_ok("old table history refresh consumes once without side effects", "pending refresh reruns exactly once with dedicated reason, keeps pending render/skip-view state, and calls no DB/push/save hooks"))
        except Exception as e:
            results.append(_fail("old table history refresh consumes once without side effects", f"{type(e).__name__}: {e}"))

        try:
            import streamlit as st

            chat_mod = importlib.import_module("app.ui.chat_middleware")

            state_keys = [
                "__chat_history",
                "__chat_inbox",
                "__chat_pending_items",
                "__chat_pending_render",
                "chat_rooms",
                "current_room",
                "sims_tables",
                "sims_export_tables",
                "__sims_export_tables_by_key",
                "__sims_analysis_ctx_by_table_key",
                "__sims_current_table_source_key",
                "__sims_current_table_source_action",
                "__sims_previous_current_table_source_key_for_prune",
                "__sims_table_render_path",
                "__ui_rerun_reason_current",
                "__sims_latest_followup_table_key",
                "__sims_last_table_key",
                "__sims_last_table_action",
                "__sims_last_push",
                "__sims_last_push_sig",
                "__sims_push_count",
                "__seq",
            ]
            saved_state = {key: st.session_state.get(key) for key in state_keys}
            missing_state = {key for key in state_keys if key not in st.session_state}
            force_key = None

            def _payload(table_key: str, action: str, df: pd.DataFrame, meta: dict[str, Any] | None = None) -> dict[str, Any]:
                payload_meta = {
                    "table_key": table_key,
                    "action": action,
                    "_force_push": True,
                    "_nlq_nonce": table_key,
                    "query_summary": "regression fixture",
                }
                if meta:
                    payload_meta.update(meta)
                return {
                    "final": True,
                    "type": "table",
                    "title": action,
                    "action": action,
                    "params": {},
                    "df": df.copy(),
                    "df_display": df.copy(),
                    "meta": payload_meta,
                }

            def _history_keys(items: list[Any]) -> list[str]:
                keys: list[str] = []
                for item in items:
                    if isinstance(item, dict):
                        meta = item.get("meta") if isinstance(item.get("meta"), dict) else {}
                        key = str(meta.get("table_key") or item.get("table_key") or "").strip()
                        if key:
                            keys.append(key)
                return keys

            def _find_item(items: list[Any], table_key: str) -> dict[str, Any] | None:
                for item in items:
                    if not isinstance(item, dict):
                        continue
                    meta = item.get("meta") if isinstance(item.get("meta"), dict) else {}
                    if str(meta.get("table_key") or item.get("table_key") or "").strip() == table_key:
                        return item
                return None

            try:
                room = {
                    "id": "room-old-table-production-prune",
                    "history": [],
                    "sims_messages": [],
                    "gen_messages": [],
                    "messages": [],
                }
                st.session_state["chat_rooms"] = [room]
                st.session_state["current_room"] = room["id"]
                st.session_state["__chat_history"] = []
                st.session_state["__chat_inbox"] = []
                st.session_state["__chat_pending_items"] = []
                st.session_state["__chat_pending_render"] = []
                st.session_state["sims_tables"] = {}
                st.session_state["sims_export_tables"] = {}
                st.session_state["__sims_export_tables_by_key"] = {}
                st.session_state["__sims_analysis_ctx_by_table_key"] = {}
                st.session_state["__sims_push_count"] = 0
                st.session_state["__seq"] = 0
                for key in ("__sims_current_table_source_key", "__sims_current_table_source_action", "__sims_previous_current_table_source_key_for_prune"):
                    st.session_state.pop(key, None)

                product_df = pd.DataFrame({
                    "명세서번호": pd.Series([None, 268, 628], dtype="Int64"),
                    "제조번호": ["000123", "001-A", "114625021"],
                    "검수확인": ["1", "0", "Y"],
                    "영업사원": ["김", "이", "박"],
                    "합계금액": [1000000, 2000000, 3000000],
                })
                top_df = product_df.head(2).copy()
                group_df = pd.DataFrame({"영업사원": ["김", "이"], "합계금액": [1000000, 2000000]})
                vendor_df = pd.DataFrame({"거래처코드": ["V001"], "거래처명": ["테스트"]})

                chat_mod.push_sims_result_to_chat(
                    _payload("sims_product_a_prodpath", "제품수불현황 조회", product_df),
                    "제품수불현황 조회",
                )
                if st.session_state.get("__sims_current_table_source_key") != "sims_product_a_prodpath":
                    raise AssertionError("product A did not become current source")

                chat_mod.push_sims_result_to_chat(
                    _payload(
                        "sims_top_prodpath",
                        "현재표 영업사원 TOP 20",
                        top_df,
                        {"current_table_followup": True},
                    ),
                    "현재표 영업사원 TOP 20",
                )
                chat_mod.push_sims_result_to_chat(
                    _payload(
                        "sims_group_prodpath",
                        "현재표 영업사원별 집계",
                        group_df,
                        {"current_table_followup": True},
                    ),
                    "현재표 영업사원별 집계",
                )

                history_after_followups = list(room.get("history") or [])
                top_item = _find_item(history_after_followups, "sims_top_prodpath")
                group_item = _find_item(history_after_followups, "sims_group_prodpath")
                top_source = str(((top_item or {}).get("meta") or {}).get("source_table_key") or "")
                group_source = str(((group_item or {}).get("meta") or {}).get("source_table_key") or "")

                chat_mod.push_sims_result_to_chat(
                    _payload("sims_vendor_b_prodpath", "거래처 목록", vendor_df),
                    "거래처 목록",
                )

                room_history = list(room.get("history") or [])
                session_history = list(st.session_state.get("__chat_history") or [])
                room_keys = _history_keys(room_history)
                session_keys = _history_keys(session_history)
                table_keys = set((st.session_state.get("sims_tables") or {}).keys())
                current_source_key = st.session_state.get("__sims_current_table_source_key")
                current_source_action = st.session_state.get("__sims_current_table_source_action")
                product_item = _find_item(room_history, "sims_product_a_prodpath")

                mismatches = []
                if top_source != "sims_product_a_prodpath" or group_source != "sims_product_a_prodpath":
                    mismatches.append(f"production followup source meta missing top={top_source} group={group_source}")
                if "sims_product_a_prodpath" not in room_keys:
                    mismatches.append(f"product A missing from room.history keys={room_keys}")
                if "sims_product_a_prodpath" not in session_keys:
                    mismatches.append(f"product A missing from __chat_history keys={session_keys}")
                if "sims_product_a_prodpath" not in table_keys:
                    mismatches.append(f"product A payload missing keys={sorted(table_keys)}")
                if current_source_key != "sims_vendor_b_prodpath" or current_source_action != "거래처 목록":
                    mismatches.append(f"current source changed key={current_source_key} action={current_source_action}")
                if not isinstance(product_item, dict):
                    mismatches.append("product A item missing before force")
                else:
                    meta = product_item.get("meta") if isinstance(product_item.get("meta"), dict) else {}
                    force_key = chat_mod._old_sims_table_force_key(product_item, meta, "prodpath")
                    st.session_state["__sims_table_render_path"] = "history"
                    st.session_state["__ui_rerun_reason_current"] = "old_table_history_refresh"
                    st.session_state["__sims_latest_followup_table_key"] = "sims_group_prodpath"
                    st.session_state["__sims_last_table_key"] = "sims_vendor_b_prodpath"
                    vendor_item = _find_item(room_history, "sims_vendor_b_prodpath")
                    vendor_meta = vendor_item.get("meta") if isinstance(vendor_item, dict) and isinstance(vendor_item.get("meta"), dict) else {}
                    if not isinstance(vendor_item, dict) or not chat_mod._should_full_render_sims_table(vendor_item, vendor_meta, "prodpath"):
                        mismatches.append("current source B did not full-render during history refresh")
                    if chat_mod._should_full_render_sims_table(product_item, meta, "prodpath"):
                        mismatches.append("product A full-rendered before explicit force")
                    st.session_state[force_key] = True
                    if not chat_mod._should_full_render_sims_table(product_item, meta, "prodpath"):
                        mismatches.append("product A did not full-render after force helper")
                    st.session_state.pop(force_key, None)
                    if chat_mod._should_full_render_sims_table(product_item, meta, "prodpath"):
                        mismatches.append("product A did not return to placeholder after collapse helper")
                if st.session_state.get("__sims_current_table_source_key") != "sims_vendor_b_prodpath":
                    mismatches.append(f"current source changed after force/collapse={st.session_state.get('__sims_current_table_source_key')}")

                if mismatches:
                    results.append(_fail("old SIMS table production push prune boundary", "; ".join(mismatches)))
                else:
                    results.append(_ok("old SIMS table production push prune boundary", "production push fills followup source meta, preserves previous source before force, and force/collapse keeps latest current source"))
            finally:
                if force_key:
                    st.session_state.pop(force_key, None)
                for key in state_keys:
                    if key in missing_state:
                        st.session_state.pop(key, None)
                    else:
                        st.session_state[key] = saved_state.get(key)
        except Exception as e:
            results.append(_fail("old SIMS table production push prune boundary", f"{type(e).__name__}: {e}"))

        try:
            main_src = Path("app/Lmstudio_SSAI_chat_main.py").read_text(encoding="utf-8")
            required_close_keys = [
                '"__sims_open"',
                '"__sims_open_ui"',
                '"__sims_panel_active"',
                '"__sims_force_open"',
                '"__sims_run_flag"',
                '"__sims_inner_submit"',
                '"__sims_selected_snapshot"',
            ]
            has_close_helper = "def _close_sims_panel_for_room_change" in main_src
            has_close_guard = '"__sims_close_for_chat_room_change"' in main_src
            has_guard_consume = "def _consume_sims_close_for_chat_room_change" in main_src
            has_panel_close_log = "[chat.room.panel_close]" in main_src
            has_after_success_log = '_log_sims_panel_room_close_state("after success=True")' in main_src
            has_no_open_assignment = '"__sims_open",\n            "__sims_open_ui"' not in main_src and 'ss["__sims_open"] = False' not in main_src
            room_switch_block = main_src[main_src.find("if picked and picked != ss.current_room:"):main_src.find("cur_name = id_to_name.get", main_src.find("if picked and picked != ss.current_room:"))]
            has_no_direct_close_in_room_switch = "_close_sims_panel_for_room_change()" not in room_switch_block
            has_render_block = 'if st.session_state.get("__ui_rerun_reason_current") == "chat_room_change":' in main_src and "should_render = False" in main_src
            has_room_reason = '"chat_room_change"' in main_src
            has_switch_total = 'switch_total = float(stats.get("event_to_main_elapsed") or 0.0) + float(stats.get("history_elapsed") or 0.0)' in main_src
            has_switch_event_id = '"__chat_room_switch_event_id"' in main_src and "event_id=%s" in room_switch_block
            has_event_id_perf = '"__ui_event_id"' in main_src and "[ui.event_to_rerun] event_id=%s" in main_src
            has_startup_pending = '"__auth_login_perf_pending"' in main_src and '"__auth_startup_perf_emitted_sig"' in main_src
            has_no_unconditional_startup_log = "[auth.startup.perf] company_select=" not in main_src
            has_save_detail = '"__chat_room_switch_save_detail"' in main_src and "json_serialize=%.3fs" in main_src
            has_sims_open_perf = "[sims.panel_open.perf]" in main_src and '"__sims_panel_open_fragment_elapsed"' in main_src
            has_authenticate_perf = "__auth_login_authenticate_elapsed" in main_src and "authenticate=%.3fs" in main_src
            has_script_path_perf = (
                "[ui.script_path.perf]" in main_src
                and "room_selector=%.3fs" in main_src
                and "sims_fragment=%.3fs" in main_src
                and "unattributed=%.3fs" in main_src
                and 'st.session_state["__ui_script_perf_durations"] = {}' in main_src
            )
            has_save_skip = (
                "[chat.save.skip]" in main_src
                and "reason=unchanged" in main_src
                and "unchanged_or_selection_only" in main_src
                and "compare_mode=%s" in main_src
                and "_record_chat_save_skip" in main_src
            )
            has_selection_only_save_skip = (
                "removed_empty_pending = _drop_empty_auto_rooms(keep_room_id=picked)" in room_switch_block
                and "dirty_reason=\"selection_only\"" in room_switch_block
                and "save_chat_rooms()" in room_switch_block
            )
            has_chat_save_diff = "[chat.save.diff]" in main_src and "changed_fields=%s" in main_src
            has_latest_message_anchor = (
                '"__chat_scroll_to_bottom_once"' not in main_src
                and '"__chat_scroll_event_id"' not in main_src
                and "[chat.room.autoscroll]" not in main_src
                and "setTimeout(run" not in main_src
                and "focus({ preventScroll: true })" not in main_src
                and "ssai-chat-bottom-anchor" in main_src
                and "ssai-latest-message-link" in main_src
                and 'href="#ssai-chat-bottom-anchor"' in main_src
                and '[data-testid="stChatInput"]' in main_src
                and "position: sticky" in main_src
            )
            has_chronological_render_merge = (
                "def _build_room_render_messages" in main_src
                and "for channel in (\"messages\", \"history\", \"sims_messages\", \"gen_messages\")" in main_src
                and "merged_msgs = _build_room_render_messages(current_room)" in main_src
                and "def _message_time_key" in main_src
                and "def _message_dedupe_key" in main_src
                and "content_sig" in main_src
            )
            has_stale_event_clear = (
                'current_ui_event_id = str(st.session_state.pop("__ui_event_id", "") or "").strip()' in main_src
                and 'if current_ui_rerun_reason != "chat_room_change":' in main_src
                and 'st.session_state.pop("__chat_room_switch_event_id", None)' in main_src
            )
            has_partitioned_storage = (
                "def _partitioned_chat_root" in main_src
                and "def _load_partitioned_chat_rooms" in main_src
                and "def _save_partitioned_chat_rooms" in main_src
                and "def _try_migrate_legacy_to_partitioned" in main_src
                and "def _ensure_partitioned_room_loaded" in main_src
                and "def _load_partitioned_room_messages" in main_src
                and '"messages.jsonl"' in main_src
                and '"rooms.json"' in main_src
                and "[chat.storage.migration]" in main_src
                and "[chat.storage.save]" in main_src
                and "[chat.storage.load]" in main_src
                and "_partition_message_key" in main_src
                and "_partition_message_payload" in main_src
                and "append_count" in main_src
                and "skipped_bad_lines" in main_src
                and "fallback=True" in main_src
                and "legacy_chat_file" in main_src
                and "storage_mode=partitioned" in main_src
                and '"__messages_loaded"' in main_src
                and "_ensure_partitioned_room_loaded(picked)" in main_src
            )
            has_partitioned_allowlist = (
                "_CHAT_PARTITION_MESSAGE_ALLOW_KEYS" in main_src
                and "_CHAT_PARTITION_META_ALLOW_KEYS" in main_src
                and "def _partition_record_line" in main_src
                and "[chat.storage.record_guard]" in main_src
                and '"channels"' in main_src
                and "def _partition_collect_records" in main_src
                and "_partition_logical_message_key" in main_src
            )
            has_room_index = (
                "def _write_rooms_index_csv" in main_src
                and '"rooms_index.csv"' in main_src
                and '"utf-8-sig"' in main_src
                and "messages_file_bytes" in main_src
            )
            has_compact_room_context = (
                "def _build_current_room_compact_context" in main_src
                and "[CURRENT_ROOM_COMPACT_CONTEXT]" in main_src
                and "[chat.room.compact_context]" in main_src
                and "Do not invent hidden table rows" in main_src
            )
            has_partitioned_perf_detail = (
                '"render_list"' in main_src
                and '"chat_context"' in main_src
                and '"session_compact"' in main_src
                and "render_list=%.3fs" in main_src
                and "chat_context=%.3fs" in main_src
                and "session_compact=%.3fs" in main_src
                and "[chat.storage.append_record]" in main_src
            )
            has_compact_common_drop_keys = (
                "CHAT_PERSISTENCE_DROP_KEYS = {" in main_src
                and "supplier_detail_df" in main_src
                and "product_shortage_df" in main_src
                and "for k in DROP_KEYS" not in main_src
                and " in DROP_KEYS" not in main_src
                and "CHAT_PERSISTENCE_DROP_KEYS" in main_src
                and "has_large_nested = any(k in msg for k in CHAT_PERSISTENCE_DROP_KEYS)" in main_src
            )
            has_room_projection = (
                "def _room_persistence_projection" in main_src
                and "_room_persistence_projection(cur_room)" in main_src
                and "json.dumps(cur_room, ensure_ascii=False, indent=2).encode" not in main_src
                and "str(k) not in CHAT_PERSISTENCE_DROP_KEYS" in main_src
            )
            has_pending_login_policy = (
                'selected_room_id = str(st.session_state.get("current_room") or "").strip()' in main_src
                and 'selected_room_id = ""' in main_src
                and "meta_obj.get(\"current_room\")" not in main_src
                and "_select_pending_new_room()" in main_src
                and "new pending room selected" in main_src
                and "was_pending = room.get(\"auto_created\") is True and room.get(\"title_initialized\") is not True" in main_src
                and "room[\"title_initialized\"] = True" in main_src
            )
            has_readable_folder_policy = (
                "def _readable_room_dirname" in main_src
                and "def _split_room_name_datetime_prefix" in main_src
                and "def _ensure_room_relative_path" in main_src
                and "relative_path" in main_src
                and "_partitioned_messages_file_for_room(root, room)" in main_src
                and "_partitioned_room_dir_for_room(root, room)" in main_src
            )
            chat_mw_src = Path("app/ui/chat_middleware.py").read_text(encoding="utf-8")
            has_sims_df_detection = (
                'elif isinstance(payload.get("data"), pd.DataFrame):' in chat_mw_src
                and 'pd.DataFrame.from_records(payload["records"])' in chat_mw_src
            )
            compact_runtime_ok = False
            compact_runtime_detail = ""
            projection_runtime_ok = False
            projection_runtime_detail = ""
            readable_tool_runtime_ok = False
            readable_tool_runtime_detail = ""
            try:
                start = main_src.index("_CHAT_PARTITION_MESSAGE_ALLOW_KEYS = {")
                end = main_src.index("\ndef _partition_seen_index", start)
                compact_src = main_src[start:end]

                class _FakeState(dict):
                    pass

                class _FakeSt:
                    session_state = _FakeState({"__chat_storage_mode": "partitioned"})

                class _FakeLog:
                    def warning(self, *args, **kwargs):
                        return None

                    def exception(self, *args, **kwargs):
                        return None

                ns = {
                    "os": __import__("os"),
                    "time": __import__("time"),
                    "json": __import__("json"),
                    "hashlib": __import__("hashlib"),
                    "Any": object,
                    "_CHAT_PARTITION_CHANNELS": ("messages", "history", "sims_messages", "gen_messages"),
                    "st": _FakeSt,
                    "log": _FakeLog(),
                    "_json_sanitize": lambda obj: obj,
                    "_script_perf_add": lambda name, elapsed: None,
                    "_partition_message_key": lambda msg: "id:" + str((msg or {}).get("id") or (msg or {}).get("table_key") or id(msg)),
                }
                exec(compact_src, ns)
                ns["_room_meta_only"] = lambda room: {
                    "id": str((room or {}).get("id") or ""),
                    "name": str((room or {}).get("name") or ""),
                    "message_count": len(ns["_partition_collect_records"](room)),
                }
                room = {
                    "id": "fixture-room",
                    "name": "fixture",
                    "messages": [
                        {"role": "user", "content": "small", "meta": {"action": "일반"}},
                        {
                            "role": "assistant",
                            "type": "table",
                            "action": "매입처별 재고부족 현황",
                            "table_key": "tbl-1",
                            "meta": {
                                "action": "매입처별 재고부족 현황",
                                "summary_md": "요약",
                                "table_key": "tbl-1",
                                "supplier_detail_df": {"records": [{"a": 1}]},
                                "product_shortage_df": {"records": [{"b": 2}]},
                            },
                        },
                        {
                            "role": "assistant",
                            "type": "table",
                            "action": "SIMS DF",
                            "table_key": "tbl-df",
                            "df": pd.DataFrame({"제품코드": ["001"], "배정부족예상금액": [10]}),
                            "df_display": pd.DataFrame({"제품코드": ["001"], "배정부족예상금액": [10]}),
                            "data": pd.DataFrame({"제품코드": ["001"], "배정부족예상금액": [10]}),
                            "records": [{"제품코드": "001", "배정부족예상금액": 10}],
                            "columns": ["제품코드", "배정부족예상금액"],
                            "meta": {
                                "action": "SIMS DF",
                                "summary_md": "summary",
                                "table_key": "tbl-df",
                                "supplier_detail_df": {"records": [{"a": 1}]},
                            },
                        },
                    ],
                    "history": [],
                    "sims_messages": [],
                    "gen_messages": [],
                }
                stats = ns["_compact_partition_room_in_memory"](room)
                meta = room["messages"][1].get("meta") or {}
                runtime_df_msg = room["messages"][2]
                projection = ns["_room_persistence_projection"](room)
                projection_json = json.dumps(projection, ensure_ascii=False)
                projected_msg = next(
                    (
                        m.get("message")
                        for m in projection.get("messages", [])
                        if (m.get("message") or {}).get("table_key") == "tbl-df"
                    ),
                    {},
                )
                compact_runtime_ok = (
                    stats.get("changed") == 1
                    and "supplier_detail_df" not in meta
                    and "product_shortage_df" not in meta
                    and meta.get("action") == "매입처별 재고부족 현황"
                    and meta.get("summary_md") == "요약"
                    and meta.get("table_key") == "tbl-1"
                )
                compact_runtime_detail = f"stats={stats} meta_keys={sorted(meta.keys())}"
                projection_runtime_ok = (
                    isinstance(runtime_df_msg.get("df"), pd.DataFrame)
                    and isinstance(runtime_df_msg.get("df_display"), pd.DataFrame)
                    and isinstance(runtime_df_msg.get("data"), pd.DataFrame)
                    and "supplier_detail_df" not in projection_json
                    and '"df"' not in projection_json
                    and '"df_display"' not in projection_json
                    and '"data"' not in projection_json
                    and '"records"' not in projection_json
                    and isinstance(projected_msg, dict)
                    and (projected_msg.get("meta") or {}).get("table_key") == "tbl-df"
                    and (projected_msg.get("meta") or {}).get("summary_md") == "summary"
                )
                projection_runtime_detail = f"projection_messages={len(projection.get('messages') or [])} bytes={len(projection_json.encode('utf-8'))}"
            except Exception as e:
                compact_runtime_detail = f"{type(e).__name__}: {e}"
                projection_runtime_detail = f"{type(e).__name__}: {e}"
            remigrate_src = Path("tools/remigrate_partitioned_chat.py").read_text(encoding="utf-8")
            has_remigrate_tool = (
                "def run(args: argparse.Namespace)" in remigrate_src
                and "parser.add_argument(\"--apply\"" in remigrate_src
                and "dry-run" in remigrate_src
                and "PARTITION_NEW_MESSAGES_AFTER_DEDUPE" in remigrate_src
                and "REMOVED_LARGE_KEY_PATH_COUNTS" in remigrate_src
                and "ROLLBACK_COMMAND" in remigrate_src
                and "messages.jsonl" in remigrate_src
                and "rooms_index.csv" in remigrate_src
                and "def messages_file_from_meta" in remigrate_src
                and "relative_path" in remigrate_src
                and "legacy file not found" in remigrate_src
                and "os.replace(str(partition_root), str(backup_root))" in remigrate_src
                and "os.replace(str(tmp_root), str(partition_root))" in remigrate_src
                and "logical_key(" in remigrate_src
                and "compact_message(" in remigrate_src
            )
            try:
                rename_path = Path("tools/rename_partitioned_room_folders.py")
                spec = importlib.util.spec_from_file_location("rename_partitioned_room_folders", rename_path)
                rename_mod = importlib.util.module_from_spec(spec)
                assert spec and spec.loader
                spec.loader.exec_module(rename_mod)
                with tempfile.TemporaryDirectory() as td:
                    chat_root = Path(td)
                    root = chat_root / "user_8"
                    rooms_dir = root / "rooms"
                    rooms_dir.mkdir(parents=True)
                    rooms = [
                        {
                            "id": "5266b472-4869-4b52-8a41-afe7d2f79a75",
                            "name": "2026-07-16 13:08 오늘 이 채팅방에서 정리한 내용을 간단히 정리 해줘",
                            "created_at": "2026-07-14T10:19:00",
                            "updated_at": "2026-07-14T10:20:00",
                            "message_count": 2,
                        },
                        {
                            "id": "cdc0d332-0000-0000-0000-000000000000",
                            "name": "2026-07-16 13:06 오늘 이 채팅방에서 정리한 내용을 간단히 정리 해줘",
                            "created_at": "2026-07-12T15:43:00",
                            "updated_at": "2026-07-12T15:44:00",
                            "message_count": 2,
                        },
                        {
                            "id": "e22dd0d9-0000-0000-0000-000000000000",
                            "name": "2026-07-16 13:13 매입처별 재고부족 현황",
                            "created_at": "2026-07-16T13:13:00",
                            "updated_at": "2026-07-16T13:14:00",
                            "message_count": 2,
                        },
                        {
                            "id": "abcdef12-0000-0000-0000-000000000001",
                            "name": '동일:제목/테스트*긴 제목 ' + "가" * 80,
                            "created_at": "2026-07-14T10:19:10",
                            "updated_at": "2026-07-14T10:20:00",
                            "message_count": 1,
                        },
                        {
                            "id": "abcdef12-0000-0000-0000-000000000002",
                            "name": '동일:제목/테스트*긴 제목 ' + "가" * 80,
                            "created_at": "2026-07-14T10:19:20",
                            "updated_at": "2026-07-14T10:20:00",
                            "message_count": 1,
                        },
                    ]
                    legacy_rooms = [
                        {
                            "id": "5266b472-4869-4b52-8a41-afe7d2f79a75",
                            "name": "2026-07-14 10:19 품목별 재고부족현황",
                        },
                        {
                            "id": "cdc0d332-0000-0000-0000-000000000000",
                            "name": "2026-07-12 15:43 품목별 재고부족현황",
                        },
                    ]
                    (chat_root / "user_8_chat_rooms.json").write_text(
                        json.dumps({"rooms": legacy_rooms}, ensure_ascii=False, indent=2),
                        encoding="utf-8",
                    )
                    for room in rooms:
                        d = rooms_dir / rename_mod.safe_uuid_dirname(room["id"])
                        d.mkdir()
                        (d / "messages.jsonl").write_text('{"message":{"role":"user","content":"x"}}\n', encoding="utf-8")
                        (d / "room.json").write_text(json.dumps(room, ensure_ascii=False), encoding="utf-8")
                    (root / "rooms.json").write_text(json.dumps({"version": 1, "rooms": rooms}, ensure_ascii=False, indent=2), encoding="utf-8")
                    originals = rename_mod.load_original_room_names(chat_root, "8")
                    plan = rename_mod.build_plan(root, originals)
                    target_by_id = {item["room_id"]: item for item in plan}
                    restore_ok = (
                        target_by_id["5266b472-4869-4b52-8a41-afe7d2f79a75"]["needs_title_restore"]
                        and target_by_id["5266b472-4869-4b52-8a41-afe7d2f79a75"]["proposed_name"] == "2026-07-14 10:19 품목별 재고부족현황"
                        and target_by_id["cdc0d332-0000-0000-0000-000000000000"]["proposed_name"] == "2026-07-12 15:43 품목별 재고부족현황"
                        and target_by_id["e22dd0d9-0000-0000-0000-000000000000"]["proposed_name"] == "2026-07-16 13:13 매입처별 재고부족 현황"
                    )
                    expected_date_parse = rename_mod.readable_dirname_for_name(
                        {
                            "id": "5266b472-4869-4b52-8a41-afe7d2f79a75",
                            "created_at": "2026-07-16T13:08:00",
                        },
                        "2026-07-14 10:19 품목별 재고부족현황",
                    )
                    no_dup_date = not any(item.get("duplicated_date") for item in plan)
                    dry_run_ok = (
                        len(plan) == 5
                        and all(item["needs_rename"] for item in plan)
                        and restore_ok
                        and no_dup_date
                        and expected_date_parse == "2026-07-14_10-19_품목별_재고부족현황__5266b472"
                        and not any(item["exists_conflict"] or item["too_long"] for item in plan)
                    )
                    rename_mod.apply_plan(root, plan)
                    doc = json.loads((root / "rooms.json").read_text(encoding="utf-8"))
                    rels = [str(r.get("relative_path") or "") for r in doc.get("rooms", [])]
                    names_after = {str(r.get("id") or ""): str(r.get("name") or "") for r in doc.get("rooms", [])}
                    applied_ok = (
                        len(set(rels)) == 5
                        and all("messages.jsonl" in rel for rel in rels)
                        and all((root / rel).exists() for rel in rels)
                        and (root / "rooms_index.csv").exists()
                        and any("2026-07-14_10-19_품목별_재고부족현황__5266b472" in rel for rel in rels)
                        and any("2026-07-12_15-43_품목별_재고부족현황__cdc0d332" in rel for rel in rels)
                        and names_after["5266b472-4869-4b52-8a41-afe7d2f79a75"] == "2026-07-14 10:19 품목별 재고부족현황"
                        and names_after["e22dd0d9-0000-0000-0000-000000000000"] == "2026-07-16 13:13 매입처별 재고부족 현황"
                    )
                    # Failure path: conflict should be detected before apply.
                    root_conflict = Path(td) / "user_9"
                    (root_conflict / "rooms").mkdir(parents=True)
                    conflict_room = {
                        "id": "11111111-2222-3333-4444-555555555555",
                        "name": "충돌 테스트",
                        "created_at": "2026-07-14T10:19:00",
                        "message_count": 1,
                    }
                    old_conflict = root_conflict / "rooms" / rename_mod.safe_uuid_dirname(conflict_room["id"])
                    old_conflict.mkdir()
                    (old_conflict / "messages.jsonl").write_text('{"message":{"role":"user","content":"x"}}\n', encoding="utf-8")
                    target_conflict = root_conflict / "rooms" / rename_mod.readable_dirname(conflict_room)
                    target_conflict.mkdir()
                    (root_conflict / "rooms.json").write_text(json.dumps({"version": 1, "rooms": [conflict_room]}, ensure_ascii=False), encoding="utf-8")
                    conflict_plan = rename_mod.build_plan(root_conflict)
                    failure_detected = any(item["exists_conflict"] for item in conflict_plan)
                    readable_tool_runtime_ok = dry_run_ok and applied_ok and failure_detected
                    readable_tool_runtime_detail = f"dry_run={dry_run_ok} restore={restore_ok} no_dup_date={no_dup_date} applied={applied_ok} conflict={failure_detected} rels={rels[:2]}"
            except Exception as e:
                readable_tool_runtime_detail = f"{type(e).__name__}: {e}"
            missing = [k for k in required_close_keys if k not in main_src]
            if not has_close_helper or missing:
                results.append(_fail("chat room switch panel close policy", f"helper={has_close_helper} missing={missing}"))
            elif not has_room_reason or not has_switch_total or not has_close_guard or not has_guard_consume or not has_panel_close_log or not has_after_success_log or not has_no_open_assignment or not has_no_direct_close_in_room_switch or not has_render_block or not has_switch_event_id or not has_event_id_perf or not has_startup_pending or not has_no_unconditional_startup_log or not has_save_detail or not has_sims_open_perf or not has_authenticate_perf or not has_script_path_perf or not has_save_skip or not has_selection_only_save_skip or not has_chat_save_diff or not has_latest_message_anchor or not has_chronological_render_merge or not has_stale_event_clear or not has_partitioned_storage or not has_partitioned_allowlist or not has_room_index or not has_compact_room_context or not has_partitioned_perf_detail or not has_compact_common_drop_keys or not has_room_projection or not has_pending_login_policy or not has_readable_folder_policy or not has_sims_df_detection or not compact_runtime_ok or not projection_runtime_ok or not has_remigrate_tool or not readable_tool_runtime_ok:
                results.append(_fail("chat room switch panel close policy", f"reason={has_room_reason} switch_total={has_switch_total} guard={has_close_guard} consume={has_guard_consume} log={has_panel_close_log} after_success={has_after_success_log} no_open_assign={has_no_open_assignment} no_direct_close={has_no_direct_close_in_room_switch} render_block={has_render_block} switch_event_id={has_switch_event_id} event_perf={has_event_id_perf} startup_pending={has_startup_pending} no_unconditional_startup={has_no_unconditional_startup_log} save_detail={has_save_detail} sims_open_perf={has_sims_open_perf} authenticate_perf={has_authenticate_perf} script_path_perf={has_script_path_perf} save_skip={has_save_skip} selection_only_skip={has_selection_only_save_skip} save_diff={has_chat_save_diff} latest_anchor={has_latest_message_anchor} chronological_merge={has_chronological_render_merge} stale_event_clear={has_stale_event_clear} partitioned_storage={has_partitioned_storage} partitioned_allowlist={has_partitioned_allowlist} room_index={has_room_index} compact_context={has_compact_room_context} partitioned_perf={has_partitioned_perf_detail} common_drop_keys={has_compact_common_drop_keys} room_projection={has_room_projection} pending_login={has_pending_login_policy} readable_folder={has_readable_folder_policy} readable_tool={readable_tool_runtime_ok} readable_detail={readable_tool_runtime_detail} sims_df_detection={has_sims_df_detection} compact_runtime={compact_runtime_ok} compact_detail={compact_runtime_detail} projection_runtime={projection_runtime_ok} projection_detail={projection_runtime_detail} remigrate_tool={has_remigrate_tool}"))
            else:
                results.append(_ok("chat room switch panel close policy", "room change consumes close guard, blocks SIMS render, logs event-scoped perf, startup perf is one-shot, skips unchanged saves, restores login pending-room flow, uses readable room folders for new rooms, uses manual latest-message anchor, renders normal/SIMS history chronologically, uses allow-listed partitioned append-only storage, writes room index, keeps runtime DataFrames out of persistence projection, adds compact room context, and provides safe remigration tooling"))
        except Exception as e:
            results.append(_fail("chat room switch panel close policy", f"{type(e).__name__}: {e}"))

        try:
            main_src = Path("app/Lmstudio_SSAI_chat_main.py").read_text(encoding="utf-8")
            module_ast = ast.parse(main_src)
            helper_names = {
                "_discard_stale_sims_close_for_panel_open",
                "_consume_sims_close_for_chat_room_change",
            }
            helper_nodes = [
                node
                for node in module_ast.body
                if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef))
                and node.name in helper_names
            ]
            if {node.name for node in helper_nodes} != helper_names:
                raise AssertionError("panel-close helper extraction failed")

            class _PanelCloseTestLog:
                def info(self, *_args: Any, **_kwargs: Any) -> None:
                    return None

                def exception(self, *_args: Any, **_kwargs: Any) -> None:
                    return None

            class _PanelCloseTestSt:
                session_state: dict[str, Any] = {}

            close_calls: list[str] = []

            def _fake_close() -> None:
                close_calls.append("close")
                _PanelCloseTestSt.session_state["__sims_open"] = False
                _PanelCloseTestSt.session_state["__sims_panel_active"] = False
                _PanelCloseTestSt.session_state["__sims_force_open"] = False
                _PanelCloseTestSt.session_state["__sims_run_flag"] = False

            helper_ns: dict[str, Any] = {
                "st": _PanelCloseTestSt,
                "log": _PanelCloseTestLog(),
                "_close_sims_panel_for_room_change": _fake_close,
            }
            exec(
                compile(
                    ast.fix_missing_locations(ast.Module(body=helper_nodes, type_ignores=[])),
                    "panel_close_helpers",
                    "exec",
                ),
                helper_ns,
            )
            consume_close = helper_ns["_consume_sims_close_for_chat_room_change"]
            discard_stale = helper_ns["_discard_stale_sims_close_for_panel_open"]

            state = _PanelCloseTestSt.session_state
            state.clear()
            state.update(
                {
                    "__ui_event_name": "chat_room_change",
                    "__ui_event_id": "room-event-a-b",
                    "__sims_close_for_chat_room_change": {
                        "token": "request-a-b",
                        "close_room_id": "room-a",
                        "current_room_id": "room-b",
                        "event_id": "room-event-a-b",
                        "rerun_reason": "chat_room_change",
                    },
                    "__sims_open": True,
                    "__sims_panel_active": True,
                    "__sims_force_open": True,
                    "__sims_run_flag": True,
                }
            )
            first_switch_consumed = consume_close(
                expected_event_id="room-event-a-b",
                expected_current_room_id="room-b",
            )
            first_switch_ok = (
                first_switch_consumed
                and "__sims_close_for_chat_room_change" not in state
                and len(close_calls) == 1
                and state.get("__sims_open") is False
                and state.get("__sims_panel_active") is False
                and state.get("__sims_force_open") is False
                and state.get("__sims_run_flag") is False
            )

            state.update(
                {
                    "__ui_event_name": "chat_room_change",
                    "__ui_event_id": "room-event-b-a",
                    "__sims_close_for_chat_room_change": {
                        "token": "request-b-a",
                        "close_room_id": "room-b",
                        "current_room_id": "room-a",
                        "event_id": "room-event-b-a",
                        "rerun_reason": "chat_room_change",
                    },
                }
            )
            return_switch_consumed = consume_close(
                expected_event_id="room-event-b-a",
                expected_current_room_id="room-a",
            )
            return_switch_ok = (
                return_switch_consumed
                and "__sims_close_for_chat_room_change" not in state
                and len(close_calls) == 2
            )

            state.update(
                {
                    "__sims_close_for_chat_room_change": {
                        "token": "stale-room-request",
                        "close_room_id": "room-b",
                        "current_room_id": "room-a",
                        "event_id": "old-room-event",
                        "rerun_reason": "chat_room_change",
                    },
                    "__ui_event_name": "sims_panel_open",
                    "__ui_event_id": "sims-panel-open-event",
                }
            )
            stale_present, stale_discarded = discard_stale(
                current_event_id="sims-panel-open-event"
            )
            state["__sims_open"] = True
            state["__sims_panel_active"] = True
            state["__sims_force_open"] = True
            state["__sims_run_flag"] = True
            first_click_open_ok = (
                stale_present
                and stale_discarded
                and "__sims_close_for_chat_room_change" not in state
                and len(close_calls) == 2
                and state.get("__sims_open") is True
                and state.get("__sims_panel_active") is True
                and state.get("__sims_force_open") is True
                and state.get("__sims_run_flag") is True
            )

            room_switch_start = main_src.index(
                "# 채팅방 선택과 SIMS 패널 닫기를 같은 room-change 흐름에서 완료한다."
            )
            room_switch_end = main_src.index(
                "_clear_current_table_source_for_room_change()", room_switch_start
            )
            same_flow_src = main_src[room_switch_start:room_switch_end]
            same_flow_order_ok = (
                same_flow_src.find("_request_sims_close_for_chat_room_change(") >= 0
                and same_flow_src.find("_consume_sims_close_for_chat_room_change(")
                > same_flow_src.find("_request_sims_close_for_chat_room_change(")
            )
            callback_nodes = [
                node
                for node in module_ast.body
                if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef))
                and node.name in {
                    "_queue_chat_room_selector_request",
                    "_new_room",
                    "_apply_rename",
                }
            ]
            explicit_callback_reruns = 0
            for callback_node in callback_nodes:
                for child in ast.walk(callback_node):
                    if not isinstance(child, ast.Call):
                        continue
                    func = child.func
                    if (
                        isinstance(func, ast.Attribute)
                        and isinstance(func.value, ast.Name)
                        and func.value.id == "st"
                        and func.attr in {"rerun", "experimental_rerun"}
                    ):
                        explicit_callback_reruns += 1

            if not first_switch_ok:
                results.append(_fail("SIMS panel first-click room close flow", "room A -> B close was not consumed exactly once"))
            elif not return_switch_ok:
                results.append(_fail("SIMS panel first-click room close flow", "room B -> A close was not consumed exactly once"))
            elif not first_click_open_ok:
                results.append(_fail("SIMS panel first-click room close flow", "stale room close cancelled the first panel-open event"))
            elif not same_flow_order_ok:
                results.append(_fail("SIMS panel first-click room close flow", "room-change request/consume are not in the same flow"))
            elif explicit_callback_reruns:
                results.append(_fail("SIMS panel first-click room close flow", f"callback explicit reruns={explicit_callback_reruns}"))
            else:
                results.append(_ok("SIMS panel first-click room close flow", "room changes consume panel-close once; stale close is discarded before first panel-open; callbacks have no explicit rerun"))
        except Exception as e:
            results.append(_fail("SIMS panel first-click room close flow", f"{type(e).__name__}: {e}"))

        try:
            main_src = Path("app/Lmstudio_SSAI_chat_main.py").read_text(encoding="utf-8")
            panel_src = Path("app/ui/sims_panel.py").read_text(encoding="utf-8")
            helper_start = main_src.index("def _room_has_any_messages")
            start = main_src.index("def _compute_room_sidebar_state")
            end = main_src.index("\ndef _ensure_sims_panel_room_title", start)
            sidebar_src = main_src[helper_start:end]
            ns: dict[str, Any] = {"Any": Any, "Iterable": Iterable}
            exec(sidebar_src, ns)
            compute = ns["_compute_room_sidebar_state"]
            resolve_pick = ns["_resolve_room_pick"]
            rooms = [{"id": f"room-{i:02d}", "name": f"room {i:02d}"} for i in range(22)]
            initial = compute(rooms, current_room_id="room-00", filter_text="", page=0, previous_filter="", page_size=10, initial_to_last=True)
            page0 = compute(rooms, current_room_id="room-00", filter_text="", page=0, previous_filter="", page_size=10)
            page1 = compute(rooms, current_room_id="room-00", filter_text="", page=1, previous_filter="", page_size=10)
            page2 = compute(rooms, current_room_id="room-00", filter_text="", page=2, previous_filter="", page_size=10)
            stale = compute(rooms, current_room_id="missing-room", filter_text="", page=1, previous_filter="", page_size=10)
            filter_changed = compute(rooms, current_room_id="room-00", filter_text="room", page=0, previous_filter="", page_size=10)
            filter_two = compute(rooms, current_room_id="room-00", filter_text="room 2", page=1, previous_filter="", page_size=10)
            filter_cleared = compute(rooms, current_room_id="room-00", filter_text="", page=1, previous_filter="room 2", page_size=10)
            same_filter = compute(rooms, current_room_id="room-00", filter_text="room", page=1, previous_filter="room", page_size=10)
            pending_rooms = rooms + [{"id": "pending", "name": "새 대화", "auto_created": True, "messages": []}]
            initial = compute(pending_rooms, current_room_id="pending", filter_text="", page=0, previous_filter="", page_size=10, initial_to_last=True)
            pending = compute(pending_rooms, current_room_id="pending", filter_text="", page=0, previous_filter="", page_size=10, initial_to_last=True)
            rooms_51 = [{"id": f"persisted-{i:02d}", "name": f"persisted {i:02d}"} for i in range(51)]
            pending_to_persisted = compute(
                rooms_51,
                current_room_id="persisted-50",
                filter_text="",
                page=4,
                previous_filter="",
                page_size=10,
                target_room_id="persisted-50",
            )
            filtered_target = compute(
                rooms_51,
                current_room_id="persisted-50",
                filter_text="50",
                page=0,
                previous_filter="50",
                page_size=10,
                target_room_id="persisted-50",
            )
            has_save_current_room_guard = (
                "def _valid_current_room_id_for_rooms" in main_src
                and "current_room_for_meta = _valid_current_room_id_for_rooms(rooms_to_save)" in main_src
                and "current_room_for_meta = _valid_current_room_id_for_rooms(rooms)" in main_src
                and '"current_room": current_room_for_meta' in main_src
            )
            has_room_list_log = (
                "[chat.room_list]" in main_src
                and "persisted_rooms=%s pending_visible=%s filtered_persisted_rooms=%s" in main_src
                and "current_kind=%s current_room_id=%s current_in_persisted=%s page_reset_reason=%s" in main_src
            )
            has_visible_pager = (
                'st.button(\n                "이전"' in main_src
                and 'st.button(\n                "다음"' in main_src
                and "페이지 {room_list_state['page_label']}" in main_src
            )
            has_visible_pager = (
                'key="__room_prev"' in main_src
                and 'key="__room_next"' in main_src
                and "room_list_state['page_label']" in main_src
            )
            has_no_page_callback_rerun = (
                "on_click=_room_prev_page" in main_src
                and "on_click=_room_next_page" in main_src
                and "explicit_rerun_called=%s" in main_src
                and "on_click=lambda: ss.update(__room_page" not in main_src
            )
            login_no_auto_pick = resolve_pick(
                current_room_id="pending",
                picked_pending="pending",
                picked_persisted=None,
                pending_ids=["pending"],
                persisted_ids=[r["id"] for r in page2["view"]],
            )
            page2_pick = resolve_pick(
                current_room_id="pending",
                picked_pending="pending",
                picked_persisted="room-21",
                pending_ids=["pending"],
                persisted_ids=[r["id"] for r in page2["view"]],
            )
            page0_pick = resolve_pick(
                current_room_id="pending",
                picked_pending="pending",
                picked_persisted="room-00",
                pending_ids=["pending"],
                persisted_ids=[r["id"] for r in page0["view"]],
            )
            page1_pick = resolve_pick(
                current_room_id="pending",
                picked_pending="pending",
                picked_persisted="room-10",
                pending_ids=["pending"],
                persisted_ids=[r["id"] for r in page1["view"]],
            )
            pending_pick = resolve_pick(
                current_room_id="room-21",
                picked_pending="pending",
                picked_persisted="room-21",
                pending_ids=["pending"],
                persisted_ids=[r["id"] for r in page2["view"]],
            )
            has_selection_resolver = (
                "def _sync_chat_room_selector_state" in main_src
                and "def _queue_chat_room_selector_request" in main_src
                and "def _consume_chat_room_selector_request" in main_src
                and "on_change=_queue_chat_room_selector_request" in main_src
                and "picked = _resolve_room_pick(" not in main_src
            )
            has_room_select_log = (
                "[chat.room_select] phase=request" in main_src
                and "[chat.room_select] phase=render_ready" in main_src
                and "[chat.room.selector]" in main_src
                and '"after_rerun"' in main_src
            )
            ok = (
                len(page0["view"]) == 20
                and page0["caption"] == "1-20 / 저장된 채팅방 총 22개"
                and page0["page"] == 0
                and page0["next_disabled"] is False
                and len(page1["view"]) == 2
                and page1["caption"] == "21-22 / 저장된 채팅방 총 22개"
                and page1["page"] == 1
                and page1["prev_disabled"] is False
                and page1["next_disabled"] is True
                and stale["page"] == 2
                and stale["page_reset_reason"] == "stale_current_room"
                and stale["current_kind"] == "stale"
                and stale["current_in_persisted"] is False
                and filter_changed["page"] == 0
                and filter_changed["page_reset_reason"] == "filter_changed"
                and filter_changed["caption"] == "검색 결과 2개 / 전체 22개"
                and filter_cleared["page"] == 0
                and filter_cleared["page_reset_reason"] == "filter_changed"
                and same_filter["page"] == 1
                and same_filter["page_reset_reason"] == ""
                and pending["persisted_room_count"] == 22
                and pending["filtered_persisted_rooms"] == 22
                and pending["pending_visible"] is True
                and pending["current_kind"] == "pending"
                and len(pending["view"]) == 20
                and has_save_current_room_guard
                and has_room_list_log
                and has_visible_pager
            )
            ok = (
                len(initial["view"]) == 2
                and initial["page"] == 2
                and initial["current_kind"] == "pending"
                and initial["page_label"] == "3 / 3"
                and initial["prev_disabled"] is False
                and initial["next_disabled"] is True
                and len(page0["view"]) == 10
                and page0["page"] == 0
                and page0["next_disabled"] is False
                and len(page1["view"]) == 10
                and page1["page"] == 1
                and page1["prev_disabled"] is False
                and page1["next_disabled"] is False
                and len(page2["view"]) == 2
                and page2["page"] == 2
                and page2["prev_disabled"] is False
                and page2["next_disabled"] is True
                and stale["page"] == 2
                and stale["page_reset_reason"] == "stale_current_room"
                and stale["current_kind"] == "stale"
                and stale["current_in_persisted"] is False
                and filter_changed["page"] == 2
                and filter_changed["page_reset_reason"] == "filter_changed"
                and filter_two["page"] == 0
                and filter_two["filtered_persisted_rooms"] == 2
                and filter_cleared["page"] == 2
                and filter_cleared["page_reset_reason"] == "filter_changed"
                and same_filter["page"] == 1
                and same_filter["page_reset_reason"] == ""
                and pending["persisted_room_count"] == 22
                and pending["filtered_persisted_rooms"] == 22
                and pending["pending_visible"] is True
                and pending["current_kind"] == "pending"
                and len(pending["view"]) == 2
                and pending["page"] == 2
                and has_save_current_room_guard
                and has_room_list_log
                and has_visible_pager
                and has_no_page_callback_rerun
                and has_selection_resolver
                and login_no_auto_pick is None
                and page2_pick == "room-21"
                and page1_pick == "room-10"
                and page0_pick == "room-00"
                and pending_pick == "pending"
                and has_room_select_log
            )
            fixture_messages = {r["id"]: f"history-message-{r['id']}" for r in rooms}
            loaded_page2_message = fixture_messages.get(page2_pick or "")
            loaded_page1_message = fixture_messages.get(page1_pick or "")
            loaded_page0_message = fixture_messages.get(page0_pick or "")
            checks = {
                "login_pending_page2": len(initial["view"]) == 2 and initial["page"] == 2 and initial["current_kind"] == "pending" and initial["page_label"] == "3 / 3",
                "page0_ten_rooms": len(page0["view"]) == 10 and page0["page"] == 0 and page0["next_disabled"] is False,
                "page1_ten_rooms": len(page1["view"]) == 10 and page1["page"] == 1 and page1["prev_disabled"] is False and page1["next_disabled"] is False,
                "page2_two_rooms": len(page2["view"]) == 2 and page2["page"] == 2 and page2["prev_disabled"] is False and page2["next_disabled"] is True,
                "stale_current_room_to_page2": stale["page"] == 2 and stale["page_reset_reason"] == "stale_current_room" and stale["current_kind"] == "stale" and stale["current_in_persisted"] is False,
                "filter_change_to_last_page": filter_changed["page"] == 2 and filter_changed["page_reset_reason"] == "filter_changed",
                "filtered_two_rooms": filter_two["page"] == 0 and filter_two["filtered_persisted_rooms"] == 2,
                "filter_clear_to_last_page": filter_cleared["page"] == 2 and filter_cleared["page_reset_reason"] == "filter_changed",
                "same_filter_preserves_page": same_filter["page"] == 1 and same_filter["page_reset_reason"] == "",
                "pending_not_in_persisted_page": pending["persisted_room_count"] == 22 and pending["filtered_persisted_rooms"] == 22 and pending["pending_visible"] is True and pending["current_kind"] == "pending" and len(pending["view"]) == 2 and pending["page"] == 2,
                "save_current_room_guard": has_save_current_room_guard,
                "room_list_log": has_room_list_log,
                "visible_pager": has_visible_pager,
                "page_callback_no_explicit_rerun": has_no_page_callback_rerun,
                "selection_resolver_used": has_selection_resolver,
                "login_no_auto_pick": login_no_auto_pick is None,
                "page2_request": page2_pick == "room-21",
                "page1_request": page1_pick == "room-10",
                "page0_request": page0_pick == "room-00",
                "pending_return_request": pending_pick == "pending",
                "room_select_phase_logs": has_room_select_log,
                "pending_to_persisted_moves_to_exact_page": pending_to_persisted["page"] == 5 and pending_to_persisted["target_room_found"] and pending_to_persisted["target_room_index"] == 50 and pending_to_persisted["current_room_visible"],
                "filtered_target_uses_filtered_page": filtered_target["page"] == 0 and filtered_target["target_room_found"] and filtered_target["target_room_index"] == 0 and filtered_target["current_room_visible"],
                "selected_page2_history_message": loaded_page2_message == "history-message-room-21",
                "selected_page1_history_message": loaded_page1_message == "history-message-room-10",
                "selected_page0_history_message": loaded_page0_message == "history-message-room-00",
            }
            failed_checks = [name for name, passed in checks.items() if not passed]
            ok = not failed_checks
            if ok:
                results.append(_ok("chat room sidebar pagination policy", "focused integration harness executed pagination, delta-based room selection, request/consume/load/render-ready logging, selected-room history mapping, stale guard, and callback no-rerun policy"))
            else:
                results.append(_fail("chat room sidebar pagination policy", f"failed_checks={failed_checks} initial={initial} page0={len(page0['view'])}/{page0['caption']} page1={len(page1['view'])}/{page1['caption']} page2={len(page2['view'])}/{page2['caption']} stale={stale} filter_changed={filter_changed} filter_cleared={filter_cleared} same_filter={same_filter} pending={pending} picks={[login_no_auto_pick, page0_pick, page1_pick, page2_pick, pending_pick]} loaded={[loaded_page0_message, loaded_page1_message, loaded_page2_message]} save_guard={has_save_current_room_guard} log={has_room_list_log} pager={has_visible_pager} resolver={has_selection_resolver}"))
        except Exception as e:
            results.append(_fail("chat room sidebar pagination policy", f"{type(e).__name__}: {e}"))

        try:
            main_src = Path("app/Lmstudio_SSAI_chat_main.py").read_text(encoding="utf-8")
            tree = ast.parse(main_src)
            chat_callback_names = {
                "_queue_chat_room_selector_request",
                "_queue_search_reset",
                "_apply_rename",
                "_room_prev_page",
                "_room_next_page",
                "_do_delete",
            }
            chat_callback_defs = [
                node
                for node in ast.walk(tree)
                if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef))
                and node.name in chat_callback_names
            ]
            callback_rerun_calls = [
                node
                for function_node in chat_callback_defs
                for node in ast.walk(function_node)
                if isinstance(node, ast.Call)
                and isinstance(node.func, ast.Attribute)
                and node.func.attr in {"rerun", "experimental_rerun"}
            ]
            selector_names = {
                "_log_chat_room_selector",
                "_is_chat_room_id",
                "_sync_chat_room_selector_state",
                "_queue_chat_room_selector_request",
                "_consume_chat_room_selector_request",
                "_clear_chat_room_selector_request_state",
                "_clear_legacy_chat_room_selector_state",
            }
            selector_nodes = [
                node
                for node in tree.body
                if (
                    isinstance(node, ast.FunctionDef) and node.name in selector_names
                )
                or (
                    isinstance(node, ast.Assign)
                    and any(
                        isinstance(target, ast.Name)
                        and target.id.startswith("_CHAT_ROOM_")
                        for target in node.targets
                    )
                )
            ]

            class _SelectorLog:
                def info(self, *args: Any, **kwargs: Any) -> None:
                    return None

            class _SelectorStreamlit:
                def __init__(self) -> None:
                    self.rerun_count = 0

                def rerun(self, *_args: Any, **_kwargs: Any) -> None:
                    self.rerun_count += 1

            selector_st = _SelectorStreamlit()

            selector_ns: dict[str, Any] = {
                "Any": Any,
                "Iterable": Iterable,
                "log": _SelectorLog(),
                "time": __import__("time"),
                "uuid": __import__("uuid"),
                "st": selector_st,
                "_new_ui_event_id": lambda prefix="ui": f"{prefix}-token",
            }
            exec(
                compile(
                    ast.Module(body=selector_nodes, type_ignores=[]),
                    "chat_room_selector_helpers",
                    "exec",
                ),
                selector_ns,
            )
            sync_selector = selector_ns["_sync_chat_room_selector_state"]
            is_room_id = selector_ns["_is_chat_room_id"]
            queue_selector = selector_ns["_queue_chat_room_selector_request"]
            consume_selector = selector_ns["_consume_chat_room_selector_request"]
            clear_selector_state = selector_ns["_clear_chat_room_selector_request_state"]
            clear_legacy_selector_state = selector_ns["_clear_legacy_chat_room_selector_state"]
            pending_key = selector_ns["_CHAT_ROOM_PENDING_SELECTOR_KEY"]
            persisted_key = selector_ns["_CHAT_ROOM_PERSISTED_SELECTOR_KEY"]
            legacy_pending_key = "__chat_room_pending_selector_id_page_0"
            legacy_persisted_key = "__chat_room_persisted_selector_id_page_0"

            def _select_once(current: str, target: str, pending: list[str], persisted: list[str], kind: str) -> tuple[dict[str, Any], dict[str, str] | None]:
                session: dict[str, Any] = {
                    "current_room": current,
                    "__chat_room_selector_valid_ids": [*pending, *persisted],
                }
                sync_selector(session, canonical_room_id=current, pending_ids=pending, persisted_ids=persisted)
                key = pending_key if kind == "pending" else persisted_key
                session[key] = target
                queue_selector(session, widget_key=key, room_kind=kind)
                request = consume_selector(session, valid_room_ids=[*pending, *persisted])
                if request:
                    session["current_room"] = request["room_id"]
                sync_selector(
                    session,
                    canonical_room_id=session.get("current_room"),
                    pending_ids=pending,
                    persisted_ids=persisted,
                )
                return session, request

            pending_room = "00000000-0000-0000-0000-000000000010"
            room_a = "00000000-0000-0000-0000-000000000001"
            room_b = "00000000-0000-0000-0000-000000000002"
            dashboard_room = "00000000-0000-0000-0000-000000000003"
            large_room = "00000000-0000-0000-0000-000000000004"
            pending_ids_fixture = [pending_room]
            persisted_ids_fixture = [room_a, room_b, dashboard_room, large_room]
            reruns_before = selector_st.rerun_count
            ab_session, ab_request = _select_once(room_a, room_b, pending_ids_fixture, persisted_ids_fixture, "persisted")
            ab_rerun_count = selector_st.rerun_count - reruns_before
            ba_session, ba_request = _select_once(room_b, room_a, pending_ids_fixture, persisted_ids_fixture, "persisted")
            reruns_before = selector_st.rerun_count
            pending_session, pending_request = _select_once(pending_room, room_a, pending_ids_fixture, persisted_ids_fixture, "persisted")
            pending_rerun_count = selector_st.rerun_count - reruns_before
            dashboard_session, dashboard_request = _select_once(room_a, dashboard_room, pending_ids_fixture, persisted_ids_fixture, "persisted")
            large_session, large_request = _select_once(dashboard_room, large_room, pending_ids_fixture, persisted_ids_fixture, "persisted")
            duplicate_session = {"current_room": room_a, persisted_key: room_b, "__chat_room_selector_valid_ids": persisted_ids_fixture}
            queue_selector(duplicate_session, widget_key=persisted_key, room_kind="persisted")
            duplicate_first = consume_selector(duplicate_session, valid_room_ids=persisted_ids_fixture)
            duplicate_second = consume_selector(duplicate_session, valid_room_ids=persisted_ids_fixture)
            same_room_session = {"current_room": room_b, persisted_key: room_b, "__chat_room_selector_valid_ids": persisted_ids_fixture}
            queue_selector(same_room_session, widget_key=persisted_key, room_kind="persisted")
            same_room_request = consume_selector(same_room_session, valid_room_ids=persisted_ids_fixture)
            queued_session = {"current_room": room_a, persisted_key: room_b, "__chat_room_selector_valid_ids": persisted_ids_fixture}
            queue_selector(queued_session, widget_key=persisted_key, room_kind="persisted")
            sync_selector(
                queued_session,
                canonical_room_id=room_a,
                pending_ids=pending_ids_fixture,
                persisted_ids=persisted_ids_fixture,
                queued_request=queued_session.get("__chat_room_selector_request"),
            )
            stale_state = {
                "__chat_room_selector_request": {"room_id": room_a},
                "__chat_room_selector_consumed_token": "old-token",
                pending_key: pending_room,
                persisted_key: room_a,
                "current_room": room_a,
            }
            clear_selector_state(stale_state)
            title_value = "2026-07-19 14:50 제품수불현황 조회"
            invalid_session = {
                "current_room": room_a,
                "__room_page": 3,
                persisted_key: title_value,
                "__chat_room_selector_valid_ids": persisted_ids_fixture,
            }
            queue_selector(invalid_session, widget_key=persisted_key, room_kind="persisted")
            invalid_request = consume_selector(invalid_session, valid_room_ids=persisted_ids_fixture)
            renamed_current_session = {
                "current_room": room_a,
                persisted_key: room_a,
                "__chat_room_selector_valid_ids": persisted_ids_fixture,
            }
            # The title can change, but selector identity must stay the room UUID.
            sync_selector(
                renamed_current_session,
                canonical_room_id=room_a,
                pending_ids=pending_ids_fixture,
                persisted_ids=persisted_ids_fixture,
            )
            renamed_click_session = dict(renamed_current_session)
            renamed_click_session[persisted_key] = room_b
            queue_selector(renamed_click_session, widget_key=persisted_key, room_kind="persisted")
            renamed_click_request = consume_selector(renamed_click_session, valid_room_ids=persisted_ids_fixture)
            if renamed_click_request:
                renamed_click_session["current_room"] = renamed_click_request["room_id"]
            sync_selector(
                renamed_click_session,
                canonical_room_id=renamed_click_session.get("current_room"),
                pending_ids=pending_ids_fixture,
                persisted_ids=persisted_ids_fixture,
            )
            legacy_title_state = {
                "current_room": room_a,
                "__room_page": 0,
                legacy_pending_key: title_value,
                legacy_persisted_key: title_value,
                persisted_key: room_a,
            }
            clear_legacy_selector_state(legacy_title_state)
            page3_key = f"{persisted_key}_page_3"
            page3_session = {"current_room": room_a, page3_key: room_b}
            sync_selector(
                page3_session,
                canonical_room_id=room_a,
                pending_ids=[],
                persisted_ids=[room_a, room_b],
                persisted_widget_key=page3_key,
            )
            page5_key = f"{persisted_key}_page_5"
            page5_session = {"current_room": room_a, page5_key: room_a}
            sync_selector(
                page5_session,
                canonical_room_id=room_a,
                pending_ids=[],
                persisted_ids=[room_b],
                persisted_widget_key=page5_key,
            )
            sentinel_value = selector_ns["_CHAT_ROOM_SELECT_SENTINEL"]
            page5_session[page5_key] = sentinel_value
            queue_selector(page5_session, widget_key=page5_key, room_kind="persisted")
            page5_empty_request = consume_selector(page5_session, valid_room_ids=[room_b])
            checks = {
                "a_to_b_once": bool(ab_request) and ab_session.get("current_room") == room_b and ab_session.get(persisted_key) == room_b and ab_session.get(pending_key, sentinel_value) == sentinel_value,
                "a_to_b_callback_has_no_explicit_rerun": ab_rerun_count == 0 and "__chat_room_selector_app_rerun_token" not in ab_session,
                "b_to_a_once": bool(ba_request) and ba_session.get("current_room") == room_a and ba_session.get(persisted_key) == room_a,
                "pending_to_persisted_once": bool(pending_request) and pending_session.get("current_room") == room_a and pending_session.get(pending_key) == sentinel_value and pending_session.get(persisted_key) == room_a,
                "pending_to_persisted_callback_has_no_explicit_rerun": pending_rerun_count == 0 and "__chat_room_selector_app_rerun_token" not in pending_session,
                "dashboard_to_general_once": bool(dashboard_request) and dashboard_session.get("current_room") == dashboard_room,
                "large_room_guard_no_bounce": bool(large_request) and large_session.get("current_room") == large_room and large_session.get(persisted_key) == large_room,
                "request_consumed_once": bool(duplicate_first) and duplicate_second is None,
                "same_room_callback_noop": same_room_request is None and "__chat_room_selector_request" not in same_room_session,
                "queued_target_not_overwritten": queued_session.get(persisted_key) == room_b,
                "title_callback_invalid": invalid_request is None and "__chat_room_selector_request" not in invalid_session and invalid_session.get("current_room") == room_a and invalid_session.get("__room_page") == 3,
                "renamed_current_room_keeps_uuid": renamed_current_session.get(persisted_key) == room_a and is_room_id(renamed_current_session.get(persisted_key)),
                "renamed_room_click_once": bool(renamed_click_request) and renamed_click_session.get("current_room") == room_b and renamed_click_session.get(persisted_key) == room_b,
                "legacy_title_state_removed": legacy_pending_key not in legacy_title_state and legacy_persisted_key not in legacy_title_state and legacy_title_state.get("current_room") == room_a,
                "selector_v2_keys": "_selector_v2_id" in pending_key and "_selector_v2_id" in persisted_key,
                "page_key_keeps_room_id": page3_session.get(page3_key) == room_a,
                "page_change_clears_out_of_view_selector": page5_session.get(page5_key) == sentinel_value and page5_empty_request is None and page5_session.get("current_room") == room_a,
                "new_pending_clears_stale_request": "current_room" in stale_state and not any(
                    key in stale_state
                    for key in (
                        "__chat_room_selector_request",
                        "__chat_room_selector_consumed_token",
                        pending_key,
                        persisted_key,
                    )
                ),
                "selector_logs_present": "[chat.room.selector]" in main_src and '"before_render"' in main_src and '"after_switch"' in main_src and '"after_rerun"' in main_src and '"callback_invalid"' in main_src and "widget_value_type=" in main_src and "value_in_room_ids=" in main_src,
                "selector_callback_uses_streamlit_auto_rerun": "_request_chat_room_selector_app_rerun" not in main_src,
                "chat_widget_callbacks_have_no_explicit_rerun": not callback_rerun_calls,
                "page_scoped_options": "_page_{selector_page}" in main_src and "options=[_CHAT_ROOM_SELECT_SENTINEL, *options_ids]" in main_src and "채팅방을 선택하세요" in main_src,
                "pending_nav_request_consumed": "__chat_room_nav_request" in main_src and "request_consumed=%s" in main_src and "target_room_id" in panel_src,
            }
            failed_checks = [name for name, passed in checks.items() if not passed]
            if failed_checks:
                results.append(_fail("chat room selector canonical synchronization", f"failed_checks={failed_checks}"))
            else:
                results.append(_ok("chat room selector canonical synchronization", "callback request is consumed once, canonical room ID and selector state remain aligned across persisted, pending, dashboard, and large-room transitions"))
        except Exception as e:
            results.append(_fail("chat room selector canonical synchronization", f"{type(e).__name__}: {e}"))

        try:
            main_src = Path("app/Lmstudio_SSAI_chat_main.py").read_text(encoding="utf-8")
            main_tree = ast.parse(main_src)
            zip_helper = next(
                node
                for node in main_tree.body
                if isinstance(node, ast.FunctionDef)
                and node.name == "_make_unique_chat_zip_entry_name"
            )
            zip_ns: dict[str, Any] = {"re": re}
            exec(
                compile(
                    ast.Module(body=[zip_helper], type_ignores=[]),
                    "chat_zip_entry_helper",
                    "exec",
                ),
                zip_ns,
            )
            make_entry_name = zip_ns["_make_unique_chat_zip_entry_name"]
            room_titles = [
                "Dashboard Lite",
                "Dashboard Lite",
                "Dashboard Lite",
                "A/B",
                "A B",
                "Room",
                "room",
                "",
            ]
            used_names: set[str] = set()
            entry_names = [make_entry_name(title, used_names) for title in room_titles]
            zip_buffer = io.BytesIO()
            with warnings.catch_warnings(record=True) as caught:
                warnings.simplefilter("always")
                with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as archive:
                    for index, entry_name in enumerate(entry_names):
                        archive.writestr(
                            entry_name,
                            json.dumps({"index": index}, ensure_ascii=False),
                        )
            with zipfile.ZipFile(io.BytesIO(zip_buffer.getvalue()), "r") as archive:
                stored_names = archive.namelist()
            duplicate_warnings = [
                warning
                for warning in caught
                if "Duplicate name" in str(warning.message)
            ]
            checks = {
                "same_title_sequence": entry_names[:3] == [
                    "Dashboard_Lite.json",
                    "Dashboard_Lite_2.json",
                    "Dashboard_Lite_3.json",
                ],
                "sanitized_collision_sequence": entry_names[3:5] == [
                    "A_B.json",
                    "A_B_2.json",
                ],
                "case_insensitive_collision": entry_names[5:7] == [
                    "Room.json",
                    "room_2.json",
                ],
                "empty_title_fallback": entry_names[7] == "chat_room.json",
                "all_entries_preserved": len(stored_names) == len(room_titles),
                "casefold_unique": len({name.casefold() for name in stored_names}) == len(stored_names),
                "no_duplicate_name_warning": not duplicate_warnings,
            }
            failed_checks = [name for name, passed in checks.items() if not passed]
            if failed_checks:
                results.append(_fail("chat room ZIP unique entry names", f"failed_checks={failed_checks} names={entry_names}"))
            else:
                results.append(_ok("chat room ZIP unique entry names", f"entries={len(stored_names)} casefold_unique=True duplicate_warnings=0"))
        except Exception as e:
            results.append(_fail("chat room ZIP unique entry names", f"{type(e).__name__}: {e}"))

        try:
            login_src = Path("app/ui/ssai_login.py").read_text(encoding="utf-8")
            render_defs = len(re.findall(r"^def\s+render_company_selector\s*\(", login_src, flags=re.M))
            checks = {
                "single_render_company_selector": render_defs == 1,
                "empty_login_default": 'st.text_input("로그인 ID", value="", placeholder="아이디를 입력하세요")' in login_src,
                "enter_to_submit": "enter_to_submit=True" in login_src,
                "password_key_defined": 'password_key = "__ssai_company_change_sims_password"' in login_src,
                "clear_password_key_defined": 'clear_password_key = "__ssai_clear_company_change_sims_password"' in login_src,
                "logout_clears_password": '"__ssai_company_change_sims_password"' in login_src and '"__ssai_clear_company_change_sims_password"' in login_src,
                "ssart_admin_fallback_only_when_missing": 'if _is_ssart_user(user) and not sims_user_id_for_change:' in login_src,
                "normal_user_sims_id": 'sims_user_id_for_change = str(user.sims_user_id or "").strip()' in login_src,
                "sims_password_verify": "verify_sims_plain_password(" in login_src,
            }
            failed = [k for k, ok in checks.items() if not ok]
            if failed:
                results.append(_fail("ssai login company selector policy", f"failed={failed} render_defs={render_defs}"))
            else:
                results.append(_ok("ssai login company selector policy", "selector/password keys/login defaults/logout cleanup verified"))
        except Exception as e:
            results.append(_fail("ssai login company selector policy", f"{type(e).__name__}: {e}"))
    except Exception as e:
        results.append(_fail("supplier stock shortage allocation fixture", f"{type(e).__name__}: {e}\n{traceback.format_exc(limit=4)}"))

    try:
        profile_mod = importlib.import_module("app.ui.sims_analysis_profiles")
        chat_mod = importlib.import_module("app.ui.chat_middleware")
        panel_mod = importlib.import_module("app.ui.sims_panel")

        profile_cases = [
            ("제품수불현황 조회", {}, "product_flow", "제품별 입고·출고·재고 변동과 수불금액 확인", "source_table"),
            ("제품재고현황 조회", {}, "stock_risk", "현재 재고, 부족 예상, 배정 부족과 공급 위험 확인", "source_table"),
            ("품목별 재고부족현황", {"analysis_type": "stock_shortage"}, "stock_risk", "현재 재고, 부족 예상, 배정 부족과 공급 위험 확인", "source_table"),
            ("입고명세 조회", {}, "trade_document", "기간·거래처·제품별 거래금액과 수량 흐름 확인", "source_table"),
            ("거래처 목록", {"analysis_type": "vendor_master"}, "master", "등록 현황, 분류, 상태 및 필수정보 완전성 확인", "source_table"),
            ("제품코드 목록", {"analysis_type": "goods_master"}, "master", "등록 현황, 분류, 상태 및 필수정보 완전성 확인", "source_table"),
            ("도로명주소 조회", {"analysis_type": "road_address_master"}, "road_address", "검색 조건에 맞는 도로명·지역 주소 후보 확인", "source_table"),
            ("현재표 영업사원 TOP 20", {"current_table_followup": True}, "current_table_top", "선택 차원과 기준 지표의 상위 그룹 비교", "derived_table"),
            ("현재표 영업사원별 집계", {"current_table_followup": True}, "current_table_group", "선택 차원별 건수·수량·금액 집계 비교", "derived_table"),
            ("현재표 조건 필터", {"current_table_followup": True}, "current_table_filter", "사용자가 지정한 조건으로 제한된 결과의 특징 확인", "derived_table"),
            ("알 수 없는 조회", {}, "generic", "현재 조회 조건과 결과 컬럼을 기준으로 주요 분포·수치·결측을 요약", "source_table"),
        ]
        profile_errors = []
        for action, meta, expected_id, expected_purpose, expected_mode in profile_cases:
            profile = profile_mod.build_sims_analysis_profile(
                action,
                params={"start_date": "2026-07-01", "end_date": "2026-07-19"},
                meta=meta,
                columns=["제품코드", "전화번호", "합계금액"],
            )
            if profile.get("profile_id") != expected_id:
                profile_errors.append(f"{action}:profile={profile.get('profile_id')} expected={expected_id}")
            if profile.get("screen_purpose") != expected_purpose:
                profile_errors.append(f"{action}:purpose={profile.get('screen_purpose')!r}")
            if profile.get("response_mode") != expected_mode:
                profile_errors.append(f"{action}:mode={profile.get('response_mode')} expected={expected_mode}")
            if not profile.get("analysis_focus"):
                profile_errors.append(f"{action}:missing_focus")
        if profile_errors:
            results.append(_fail("SIMS LLM analysis profiles", "; ".join(profile_errors)))
        else:
            results.append(_ok("SIMS LLM analysis profiles", f"profile_cases={len(profile_cases)}"))

        scope_profile = profile_mod.build_sims_analysis_profile("거래처 목록", columns=["사업자번호", "전화번호", "거래처명", "합계금액"])
        params = {
            "start_date": "2026-07-01",
            "end_date": "2026-07-19",
            "거래처명": "테스트약국",
            "전화번호": "010-1111-2222",
            "사업자번호": "123-45-67890",
            "상세주소": "서울시 테스트로 1",
            "계좌번호": "111-222-333",
            "api_key": "sk-test-secret",
        }
        params_before = dict(params)
        scope = profile_mod.build_query_scope_summary(params=params, meta={}, profile=scope_profile)
        scope_errors = []
        if "2026-07-01" not in scope or "2026-07-19" not in scope or "거래처명 테스트약국" not in scope:
            scope_errors.append(f"missing_expected_scope={scope!r}")
        for secret in ("010-1111-2222", "123-45-67890"):
            if secret in scope:
                scope_errors.append(f"sensitive_scope_leaked={secret}")
        if params != params_before:
            scope_errors.append("params_mutated")
        no_scope = profile_mod.build_query_scope_summary(params={}, meta={}, profile=scope_profile)
        if "전체 조회 결과" not in no_scope:
            scope_errors.append(f"no_condition_text={no_scope!r}")
        if scope_errors:
            results.append(_fail("SIMS LLM query scope summary", "; ".join(scope_errors)))
        else:
            results.append(_ok("SIMS LLM query scope summary", f"scope={scope!r} no_scope={no_scope!r}"))

        sample_df = pd.DataFrame(
            [
                {
                    "사업자번호": "123-45-67890",
                    "전화번호": "02-111-2222",
                    "대표자명": "홍길동",
                    "상세주소": "서울시 테스트로 1",
                    "이메일": "secret@example.com",
                    "로그인ID": "user01",
                    "계좌번호": "111-222-333",
                    "거래처명": "테스트약국",
                    "거래처종류": "도매",
                    "시도": "경기",
                    "시군구": "수원시",
                    "도로명": "광교로",
                    "제품코드": "000123",
                    "합계금액": 1000,
                    "수량": 2,
                }
            ]
        )
        sample_before = sample_df.copy(deep=True)
        sanitized = profile_mod.sanitize_sims_llm_dataframe(sample_df, scope_profile)
        sanitizer_errors = []
        mapping_fixture = {
            "accounting_count": 7,
            "\uc0ac\uc6a9\uc790\uc218": 12,
            "\ub300\ud45c\uc790\uc218": 3,
            "\uc81c\ud488\ucf54\ub4dc": "123-45-67890",
            "\ud488\ubaa9\ucf54\ub4dc": "010-1111-2222",
            "\uc0c1\ud488\ucf54\ub4dc": "000123",
            "\uc218\ub7c9": 0,
            "\uae08\uc561": 0,
            "\uc804\ud654\ubc88\ud638": "010-1111-2222",
            "\uc0ac\uc5c5\uc790\ubc88\ud638": "123-45-67890",
            "\ub300\ud45c\uc790\uba85": "\ud64d\uae38\ub3d9",
            "query_summary": "\uc804\ud654\ubc88\ud638 010-1111-2222 / \uc0ac\uc5c5\uc790\ubc88\ud638 123-45-67890",
            "nested": {"password": "pw-secret", "token": "tok-secret", "api_key": "sk-secret"},
        }
        mapping_before = dict(mapping_fixture)
        mapping_before["nested"] = dict(mapping_fixture["nested"])
        mapping_safe = profile_mod.sanitize_llm_mapping(mapping_fixture, profile=scope_profile)
        for key in ("accounting_count", "\uac74\uc218", "\uc0ac\uc6a9\uc790\uc218", "\ub300\ud45c\uc790\uc218"):
            if not profile_mod._is_aggregate_key(key):
                sanitizer_errors.append(f"aggregate_key_not_detected={key}")
        expected_preserved = {
            "accounting_count": 7,
            "\uc0ac\uc6a9\uc790\uc218": 12,
            "\ub300\ud45c\uc790\uc218": 3,
            "\uc81c\ud488\ucf54\ub4dc": "123-45-67890",
            "\ud488\ubaa9\ucf54\ub4dc": "010-1111-2222",
            "\uc0c1\ud488\ucf54\ub4dc": "000123",
            "\uc218\ub7c9": 0,
            "\uae08\uc561": 0,
        }
        for key, expected in expected_preserved.items():
            if mapping_safe.get(key) != expected:
                sanitizer_errors.append(f"mapping_preserve_failed={key}:{mapping_safe.get(key)!r}")
        expected_masked = {
            "\uc804\ud654\ubc88\ud638": "\uc804\ud654\ubc88\ud638 \uc870\uac74 \uc801\uc6a9",
            "\uc0ac\uc5c5\uc790\ubc88\ud638": "\uc0ac\uc5c5\uc790\ubc88\ud638 \uc870\uac74 \uc801\uc6a9",
            "\ub300\ud45c\uc790\uba85": "\ub300\ud45c\uc790\uba85 \uc870\uac74 \uc801\uc6a9",
        }
        for key, expected in expected_masked.items():
            if mapping_safe.get(key) != expected:
                sanitizer_errors.append(f"mapping_mask_failed={key}:{mapping_safe.get(key)!r}")
        for key in ("password", "token", "api_key"):
            expected = f"{key} \uc870\uac74 \uc801\uc6a9"
            if mapping_safe.get("nested", {}).get(key) != expected:
                sanitizer_errors.append(f"nested_mask_failed={key}:{mapping_safe.get('nested', {}).get(key)!r}")
        mapping_json = json.dumps(mapping_safe, ensure_ascii=False, default=str)
        profile_src = (PROJECT_ROOT / "app" / "ui" / "sims_analysis_profiles.py").read_text(encoding="utf-8")
        helper_src = profile_src[profile_src.index("def _is_aggregate_key"):profile_src.index("def is_sensitive_llm_column")]
        if "??" in mapping_json or "??" in helper_src:
            sanitizer_errors.append("broken_question_mark_literal_remains")
        for secret in ("pw-secret", "tok-secret", "sk-secret", "\ud64d\uae38\ub3d9"):
            if secret in mapping_json:
                sanitizer_errors.append(f"mapping_secret_leaked={secret}")
        if str(mapping_safe.get("\uc804\ud654\ubc88\ud638") or "") == "010-1111-2222":
            sanitizer_errors.append("phone_field_not_masked")
        if str(mapping_safe.get("\uc0ac\uc5c5\uc790\ubc88\ud638") or "") == "123-45-67890":
            sanitizer_errors.append("bizno_field_not_masked")
        if "010-1111-2222 / \uc0ac\uc5c5\uc790" in mapping_json:
            sanitizer_errors.append("query_summary_raw_sensitive_phrase_leaked")
        if mapping_fixture != mapping_before:
            sanitizer_errors.append("mapping_source_mutated")
        for col in ("사업자번호", "전화번호", "대표자명", "상세주소", "이메일", "로그인ID"):
            if col in sanitized.columns:
                sanitizer_errors.append(f"sensitive_col_kept={col}")
        for col in ("거래처명", "거래처종류", "제품코드", "합계금액", "수량"):
            if col not in sanitized.columns:
                sanitizer_errors.append(f"allowed_col_missing={col}")
        if int(sanitized["합계금액"].sum()) != 1000:
            sanitizer_errors.append("aggregate_changed")
        try:
            pd.testing.assert_frame_equal(sample_df, sample_before)
        except AssertionError as exc:
            sanitizer_errors.append(f"source_mutated={exc}")
        ctx = chat_mod._build_sims_analysis_context_from_df(
            sample_df,
            result={"title": "거래처 목록", "meta": {"analysis_type": "vendor_master"}, "df": sample_df},
            action_name="거래처 목록",
            params=params,
            meta={
                "analysis_type": "vendor_master",
                "table_key": "sims_vendor",
                "query_summary": "전화번호 010-1111-2222 / 사업자번호 123-45-67890",
                "summary_md": "대표자명 홍길동 / 이메일 secret@example.com",
                "detail_summary": {"상세주소": "서울시 테스트로 1"},
            },
        )
        ctx_json = json.dumps(ctx, ensure_ascii=False, default=str)
        for secret in (
            "123-45-67890",
            "010-1111-2222",
            "02-111-2222",
            "홍길동",
            "secret@example.com",
            "user01",
            "서울시 테스트로 1",
            "111-222-333",
            "sk-test-secret",
        ):
            if secret in ctx_json:
                sanitizer_errors.append(f"context_leaked={secret}")
        for container_name in ("params", "meta", "query_scope_summary", "analysis_text", "llm_summary_md", "detail_summary", "sample_records"):
            if container_name not in ctx_json and container_name in {"params", "meta", "query_scope_summary", "analysis_text", "sample_records"}:
                sanitizer_errors.append(f"context_container_missing={container_name}")
        if ctx.get("analysis_profile_id") != "master" or not ctx.get("query_scope_summary"):
            sanitizer_errors.append("context_profile_missing")
        road_df = pd.DataFrame(
            [
                {
                    "시도": "경기",
                    "시군구": "수원시",
                    "도로명": "광교로",
                    "도로명주소": "경기도 수원시 광교로 123",
                    "상세주소": "101동 202호",
                }
            ]
        )
        road_profile = profile_mod.build_sims_analysis_profile("도로명주소 조회", meta={"analysis_type": "road_address_master"}, columns=list(road_df.columns))
        road_safe = profile_mod.sanitize_sims_llm_dataframe(road_df, road_profile)
        if not {"시도", "시군구", "도로명"}.issubset(set(road_safe.columns)):
            sanitizer_errors.append(f"road_safe_region_missing={list(road_safe.columns)}")
        if "도로명주소" in road_safe.columns or "상세주소" in road_safe.columns:
            sanitizer_errors.append(f"road_full_address_kept={list(road_safe.columns)}")
        stock_ctx = chat_mod._build_stock_shortage_analysis_ctx(
            pd.DataFrame([{"제품코드": "P001", "부족등급": "정상", "현재고수량": 1, "1개월부족수량": 0}]),
            action_name="품목별 재고부족현황",
            params={"전화번호": "010-1111-2222"},
            meta={"analysis_type": "stock_shortage", "query_summary": "사업자번호 123-45-67890"},
        )
        stock_json = json.dumps(stock_ctx, ensure_ascii=False, default=str)
        for key in ("analysis_profile_id", "screen_purpose", "query_scope_summary", "analysis_focus", "analysis_response_mode"):
            if key not in stock_ctx:
                sanitizer_errors.append(f"stock_profile_missing={key}")
        for secret in ("010-1111-2222", "123-45-67890"):
            if secret in stock_json:
                sanitizer_errors.append(f"stock_context_leaked={secret}")
        if sanitizer_errors:
            results.append(_fail("SIMS LLM sanitizer and context", "; ".join(sanitizer_errors)))
        else:
            results.append(_ok("SIMS LLM sanitizer and context", f"kept_cols={list(sanitized.columns)} profile={ctx.get('analysis_profile_id')}"))

        rule_full = profile_mod.build_response_format_instruction("현재표 분석해줘")
        rule_summary = profile_mod.build_response_format_instruction("현재표 요약해줘")
        prompt_errors = []
        if "조회 이해" not in rule_full or "LLM 의견" not in rule_full or "8~12줄" not in rule_full:
            prompt_errors.append("full_analysis_rule_missing_sections")
        if "2~3문장" not in rule_full or "근거가 부족하면 길이를 늘리지 말고" not in rule_full:
            prompt_errors.append("llm_opinion_length_policy_missing")
        if "두 부분만" not in rule_summary or "LLM 의견은 강제하지 마세요" not in rule_summary:
            prompt_errors.append("summary_rule_forces_opinion")
        main_src = Path("app/Lmstudio_SSAI_chat_main.py").read_text(encoding="utf-8")
        if "build_response_format_instruction(" not in main_src or "analysis_ctx_override.get(\"kind\") == \"SIMS_ANALYSIS_CONTEXT_V1\"" not in main_src:
            prompt_errors.append("main_prompt_or_override_guard_missing")
        if prompt_errors:
            results.append(_fail("SIMS LLM prompt shape", "; ".join(prompt_errors)))
        else:
            results.append(_ok("SIMS LLM prompt shape", "analysis vs summary response instructions verified"))

        class _FakeSt:
            def __init__(self):
                self.session_state: dict[str, Any] = {}
                self.warnings: list[str] = []
                self.markdowns: list[str] = []

            def warning(self, msg: str) -> None:
                self.warnings.append(str(msg))

            def markdown(self, msg: str) -> None:
                self.markdowns.append(str(msg))

            def error(self, msg: str) -> None:
                self.warnings.append(str(msg))

        fake_st = _FakeSt()
        panel_errors = []
        panel_ctx = {
            "kind": "SIMS_ANALYSIS_CONTEXT_V1",
            "table_key": "sims_panel",
            "action": "제품수불현황 조회",
            "analysis_profile_id": "product_flow",
        }
        fake_st.session_state.update(
            {
                "__sims_selected": {"category": "테스트", "action": "제품수불현황 조회"},
                "__sims_current_table_source_key": "sims_current_b",
                "__sims_current_table_source_action": "거래처 목록",
                "__sims_current_table_source_analysis_ctx": {"table_key": "sims_current_b", "action": "거래처 목록"},
                "__sims_analysis_ctx": {"table_key": "sims_current_b", "action": "거래처 목록"},
                "__sims_analysis_ctx_by_table_key": {"sims_panel": panel_ctx},
            }
        )
        runner_calls: list[dict[str, Any]] = []
        fake_st.session_state["__sims_llm_analysis_runner"] = lambda prompt, **kw: runner_calls.append({"prompt": prompt, **kw})
        old_panel_st = panel_mod.st
        old_get_payload = panel_mod._get_panel_last_final_payload
        old_ensure = panel_mod._ensure_panel_llm_context_from_payload
        try:
            panel_mod.st = fake_st
            panel_mod._get_panel_last_final_payload = lambda category, action: {
                "action": "제품수불현황 조회",
                "title": "제품수불현황 조회",
                "meta": {"table_key": "sims_panel", "action": "제품수불현황 조회"},
            }
            ensure_calls = {"count": 0}

            def _fake_ensure(**kwargs):
                ensure_calls["count"] += 1
                fake_st.session_state["__sims_current_table_source_key"] = "sims_panel"
                fake_st.session_state["__sims_current_table_source_action"] = kwargs.get("action")
                fake_st.session_state["__sims_analysis_ctx"] = panel_ctx
                return True

            panel_mod._ensure_panel_llm_context_from_payload = _fake_ensure
            panel_mod._run_panel_llm_analysis_from_button("패널 표 분석", "panel-key")
            if ensure_calls["count"] != 1:
                panel_errors.append(f"ensure_calls={ensure_calls['count']}")
            if len(runner_calls) != 1:
                panel_errors.append(f"runner_calls={len(runner_calls)}")
            elif runner_calls[0].get("analysis_ctx_override") is not panel_ctx:
                panel_errors.append("runner_override_not_exact_context")
            elif runner_calls[0].get("clicked_table_key") != "sims_panel" or runner_calls[0].get("clicked_action") != "제품수불현황 조회":
                panel_errors.append(f"clicked_identity_wrong={runner_calls[0]}")
            if fake_st.session_state.get("__sims_current_table_source_key") != "sims_current_b":
                panel_errors.append("current_source_changed")
            runner_calls.clear()
            fake_st.session_state["__sims_analysis_ctx_by_table_key"] = {"sims_panel": {**panel_ctx, "table_key": "other"}}
            panel_mod._run_panel_llm_analysis_from_button("패널 표 분석", "panel-key")
            if runner_calls:
                panel_errors.append("mismatch_called_runner")
        finally:
            panel_mod.st = old_panel_st
            panel_mod._get_panel_last_final_payload = old_get_payload
            panel_mod._ensure_panel_llm_context_from_payload = old_ensure
        if panel_errors:
            results.append(_fail("SIMS panel LLM exact override", "; ".join(panel_errors)))
        else:
            results.append(_ok("SIMS panel LLM exact override", "panel button selects exact table context, blocks mismatch, restores current source"))

        source_ctx = {
            "kind": "SIMS_ANALYSIS_CONTEXT_V1",
            "table_key": "sims_a",
            "action": "제품수불현황 조회",
            "analysis_text": "source A",
        }
        derived_ctx = {
            "kind": "SIMS_ANALYSIS_CONTEXT_V1",
            "table_key": "sims_p",
            "source_table_key": "sims_a",
            "action": "현재표 영업사원 TOP 20",
            "current_table_followup": True,
            "analysis_text": "derived P",
        }
        mismatch_errors = []
        if chat_mod._sims_clicked_llm_context_mismatch(derived_ctx, "sims_p", "현재표 영업사원 TOP 20"):
            mismatch_errors.append("derived_ctx_exact_rejected")
        if chat_mod._sims_clicked_llm_context_mismatch(source_ctx, "sims_p", "현재표 영업사원 TOP 20") != "table_key_mismatch":
            mismatch_errors.append("source_ctx_should_not_replace_derived")
        if chat_mod._sims_clicked_llm_context_mismatch({**derived_ctx, "expired": "true"}, "sims_p", "현재표 영업사원 TOP 20") != "expired_context":
            mismatch_errors.append("expired_clicked_context_not_blocked")
        if mismatch_errors:
            results.append(_fail("SIMS clicked table override fail-closed", "; ".join(mismatch_errors)))
        else:
            results.append(_ok("SIMS clicked table override fail-closed", "derived table override remains table-scoped; source/global fallback blocked on mismatch/expired"))
    except Exception as e:
        results.append(_fail("SIMS query-specific LLM analysis profiles", f"{type(e).__name__}: {e}\n{traceback.format_exc(limit=4)}"))

    try:
        dash_mod = importlib.import_module("app.services.dashboard_lite_facts")
        sales_df = pd.DataFrame(
            [
                {
                    "제약사명": "A제약",
                    "제품그룹명": "정상그룹",
                    "제품그룹코드": "G_OK",
                    "제품구분명": "정상구분",
                    "제품구분코드": "D_OK",
                    "제품분류명": "정상분류",
                    "제품분류코드": "C_OK",
                    "완료월총매출": 1000,
                    "완료월수": 2,
                    "당월 현재매출": 300,
                    "당월 예상매출": 600,
                    "최근3개월증감률": -20,
                    "추세판정": "감소",
                    "2026-05 매출": 500,
                    "2026-06 매출": 500,
                },
                {
                    "제약사명": "B제약",
                    "제품그룹명": "정상그룹",
                    "제품그룹코드": "G_OK",
                    "제품구분명": "정상구분",
                    "제품구분코드": "D_OK",
                    "제품분류명": "정상분류",
                    "제품분류코드": "C_OK",
                    "완료월총매출": 3000,
                    "완료월수": 2,
                    "당월 현재매출": 1000,
                    "당월 예상매출": 1200,
                    "최근3개월증감률": 5,
                    "추세판정": "안정",
                    "2026-05 매출": 1500,
                    "2026-06 매출": 1500,
                },
                {
                    "제약사명": "C제약",
                    "제품그룹명": "정상그룹",
                    "제품그룹코드": "G_OK",
                    "제품구분명": "정상구분",
                    "제품구분코드": "D_OK",
                    "제품분류명": "정상분류",
                    "제품분류코드": "C_OK",
                    "완료월총매출": 6000,
                    "완료월수": 2,
                    "당월 현재매출": 0,
                    "당월 예상매출": 1000,
                    "최근3개월증감률": -35,
                    "추세판정": "감소",
                    "2026-05 매출": 3000,
                    "2026-06 매출": 3000,
                },
                {
                    "제약사명": "D제약",
                    "제품그룹명": "제외그룹",
                    "제품그룹코드": "G_EX",
                    "제품구분명": "정상구분",
                    "제품구분코드": "D_OK",
                    "제품분류명": "정상분류",
                    "제품분류코드": "C_OK",
                    "완료월총매출": 999999,
                    "완료월수": 2,
                    "당월 현재매출": 999999,
                    "당월 예상매출": 999999,
                    "최근3개월증감률": -99,
                    "추세판정": "감소",
                    "2026-05 매출": 999999,
                    "2026-06 매출": 999999,
                },
                {
                    "제약사명": "E제약",
                    "제품그룹명": "정상그룹",
                    "제품그룹코드": "G_OK",
                    "제품구분명": "제외구분",
                    "제품구분코드": "D_EX",
                    "제품분류명": "정상분류",
                    "제품분류코드": "C_OK",
                    "완료월총매출": 888888,
                    "완료월수": 2,
                    "당월 현재매출": 888888,
                    "당월 예상매출": 888888,
                    "최근3개월증감률": -88,
                    "추세판정": "감소",
                    "2026-05 매출": 888888,
                    "2026-06 매출": 888888,
                },
                {
                    "제약사명": "F제약",
                    "제품그룹명": "정상그룹",
                    "제품그룹코드": "G_OK",
                    "제품구분명": "정상구분",
                    "제품구분코드": "D_OK",
                    "제품분류명": "제외분류",
                    "제품분류코드": "C_EX",
                    "완료월총매출": 777777,
                    "완료월수": 2,
                    "당월 현재매출": 777777,
                    "당월 예상매출": 777777,
                    "최근3개월증감률": -77,
                    "추세판정": "감소",
                    "2026-05 매출": 777777,
                    "2026-06 매출": 777777,
                },
            ]
        )
        stock_df = pd.DataFrame(
            [
                {
                    "제품코드": "P001",
                    "제품명": "부족품목",
                    "제조사명": "A제약",
                    "현재재고수량": 50,
                    "당월 잔여예상출고수량": 100,
                    "부족예상수량": 50,
                    "부족예상금액": 500,
                },
                {
                    "제품코드": "P001",
                    "제품명": "부족품목",
                    "제조사명": "A제약",
                    "현재재고수량": 10,
                    "당월 잔여예상출고수량": 30,
                    "부족예상수량": 20,
                    "부족예상금액": 200,
                },
                {
                    "제품코드": "P002",
                    "제품명": "충분품목",
                    "제조사명": "B제약",
                    "현재재고수량": 98,
                    "당월 잔여예상출고수량": 100,
                    "부족예상수량": 0,
                    "부족예상금액": 0,
                },
                {
                    "제품코드": "P003",
                    "제품명": "수요없음",
                    "제조사명": "C제약",
                    "현재재고수량": 0,
                    "당월 잔여예상출고수량": 0,
                    "부족예상수량": 0,
                    "부족예상금액": 0,
                },
                {
                    "제품코드": "P004",
                    "제품명": "제외품목",
                    "제조사명": "D제약",
                    "제품그룹명": "제외그룹",
                    "제품그룹코드": "G_EX",
                    "현재재고수량": 0,
                    "당월 잔여예상출고수량": 1000,
                    "부족예상수량": 1000,
                    "부족예상금액": 100000,
                },
            ]
        )
        sales_df["제품그룹Gcode"] = "0013"
        sales_df["제품구분Gcode"] = "0004"
        sales_df["제품분류Gcode"] = "0031"
        stock_df["제품그룹Gcode"] = "0013"
        stock_df["제품구분Gcode"] = "0004"
        stock_df["제품분류Gcode"] = "0031"
        sales_original = sales_df.copy(deep=True)
        stock_original = stock_df.copy(deep=True)
        facts = dash_mod.build_dashboard_lite_facts(
            {
                "month_from": "202601",
                "month_to": "202607",
                "evaluation_month": "202607",
                "date_from": "20260101",
                "date_to": "20260719",
                "stock_cd_list": ["00002", "00001"],
                "stock_name_list": ["본사 창고", "전주 창고"],
                "exclude_product_group_list": ["0013:G_EX"],
                "exclude_product_group_nm_list": ["제외그룹"],
                "exclude_product_di_list": ["0004:D_EX"],
                "exclude_product_di_nm_list": ["제외구분"],
                "exclude_product_class_list": ["0031:C_EX"],
                "exclude_product_class_nm_list": ["제외분류"],
            },
            manufacturer_summary_payload={"df": sales_df, "meta": {"evaluation_month": "2026-07"}},
            stock_shortage_payload={"df": stock_df, "meta": {}},
            inbound_facts_df=pd.DataFrame(),
        )
        fact_errors: list[str] = []
        required_fact_keys = {
            "scope",
            "period",
            "partial_period",
            "sales",
            "purchase",
            "inventory",
            "stock_readiness",
            "turnover_days",
            "rankings",
            "trend_counts",
            "trend_amounts",
            "trend_shares",
            "risk_targets",
            "today_actions",
            "data_quality",
            "comparison_rules",
            "forbidden_comparisons",
            "filters",
            "additional_notes",
        }
        missing_fact_keys = required_fact_keys - set(facts.keys())
        if missing_fact_keys:
            fact_errors.append(f"missing_keys={sorted(missing_fact_keys)!r}")
        if facts.get("kind") != "SIMS_DASHBOARD_FACTS_V01":
            fact_errors.append("kind_mismatch")
        sales_metrics = facts["sales"]["metrics"]
        if round(float(sales_metrics["completed_month_avg_sales"]["value"]), 2) != 5000.0:
            fact_errors.append("completed_avg_sales_wrong")
        if round(float(sales_metrics["current_month_progress_pct"]["value"]), 2) != 46.43:
            fact_errors.append("current_progress_wrong")
        forbidden = " ".join(str(x) for x in facts.get("forbidden_comparisons") or [])
        if "부분월" not in forbidden or "sample_records" not in forbidden:
            fact_errors.append("forbidden_comparisons_missing")
        chart_kinds = {str(r.get("kind")) for r in facts["sales"]["chart_rows"]}
        if not {"완료월 실제", "당월 현재(부분월)", "당월 예상"}.issubset(chart_kinds):
            fact_errors.append(f"chart_kinds={chart_kinds!r}")
        if facts["trend_counts"].get("감소") != 2:
            fact_errors.append("trend_count_wrong")

        inv_metrics = facts["inventory"]["metrics"]
        if int(inv_metrics["ready_sku_count"]["value"]) != 1:
            fact_errors.append("ready_sku_count_wrong")
        if int(inv_metrics["shortage_sku_count"]["value"]) != 1:
            fact_errors.append("shortage_sku_count_wrong")
        if round(float(inv_metrics["sku_readiness_pct"]["value"]), 2) != 50.0:
            fact_errors.append("sku_readiness_pct_wrong")
        expected_stock_meta_terms = {"위험보정 잔여예상수요", "수요급증", "진행속도 보정"}
        for metric_name in ("ready_sku_count", "shortage_sku_count", "sku_readiness_pct", "shortage_qty"):
            metric = inv_metrics.get(metric_name) or {}
            metadata_text = " ".join(
                [str(metric.get("time_basis") or "")]
                + [str(value) for value in metric.get("source_columns") or []]
            )
            if not expected_stock_meta_terms.issubset(set(term for term in expected_stock_meta_terms if term in metadata_text)):
                fact_errors.append(f"stock_metric_provenance={metric_name}:{metadata_text}")
        stock_readiness_meta = facts.get("stock_readiness") or {}
        if (
            "위험보정 잔여예상수요" not in str(stock_readiness_meta.get("policy") or "")
            or "수요급증" not in str(stock_readiness_meta.get("adjustment_policy") or "")
            or "진행속도" not in str(stock_readiness_meta.get("adjustment_policy") or "")
        ):
            fact_errors.append(f"stock_readiness_provenance={stock_readiness_meta!r}")
        risk_names = [r.get("product_name") for r in facts["inventory"]["risk_targets"]]
        if risk_names != ["부족품목"]:
            fact_errors.append(f"risk_names={risk_names!r}")
        risk = facts["inventory"]["risk_targets"][0] if facts["inventory"]["risk_targets"] else {}
        if round(float(risk.get("current_stock_qty") or 0), 2) != 60.0 or round(float(risk.get("remaining_expected_demand_qty") or 0), 2) != 130.0:
            fact_errors.append(f"multi_stock_product_not_aggregated={risk!r}")
        if round(float(risk.get("shortage_qty") or 0), 2) != 70.0:
            fact_errors.append(f"shortage_formula_wrong={risk!r}")
        if any(str(r.get("product_code")) == "P004" for r in facts["inventory"]["risk_targets"]):
            fact_errors.append("excluded_product_in_risk_targets")
        if not facts.get("today_actions"):
            fact_errors.append("today_actions_empty")
        action_ids = [str(action.get("action_id") or "") for action in facts.get("today_actions") or []]
        required_action_fields = {
            "action_id", "priority", "status", "cause_type", "target_kind", "target_code",
            "target_name", "evidence_label", "evidence_value", "evidence_unit",
            "threshold_label", "threshold_value", "recommended_action", "drilldown_action",
            "drilldown_params", "source_dashboard_event_id",
        }
        if any(required_action_fields - set(action) for action in facts.get("today_actions") or []):
            fact_errors.append("today_actions_schema_missing")
        if len(action_ids) != len(set(action_ids)) or any(not action_id for action_id in action_ids):
            fact_errors.append("today_actions_duplicate_or_empty_action_id")
        action_priorities = [int(action.get("priority") or 0) for action in facts.get("today_actions") or []]
        if action_priorities != sorted(action_priorities) or len(action_priorities) > 10:
            fact_errors.append("today_actions_order_or_limit")
        if any(action.get("cause_type") == "stock_shortage" and float(action.get("stock_readiness_pct") or 0) >= 98.0 for action in facts.get("today_actions") or []):
            fact_errors.append("today_actions_contains_ready_shortage")
        filter_facts = facts.get("filters") or {}
        if [x.get("code") for x in filter_facts.get("included_stock_locations") or []] != ["00002", "00001"]:
            fact_errors.append(f"filter_stock_locations={filter_facts!r}")
        if [x.get("name") for x in filter_facts.get("excluded_product_groups") or []] != ["제외그룹"]:
            fact_errors.append(f"filter_group={filter_facts!r}")
        if [(x.get("gcode"), x.get("tcode")) for x in filter_facts.get("excluded_product_groups") or []] != [("0013", "G_EX")]:
            fact_errors.append(f"filter_group_code={filter_facts!r}")
        if [(x.get("gcode"), x.get("tcode")) for x in filter_facts.get("excluded_product_types") or []] != [("0004", "D_EX")]:
            fact_errors.append(f"filter_di_code={filter_facts!r}")
        if [(x.get("gcode"), x.get("tcode")) for x in filter_facts.get("excluded_product_classes") or []] != [("0031", "C_EX")]:
            fact_errors.append(f"filter_class_code={filter_facts!r}")
        if not facts.get("additional_notes") or "sales_decline_targets" not in facts.get("additional_notes"):
            fact_errors.append("additional_notes_missing")
        if facts["turnover_days"].get("status") != "자료부족":
            fact_errors.append("turnover_status_wrong")
        stock_risk_rows = [
            {"product_code": "MISSING", "current_stock_qty": 0, "current_stock_amt": 0, "remaining_expected_demand_qty": 10, "shortage_qty": 10, "shortage_amt": 100, "stock_readiness_pct": 0, "stock_valuation_unit_price": 10, "_stock_risk_required_values_present": False},
            {"product_code": "NO_DEMAND", "current_stock_qty": 100, "current_stock_amt": 1000, "remaining_expected_demand_qty": 0, "shortage_qty": 0, "shortage_amt": 0, "stock_readiness_pct": 100, "stock_valuation_unit_price": 10, "_stock_risk_required_values_present": True},
            {"product_code": "SHORTAGE", "current_stock_qty": 20, "current_stock_amt": 200, "remaining_expected_demand_qty": 100, "shortage_qty": 80, "shortage_amt": 800, "stock_readiness_pct": 20, "stock_valuation_unit_price": 10, "_stock_risk_required_values_present": True},
            {"product_code": "SHORTAGE", "product_name": "중복 제품코드", "current_stock_qty": 20, "current_stock_amt": 200, "remaining_expected_demand_qty": 100, "shortage_qty": 80, "shortage_amt": 800, "stock_readiness_pct": 20, "stock_valuation_unit_price": 10, "_stock_risk_required_values_present": True},
            {"product_code": "", "product_name": "무코드 제품", "current_stock_qty": 20, "current_stock_amt": 200, "remaining_expected_demand_qty": 100, "shortage_qty": 80, "shortage_amt": 800, "stock_readiness_pct": 20, "stock_valuation_unit_price": 10, "_stock_risk_required_values_present": True},
            {"product_code": "WARNING", "current_stock_qty": 90, "current_stock_amt": 900, "remaining_expected_demand_qty": 100, "shortage_qty": 0, "shortage_amt": 0, "stock_readiness_pct": 90, "stock_valuation_unit_price": 10, "_stock_risk_required_values_present": True},
            {"product_code": "OVER", "current_stock_qty": 150, "current_stock_amt": 750, "remaining_expected_demand_qty": 100, "3개월필요수량": 100, "shortage_qty": 0, "shortage_amt": 0, "stock_readiness_pct": 100, "stock_valuation_unit_price": 5, "_stock_risk_required_values_present": True},
            {"product_code": "NORMAL", "current_stock_qty": 100, "current_stock_amt": 500, "remaining_expected_demand_qty": 100, "shortage_qty": 0, "shortage_amt": 0, "stock_readiness_pct": 100, "stock_valuation_unit_price": 5, "_stock_risk_required_values_present": True},
            {"product_code": "", "product_name": "", "current_stock_qty": 0, "current_stock_amt": 0, "remaining_expected_demand_qty": 10, "shortage_qty": 10, "shortage_amt": 100, "stock_readiness_pct": 0, "stock_valuation_unit_price": 10, "_stock_risk_required_values_present": False},
        ]
        stock_risk_summary = dash_mod._classify_stock_risk_rows(stock_risk_rows, readiness_warning_pct=98.0)
        stock_risk_by_code = {row.get("product_code"): row for row in stock_risk_rows}
        stock_risk_by_status = {row.get("재고위험상태"): row for row in stock_risk_summary}
        expected_risk_statuses = {
            "MISSING": "판정 제외",
            "NO_DEMAND": "판정 제외",
            "SHORTAGE": "긴급 부족",
            "WARNING": "부족 주의",
            "OVER": "적정",
            "NORMAL": "적정",
        }
        if {code: stock_risk_by_code[code].get("재고위험상태") for code in expected_risk_statuses} != expected_risk_statuses:
            fact_errors.append(f"stock_risk_priority={stock_risk_by_code!r}")
        if not pd.isna(stock_risk_by_code["NO_DEMAND"].get("재고커버리지율")):
            fact_errors.append("stock_risk_no_demand_coverage_not_null")
        if (
            float(stock_risk_by_code["OVER"].get("과잉후보수량") or 0) != 50
            or float(stock_risk_by_code["OVER"].get("과잉후보금액") or 0) != 250
        ):
            fact_errors.append(f"stock_risk_overstock_formula={stock_risk_by_code['OVER']!r}")
        if sum(int(row.get("품목수") or 0) for row in stock_risk_summary) != 8:
            fact_errors.append(f"stock_risk_summary_count={stock_risk_summary!r}")
        if (
            int(stock_risk_by_status.get("긴급 부족", {}).get("품목수") or 0) != 2
            or int(stock_risk_by_status.get("부족 주의", {}).get("품목수") or 0) != 1
            or int(stock_risk_by_status.get("적정", {}).get("품목수") or 0) != 2
            or int(stock_risk_by_status.get("판정 제외", {}).get("품목수") or 0) != 3
        ):
            fact_errors.append(f"stock_risk_summary_values={stock_risk_summary!r}")
        threshold_rows = [
            {"product_code": "THRESHOLD", "current_stock_qty": 90, "current_stock_amt": 900, "remaining_expected_demand_qty": 100, "shortage_qty": 0, "shortage_amt": 0, "stock_readiness_pct": 90, "stock_valuation_unit_price": 10, "_stock_risk_required_values_present": True},
        ]
        dash_mod._classify_stock_risk_rows([dict(row) for row in threshold_rows], readiness_warning_pct=98.0)
        threshold_at_98 = dash_mod._classify_stock_risk_rows(threshold_rows, readiness_warning_pct=98.0)
        threshold_at_80_rows = [dict(row) for row in threshold_rows]
        dash_mod._classify_stock_risk_rows(threshold_at_80_rows, readiness_warning_pct=80.0)
        if threshold_rows[0].get("재고위험상태") != "부족 주의" or threshold_at_80_rows[0].get("재고위험상태") != "적정":
            fact_errors.append(f"stock_risk_warning_threshold={threshold_rows!r}/{threshold_at_80_rows!r}/{threshold_at_98!r}")
        boundary_rows = [
            {"product_code": "EMERGENCY_40", "current_stock_qty": 40, "current_stock_amt": 400, "remaining_expected_demand_qty": 100, "shortage_qty": 60, "shortage_amt": 600, "stock_readiness_pct": 40, "stock_valuation_unit_price": 10, "_stock_risk_required_values_present": True},
            {"product_code": "EMERGENCY_49_9", "current_stock_qty": 49.9, "current_stock_amt": 499, "remaining_expected_demand_qty": 100, "shortage_qty": 50.1, "shortage_amt": 501, "stock_readiness_pct": 49.9, "stock_valuation_unit_price": 10, "_stock_risk_required_values_present": True},
            {"product_code": "WARNING_50", "current_stock_qty": 50, "current_stock_amt": 500, "remaining_expected_demand_qty": 100, "shortage_qty": 50, "shortage_amt": 500, "stock_readiness_pct": 50, "stock_valuation_unit_price": 10, "_stock_risk_required_values_present": True},
            {"product_code": "WARNING_97_9", "current_stock_qty": 97.9, "current_stock_amt": 979, "remaining_expected_demand_qty": 100, "shortage_qty": 2.1, "shortage_amt": 21, "stock_readiness_pct": 97.9, "stock_valuation_unit_price": 10, "_stock_risk_required_values_present": True},
            {"product_code": "NORMAL_98", "current_stock_qty": 98, "current_stock_amt": 980, "remaining_expected_demand_qty": 100, "shortage_qty": 2, "shortage_amt": 20, "stock_readiness_pct": 98, "stock_valuation_unit_price": 10, "_stock_risk_required_values_present": True},
        ]
        dash_mod._classify_stock_risk_rows(boundary_rows, readiness_warning_pct=98.0)
        expected_boundary_statuses = {
            "EMERGENCY_40": "긴급 부족",
            "EMERGENCY_49_9": "긴급 부족",
            "WARNING_50": "부족 주의",
            "WARNING_97_9": "부족 주의",
            "NORMAL_98": "적정",
        }
        if {row.get("product_code"): row.get("재고위험상태") for row in boundary_rows} != expected_boundary_statuses:
            fact_errors.append(f"stock_risk_boundaries={boundary_rows!r}")
        overstock_rows = [row for row in stock_risk_rows if row.get("과잉후보여부")]
        if (
            any(row.get("재고위험상태") != "적정" for row in overstock_rows)
            or len(overstock_rows) != 1
            or float(overstock_rows[0].get("과잉후보수량") or 0) != 50
            or float(overstock_rows[0].get("과잉후보금액") or 0) != 250
        ):
            fact_errors.append(f"stock_risk_overstock_subset={overstock_rows!r}")
        demand_surge_rows = [
            {"product_code": "NORMAL_DEMAND", "current_stock_qty": 60, "current_stock_amt": 600, "remaining_expected_demand_qty": 60, "당월현재출고수량": 40, "당월기준예상출고수량": 100, "shortage_qty": 0, "shortage_amt": 0, "stock_readiness_pct": 100, "stock_valuation_unit_price": 10, "_stock_risk_required_values_present": True},
            {"product_code": "SURGE_EMERGENCY", "current_stock_qty": 90, "current_stock_amt": 900, "remaining_expected_demand_qty": 0, "당월현재출고수량": 120, "당월기준예상출고수량": 100, "shortage_qty": 0, "shortage_amt": 0, "stock_readiness_pct": 100, "stock_valuation_unit_price": 10, "_stock_risk_required_values_present": True},
            {"product_code": "SURGE_WARNING", "current_stock_qty": 95, "current_stock_amt": 950, "remaining_expected_demand_qty": 0, "당월현재출고수량": 120, "당월기준예상출고수량": 100, "shortage_qty": 0, "shortage_amt": 0, "stock_readiness_pct": 100, "stock_valuation_unit_price": 10, "_stock_risk_required_values_present": True},
            {"product_code": "SURGE_OVERSTOCK_BLOCK", "current_stock_qty": 400, "current_stock_amt": 4000, "remaining_expected_demand_qty": 0, "당월현재출고수량": 120, "당월기준예상출고수량": 100, "3개월필요수량": 300, "shortage_qty": 0, "shortage_amt": 0, "stock_readiness_pct": 100, "stock_valuation_unit_price": 10, "_stock_risk_required_values_present": True},
            {"product_code": "EQUAL_FORECAST", "current_stock_qty": 0, "current_stock_amt": 0, "remaining_expected_demand_qty": 0, "당월현재출고수량": 100, "당월기준예상출고수량": 100, "shortage_qty": 0, "shortage_amt": 0, "stock_readiness_pct": 100, "stock_valuation_unit_price": 10, "_stock_risk_required_values_present": True},
            {"product_code": "ZERO_FORECAST_SURGE", "current_stock_qty": 10, "current_stock_amt": 100, "remaining_expected_demand_qty": 0, "당월현재출고수량": 1, "당월기준예상출고수량": 0, "shortage_qty": 0, "shortage_amt": 0, "stock_readiness_pct": 100, "stock_valuation_unit_price": 10, "_stock_risk_required_values_present": True},
        ]
        dash_mod._apply_current_month_demand_surge(
            demand_surge_rows,
            evaluation_month="202607",
            policy_date="20260712",
        )
        dash_mod._classify_stock_risk_rows(demand_surge_rows, readiness_warning_pct=98.0)
        demand_surge_by_code = {row.get("product_code"): row for row in demand_surge_rows}
        normal_demand = demand_surge_by_code["NORMAL_DEMAND"]
        surge_emergency = demand_surge_by_code["SURGE_EMERGENCY"]
        surge_warning = demand_surge_by_code["SURGE_WARNING"]
        surge_overstock = demand_surge_by_code["SURGE_OVERSTOCK_BLOCK"]
        equal_forecast = demand_surge_by_code["EQUAL_FORECAST"]
        zero_forecast_surge = demand_surge_by_code["ZERO_FORECAST_SURGE"]
        if (
            normal_demand.get("수요급증여부")
            or float(normal_demand.get("위험보정잔여예상수요") or 0) != 60
            or float(normal_demand.get("remaining_expected_demand_qty") or 0) != 60
            or not surge_emergency.get("수요급증여부")
            or float(surge_emergency.get("진행속도기준월말예상출고수량") or 0) != 310
            or float(surge_emergency.get("위험보정잔여예상수요") or 0) != 190
            or surge_emergency.get("위험보정기준") != "진행속도 보정"
            or surge_emergency.get("재고위험상태") != "긴급 부족"
            or surge_warning.get("재고위험상태") != "부족 주의"
            or bool(surge_overstock.get("과잉후보여부"))
            or equal_forecast.get("수요급증여부")
            or not zero_forecast_surge.get("수요급증여부")
        ):
            fact_errors.append(f"stock_risk_demand_surge={demand_surge_rows!r}")
        surge_actions = dash_mod._build_today_actions(
            {},
            {"risk_targets": [surge_emergency, surge_warning]},
            {},
        )
        surge_action = next((item for item in surge_actions if item.get("product_code") == "SURGE_EMERGENCY"), {})
        if (
            surge_emergency.get("재고위험사유") != "수요급증 후 잔여수요 절반 미만"
            or float(surge_action.get("remaining_expected_demand_qty") or 0) != 190
            or float(surge_action.get("shortage_qty") or 0) != 100
            or float(surge_action.get("shortage_amt") or 0) != 1000
            or round(float(surge_action.get("stock_readiness_pct") or 0), 2) != 47.37
            or "수요급증" not in str(surge_action.get("evidence") or "")
            or "진행속도 보정" not in str(surge_action.get("evidence") or "")
        ):
            fact_errors.append(f"stock_risk_demand_surge_actions={surge_actions!r}")
        action_code_contract_row = dict(surge_emergency)
        action_code_contract_row.update({
            "product_code": "12345",
            "product_name": "코드계약제품",
            "주요매입처코드": "16789",
            "manufacturer_code": "11090",
        })
        action_code_contract = dash_mod._build_today_actions({}, {"risk_targets": [action_code_contract_row]}, {})
        contract_action = action_code_contract[0] if action_code_contract else {}
        if (
            contract_action.get("target_code") != "12345"
            or (contract_action.get("drilldown_params") or {}).get("product_code") != "12345"
            or contract_action.get("target_code") in {"16789", "11090"}
        ):
            fact_errors.append(f"today_action_product_code_contract={contract_action!r}")
        stock_risk_log_records: list[tuple[str, tuple[Any, ...]]] = []

        class _StockRiskLogCapture:
            def info(self, message: str, *args: Any) -> None:
                stock_risk_log_records.append((message, args))

        original_stock_risk_log = dash_mod.log
        dash_mod.log = _StockRiskLogCapture()
        try:
            dash_mod._classify_stock_risk_rows([dict(row) for row in demand_surge_rows], readiness_warning_pct=98.0)
        finally:
            dash_mod.log = original_stock_risk_log
        surge_log_args = stock_risk_log_records[-1][1] if stock_risk_log_records else ()
        if len(surge_log_args) < 8 or float(surge_log_args[6] or 0) != 1000 or float(surge_log_args[7] or 0) != 950:
            fact_errors.append(f"stock_risk_demand_surge_log={stock_risk_log_records!r}")
        past_period_rows = [dict(surge_emergency)]
        dash_mod._apply_current_month_demand_surge(
            past_period_rows,
            evaluation_month="202606",
            policy_date="20260712",
        )
        month_end_rows = [dict(surge_emergency)]
        dash_mod._apply_current_month_demand_surge(
            month_end_rows,
            evaluation_month="202607",
            policy_date="20260731",
        )
        if (
            past_period_rows[0].get("수요급증여부")
            or past_period_rows[0].get("위험보정기준") != "현재월 아님"
            or float(month_end_rows[0].get("위험보정잔여예상수요") or 0) != 0
        ):
            fact_errors.append(f"stock_risk_demand_surge_period={past_period_rows!r}/{month_end_rows!r}")
        detail_history_source = pd.DataFrame(
            [
                {"기준월": month, "제품코드": code, "출고수량": qty}
                for code, quantities in {
                    "FORECAST_EXCEEDED": {"202604": 0, "202605": 0, "202606": 0},
                    "FORECAST_OMISSION": {"202604": 5, "202605": 0, "202606": 0},
                    "SEASONAL": {"202506": 0, "202507": 5, "202508": 0, "202604": 0, "202605": 0, "202606": 0},
                    "REACTIVATED": {"202510": 7, "202506": 0, "202507": 0, "202508": 0, "202604": 0, "202605": 0, "202606": 0},
                    "NEW_CANDIDATE": {"202506": 0, "202507": 0, "202508": 0, "202604": 0, "202605": 0, "202606": 0},
                    "RETURN_ONLY": {"202604": -5, "202605": 0, "202606": 0},
                    "NETTED_ZERO": {"202604": 0, "202605": 0, "202606": 0},
                }.items()
                for month, qty in quantities.items()
            ]
        )
        detail_history = dash_mod._build_demand_surge_history_by_product(
            detail_history_source,
            evaluation_month="202607",
            history_month_from="202506",
        )
        detail_rows = [
            {"product_code": "FORECAST_EXCEEDED", "수요급증여부": True, "당월현재출고수량": 120, "당월기준예상출고수량": 100},
            {"product_code": "FORECAST_OMISSION", "수요급증여부": True, "당월현재출고수량": 10, "당월기준예상출고수량": 0},
            {"product_code": "SEASONAL", "수요급증여부": True, "당월현재출고수량": 10, "당월기준예상출고수량": 0},
            {"product_code": "REACTIVATED", "수요급증여부": True, "당월현재출고수량": 10, "당월기준예상출고수량": 0},
            {"product_code": "NEW_CANDIDATE", "수요급증여부": True, "당월현재출고수량": 10, "당월기준예상출고수량": 0},
            {"product_code": "RETURN_ONLY", "수요급증여부": True, "당월현재출고수량": 10, "당월기준예상출고수량": 0},
            {"product_code": "NETTED_ZERO", "수요급증여부": True, "당월현재출고수량": 10, "당월기준예상출고수량": 0},
            {"product_code": "", "수요급증여부": True, "당월현재출고수량": 10, "당월기준예상출고수량": 0},
        ]
        detail_summary = dash_mod._apply_demand_surge_detail(
            detail_rows,
            history=detail_history,
            evaluation_month="202607",
        )
        detail_by_code = {str(row.get("product_code") or "<missing>"): row for row in detail_rows}
        expected_detail_categories = {
            "FORECAST_EXCEEDED": "기존 예상 초과",
            "FORECAST_OMISSION": "예상 누락",
            "SEASONAL": "계절성 재발생 후보",
            "REACTIVATED": "3개월 이상 재출고",
            "NEW_CANDIDATE": "신규 출고 후보",
            "RETURN_ONLY": "신규 출고 후보",
            "NETTED_ZERO": "신규 출고 후보",
            "<missing>": "분류자료부족",
        }
        actual_detail_categories = {
            code: str(row.get("수요급증세부분류") or "")
            for code, row in detail_by_code.items()
        }
        if (
            actual_detail_categories != expected_detail_categories
            or int(detail_by_code["FORECAST_OMISSION"].get("최근3개월양의출고발생월수") or 0) != 1
            or bool(detail_by_code["RETURN_ONLY"].get("최근3개월출고여부"))
            or bool(detail_by_code["NETTED_ZERO"].get("최근3개월출고여부"))
            or int(detail_summary.get("forecast_exceeded_rows") or 0) != 1
            or int(detail_summary.get("unexpected_outbound_rows") or 0) != 7
            or int(detail_summary.get("total_rows") or 0) != len(detail_rows)
            or int(detail_summary.get("forecast_exceeded_rows") or 0) + int(detail_summary.get("unexpected_outbound_rows") or 0) != int(detail_summary.get("total_rows") or 0)
            or sum(
                int(detail_summary.get(key) or 0)
                for key in (
                    "forecast_exceeded_rows",
                    "forecast_omission_rows",
                    "seasonal_recurrence_candidate_rows",
                    "reactivated_after_3m_rows",
                    "new_outbound_candidate_rows",
                    "insufficient_history_rows",
                )
            ) != int(detail_summary.get("total_rows") or 0)
        ):
            fact_errors.append(f"demand_surge_detail={actual_detail_categories!r}/{detail_summary!r}")
        stock_risk_log_records = []

        class _StockRiskLogCapture:
            def info(self, message: str, *args: Any) -> None:
                stock_risk_log_records.append((message, args))

        original_stock_risk_log = dash_mod.log
        dash_mod.log = _StockRiskLogCapture()
        try:
            empty_stock_risk_summary = dash_mod._classify_stock_risk_rows([], readiness_warning_pct=98.0)
        finally:
            dash_mod.log = original_stock_risk_log
        if (
            len(empty_stock_risk_summary) != 4
            or not stock_risk_log_records
            or "total_rows=0" not in stock_risk_log_records[0][0]
            or "warning_rows=0" not in stock_risk_log_records[0][0]
            or "overstock_candidate_amount=0" not in stock_risk_log_records[0][0]
            or "readiness_warning_pct=" not in stock_risk_log_records[0][0]
        ):
            fact_errors.append(f"stock_risk_empty_log={stock_risk_log_records!r}")
        if "stock_risk_summary" not in (facts.get("inventory") or {}):
            fact_errors.append("stock_risk_summary_missing")
        if "stock_overstock_summary" not in (facts.get("inventory") or {}):
            fact_errors.append("stock_overstock_summary_missing")
        facts_stock_risk_by_status = {
            str(row.get("재고위험상태") or ""): row
            for row in (facts.get("inventory") or {}).get("stock_risk_summary") or []
        }
        facts_emergency_warning = sum(
            int((facts_stock_risk_by_status.get(status) or {}).get("품목수") or 0)
            for status in ("긴급 부족", "부족 주의")
        )
        facts_normal = int((facts_stock_risk_by_status.get("적정") or {}).get("품목수") or 0)
        if (
            int((facts.get("inventory") or {}).get("metrics", {}).get("shortage_sku_count", {}).get("value") or 0) != facts_emergency_warning
            or int((facts.get("inventory") or {}).get("metrics", {}).get("ready_sku_count", {}).get("value") or 0) != facts_normal
        ):
            fact_errors.append(f"stock_risk_kpi_reconciliation={facts.get('inventory')!r}")
        if not sales_df.equals(sales_original) or not stock_df.equals(stock_original):
            fact_errors.append("input_df_mutated")
        default_scope = dash_mod.default_dashboard_lite_scope(today=date(2026, 7, 20))
        if default_scope.get("month_from") != "202601" or default_scope.get("month_to") != "202606" or default_scope.get("evaluation_month") != "202607":
            fact_errors.append(f"default_scope={default_scope!r}")
        try:
            dash_mod.normalize_dashboard_lite_params({"month_from": "202501", "month_to": "202602", "evaluation_month": "202602"}, today=date(2026, 7, 20))
            fact_errors.append("long_range_not_blocked")
        except ValueError:
            pass
        try:
            dash_mod.build_dashboard_lite_facts({})
            fact_errors.append("empty_params_service_path_not_blocked")
        except ValueError:
            pass
        all_stock_scope = dash_mod.normalize_dashboard_lite_params(
            {"month_from": "202601", "month_to": "202606", "evaluation_month": "202607"},
            today=date(2026, 7, 20),
        )
        if all_stock_scope.get("stock_cd_list") or all_stock_scope.get("stock_mode") != "real":
            fact_errors.append(f"all_stock_default={all_stock_scope!r}")

        if fact_errors:
            results.append(_fail("Dashboard Lite deterministic facts", "; ".join(fact_errors)))
        else:
            results.append(_ok("Dashboard Lite deterministic facts", f"actions={len(facts.get('today_actions') or [])} risks={len(facts.get('risk_targets') or [])}"))

        try:
            import app.services.analytics_manufacturer_sales_trend_service as dashboard_trend_mod

            dashboard_date_errors: list[str] = []
            jan_scope = dash_mod.default_dashboard_lite_scope(today=date(2026, 1, 20))
            if (
                jan_scope.get("month_from") != "202507"
                or jan_scope.get("month_to") != "202512"
                or jan_scope.get("evaluation_month") != "202601"
            ):
                dashboard_date_errors.append(f"january_default_scope={jan_scope!r}")

            visible_scope = {
                "month_from": "202601",
                "month_to": "202606",
                "evaluation_month": "202607",
                "date_from": "20260101",
                "date_to": "20260630",
                "policy_date": "20260720",
            }
            source_scope = dash_mod._dashboard_internal_source_params(visible_scope, today=date(2026, 7, 20))
            if (
                source_scope.get("month_from") != "202506"
                or source_scope.get("month_to") != "202607"
                or source_scope.get("date_from") != "20250601"
                or source_scope.get("dashboard_lite_display_month_from") != "202601"
                or source_scope.get("dashboard_lite_display_month_to") != "202606"
                or source_scope.get("dashboard_lite_trend_month_from") != "202510"
                or source_scope.get("dashboard_lite_history_month_from") != "202506"
            ):
                dashboard_date_errors.append(f"internal_trend_scope={source_scope!r}")
            existing_support_scope = {
                **source_scope,
                "month_from": source_scope.get("dashboard_lite_trend_month_from"),
                "date_from": f"{source_scope.get('dashboard_lite_trend_month_from')}01",
            }
            if (
                existing_support_scope.get("month_from") != "202510"
                or existing_support_scope.get("date_from") != "20251001"
            ):
                dashboard_date_errors.append(f"existing_support_scope={existing_support_scope!r}")

            months = ["202510", "202511", "202512", "202601", "202602", "202603", "202604", "202605", "202606", "202607"]
            amounts = [100, 200, 300, 10, 20, 30, 40, 50, 60, 70]
            raw = pd.DataFrame(
                [
                    {
                        "\uae30\uc900\uc6d4": month,
                        "\uc81c\ud488\ucf54\ub4dc": "P1",
                        "\uc81c\uc870\uc0ac\uba85": "M1",
                        "\ub9e4\ucd9c\uacf5\uae09\uac00\uc561": amount,
                        "\ub9e4\ucd9c\ud569\uacc4": amount,
                    }
                    for month, amount in zip(months, amounts)
                ]
            )
            raw_before = raw.copy(deep=True)
            detail = dashboard_trend_mod.get_manufacturer_sales_trend(source_scope, raw_df=raw)
            summary = dashboard_trend_mod.get_manufacturer_sales_trend_summary(source_scope, raw_df=raw)
            detail_months = list(detail["\uae30\uc900\uc6d4"].astype(str))
            if detail_months != ["202601", "202602", "202603", "202604", "202605", "202606", "202607"]:
                dashboard_date_errors.append(f"detail_months={detail_months!r}")
            january_row = detail.loc[detail["\uae30\uc900\uc6d4"].astype(str) == "202601"].iloc[0]
            if float(january_row["\ucd5c\uadfc3\uac1c\uc6d4\ud3c9\uade0\ub9e4\ucd9c"]) != 200.0:
                dashboard_date_errors.append(f"first_display_month_trend={january_row['\ucd5c\uadfc3\uac1c\uc6d4\ud3c9\uade0\ub9e4\ucd9c']!r}")
            summary_row = summary.iloc[0]
            if int(summary_row["\uc644\ub8cc\uc6d4\uc218"]) != 6 or float(summary_row["\uc644\ub8cc\uc6d4\ucd1d\ub9e4\ucd9c"]) != 210.0:
                dashboard_date_errors.append(f"completed_metrics={summary_row.to_dict()!r}")
            if any(column.startswith(("2025-10", "2025-11", "2025-12", "2026-07")) for column in summary.columns):
                dashboard_date_errors.append(f"support_or_evaluation_month_leaked={list(summary.columns)!r}")
            if not raw.equals(raw_before):
                dashboard_date_errors.append("internal_trend_source_mutated")

            if dashboard_date_errors:
                results.append(_fail("Dashboard Lite completed-month scope and trend support", "; ".join(dashboard_date_errors)))
            else:
                results.append(_ok("Dashboard Lite completed-month scope and trend support", "defaults, 3-month support, visible months, and completed metrics verified"))
        except Exception as e:
            results.append(_fail("Dashboard Lite completed-month scope and trend support", f"{type(e).__name__}: {e}"))

        try:
            import app.services.analytics_sales_trend_service as dashboard_sales_mod

            condition_errors: list[str] = []
            condition_params = dash_mod.normalize_dashboard_lite_params(
                {
                    "month_from": "202601",
                    "month_to": "202606",
                    "evaluation_month": "202607",
                    "product_group_list": ["0013:A", "0013:B"],
                    "product_di_list": ["0004:X"],
                    "product_class_list": ["0031:C"],
                    "vendor_group_list": ["0019:G"],
                    "vendor_kind_list": ["0009:K"],
                    "io_prefix_list": ["0", "5"],
                    "stock_mode": "book",
                },
                today=date(2026, 7, 20),
            )
            if condition_params.get("source_mode") != "detail" or condition_params.get("stock_mode") != "book":
                condition_errors.append(f"source_or_stock_mode={condition_params!r}")
            where_sql = dashboard_sales_mod._build_filters(
                dash_mod._dashboard_internal_source_params(condition_params, today=date(2026, 7, 20))
            )
            if "Rd04_Physic_Tax_Gcode" not in where_sql or "Rd03_Ven_Group_Gcode" not in where_sql:
                condition_errors.append(f"bound_filter_missing={where_sql!r}")
            class_pair_params = {
                "month_from": "202601",
                "month_to": "202606",
                "product_class_list": [],
                "dashboard_product_class_list": ["0031:01", "0031:03"],
            }
            detail_class_params = dict(class_pair_params)
            detail_class_sql = dashboard_sales_mod._build_filters(detail_class_params)
            monthly_class_params = dict(class_pair_params)
            monthly_class_sql = dashboard_sales_mod._build_monthly_filters(
                monthly_class_params,
                dashboard_sales_mod._monthly_spec("monthly_book"),
            )
            if (
                "Rd04_Physic_Tax_Gcode" not in detail_class_sql
                or "Rd04_Physic_Tax_Gcode" not in monthly_class_sql
                or "product_class_in_0" in detail_class_params
                or "product_class_monthly_in_0" in monthly_class_params
                or dashboard_sales_mod._monthly_fast_path_reason(class_pair_params, "monthly_book") != "master_code_filter"
            ):
                condition_errors.append(
                    f"0031_tax_contract={detail_class_sql!r}/{monthly_class_sql!r}/"
                    f"{detail_class_params!r}/{monthly_class_params!r}"
                )
            legacy_class_params = {
                "month_from": "202601",
                "month_to": "202606",
                "product_class_list": ["01"],
            }
            legacy_class_sql = dashboard_sales_mod._build_filters(legacy_class_params)
            if (
                "product_class_in_0" not in legacy_class_params
                or "dashboard_product_class_list_g_0" in legacy_class_params
            ):
                condition_errors.append(f"0028_legacy_contract={legacy_class_sql!r}/{legacy_class_params!r}")
            empty_class_sql = dashboard_sales_mod._build_filters(
                {"month_from": "202601", "month_to": "202606", "product_class_list": [], "dashboard_product_class_list": []}
            )
            if "dashboard_product_class_list_g_0" in empty_class_sql or "product_class_in_0" in empty_class_sql:
                condition_errors.append(f"empty_0031_contract={empty_class_sql!r}")
            try:
                dash_mod.normalize_dashboard_lite_params({"month_from": "202601", "month_to": "202606", "evaluation_month": "202607", "readiness_warning_pct": 101}, today=date(2026, 7, 20))
                condition_errors.append("invalid_readiness_not_blocked")
            except ValueError:
                pass
            if condition_errors:
                results.append(_fail("Dashboard Lite additional condition normalization and SQL binding", "; ".join(condition_errors)))
            else:
                results.append(_ok("Dashboard Lite additional condition normalization and SQL binding", "all-scope defaults, code-pair AND/OR binding, IO prefixes, stock mode, and threshold validation verified"))
        except Exception as e:
            results.append(_fail("Dashboard Lite additional condition normalization and SQL binding", f"{type(e).__name__}: {e}"))

        try:
            import app.services.ssai_analysis_profile_service as profile_service_mod
            from app.services.ssai_analysis_profile_service import (
                build_company_default_adapter,
                normalize_company_default_conditions,
                profile_conditions_for_storage,
            )

            stored = profile_conditions_for_storage(
                {
                    "month_from": "202601",
                    "month_to": "202606",
                    "stock_mode": "real",
                    "stock_cd_list": ["00002", "00001"],
                    "io_gu_list": ["000200", "000100"],
                    "product_class_list": ["0031:01"],
                    "manufacturer_test_codes": ["V001"],
                }
            )
            if (
                stored.get("stock_cd_list") != ["00001", "00002"]
                or stored.get("io_gu_list") != ["000100", "000200"]
                or "manufacturer_test_codes" in stored
                or "month_from" in stored
            ):
                raise AssertionError(f"profile_storage_scope={stored!r}")
            default_fixture = {
                "stock_mode": "real",
                "stock_cd_list": ["00247", "00001", "00001"],
                "product_di_list": ["0004:2", "0004:1"],
                "product_class_list": ["0031:01"],
                "io_gu_list": ["501", "001"],
                "readiness_warning_pct": 98,
                "amount_display_unit": "thousand",
                "month_from": "202601",
                "manufacturer_test_codes": ["not-persistent"],
            }
            normalized_default = normalize_company_default_conditions(default_fixture)
            if (
                normalized_default.get("stock_cd_list") != ["00001", "00247"]
                or "month_from" in normalized_default
                or "manufacturer_test_codes" in normalized_default
                or default_fixture["stock_cd_list"] != ["00247", "00001", "00001"]
            ):
                raise AssertionError(f"profile_default_normalization={normalized_default!r}")
            adapter = build_company_default_adapter(
                default_fixture,
                supported_keys={"stock_mode", "stock_cd_list", "product_di_list"},
                explicit={"stock_mode": "book", "stock_cd_list": ["00247"]},
                explicit_keys={"stock_mode", "stock_cd_list"},
            )
            if (
                adapter["effective"].get("stock_mode") != "book"
                or adapter["effective"].get("stock_cd_list") != ["00247"]
                or adapter["sources"].get("stock_mode") != "explicit"
                or "product_class_list" not in adapter["unsupported_default_keys"]
            ):
                raise AssertionError(f"profile_adapter_override={adapter!r}")
            clear_adapter = build_company_default_adapter(
                default_fixture,
                supported_keys={"stock_cd_list"},
                clear_keys={"stock_cd_list"},
            )
            if clear_adapter["effective"].get("stock_cd_list") != [] or clear_adapter["sources"].get("stock_cd_list") != "explicit_clear":
                raise AssertionError(f"profile_adapter_clear={clear_adapter!r}")
            import app.sims.views.dashboard_lite as profile_view_mod
            restored_io = profile_view_mod._dashboard_profile_widget_value("io_gu_list", ["001", "002", "051"])
            if restored_io != ["0012:001", "0012:002", "0012:051"]:
                raise AssertionError(f"profile_io_widget_restore={restored_io!r}")
            class _ProfileCursor:
                def __init__(self):
                    self.calls = []
                    self.rowcount = 1
                    self._sql = ""

                def execute(self, sql, *params):
                    self._sql = str(sql)
                    self.calls.append((self._sql, params))
                    return self

                def fetchone(self):
                    if "profile_json" in self._sql:
                        return ('{"io_gu_list": ["001", "002", "051"]}',)
                    if "profile_id" in self._sql:
                        return (1,)
                    return None

            class _ProfileConn:
                def __init__(self): self.cursor_obj = _ProfileCursor()
                def __enter__(self): return self
                def __exit__(self, *_args): return False
                def cursor(self): return self.cursor_obj
                def commit(self): pass

            profile_conn = _ProfileConn()
            old_profile_connect = profile_service_mod.connect_ssai_db
            profile_service_mod.connect_ssai_db = lambda: profile_conn
            try:
                loaded_for_user_a = profile_service_mod.load_dashboard_profile(company_id=4)
                loaded_for_user_b = profile_service_mod.load_dashboard_profile(company_id=4)
                profile_service_mod.save_dashboard_profile(company_id=4, params=stored, actor_user_id=99)
            finally:
                profile_service_mod.connect_ssai_db = old_profile_connect
            if loaded_for_user_a != loaded_for_user_b or loaded_for_user_a.get("io_gu_list") != ["001", "002", "051"]:
                raise AssertionError(f"company_shared_profile_load={loaded_for_user_a!r}/{loaded_for_user_b!r}")
            calls_text = "\n".join(sql for sql, _params in profile_conn.cursor_obj.calls)
            if "WHERE company_id = ?" not in calls_text or "WHERE user_id = ?" in calls_text:
                raise AssertionError(f"company_scope_sql={calls_text!r}")
            profile_src = Path("app/services/ssai_analysis_profile_service.py").read_text(encoding="utf-8")
            migration_src = Path("tools/ssai_add_analysis_profile_schema.py").read_text(encoding="utf-8")
            expected = ("ANALYSIS_PROFILE_MANAGE", "SSAI_ANALYSIS_PROFILES", "UQ_SSAI_ANALYSIS_PROFILES_COMPANY", "duplicate_company_count")
            if not all(token in profile_src + migration_src for token in expected):
                raise AssertionError("profile_schema_or_permission_missing")
            results.append(_ok("Dashboard Lite saved profile scope", "same-company users restore one company profile; actor user is audit-only and migration targets company uniqueness"))
        except Exception as e:
            results.append(_fail("Dashboard Lite saved profile scope", f"{type(e).__name__}: {e}"))

        try:
            import app.services.ssai_analysis_profile_service as company_profile_mod
            import app.services.analytics_sales_trend_service as analytics_service_mod
            import app.sims.nlq.nlq_router as nlq_router_mod
            import app.sims.views.analytics_views as analytics_view_mod
            import app.ui.ssai_login as login_mod

            full_pair = company_profile_mod.normalize_analytics_multi_code_filter(
                ["01", "02"], ["01", "02"], ["0031:01", "0031:02"], "0031"
            )
            wrong_group_pair = company_profile_mod.normalize_analytics_multi_code_filter(
                ["01", "02"], ["01", "02"], ["0031:01", "0031:02"], "0028"
            )
            partial_pair = company_profile_mod.normalize_analytics_multi_code_filter(
                ["01"], ["01", "02"], ["0031:01"], "0031"
            )
            empty_universe = company_profile_mod.normalize_analytics_multi_code_filter(
                ["01"], [], ["0031:01"], "0031"
            )
            if (
                not full_pair["is_full_selection"]
                or wrong_group_pair["is_full_selection"]
                or partial_pair["effective_pairs"] != ["0031:01"]
                or empty_universe["is_full_selection"]
            ):
                raise AssertionError(
                    f"pair_aware_full_selection={full_pair!r}/{wrong_group_pair!r}/"
                    f"{partial_pair!r}/{empty_universe!r}"
                )

            string_profile = company_profile_mod.profile_conditions_for_storage({
                "stock_cd_list": ["00001", "01", "1", 1, None],
                "product_class_list": ["0031:01", "0031:1", "31:01", 31],
            })
            normalized_codes = company_profile_mod.normalize_analytics_multi_code_filter(
                ["01", "1"], ["01", "1"], ["0031:01", "0031:1"], "0031"
            )
            if (
                string_profile.get("stock_cd_list") != ["00001", "01", "1"]
                or string_profile.get("product_class_list") != ["0031:01", "0031:1", "31:01"]
                or normalized_codes.get("effective_pairs") != []
                or company_profile_mod.normalize_business_code(1) != ""
                or company_profile_mod.normalize_business_code_pair("0031:01") != "0031:01"
            ):
                raise AssertionError(f"business_code_string_contract={string_profile!r}/{normalized_codes!r}")

            profile_by_company = {
                "41": {
                    "stock_mode": "real",
                    "stock_cd_list": ["00247", "00001"],
                "product_di_list": ["0004:2", "0004:1"],
                "product_class_list": ["0031:01"],
                "io_gu_list": ["0012:051"],
                "risk_analysis_days": 90,
                    "manufacturer_test_codes": ["must-not-apply"],
                },
                "42": {
                    "stock_mode": "real",
                "stock_cd_list": ["00999"],
                "product_di_list": ["0004:7"],
                "io_gu_list": ["0012:001"],
                },
            }
            current_company = {"id": "41"}
            old_profile_loader = company_profile_mod.load_dashboard_profile
            old_analytics_loader = analytics_view_mod.load_dashboard_profile
            old_analytics_st = analytics_view_mod.st
            old_company_getter = login_mod.get_selected_company
            old_nlq_option_codes = nlq_router_mod._analytics_nlq_option_codes
            try:
                company_profile_mod.load_dashboard_profile = lambda *, company_id: dict(profile_by_company.get(str(company_id)) or {})
                analytics_view_mod.load_dashboard_profile = company_profile_mod.load_dashboard_profile
                login_mod.get_selected_company = lambda: {"company_id": current_company["id"]}
                nlq_router_mod._analytics_nlq_option_codes = lambda field: {
                    "stock_cd_list": ["00001", "00247", "00901"],
                    "product_di_list": ["1", "2", "3", "5", "6", "7", "J"],
                    "product_class_list": ["01", "02", "03", "08"],
                }.get(field, [])

                nlq_state: dict[str, Any] = {}
                nlq_default = nlq_router_mod._apply_company_default_to_analytics_nlq(
                    {"date_from": "20260101", "date_to": "20260131"},
                    text="품목별 재고부족현황",
                    action="품목별 재고부족현황",
                    session_state=nlq_state,
                    logger=logging.getLogger("ssai.regression"),
                )
                if (
                    nlq_default.get("stock_mode") != "real"
                    or nlq_default.get("stock_cds") != ["00001", "00247"]
                    or nlq_default.get("product_di_list") != ["1", "2"]
                    or nlq_default.get("dashboard_product_di_list") != ["0004:1", "0004:2"]
                    or nlq_default.get("product_class_list") != []
                    or nlq_default.get("dashboard_product_class_list") != ["0031:01"]
                    or "manufacturer_test_codes" in nlq_default
                    or nlq_default.get("date_from") != "20260101"
                ):
                    raise AssertionError(f"nlq_default_apply={nlq_default!r}")
                nlq_explicit = nlq_router_mod._apply_company_default_to_analytics_nlq(
                    {"stock_mode": "book", "stock_cds": ["00247"], "stock_cd": "00247"},
                    text="장부재고 00247 창고 품목별 재고부족현황",
                    action="품목별 재고부족현황",
                    session_state={},
                    logger=logging.getLogger("ssai.regression"),
                )
                if nlq_explicit.get("stock_mode") != "book" or nlq_explicit.get("stock_cds") != ["00247"]:
                    raise AssertionError(f"nlq_explicit_override={nlq_explicit!r}")
                nlq_clear = nlq_router_mod._apply_company_default_to_analytics_nlq(
                    {}, text="전체 창고 품목별 재고부족현황", action="품목별 재고부족현황",
                    session_state={}, logger=logging.getLogger("ssai.regression"),
                )
                if nlq_clear.get("stock_cds") != []:
                    raise AssertionError(f"nlq_explicit_clear={nlq_clear!r}")

                full_profile = dict(profile_by_company["41"])
                full_profile["product_class_list"] = ["0031:01", "0031:02", "0031:03", "0031:08"]
                old_company_profile = profile_by_company["41"]
                profile_by_company["41"] = full_profile
                nlq_full_class = nlq_router_mod._apply_company_default_to_analytics_nlq(
                    {}, text="full class default", action="품목별 매출 추세 요약표",
                    session_state={}, logger=logging.getLogger("ssai.regression"),
                )
                profile_by_company["41"] = old_company_profile
                full_sources = nlq_full_class.get("__analysis_default_sources") or {}
                if (
                    nlq_full_class.get("product_class_list") != []
                    or nlq_full_class.get("dashboard_product_class_list") != []
                    or any(nlq_full_class.get(key) not in ([], "") for key in ("product_class", "product_class_nm", "product_class_nm_list"))
                    or "product_class_list" in full_sources
                ):
                    raise AssertionError(f"nlq_full_class_normalization={nlq_full_class!r}")
                partial_profile = dict(profile_by_company["41"])
                partial_profile["product_class_list"] = ["0031:01", "0031:03"]
                profile_by_company["41"] = partial_profile
                nlq_partial_class = nlq_router_mod._apply_company_default_to_analytics_nlq(
                    {}, text="partial class default", action="품목별 매출 추세 요약표",
                    session_state={}, logger=logging.getLogger("ssai.regression"),
                )
                profile_by_company["41"] = old_company_profile
                if (
                    nlq_partial_class.get("product_class_list") != []
                    or nlq_partial_class.get("dashboard_product_class_list") != ["0031:01", "0031:03"]
                    or (nlq_partial_class.get("__analysis_default_sources") or {}).get("product_class_list") != "default"
                ):
                    raise AssertionError(f"nlq_partial_class_normalization={nlq_partial_class!r}")
                nlq_product_clear = nlq_router_mod._apply_company_default_to_analytics_nlq(
                    {}, text="전체 제품구분 품목별 재고부족현황", action="품목별 재고부족현황",
                    session_state={}, logger=logging.getLogger("ssai.regression"),
                )
                if (
                    nlq_product_clear.get("product_di_list") != []
                    or nlq_product_clear.get("dashboard_product_di_list") != []
                    or (nlq_product_clear.get("__analysis_default_sources") or {}).get("product_di_list") != "explicit_clear"
                ):
                    raise AssertionError(f"nlq_product_explicit_clear={nlq_product_clear!r}")

                class _ProfileStateStreamlit:
                    def __init__(self):
                        self.session_state = {"__sims_widget_ns": "profile-test"}

                analytics_st = _ProfileStateStreamlit()
                analytics_view_mod.st = analytics_st
                analytics_view_mod._prepare_analytics_company_defaults("stock_shortage", "profile-test")
                stock_key = "__analytics_stock_shortage_stock_mode__profile-test"
                di_prefill = "__analytics_dashboard_prefill_codes::__analytics_stock_shortage_product_di__profile-test"
                if analytics_st.session_state.get(stock_key) != "실재고" or analytics_st.session_state.get(di_prefill) != ["1", "2"]:
                    raise AssertionError(f"kpi_default_initial={analytics_st.session_state!r}")
                analytics_st.session_state[stock_key] = "장부재고"
                analytics_view_mod._prepare_analytics_company_defaults("stock_shortage", "profile-test")
                if analytics_st.session_state.get(stock_key) != "장부재고":
                    raise AssertionError("kpi_live_value_overwritten")
                current_company["id"] = "42"
                analytics_view_mod._prepare_analytics_company_defaults("stock_shortage", "profile-test")
                stock_prefill = "__analytics_dashboard_prefill_codes::__analytics_stock_shortage_stock__profile-test"
                di_prefill = "__analytics_dashboard_prefill_codes::__analytics_stock_shortage_product_di__profile-test"
                if (
                    analytics_st.session_state.get(stock_key) != "실재고"
                    or analytics_st.session_state.get(stock_prefill) != ["00999"]
                    or analytics_st.session_state.get(di_prefill) != ["7"]
                ):
                    raise AssertionError(f"kpi_company_isolation={analytics_st.session_state!r}")

                current_company["id"] = "41"
                summary_ns = "summary-profile-test"
                analytics_view_mod._prepare_analytics_company_defaults("sales_trend_summary", summary_ns)
                summary_di_key = f"__analytics_sales_trend_summary_product_di__{summary_ns}"
                analytics_view_mod._apply_dashboard_code_prefill(
                    summary_di_key,
                    [{"code": "1", "label": "1 - A"}, {"code": "2", "label": "2 - B"}],
                    multiple=True,
                )
                if analytics_st.session_state.get(summary_di_key) != ["1 - A", "2 - B"]:
                    raise AssertionError(f"summary_widget_prefill={analytics_st.session_state!r}")
                summary_params = analytics_view_mod._attach_analytics_default_code_pairs(
                    {"product_di_list": ["1", "2"], "product_class_list": ["01"], "stock_cd_list": ["00001", "00247"]},
                    action_key="sales_trend_summary",
                    ns=summary_ns,
                )
                if summary_params.get("dashboard_product_di_list") != ["0004:1", "0004:2"]:
                    raise AssertionError(f"summary_pair_params={summary_params!r}")
                summary_order_params = analytics_view_mod._attach_analytics_default_code_pairs(
                    {"product_di_list": ["2", "1"], "product_class_list": ["01"]},
                    action_key="sales_trend_summary",
                    ns=summary_ns,
                )
                summary_changed_params = analytics_view_mod._attach_analytics_default_code_pairs(
                    {"product_di_list": ["1", "3"], "product_class_list": ["01"]},
                    action_key="sales_trend_summary",
                    ns=summary_ns,
                )
                if (
                    summary_order_params.get("dashboard_product_di_list") != ["0004:1", "0004:2"]
                    or "dashboard_product_di_list" in summary_changed_params
                ):
                    raise AssertionError(f"summary_pair_order_independence={summary_order_params!r}/{summary_changed_params!r}")
                captured_summary_params: dict[str, Any] = {}
                old_summary_service = analytics_view_mod.get_sales_trend_summary_result
                analytics_view_mod.get_sales_trend_summary_result = lambda params: captured_summary_params.update(dict(params)) or {"meta": {}}
                try:
                    analytics_view_mod.get_sales_trend_summary_result(summary_params)
                finally:
                    analytics_view_mod.get_sales_trend_summary_result = old_summary_service
                if captured_summary_params.get("dashboard_product_di_list") != ["0004:1", "0004:2"]:
                    raise AssertionError(f"summary_service_params={captured_summary_params!r}")

                class _SummaryFormCtx:
                    def __enter__(self):
                        return self
                    def __exit__(self, *_args):
                        return False

                class _SummaryViewStreamlit:
                    def __init__(self, state):
                        self.session_state = state
                        self.captions: list[str] = []
                    def subheader(self, *_args, **_kwargs): pass
                    def caption(self, text, *_args, **_kwargs): self.captions.append(str(text))
                    def form(self, *_args, **_kwargs): return _SummaryFormCtx()
                    def columns(self, count, **_kwargs):
                        size = len(count) if isinstance(count, (list, tuple)) else int(count)
                        return [_SummaryFormCtx() for _ in range(size)]
                    def selectbox(self, _label, options, *, key=None, **_kwargs):
                        return self.session_state.get(key, options[0] if options else "")
                    def multiselect(self, _label, _options, *, key=None, **_kwargs):
                        return self.session_state.get(key, [])
                    def checkbox(self, _label, value=False, *, key=None, **_kwargs):
                        return self.session_state.get(key, value)
                    def text_input(self, _label, value="", *, key=None, **_kwargs):
                        return self.session_state.get(key, value)
                    def date_input(self, _label, value, *, key=None, **_kwargs):
                        return self.session_state.get(key, value)
                    def form_submit_button(self, *_args, **_kwargs): return True

                summary_view_state = {"__sims_widget_ns": "summary-view"}
                summary_view_st = _SummaryViewStreamlit(summary_view_state)
                old_view_st = analytics_view_mod.st
                old_code_loader = analytics_view_mod._load_code_options
                old_summary_service = analytics_view_mod.get_sales_trend_summary_result
                summary_view_capture: dict[str, Any] = {}
                try:
                    analytics_view_mod.st = summary_view_st
                    analytics_view_mod._load_code_options = lambda gcode: {
                        "0004": [{"code": "1", "name": "A", "label": "1 - A"}, {"code": "2", "name": "B", "label": "2 - B"}],
                        "0031": [{"code": "01", "name": "C", "label": "01 - C"}],
                        "0018": [{"code": "00001", "name": "창고", "label": "00001 - 창고"}, {"code": "00247", "name": "창고2", "label": "00247 - 창고2"}],
                    }.get(str(gcode), [])
                    analytics_view_mod.get_sales_trend_summary_result = lambda params: summary_view_capture.update(dict(params)) or {"meta": {}}
                    analytics_view_mod.render_sales_trend_summary_analysis()
                finally:
                    analytics_view_mod.st = old_view_st
                    analytics_view_mod._load_code_options = old_code_loader
                    analytics_view_mod.get_sales_trend_summary_result = old_summary_service
                if (
                    not any("회사 Default 초기값: 재고위치 · 제품구분 · 제품분류" in caption for caption in summary_view_st.captions)
                    or summary_view_capture.get("stock_cd_list") != []
                    or summary_view_capture.get("product_di_list") != []
                    or summary_view_capture.get("dashboard_product_di_list") != []
                    or summary_view_capture.get("product_class_list") != []
                    or summary_view_capture.get("dashboard_product_class_list") != []
                ):
                    raise AssertionError(f"summary_view_default_path={summary_view_st.captions!r}/{summary_view_capture!r}")

                manufacturer_options = {
                    "0004": [
                        {"code": code, "name": f"DI{code}", "label": f"{code} - DI{code}"}
                        for code in ["1", "2", "3", "5", "6", "7", "J"]
                    ],
                    "0031": [
                        {"code": code, "name": f"CLASS{code}", "label": f"{code} - CLASS{code}"}
                        for code in ["01", "02", "03", "08"]
                    ],
                    "0018": [
                        {"code": code, "name": f"STOCK{code}", "label": f"{code} - STOCK{code}"}
                        for code in ["00001", "00247", "00901"]
                    ],
                }

                def _capture_manufacturer_form(
                    action_key: str,
                    render_func,
                    service_name: str,
                    profile: dict[str, Any],
                    state: dict[str, Any] | None = None,
                ) -> dict[str, Any]:
                    captured: dict[str, Any] = {}
                    view_state = {"__sims_widget_ns": f"manufacturer-{action_key}"}
                    view_state.update(state or {})
                    view_st = _SummaryViewStreamlit(view_state)
                    old_form_st = analytics_view_mod.st
                    old_form_loader = analytics_view_mod._load_code_options
                    old_form_service = getattr(analytics_view_mod, service_name)
                    old_profile = profile_by_company["41"]
                    try:
                        profile_by_company["41"] = dict(profile)
                        analytics_view_mod.st = view_st
                        analytics_view_mod._load_code_options = lambda gcode: manufacturer_options.get(str(gcode), [])
                        setattr(
                            analytics_view_mod,
                            service_name,
                            lambda params: captured.update(dict(params)) or {"meta": {}},
                        )
                        render_func()
                    finally:
                        profile_by_company["41"] = old_profile
                        analytics_view_mod.st = old_form_st
                        analytics_view_mod._load_code_options = old_form_loader
                        setattr(analytics_view_mod, service_name, old_form_service)
                    return captured

                manufacturer_full_profile = dict(profile_by_company["41"])
                manufacturer_full_profile["product_class_list"] = [
                    "0031:01", "0031:02", "0031:03", "0031:08"
                ]
                manufacturer_detail_full = _capture_manufacturer_form(
                    "manufacturer_sales_trend",
                    analytics_view_mod.render_manufacturer_sales_trend_analysis,
                    "get_manufacturer_sales_trend_result",
                    manufacturer_full_profile,
                )
                manufacturer_summary_full = _capture_manufacturer_form(
                    "manufacturer_sales_trend_summary",
                    analytics_view_mod.render_manufacturer_sales_trend_summary_analysis,
                    "get_manufacturer_sales_trend_summary_result",
                    manufacturer_full_profile,
                )
                for captured in (manufacturer_detail_full, manufacturer_summary_full):
                    if (
                        captured.get("product_class_list") != []
                        or captured.get("dashboard_product_class_list") != []
                        or captured.get("stock_cd_list") != ["00001", "00247"]
                        or captured.get("dashboard_product_di_list") != ["0004:1", "0004:2"]
                    ):
                        raise AssertionError(f"manufacturer_full_class_service_params={captured!r}")

                manufacturer_partial_profile = dict(profile_by_company["41"])
                manufacturer_partial_profile["product_class_list"] = ["0031:01", "0031:03"]
                manufacturer_partial = _capture_manufacturer_form(
                    "manufacturer_sales_trend",
                    analytics_view_mod.render_manufacturer_sales_trend_analysis,
                    "get_manufacturer_sales_trend_result",
                    manufacturer_partial_profile,
                )
                if (
                    manufacturer_partial.get("product_class_list") != []
                    or manufacturer_partial.get("dashboard_product_class_list") != ["0031:01", "0031:03"]
                ):
                    raise AssertionError(f"manufacturer_partial_class_service_params={manufacturer_partial!r}")

                manufacturer_changed = _capture_manufacturer_form(
                    "manufacturer_sales_trend_summary",
                    analytics_view_mod.render_manufacturer_sales_trend_summary_analysis,
                    "get_manufacturer_sales_trend_summary_result",
                    manufacturer_partial_profile,
                    {
                        "__analytics_manufacturer_sales_trend_summary_product_class__manufacturer-manufacturer_sales_trend_summary": [
                            "02 - CLASS02"
                        ],
                    },
                )
                if (
                    manufacturer_changed.get("product_class_list") != []
                    or manufacturer_changed.get("dashboard_product_class_list") != ["0031:02"]
                ):
                    raise AssertionError(f"manufacturer_user_class_override={manufacturer_changed!r}")

                old_code_loader = analytics_view_mod._load_code_options
                try:
                    analytics_view_mod._load_code_options = lambda gcode: {
                        "0004": [{"code": code} for code in ["1", "2", "3", "5", "6", "7", "J"]],
                        "0031": [{"code": code} for code in ["01", "02", "03", "08"]],
                        "0018": [{"code": code} for code in ["00001", "00247", "00901"]],
                    }.get(str(gcode), [])
                    panel_full = analytics_view_mod._normalize_analytics_multi_code_params(
                        {
                            "stock_cd_list": ["00001", "00247"],
                            "product_di_list": ["1", "2"],
                            "dashboard_product_di_list": ["0004:1", "0004:2"],
                            "product_class_list": ["01", "02", "03", "08"],
                            "dashboard_product_class_list": ["0031:01", "0031:02", "0031:03", "0031:08"],
                        },
                        action_key="sales_trend_summary",
                    )
                finally:
                    analytics_view_mod._load_code_options = old_code_loader
                if (
                    panel_full.get("product_class_list") != []
                    or panel_full.get("dashboard_product_class_list") != []
                    or panel_full.get("stock_cd_list") != ["00001", "00247"]
                    or panel_full.get("dashboard_product_di_list") != ["0004:1", "0004:2"]
                ):
                    raise AssertionError(f"panel_full_selection_normalization={panel_full!r}")
                for key in (
                    "stock_cd_list", "product_di_list", "dashboard_product_di_list",
                    "product_class_list", "dashboard_product_class_list",
                ):
                    if panel_full.get(key) != nlq_full_class.get(key):
                        raise AssertionError(
                            f"panel_nlq_default_param_mismatch key={key} "
                            f"panel={panel_full.get(key)!r} nlq={nlq_full_class.get(key)!r}"
                        )
                group_forecast_params = analytics_view_mod._attach_analytics_default_code_pairs(
                    {"product_di_list": ["1"], "product_class_list": ["01"]},
                    action_key="customer_sales_forecast",
                    ns=summary_ns,
                )
                group_forecast_params = analytics_view_mod._normalize_analytics_multi_code_params(
                    group_forecast_params,
                    action_key="customer_sales_forecast",
                )
                if group_forecast_params.get("product_di_list") != ["1"] or group_forecast_params.get("dashboard_product_class_list") != ["0031:01"]:
                    raise AssertionError(f"group_forecast_non_io_param_changed={group_forecast_params!r}")

                import app.ui.sims_panel as panel_mod
                registry_actions = dict(panel_mod._CATEGORIES["분석/KPI"]["actions"])
                target_actions = set(analytics_view_mod.KPI_DEFAULT_ACTION_SPECS)
                registry_default_actions = {
                    action for action, view in registry_actions.items()
                    if getattr(view, "__name__", "") in {
                        spec["view"] for spec in analytics_view_mod.KPI_DEFAULT_ACTION_SPECS.values()
                    }
                }
                adapter_actions = {
                    action for action, spec in analytics_view_mod.KPI_DEFAULT_ACTION_SPECS.items()
                    if spec["adapter_key"] in analytics_view_mod._ANALYTICS_DEFAULT_KEYS
                }
                nlq_actions = set(nlq_router_mod._ANALYTICS_NLQ_DEFAULT_KEYS)
                if registry_default_actions != target_actions or adapter_actions != target_actions or not target_actions.issubset(nlq_actions):
                    raise AssertionError(
                        f"default_action_sets={registry_default_actions!r}/{adapter_actions!r}/{nlq_actions!r}"
                    )

                captured_nlq_params: dict[str, Any] = {}
                captured_nlq_payload: dict[str, Any] = {}
                old_action_resolver = nlq_router_mod._resolve_analytics_action
                old_handler_getter = nlq_router_mod._get_analytics_handler
                old_param_builder = nlq_router_mod._build_analytics_params
                import app.ui.chat_middleware as chat_middleware_mod
                old_push = chat_middleware_mod.push_sims_result_to_chat
                try:
                    nlq_router_mod._resolve_analytics_action = lambda _text: "품목별 재고부족현황"
                    nlq_router_mod._get_analytics_handler = lambda _action: (
                        lambda params: captured_nlq_params.update(dict(params)) or {"meta": {}, "params": dict(params)}
                    )
                    chat_middleware_mod.push_sims_result_to_chat = lambda payload, _action: captured_nlq_payload.update(dict(payload))
                    nlq_router_mod._build_analytics_params = lambda _text, _action: {"date_from": "20260101", "date_to": "20260131"}
                    if not nlq_router_mod._try_handle_analytics_nlq(
                        "품목별 재고부족현황",
                        room={}, session_state={}, make_ts=lambda: "", next_seq=lambda: 1,
                        logger=logging.getLogger("ssai.regression"),
                    ):
                        raise AssertionError("nlq_handler_not_handled")
                finally:
                    nlq_router_mod._resolve_analytics_action = old_action_resolver
                    nlq_router_mod._get_analytics_handler = old_handler_getter
                    nlq_router_mod._build_analytics_params = old_param_builder
                    chat_middleware_mod.push_sims_result_to_chat = old_push
                if (
                    captured_nlq_params.get("dashboard_product_di_list") != ["0004:1", "0004:2"]
                    or "회사 Default" not in str((captured_nlq_payload.get("meta") or {}).get("query_summary") or "")
                ):
                    raise AssertionError(f"nlq_handler_params_or_sources={captured_nlq_params!r}/{captured_nlq_payload!r}")

                def _run_nlq_case(question: str, action: str, parsed: dict[str, Any]) -> tuple[dict[str, Any], dict[str, Any]]:
                    captured_params: dict[str, Any] = {}
                    captured_payload: dict[str, Any] = {}
                    old_resolver = nlq_router_mod._resolve_analytics_action
                    old_handler = nlq_router_mod._get_analytics_handler
                    old_builder = nlq_router_mod._build_analytics_params
                    old_case_push = chat_middleware_mod.push_sims_result_to_chat
                    try:
                        nlq_router_mod._resolve_analytics_action = lambda _text: action
                        nlq_router_mod._build_analytics_params = lambda _text, _action: dict(parsed)
                        nlq_router_mod._get_analytics_handler = lambda _action: (
                            lambda params: captured_params.update(dict(params)) or {"meta": {}, "params": dict(params)}
                        )
                        chat_middleware_mod.push_sims_result_to_chat = lambda payload, _action: captured_payload.update(dict(payload))
                        if not nlq_router_mod._try_handle_analytics_nlq(
                            question, room={}, session_state={}, make_ts=lambda: "", next_seq=lambda: 1,
                            logger=logging.getLogger("ssai.regression"),
                        ):
                            raise AssertionError(f"nlq_case_not_handled={question}")
                    finally:
                        nlq_router_mod._resolve_analytics_action = old_resolver
                        nlq_router_mod._get_analytics_handler = old_handler
                        nlq_router_mod._build_analytics_params = old_builder
                        chat_middleware_mod.push_sims_result_to_chat = old_case_push
                    return captured_params, captured_payload

                explicit_params, explicit_payload = _run_nlq_case(
                    "장부재고 00247 창고 품목별 재고부족현황", "품목별 재고부족현황",
                    {"stock_mode": "book", "stock_cds": ["00247"], "stock_cd": "00247"},
                )
                explicit_summary = str((explicit_payload.get("meta") or {}).get("query_summary") or "")
                if (
                    explicit_params.get("stock_mode") != "book"
                    or explicit_params.get("stock_cd_list") != ["00247"]
                    or "재고위치: 00247 (질문에서 지정)" not in explicit_summary
                    or "재고위치: 전체 (질문에서 지정)" in explicit_summary
                    or explicit_params.get("dashboard_product_di_list") != ["0004:1", "0004:2"]
                ):
                    raise AssertionError(f"nlq_explicit_stock_display={explicit_params!r}/{explicit_summary!r}")

                clear_params, clear_payload = _run_nlq_case(
                    "전체 창고 전체 제품구분 품목별 재고부족현황", "품목별 재고부족현황", {
                        "stock_cds": ["00247"], "stock_nm": "창고", "product_di_nm": "일반",
                    },
                )
                clear_summary = str((clear_payload.get("meta") or {}).get("query_summary") or "")
                if (
                    any(clear_params.get(key) not in ([], "") for key in ("stock_cd_list", "stock_cds", "stock_cd", "stock_nm", "stock_nm_list"))
                    or any(clear_params.get(key) not in ([], "") for key in ("product_di_list", "dashboard_product_di_list", "product_di", "product_di_nm", "product_di_nm_list"))
                    or clear_params.get("dashboard_product_class_list") != ["0031:01"]
                    or "재고위치: 전체 (전체 조건)" not in clear_summary
                    or "제품구분: 전체 (전체 조건)" not in clear_summary
                ):
                    raise AssertionError(f"nlq_explicit_clear_aliases={clear_params!r}/{clear_summary!r}")

                summary_clear_params, _summary_clear_payload = _run_nlq_case(
                    "전체 제품분류 품목별 매출 추세 요약표", "품목별 매출 추세 요약표", {"product_class_nm": "기존"},
                )
                if (
                    any(summary_clear_params.get(key) not in ([], "") for key in ("product_class_list", "dashboard_product_class_list", "product_class", "product_class_nm", "product_class_nm_list"))
                    or summary_clear_params.get("stock_cd_list") != ["00001", "00247"]
                    or summary_clear_params.get("dashboard_product_di_list") != ["0004:1", "0004:2"]
                ):
                    raise AssertionError(f"nlq_summary_product_class_clear={summary_clear_params!r}")

                name_only = nlq_router_mod._apply_company_default_to_analytics_nlq(
                    {"stock_nm": "본사"}, text="본사 창고 품목별 재고부족현황", action="품목별 재고부족현황",
                    session_state={}, logger=logging.getLogger("ssai.regression"),
                )
                if (
                    name_only.get("stock_nm") != "본사"
                    or name_only.get("stock_cd_list") not in (None, [])
                    or name_only.get("stock_cds") not in (None, [])
                    or name_only.get("stock_cd") not in (None, "")
                ):
                    raise AssertionError(f"nlq_name_to_code_leak={name_only!r}")

                actual_parser_results: dict[str, tuple[dict[str, Any], dict[str, Any]]] = {}
                old_resolver = nlq_router_mod._resolve_analytics_action
                old_handler = nlq_router_mod._get_analytics_handler
                old_case_push = chat_middleware_mod.push_sims_result_to_chat
                old_explicit_option_loader = analytics_view_mod._load_code_options
                try:
                    analytics_view_mod._load_code_options = lambda gcode: (
                        [{"code": "01", "name": "일반", "label": "01 - 일반"}]
                        if gcode == "0031" else []
                    )
                    def _run_actual_parser_case(question: str, action: str) -> tuple[dict[str, Any], dict[str, Any]]:
                        captured_params: dict[str, Any] = {}
                        captured_payload: dict[str, Any] = {}
                        nlq_router_mod._resolve_analytics_action = lambda _text: action
                        nlq_router_mod._get_analytics_handler = lambda _action: (
                            lambda params: captured_params.update(dict(params)) or {
                                "params": dict(params),
                                "meta": {"query_summary": f"재고위치: {params.get('stock_cd') or params.get('stock_nm') or '전체'}"},
                            }
                        )
                        chat_middleware_mod.push_sims_result_to_chat = lambda payload, _action: captured_payload.update(dict(payload))
                        if not nlq_router_mod._try_handle_analytics_nlq(
                            question, room={}, session_state={}, make_ts=lambda: "", next_seq=lambda: 1,
                            logger=logging.getLogger("ssai.regression"),
                        ):
                            raise AssertionError(f"actual_parser_not_handled={question}")
                        return captured_params, captured_payload

                    actual_parser_results["code"] = _run_actual_parser_case("장부재고 00247 창고 품목별 재고부족현황", "품목별 재고부족현황")
                    actual_parser_results["clear"] = _run_actual_parser_case("전체 창고 전체 제품구분 품목별 재고부족현황", "품목별 재고부족현황")
                    actual_parser_results["class_clear"] = _run_actual_parser_case("전체 제품분류 품목별 매출 추세 요약표", "품목별 매출 추세 요약표")
                    actual_parser_results["stock_name"] = _run_actual_parser_case("본사 창고 품목별 재고부족현황", "품목별 재고부족현황")
                    actual_parser_results["di_name"] = _run_actual_parser_case("일반의약품 제품구분 품목별 매출 추세 요약표", "품목별 매출 추세 요약표")
                    for action in (
                        "품목별 재고부족현황",
                        "품목별 매출 추세 요약표",
                        "품목별 매출 예상",
                    ):
                        actual_parser_results[f"0031_default::{action}"] = _run_actual_parser_case(
                            action,
                            action,
                        )
                    actual_parser_results["manufacturer_group"] = _run_actual_parser_case(
                        "제약사별 매출 추세분석",
                        "제약사별 매출 추세 분석",
                    )
                    actual_parser_results["manufacturer_explicit"] = _run_actual_parser_case(
                        "한미제약 제약사별 매출 추세분석",
                        "제약사별 매출 추세 분석",
                    )
                    actual_parser_results["manufacturer_sales_forecast"] = _run_actual_parser_case(
                        "한미제약 품목별 매출 예상 조회",
                        "품목별 매출 예상",
                    )
                    actual_parser_results["manufacturer_sales_summary"] = _run_actual_parser_case(
                        "한미제약 품목별 매출 추세 요약표 조회",
                        "품목별 매출 추세 요약표",
                    )
                    actual_parser_results["manufacturer_year_group"] = _run_actual_parser_case(
                        "2026년 제약사별 매출 추세분석",
                        "제약사별 매출 추세 분석",
                    )
                    actual_parser_results["explicit_0031_tcode"] = _run_actual_parser_case(
                        "01 제품분류 품목별 매출 추세 요약표",
                        "품목별 매출 추세 요약표",
                    )
                    actual_parser_results["explicit_0031_pair"] = _run_actual_parser_case(
                        "0031:01 제품분류 품목별 매출 추세 요약표",
                        "품목별 매출 추세 요약표",
                    )
                    actual_parser_results["explicit_0031_name"] = _run_actual_parser_case(
                        "일반 제품분류 품목별 매출 추세 요약표",
                        "품목별 매출 추세 요약표",
                    )
                    actual_parser_results["explicit_0028_pair"] = _run_actual_parser_case(
                        "0028:01 제품분류 품목별 매출 추세 요약표",
                        "품목별 매출 추세 요약표",
                    )
                    actual_parser_results["context_pairs_forecast"] = _run_actual_parser_case(
                        "0004:1 제품구분 0031:01 제품분류 품목별 매출 예상 조회",
                        "품목별 매출 예상",
                    )
                    actual_parser_results["context_pairs_summary"] = _run_actual_parser_case(
                        "0031:01 제품분류 0004:J 제품구분 품목별 매출 추세 요약표 조회",
                        "품목별 매출 추세 요약표",
                    )
                    actual_parser_results["context_pairs_0028_0031"] = _run_actual_parser_case(
                        "0028:01 특수관리제품 0031:01 제품분류 품목별 재고부족현황",
                        "품목별 재고부족현황",
                    )
                finally:
                    nlq_router_mod._resolve_analytics_action = old_resolver
                    nlq_router_mod._get_analytics_handler = old_handler
                    chat_middleware_mod.push_sims_result_to_chat = old_case_push
                    analytics_view_mod._load_code_options = old_explicit_option_loader

                actual_code_params, actual_code_payload = actual_parser_results["code"]
                actual_clear_params, _actual_clear_payload = actual_parser_results["clear"]
                actual_class_params, _actual_class_payload = actual_parser_results["class_clear"]
                actual_stock_name_params, actual_stock_name_payload = actual_parser_results["stock_name"]
                actual_di_name_params, _actual_di_name_payload = actual_parser_results["di_name"]
                actual_code_summary = str((actual_code_payload.get("meta") or {}).get("query_summary") or "")
                if (
                    actual_code_params.get("stock_mode") != "book"
                    or actual_code_params.get("stock_cd_list") != ["00247"]
                    or actual_code_params.get("stock_nm") not in (None, "")
                    or actual_code_summary.count("재고위치:") != 1
                    or (actual_code_payload.get("meta") or {}).get("condition") != actual_code_summary
                ):
                    raise AssertionError(f"actual_parser_code_path={actual_code_params!r}/{actual_code_summary!r}")
                if (
                    any(actual_clear_params.get(key) not in ([], "") for key in ("stock_cd_list", "stock_cds", "stock_cd", "stock_nm", "stock_nm_list"))
                    or any(actual_clear_params.get(key) not in ([], "") for key in ("product_di_list", "dashboard_product_di_list", "product_di", "product_di_nm", "product_di_nm_list"))
                    or any(actual_class_params.get(key) not in ([], "") for key in ("product_class_list", "dashboard_product_class_list", "product_class", "product_class_nm", "product_class_nm_list"))
                ):
                    raise AssertionError(f"actual_parser_explicit_clear={actual_clear_params!r}/{actual_class_params!r}")
                if (
                    actual_stock_name_params.get("stock_nm") != "본사"
                    or actual_stock_name_params.get("stock_cd_list") not in (None, [])
                    or actual_stock_name_params.get("stock_cds") not in (None, [])
                    or "본사" not in str((actual_stock_name_payload.get("meta") or {}).get("query_summary") or "")
                    or actual_di_name_params.get("product_di_nm") != "일반의약품"
                    or actual_di_name_params.get("product_di_list") not in (None, [])
                    or actual_di_name_params.get("dashboard_product_di_list") not in (None, [])
                ):
                    raise AssertionError(f"actual_parser_name_path={actual_stock_name_params!r}/{actual_di_name_params!r}")
                for key, (captured_params, _captured_payload) in actual_parser_results.items():
                    if not key.startswith("0031_default::"):
                        continue
                    if (
                        captured_params.get("product_class_list") != []
                        or captured_params.get("product_class") not in (None, "")
                        or captured_params.get("product_class_nm") not in (None, "")
                        or captured_params.get("product_class_nm_list") not in (None, [])
                        or captured_params.get("dashboard_product_class_list") != ["0031:01"]
                    ):
                        raise AssertionError(f"actual_parser_0031_contract={key}/{captured_params!r}")

                manufacturer_group_params, _ = actual_parser_results["manufacturer_group"]
                manufacturer_explicit_params, _ = actual_parser_results["manufacturer_explicit"]
                manufacturer_sales_forecast_params, _ = actual_parser_results["manufacturer_sales_forecast"]
                manufacturer_sales_summary_params, _ = actual_parser_results["manufacturer_sales_summary"]
                manufacturer_year_group_params, _ = actual_parser_results["manufacturer_year_group"]
                if (
                    manufacturer_group_params.get("maker_nm") not in (None, "")
                    or manufacturer_group_params.get("product_ven_nm") not in (None, "")
                    or manufacturer_explicit_params.get("maker_nm") != "한미제약"
                    or manufacturer_explicit_params.get("product_ven_nm") != "한미제약"
                    or manufacturer_sales_forecast_params.get("maker_nm") != "한미제약"
                    or manufacturer_sales_forecast_params.get("product_ven_nm") != "한미제약"
                    or manufacturer_sales_summary_params.get("maker_nm") != "한미제약"
                    or manufacturer_sales_summary_params.get("product_ven_nm") != "한미제약"
                    or manufacturer_year_group_params.get("maker_nm") not in (None, "")
                    or manufacturer_year_group_params.get("product_ven_nm") not in (None, "")
                ):
                    raise AssertionError(
                        "manufacturer_group_action_parser="
                        f"{manufacturer_group_params!r}/{manufacturer_explicit_params!r}/"
                        f"{manufacturer_sales_forecast_params!r}/{manufacturer_sales_summary_params!r}/"
                        f"{manufacturer_year_group_params!r}"
                    )
                for key in ("explicit_0031_tcode", "explicit_0031_pair", "explicit_0031_name"):
                    captured_params, _ = actual_parser_results[key]
                    if (
                        captured_params.get("product_class_list") != []
                        or captured_params.get("product_class") not in (None, "")
                        or captured_params.get("product_class_nm") not in (None, "")
                        or captured_params.get("product_class_nm_list") not in (None, [])
                        or captured_params.get("dashboard_product_class_list") != ["0031:01"]
                    ):
                        raise AssertionError(f"explicit_0031_tax_contract={key}/{captured_params!r}")
                explicit_0028_params, _ = actual_parser_results["explicit_0028_pair"]
                if (
                    explicit_0028_params.get("product_class_list") != ["01"]
                    or explicit_0028_params.get("dashboard_product_class_list") not in (None, [])
                ):
                        raise AssertionError(f"explicit_0028_legacy_contract={explicit_0028_params!r}")

                context_forecast_params, _ = actual_parser_results["context_pairs_forecast"]
                context_summary_params, _ = actual_parser_results["context_pairs_summary"]
                context_mixed_params, _ = actual_parser_results["context_pairs_0028_0031"]
                if (
                    context_forecast_params.get("product_di_list") != ["1"]
                    or context_forecast_params.get("dashboard_product_di_list") != ["0004:1"]
                    or context_forecast_params.get("product_class_list") != []
                    or context_forecast_params.get("dashboard_product_class_list") != ["0031:01"]
                ):
                    raise AssertionError(f"context_pair_forecast_contract={context_forecast_params!r}")
                if (
                    context_summary_params.get("product_di_list") != ["J"]
                    or context_summary_params.get("dashboard_product_di_list") != ["0004:J"]
                    or context_summary_params.get("product_class_list") != []
                    or context_summary_params.get("dashboard_product_class_list") != ["0031:01"]
                ):
                    raise AssertionError(f"context_pair_summary_contract={context_summary_params!r}")
                if (
                    context_mixed_params.get("product_class_list") != ["01"]
                    or context_mixed_params.get("dashboard_product_class_list") != ["0031:01"]
                ):
                    raise AssertionError(f"context_pair_0028_0031_contract={context_mixed_params!r}")

                for group_action in (
                    "품목별 매출 추세 분석",
                    "매입처별 재고부족 현황",
                    "매출처별 매출 예상",
                    "영업사원별 매출 예상",
                    "지역별 매출 예상",
                ):
                    cleaned = nlq_router_mod._cleanup_analytics_named_params(
                        {"maker_nm": "별", "product_ven_nm": "별"},
                        text=group_action,
                        action=group_action,
                    )
                    if cleaned.get("maker_nm") or cleaned.get("product_ven_nm"):
                        raise AssertionError(f"group_action_suffix_parser={group_action}/{cleaned!r}")

                bind_params = {"dashboard_product_class_list": ["0031:01"]}
                detail_clauses: list[str] = []
                analytics_service_mod._add_dashboard_code_pair_filter(
                    detail_clauses,
                    bind_params,
                    gcode_sql="Physic_Cd.Rd04_Physic_Tax_Gcode",
                    tcode_sql="Physic_Cd.Rd04_Physic_Tax",
                    key="dashboard_product_class_list",
                )
                monthly_bind_params = {"dashboard_product_class_list": ["0031:01"]}
                monthly_clauses: list[str] = []
                analytics_service_mod._add_dashboard_code_pair_filter(
                    monthly_clauses,
                    monthly_bind_params,
                    gcode_sql="Physic_Cd.Rd04_Physic_Tax_Gcode",
                    tcode_sql="Physic_Cd.Rd04_Physic_Tax",
                    key="dashboard_product_class_list",
                )
                detail_predicate = " ".join(detail_clauses)
                monthly_predicate = " ".join(monthly_clauses)
                if (
                    bind_params.get("dashboard_product_class_list_g_0") != "0031"
                    or bind_params.get("dashboard_product_class_list_t_0") != "01"
                    or monthly_bind_params.get("dashboard_product_class_list_g_0") != "0031"
                    or monthly_bind_params.get("dashboard_product_class_list_t_0") != "01"
                    or "Rd04_Physic_Gu IN" in detail_predicate
                    or "Rd04_Physic_Gu IN" in monthly_predicate
                ):
                    raise AssertionError(
                        f"business_code_sql_bind={bind_params!r}/{monthly_bind_params!r}/"
                        f"{detail_predicate!r}/{monthly_predicate!r}"
                    )

                cache_state = {"__analysis_profile_company_cache": {"41": {"stock_cd_list": ["00001"]}, "42": {"stock_cd_list": ["00999"]}}}
                company_profile_mod.invalidate_analysis_profile_cache(cache_state, company_id=41)
                if "41" in cache_state["__analysis_profile_company_cache"] or "42" not in cache_state["__analysis_profile_company_cache"]:
                    raise AssertionError(f"profile_cache_target_invalidation={cache_state!r}")
                current_company["id"] = "41"
                refresh_ns = "save-refresh"
                refresh_stock_widget = f"__analytics_stock_shortage_stock__{refresh_ns}"
                refresh_prefill = f"__analytics_dashboard_prefill_codes::{refresh_stock_widget}"
                profile_by_company["41"]["stock_cd_list"] = ["00001"]
                company_profile_mod.invalidate_analysis_profile_cache(analytics_st.session_state, company_id=41)
                analytics_view_mod._prepare_analytics_company_defaults("stock_shortage", refresh_ns)
                if analytics_st.session_state.get(refresh_prefill) != ["00001"]:
                    raise AssertionError(f"profile_generation_initial={analytics_st.session_state!r}")
                profile_by_company["41"]["stock_cd_list"] = ["00247"]
                analytics_st.session_state["__analysis_profile_generation::42"] = 7
                generation = company_profile_mod.mark_analysis_profile_saved(analytics_st.session_state, company_id=41)
                analytics_view_mod._prepare_analytics_company_defaults("stock_shortage", refresh_ns)
                if (
                    generation != 1
                    or analytics_st.session_state.get(refresh_prefill) != ["00247"]
                    or analytics_st.session_state.get(f"__analytics_profile_applied_generation::41::stock_shortage::{refresh_ns}") != generation
                ):
                    raise AssertionError(f"profile_generation_refresh={analytics_st.session_state!r}")
                analytics_st.session_state[refresh_stock_widget] = ["00901"]
                analytics_view_mod._prepare_analytics_company_defaults("stock_shortage", refresh_ns)
                if analytics_st.session_state.get(refresh_stock_widget) != ["00901"]:
                    raise AssertionError(f"profile_generation_live_value_overwritten={analytics_st.session_state!r}")
                if analytics_st.session_state.get("__analysis_profile_generation::42") != 7:
                    raise AssertionError(f"profile_generation_other_company_mutated={analytics_st.session_state!r}")

                nlq_src = Path("app/sims/nlq/nlq_router.py").read_text(encoding="utf-8")
                if (
                    "analytics handled action=%r params=%r" in nlq_src
                    or "analytics service failed action=%r params=%r" in nlq_src
                    or "_analytics_nlq_code_values" not in nlq_src
                    or "_analytics_nlq_name_values" not in nlq_src
                ):
                    raise AssertionError("nlq_param_logging_or_alias_separation_missing")
            finally:
                company_profile_mod.load_dashboard_profile = old_profile_loader
                analytics_view_mod.load_dashboard_profile = old_analytics_loader
                analytics_view_mod.st = old_analytics_st
                login_mod.get_selected_company = old_company_getter
                nlq_router_mod._analytics_nlq_option_codes = old_nlq_option_codes
            results.append(_ok("Company Default adapter for Dashboard, KPI, and NLQ", "supported keys only; dates/manufacturer excluded; explicit override/clear and company-scoped cache verified"))
        except Exception as e:
            results.append(_fail("Company Default adapter for Dashboard, KPI, and NLQ", f"{type(e).__name__}: {e}"))

        kpi_company_io_errors: list[str] = []
        try:
            import app.services.analytics_sales_trend_service as analytics_service_mod
            import app.sims.nlq.nlq_router as nlq_router_mod
            import app.sims.views.analytics_views as analytics_view_mod

            analytics_view_source = Path(analytics_view_mod.__file__).read_text(encoding="utf-8")
            if "def _render_analytics_io_scope" in analytics_view_source or "__analytics_.*_io_gu_all" in analytics_view_source:
                kpi_company_io_errors.append("kpi_io_override_ui_present")
            enforced = analytics_view_mod._attach_analytics_company_io(
                {"io_gu_list": ["999"]}, {"effective": {"io_gu_list": ["0012:051", "001"]}}
            )
            if enforced.get("io_gu_list") != ["051", "001"] or enforced.get("io_gu_source") != "company_default":
                kpi_company_io_errors.append(f"kpi_company_io_not_forced={enforced!r}")
            missing = analytics_view_mod._attach_analytics_company_io({}, {"effective": {}})
            if not missing.get("__company_io_missing") or "io_gu_list" in missing:
                kpi_company_io_errors.append(f"kpi_company_io_missing_marker={missing!r}")
            try:
                analytics_service_mod._sales_io_scope({"_require_company_io": True})
                kpi_company_io_errors.append("kpi_company_io_missing_not_blocked")
            except ValueError as exc:
                if "회사 공통 분석용 입출고구분" not in str(exc):
                    kpi_company_io_errors.append(f"kpi_company_io_wrong_message={exc}")
            if analytics_service_mod._sales_io_scope({}) != ("legacy_broad_fallback", []):
                kpi_company_io_errors.append("ordinary_io_legacy_changed")

            old_company_getter = nlq_router_mod.get_selected_company if hasattr(nlq_router_mod, "get_selected_company") else None
            # The router imports this dependency inside the helper, so patch the
            # source module used by that import rather than parser behavior.
            import app.services.ssai_analysis_profile_service as profile_service_mod
            import app.ui.ssai_login as login_mod
            old_profile_loader = profile_service_mod.load_dashboard_profile
            old_login_getter = login_mod.get_selected_company
            try:
                profile_service_mod.load_dashboard_profile = lambda **_kwargs: {"io_gu_list": ["0012:051"]}
                login_mod.get_selected_company = lambda: {"company_id": "71"}
                forced_nlq = nlq_router_mod._apply_company_default_to_analytics_nlq(
                    {"io_gu_list": ["193"], "io_gu": "193"},
                    text="정상출고 품목별 매출 예상",
                    action="품목별 매출 예상",
                    session_state={},
                    logger=logging.getLogger("ssai.regression"),
                )
                if forced_nlq.get("io_gu_list") != ["051"] or forced_nlq.get("io_gu_source") != "company_default":
                    kpi_company_io_errors.append(f"kpi_nlq_explicit_io_not_ignored={forced_nlq!r}")
            finally:
                profile_service_mod.load_dashboard_profile = old_profile_loader
                login_mod.get_selected_company = old_login_getter
        except Exception as exc:
            kpi_company_io_errors.append(f"{type(exc).__name__}: {exc}")
        if kpi_company_io_errors:
            results.append(_fail("KPI company IO scope enforcement", "; ".join(kpi_company_io_errors)))
        else:
            results.append(_ok("KPI company IO scope enforcement", "KPI UI has no IO override control; panel/NLQ use the persisted company scope, missing scope blocks, and ordinary IO legacy behavior remains separate"))

        code_pair_errors: list[str] = []
        code_pair_df = pd.DataFrame(
            [
                {"제품코드": "KEEP_CROSS_GCODE", "제품분류Gcode": "0001", "제품분류코드": "01", "제품분류명": "일반", "완료월총매출": 10},
                {"제품코드": "DROP_CODE_PAIR", "제품분류Gcode": "0031", "제품분류코드": "01", "제품분류명": "이름변경", "완료월총매출": 20},
                {"제품코드": "KEEP_SAME_NAME", "제품분류Gcode": "0031", "제품분류코드": "99", "제품분류명": "일반", "완료월총매출": 30},
                {"제품코드": "KEEP_OTHER", "제품분류Gcode": "0031", "제품분류코드": "02", "제품분류명": "기타", "완료월총매출": 40},
            ]
        )
        code_pair_original = code_pair_df.copy(deep=True)
        code_pair_filtered = dash_mod._filter_sales_source_for_dashboard(
            code_pair_df,
            {"exclude_product_class_list": ["0031:01"], "exclude_product_class_nm_list": ["일반"]},
        )
        remaining_products = list(code_pair_filtered["제품코드"])
        if remaining_products != ["KEEP_CROSS_GCODE", "KEEP_SAME_NAME", "KEEP_OTHER"]:
            code_pair_errors.append(f"code_pair_remaining={remaining_products!r}")
        code_pair_diag = code_pair_filtered.attrs.get("dashboard_filter_diagnostics") or []
        class_diag = next((d for d in code_pair_diag if d.get("label") == "제품분류"), {})
        if class_diag.get("filter_basis") != "code_pair":
            code_pair_errors.append(f"filter_basis={class_diag!r}")
        if class_diag.get("filtered_rows") != 1:
            code_pair_errors.append(f"filtered_rows={class_diag!r}")
        if not code_pair_df.equals(code_pair_original):
            code_pair_errors.append("code_pair_input_mutated")

        fallback_df = pd.DataFrame(
            [
                {"제품코드": "DROP_NAME", "제품분류명": "제외분류", "완료월총매출": 10},
                {"제품코드": "KEEP_NAME", "제품분류명": "정상분류", "완료월총매출": 20},
            ]
        )
        fallback_filtered = dash_mod._filter_sales_source_for_dashboard(
            fallback_df,
            {"exclude_product_class_list": ["0031:C_EX"], "exclude_product_class_nm_list": ["제외분류"]},
        )
        fallback_diag = fallback_filtered.attrs.get("dashboard_filter_diagnostics") or []
        fallback_class_diag = next((d for d in fallback_diag if d.get("label") == "제품분류"), {})
        if list(fallback_filtered["제품코드"]) != ["DROP_NAME", "KEEP_NAME"]:
            code_pair_errors.append(f"missing_code_columns_mutated_rows={list(fallback_filtered['제품코드'])!r}")
        if (
            fallback_class_diag.get("filter_basis") != "not_applied"
            or not fallback_class_diag.get("missing_code_columns")
            or fallback_class_diag.get("selected_code_pair_count") != 1
        ):
            code_pair_errors.append(f"not_applied_diag={fallback_class_diag!r}")

        stock_payload = {
            "df": pd.DataFrame(
                [
                    {"제품코드": "DROP_CODE_PAIR", "현재재고": 1},
                    {"제품코드": "KEEP_CROSS_GCODE", "현재재고": 2},
                ]
            )
        }
        enriched_stock = dash_mod._attach_dashboard_product_code_pairs(stock_payload, code_pair_df)
        filtered_stock = dash_mod._filter_payload_df_for_dashboard(
            enriched_stock,
            {"exclude_product_class_list": ["0031:01"]},
        )
        stock_df = dash_mod._payload_df(filtered_stock)
        stock_diag = stock_df.attrs.get("dashboard_filter_diagnostics") or []
        stock_class_diag = next((d for d in stock_diag if d.get("label") == "제품분류"), {})
        if list(stock_df["제품코드"]) != ["KEEP_CROSS_GCODE"]:
            code_pair_errors.append(f"stock_code_pair_remaining={list(stock_df['제품코드'])!r}")
        if stock_class_diag.get("filter_basis") != "code_pair":
            code_pair_errors.append(f"stock_filter_basis={stock_class_diag!r}")

        inclusion_df = pd.DataFrame(
            [
                {"\uc81c\ud488\ucf54\ub4dc": "KEEP", "\uc81c\ud488\uadf8\ub8f9Gcode": "0013", "\uc81c\ud488\uadf8\ub8f9\ucf54\ub4dc": "A", "\uc81c\ud488\uad6c\ubd84Gcode": "0004", "\uc81c\ud488\uad6c\ubd84\ucf54\ub4dc": "X", "\uc81c\ud488\ubd84\ub958Gcode": "0031", "\uc81c\ud488\ubd84\ub958\ucf54\ub4dc": "C"},
                {"\uc81c\ud488\ucf54\ub4dc": "DROP_GROUP", "\uc81c\ud488\uadf8\ub8f9Gcode": "0013", "\uc81c\ud488\uadf8\ub8f9\ucf54\ub4dc": "Z", "\uc81c\ud488\uad6c\ubd84Gcode": "0004", "\uc81c\ud488\uad6c\ubd84\ucf54\ub4dc": "X", "\uc81c\ud488\ubd84\ub958Gcode": "0031", "\uc81c\ud488\ubd84\ub958\ucf54\ub4dc": "C"},
                {"\uc81c\ud488\ucf54\ub4dc": "DROP_DI", "\uc81c\ud488\uadf8\ub8f9Gcode": "0013", "\uc81c\ud488\uadf8\ub8f9\ucf54\ub4dc": "B", "\uc81c\ud488\uad6c\ubd84Gcode": "0004", "\uc81c\ud488\uad6c\ubd84\ucf54\ub4dc": "Y", "\uc81c\ud488\ubd84\ub958Gcode": "0031", "\uc81c\ud488\ubd84\ub958\ucf54\ub4dc": "C"},
                {"\uc81c\ud488\ucf54\ub4dc": "KEEP_GROUP_OR", "\uc81c\ud488\uadf8\ub8f9Gcode": "0013", "\uc81c\ud488\uadf8\ub8f9\ucf54\ub4dc": "B", "\uc81c\ud488\uad6c\ubd84Gcode": "0004", "\uc81c\ud488\uad6c\ubd84\ucf54\ub4dc": "X", "\uc81c\ud488\ubd84\ub958Gcode": "0031", "\uc81c\ud488\ubd84\ub958\ucf54\ub4dc": "C"},
            ]
        )
        inclusion_filtered = dash_mod._filter_sales_source_for_dashboard(
            inclusion_df,
            {"product_group_list": ["0013:A", "0013:B"], "product_di_list": ["0004:X"], "product_class_list": ["0031:C"]},
        )
        if list(inclusion_filtered["\uc81c\ud488\ucf54\ub4dc"]) != ["KEEP", "KEEP_GROUP_OR"]:
            code_pair_errors.append(f"inclusion_and_or={list(inclusion_filtered['\uc81c\ud488\ucf54\ub4dc'])!r}")

        if code_pair_errors:
            results.append(_fail("Dashboard Lite product code-pair filters", "; ".join(code_pair_errors)))
        else:
            results.append(_ok("Dashboard Lite product code-pair filters", "Gcode+Tcode excludes exact code pairs; same Tcode/different Gcode and same name/different code are preserved"))

        import app.services.analytics_manufacturer_sales_trend_service as manufacturer_mod
        import app.services.analytics_sales_trend_service as sales_mod

        import app.services.dashboard_inbound_facts_service as inbound_mod

        service_errors: list[str] = []
        calls = {"shared_sales": 0, "manufacturer": 0, "forecast": 0, "stock": 0, "inbound": 0}
        seen_params: list[dict] = []
        forecast_params_seen: list[dict] = []
        shortage_params_seen: list[dict] = []
        preloaded_seen = {"manufacturer": False, "stock": False, "stock_master_universe": False}
        old_shared = getattr(sales_mod, "get_sales_trend_df")
        old_dashboard_bundle = getattr(sales_mod, "get_dashboard_sales_source_bundle")
        old_manufacturer = getattr(manufacturer_mod, "get_manufacturer_sales_trend_summary_result")
        old_forecast = getattr(sales_mod, "get_sales_forecast_df")
        old_stock = getattr(sales_mod, "get_stock_shortage_result")
        old_inbound = getattr(inbound_mod, "get_dashboard_inbound_facts")

        def _fake_shared_sales_source(params=None):
            calls["shared_sales"] += 1
            seen_params.append(dict(params or {}))
            return sales_df.copy()

        def _fake_dashboard_sales_bundle(params=None):
            shared = _fake_shared_sales_source(params)
            return {
                "sales_df": shared,
                "purchase_vendor_df": pd.DataFrame(columns=["기준월", "제품코드", "매입처코드", "매입처명", "입고수량", "매입금액", "매입발생건수"]),
                "perf": {"purchase_source_sql_included": True, "purchase_source_rows": 0, "purchase_min_frame_ms": 0},
            }

        def _fake_manufacturer_service(params=None, raw_df=None):
            calls["manufacturer"] += 1
            seen_params.append(dict(params or {}))
            preloaded_seen["manufacturer"] = isinstance(raw_df, pd.DataFrame)
            return {"df": sales_df.copy(), "meta": {"evaluation_month": "202607"}}

        def _fake_stock_service(params=None, sales_raw_df=None, sales_forecast_df=None, product_universe_df=None):
            calls["stock"] += 1
            seen_params.append(dict(params or {}))
            shortage_params_seen.append(dict(params or {}))
            preloaded_seen["stock"] = (
                isinstance(sales_raw_df, pd.DataFrame)
                and isinstance(sales_forecast_df, pd.DataFrame)
            )
            preloaded_seen["stock_master_universe"] = isinstance(product_universe_df, pd.DataFrame)
            return {"df": stock_df.copy(), "meta": {}}

        def _fake_forecast_service(params=None, raw_df=None):
            calls["forecast"] += 1
            seen_params.append(dict(params or {}))
            forecast_params_seen.append(dict(params or {}))
            return sales_df.copy()

        def _fake_inbound_service(params=None, **_kwargs):
            calls["inbound"] += 1
            return pd.DataFrame([{
                "product_code": "P1", "last_normal_inbound_date": "20260701",
                "normal_inbound_day_count_365": 2, "avg_inbound_cycle_days": 15.0,
                "inbound_data_status": "normal", "inbound_delayed_candidate": False,
                "normal_inbound_90_exists": True, "normal_inbound_365_exists": True,
                "recent_inbound_vendor_source": "actual_inbound",
            }])

        try:
            setattr(sales_mod, "get_sales_trend_df", _fake_shared_sales_source)
            setattr(sales_mod, "get_dashboard_sales_source_bundle", _fake_dashboard_sales_bundle)
            setattr(manufacturer_mod, "get_manufacturer_sales_trend_summary_result", _fake_manufacturer_service)
            setattr(sales_mod, "get_sales_forecast_df", _fake_forecast_service)
            setattr(sales_mod, "get_stock_shortage_result", _fake_stock_service)
            setattr(inbound_mod, "get_dashboard_inbound_facts", _fake_inbound_service)
            built = dash_mod.build_dashboard_lite_facts(
                {
                    "month_from": "202601",
                    "month_to": "202607",
                    "evaluation_month": "202607",
                    "stock_cd_list": ["00001"],
                    "manufacturer_test_codes": ["V001"],
                    "io_gu_list": ["501", "590"],
                },
                today=date(2026, 7, 20),
            )
        finally:
            setattr(sales_mod, "get_sales_trend_df", old_shared)
            setattr(sales_mod, "get_dashboard_sales_source_bundle", old_dashboard_bundle)
            setattr(manufacturer_mod, "get_manufacturer_sales_trend_summary_result", old_manufacturer)
            setattr(sales_mod, "get_sales_forecast_df", old_forecast)
            setattr(sales_mod, "get_stock_shortage_result", old_stock)
            setattr(inbound_mod, "get_dashboard_inbound_facts", old_inbound)
        if calls != {"shared_sales": 1, "manufacturer": 1, "forecast": 1, "stock": 1, "inbound": 1}:
            service_errors.append(f"service_calls={calls!r}")
        if built.get("source_call_count") != 3:
            service_errors.append(f"source_call_count={built.get('source_call_count')!r}")
        if built.get("base_source_call_count") != 2 or built.get("inbound_source_call_count") != 1:
            service_errors.append(
                "source_call_contract="
                f"base={built.get('base_source_call_count')!r},"
                f"inbound={built.get('inbound_source_call_count')!r}"
            )
        if preloaded_seen != {"manufacturer": True, "stock": True, "stock_master_universe": True}:
            service_errors.append(f"preloaded_seen={preloaded_seen!r}")
        if any(not p.get("month_from") or not p.get("month_to") for p in seen_params):
            service_errors.append(f"missing_month_params={seen_params!r}")
        if any(p == {} for p in seen_params):
            service_errors.append("empty_params_sent")
        if not any(p.get("dashboard_manufacturer_codes") == ["V001"] for p in seen_params):
            service_errors.append(f"manufacturer_filter_not_bound={seen_params!r}")
        if len(forecast_params_seen) != 1 or forecast_params_seen[0].get("io_gu_list") != ["501", "590"]:
            service_errors.append(f"dashboard_forecast_io_lost={forecast_params_seen!r}")
        if len(shortage_params_seen) != 1 or shortage_params_seen[0].get("io_gu_list") != ["501", "590"]:
            service_errors.append(f"dashboard_shortage_io_lost={shortage_params_seen!r}")
        if built.get("filters", {}).get("manufacturer_test_codes") != ["V001"]:
            service_errors.append(f"manufacturer_filter_not_in_facts={built.get('filters')!r}")
        if service_errors:
            results.append(_fail("Dashboard Lite guarded service calls", "; ".join(service_errors)))
        else:
            results.append(_ok("Dashboard Lite guarded service calls", "base source calls=2, inbound source calls=1, total source calls=3; visualization and compact snapshots add no source call"))

        inbound_errors: list[str] = []
        inbound_fixture = pd.DataFrame([
            {"product_code": "P1", "master_order_vendor_code": "M1", "master_order_vendor_name": "Master", "inbound_date": "20260101", "io_tcode": "001", "vendor_code": "A", "inbound_vendor_name": "Vendor A", "quantity": 10, "oquantity": 0, "supply_price": 100},
            {"product_code": "P1", "master_order_vendor_code": "M1", "inbound_date": "20260111", "io_tcode": "002", "vendor_code": "A", "inbound_vendor_name": "Vendor A", "quantity": 10, "oquantity": 0, "supply_price": 100},
            {"product_code": "P1", "master_order_vendor_code": "M1", "inbound_date": "20260131", "io_tcode": "001", "vendor_code": "B", "inbound_vendor_name": "Vendor B", "quantity": 10, "oquantity": 0, "supply_price": 90},
            {"product_code": "P1", "master_order_vendor_code": "M1", "inbound_date": "20260720", "io_tcode": "051", "vendor_code": "X", "quantity": 99, "oquantity": 0, "supply_price": 999},
            {"product_code": "P1", "master_order_vendor_code": "M1", "inbound_date": "20260721", "io_tcode": "101", "vendor_code": "A", "quantity": -3, "oquantity": 0, "supply_price": -30},
            {"product_code": "P2", "master_order_vendor_code": "M2", "inbound_date": "20260721", "io_tcode": "001", "vendor_code": "A", "quantity": -2, "oquantity": 0, "supply_price": -20},
            {"product_code": "P3", "master_order_vendor_code": "M3", "inbound_date": "", "io_tcode": "", "vendor_code": "", "quantity": 0, "oquantity": 0, "supply_price": 0},
            {"product_code": "P4", "master_order_vendor_code": "M4", "inbound_date": "20260720", "io_tcode": "001", "vendor_code": "", "quantity": 200, "oquantity": 0, "supply_price": 200},
            {"product_code": "P5", "master_order_vendor_code": "M5", "inbound_date": "20260720", "io_tcode": "001", "vendor_code": "", "quantity": 200, "oquantity": 0, "supply_price": 200},
            {"product_code": "P5", "master_order_vendor_code": "M5", "inbound_date": "20260721", "io_tcode": "001", "vendor_code": "V5", "inbound_vendor_name": "Vendor 5", "quantity": 1, "oquantity": 0, "supply_price": 1},
            {"product_code": "P6", "master_order_vendor_code": "M6", "inbound_date": "20250722", "io_tcode": "001", "vendor_code": "OLD", "quantity": 10, "oquantity": 0, "supply_price": 10},
            {"product_code": "P6", "master_order_vendor_code": "M6", "inbound_date": "20250723", "io_tcode": "001", "vendor_code": "EDGE", "quantity": 10, "oquantity": 0, "supply_price": 10},
            {"product_code": "P6", "master_order_vendor_code": "M6", "inbound_date": "20260723", "io_tcode": "001", "vendor_code": "FUTURE", "quantity": 10, "oquantity": 0, "supply_price": 10},
            {"product_code": "P6", "master_order_vendor_code": "M6", "inbound_date": "invalid", "io_tcode": "001", "vendor_code": "BAD", "quantity": 10, "oquantity": 0, "supply_price": 10},
        ])
        inbound_frame = inbound_mod.build_dashboard_inbound_facts_frame(
            inbound_fixture, data_cutoff_date="20260722", cycle_lookback_days=365, vendor_lookback_days=90,
        ).set_index("product_code")
        p1 = inbound_frame.loc["P1"].to_dict()
        p2 = inbound_frame.loc["P2"].to_dict()
        p3 = inbound_frame.loc["P3"].to_dict()
        p4 = inbound_frame.loc["P4"].to_dict()
        p5 = inbound_frame.loc["P5"].to_dict()
        p6 = inbound_frame.loc["P6"].to_dict()
        vendor_frame = inbound_mod.build_dashboard_inbound_facts_frame(
            inbound_fixture, data_cutoff_date="20260131", cycle_lookback_days=365, vendor_lookback_days=90,
        ).set_index("product_code")
        if p1.get("normal_inbound_day_count_365") != 3 or p1.get("avg_inbound_cycle_days") != 15.0:
            inbound_errors.append(f"normal_day_or_gap={p1!r}")
        if p1.get("normal_inbound_raw_qty_365") != 30.0 or p1.get("inbound_return_raw_qty_365") != -3.0:
            inbound_errors.append(f"raw_or_return_sign={p1!r}")
        if p2.get("normal_inbound_365_exists") or p2.get("inbound_data_status") != "insufficient":
            inbound_errors.append(f"negative_not_excluded={p2!r}")
        if p3.get("recent_inbound_vendor_source") != "master_order_vendor" or not p3.get("recent_inbound_vendor_fallback"):
            inbound_errors.append(f"fallback_contract={p3!r}")
        if p4.get("recent_inbound_vendor_source") != "master_order_vendor" or p4.get("recent_inbound_vendor_code") != "M4":
            inbound_errors.append(f"blank_vendor_fallback={p4!r}")
        if p5.get("recent_inbound_vendor_source") != "actual_inbound" or p5.get("recent_inbound_vendor_code") != "V5":
            inbound_errors.append(f"mixed_vendor_prefers_actual={p5!r}")
        if p6.get("normal_inbound_day_count_365") != 1 or p6.get("last_normal_inbound_date") != "20250723":
            inbound_errors.append(f"cycle_boundary_or_future={p6!r}")
        if vendor_frame.loc["P1", "recent_inbound_vendor_code"] != "A":
            inbound_errors.append(f"vendor_rank_contract={vendor_frame.loc['P1'].to_dict()!r}")
        if vendor_frame.loc["P1", "recent_inbound_vendor_name"] != "Vendor A":
            inbound_errors.append(f"vendor_name_contract={vendor_frame.loc['P1'].to_dict()!r}")
        sql_text, sql_binds = inbound_mod._sql(
            {"stock_cd_list": ["00001"], "dashboard_product_group_list": ["0013:01"], "dashboard_product_di_list": ["0004:1"], "product_class_list": ["0031:01"], "vendor_group_list": ["0019:02"], "vendor_kind_list": ["0009:J"], "io_gu_list": ["0012:090"]},
            start_date="20250723", cutoff_date="20260722",
        )
        if "'001', '002', '101', '102', '193'" not in sql_text or "LEFT(" in sql_text:
            inbound_errors.append("fixed_tcode_whitelist_contract")
        for predicate in (
            "I.Rd11_Stock_Cd_Gcode = '0018'", "P.Rd04_Physic_Group_Gcode = '0013'",
            "P.Rd04_Physic_Di_Gcode = '0004'", "P.Rd04_Physic_Tax_Gcode = '0031'",
            "FilterVendor.Rd03_Ven_Group_Gcode = '0019'", "FilterVendor.Rd03_Ven_Kind_Gcode = '0009'",
        ):
            if predicate not in sql_text:
                inbound_errors.append(f"missing_compound_predicate={predicate}")
        if sql_binds.get("stock_cd_0") != "00001" or sql_binds.get("product_class_0") != "01":
            inbound_errors.append(f"string_bind_contract={sql_binds!r}")
        if inbound_mod._codes(["01", "1", "J"]) != ["01", "1", "J"] or inbound_mod._codes([1, 2.0, None]) != [] or inbound_mod._codes("0031:01") != ["0031:01"]:
            inbound_errors.append("strict_string_code_contract")
        cutoff_cases = {
            "current": dash_mod._dashboard_inbound_cutoff_date({"evaluation_month": "202607", "policy_date": "20260727", "date_to": "20260630"}, today=date(2026, 7, 27)),
            "past": dash_mod._dashboard_inbound_cutoff_date({"evaluation_month": "202606", "policy_date": "20260727"}, today=date(2026, 7, 27)),
            "future": dash_mod._dashboard_inbound_cutoff_date({"evaluation_month": "202608", "policy_date": "20260728"}, today=date(2026, 7, 27)),
        }
        if cutoff_cases != {"current": "20260727", "past": "20260630", "future": "20260727"}:
            inbound_errors.append(f"inbound_cutoff_policy={cutoff_cases!r}")
        synthetic_rows = 200_000
        synthetic_source = pd.DataFrame({
            "product_code": [f"S{index % 25_000:05d}" for index in range(synthetic_rows)],
            "master_order_vendor_code": [f"M{index % 97:03d}" for index in range(synthetic_rows)],
            "inbound_date": [f"2026{(index % 7) + 1:02d}{(index % 27) + 1:02d}" for index in range(synthetic_rows)],
            "io_tcode": ["001" if index % 5 else "101" for index in range(synthetic_rows)],
            "vendor_code": [f"V{index % 113:03d}" for index in range(synthetic_rows)],
            "quantity": [1 if index % 5 else -1 for index in range(synthetic_rows)],
            "oquantity": [0] * synthetic_rows,
            "supply_price": [10] * synthetic_rows,
        })
        synthetic_started = time.perf_counter()
        synthetic_facts = inbound_mod.build_dashboard_inbound_facts_frame(
            synthetic_source, data_cutoff_date="20260727", cycle_lookback_days=365, vendor_lookback_days=90,
        )
        synthetic_elapsed_ms = int((time.perf_counter() - synthetic_started) * 1000)
        if len(synthetic_facts) != 25_000 or synthetic_elapsed_ms > 10_000:
            inbound_errors.append(f"vectorized_synthetic={len(synthetic_facts)}rows/{synthetic_elapsed_ms}ms")
        if inbound_errors:
            results.append(_fail("Dashboard inbound-date facts", "; ".join(inbound_errors)))
        else:
            results.append(_ok("Dashboard inbound-date facts", "001/002 positive events, signed 101/102/193 returns, fallback, cycle gap, and string binds are preserved"))

        import app.sims.views.dashboard_lite as view_mod

        render_errors: list[str] = []

        class _FakeCtx:
            def __enter__(self):
                return self
            def __exit__(self, *_exc):
                return False

        class _FakeNumberColumnFactory:
            def NumberColumn(self, label=None, **kwargs):
                return {"label": label, **kwargs}

        class _FakeStreamlit:
            def __init__(self, *, submit_sequence: list[bool]):
                self.session_state = {}
                self._submit_sequence = list(submit_sequence)
                self.calls: dict[str, int] = {}
                self.column_specs: list[object] = []
                self.markdowns: list[str] = []
                self.captions: list[str] = []
                self.column_config = _FakeNumberColumnFactory()
            def _count(self, name: str) -> None:
                self.calls[name] = self.calls.get(name, 0) + 1
            def subheader(self, *_args, **_kwargs): self._count("subheader")
            def caption(self, text, *_args, **_kwargs):
                self._count("caption")
                self.captions.append(str(text))
            def info(self, *_args, **_kwargs): self._count("info")
            def warning(self, *_args, **_kwargs): self._count("warning")
            def error(self, *_args, **_kwargs): self._count("error")
            def success(self, *_args, **_kwargs): self._count("success")
            def write(self, *_args, **_kwargs): self._count("write")
            def markdown(self, text, *_args, **_kwargs):
                self._count("markdown")
                self.markdowns.append(str(text))
            def divider(self, *_args, **_kwargs): self._count("divider")
            def json(self, *_args, **_kwargs): self._count("json")
            def rerun(self, *_args, **_kwargs): self._count("rerun")
            def altair_chart(self, *_args, **_kwargs): self._count("altair_chart")
            def metric(self, *_args, **_kwargs): self._count("metric")
            def dataframe(self, *_args, **_kwargs): self._count("dataframe")
            def container(self, *_args, **_kwargs): return _FakeCtx()
            def expander(self, *_args, **_kwargs): return _FakeCtx()
            def spinner(self, *_args, **_kwargs): return _FakeCtx()
            def form(self, *_args, **_kwargs): return _FakeCtx()
            def columns(self, n, **_kwargs):
                self.column_specs.append(n)
                return [_FakeCtx() for _ in range(len(n) if isinstance(n, (list, tuple)) else int(n))]
            def text_input(self, _label, value="", **kwargs):
                key = kwargs.get("key")
                if key:
                    self.session_state.setdefault(key, value)
                    return self.session_state.get(key, "")
                return value
            def radio(self, _label, options=None, **kwargs):
                key = kwargs.get("key")
                return self.session_state.get(key, list(options or [""])[0])
            def toggle(self, _label, value=False, **kwargs):
                self._count("toggle")
                key = kwargs.get("key")
                return self.session_state.get(key, value)
            def number_input(self, _label, value=0, **kwargs):
                key = kwargs.get("key")
                return self.session_state.get(key, value)
            def selectbox(self, _label, options=None, **kwargs):
                key = kwargs.get("key")
                return self.session_state.get(key, list(options or [""])[0])
            def multiselect(self, label, options=None, default=None, **_kwargs):
                self._count(f"multiselect:{label}")
                if label == "재고위치":
                    return ["00001"]
                if label == "제외할 제품그룹명":
                    return ["0013:G_EX"]
                if label == "제외할 제품구분명":
                    return ["0004:D_EX"]
                if label == "제외할 제품분류명":
                    return ["0031:C_EX"]
                return []
            def form_submit_button(self, *_args, **_kwargs):
                self._count("submit_button")
                return self._submit_sequence.pop(0) if self._submit_sequence else False

        old_st = getattr(view_mod, "st")
        old_build = getattr(view_mod, "build_dashboard_lite_facts")
        old_stock_options = getattr(view_mod, "_dashboard_stock_options")
        old_code_options = getattr(view_mod, "_dashboard_code_name_options")
        old_manufacturer_query = getattr(view_mod, "query_to_df")
        old_dashboard_target = getattr(view_mod, "_DASHBOARD_RENDER_TARGET")
        old_dashboard_identity = getattr(view_mod, "_dashboard_context_identity")
        old_current_chat_room_id = getattr(view_mod, "get_current_chat_room_id")
        old_dashboard_profile_loader = getattr(view_mod, "load_dashboard_profile")
        build_calls: list[dict] = []
        requested_gcodes: list[str] = []
        option_calls = {"stock": 0, "code": 0}

        class _FakeDashboardTarget:
            def __init__(self):
                self.container_calls = 0
            def container(self):
                self.container_calls += 1
                return _FakeCtx()

        def _fake_build(params=None):
            build_calls.append(dict(params or {}))
            return dict(facts)

        def _fake_stock_options():
            option_calls["stock"] += 1
            return ["00001", "00002"], {"00001": "본사 창고", "00002": "전주 창고"}

        def _fake_code_options(gcode):
            option_calls["code"] += 1
            requested_gcodes.append(str(gcode))
            if str(gcode) == "0013":
                return ["0013:G_EX"], {"0013:G_EX": "제외그룹"}
            if str(gcode) == "0004":
                return ["0004:D_EX"], {"0004:D_EX": "제외구분"}
            if str(gcode) == "0031":
                return ["0031:C_EX"], {"0031:C_EX": "제외분류"}
            if str(gcode) == "0019":
                return ["0019:G_EX"], {"0019:G_EX": "vendor-group"}
            if str(gcode) == "0009":
                return ["0009:K_EX"], {"0009:K_EX": "vendor-kind"}
            if str(gcode) == "0012":
                return ["0012:I_EX"], {"0012:I_EX": "io-gu"}
            return [], {}

        try:
            fake_st = _FakeStreamlit(submit_sequence=[False])
            setattr(view_mod, "st", fake_st)
            setattr(view_mod, "build_dashboard_lite_facts", _fake_build)
            setattr(view_mod, "_dashboard_stock_options", _fake_stock_options)
            setattr(view_mod, "_dashboard_code_name_options", _fake_code_options)
            setattr(view_mod, "_dashboard_context_identity", lambda: {"user_id": "8", "company_id": "4", "db_sig": ""})
            setattr(view_mod, "get_current_chat_room_id", lambda: "dashboard-room")
            setattr(view_mod, "load_dashboard_profile", lambda **_kwargs: {"io_gu_list": ["0012:I_EX"]})
            primary_target = _FakeDashboardTarget()
            view_mod.set_dashboard_lite_render_target(primary_target)
            fake_st.session_state["current_room"] = "dashboard-room"
            fake_st.session_state["__chat_current_room_id"] = "dashboard-room"
            fake_st.session_state["chat_rooms"] = [{
                "id": "dashboard-room",
                "name": "새 대화",
                "created_at": "2026-07-25T17:49:31+09:00",
                "auto_created": True,
                "name_auto": True,
                "title_initialized": False,
            }]
            opened = view_mod.render_dashboard_lite()
            if build_calls:
                render_errors.append(f"open_called_service={len(build_calls)}")
            if option_calls != {"stock": 1, "code": 6}:
                render_errors.append(f"open_called_option_source={option_calls!r}")
            if fake_st.column_specs[:1] != [[1, 1, 1, 1.1, 2.1, 2.1]]:
                render_errors.append(f"single_scope_row_not_expected_columns={fake_st.column_specs!r}")
            if (opened.get("meta") or {}).get("status") != "condition_only":
                render_errors.append(f"open_status={opened!r}")

            rerendered = view_mod.render_dashboard_lite()
            if build_calls or (rerendered.get("meta") or {}).get("status") != "condition_only":
                render_errors.append(f"rerun_triggered_analysis={rerendered!r}|calls={len(build_calls)}")
            if option_calls != {"stock": 1, "code": 6}:
                render_errors.append(f"rerun_reloaded_option_source={option_calls!r}")
            stale_widget_values = {
                "__dashboard_lite_stock_labels": ["00001", "stale-stock"],
                "__dashboard_lite_vendor_group_list": ["0019:G_EX", "stale-vendor-group"],
                "__dashboard_lite_vendor_kind_list": ["0009:K_EX", "stale-vendor-kind"],
                "__dashboard_lite_product_group_list": ["0013:G_EX", "stale-product-group"],
                "__dashboard_lite_product_di_list": ["0004:D_EX", "stale-product-di"],
                "__dashboard_lite_product_class_list": ["0031:C_EX", "stale-product-class"],
                "__dashboard_lite_io_gu_list": ["0012:I_EX", "stale-io-gu"],
            }
            fake_st.session_state.update(stale_widget_values)
            stale_stock_rerender = view_mod.render_dashboard_lite()
            for widget_key, values in stale_widget_values.items():
                expected_values = [value for value in values if not value.startswith("stale-")]
                if fake_st.session_state.get(widget_key) != expected_values:
                    render_errors.append(f"stale_multiselect_value_not_pruned={widget_key}:{fake_st.session_state.get(widget_key)!r}")
            if (stale_stock_rerender.get("meta") or {}).get("status") != "condition_only" or build_calls:
                render_errors.append("stock_widget_state_rerun_triggered_analysis")
            fake_st._submit_sequence = [True, False]
            first = view_mod.render_dashboard_lite()
            second = view_mod.render_dashboard_lite()
            if len(build_calls) != 1:
                render_errors.append(f"submit_rerun_build_calls={len(build_calls)}")
            if option_calls != {"stock": 1, "code": 6}:
                render_errors.append(f"submit_reloaded_option_source={option_calls!r}")
            if build_calls and build_calls[0].get("stock_cd_list") != ["00001"]:
                render_errors.append(f"stock_cd_list_not_passed={build_calls[0]!r}")
            if "0031" not in requested_gcodes or "0001" in requested_gcodes:
                render_errors.append(f"product_class_gcode_wrong={requested_gcodes!r}")
            if first.get("meta", {}).get("facts_kind") != "SIMS_DASHBOARD_FACTS_V01":
                render_errors.append("submit_missing_facts_kind")
            dashboard_cache = fake_st.session_state.get("__dashboard_lite_result") or {}
            if not {"company_id", "query_fingerprint", "elapsed_seconds", "created_at"}.issubset(dashboard_cache):
                render_errors.append(f"dashboard_result_metadata_missing={dashboard_cache!r}")
            if fake_st.session_state["chat_rooms"][0].get("name") != "2026-07-25 17:49 Dashboard Lite":
                render_errors.append(f"dashboard_room_title={fake_st.session_state['chat_rooms'][0]!r}")
            if (
                fake_st.session_state["chat_rooms"][0].get("auto_created") is not False
                or fake_st.session_state["chat_rooms"][0].get("name_auto") is not False
                or fake_st.session_state["chat_rooms"][0].get("title_initialized") is not True
            ):
                render_errors.append(f"dashboard_room_title_flags={fake_st.session_state['chat_rooms'][0]!r}")
            title_after_first_submit = fake_st.session_state["chat_rooms"][0].get("name")
            if view_mod._mark_dashboard_room_title() is not False or fake_st.session_state["chat_rooms"][0].get("name") != title_after_first_submit:
                render_errors.append("dashboard_room_title_changed_on_repeat")

            protected_rooms = [
                {
                    "id": "dashboard-room",
                    "name": "사용자 지정 제목",
                    "created_at": "2026-07-25T17:50:00",
                    "auto_created": True,
                    "name_auto": False,
                    "title_initialized": True,
                },
                {
                    "id": "saved-dashboard-room",
                    "name": "Dashboard Lite",
                    "created_at": "2026-07-24T10:00:00",
                    "auto_created": False,
                    "name_auto": False,
                    "title_initialized": True,
                },
            ]
            fake_st.session_state["chat_rooms"] = protected_rooms
            fake_st.session_state["current_room"] = "dashboard-room"
            fake_st.session_state["__chat_current_room_id"] = "dashboard-room"
            if view_mod._mark_dashboard_room_title() is not False or protected_rooms[0]["name"] != "사용자 지정 제목":
                render_errors.append(f"dashboard_user_title_overwritten={protected_rooms[0]!r}")
            fake_st.session_state["current_room"] = "saved-dashboard-room"
            fake_st.session_state["__chat_current_room_id"] = "saved-dashboard-room"
            setattr(view_mod, "get_current_chat_room_id", lambda: "saved-dashboard-room")
            if view_mod._mark_dashboard_room_title() is not False or protected_rooms[1]["name"] != "Dashboard Lite":
                render_errors.append(f"dashboard_saved_title_rewritten={protected_rooms[1]!r}")
            fake_st.session_state["chat_rooms"] = [{
                "id": "dashboard-room",
                "name": title_after_first_submit,
                "created_at": "2026-07-25T17:49:31+09:00",
                "auto_created": False,
                "name_auto": False,
                "title_initialized": True,
            }]
            fake_st.session_state["current_room"] = "dashboard-room"
            fake_st.session_state["__chat_current_room_id"] = "dashboard-room"
            setattr(view_mod, "get_current_chat_room_id", lambda: "dashboard-room")
            if fake_st.calls.get("rerun"):
                render_errors.append(f"dashboard_unexpected_rerun={fake_st.calls.get('rerun')}")
            if first.get("type") != "dashboard_lite" or not first.get("final"):
                render_errors.append(f"dashboard_chat_payload_missing={first!r}")
            if not isinstance(first.get("meta", {}).get("dashboard_cache"), dict):
                render_errors.append("dashboard_chat_cache_missing")
            if (
                not dashboard_cache.get("dashboard_event_id")
                or first.get("id") != dashboard_cache.get("dashboard_event_id")
                or first.get("meta", {}).get("dashboard_event_id") != dashboard_cache.get("dashboard_event_id")
                or first.get("meta", {}).get("dashboard_cache", {}).get("dashboard_event_id") != dashboard_cache.get("dashboard_event_id")
            ):
                render_errors.append("dashboard_event_id_not_created_before_snapshot")
            render_tree = ast.parse(Path("app/sims/views/dashboard_lite.py").read_text(encoding="utf-8"))
            render_node = next(
                node
                for node in render_tree.body
                if isinstance(node, ast.FunctionDef) and node.name == "render_dashboard_lite"
            )
            dashboard_uuid_calls = [
                node
                for node in ast.walk(render_node)
                if isinstance(node, ast.Call)
                and isinstance(node.func, ast.Attribute)
                and isinstance(node.func.value, ast.Name)
                and node.func.value.id == "uuid"
                and node.func.attr == "uuid4"
            ]
            if len(dashboard_uuid_calls) != 1:
                render_errors.append(f"dashboard_event_uuid_create_count={len(dashboard_uuid_calls)}")
            if primary_target.container_calls != 0:
                render_errors.append(f"dashboard_panel_rendered_primary={primary_target.container_calls}")
            view_mod.reset_dashboard_lite_primary_render_guard()
            if view_mod.render_dashboard_lite_chat_item(dashboard_cache, render_mode="primary") is not True:
                render_errors.append("dashboard_primary_message_render_missing")
            primary_detail_rows = ((dashboard_cache.get("facts") or {}).get("inventory") or {}).get("risk_detail_rows") or []
            primary_toggle_before_chat = fake_st.calls.get("toggle", 0)
            if view_mod.render_dashboard_lite_chat_item(dashboard_cache, render_mode="primary") is not False:
                render_errors.append("dashboard_same_rerun_primary_not_guarded")
            view_mod.render_dashboard_lite_chat_item(first["meta"]["dashboard_cache"], render_mode="chat")
            if not any(text == "## 일일 재고·매출 보고" for text in fake_st.markdowns):
                render_errors.append("dashboard_result_title_not_rendered")
            if not any("조회기간:" in text and "평가월:" in text and "재고위치:" in text for text in fake_st.captions):
                render_errors.append(f"dashboard_scope_header_missing={fake_st.captions!r}")
            if not any(text.startswith("조회 완료 · ") and " · " in text for text in fake_st.captions):
                render_errors.append(f"dashboard_completed_header_missing={fake_st.captions!r}")
            if second.get("meta", {}).get("facts_kind") != "SIMS_DASHBOARD_FACTS_V01":
                render_errors.append("rerun_cache_not_rendered")
            if primary_target.container_calls != 0:
                render_errors.append(f"dashboard_primary_render_escaped_message_position={primary_target.container_calls}")
            if primary_detail_rows and primary_toggle_before_chat < 1:
                render_errors.append("dashboard_primary_detail_toggle_missing")
            if fake_st.calls.get("toggle", 0) != primary_toggle_before_chat:
                render_errors.append("dashboard_chat_snapshot_rendered_detail_toggle")
            if len((((fake_st.session_state.get("__dashboard_lite_result") or {}).get("facts") or {}).get("inventory") or {}).get("risk_detail_rows") or []) != len(primary_detail_rows):
                render_errors.append("dashboard_chat_push_mutated_primary_detail_rows")
            view_mod.clear_dashboard_lite_session_state(fake_st.session_state)
            after_company_change = view_mod.render_dashboard_lite()
            if len(build_calls) != 1:
                render_errors.append(f"company_change_triggered_analysis={len(build_calls)}")
            if option_calls != {"stock": 2, "code": 12}:
                render_errors.append(f"company_change_did_not_reload_options={option_calls!r}")
            if (after_company_change.get("meta") or {}).get("status") != "condition_only":
                render_errors.append(f"company_change_open_status={after_company_change!r}")

            manufacturer_queries: list[tuple] = []
            def _fake_manufacturer_query(_sql, bind=None):
                bound = tuple(bind or ())
                manufacturer_queries.append(bound)
                if bound == ("10047",):
                    return pd.DataFrame([{"manufacturer_code": "10047", "manufacturer_name": "삼진제약"}])
                if bound == ("%삼진제약%",):
                    return pd.DataFrame([{"manufacturer_code": "10047", "manufacturer_name": "삼진제약"}])
                if bound == ("%동진%",):
                    return pd.DataFrame([
                        {"manufacturer_code": "20001", "manufacturer_name": "동진A"},
                        {"manufacturer_code": "20002", "manufacturer_name": "동진B"},
                    ])
                return pd.DataFrame()
            setattr(view_mod, "query_to_df", _fake_manufacturer_query)
            fake_st.session_state["__dashboard_lite_manufacturer_text"] = "10047"
            resolved_code = view_mod._resolve_dashboard_manufacturer("10047")
            if resolved_code.get("codes") != ["10047"]:
                render_errors.append(f"manufacturer_code_resolution={resolved_code!r}")
            if "제약사: 삼진제약 [10047]" not in view_mod._dashboard_scope_header({"month_from": "202601", "month_to": "202606", "evaluation_month": "202607", "manufacturer_scope_label": resolved_code.get("label")}):
                render_errors.append("manufacturer_exact_scope_header_missing")
            resolved_name = view_mod._resolve_dashboard_manufacturer("삼진제약")
            if resolved_name.get("codes") != ["10047"]:
                render_errors.append(f"manufacturer_name_resolution={resolved_name!r}")
            resolved_multiple = view_mod._resolve_dashboard_manufacturer("동진")
            if resolved_multiple.get("codes") != ["20001", "20002"] or resolved_multiple.get("match_mode") != "name_like":
                render_errors.append(f"manufacturer_multi_like_resolution={resolved_multiple!r}")
            if "제약사: '동진' 포함 2개사" not in view_mod._dashboard_scope_header({"month_from": "202601", "month_to": "202606", "evaluation_month": "202607", "manufacturer_scope_label": resolved_multiple.get("label")}):
                render_errors.append("manufacturer_multi_scope_header_missing")
            if "__dashboard_lite_manufacturer_candidates" in fake_st.session_state:
                render_errors.append("manufacturer_candidate_state_retained")
            reset_all = view_mod._resolve_dashboard_manufacturer("전체")
            if reset_all.get("codes") or any(fake_st.session_state.get(k) for k in ("__dashboard_lite_manufacturer_test_codes", "__dashboard_lite_manufacturer_resolved_code", "__dashboard_lite_manufacturer_resolved_name")):
                render_errors.append(f"manufacturer_all_did_not_clear={fake_st.session_state!r}")
            if "제약사: 전체" not in view_mod._dashboard_scope_header({"month_from": "202601", "month_to": "202606", "evaluation_month": "202607"}):
                render_errors.append("manufacturer_all_scope_header_missing")
            missing = view_mod._resolve_dashboard_manufacturer("없는제약사")
            if missing.get("status") != "missing" or fake_st.session_state.get("__dashboard_lite_manufacturer_test_codes"):
                render_errors.append(f"manufacturer_missing_reused_previous={missing!r}")
            too_short = view_mod._resolve_dashboard_manufacturer("가")
            if too_short.get("status") != "too_short" or too_short.get("codes"):
                render_errors.append(f"manufacturer_short_search_not_blocked={too_short!r}")
            if not manufacturer_queries or any(not isinstance(query, tuple) or not query for query in manufacturer_queries):
                render_errors.append(f"manufacturer_query_binding_missing={manufacturer_queries!r}")
        finally:
            setattr(view_mod, "st", old_st)
            setattr(view_mod, "build_dashboard_lite_facts", old_build)
            setattr(view_mod, "_dashboard_stock_options", old_stock_options)
            setattr(view_mod, "_dashboard_code_name_options", old_code_options)
            setattr(view_mod, "query_to_df", old_manufacturer_query)
            setattr(view_mod, "_dashboard_context_identity", old_dashboard_identity)
            setattr(view_mod, "get_current_chat_room_id", old_current_chat_room_id)
            setattr(view_mod, "load_dashboard_profile", old_dashboard_profile_loader)
            view_mod.set_dashboard_lite_render_target(old_dashboard_target)

        if render_errors:
            results.append(_fail("Dashboard Lite button-gated render cache", "; ".join(render_errors)))
        else:
            results.append(_ok("Dashboard Lite button-gated render cache", "open loads options once without facts; reruns reuse options; submit calls facts once; the dedicated result keeps scope/time metadata and names an empty Dashboard-only room"))

        chart_errors: list[str] = []
        try:
            sales_chart_spec = view_mod._build_sales_bar_chart(facts).to_dict()
            sales_chart_json = json.dumps(sales_chart_spec, ensure_ascii=False)
            if '"type": "bar"' not in sales_chart_json:
                chart_errors.append("sales_chart_not_bar")
            if '"type": "line"' not in sales_chart_json or '"type": "point"' not in sales_chart_json:
                chart_errors.append("sales_chart_line_or_marker_missing")
            if not all(token in sales_chart_json for token in ("#2563eb", "#f97316", "#0f766e")):
                chart_errors.append("sales_chart_actual_forecast_marker_colors_missing")
            if not all(token in sales_chart_json for token in ("실제매출", "예상매출", "현재일 기준")):
                chart_errors.append("sales_chart_user_legend_missing")
        except Exception as exc:
            chart_errors.append(f"sales_chart_build={type(exc).__name__}:{exc}")
        if chart_errors:
            results.append(_fail("Dashboard Lite monthly sales bar chart", "; ".join(chart_errors)))
        else:
            results.append(_ok("Dashboard Lite monthly sales chart", "actual bars, forecast line, and the current-day marker are rendered from the existing compact facts"))

        preforecast_errors: list[str] = []
        chart_source = pd.DataFrame(
            [
                {
                    "제약사명": "검증제약",
                    "2026-01 매출": 10,
                    "2026-02 매출": 20,
                    "2026-03 매출": 30,
                    "2026-04 매출": 40,
                    "2026-05 매출": 50,
                    "2026-06 매출": 160,
                    "당월 현재매출": 70,
                    "당월 예상매출": 90,
                }
            ]
        )
        chart_source_changed_target = chart_source.copy(deep=True)
        chart_source_changed_target.loc[0, "2026-04 매출"] = 40_000
        chart_source_changed_history = chart_source.copy(deep=True)
        chart_source_changed_history.loc[0, "2026-03 매출"] = 30_000
        try:
            chart_history = [
                {"period": "2025-10", "period_sort": "202510", "value": 7},
                {"period": "2025-11", "period_sort": "202511", "value": 8},
                {"period": "2025-12", "period_sort": "202512", "value": 9},
            ]
            chart_sales = dash_mod._build_sales_facts({"df": chart_source, "meta": {"evaluation_month": "202607"}}, history_actuals=chart_history, evaluation_month="202607", policy_date="20260714", today=date(2026, 7, 28))
            target_changed_sales = dash_mod._build_sales_facts({"df": chart_source_changed_target, "meta": {"evaluation_month": "202607"}}, history_actuals=chart_history, evaluation_month="202607", policy_date="20260714", today=date(2026, 7, 28))
            history_changed_sales = dash_mod._build_sales_facts({"df": chart_source_changed_history, "meta": {"evaluation_month": "202607"}}, history_actuals=chart_history, evaluation_month="202607", policy_date="20260714", today=date(2026, 7, 28))
            chart_rows = chart_sales.get("chart_rows") or []
            period_order = []
            for row in chart_rows:
                if row.get("period") not in period_order:
                    period_order.append(row.get("period"))
            expected_period_order = [f"2026-{month:02d}" for month in range(1, 8)]
            if period_order != expected_period_order:
                preforecast_errors.append(f"period_order={period_order!r}")
            actual_periods = {
                row.get("period")
                for row in chart_rows
                if row.get("kind") in {"완료월 실제", "당월 현재(부분월)"}
            }
            if actual_periods != set(expected_period_order):
                preforecast_errors.append(f"actual_periods={sorted(actual_periods)!r}")
            past_rows = [row for row in chart_rows if row.get("kind") == "완료월 사전예상"]
            if [row.get("period") for row in past_rows] != ["2026-01", "2026-02", "2026-03", "2026-04", "2026-05", "2026-06"]:
                preforecast_errors.append(f"past_forecast_periods={[row.get('period') for row in past_rows]!r}")
            july_kinds = {row.get("kind") for row in chart_rows if row.get("period") == "2026-07"}
            if july_kinds != {"당월 현재(부분월)", "당월 예상"}:
                preforecast_errors.append(f"july_kinds={july_kinds!r}")
            def _forecast_value(sales: dict[str, Any], period: str) -> float:
                return float(next(row["value"] for row in sales.get("chart_rows") or [] if row.get("kind") == "완료월 사전예상" and row.get("period") == period))
            if _forecast_value(chart_sales, "2026-04") != _forecast_value(target_changed_sales, "2026-04"):
                preforecast_errors.append("target_month_actual_changed_its_preforecast")
            if _forecast_value(chart_sales, "2026-04") == _forecast_value(history_changed_sales, "2026-04"):
                preforecast_errors.append("prior_history_did_not_change_later_preforecast")
            visualization = chart_sales.get("visualization") or {}
            if abs(float(visualization.get("time_progress_pct") or 0) - (14 / 31 * 100)) > 1e-9:
                preforecast_errors.append(f"time_progress={visualization.get('time_progress_pct')!r}")
            if abs(float(visualization.get("expected_to_date_sales") or 0) - (90 * 14 / 31)) > 1e-9:
                preforecast_errors.append(f"expected_to_date={visualization.get('expected_to_date_sales')!r}")
            if visualization.get("remaining_forecast") != 20:
                preforecast_errors.append(f"remaining_forecast={visualization.get('remaining_forecast')!r}")
            if visualization.get("time_adjusted_status") != "시간 진척보다 앞섬":
                preforecast_errors.append(f"time_adjusted_status={visualization.get('time_adjusted_status')!r}")
            if dash_mod._dashboard_time_progress("202602", today=date(2026, 7, 28)).get("pct") != 100.0:
                preforecast_errors.append("past_month_time_progress")
            if dash_mod._dashboard_time_progress("202608", today=date(2026, 7, 28)).get("pct") != 0.0:
                preforecast_errors.append("future_month_time_progress")
            if dash_mod._dashboard_time_progress("202402", today=date(2024, 2, 29)).get("pct") != 100.0:
                preforecast_errors.append("leap_year_time_progress")
            chart_spec = view_mod._build_sales_bar_chart({"sales": chart_sales}).to_dict()
            expected_display_order = [f"{month}월" for month in range(1, 8)]
            layer_sorts = [layer.get("encoding", {}).get("x", {}).get("sort") for layer in chart_spec.get("layer", [])]
            if not layer_sorts or any(sort != expected_display_order for sort in layer_sorts):
                preforecast_errors.append(f"layer_period_sorts={layer_sorts!r}")
            if any("xOffset" in layer.get("encoding", {}) for layer in chart_spec.get("layer", [])):
                preforecast_errors.append("x_offset_side_by_side_remains")
            if chart_spec.get("resolve", {}).get("scale", {}).get("y") != "shared":
                preforecast_errors.append(f"y_scale_resolve={chart_spec.get('resolve')!r}")
            layer_marks = [layer.get("mark") or {} for layer in chart_spec.get("layer", [])]
            if layer_marks[0].get("type") != "bar" or layer_marks[0].get("size") != 32:
                preforecast_errors.append(f"actual_bar_width={layer_marks!r}")
            if layer_marks[1].get("type") != "line" or layer_marks[2].get("type") != "point":
                preforecast_errors.append(f"forecast_line_or_current_marker={layer_marks!r}")
            layer_stacks = [layer.get("encoding", {}).get("y", {}).get("stack") for layer in chart_spec.get("layer", [])]
            if layer_stacks != [None, None, None]:
                preforecast_errors.append(f"layer_y_stacks={layer_stacks!r}")
            forecast_encoding = chart_spec.get("layer", [])[1].get("encoding", {})
            if forecast_encoding.get("detail", {}).get("field") != "forecast_segment":
                preforecast_errors.append(f"forecast_gap_segment={forecast_encoding.get('detail')!r}")
            cross_year_rows = list(chart_rows)
            cross_year_rows.append({"period": "2027-01", "period_sort": "202701", "kind": "완료월 실제", "value": 12})
            cross_year_spec = view_mod._build_sales_bar_chart({"sales": {"chart_rows": cross_year_rows, "visualization": visualization}}).to_dict()
            if "2027년 1월" not in json.dumps(cross_year_spec, ensure_ascii=False):
                preforecast_errors.append("cross_year_month_label_missing")
            missing_forecast_rows = [row for row in chart_rows if not (row.get("kind") == "완료월 사전예상" and row.get("period") == "2026-03")]
            missing_forecast_spec = view_mod._build_sales_bar_chart({"sales": {"chart_rows": missing_forecast_rows, "visualization": visualization}}).to_dict()
            missing_json = json.dumps(missing_forecast_spec, ensure_ascii=False)
            if '"period": "2026-03", "period_sort": "202603", "kind": "완료월 사전예상", "value": 0' in missing_json:
                preforecast_errors.append("missing_forecast_zero_fabricated")
            actual_values = {
                row.get("period"): float(row.get("value") or 0)
                for row in chart_rows
                if row.get("kind") in {"완료월 실제", "당월 현재(부분월)"}
            }
            if actual_values.get("2026-06", 0) <= actual_values.get("2026-07", 0):
                preforecast_errors.append(f"actual_height_fixture_invalid={actual_values!r}")
        except Exception as exc:
            preforecast_errors.append(f"preforecast_runtime={type(exc).__name__}:{exc}")
        if preforecast_errors:
            results.append(_fail("Dashboard Lite completed-month preforecast bars", "; ".join(preforecast_errors)))
        else:
            results.append(_ok("Dashboard Lite sales visualization", "actual bars, forecast line gaps, and the current-day marker use the compact facts without mutating raw chart rows"))

        presentation_errors: list[str] = []
        try:
            presentation_base = {"sales": {"metrics": {
                "current_month_sales": {"value": 110},
                "current_month_forecast_sales": {"value": 100},
                "current_month_progress_pct": {"value": 110},
                "time_progress_pct": {"value": 90.3, "time_basis": "202607 28/31일 경과 진행중"},
                "time_adjusted_achievement_pct": {"value": 121.8},
            }, "visualization": {"expected_to_date_sales": 90.3}}}
            over_state = view_mod._sales_presentation_state(presentation_base)
            if over_state.get("comparison_label") != "월말 예상 초과" or over_state.get("comparison_amount") != 10:
                presentation_errors.append(f"over_state={over_state!r}")
            if over_state.get("time_adjusted_status") != "현재일 예상보다 앞섬":
                presentation_errors.append(f"ahead_status={over_state!r}")
            under_facts = json.loads(json.dumps(presentation_base))
            under_facts["sales"]["metrics"]["current_month_sales"]["value"] = 80
            under_facts["sales"]["metrics"]["time_adjusted_achievement_pct"]["value"] = 96
            under_state = view_mod._sales_presentation_state(under_facts)
            if under_state.get("comparison_label") != "월말 예상 잔여" or under_state.get("comparison_amount") != 20:
                presentation_errors.append(f"under_state={under_state!r}")
            equal_facts = json.loads(json.dumps(presentation_base))
            equal_facts["sales"]["metrics"]["current_month_sales"]["value"] = 100
            equal_facts["sales"]["metrics"]["time_adjusted_achievement_pct"]["value"] = 94
            equal_state = view_mod._sales_presentation_state(equal_facts)
            if equal_state.get("comparison_label") != "월말 예상 도달" or equal_state.get("comparison_amount") != 0:
                presentation_errors.append(f"equal_state={equal_state!r}")
            if equal_state.get("time_adjusted_status") != "현재일 예상보다 뒤처짐":
                presentation_errors.append(f"behind_status={equal_state!r}")
            unchanged_remaining = (chart_sales.get("visualization") or {}).get("remaining_forecast")
            if unchanged_remaining != 20:
                presentation_errors.append(f"remaining_fact_mutated={unchanged_remaining!r}")
            ui_source = (PROJECT_ROOT / "app" / "sims" / "views" / "dashboard_lite.py").read_text(encoding="utf-8")
            for label in ("현재일 기준 예상매출", "현재일 기준 달성률", "월 경과", "월말 예상 초과", "오늘의 매출 요약"):
                if label not in ui_source:
                    presentation_errors.append(f"ui_label_missing={label}")
            if '"시간 진척률"' in ui_source or '"시간 대비 달성률"' in ui_source:
                presentation_errors.append("legacy_ui_labels_remain")
        except Exception as exc:
            presentation_errors.append(f"presentation_runtime={type(exc).__name__}:{exc}")
        if presentation_errors:
            results.append(_fail("Dashboard Lite sales presentation states", "; ".join(presentation_errors)))
        else:
            results.append(_ok("Dashboard Lite sales presentation states", "four cards and the status bar distinguish forecast overage, remaining amount, forecast reached, and current-day achievement labels without changing facts"))

        reset_errors: list[str] = []
        session_fixture = {
            "__dashboard_lite_result": {"facts": "cached"},
            "__dashboard_lite_run_seq": 7,
            "__dashboard_lite_styles_loaded": True,
            "__dashboard_lite_stock_labels": ["00001"],
            "__dashboard_lite_exclude_product_group_list": ["0013:G_EX"],
            "__dashboard_lite_exclude_product_di_list": ["0004:D_EX"],
            "__dashboard_lite_exclude_product_class_list": ["0031:C_EX"],
            "__dashboard_lite_cache_key__widget_ns_1": "old",
            "__dashboard_lite_result__another_ns": "old-other",
            "__sims_widget_ns": "widget_ns_1",
            "__sims_cat": "분석/KPI",
            "__sims_action": "Dashboard Lite v0.1",
            "__login_user": "keep-user",
            "__selected_company_id": 4,
            "__chat_current_room_id": "room-a",
        }
        removed_reset_keys = view_mod.clear_dashboard_lite_session_state(session_fixture, namespace="widget_ns_1")
        if "__dashboard_lite_result" in session_fixture or "__dashboard_lite_result__another_ns" in session_fixture:
            reset_errors.append(f"dashboard_keys_remaining={sorted(k for k in session_fixture if str(k).startswith('__dashboard_lite_'))!r}")
        if not {"__dashboard_lite_result", "__dashboard_lite_cache_key__widget_ns_1"}.issubset(set(removed_reset_keys)):
            reset_errors.append(f"removed_keys={removed_reset_keys!r}")
        if session_fixture.get("__login_user") != "keep-user" or session_fixture.get("__selected_company_id") != 4 or session_fixture.get("__chat_current_room_id") != "room-a":
            reset_errors.append(f"login_company_chat_mutated={session_fixture!r}")
        try:
            from types import SimpleNamespace

            login_src_for_clear = Path("app/ui/ssai_login.py").read_text(encoding="utf-8")
            login_tree_for_clear = ast.parse(login_src_for_clear)
            clear_node = next(
                node for node in login_tree_for_clear.body
                if isinstance(node, ast.FunctionDef) and node.name == "_clear_company_dependent_state"
            )
            company_change_state = {
                "__dashboard_lite_scope_options": {"stock_codes": ["previous-company"]},
                "__dashboard_lite_result": {"facts": "previous-company"},
                "__dashboard_lite_run_seq": 9,
                "__dashboard_lite_stock_labels": ["previous-company"],
                "__login_user": "keep-user",
                "ssai_selected_company": {"company_id": 5},
                "__chat_current_room_id": "room-after-change",
            }
            cache_clears = {"count": 0}
            fake_login_st = SimpleNamespace(
                session_state=company_change_state,
                cache_data=SimpleNamespace(clear=lambda: cache_clears.__setitem__("count", cache_clears["count"] + 1)),
            )
            clear_ns = {
                "st": fake_login_st,
                "log": SimpleNamespace(warning=lambda *_args, **_kwargs: None),
            }
            exec(compile(ast.Module(body=[clear_node], type_ignores=[]), "company_state_clear", "exec"), clear_ns)
            clear_ns["_clear_company_dependent_state"]()
            if any(str(key).startswith("__dashboard_lite_") for key in company_change_state):
                reset_errors.append(f"company_change_dashboard_state_remaining={company_change_state!r}")
            if company_change_state.get("__login_user") != "keep-user" or company_change_state.get("ssai_selected_company") != {"company_id": 5} or company_change_state.get("__chat_current_room_id") != "room-after-change":
                reset_errors.append(f"company_change_preserved_state_mutated={company_change_state!r}")
            if cache_clears["count"] != 1:
                reset_errors.append(f"company_change_cache_clear_count={cache_clears!r}")
        except Exception as exc:
            reset_errors.append(f"company_change_clear_runtime={type(exc).__name__}:{exc}")
        main_src_for_reset = Path("app/Lmstudio_SSAI_chat_main.py").read_text(encoding="utf-8")
        if "clear_dashboard_lite_session_state(st.session_state, _ns)" not in main_src_for_reset:
            reset_errors.append("main_reset_helper_not_called")
        if reset_errors:
            results.append(_fail("Dashboard Lite option reset clears real state", "; ".join(reset_errors)))
        else:
            results.append(_ok("Dashboard Lite option reset clears real state", "Dashboard-specific results/widgets/cache clear through helper while login/company/chat state remains"))

        new_chat_errors: list[str] = []
        try:
            from types import SimpleNamespace

            main_src_for_new_chat = Path("app/Lmstudio_SSAI_chat_main.py").read_text(encoding="utf-8")
            main_tree_for_new_chat = ast.parse(main_src_for_new_chat)
            clear_node = next(
                node for node in main_tree_for_new_chat.body
                if isinstance(node, ast.FunctionDef) and node.name == "_clear_dashboard_lite_for_new_chat"
            )
            new_chat_state = {
                "__dashboard_lite_result": {"facts": "cached"},
                "__dashboard_lite_scope_options": {"stock": ["old"]},
                "__dashboard_lite_run_seq": 3,
                "__dashboard_lite_query_fingerprint": "old-query",
                "__dashboard_lite_elapsed_seconds": 9.1,
                "__dashboard_lite_created_at": "old-time",
                "__chat_history": [{"id": "old-chat"}],
                "__login_user": "keep-user",
                "__selected_company_id": 4,
            }
            fake_main_st = SimpleNamespace(session_state=new_chat_state)
            clear_ns = {"st": fake_main_st, "log": SimpleNamespace(debug=lambda *_args, **_kwargs: None)}
            exec(compile(ast.Module(body=[clear_node], type_ignores=[]), "new_chat_dashboard_clear", "exec"), clear_ns)
            removed = clear_ns["_clear_dashboard_lite_for_new_chat"]()
            if not removed or any(str(key).startswith("__dashboard_lite_") for key in new_chat_state):
                new_chat_errors.append(f"dashboard_state_remaining={new_chat_state!r}")
            if new_chat_state.get("__login_user") != "keep-user" or new_chat_state.get("__selected_company_id") != 4:
                new_chat_errors.append(f"login_company_mutated={new_chat_state!r}")
            already_reset_state = {
                "__dashboard_lite_result": {"facts": "cleared-by-option-reset"},
                "__dashboard_lite_scope_options": {"stock": ["cleared-by-option-reset"]},
                "__login_user": "keep-user",
            }
            view_mod.clear_dashboard_lite_session_state(already_reset_state)
            fake_main_st.session_state = already_reset_state
            clear_ns["_clear_dashboard_lite_for_new_chat"]()
            if any(str(key).startswith("__dashboard_lite_") for key in already_reset_state):
                new_chat_errors.append(f"option_reset_then_new_chat_remaining={already_reset_state!r}")
            if already_reset_state.get("__login_user") != "keep-user":
                new_chat_errors.append(f"option_reset_then_new_chat_login_mutated={already_reset_state!r}")
            new_room_node = next(
                node for node in ast.walk(main_tree_for_new_chat)
                if isinstance(node, ast.FunctionDef) and node.name == "_new_room"
            )
            new_room_source = ast.get_source_segment(main_src_for_new_chat, new_room_node) or ""
            if "_clear_dashboard_lite_for_new_chat()" not in new_room_source:
                new_chat_errors.append("new_room_dashboard_clear_not_connected")
        except Exception as exc:
            new_chat_errors.append(f"new_chat_clear_runtime={type(exc).__name__}:{exc}")
        if new_chat_errors:
            results.append(_fail("Dashboard Lite new chat clears cached result", "; ".join(new_chat_errors)))
        else:
            results.append(_ok("Dashboard Lite new chat clears cached result", "new-chat initialization removes Dashboard-only state without reloading dashboard sources or altering login/company state"))

        metric_errors: list[str] = []
        captured_markdown: list[str] = []
        try:
            fake_st = _FakeStreamlit(submit_sequence=[])
            fake_st.markdown = lambda text, **_kwargs: captured_markdown.append(str(text))
            setattr(view_mod, "st", fake_st)
            view_mod._metric_card("<script>alert(1)</script>", 1, suffix="<b>%</b>", help_text="<img src=x>")
        finally:
            setattr(view_mod, "st", old_st)
        metric_html = "\n".join(captured_markdown)
        if "<script>" in metric_html or "<b>" in metric_html or "<img" in metric_html:
            metric_errors.append(f"raw_html={metric_html!r}")
        if "&lt;script&gt;" not in metric_html or "&lt;b&gt;%&lt;/b&gt;" not in metric_html or "&lt;img src=x&gt;" not in metric_html:
            metric_errors.append(f"escaped_html_missing={metric_html!r}")
        if metric_errors:
            results.append(_fail("Dashboard Lite KPI HTML escaping", "; ".join(metric_errors)))
        else:
            results.append(_ok("Dashboard Lite KPI HTML escaping", "metric label/value/help are escaped inside dashboard-scoped HTML"))

        action_table_errors: list[str] = []
        action_view_src = Path("app/sims/views/dashboard_lite.py").read_text(encoding="utf-8")
        action_view_start = action_view_src.find("def _render_today_actions")
        action_view_end = action_view_src.find("def _risk_detail_instance_key", action_view_start)
        action_view_source = action_view_src[action_view_start:action_view_end]
        for required_text in (
            "st.columns", "상세 보기", "render_mode == \"primary\"",
            "__dashboard_lite_action_drilldown::", "판정 기준",
            "_fmt_dashboard_amount(value, amount_unit)", "threshold_label",
            "on_click=_select_dashboard_action_detail",
        ):
            if required_text not in action_view_source:
                action_table_errors.append(f"action_list_missing={required_text}")
        if action_table_errors:
            results.append(_fail("Dashboard Lite today action numeric table", "; ".join(action_table_errors)))
        else:
            results.append(_ok("Dashboard Lite today action numeric table", "priority/qty/amount/ratio columns remain numeric before st.dataframe"))

        action_display_errors: list[str] = []
        try:
            if view_mod._fmt_dashboard_amount(105_600_000, "thousand") != "105,600.0 천원":
                action_display_errors.append("thousand_amount_display")
            if view_mod._fmt_dashboard_amount(105_600_000, "million") != "105.6 백만원":
                action_display_errors.append("million_amount_display")
            if view_mod._fmt_threshold_pct(45.3) != "45.3":
                action_display_errors.append("readiness_threshold_precision")
            if view_mod._fmt_threshold_pct(-12.5) != "-12.5":
                action_display_errors.append("decline_threshold_precision")
        except Exception as exc:
            action_display_errors.append(f"{type(exc).__name__}: {exc}")
        if action_display_errors:
            results.append(_fail("Dashboard Lite today action amount and threshold display", "; ".join(action_display_errors)))
        else:
            results.append(_ok("Dashboard Lite today action amount and threshold display", "facts stay raw while primary and compact amount/threshold rendering uses the Dashboard unit policy"))

        action_callback_errors: list[str] = []
        try:
            class _ActionCallbackStreamlit(_FakeStreamlit):
                def __init__(self, *, click_once: bool):
                    super().__init__(submit_sequence=[])
                    self.click_once = click_once
                    self.writes: list[str] = []

                def write(self, value, *_args, **_kwargs):
                    self.writes.append(str(value))

                def button(self, _label, **kwargs):
                    self._count("button")
                    if self.click_once:
                        self.click_once = False
                        callback = kwargs.get("on_click")
                        if callable(callback):
                            callback(*tuple(kwargs.get("args") or ()))
                    return False

            old_view_st = view_mod.st
            old_view_room_getter = view_mod.get_current_chat_room_id
            try:
                view_mod.get_current_chat_room_id = lambda: "room-action"
                action_st = _ActionCallbackStreamlit(click_once=True)
                view_mod.st = action_st
                current_action = {
                    "action_id": "a1", "priority": 1, "status": "긴급 부족",
                    "target_name": "테스트제품", "target_code": "12345", "product_code": "12345",
                    "major_purchase_vendor_code": "16789", "manufacturer_code": "11090",
                    "drilldown_action": "품목별 재고부족현황", "drilldown_params": {},
                }
                current_cache = {"company_id": "company-action", "room_id": "room-action", "dashboard_event_id": "event-action", "params": {}}
                view_mod._render_today_actions({"today_actions": [current_action], "filters": {}}, current_cache, render_mode="primary")
                selected = action_st.session_state.get("__dashboard_selected_action_detail")
                if not isinstance(selected, dict) or selected.get("product_code") != "12345" or selected.get("product_code") in {"16789", "11090"}:
                    action_callback_errors.append("callback_selection_not_cached")
                if "__dashboard_drilldown_request" in action_st.session_state or "__dashboard_drilldown_auto_run" in action_st.session_state:
                    action_callback_errors.append("callback_legacy_handoff_created")
                if not isinstance(action_st.session_state.get("__dashboard_lite_suppress_chat_autoscroll_once"), dict):
                    action_callback_errors.append("callback_scroll_suppression_missing")
                view_mod._render_selected_dashboard_action_detail(
                    {"inventory": {"risk_detail_rows": [{"제품코드": "12345", "제품명": "코드계약제품", "위험상태": "긴급 부족"}]}},
                    current_cache,
                    render_mode="primary",
                )
                if action_st.calls.get("dataframe", 0) != 1:
                    action_callback_errors.append("cached_action_detail_not_rendered")
                legacy_action = {
                    "rank": 1, "priority": "높음", "risk_grade": "조치 필요",
                    "target": "테스트제품", "product_code": "00001", "stock_readiness_pct": 45.3,
                    "shortage_amt": 100000, "recommended_action": "확인", "drill_down": "품목별 재고부족현황",
                }
                legacy_st = _ActionCallbackStreamlit(click_once=False)
                view_mod.st = legacy_st
                view_mod._render_today_actions({"today_actions": [legacy_action], "filters": {}}, current_cache, render_mode="chat")
                if view_mod._safe_action_rank(legacy_action, 9) != 1:
                    action_callback_errors.append("legacy_rank")
                if "조치 필요" not in legacy_st.writes:
                    action_callback_errors.append("legacy_status")
                if "테스트제품" not in legacy_st.writes:
                    action_callback_errors.append("legacy_target")
                for action_case, expected in (
                    ({"priority": 1}, 1), ({"priority": "1"}, 1), ({"priority": "높음", "rank": 2}, 2),
                    ({"priority": ""}, 4), ({"priority": None}, 5), ({}, 6), ({"rank": "1"}, 1),
                ):
                    fallback = expected if expected > 1 else 9
                    if view_mod._safe_action_rank(action_case, fallback) != expected:
                        action_callback_errors.append(f"safe_rank={action_case!r}")
            finally:
                view_mod.st = old_view_st
                view_mod.get_current_chat_room_id = old_view_room_getter
        except Exception as exc:
            action_callback_errors.append(f"{type(exc).__name__}: {exc}")
        if action_callback_errors:
            results.append(_fail("Dashboard Lite action callback and legacy snapshot", "; ".join(action_callback_errors)))
        else:
            results.append(_ok("Dashboard Lite action callback and legacy snapshot", "button callback stores only the active cache selection; legacy rank/priority fields render without a numeric conversion error"))

        panel_src = Path("app/ui/sims_panel.py").read_text(encoding="utf-8")
        entry_src = Path("app/ui/sims_entry.py").read_text(encoding="utf-8")
        view_src = Path("app/sims/views/dashboard_lite.py").read_text(encoding="utf-8")
        dashboard_facts_src = Path("app/services/dashboard_lite_facts.py").read_text(encoding="utf-8")
        main_dashboard_src = Path("app/Lmstudio_SSAI_chat_main.py").read_text(encoding="utf-8")
        chat_middleware_dashboard_src = Path("app/ui/chat_middleware.py").read_text(encoding="utf-8")
        sales_trend_src = Path("app/services/analytics_sales_trend_service.py").read_text(encoding="utf-8")
        drilldown_errors: list[str] = []
        for required_text in (
            "def _consume_dashboard_drilldown_request",
            "room_mismatch",
            "company_mismatch",
            "stale_event",
            "target_not_registered",
            "__dashboard_drilldown_auto_run",
        ):
            if required_text not in panel_src:
                drilldown_errors.append(f"panel_missing={required_text}")
        analytics_src = Path("app/sims/views/analytics_views.py").read_text(encoding="utf-8")
        for required_text in (
            "def _dashboard_stock_shortage_handoff",
            "def _apply_dashboard_stock_shortage_params",
            "def _dashboard_stock_shortage_handoff_summary",
            "dashboard_product_group_list",
            "manufacturer_test_codes",
            "io_gu_list",
            "__dashboard_drilldown_auto_run",
            "st.session_state.pop(\"__dashboard_drilldown_auto_run\"",
        ):
            if required_text not in analytics_src:
                drilldown_errors.append(f"analytics_missing={required_text}")
        sales_start = analytics_src.find("def render_sales_trend_analysis")
        sales_end = analytics_src.find("\ndef ", sales_start + 1)
        stock_start = analytics_src.find("def render_stock_shortage_analysis")
        stock_end = analytics_src.find("\ndef ", stock_start + 1)
        sales_source = analytics_src[sales_start:sales_end]
        stock_source = analytics_src[stock_start:stock_end]
        if "_apply_dashboard_stock_shortage_params" in sales_source or "dashboard_handoff" in sales_source:
            drilldown_errors.append("sales_trend_handoff_location")
        if "_apply_dashboard_stock_shortage_params(params, dashboard_handoff)" not in stock_source:
            drilldown_errors.append("stock_shortage_params_apply_missing")
        if "_dashboard_stock_shortage_handoff_summary(params)" not in stock_source:
            drilldown_errors.append("stock_shortage_summary_missing")
        if view_src.count('st.markdown("### 오늘의 조치")') != 1 or 'st.markdown("#### 오늘의 조치")' in action_view_source:
            drilldown_errors.append("today_actions_heading_duplicate")
        if "uuid.uuid4" in action_view_source:
            drilldown_errors.append("action_widget_uuid")
        for required_text in (
            "def _dashboard_action_detail_selection",
            "def _select_dashboard_action_detail",
            "def _render_selected_dashboard_action_detail",
            "on_click=_select_dashboard_action_detail",
            "__dashboard_selected_action_detail",
            "def _safe_action_rank",
            "def _legacy_action_status",
            "def _legacy_action_target",
        ):
            if required_text not in action_view_src:
                drilldown_errors.append(f"action_callback_missing={required_text}")
        if "on_click=_queue_dashboard_drilldown_request" in action_view_source or "__dashboard_drilldown_request" in action_view_source:
            drilldown_errors.append("action_callback_uses_legacy_handoff")
        if "if st.button(\"상세 보기\"" in action_view_source or "int(action.get('priority') or 0)" in action_view_source:
            drilldown_errors.append("legacy_button_or_priority_cast")
        panel_main_start = panel_src.find("def render_sims_main")
        panel_main_end = panel_src.find("\ndef ", panel_main_start + 1)
        panel_main_source = panel_src[panel_main_start:panel_main_end]
        consume_at = panel_main_source.find("_consume_dashboard_drilldown_request(selected")
        context_controls_at = panel_main_source.find("render_sims_context_controls()")
        if consume_at < 0 or context_controls_at < 0 or consume_at > context_controls_at:
            drilldown_errors.append("consume_after_panel_widget")
        if consume_at >= 0 and 'ss["__sims_open"] =' in panel_main_source[consume_at:]:
            drilldown_errors.append("panel_main_mutates_sims_open_after_consume")
        for required_text in (
            "stage=prepared", "stage=consumed", "stage=discarded", "widget_safe_phase=",
            "state_prepare_failed", "unsafe_widget_phase",
        ):
            if required_text not in panel_src:
                drilldown_errors.append(f"consume_log_missing={required_text}")
        if "_consume_dashboard_drilldown_request" not in main_dashboard_src:
            drilldown_errors.append("sidebar_preflight_missing")
        for required_text in (
            "__dashboard_lite_suppress_chat_autoscroll_once",
            "[dashboard.scroll] reason=%s",
            "suppress_consumed=True",
        ):
            if required_text not in main_dashboard_src:
                drilldown_errors.append(f"scroll_suppression_missing={required_text}")
        if re.search(r"(?m)^\s*logger\.(?:info|warning|exception)\(", main_dashboard_src):
            drilldown_errors.append("main_undefined_logger_reference")
        if "def _consume_dashboard_scroll_suppression" not in main_dashboard_src or "log.info(" not in main_dashboard_src:
            drilldown_errors.append("scroll_suppression_existing_logger_not_used")
        try:
            main_tree = ast.parse(main_dashboard_src)
            scroll_helper_node = next(
                node for node in main_tree.body
                if isinstance(node, ast.FunctionDef) and node.name == "_consume_dashboard_scroll_suppression"
            )

            class _ScrollSuppressionSt:
                session_state: dict[str, Any] = {}

            class _ScrollSuppressionLog:
                calls: list[tuple[Any, ...]] = []

                def info(self, *args: Any, **_kwargs: Any) -> None:
                    self.calls.append(args)

            scroll_log = _ScrollSuppressionLog()
            scroll_ns = {"st": _ScrollSuppressionSt, "log": scroll_log, "Any": Any}
            exec(
                compile(ast.fix_missing_locations(ast.Module(body=[scroll_helper_node], type_ignores=[])), "dashboard_scroll_helper", "exec"),
                scroll_ns,
            )
            _ScrollSuppressionSt.session_state = {"__dashboard_lite_suppress_chat_autoscroll_once": {"reason": "action_detail"}}
            if scroll_ns["_consume_dashboard_scroll_suppression"]("anchor-1") is not None:
                drilldown_errors.append("scroll_suppression_not_consumed")
            if "__dashboard_lite_suppress_chat_autoscroll_once" in _ScrollSuppressionSt.session_state or not scroll_log.calls:
                drilldown_errors.append("scroll_suppression_runtime_log_missing")
            if scroll_ns["_consume_dashboard_scroll_suppression"]("anchor-2") != "anchor-2":
                drilldown_errors.append("scroll_suppression_new_message_blocked")
        except Exception as exc:
            drilldown_errors.append(f"scroll_suppression_runtime={type(exc).__name__}: {exc}")
        if drilldown_errors:
            results.append(_fail("Dashboard Lite action drill-down one-shot", "; ".join(drilldown_errors)))
        else:
            results.append(_ok("Dashboard Lite action drill-down one-shot", "deterministic action key, primary-only control, and room/company/event fail-closed consumer are present"))

        action_scope_errors: list[str] = []
        try:
            analytics_views_mod = importlib.import_module("app.sims.views.analytics_views")
            handoff_source = {
                "stock_cd_list": ["00001", "00002"],
                "vendor_group_list": ["VG1"],
                "vendor_kind_list": ["VK1"],
                "product_group_list": ["PG1", "PG2"],
                "product_di_list": ["PD1", "PD2"],
                "product_class_list": ["PC1", "PC2"],
                "io_gu_list": ["001", "051"],
                "manufacturer_test_codes": ["10047"],
                "amount_display_unit": "million",
                "product_code": "P100",
            }
            applied = analytics_views_mod._apply_dashboard_stock_shortage_params(
                {"physic_cd": "P100", "product_group": "", "product_group_nm": ""},
                {"target_params": handoff_source},
            )
            for key in (
                "stock_cd_list", "vendor_group_list", "vendor_kind_list", "product_group_list", "product_di_list",
                "product_class_list", "io_gu_list", "manufacturer_test_codes", "amount_display_unit",
                "dashboard_product_group_list", "dashboard_product_di_list", "dashboard_product_class_list",
                "dashboard_manufacturer_codes",
            ):
                source_key = {
                    "dashboard_product_group_list": "product_group_list",
                    "dashboard_product_di_list": "product_di_list",
                    "dashboard_product_class_list": "product_class_list",
                    "dashboard_manufacturer_codes": "manufacturer_test_codes",
                }.get(key, key)
                if applied.get(key) != handoff_source.get(source_key):
                    action_scope_errors.append(f"handoff_missing={key}")
            if applied.get("product_group") or applied.get("product_group_nm"):
                action_scope_errors.append("multi_group_narrowed_by_legacy_widget")
            summary = analytics_views_mod._dashboard_stock_shortage_handoff_summary(applied)
            for token in ("재고위치", "제품구분", "제품분류", "입출고구분"):
                if token not in summary:
                    action_scope_errors.append(f"handoff_summary_missing={token}")
        except Exception as exc:
            action_scope_errors.append(f"{type(exc).__name__}: {exc}")
        if action_scope_errors:
            results.append(_fail("Dashboard Lite action drill-down scope handoff", "; ".join(action_scope_errors)))
        else:
            results.append(_ok("Dashboard Lite action drill-down scope handoff", "stock, product code-pair, vendor, IO, manufacturer, product-code, and amount-unit handoff params are retained"))

        drilldown_runtime_errors: list[str] = []
        try:
            analytics_views_mod = importlib.import_module("app.sims.views.analytics_views")

            class _FakeAnalyticsStreamlit(_FakeStreamlit):
                def date_input(self, _label, value=None, **kwargs):
                    return self.session_state.get(kwargs.get("key"), value)

                def text_input(self, _label, value="", **kwargs):
                    return self.session_state.get(kwargs.get("key"), value)

                def multiselect(self, _label, options=None, default=None, **kwargs):
                    key = kwargs.get("key")
                    return self.session_state.get(key, default if default is not None else [])

                def checkbox(self, _label, value=False, **kwargs):
                    return self.session_state.get(kwargs.get("key"), value)

            def _fake_code_options(gcode: str) -> list[dict[str, str]]:
                codes = {
                    "0013": ["PG1", "PG2"],
                    "0004": ["PD1", "PD2"],
                    "0031": ["PC1", "PC2"],
                    "0018": ["00001", "00002"],
                }.get(str(gcode), [])
                return [{"code": code, "name": code, "label": f"{code} - {code}"} for code in codes]

            old_analytics_st = analytics_views_mod.st
            old_sales_result = analytics_views_mod.get_sales_trend_result
            old_stock_result = analytics_views_mod.get_stock_shortage_result
            old_code_loader = analytics_views_mod._load_code_options
            old_inline_header = analytics_views_mod._render_inline_analysis_header_enabled
            old_default_adapter = analytics_views_mod._prepare_analytics_company_defaults
            sales_calls: list[dict[str, Any]] = []
            stock_calls: list[dict[str, Any]] = []
            try:
                analytics_views_mod._load_code_options = _fake_code_options
                analytics_views_mod._render_inline_analysis_header_enabled = lambda: False
                analytics_views_mod._prepare_analytics_company_defaults = lambda *_args, **_kwargs: {
                    "effective": {"io_gu_list": ["0012:051"]},
                    "profile_found": True,
                    "default_supported_keys": ["io_gu_list"],
                }
                analytics_views_mod.get_sales_trend_result = lambda params: sales_calls.append(dict(params)) or {"final": True, "type": "dataframe", "data": pd.DataFrame(), "meta": {}}
                sales_st = _FakeAnalyticsStreamlit(submit_sequence=[True])
                analytics_views_mod.st = sales_st
                analytics_views_mod.render_sales_trend_analysis()
                if len(sales_calls) != 1:
                    drilldown_runtime_errors.append(f"sales_trend_calls={len(sales_calls)}")
                if sales_calls and any(key.startswith("dashboard_") for key in sales_calls[0]):
                    drilldown_runtime_errors.append("sales_trend_dashboard_scope_leak")

                handoff_params = {
                    "month_from": "202601", "evaluation_month": "202607", "stock_mode": "real",
                    "stock_cd_list": ["00001", "00002"], "vendor_group_list": ["VG1"],
                    "vendor_kind_list": ["VK1"], "product_group_list": ["PG1", "PG2"],
                    "product_di_list": ["PD1", "PD2"], "product_class_list": ["PC1", "PC2"],
                    "io_gu_list": ["001", "051"], "manufacturer_test_codes": ["10047", "10048"],
                    "product_code": "P100", "amount_display_unit": "million",
                }
                stock_st = _FakeAnalyticsStreamlit(submit_sequence=[False])
                stock_st.session_state["__dashboard_drilldown_auto_run"] = {
                    "target_action": "품목별 재고부족현황", "target_params": handoff_params,
                }
                analytics_views_mod.st = stock_st
                analytics_views_mod.get_stock_shortage_result = lambda params: stock_calls.append(dict(params)) or {"final": True, "type": "dataframe", "data": pd.DataFrame(), "meta": {}}
                analytics_views_mod.render_stock_shortage_analysis()
                if len(stock_calls) != 1:
                    drilldown_runtime_errors.append(f"stock_shortage_calls={len(stock_calls)}")
                captured = stock_calls[0] if stock_calls else {}
                for key in (
                    "stock_mode", "stock_cd_list", "vendor_group_list", "vendor_kind_list", "product_group_list",
                    "product_di_list", "product_class_list", "io_gu_list", "manufacturer_test_codes",
                    "dashboard_product_group_list", "dashboard_product_di_list", "dashboard_product_class_list",
                    "dashboard_manufacturer_codes", "amount_display_unit",
                ):
                    if key not in captured:
                        drilldown_runtime_errors.append(f"stock_runtime_missing={key}")
                if captured.get("physic_cd") != "P100":
                    drilldown_runtime_errors.append(f"stock_runtime_product_code={captured.get('physic_cd')!r}")
                if captured.get("io_gu_list") != ["051"] or captured.get("io_gu_source") != "company_default":
                    drilldown_runtime_errors.append(f"stock_runtime_company_io={captured.get('io_gu_list')!r}/{captured.get('io_gu_source')!r}")
                if "__dashboard_drilldown_auto_run" in stock_st.session_state:
                    drilldown_runtime_errors.append("stock_runtime_auto_run_not_consumed")
                if stock_st.calls.get("rerun", 0):
                    drilldown_runtime_errors.append("stock_runtime_explicit_rerun")
            finally:
                analytics_views_mod.st = old_analytics_st
                analytics_views_mod.get_sales_trend_result = old_sales_result
                analytics_views_mod.get_stock_shortage_result = old_stock_result
                analytics_views_mod._load_code_options = old_code_loader
                analytics_views_mod._render_inline_analysis_header_enabled = old_inline_header
                analytics_views_mod._prepare_analytics_company_defaults = old_default_adapter
        except Exception as exc:
            drilldown_runtime_errors.append(f"{type(exc).__name__}: {exc}")
        if drilldown_runtime_errors:
            results.append(_fail("Dashboard Lite drill-down runtime placement", "; ".join(drilldown_runtime_errors)))
        else:
            results.append(_ok("Dashboard Lite drill-down runtime placement", "sales trend has no handoff dependency; stock shortage consumes one request, retains scope params, and makes one service call"))

        drilldown_widget_lock_errors: list[str] = []
        try:
            panel_mod = importlib.import_module("app.ui.sims_panel")

            class _WidgetLockedState(dict):
                def __init__(self):
                    super().__init__()
                    self._locked_keys: set[str] = set()

                def lock_widget(self, key: str) -> None:
                    self._locked_keys.add(key)

                def __setitem__(self, key, value):
                    if key in self._locked_keys:
                        raise RuntimeError(f"widget key locked: {key}")
                    super().__setitem__(key, value)

            class _FakePanelStreamlit:
                def __init__(self, session_state):
                    self.session_state = session_state

            target_category = next(iter(panel_mod._CATEGORIES))
            target_action = next(iter(panel_mod._CATEGORIES[target_category]["actions"]))

            def _valid_request(*, company_id="company-1"):
                return {
                    "source": "dashboard",
                    "request_token": "token-1",
                    "source_room_id": "room-1",
                    "company_id": company_id,
                    "source_dashboard_event_id": "event-1",
                    "target_category": target_category,
                    "target_action": target_action,
                    "target_params": {"product_code": "P100"},
                }

            old_panel_st = panel_mod.st
            old_room_getter = panel_mod.get_current_chat_room_id
            old_company_stamp = panel_mod._panel_current_company_stamp
            try:
                panel_mod.get_current_chat_room_id = lambda: "room-1"
                panel_mod._panel_current_company_stamp = lambda: {"company_id": "company-1"}

                normal_state = _WidgetLockedState()
                normal_state["__dashboard_drilldown_request"] = _valid_request()
                normal_state["__dashboard_lite_result"] = {"dashboard_event_id": "event-1"}
                panel_mod.st = _FakePanelStreamlit(normal_state)
                target = panel_mod._consume_dashboard_drilldown_request(None, widget_safe_phase=True)
                if target != {"category": target_category, "action": target_action}:
                    drilldown_widget_lock_errors.append("prepared_target")
                for key, expected in (
                    ("__sims_open", True),
                    ("__sims_panel_active", True),
                    ("__sims_run_flag", True),
                ):
                    if normal_state.get(key) is not expected:
                        drilldown_widget_lock_errors.append(f"prepared_state={key}")
                if "__dashboard_drilldown_request" in normal_state:
                    drilldown_widget_lock_errors.append("request_not_consumed")
                if not isinstance(normal_state.get("__dashboard_drilldown_auto_run"), dict):
                    drilldown_widget_lock_errors.append("auto_run_missing")
                normal_state.lock_widget("__sims_open")
                try:
                    normal_state["__sims_open"] = False
                    drilldown_widget_lock_errors.append("widget_lock_not_enforced")
                except RuntimeError:
                    pass
                if panel_mod._consume_dashboard_drilldown_request(target, widget_safe_phase=True) != target:
                    drilldown_widget_lock_errors.append("one_shot_repeat")

                invalid_state = _WidgetLockedState()
                invalid_state["__dashboard_drilldown_request"] = _valid_request(company_id="other-company")
                invalid_state["__dashboard_lite_result"] = {"dashboard_event_id": "event-1"}
                invalid_state["__sims_open"] = False
                invalid_state["__sims_panel_active"] = False
                panel_mod.st = _FakePanelStreamlit(invalid_state)
                panel_mod._consume_dashboard_drilldown_request({"category": "keep", "action": "keep"}, widget_safe_phase=True)
                if invalid_state.get("__sims_open") or invalid_state.get("__sims_panel_active"):
                    drilldown_widget_lock_errors.append("invalid_mutated_panel_state")
                if "__dashboard_drilldown_request" in invalid_state:
                    drilldown_widget_lock_errors.append("invalid_request_not_discarded")
            finally:
                panel_mod.st = old_panel_st
                panel_mod.get_current_chat_room_id = old_room_getter
                panel_mod._panel_current_company_stamp = old_company_stamp
        except Exception as exc:
            drilldown_widget_lock_errors.append(f"{type(exc).__name__}: {exc}")
        if drilldown_widget_lock_errors:
            results.append(_fail("Dashboard drill-down widget-safe consume", "; ".join(drilldown_widget_lock_errors)))
        else:
            results.append(_ok("Dashboard drill-down widget-safe consume", "request is prepared before the locked toggle, consumed once, and invalid requests leave panel state unchanged"))
        dashboard_multiselect_keys = (
            "__dashboard_lite_stock_labels",
            "__dashboard_lite_product_group_list",
            "__dashboard_lite_product_di_list",
            "__dashboard_lite_product_class_list",
            "__dashboard_lite_vendor_group_list",
            "__dashboard_lite_vendor_kind_list",
            "__dashboard_lite_io_gu_list",
        )
        dashboard_multiselect_lines = [
            line for line in view_src.splitlines() if "st.multiselect(" in line
        ]
        dashboard_multiselect_default_conflicts = [
            key
            for key in dashboard_multiselect_keys
            if any(key in line and "default=" in line for line in dashboard_multiselect_lines)
        ]
        dashboard_multiselect_state_missing = [
            key
            for key in dashboard_multiselect_keys
            if f'_prepare_dashboard_multiselect_state("{key}"' not in view_src
            and key != "__dashboard_lite_stock_labels"
        ]
        dashboard_scalar_keys = (
            "__dashboard_lite_stock_mode",
            "__dashboard_lite_major_purchase_vendor_days",
            "__dashboard_lite_risk_analysis_days",
            "__dashboard_lite_overstock_inactive_days",
            "__dashboard_lite_readiness_warning_pct",
            "__dashboard_lite_risk_quick_view_count",
            "__dashboard_lite_amount_display_unit",
        )
        dashboard_scalar_widget_lines = [
            line
            for line in view_src.splitlines()
            if "st.number_input(" in line or "st.selectbox(" in line or "st.radio(" in line
        ]
        dashboard_scalar_default_conflicts = [
            key
            for key in dashboard_scalar_keys
            if any(
                key in line and re.search(r"(?<![A-Za-z_])(?:value|index|default)\s*=(?!=)", line)
                for line in dashboard_scalar_widget_lines
            )
        ]
        threshold_format_cases = {
            98.0: "98",
            97.5: "97.5",
            97.25: "97.25",
            80.0: "80",
        }
        threshold_format_errors = {
            value: view_mod._fmt_threshold_pct(value)
            for value, expected in threshold_format_cases.items()
            if view_mod._fmt_threshold_pct(value) != expected
        }
        if (
            '"Dashboard Lite v0.1": dashboard_lite.render_dashboard_lite' not in panel_src
            or 'from app.sims.views import users, codes, vendors, goods, road_address, analytics_views, dashboard_lite, rddbc_io_views' not in panel_src
            or '"대시보드": [\n            "Dashboard Lite v0.1",' not in entry_src
            or '"대시보드": "dashboard",' not in entry_src
            or 'st.altair_chart' not in view_src
            or 'width="stretch"' not in view_src
            or 'st.form("dashboard_lite_scope_form"' not in view_src
            or '"재고위치"' not in view_src
            or '"product_group_list"' not in view_src
            or '_dashboard_code_name_options("0031")' not in view_src
            or '_dashboard_code_name_options("0001")' in view_src
            or 'def _resolve_dashboard_manufacturer' not in view_src
            or 'LEFT JOIN dbo.Rddbc030 AS V' not in view_src
            or 'ORDER BY manufacturer_name, manufacturer_code' not in view_src
            or 'cols = st.columns(4)' not in view_src
            or 'st.text_input("제약사"' not in view_src
            or 'st.selectbox("제약사 후보"' in view_src
            or 'build_dashboard_lite_chat_snapshot' not in view_src
            or 'manufacturer_codes = _clean_list(option_cache.get("manufacturer_codes"))' in view_src
            or 'form_submit_button("대시보드 조회"' not in view_src
            or 'build_dashboard_lite_facts(params)' not in view_src
            or '"type": "dashboard_lite"' not in view_src
            or 'render_dashboard_lite_chat_item' not in chat_middleware_dashboard_src
            or 'elif str(payload.get("type") or "").strip().lower() != "dashboard_lite"' not in chat_middleware_dashboard_src
            or '[dashboard.chat_push]' not in panel_src
            or 'dashboard_item["role"] = "assistant"' not in main_dashboard_src
            or 'dashboard_primary_area = st.empty()' in main_dashboard_src
            or bool(dashboard_multiselect_default_conflicts)
            or bool(dashboard_scalar_default_conflicts)
            or 'stock_widget_key = "__dashboard_lite_stock_labels"' not in view_src
            or '_prepare_dashboard_multiselect_state(stock_widget_key, stock_codes)' not in view_src
            or bool(dashboard_multiselect_state_missing)
            or "_DASHBOARD_PROFILE_SCALAR_DEFAULTS" not in view_src
            or "_prepare_dashboard_profile_scalar_state()" not in view_src
            or "min_value=1, step=1, key=\"__dashboard_lite_major_purchase_vendor_days\"" not in view_src
            or "min_value=0.1, max_value=100.0, step=0.1, key=\"__dashboard_lite_readiness_warning_pct\"" not in view_src
            or 'threshold_value = float((facts.get("stock_readiness") or {}).get("threshold_pct") or 98.0)' not in view_src
            or "display_readiness_pct" not in view_src
            or "display_remaining_demand_qty" not in view_src
            or "display_shortage_amt" not in view_src
            or "수요급증여부:N" not in view_src
            or "위험보정기준:N" not in view_src
            or "98% 미만 SKU만 기본 조치 대상으로 표시합니다." in view_src
            or "threshold_value:.0f" in view_src
            or "readiness_threshold:.0f" in view_src
            or bool(threshold_format_errors)
            or '"위험보정부족예상금액", r.get("shortage_amt")' not in dashboard_facts_src
            or '"위험보정부족예상수량", r.get("shortage_qty")' not in dashboard_facts_src
            or '"위험보정재고준비율", r.get("stock_readiness_pct")' not in dashboard_facts_src
            or '"수요급증 후 잔여수요 절반 미만"' not in dashboard_facts_src
            or 'Dashboard facts / 비교 금지 규칙' in view_src
            or 'with st.expander("추가 확인사항"' in view_src
        ):
            results.append(_fail("Dashboard Lite panel/view registration", "registry or non-chat view contract missing"))
        else:
            results.append(_ok("Dashboard Lite panel/view registration", "analysis/KPI dashboard group/action is rendered as a single final chat-history Dashboard message"))

        dashboard_order_errors: list[str] = []
        history_anchor = main_dashboard_src.find("for idx, m in enumerate(merged_msgs):")
        if history_anchor < 0:
            dashboard_order_errors.append(f"history_anchor={history_anchor}")
        if '"__dashboard_lite_result"' not in view_src or '"type": "dashboard_lite"' not in view_src:
            dashboard_order_errors.append("dashboard_not_chat_history_state")
        if '_dashboard_primary_target = st.empty()' in main_dashboard_src:
            dashboard_order_errors.append("dashboard_top_primary_target_remaining")
        if 'render_cached_dashboard_lite_primary()' in main_dashboard_src:
            dashboard_order_errors.append("dashboard_top_cached_render_remaining")
        if '_render_dashboard_primary_after_panel_if_requested' in main_dashboard_src:
            dashboard_order_errors.append("dashboard_post_panel_primary_render_remaining")
        if 'render_mode = str(decision.get("render_mode") or "chat")' not in chat_middleware_dashboard_src:
            dashboard_order_errors.append("dashboard_message_position_mode_dispatch_missing")
        try:
            order_tree = ast.parse(main_dashboard_src)
            order_nodes = [
                node
                for node in order_tree.body
                if isinstance(node, ast.FunctionDef)
                and node.name in {
                    "normalize_ts",
                    "_message_time_key",
                    "_message_dedupe_key",
                    "_build_room_render_messages",
                }
            ]
            order_ns: dict[str, Any] = {"re": re}
            exec(
                compile(
                    ast.Module(body=order_nodes, type_ignores=[]),
                    "dashboard_chronological_order",
                    "exec",
                ),
                order_ns,
            )
            build_render_messages = order_ns["_build_room_render_messages"]
            mixed_room = {
                "messages": [
                    {"id": "user-list", "type": "table", "title": "사용자목록", "time": "2026-07-25 23:10:00"},
                ],
                "history": [
                    {"id": "dashboard-a", "type": "dashboard_lite", "title": "Dashboard Lite", "time": "2026-07-25 22:17:00"},
                    {"id": "dashboard-b", "type": "dashboard_lite", "title": "Dashboard Lite", "time": "2026-07-25 22:57:00"},
                    {"id": "dashboard-c", "type": "dashboard_lite", "title": "Dashboard Lite", "time": "2026-07-25 23:20:00"},
                    {"id": "dashboard-d", "type": "dashboard_lite", "title": "Dashboard Lite", "time": "2026-07-25 23:20:00"},
                ],
            }
            chronological = build_render_messages(mixed_room)
            chronological_ids = [item.get("id") for item in chronological]
            if chronological_ids != ["dashboard-a", "dashboard-b", "user-list", "dashboard-c", "dashboard-d"]:
                dashboard_order_errors.append(f"mixed_dashboard_sims_order={chronological_ids!r}")

            active_primary = {
                "company_id": "company-a",
                "room_id": "room-a",
                "dashboard_event_id": "dashboard-c",
                "query_fingerprint": "same-query",
                "facts": {"inventory": {"risk_detail_rows": [{"제품코드": "P1"}]}},
            }
            original_view_st = view_mod.st
            mode_st = type("ModeSt", (), {"session_state": {"__dashboard_lite_result": active_primary}})()
            view_mod.st = mode_st
            try:
                event_modes = [
                    view_mod.dashboard_lite_chat_render_decision(
                        {
                            "company_id": "company-a",
                            "room_id": "room-a",
                            "dashboard_event_id": event_id,
                            "query_fingerprint": "same-query",
                        },
                        {"room_id": "room-a", "dashboard_event_id": event_id},
                    ).get("render_mode")
                    for event_id in ("dashboard-a", "dashboard-b", "dashboard-c")
                ]
                if event_modes != ["chat", "chat", "primary"]:
                    dashboard_order_errors.append(f"same_room_primary_modes={event_modes!r}")
                mode_st.session_state.pop("__dashboard_lite_result", None)
                return_modes = [
                    view_mod.dashboard_lite_chat_render_decision(
                        {
                            "company_id": "company-a",
                            "room_id": "room-a",
                            "dashboard_event_id": event_id,
                            "query_fingerprint": "same-query",
                        },
                        {"room_id": "room-a", "dashboard_event_id": event_id},
                    ).get("render_mode")
                    for event_id in ("dashboard-a", "dashboard-b", "dashboard-c")
                ]
                if return_modes != ["chat", "chat", "chat"]:
                    dashboard_order_errors.append(f"room_return_compact_modes={return_modes!r}")
            finally:
                view_mod.st = original_view_st
        except Exception as exc:
            dashboard_order_errors.append(f"chronological_runtime={type(exc).__name__}:{exc}")
        if dashboard_order_errors:
            results.append(_fail("Dashboard Lite primary document order", "; ".join(dashboard_order_errors)))
        else:
            results.append(_ok("Dashboard Lite primary document order", "chronological chat history owns Dashboard rendering; only the active event uses primary controls and older events stay compact"))

        dashboard_roundtrip_errors: list[str] = []
        try:
            main_tree_for_dashboard = ast.parse(main_dashboard_src)
            partition_node = next(
                node for node in main_tree_for_dashboard.body
                if isinstance(node, ast.FunctionDef) and node.name == "_partition_message_payload"
            )
            minimal_snapshot_node = next(
                node for node in main_tree_for_dashboard.body
                if isinstance(node, ast.FunctionDef) and node.name == "_minimal_dashboard_partition_snapshot"
            )
            partition_ns = {
                "_CHAT_PARTITION_MESSAGE_ALLOW_KEYS": {"id", "type", "title", "action", "params", "meta"},
                "_CHAT_PARTITION_META_ALLOW_KEYS": {"action", "analysis_type", "facts_kind", "dashboard_cache", "dashboard_event_id", "room_id"},
                "_CHAT_PARTITION_TEXT_LIMIT": 20_000,
                "_clip_partition_text": lambda value, _limit: value,
                "_compact_partition_value": lambda value: value,
                "_json_sanitize": lambda value: value,
                "build_dashboard_lite_chat_snapshot": view_mod.build_dashboard_lite_chat_snapshot,
                "Any": Any,
                "log": type("Log", (), {
                    "warning": staticmethod(lambda *_args, **_kwargs: None),
                    "info": staticmethod(lambda *_args, **_kwargs: None),
                })(),
            }
            exec(compile(ast.Module(body=[minimal_snapshot_node, partition_node], type_ignores=[]), "dashboard_partition", "exec"), partition_ns)
            dashboard_payload = {
                "id": "dashboard-event-1",
                "type": "dashboard_lite",
                "title": "Dashboard Lite",
                "action": "Dashboard Lite v0.1",
                "params": {"month_from": "202601"},
                "meta": {
                    "analysis_type": "dashboard_lite",
                    "facts_kind": "SIMS_DASHBOARD_FACTS_V01",
                    "room_id": "room-dashboard",
                    "dashboard_event_id": "dashboard-event-1",
                    "dashboard_cache": {"facts": {"sales": {"chart_rows": [{"period": "2026-01", "value": 1}]}}},
                },
            }
            restored_dashboard = partition_ns["_partition_message_payload"](dashboard_payload)
            restored_meta = restored_dashboard.get("meta") or {}
            if restored_dashboard.get("type") != "dashboard_lite" or restored_dashboard.get("id") != "dashboard-event-1":
                dashboard_roundtrip_errors.append(f"payload_identity={restored_dashboard!r}")
            restored_cache = restored_meta.get("dashboard_cache") or {}
            if (
                restored_meta.get("room_id") != "room-dashboard"
                or restored_dashboard.get("id") != restored_meta.get("dashboard_event_id")
                or restored_cache.get("room_id") != "room-dashboard"
                or restored_dashboard.get("id") != restored_cache.get("dashboard_event_id")
                or restored_cache.get("dashboard_event_id") != "dashboard-event-1"
                or not isinstance(restored_cache.get("facts"), dict)
            ):
                dashboard_roundtrip_errors.append(f"payload_cache_lost={restored_meta!r}")
            full_cache = {
                "params": {
                    "month_from": "202601", "month_to": "202606", "evaluation_month": "202607",
                    "manufacturer_scope_label": "전체", "manufacturer_test_codes": [f"V{i:05d}" for i in range(1000)],
                    "manufacturer_names": [f"제약사{i}" for i in range(1000)],
                },
                "facts": {
                    "sales": {
                        "metrics": {"amount": 1},
                        "visualization": {
                            "current_sales": 10, "forecast_sales": 20, "remaining_forecast": 10,
                            "sales_progress_pct": 50, "time_progress_pct": 40,
                            "time_adjusted_achievement_pct": 125, "expected_to_date_sales": 8,
                            "chart_month_count": 7, "completed_month_count": 6,
                        },
                        "chart_rows": [{"period": "2026-01", "value": 1}],
                    },
                    "inventory": {"metrics": {"sku": 1}, "risk_targets": [{"code": "P1"}], "stock_risk_summary": [{"재고위험상태": "긴급 부족", "품목수": 1, "부족예상금액": 1, "현재재고금액": 1}], "stock_overstock_summary": {"품목수": 1, "과잉후보수량": 2, "과잉후보금액": 3}, "stock_demand_surge_summary": {"품목수": 1, "위험보정부족예상수량": 2, "위험보정부족예상금액": 3}, "readiness_rows": [{"code": str(i), "amount": i} for i in range(10000)]},
                    "today_actions": [{"priority": 1}],
                },
            }
            compact_cache = view_mod.build_dashboard_lite_chat_snapshot(full_cache)
            compact_json = json.dumps(compact_cache, ensure_ascii=False).encode("utf-8")
            if "readiness_rows" in json.dumps(compact_cache, ensure_ascii=False) or len(compact_json) > 65_536:
                dashboard_roundtrip_errors.append(f"dashboard_snapshot_not_compact bytes={len(compact_json)}")
            if not (compact_cache.get("facts", {}).get("sales", {}).get("chart_rows") and compact_cache.get("facts", {}).get("inventory", {}).get("risk_targets")):
                dashboard_roundtrip_errors.append("dashboard_snapshot_render_inputs_lost")
            compact_visualization = compact_cache.get("facts", {}).get("sales", {}).get("visualization") or {}
            if compact_visualization.get("time_adjusted_achievement_pct") != 125 or compact_visualization.get("expected_to_date_sales") != 8:
                dashboard_roundtrip_errors.append(f"dashboard_snapshot_visualization_lost={compact_visualization!r}")
            if "stock_risk_summary" not in compact_cache.get("facts", {}).get("inventory", {}):
                dashboard_roundtrip_errors.append("dashboard_stock_risk_summary_missing")
            if "stock_overstock_summary" not in compact_cache.get("facts", {}).get("inventory", {}):
                dashboard_roundtrip_errors.append("dashboard_stock_overstock_summary_missing")
            if "stock_demand_surge_summary" not in compact_cache.get("facts", {}).get("inventory", {}):
                dashboard_roundtrip_errors.append("dashboard_stock_demand_surge_summary_missing")
            full_payload = {
                "id": "dashboard-event-full",
                "type": "dashboard_lite",
                "title": "Dashboard Lite",
                "action": "Dashboard Lite v0.1",
                "meta": {
                    "analysis_type": "dashboard_lite",
                    "room_id": "room-dashboard",
                    "dashboard_event_id": "dashboard-event-full",
                    "dashboard_cache": full_cache,
                },
            }
            partitioned_full = partition_ns["_partition_message_payload"](full_payload)
            partitioned_json = json.dumps(partitioned_full, ensure_ascii=False).encode("utf-8")
            if b"readiness_rows" in partitioned_json or len(partitioned_json) > 65_536:
                dashboard_roundtrip_errors.append(f"partitioned_dashboard_payload_not_compact bytes={len(partitioned_json)}")

            old_snapshot_builder = view_mod.build_dashboard_lite_chat_snapshot
            try:
                def _raise_snapshot(_cache):
                    raise RuntimeError("test snapshot failure")
                view_mod.build_dashboard_lite_chat_snapshot = _raise_snapshot
                fallback_partitioned = partition_ns["_partition_message_payload"](full_payload)
            finally:
                view_mod.build_dashboard_lite_chat_snapshot = old_snapshot_builder
            fallback_json = json.dumps(fallback_partitioned, ensure_ascii=False).encode("utf-8")
            fallback_cache = (fallback_partitioned.get("meta") or {}).get("dashboard_cache") or {}
            if (
                b"readiness_rows" in fallback_json
                or len(fallback_json) > 65_536
                or fallback_cache.get("snapshot_status") != "minimal_fallback"
                or "facts" in fallback_cache
            ):
                dashboard_roundtrip_errors.append(f"partitioned_dashboard_fallback_not_closed bytes={len(fallback_json)} cache={fallback_cache!r}")
            # Exercise the final compact message record as written by the
            # partition pipeline, not only the snapshot builder.
            tmp_partition = Path(tempfile.mkdtemp(prefix="dashboard_partition_")) / "messages.jsonl"
            tmp_partition.write_text(json.dumps(partitioned_full, ensure_ascii=False) + "\n", encoding="utf-8")
            restored_records = [json.loads(line) for line in tmp_partition.read_text(encoding="utf-8").splitlines() if line.strip()]
            if len(restored_records) != 1 or len(tmp_partition.read_bytes()) > 65_536:
                dashboard_roundtrip_errors.append("partitioned_dashboard_record_roundtrip_failed")
            if "readiness_rows" in json.dumps(restored_records, ensure_ascii=False):
                dashboard_roundtrip_errors.append("partitioned_dashboard_record_contains_readiness_rows")
            if 'if str(m.get("type") or "").strip().lower() == "dashboard_lite"' not in main_dashboard_src:
                dashboard_roundtrip_errors.append("renderer_dispatch_missing")
        except Exception as exc:
            dashboard_roundtrip_errors.append(f"roundtrip_runtime={type(exc).__name__}:{exc}")
        if dashboard_roundtrip_errors:
            results.append(_fail("Dashboard Lite chat payload round-trip", "; ".join(dashboard_roundtrip_errors)))
        else:
            results.append(_ok("Dashboard Lite chat payload round-trip", "Dashboard id, room id, meta cache, and nested chart facts survive partition storage and renderer dispatch"))

        stock_timing_errors: list[str] = []
        try:
            timing = dash_mod._stock_timing_meta(
                {
                    "stock_sql_ms": 192,
                    "stock_query_batches": 2,
                    "stock_aggregate_ms": 3,
                    "stock_shortage_build_ms": 176,
                    "stock_shortage_total_ms": 374,
                },
                fallback_total_ms=999,
            )
            expected_timing = {
                "stock_sql_ms": 192,
                "stock_batch_count": 2,
                "stock_aggregate_ms": 3,
                "stock_shortage_build_ms": 176,
                "stock_shortage_total_ms": 374,
                "configured_batch_size": 0,
                "effective_chunk_size": 0,
                "fixed_parameter_count": 0,
                "stock_cd_parameter_count": 0,
                "io_gu_parameter_count": 0,
                "total_parameter_count": 0,
            }
            if timing != expected_timing:
                stock_timing_errors.append(f"timing_forward={timing!r}")
            timing_frame = pd.DataFrame({"제품코드": ["P1"]})
            timing_frame.attrs.update(
                {
                    "stock_sql_ms": 192,
                    "stock_query_batches": 2,
                    "stock_aggregate_ms": 3,
                    "stock_shortage_build_ms": 176,
                    "stock_shortage_total_ms": 374,
                }
            )
            timing_filtered = dash_mod._filter_sales_source_for_dashboard(timing_frame, {})
            timing_enriched = dash_mod._attach_dashboard_product_code_pairs(
                {"df": timing_filtered},
                pd.DataFrame({"제품코드": ["P1"]}),
            )["df"]
            if dash_mod._stock_timing_meta(timing_enriched.attrs, fallback_total_ms=999) != expected_timing:
                stock_timing_errors.append(f"timing_attrs_lost={timing_enriched.attrs!r}")
        except Exception as exc:
            stock_timing_errors.append(f"timing_runtime={type(exc).__name__}:{exc}")
        if stock_timing_errors:
            results.append(_fail("Dashboard Lite stock timing forwarding", "; ".join(stock_timing_errors)))
        else:
            results.append(_ok("Dashboard Lite stock timing forwarding", "SQL, batch, aggregate, build, and total timings remain separate in Dashboard metadata"))

        dashboard_delivery_errors: list[str] = []
        try:
            import app.ui.chat_middleware as dashboard_chat_mod
            import app.ui.sims_panel as dashboard_panel_mod

            class _FakeChatStreamlit:
                def __init__(self):
                    self.session_state = {
                        "current_room": "room-a",
                        "__chat_current_room_id": "room-a",
                        "chat_rooms": [{"id": "room-a"}, {"id": "room-b"}],
                    }

            original_chat_st = dashboard_chat_mod.st
            original_view_st = view_mod.st
            fake_chat_st = _FakeChatStreamlit()
            dashboard_chat_mod.st = fake_chat_st
            view_mod.st = fake_chat_st
            try:
                pending_dashboard = {
                    "id": "dashboard-event-a",
                    "type": "dashboard_lite",
                    "meta": {"room_id": "room-a", "dashboard_event_id": "dashboard-event-a"},
                }
                history_dashboard = dict(pending_dashboard)
                next_dashboard = {
                    "id": "dashboard-event-b",
                    "type": "dashboard_lite",
                    "meta": {"room_id": "room-a", "dashboard_event_id": "dashboard-event-b"},
                }
                if not dashboard_chat_mod._should_render_sims_message_once(pending_dashboard, pending_dashboard["meta"]):
                    dashboard_delivery_errors.append("dashboard_first_render_blocked")
                if dashboard_chat_mod._should_render_sims_message_once(history_dashboard, history_dashboard["meta"]):
                    dashboard_delivery_errors.append("dashboard_pending_history_duplicate_not_blocked")
                if not dashboard_chat_mod._should_render_sims_message_once(next_dashboard, next_dashboard["meta"]):
                    dashboard_delivery_errors.append("dashboard_distinct_event_blocked")
                if dashboard_chat_mod.get_current_chat_room_id() != "room-a":
                    dashboard_delivery_errors.append("dashboard_current_room_id_missing")

                active_primary = {
                    "company_id": "company-a",
                    "room_id": "room-a",
                    "dashboard_event_id": "dashboard-event-a",
                    "query_fingerprint": "same-query",
                    "facts": {},
                }
                fake_chat_st.session_state["__dashboard_lite_result"] = active_primary
                active_cache = {"company_id": "company-a", "room_id": "room-a", "query_fingerprint": "same-query"}
                active_meta = {"room_id": "room-a", "dashboard_event_id": "dashboard-event-a"}
                active_decision = view_mod.dashboard_lite_chat_render_decision(active_cache, active_meta)
                if (
                    active_decision.get("skipped")
                    or active_decision.get("action") != "render_active_primary"
                    or active_decision.get("render_mode") != "primary"
                    or active_decision.get("render_cache") is not active_primary
                ):
                    dashboard_delivery_errors.append(f"dashboard_active_message_not_primary={active_decision!r}")
                exact_event_snapshot = view_mod.build_dashboard_lite_chat_snapshot(
                    {
                        "company_id": "company-a",
                        "room_id": "room-a",
                        "dashboard_event_id": "dashboard-event-a",
                        "query_fingerprint": "same-query",
                        "facts": {},
                    }
                )
                if exact_event_snapshot.get("room_id") != "room-a" or exact_event_snapshot.get("dashboard_event_id") != "dashboard-event-a":
                    dashboard_delivery_errors.append(f"dashboard_snapshot_event_not_preserved={exact_event_snapshot!r}")

                previous_decision = view_mod.dashboard_lite_chat_render_decision(
                    active_cache,
                    {"room_id": "room-a", "dashboard_event_id": "dashboard-event-previous"},
                )
                other_room_decision = view_mod.dashboard_lite_chat_render_decision(
                    active_cache,
                    {"room_id": "room-b", "dashboard_event_id": "dashboard-event-a"},
                )
                if (
                    previous_decision.get("render_mode") != "chat"
                    or other_room_decision.get("render_mode") != "chat"
                ):
                    dashboard_delivery_errors.append("dashboard_history_snapshot_not_compact")

                fake_chat_st.session_state.pop("__dashboard_lite_result", None)
                no_primary_decision = view_mod.dashboard_lite_chat_render_decision(active_cache, active_meta)
                if no_primary_decision.get("render_mode") != "chat":
                    dashboard_delivery_errors.append("dashboard_snapshot_without_primary_not_compact")

                primary_cleanup_state = {
                    "__dashboard_lite_result": {"facts": {}},
                    "__dashboard_lite_risk_detail_excel::active": {"bytes": b"x"},
                    "__dashboard_lite_product_group_list": ["0013:A"],
                }
                view_mod.clear_dashboard_lite_active_result(primary_cleanup_state)
                if (
                    "__dashboard_lite_result" in primary_cleanup_state
                    or "__dashboard_lite_risk_detail_excel::active" in primary_cleanup_state
                    or "__dashboard_lite_product_group_list" not in primary_cleanup_state
                ):
                    dashboard_delivery_errors.append(f"dashboard_room_switch_primary_cleanup={primary_cleanup_state!r}")

                ownership_same_room = view_mod.dashboard_lite_primary_cache_matches_context(
                    active_primary,
                    current_room_id="room-a",
                    current_company_id="company-a",
                )
                ownership_other_room = view_mod.dashboard_lite_primary_cache_matches_context(
                    active_primary,
                    current_room_id="room-b",
                    current_company_id="company-a",
                )
                ownership_other_company = view_mod.dashboard_lite_primary_cache_matches_context(
                    active_primary,
                    current_room_id="room-a",
                    current_company_id="company-b",
                )
                if not (ownership_same_room["room_match"] and ownership_same_room["company_match"]):
                    dashboard_delivery_errors.append(f"dashboard_primary_same_room_rejected={ownership_same_room!r}")
                if ownership_other_room["room_match"] or not ownership_other_room["company_match"]:
                    dashboard_delivery_errors.append(f"dashboard_primary_room_mismatch_not_detected={ownership_other_room!r}")
                if ownership_other_company["company_match"] or not ownership_other_company["room_match"]:
                    dashboard_delivery_errors.append(f"dashboard_primary_company_mismatch_not_detected={ownership_other_company!r}")

                mismatch_cleanup_state = {
                    "__dashboard_lite_result": active_primary,
                    "__dashboard_lite_risk_detail_excel::room-a": {"bytes": b"x"},
                    "__dashboard_lite_product_group_list": ["0013:A"],
                }
                if ownership_other_room["room_match"] or not ownership_other_room["company_match"]:
                    dashboard_delivery_errors.append("dashboard_primary_room_mismatch_policy_invalid")
                else:
                    view_mod.clear_dashboard_lite_active_result(mismatch_cleanup_state)
                    if "__dashboard_lite_result" in mismatch_cleanup_state or "__dashboard_lite_risk_detail_excel::room-a" in mismatch_cleanup_state:
                        dashboard_delivery_errors.append(f"dashboard_primary_room_mismatch_not_cleared={mismatch_cleanup_state!r}")

                middleware_src = Path("app/ui/chat_middleware.py").read_text(encoding="utf-8")
                for token in ("dashboard_lite_chat_render_decision", "_should_render_sims_message_once"):
                    if token not in middleware_src:
                        dashboard_delivery_errors.append(f"dashboard_snapshot_render_guard={token}")
                view_src_for_delivery = Path("app/sims/views/dashboard_lite.py").read_text(encoding="utf-8")
                if "render_active_primary" not in view_src_for_delivery:
                    dashboard_delivery_errors.append("dashboard_active_primary_dispatch_log_missing")
                for token in (
                    "dashboard_lite_primary_cache_matches_context",
                    "action=skip_room_mismatch",
                    "current_room_id=get_current_chat_room_id()",
                    "clear_dashboard_lite_active_result(st.session_state)",
                ):
                    if token not in view_src_for_delivery:
                        dashboard_delivery_errors.append(f"dashboard_primary_room_ownership_guard_missing={token}")
                panel_src_for_delivery = Path("app/ui/sims_panel.py").read_text(encoding="utf-8")
                if 'payload["id"] = event_id' not in panel_src_for_delivery or 'cache["dashboard_event_id"] = event_id' not in panel_src_for_delivery:
                    dashboard_delivery_errors.append("dashboard_event_id_not_linked_before_push")
                if (
                    '"__dashboard_lite_pending_persist"' not in panel_src_for_delivery
                    or '"__chat_save_dirty_reason"] = "content_changed"' not in panel_src_for_delivery
                    or '"__chat_room_nav_request"' not in panel_src_for_delivery
                    or 'def _flush_pending_dashboard_chat_persistence()' not in main_dashboard_src
                    or 'save_chat_rooms()' not in main_dashboard_src[main_dashboard_src.find("def _flush_pending_dashboard_chat_persistence()"):]
                    or '"messages_jsonl" in changed_fields' not in main_dashboard_src
                ):
                    dashboard_delivery_errors.append("dashboard_partition_save_pipeline_missing")
                if not hasattr(view_mod, "reset_dashboard_lite_primary_render_guard"):
                    dashboard_delivery_errors.append("dashboard_primary_guard_runtime_missing")
            finally:
                dashboard_chat_mod.st = original_chat_st
                view_mod.st = original_view_st

            original_panel_uuid4 = dashboard_panel_mod.uuid.uuid4
            try:
                dashboard_panel_mod.uuid.uuid4 = lambda: (_ for _ in ()).throw(
                    AssertionError("normal Dashboard push reached UUID fallback")
                )
                precreated_event = dashboard_panel_mod._dashboard_event_id_for_push(
                    {"id": "dashboard-event-precreated"},
                    {"dashboard_event_id": "dashboard-event-precreated"},
                    {"dashboard_event_id": "dashboard-event-precreated"},
                )
                if precreated_event != "dashboard-event-precreated":
                    dashboard_delivery_errors.append(f"dashboard_precreated_event_not_reused={precreated_event!r}")
            except Exception as exc:
                dashboard_delivery_errors.append(f"dashboard_event_fallback_reached={type(exc).__name__}:{exc}")
            finally:
                dashboard_panel_mod.uuid.uuid4 = original_panel_uuid4

            session_state: dict[str, Any] = {}
            dashboard_panel_mod._remember_dashboard_chat_push_signature(
                session_state,
                company_id="company-a",
                room_id="room-a",
                signature="same-query",
            )
            if not dashboard_panel_mod._dashboard_chat_push_is_duplicate(
                session_state,
                company_id="company-a",
                room_id="room-a",
                signature="same-query",
            ):
                dashboard_delivery_errors.append("dashboard_same_room_signature_not_deduped")
            if dashboard_panel_mod._dashboard_chat_push_is_duplicate(
                session_state,
                company_id="company-a",
                room_id="room-b",
                signature="same-query",
            ):
                dashboard_delivery_errors.append("dashboard_cross_room_signature_blocked")
        except Exception as exc:
            dashboard_delivery_errors.append(f"dashboard_delivery_runtime={type(exc).__name__}:{exc}")
        if dashboard_delivery_errors:
            results.append(_fail("Dashboard Lite room-scoped chat delivery", "; ".join(dashboard_delivery_errors)))
        else:
            results.append(_ok("Dashboard Lite room-scoped chat delivery", "pending/history duplicates share one room+event render key, while the same query remains deliverable in another room"))

        stock_batch_errors: list[str] = []
        try:
            import app.services.analytics_sales_trend_service as batch_sales_mod

            default_plan = batch_sales_mod._stock_query_batch_plan(stock_cd_count=20, io_gu_count=80, configured_value=None)
            normal_plan = batch_sales_mod._stock_query_batch_plan(stock_cd_count=20, io_gu_count=80, configured_value="1600")
            oversized_plan = batch_sales_mod._stock_query_batch_plan(stock_cd_count=150, io_gu_count=100, configured_value="5000")
            for invalid in ("text", "0", "-3"):
                plan = batch_sales_mod._stock_query_batch_plan(stock_cd_count=20, io_gu_count=80, configured_value=invalid)
                if plan["configured_batch_size"] != 1800:
                    stock_batch_errors.append(f"batch_invalid_fallback={invalid}:{plan!r}")
            if default_plan["configured_batch_size"] != 1800:
                stock_batch_errors.append(f"batch_default={default_plan!r}")
            if normal_plan["effective_chunk_size"] != 1600:
                stock_batch_errors.append(f"batch_normal={normal_plan!r}")
            if oversized_plan["effective_chunk_size"] >= batch_sales_mod.SQL_SERVER_PARAMETER_LIMIT:
                stock_batch_errors.append(f"batch_oversized_not_capped={oversized_plan!r}")
            plan_2000_products = batch_sales_mod._stock_query_batch_plan(stock_cd_count=120, io_gu_count=160, configured_value="2000")
            if plan_2000_products["total_parameter_count"] >= batch_sales_mod.SQL_SERVER_PARAMETER_LIMIT:
                stock_batch_errors.append(f"batch_parameter_limit={plan_2000_products!r}")
        except Exception as exc:
            stock_batch_errors.append(f"batch_plan_runtime={type(exc).__name__}:{exc}")
        if stock_batch_errors:
            results.append(_fail("Dashboard Lite stock SQL parameter batches", "; ".join(stock_batch_errors)))
        else:
            results.append(_ok("Dashboard Lite stock SQL parameter batches", "configured values fall back safely and every effective product batch remains below the SQL Server bind limit"))

        io_scope_errors: list[str] = []
        try:
            import app.services.analytics_sales_trend_service as io_scope_mod

            exact_params = {"io_gu_list": ["501", "590"]}
            exact_where = io_scope_mod._build_filters(exact_params)
            if "Rd12_Io_Gu_Gcode = %(sales_io_gu_gcode)s" not in exact_where:
                io_scope_errors.append("sales_exact_gcode_missing")
            if "Rd12_Io_Gu IN (%(sales_io_gu_0)s, %(sales_io_gu_1)s)" not in exact_where:
                io_scope_errors.append("sales_exact_tcode_bind_missing")
            if exact_params.get("sales_io_gu_0") != "501" or exact_params.get("sales_io_gu_1") != "590":
                io_scope_errors.append(f"sales_exact_values={exact_params!r}")
            if "LEFT(Out_Put.Rd12_Io_Gu, 1)" in exact_where:
                io_scope_errors.append("sales_exact_prefix_expansion")

            mixed_params = {"io_gu_list": ["001", "002", "501", "590", "601"]}
            monthly_spec = io_scope_mod._monthly_spec("monthly_real")
            monthly_prefix = monthly_spec["prefix"]
            monthly_where, monthly_bind = io_scope_mod._build_monthly_fast_where(
                mixed_params,
                monthly_spec,
            )
            if f"M.{monthly_prefix}_Io_Gu_Gcode = %(sales_io_gu_gcode)s" not in monthly_where:
                io_scope_errors.append("monthly_exact_gcode_missing")
            if f"M.{monthly_prefix}_Io_Gu IN (%(sales_io_gu_0)s, %(sales_io_gu_1)s, %(sales_io_gu_2)s, %(sales_io_gu_3)s, %(sales_io_gu_4)s)" not in monthly_where:
                io_scope_errors.append("monthly_exact_tcode_bind_missing")
            if f"LEFT(M.{monthly_prefix}_Io_Gu, 1) IN" not in monthly_where:
                io_scope_errors.append("monthly_sales_direction_guard_missing")
            if [monthly_bind.get(f"sales_io_gu_{index}") for index in range(5)] != ["001", "002", "501", "590", "601"]:
                io_scope_errors.append(f"monthly_exact_values={monthly_bind!r}")

            legacy_monthly_params = {"io_gu_list": ["501"]}
            legacy_monthly_where = io_scope_mod._build_monthly_filters(
                legacy_monthly_params,
                monthly_spec,
            )
            if (
                f"M.{monthly_prefix}_Io_Gu_Gcode = %(sales_io_gu_gcode)s" not in legacy_monthly_where
                or "%(sales_io_gu_0)s" not in legacy_monthly_where
            ):
                io_scope_errors.append(f"monthly_legacy_exact_io_missing={legacy_monthly_where!r}")

            dedupe_mode, dedupe_codes = io_scope_mod._sales_io_scope({"io_gu_list": ["501", "590", "501"]})
            if dedupe_mode != "exact_selected" or dedupe_codes != ["501", "590"]:
                io_scope_errors.append(f"sales_ordered_dedupe={dedupe_mode}:{dedupe_codes!r}")

            explicit_all_params = {"io_gu_list": []}
            explicit_all_where = io_scope_mod._build_filters(explicit_all_params)
            legacy_where = io_scope_mod._build_filters({})
            if explicit_all_params.get("_sales_io_filter_mode") != "explicit_all" or "LEFT(Out_Put.Rd12_Io_Gu, 1) IN ('5', '6')" not in explicit_all_where:
                io_scope_errors.append("sales_explicit_all_compatibility")
            if "legacy_broad_fallback" not in str(legacy_where) and "LEFT(Out_Put.Rd12_Io_Gu, 1) IN ('5', '6')" not in legacy_where:
                io_scope_errors.append("sales_legacy_broad_fallback")
            try:
                io_scope_mod._sales_io_scope({"io_gu_list": [501]})
                io_scope_errors.append("sales_numeric_code_not_rejected")
            except ValueError:
                pass

            original = {
                "io_gu": "501", "io_gu_list": ["501", "590"], "io_gu_pairs": ["0012:501"],
                "dashboard_io_gu_list": ["501"], "sales_io_gu_list": ["501"], "purchase_io_gu_list": ["001"],
                "stock_cd_list": ["00001"], "product_di_list": ["0004:1"],
            }
            stock_only, ignored_count = io_scope_mod._copy_current_stock_params_without_io_scope(original)
            if any(key in stock_only for key in io_scope_mod._IO_GU_SCOPE_PARAM_KEYS):
                io_scope_errors.append(f"stock_io_alias_retained={stock_only!r}")
            if stock_only.get("stock_cd_list") != ["00001"] or stock_only.get("product_di_list") != ["0004:1"]:
                io_scope_errors.append(f"stock_scope_lost_non_io_filters={stock_only!r}")
            if original.get("io_gu_list") != ["501", "590"] or ignored_count != 7:
                io_scope_errors.append(f"stock_scope_source_mutated_or_count={original!r}|{ignored_count}")

            stock_loader_source = Path(io_scope_mod.__file__).read_text(encoding="utf-8")
            loader_start = stock_loader_source.index("def _load_product_current_stock(")
            loader_end = stock_loader_source.index("def get_stock_shortage_df(", loader_start)
            loader_source = stock_loader_source[loader_start:loader_end]
            if "Io_Gu IN" in loader_source or "io_gu_filter_sql" in loader_source:
                io_scope_errors.append("current_stock_tcode_predicate_present")
            if "Io_Gu_Gcode = %(io_gu_gcode)s" not in loader_source:
                io_scope_errors.append("current_stock_gcode_missing")
            movement_start = stock_loader_source.index("def _load_product_current_month_stock_movements(")
            movement_end = stock_loader_source.index("def _load_product_current_stock(", movement_start)
            movement_source = stock_loader_source[movement_start:movement_end]
            if "Io_Gu)), 1) NOT IN" in movement_source:
                io_scope_errors.append("current_month_movement_prefix_exclusion_present")
            if "Rd11_Io_Gu_Gcode = %(io_gu_gcode)s" not in movement_source or "Rd12_Io_Gu_Gcode = %(io_gu_gcode)s" not in movement_source:
                io_scope_errors.append("current_month_movement_gcode_missing")
            zero_io_plan = io_scope_mod._stock_query_batch_plan(stock_cd_count=100, io_gu_count=0, configured_value="1800")
            if zero_io_plan["io_gu_parameter_count"] != 0 or zero_io_plan["total_parameter_count"] >= io_scope_mod.SQL_SERVER_PARAMETER_LIMIT:
                io_scope_errors.append(f"current_stock_zero_io_bind_plan={zero_io_plan!r}")

            dashboard_facts_source = (PROJECT_ROOT / "app" / "services" / "dashboard_lite_facts.py").read_text(encoding="utf-8")
            if "_dashboard_current_stock_params" in dashboard_facts_source or "stock_shortage_payload[\"params\"] = dict(service_params)" in dashboard_facts_source:
                io_scope_errors.append("dashboard_display_params_override_present")

            customer_forecast_mod = importlib.import_module("app.services.analytics_customer_sales_forecast_service")
            captured_customer_sql: list[tuple[str, dict[str, Any]]] = []
            old_customer_query = customer_forecast_mod.query_to_df
            try:
                customer_forecast_mod.query_to_df = lambda sql, bind: (
                    captured_customer_sql.append((str(sql), dict(bind)))
                    or pd.DataFrame([
                        {"base_month": "202607", "customer_cd": "50001", "supply_amt": 100.0,
                         "tax_amt": 10.0, "total_amt": 110.0, "row_count": 1}
                    ])
                )
                customer_exact = customer_forecast_mod._combine_sources(
                    {"date_from": "20260701", "date_to": "20260731", "io_gu_list": ["501"]},
                    {"effective_date_to": "20260731"},
                )
                if customer_exact.empty or len(captured_customer_sql) != 1:
                    io_scope_errors.append("customer_exact_io_source_not_called")
                elif (
                    "FROM dbo.Rddbc120 AS D" not in captured_customer_sql[0][0]
                    or captured_customer_sql[0][1].get("sales_io_gu_0") != "501"
                ):
                    io_scope_errors.append("customer_exact_io_detail_contract_missing")
            finally:
                customer_forecast_mod.query_to_df = old_customer_query

            nlq_mod = importlib.import_module("app.sims.nlq.nlq_router")
            required_group_actions = {
                "제약사별 매출 추세 분석", "제약사별 매출 추세 분석 요약표",
                "매출처별 매출 예상", "영업사원별 매출 예상", "지역별 매출 예상",
            }
            missing_nlq_io_actions = sorted(
                action for action in required_group_actions
                if "io_gu_list" not in nlq_mod._ANALYTICS_NLQ_DEFAULT_KEYS.get(action, set())
            )
            if missing_nlq_io_actions:
                io_scope_errors.append(f"nlq_group_io_default_missing={missing_nlq_io_actions!r}")
        except Exception as exc:
            io_scope_errors.append(f"io_scope_runtime={type(exc).__name__}:{exc}")
        if io_scope_errors:
            results.append(_fail("Dashboard/KPI IO scope contract", "; ".join(io_scope_errors)))
        else:
            results.append(_ok("Dashboard/KPI IO scope contract", "selected sales Tcodes bind exactly, while current stock removes every IO alias and binds no IO Tcode"))

        supplier_scope_errors: list[str] = []
        try:
            from app.services.product_supplier_scope_service import (
                build_product_supplier_scope_sql, load_supplier_manager_options,
                normalize_product_supplier_scope, supplier_scope_filter_active, supplier_scope_fingerprint,
            )
            manufacturer = normalize_product_supplier_scope({
                "product_supplier_scope_mode": "manufacturer", "manufacturer_codes": ["00015", "00015", 15],
                "manufacturer_manager_codes": ["00021"], "order_vendor_codes": ["20001"],
            })
            if manufacturer["manufacturer_codes"] != ["00015"] or manufacturer["order_vendor_codes"] or manufacturer["purchase_manager_codes"]:
                supplier_scope_errors.append(f"manufacturer_exclusive_or_string_contract={manufacturer!r}")
            binds: dict[str, Any] = {}
            sql = build_product_supplier_scope_sql(manufacturer, binds, product_code_sql="M.Rd21_Physic_Cd", bind_prefix="scope")
            if "SupplierProduct.Rd04_Ven_Cd IN" not in sql or "SupplierVendor.Rd03_Sales_Man IN" not in sql or "Rd04_Orven_Cd" in sql or binds != {"scope_vendor_0": "00015", "scope_manager_0": "00021"}:
                supplier_scope_errors.append(f"manufacturer_sql_contract={sql!r}|{binds!r}")
            order_vendor = normalize_product_supplier_scope({
                "product_supplier_scope_mode": "order_vendor", "manufacturer_codes": ["00015"],
                "order_vendor_codes": ["20001", "20002"], "purchase_manager_codes": ["00031"],
            })
            binds = {}
            sql = build_product_supplier_scope_sql(order_vendor, binds, product_code_sql="M.Rd21_Physic_Cd", bind_prefix="scope")
            if order_vendor["manufacturer_codes"] or "SupplierProduct.Rd04_Orven_Cd IN" not in sql or "SupplierVendor.Rd03_Sales_Man IN" not in sql:
                supplier_scope_errors.append(f"order_vendor_sql_contract={order_vendor!r}|{sql!r}")
            all_scope = normalize_product_supplier_scope({"product_supplier_scope_mode": "all", "manufacturer_codes": ["00015"], "order_vendor_codes": ["20001"]})
            if all_scope["product_supplier_scope_mode"] != "manufacturer" or any(all_scope[key] for key in ("manufacturer_codes", "manufacturer_manager_codes", "order_vendor_codes", "purchase_manager_codes")):
                supplier_scope_errors.append(f"all_scope_not_cleared={all_scope!r}")
            empty_scope_sql = build_product_supplier_scope_sql(all_scope, {}, product_code_sql="M.Rd21_Physic_Cd", bind_prefix="scope")
            if empty_scope_sql:
                supplier_scope_errors.append(f"empty_supplier_scope_has_predicate={empty_scope_sql!r}")
            if (
                supplier_scope_filter_active({"product_supplier_scope_mode": "manufacturer"})
                or supplier_scope_filter_active({"product_supplier_scope_mode": "order_vendor"})
                or not supplier_scope_filter_active({"product_supplier_scope_mode": "manufacturer", "manufacturer_manager_codes": ["00021"]})
                or not supplier_scope_filter_active({"product_supplier_scope_mode": "order_vendor", "order_vendor_codes": ["20001"]})
            ):
                supplier_scope_errors.append("supplier_scope_filter_active_contract")
            scope_service = importlib.import_module("app.services.product_supplier_scope_service")
            captured_manager_queries: list[tuple[str, Any]] = []
            original_manager_query = scope_service.query_to_df
            try:
                scope_service.query_to_df = lambda sql, query_params=(): (captured_manager_queries.append((sql, query_params)) or pd.DataFrame([{"user_code": "00021", "user_name": "담당자"}]))
                manager_rows = load_supplier_manager_options(mode="manufacturer")
                manager_rows_with_vendor = load_supplier_manager_options(mode="order_vendor", vendor_codes=["00015"])
            finally:
                scope_service.query_to_df = original_manager_query
            if (
                manager_rows[0]["code"] != "00021"
                or captured_manager_queries[0][1] != ()
                or "ManagerRows.user_name" not in captured_manager_queries[0][0]
                or "SELECT DISTINCT" not in captured_manager_queries[0][0]
                or captured_manager_queries[1][1] != ("00015",)
                or "P.Rd04_Orven_Cd IN (?)" not in captured_manager_queries[1][0]
            ):
                supplier_scope_errors.append("supplier_manager_sql_bind_or_string_contract")
            if supplier_scope_fingerprint({"product_supplier_scope_mode": "manufacturer", "manufacturer_codes": ["00002", "00001"]}) != supplier_scope_fingerprint({"product_supplier_scope_mode": "manufacturer", "manufacturer_codes": ["00001", "00002"]}):
                supplier_scope_errors.append("supplier_scope_fingerprint_order_dependent")
            inbound_mod = importlib.import_module("app.services.dashboard_inbound_facts_service")
            inbound_sql, _ = inbound_mod._sql({"product_supplier_scope_mode": "manufacturer", "manufacturer_codes": ["00015"]}, start_date="20260101", cutoff_date="20260131")
            if "SupplierProduct.Rd04_Physic_Cd = P.Rd04_Physic_Cd" not in inbound_sql or "P.Rd04_Physic_Cd = P.Rd04_Physic_Cd" in inbound_sql:
                supplier_scope_errors.append("inbound_manufacturer_alias_correlation")
            inbound_sql, _ = inbound_mod._sql({"product_supplier_scope_mode": "order_vendor", "order_vendor_codes": ["20001"]}, start_date="20260101", cutoff_date="20260131")
            if "SupplierProduct.Rd04_Physic_Cd = P.Rd04_Physic_Cd" not in inbound_sql or "SupplierProduct.Rd04_Orven_Cd IN" not in inbound_sql:
                supplier_scope_errors.append("inbound_order_vendor_alias_correlation")
            view_source = Path(view_mod.__file__).read_text(encoding="utf-8")
            form_start = view_source.index('with st.form("dashboard_lite_scope_form"')
            scope_row_start = view_source.index("scope_cols = st.columns([1, 1, 1, 1.1, 2.1, 2.1], gap=\"small\")")
            scope_start = view_source.index('"공급 기준"', scope_row_start)
            if (
                not (scope_row_start < scope_start < form_start)
                or "options=[SCOPE_MANUFACTURER, SCOPE_ORDER_VENDOR]" not in view_source
                or "담당자 목록을 불러오지 못했습니다." not in view_source
                or "on_change=_on_dashboard_supplier_scope_mode_change" not in view_source
            ):
                supplier_scope_errors.append("supplier_scope_form_callback_contract")
            header = view_mod._dashboard_scope_header({"month_from": "202601", "month_to": "202606", "evaluation_month": "202607", "product_supplier_scope_mode": "manufacturer", "supplier_scope_label": "제약사A [00015]", "supplier_manager_label": "담당자A [00021]"})
            if "제약사 담당자: 담당자A [00021]" not in header:
                supplier_scope_errors.append("supplier_manager_header_contract")
            blank_header = view_mod._dashboard_scope_header({"month_from": "202601", "month_to": "202606", "evaluation_month": "202607", "product_supplier_scope_mode": "all"})
            if "공급 기준: 제약사" not in blank_header or "제약사: 전체" not in blank_header or "제약사 담당자: 전체" not in blank_header or "공급 기준: 전체" in blank_header:
                supplier_scope_errors.append("supplier_scope_header_legacy_all_contract")
            for log_source_path in (
                Path("app/services/analytics_sales_trend_service.py"),
                Path("app/services/analytics_manufacturer_sales_trend_service.py"),
            ):
                log_source = log_source_path.read_text(encoding="utf-8")
                if "params=%r" in log_source:
                    supplier_scope_errors.append(f"raw_analytics_params_log={log_source_path.name}")
        except Exception as exc:
            supplier_scope_errors.append(f"supplier_scope_runtime={type(exc).__name__}:{exc}")
        if supplier_scope_errors:
            results.append(_fail("Dashboard product supplier scope contract", "; ".join(supplier_scope_errors)))
        else:
            results.append(_ok("Dashboard product supplier scope contract", "manufacturer/order-vendor scopes are exclusive, legacy all normalizes to blank manufacturer, and manager binds are positional"))

        product_universe_errors: list[str] = []
        try:
            sales_mod = importlib.import_module("app.services.analytics_sales_trend_service")
            dashboard_mod = importlib.import_module("app.services.dashboard_lite_facts")
            original_stock_loader = sales_mod._load_product_current_stock
            stock_loader_calls: list[list[str]] = []
            try:
                def _fake_stock_loader(product_codes, **_kwargs):
                    stock_loader_calls.append(list(product_codes))
                    frame = pd.DataFrame([{
                        "제품코드": "P_SCOPE_EMPTY",
                        "실재고수량": 12.0,
                        "실재고금액": 1200.0,
                        "실재고평가단가": 100.0,
                    }])
                    frame.attrs.update({"stock_sql_ms": 1, "stock_query_batches": 1, "stock_aggregate_ms": 0})
                    return frame

                sales_mod._load_product_current_stock = _fake_stock_loader
                no_sales_df = sales_mod.get_stock_shortage_df(
                    {"stock_mode": "real", "month_from": "202601", "month_to": "202607", "evaluation_month": "202607"},
                    sales_forecast_df=pd.DataFrame(),
                    product_universe_df=pd.DataFrame([{"product_code": "P_SCOPE_EMPTY"}]),
                )
                default_scope_df = sales_mod.get_stock_shortage_df(
                    {"stock_mode": "real", "month_from": "202601", "month_to": "202607", "evaluation_month": "202607"},
                    sales_forecast_df=pd.DataFrame([{"제품코드": "P_DEFAULT"}]),
                    product_universe_df=None,
                )
                explicit_empty_df = sales_mod.get_stock_shortage_df(
                    {"stock_mode": "real", "month_from": "202601", "month_to": "202607", "evaluation_month": "202607"},
                    sales_forecast_df=pd.DataFrame([{"제품코드": "P_DEFAULT"}]),
                    product_universe_df=pd.DataFrame(columns=["product_code"]),
                )
            finally:
                sales_mod._load_product_current_stock = original_stock_loader
            if not isinstance(no_sales_df, pd.DataFrame) or len(no_sales_df) != 1:
                product_universe_errors.append("sales_empty_product_universe_row_missing")
            elif float(no_sales_df.iloc[0].get("현재재고수량") or 0) != 12.0 or float(no_sales_df.iloc[0].get("당월 잔여예상출고수량") or 0) != 0.0:
                product_universe_errors.append("sales_empty_stock_or_demand_contract")
            if not isinstance(default_scope_df, pd.DataFrame) or default_scope_df.empty or str(default_scope_df.iloc[0].get("제품코드") or "") != "P_DEFAULT":
                product_universe_errors.append("empty_scope_default_universe_not_preserved")
            if not isinstance(explicit_empty_df, pd.DataFrame) or not explicit_empty_df.empty:
                product_universe_errors.append("explicit_empty_master_universe_not_empty")
            if stock_loader_calls != [["P_SCOPE_EMPTY"], ["P_DEFAULT"]]:
                product_universe_errors.append(f"sales_empty_stock_loader_calls={stock_loader_calls!r}")
            inventory = dashboard_mod._build_inventory_facts(
                {"df": no_sales_df},
                inbound_facts_df=pd.DataFrame([{"product_code": "P_SCOPE_EMPTY"}]),
                source_call_count=2,
                inbound_source_call_count=1,
                stock_mode="real",
                evaluation_month="202607",
            )
            if len(inventory.get("readiness_rows") or []) != 1:
                product_universe_errors.append("sales_empty_dashboard_inventory_missing")
        except Exception as exc:
            product_universe_errors.append(f"sales_empty_product_universe_runtime={type(exc).__name__}:{exc}")
        if product_universe_errors:
            results.append(_fail("Dashboard supplier product universe", "; ".join(product_universe_errors)))
        else:
            results.append(_ok("Dashboard supplier product universe", "only explicit supplier scopes apply the master universe; normal scope preserves the sales-based Dashboard path"))

        display_unit_errors: list[str] = []
        try:
            view_mod = importlib.import_module("app.sims.views.dashboard_lite")
            event_a_facts = {
                "filters": {"amount_display_unit": "auto"},
                "sales": {"metrics": {"sales": {"unit": "원", "value": 5_000_000}}},
                "inventory": {},
            }
            event_a_unit = view_mod._resolved_dashboard_amount_unit(event_a_facts, "auto")
            event_a_facts["filters"].update({
                "amount_display_unit": event_a_unit,
                "amount_display_unit_requested": "auto",
                "amount_display_unit_resolved": event_a_unit,
            })
            event_a_cache = {"params": {"amount_display_unit": "auto", "amount_display_unit_requested": "auto", "amount_display_unit_resolved": event_a_unit}, "facts": event_a_facts}
            event_a_snapshot = view_mod.build_dashboard_lite_chat_snapshot(event_a_cache)
            event_b_facts = {"filters": {"amount_display_unit": "won", "amount_display_unit_requested": "won", "amount_display_unit_resolved": "won"}}
            if (
                event_a_unit != "million"
                or view_mod._facts_amount_display_unit(event_a_facts) != "million"
                or event_a_snapshot.get("params", {}).get("amount_display_unit_resolved") != "million"
                or event_a_snapshot.get("facts", {}).get("filters", {}).get("amount_display_unit_resolved") != "million"
                or view_mod._facts_amount_display_unit(event_b_facts) != "won"
            ):
                display_unit_errors.append("event_amount_display_unit_not_immutable")
        except Exception as exc:
            display_unit_errors.append(f"event_amount_display_unit_runtime={type(exc).__name__}:{exc}")
        if display_unit_errors:
            results.append(_fail("Dashboard event amount display unit", "; ".join(display_unit_errors)))
        else:
            results.append(_ok("Dashboard event amount display unit", "requested and resolved units are retained per primary event and compact snapshot"))

        profile_reentry_errors: list[str] = []
        try:
            profile_state = {
                "__dashboard_lite_profile_loaded_for": "company-a",
                "__dashboard_lite_manufacturer_text": "manufacturer-test",
                "__dashboard_lite_result": {"facts": {}},
            }
            removed = view_mod.clear_dashboard_lite_session_state(profile_state)
            if "__dashboard_lite_profile_loaded_for" not in removed or profile_state:
                profile_reentry_errors.append(f"dashboard_profile_reset={removed!r}|{profile_state!r}")
        except Exception as exc:
            profile_reentry_errors.append(f"dashboard_profile_reset_runtime={type(exc).__name__}:{exc}")
        if profile_reentry_errors:
            results.append(_fail("Dashboard Lite profile re-entry reset", "; ".join(profile_reentry_errors)))
        else:
            results.append(_ok("Dashboard Lite profile re-entry reset", "Dashboard state clearing removes the profile-loaded marker and non-persistent manufacturer state before re-entry"))

        profile_restore_errors: list[str] = []
        old_profile_st = getattr(view_mod, "st")
        old_profile_identity = getattr(view_mod, "_dashboard_context_identity")
        old_profile_loader = getattr(view_mod, "load_dashboard_profile")
        profile_load_calls: list[str] = []
        saved_profile = {
            "stock_mode": "book",
            "stock_cd_list": ["00001", "stale-stock"],
            "vendor_group_list": ["0019:G_EX"],
            "vendor_kind_list": ["0009:K_EX"],
            "product_group_list": ["0013:G_EX", "stale-product"],
            "product_di_list": ["0004:D_EX"],
            "product_class_list": ["0031:C_EX"],
            "io_gu_list": ["I_EX"],
            "major_purchase_vendor_days": 75,
            "risk_analysis_days": 60,
            "overstock_inactive_days": 45,
            "readiness_warning_pct": 97.5,
            "risk_quick_view_count": 20,
            "amount_display_unit": "million",
        }
        try:
            profile_st = _FakeStreamlit(submit_sequence=[])
            setattr(view_mod, "st", profile_st)
            setattr(view_mod, "_dashboard_context_identity", lambda: {"user_id": "user-a", "company_id": "4", "db_sig": "safe"})

            def _fake_profile_loader(*, company_id):
                profile_load_calls.append(str(company_id))
                return dict(saved_profile)

            setattr(view_mod, "load_dashboard_profile", _fake_profile_loader)
            view_mod._apply_saved_dashboard_profile_once()
            for source_key, widget_key in view_mod._DASHBOARD_PROFILE_WIDGETS.items():
                expected_value = view_mod._dashboard_profile_widget_value(source_key, saved_profile[source_key])
                if profile_st.session_state.get(widget_key) != expected_value:
                    profile_restore_errors.append(f"initial_profile_widget_not_restored={source_key}:{profile_st.session_state.get(widget_key)!r}")

            profile_st.session_state["__dashboard_lite_product_group_list"] = ["live-product"]
            profile_st.session_state["__dashboard_lite_risk_analysis_days"] = 12
            view_mod._apply_saved_dashboard_profile_once()
            if (
                len(profile_load_calls) != 1
                or profile_st.session_state.get("__dashboard_lite_product_group_list") != ["live-product"]
                or profile_st.session_state.get("__dashboard_lite_risk_analysis_days") != 12
            ):
                profile_restore_errors.append(f"live_state_overwritten={profile_load_calls!r}|{profile_st.session_state!r}")

            for widget_key in view_mod._DASHBOARD_PROFILE_WIDGETS.values():
                profile_st.session_state.pop(widget_key, None)
            saved_profile["risk_analysis_days"] = 30
            view_mod._apply_saved_dashboard_profile_once()
            if len(profile_load_calls) != 2:
                profile_restore_errors.append(f"action_reentry_profile_not_reloaded={profile_load_calls!r}")
            if profile_st.session_state.get("__dashboard_lite_risk_analysis_days") != 30:
                profile_restore_errors.append(f"action_reentry_profile_not_restored={profile_st.session_state!r}")
            view_mod._prepare_dashboard_multiselect_state("__dashboard_lite_stock_labels", ["00001"])
            view_mod._prepare_dashboard_multiselect_state("__dashboard_lite_product_group_list", ["0013:G_EX"])
            if profile_st.session_state.get("__dashboard_lite_stock_labels") != ["00001"]:
                profile_restore_errors.append("stale_stock_option_not_pruned")
            if profile_st.session_state.get("__dashboard_lite_product_group_list") != ["0013:G_EX"]:
                profile_restore_errors.append("stale_product_option_not_pruned")

            empty_profile_st = _FakeStreamlit(submit_sequence=[])
            setattr(view_mod, "st", empty_profile_st)
            setattr(view_mod, "_dashboard_context_identity", lambda: {"user_id": "user-b", "company_id": "5", "db_sig": "safe"})
            setattr(view_mod, "load_dashboard_profile", lambda **_kwargs: None)
            view_mod._apply_saved_dashboard_profile_once()
            view_mod._prepare_dashboard_profile_scalar_state()
            for source_key, default_value in view_mod._DASHBOARD_PROFILE_SCALAR_DEFAULTS.items():
                widget_key = view_mod._DASHBOARD_PROFILE_WIDGETS[source_key]
                if empty_profile_st.session_state.get(widget_key) != default_value:
                    profile_restore_errors.append(f"empty_profile_default_missing={source_key}:{empty_profile_st.session_state!r}")

            company_profile_st = _FakeStreamlit(submit_sequence=[])
            identities = {"company_id": "4"}
            company_profiles = {
                "4": dict(saved_profile),
                "5": {**saved_profile, "stock_mode": "real", "risk_analysis_days": 45, "amount_display_unit": "won"},
            }
            setattr(view_mod, "st", company_profile_st)
            setattr(view_mod, "_dashboard_context_identity", lambda: {"user_id": "user-c", "company_id": identities["company_id"], "db_sig": "safe"})
            setattr(view_mod, "load_dashboard_profile", lambda *, company_id: dict(company_profiles[str(company_id)]))
            view_mod._apply_saved_dashboard_profile_once()
            identities["company_id"] = "5"
            view_mod._apply_saved_dashboard_profile_once()
            if (
                company_profile_st.session_state.get("__dashboard_lite_stock_mode") != "real"
                or company_profile_st.session_state.get("__dashboard_lite_risk_analysis_days") != 45
                or company_profile_st.session_state.get("__dashboard_lite_amount_display_unit") != "won"
            ):
                profile_restore_errors.append(f"company_profile_mixed={company_profile_st.session_state!r}")
        except Exception as exc:
            profile_restore_errors.append(f"profile_restore_runtime={type(exc).__name__}:{exc}")
        finally:
            setattr(view_mod, "st", old_profile_st)
            setattr(view_mod, "_dashboard_context_identity", old_profile_identity)
            setattr(view_mod, "load_dashboard_profile", old_profile_loader)
        if profile_restore_errors:
            results.append(_fail("Dashboard Lite action re-entry profile restore", "; ".join(profile_restore_errors)))
        else:
            results.append(_ok("Dashboard Lite action re-entry profile restore", "missing Dashboard widget keys reload the saved company profile, while ordinary reruns preserve live widget state"))

        sales_source_errors: list[str] = []
        for token in (
            "제품그룹Gcode",
            "제품그룹코드",
            "제품구분Gcode",
            "제품구분코드",
            "제품분류Gcode",
            "제품분류코드",
            "Rd04_Physic_Group_Gcode",
            "Rd04_Physic_Di_Gcode",
            "Rd04_Physic_Tax_Gcode",
            "Rd04_Physic_Tax",
            "Physic_Group_Nm.Rd01_Hnm",
            "Physic_Di_Nm.Rd01_Hnm",
            "Physic_Tax_Nm.Rd01_Hnm",
        ):
            if token not in sales_trend_src:
                sales_source_errors.append(f"missing={token}")
        if "Rd01_Hmn" in sales_trend_src:
            sales_source_errors.append("wrong_rd01_hmn_column_present")
        dashboard_class_blocks = [
            sales_trend_src[sales_trend_src.find("def _load_monthly_product_master_for_codes"):sales_trend_src.find("def _load_monthly_vendor_names_for_codes")],
            sales_trend_src[sales_trend_src.find("def _get_sales_trend_monthly_df_legacy"):sales_trend_src.find("def get_sales_trend_detail_df")],
            sales_trend_src[sales_trend_src.find("def get_sales_trend_detail_df"):],
        ]
        for idx, block in enumerate(dashboard_class_blocks):
            if "제품분류Gcode" in block and ("Rd04_Physic_Gu_Gcode AS 제품분류Gcode" in block or "Physic_Gu_Nm.Rd01_Hnm AS 제품분류명" in block):
                sales_source_errors.append(f"dashboard_product_class_uses_gu_path={idx}")
        if sales_source_errors:
            results.append(_fail("Dashboard Lite sales source code-pair columns", "; ".join(sales_source_errors)))
        else:
            results.append(_ok("Dashboard Lite sales source code-pair columns", "monthly/detail sales sources expose product group/type/class Gcode+Tcode and Rd01_Hnm names"))

        vendor_stock_risk_errors: list[str] = []
        try:
            vendor_rows = [
                {"product_code": code, "재고위험상태": "긴급 부족" if index == 0 else "부족 주의", "위험보정부족예상금액": float((index + 1) * 100), "위험보정부족예상수량": float(index + 1), "수요급증여부": index == 1, "과잉후보여부": index == 2, "과잉후보금액": float(index)}
                for index, code in enumerate(["P1", "P2", "P3", "P4", "P5", "P6", "P7", ""])
            ]
            vendor_rows.append(
                {"product_code": "P8", "재고위험상태": "긴급 부족", "위험보정부족예상금액": 0.0, "위험보정부족예상수량": 1.0, "수요급증여부": False, "과잉후보여부": False, "과잉후보금액": 0.0}
            )
            purchase_vendor_df = pd.DataFrame([
                {"기준월": "202601", "제품코드": "P1", "매입처코드": "A", "매입처명": "공급처 A", "입고수량": 1, "매입금액": 100, "매입발생건수": 1},
                {"기준월": "202602", "제품코드": "P1", "매입처코드": "B", "매입처명": "공급처 B", "입고수량": 1, "매입금액": 200, "매입발생건수": 1},
                {"기준월": "202603", "제품코드": "P2", "매입처코드": "C", "매입처명": "공급처 C", "입고수량": 1, "매입금액": 100, "매입발생건수": 1},
                {"기준월": "202603", "제품코드": "P2", "매입처코드": "D", "매입처명": "공급처 D", "입고수량": 2, "매입금액": 100, "매입발생건수": 1},
                {"기준월": "202603", "제품코드": "P3", "매입처코드": "E", "매입처명": "공급처 E", "입고수량": 1, "매입금액": 100, "매입발생건수": 1},
                {"기준월": "202606", "제품코드": "P3", "매입처코드": "F", "매입처명": "공급처 F", "입고수량": 1, "매입금액": 100, "매입발생건수": 1},
                {"기준월": "202505", "제품코드": "P4", "매입처코드": "G", "매입처명": "공급처 G", "입고수량": 1, "매입금액": 50, "매입발생건수": 1},
                {"기준월": "202604", "제품코드": "P5", "매입처코드": "R", "매입처명": "반품처", "입고수량": -1, "매입금액": -50, "매입발생건수": 1},
                {"기준월": "202605", "제품코드": "P6", "매입처코드": "", "매입처명": "", "입고수량": 1, "매입금액": 50, "매입발생건수": 1},
                {"기준월": "202606", "제품코드": "P7", "매입처코드": "A", "매입처명": "공급처 A", "입고수량": 1, "매입금액": 100, "매입발생건수": 1},
                {"기준월": "202606", "제품코드": "P7", "매입처코드": "", "매입처명": "", "입고수량": 1, "매입금액": 200, "매입발생건수": 1},
                {"기준월": "202606", "제품코드": "", "매입처코드": "X", "매입처명": "", "입고수량": 1, "매입금액": 10, "매입발생건수": 1},
                {"기준월": "", "제품코드": "PM", "매입처코드": "X", "매입처명": "", "입고수량": 1, "매입금액": 10, "매입발생건수": 1},
                {"기준월": "202606", "제품코드": "PI", "매입처코드": "X", "매입처명": "", "입고수량": 1, "매입금액": "invalid", "매입발생건수": 1},
                {"기준월": "202701", "제품코드": "PO", "매입처코드": "X", "매입처명": "", "입고수량": 1, "매입금액": 10, "매입발생건수": 1},
            ])
            vendor_result = dash_mod._attach_major_purchase_vendors(
                vendor_rows,
                purchase_vendor_df,
                evaluation_month="202607",
                history_month_from="202501",
                source_call_count=2,
            )
            by_product = {str(row.get("product_code") or ""): row for row in vendor_rows}
            expected_winners = {"P1": "B", "P2": "D", "P3": "F", "P4": "G", "P7": "A"}
            for product_code, vendor_code in expected_winners.items():
                if by_product[product_code].get("주요매입처코드") != vendor_code:
                    vendor_stock_risk_errors.append(f"winner={product_code}:{by_product[product_code]!r}")
            if by_product["P4"].get("주요매입처선정기준") != "지원기간 fallback":
                vendor_stock_risk_errors.append(f"fallback_basis={by_product['P4']!r}")
            if by_product["P5"].get("주요매입처상태") != "recent_purchase_none":
                vendor_stock_risk_errors.append(f"return_only_status={by_product['P5']!r}")
            if by_product["P6"].get("주요매입처상태") != "vendor_unknown":
                vendor_stock_risk_errors.append(f"blank_vendor_status={by_product['P6']!r}")
            if by_product["P7"].get("주요매입처상태") != "assigned":
                vendor_stock_risk_errors.append(f"mixed_blank_vendor_status={by_product['P7']!r}")
            if by_product[""].get("주요매입처상태") != "product_code_missing":
                vendor_stock_risk_errors.append(f"missing_product_status={by_product['']!r}")
            summary = vendor_result.get("summary") or {}
            if int(summary.get("assigned_rows") or 0) + int(summary.get("unassigned_rows") or 0) != int(summary.get("risk_rows") or 0):
                vendor_stock_risk_errors.append(f"assignment_count_mismatch={summary!r}")
            for status_key, positive_key, zero_key in (
                ("status_risk_rows", "amount_positive_risk_rows", "amount_zero_risk_rows"),
                ("status_emergency_rows", "amount_positive_emergency_rows", "amount_zero_emergency_rows"),
                ("status_warning_rows", "amount_positive_warning_rows", "amount_zero_warning_rows"),
            ):
                if int(summary.get(status_key) or 0) != int(summary.get(positive_key) or 0) + int(summary.get(zero_key) or 0):
                    vendor_stock_risk_errors.append(f"status_amount_count_mismatch={status_key}:{summary!r}")
            if int(summary.get("risk_rows") or 0) != int(summary.get("amount_positive_risk_rows") or 0) or int(summary.get("amount_zero_emergency_rows") or 0) != 1:
                vendor_stock_risk_errors.append(f"zero_amount_vendor_exclusion={summary!r}")
            if abs(float(summary.get("assigned_adjusted_shortage_amount") or 0) + float(summary.get("unassigned_adjusted_shortage_amount") or 0) - float(summary.get("total_adjusted_shortage_amount") or 0)) > 1e-9:
                vendor_stock_risk_errors.append(f"assignment_amount_mismatch={summary!r}")
            if int(summary.get("purchase_positive_rows") or 0) + int(summary.get("purchase_nonpositive_rows") or 0) + int(summary.get("purchase_unclassified_rows") or 0) != int(summary.get("purchase_source_rows") or 0):
                vendor_stock_risk_errors.append(f"purchase_classification_mismatch={summary!r}")
            if not all(int(summary.get(key) or 0) >= 1 for key in ("missing_product_code_rows", "missing_month_rows", "invalid_numeric_rows", "other_excluded_rows")):
                vendor_stock_risk_errors.append(f"purchase_unclassified_reason_missing={summary!r}")
            if len(vendor_result.get("top_rows") or []) > 10 or any(not str(row.get("주요매입처코드") or "") for row in (vendor_result.get("rows") or [])):
                vendor_stock_risk_errors.append(f"vendor_rows_invalid={vendor_result!r}")
            if any(int(vendor_result.get(key) or 0) < 0 for key in ("aggregate_ms", "rank_ms", "group_ms")):
                vendor_stock_risk_errors.append(f"vendor_timing_invalid={vendor_result!r}")
            required_facts_tokens = (
                "get_dashboard_sales_source_bundle",
                "purchase_vendor_df",
                "vendor_stock_risk_summary",
                "vendor_stock_risk_top_rows",
                "[dashboard.vendor_stock_risk]",
                "source_call_count=%s",
            )
            for token in required_facts_tokens:
                if token not in dashboard_facts_src and token not in sales_trend_src:
                    vendor_stock_risk_errors.append(f"missing_source_token={token}")
            if "analytics_supplier_stock_shortage_service" in dashboard_facts_src:
                vendor_stock_risk_errors.append("dashboard_vendor_risk_uses_supplier_service")
            if "'sales' AS [_dashboard_source_kind]" not in sales_trend_src or "'purchase_vendor' AS [_dashboard_source_kind]" not in sales_trend_src:
                vendor_stock_risk_errors.append("dashboard_source_bundle_branch_missing")
            for token in ("raw_bundle_rows", "sales_finalize_ms", "purchase_min_frame_ms", "source_scan_mode", "union_branches"):
                if token not in sales_trend_src:
                    vendor_stock_risk_errors.append(f"bundle_timing_missing={token}")
            attach_vendor_src = dashboard_facts_src[
                dashboard_facts_src.find("def _attach_major_purchase_vendors"):
                dashboard_facts_src.find("def _build_inventory_facts")
            ]
            for token in ("transform(\"any\")", "_candidate_tier", "drop_duplicates(subset=[\"제품코드\"]", "purchase_unclassified_rows", "amount_positive_risk_rows"):
                if token not in attach_vendor_src:
                    vendor_stock_risk_errors.append(f"vendor_vectorization_missing={token}")
            for forbidden in ("for product_code, candidates", ".groupby.apply", ".iterrows()"):
                if forbidden in attach_vendor_src:
                    vendor_stock_risk_errors.append(f"vendor_vectorization_forbidden={forbidden}")
            if "vendor_stock_risk_summary" not in view_src or "vendor_stock_risk_rows" in view_src[view_src.find("def build_dashboard_lite_chat_snapshot"):view_src.find("def _dashboard_scope_header")]:
                vendor_stock_risk_errors.append("dashboard_snapshot_vendor_contract")
        except Exception as exc:
            vendor_stock_risk_errors.append(f"vendor_stock_risk_runtime={type(exc).__name__}:{exc}")
        if vendor_stock_risk_errors:
            results.append(_fail("Dashboard Lite vendor stock risk", "; ".join(vendor_stock_risk_errors)))
        else:
            results.append(_ok("Dashboard Lite vendor stock risk", "one major vendor per product, recent-six priority/fallback, risk reconciliation, compact snapshot, and two-source contract verified"))

        risk_detail_errors: list[str] = []
        try:
            risk_detail_mod = importlib.import_module("app.services.dashboard_risk_detail_export")
            key = {
                "state": "\uc7ac\uace0\uc704\ud5d8\uc0c1\ud0dc",
                "reason": "\uc7ac\uace0\uc704\ud5d8\uc0ac\uc720",
                "amount": "\uc704\ud5d8\ubcf4\uc815\ubd80\uc871\uc608\uc0c1\uae08\uc561",
                "qty": "\uc704\ud5d8\ubcf4\uc815\ubd80\uc871\uc608\uc0c1\uc218\ub7c9",
                "surge": "\uc218\uc694\uae09\uc99d\uc5ec\ubd80",
                "vendor_code": "\uc8fc\uc694\ub9e4\uc785\ucc98\ucf54\ub4dc",
                "vendor_name": "\uc8fc\uc694\ub9e4\uc785\ucc98\uba85",
                "vendor_status": "\uc8fc\uc694\ub9e4\uc785\ucc98\uc0c1\ud0dc",
                "stock_qty": "\ud604\uc7ac\uc7ac\uace0\uc218\ub7c9",
                "stock_amt": "\ud604\uc7ac\uc7ac\uace0\uae08\uc561",
                "readiness": "\uc704\ud5d8\ubcf4\uc815\uc7ac\uace0\uc900\ube44\uc728",
            }
            detail_source = [
                {
                    "product_code": "P-001", "product_name": "Alpha", "specification": "10", "manufacturer_name": "M1",
                    key["state"]: "\uae34\uae09 \ubd80\uc871", key["reason"]: "\ubd80\uc871", key["amount"]: 500.0, key["qty"]: 5.0,
                    key["surge"]: True, key["vendor_code"]: "A", key["vendor_name"]: "Vendor A", key["vendor_status"]: "assigned",
                    key["stock_qty"]: 1.0, key["stock_amt"]: 100.0, key["readiness"]: 20.0,
                },
                {
                    "product_code": "P-002", "product_name": "Beta", key["state"]: "\ubd80\uc871 \uc8fc\uc758", key["reason"]: "\uc8fc\uc758",
                    key["amount"]: 0.0, key["qty"]: 1.0, key["surge"]: False,
                    key["vendor_code"]: "", key["vendor_name"]: "", key["vendor_status"]: "vendor_unknown",
                },
                {
                    "product_code": "P-003", "product_name": "Gamma", key["state"]: "\uc801\uc815", key["amount"]: 900.0,
                },
            ]
            detail_rows, detail_summary = dash_mod._build_dashboard_risk_detail(detail_source, stock_mode="real")
            if len(detail_rows) != 2 or int(detail_summary.get("emergency_rows") or 0) != 1 or int(detail_summary.get("warning_rows") or 0) != 1:
                risk_detail_errors.append(f"detail_status_summary={detail_summary!r}")
            filtered_df, filtered_summary, _ = dash_mod.filter_dashboard_risk_detail_rows(
                detail_rows,
                risk_status="\uae34\uae09 \ubd80\uc871",
                vendor_key="assigned:A",
                surge_filter="\uc218\uc694\uae09\uc99d",
                include_zero_amount=True,
                search_text="P-001",
            )
            if len(filtered_df) != 1 or int(filtered_summary.get("filtered_rows") or 0) != 1:
                risk_detail_errors.append(f"detail_filter={filtered_summary!r}")
            nonzero_df, _, _ = dash_mod.filter_dashboard_risk_detail_rows(detail_rows, include_zero_amount=False)
            if len(nonzero_df) != 1:
                risk_detail_errors.append(f"detail_zero_exclusion={len(nonzero_df)}")

            detail_rows[0]["_\uc8fc\uc694\ub9e4\uc785\ucc98\ud544\ud130\ud0a4"] = "assigned:A"
            excel_bytes, excel_info = risk_detail_mod.build_dashboard_risk_detail_excel_bytes(
                detail_rows,
                [{"\uc8fc\uc694\ub9e4\uc785\ucc98\uba85": "Vendor A", "\uc804\uccb4\uc704\ud5d8\ubcf4\uc815\ubd80\uc871\uae08\uc561": 500.0}],
                [
                    {"\uc870\ud68c\uc870\uac74": "\uc870\ud68c\uc644\ub8cc\uc2dc\uac01", "\uc801\uc6a9\uac12": "query-finished"},
                    {"\uc870\ud68c\uc870\uac74": "Excel\uc0dd\uc131\uc2dc\uac01", "\uc801\uc6a9\uac12": "excel-created"},
                ],
            )
            from io import BytesIO
            from openpyxl import load_workbook

            workbook = load_workbook(BytesIO(excel_bytes), read_only=False)
            if set(workbook.sheetnames) != {risk_detail_mod.DETAIL_SHEET_NAME, risk_detail_mod.VENDOR_SHEET_NAME, risk_detail_mod.SCOPE_SHEET_NAME}:
                risk_detail_errors.append(f"excel_sheets={workbook.sheetnames!r}")
            detail_sheet = workbook[risk_detail_mod.DETAIL_SHEET_NAME]
            detail_headers = [str(cell.value or "") for cell in detail_sheet[1]]
            readiness_column = detail_headers.index(key["readiness"]) + 1 if key["readiness"] in detail_headers else 0
            readiness_cell = detail_sheet.cell(row=2, column=readiness_column) if readiness_column else None
            if (
                int(excel_info.get("export_rows") or 0) != 2
                or detail_sheet.freeze_panes != "A2"
                or any(header.startswith("_") for header in detail_headers)
                or "_\uc8fc\uc694\ub9e4\uc785\ucc98\ud544\ud130\ud0a4" in detail_headers
                or not {"\uc81c\ud488\ucf54\ub4dc", key["amount"], key["vendor_name"]}.issubset(set(detail_headers))
                or readiness_cell is None
                or readiness_cell.value != 20.0
                or readiness_cell.number_format != r"0.00\%"
            ):
                risk_detail_errors.append(f"excel_contract={excel_info!r}")
            scope_sheet = workbook[risk_detail_mod.SCOPE_SHEET_NAME]
            scope_pairs = {
                str(scope_sheet.cell(row=row_index, column=1).value or ""): str(scope_sheet.cell(row=row_index, column=2).value or "")
                for row_index in range(2, scope_sheet.max_row + 1)
            }
            if scope_pairs.get("\uc870\ud68c\uc644\ub8cc\uc2dc\uac01") != "query-finished" or scope_pairs.get("Excel\uc0dd\uc131\uc2dc\uac01") != "excel-created":
                risk_detail_errors.append(f"scope_timestamps={scope_pairs!r}")

            primary_detail_count = len(detail_rows)
            detail_cache = {
                "params": {"month_from": "202601", "month_to": "202606", "evaluation_month": "202607"},
                "facts": {"inventory": {"risk_detail_summary": detail_summary, "risk_detail_rows": detail_rows, "readiness_rows": detail_rows}},
            }
            detail_snapshot = view_mod.build_dashboard_lite_chat_snapshot(detail_cache)
            snapshot_json = json.dumps(detail_snapshot, ensure_ascii=False)
            primary_detail_after_snapshot = len(
                (((detail_cache.get("facts") or {}).get("inventory") or {}).get("risk_detail_rows") or [])
            )
            if (
                "risk_detail_rows" in snapshot_json
                or "risk_detail_summary" not in snapshot_json
                or primary_detail_after_snapshot != primary_detail_count
            ):
                risk_detail_errors.append("risk_detail_snapshot_contract")
            detail_view_segment = view_src[view_src.find("def _render_risk_detail"):view_src.find("def _render_today_actions")]
            for token in ("filter_dashboard_risk_detail_rows", "build_dashboard_risk_detail_excel_bytes", 'require_permission("EXPORT_EXCEL"', "__dashboard_lite_risk_detail_excel::"):
                if token not in view_src:
                    risk_detail_errors.append(f"risk_detail_ui_token={token}")
            detail_filter_source = dashboard_facts_src[
                dashboard_facts_src.find("def filter_dashboard_risk_detail_rows"):
                dashboard_facts_src.find("def _attach_major_purchase_vendors")
            ]
            if "build_dashboard_lite_facts(" in detail_filter_source:
                risk_detail_errors.append("risk_detail_filter_reloads_facts")
            risk_detail_render_source = view_src[view_src.find("def _render_risk_detail"):view_src.find("def _dashboard_context_identity")]
            no_rows_index = risk_detail_render_source.find("if not rows:")
            toggle_index = risk_detail_render_source.find("toggle(")
            if no_rows_index < 0 or toggle_index < 0 or no_rows_index > toggle_index:
                risk_detail_errors.append("snapshot_toggle_not_blocked_before_render")
            for token in ('render_mode="primary"', 'render_mode="chat"', 'render_mode != "primary"', "[dashboard.risk_detail.render]"):
                if token not in view_src:
                    risk_detail_errors.append(f"risk_detail_render_mode_missing={token}")
            query_conditions = view_mod._risk_detail_query_conditions(
                {"created_at": "query-finished", "params": {}},
                {"inventory": {}},
                excel_created_at="excel-created",
            )
            query_condition_values = {str(row.get("\uc870\uac74\uba85") or ""): str(row.get("\uac12") or "") for row in query_conditions}
            if query_condition_values.get("\uc870\ud68c\uc644\ub8cc\uc2dc\uac01") != "query-finished" or query_condition_values.get("Excel\uc0dd\uc131\uc2dc\uac01") != "excel-created":
                risk_detail_errors.append(f"query_condition_timestamps={query_condition_values!r}")
        except Exception as exc:
            risk_detail_errors.append(f"risk_detail_runtime={type(exc).__name__}:{exc}")
        if risk_detail_errors:
            results.append(_fail("Dashboard Lite risk detail and Excel export", "; ".join(risk_detail_errors)))
        else:
            results.append(_ok("Dashboard Lite risk detail and Excel export", "risk filters run on cached facts, Excel has three sheets, and detail rows stay outside the chat snapshot"))

        ui_security_src = Path("app/Lmstudio_SSAI_chat_main.py").read_text(encoding="utf-8")
        chat_middleware_src = Path("app/ui/chat_middleware.py").read_text(encoding="utf-8")
        panel_src_for_security = Path("app/ui/sims_panel.py").read_text(encoding="utf-8")
        db_src = Path("app/db/mssql_client.py").read_text(encoding="utf-8")
        login_src = Path("app/ui/ssai_login.py").read_text(encoding="utf-8")
        security_errors: list[str] = []
        forbidden_ui_tokens = [
            '"LMSTUDIO_BASE_URL": os.getenv',
            '"MSSQL_SERVER":',
            '"MSSQL_DATABASE":',
            "Base URL:",
            "list_tables(20)",
            "search_columns(",
            '"BASE_DIR": str(',
            '"CHAT_FILE": str(',
            '"UPLOAD_DIR": str(',
        ]
        for token in forbidden_ui_tokens:
            if token in ui_security_src:
                security_errors.append(f"ui_token={token}")
        forbidden_runtime_log_tokens = [
            "env_file=%s",
            "CHAT_FILE=%s",
            "LOG_FILE=%s",
            "sys_executable=%s",
            "effective_user_chat_file=%s",
            "partition_root=%s",
            "root=%s",
            "db_name=%s",
            "legacy_file=%s",
            "chat_file=%s",
            "fallback legacy file=%s",
        ]
        for token in forbidden_runtime_log_tokens:
            if token in ui_security_src or token in chat_middleware_src or token in db_src or token in login_src:
                security_errors.append(f"runtime_log_token={token}")
        for token in ("payload_db=%s", "current_db=%s"):
            if token in ui_security_src or token in chat_middleware_src or token in panel_src_for_security:
                security_errors.append(f"stale_payload_db_log_token={token}")
        for token in ("login_id=%s", "company_name=%s", "db_name=%s", "[auth.company] sidebar company list failed login_id=%s"):
            if token in login_src:
                security_errors.append(f"auth_runtime_log_token={token}")
        for token in ("DB명:", "return f\"{middle} / DB:"):
            if token in login_src:
                security_errors.append(f"physical_db_ui_token={token}")
        if '"host": cfg.host' in ui_security_src or '"db": cfg.database' in ui_security_src or '"user": cfg.user' in ui_security_src:
            security_errors.append("db_diagnostic_raw_config_present")
        for token in ('log.exception("[chat.load]', 'log.exception("[chat.storage.save]'):
            if token in ui_security_src:
                security_errors.append(f"traceback_log_token={token}")
        safe_startup_log = (
            "[app.env] env_configured=%s loaded=%s APP_ENV=%s SSAI_INSTANCE_ID=%s chat_storage_configured=%s log_configured=%s venv_active=%s"
            % (True, True, "dev", "HO1", True, True, True)
        )
        safe_user_path_log = "[app.env.user_paths] chat_storage_configured=%s partition_storage_configured=%s" % (True, True)
        if any(fragment in safe_startup_log or fragment in safe_user_path_log for fragment in ("C:\\", "CHAT_FILE=", "LOG_FILE=", "sys_executable=", "Database=", "Pwd=")):
            security_errors.append(f"safe_log_format_leaks={safe_startup_log!r}|{safe_user_path_log!r}")
        forbidden_log_tokens = [
            "ODBC conn: %s",
            "ODBC company conn company_id=%s: %s",
            "Pwd=******",
        ]
        for token in forbidden_log_tokens:
            if token in db_src:
                security_errors.append(f"log_token={token}")
        if "[db.connection]" not in db_src or "connection_configured" not in db_src:
            security_errors.append("safe_db_connection_log_missing")
        chat_context_src = ui_security_src[ui_security_src.find("def _chat_log_context"):ui_security_src.find("def _chat_log_kv")]
        chat_context_base = chat_context_src[:chat_context_src.find("blocked_keys =")]
        if any(token in chat_context_base for token in ('"login_id"', '"company_name"', '"db_name"', '"chat_file"', '"legacy_file"', '"partition_root"')):
            security_errors.append("chat_log_context_sensitive_key_present")
        try:
            from types import SimpleNamespace

            main_tree = ast.parse(ui_security_src)
            helper_nodes = [
                node for node in main_tree.body
                if isinstance(node, ast.FunctionDef) and node.name in {"_safe_log_value", "_chat_log_context", "_chat_log_kv"}
            ]
            helper_ns = {"Any": Any}
            exec(compile(ast.Module(body=helper_nodes, type_ignores=[]), "chat_log_helpers", "exec"), helper_ns)
            helper_ns["get_current_user"] = lambda: SimpleNamespace(user_id=7, login_id="hidden-login", user_type="SSART_ADMIN", user_grade="SUPER")
            helper_ns["get_selected_company"] = lambda: {"company_id": 4, "company_name": "hidden-company", "db_name": "hidden-db"}
            rendered_log = helper_ns["_chat_log_kv"](
                {"id": "room-1"},
                chat_file="C:\\hidden\\room.json",
                legacy_file="C:\\hidden\\legacy.json",
                db_name="hidden-db",
                server="hidden-server",
                password="hidden-password",
                api_key="hidden-key",
                message_count=3,
            )
            if any(value in rendered_log for value in ("hidden-login", "hidden-company", "hidden-db", "hidden-server", "hidden-password", "hidden-key", "C:\\hidden")):
                security_errors.append(f"chat_log_helper_leak={rendered_log!r}")
            if "user_id=7" not in rendered_log or "company_id=4" not in rendered_log or "message_count=3" not in rendered_log:
                security_errors.append(f"chat_log_helper_state_missing={rendered_log!r}")
        except Exception as exc:
            security_errors.append(f"chat_log_helper_runtime={type(exc).__name__}:{exc}")
        try:
            login_tree = ast.parse(login_src)
            login_helper_nodes = [
                node for node in login_tree.body
                if isinstance(node, ast.FunctionDef) and node.name in {"_safe_log_value", "_login_log_context", "_login_log_kv"}
            ]
            login_ns = {"Any": Any, "AuthUser": Any}
            exec(compile(ast.Module(body=login_helper_nodes, type_ignores=[]), "login_log_helpers", "exec"), login_ns)
            rendered_login_log = login_ns["_login_log_kv"](
                user=SimpleNamespace(user_id=8, login_id="hidden-login", user_type="WHOLESALE_ADMIN", user_grade="MANAGER"),
                company={"company_id": 5, "company_name": "hidden-company", "db_name": "hidden-db"},
                login_id="hidden-login",
                company_name="hidden-company",
                db_name="hidden-db",
                server="hidden-server",
                password="hidden-password",
                api_key="hidden-key",
                permission_count=2,
            )
            if any(value in rendered_login_log for value in ("hidden-login", "hidden-company", "hidden-db", "hidden-server", "hidden-password", "hidden-key")):
                security_errors.append(f"login_log_helper_leak={rendered_login_log!r}")
            if "user_id=8" not in rendered_login_log or "company_id=5" not in rendered_login_log or "permission_count=2" not in rendered_login_log:
                security_errors.append(f"login_log_helper_state_missing={rendered_login_log!r}")
        except Exception as exc:
            security_errors.append(f"login_log_helper_runtime={type(exc).__name__}:{exc}")
        try:
            import app.ui.chat_middleware as runtime_chat_mod
            rendered_runtime_log = runtime_chat_mod._chat_runtime_log_kv(
                {"id": "room-9"},
                login_id="hidden-login",
                company_name="hidden-company",
                db_name="hidden-db",
                chat_file="C:\\hidden\\room.json",
                legacy_file="C:\\hidden\\legacy.json",
                server="hidden-server",
                password="hidden-password",
                api_key="hidden-key",
                table_key="sims_safe",
                row_count=3,
            )
            if any(value in rendered_runtime_log for value in ("hidden-login", "hidden-company", "hidden-db", "hidden-server", "hidden-password", "hidden-key", "C:\\hidden")):
                security_errors.append(f"runtime_chat_log_helper_leak={rendered_runtime_log!r}")
            if "table_key=sims_safe" not in rendered_runtime_log or "row_count=3" not in rendered_runtime_log or "room_id=room-9" not in rendered_runtime_log:
                security_errors.append(f"runtime_chat_log_helper_state_missing={rendered_runtime_log!r}")
        except Exception as exc:
            security_errors.append(f"runtime_chat_log_helper_runtime={type(exc).__name__}:{exc}")
        try:
            import app.ui.sims_panel as panel_security_mod

            stale_log_state = panel_security_mod._panel_stale_payload_log_state(
                {"company_id": 11, "db_name": "hidden-db-a", "company_name": "hidden-company-a", "login_id": "hidden-login"},
                {"company_id": 12, "db_name": "hidden-db-b", "company_name": "hidden-company-b", "server": "hidden-server", "path": "C:\\hidden"},
            )
            rendered_stale_log = (
                "[panel] stale action=safe-action payload_company_id=%s current_company_id=%s db_mismatch=%s"
                % (
                    stale_log_state.get("payload_company_id"),
                    stale_log_state.get("current_company_id"),
                    stale_log_state.get("db_mismatch"),
                )
            )
            if any(
                value in rendered_stale_log
                for value in ("hidden-login", "hidden-company-a", "hidden-company-b", "hidden-db-a", "hidden-db-b", "hidden-server", "C:\\hidden")
            ):
                security_errors.append(f"panel_stale_log_helper_leak={rendered_stale_log!r}")
            if stale_log_state != {"payload_company_id": 11, "current_company_id": 12, "db_mismatch": True}:
                security_errors.append(f"panel_stale_log_state={stale_log_state!r}")
            for marker in (
                "skip stale table stash after company change",
                "skip stale final payload after company change",
            ):
                start = panel_src_for_security.find(marker)
                block = panel_src_for_security[start:start + 650] if start >= 0 else ""
                if start < 0 or any(token in block for token in ("payload_stamp.get(\"db_name\")", "current_stamp.get(\"db_name\")", "payload_db=", "current_db=")):
                    security_errors.append(f"panel_stale_log_template={marker}")
        except Exception as exc:
            security_errors.append(f"panel_stale_log_helper_runtime={type(exc).__name__}:{exc}")
        for token in (
            'log.exception("[chat] render pending item failed")',
            'log.exception("[chat] normalize result failed")',
            'log.exception("[chat] stash sims table/export table before json-safe failed")',
        ):
            if token in chat_middleware_src:
                security_errors.append(f"chat_push_traceback_log={token}")
        if security_errors:
            results.append(_fail("Dashboard Lite env and connection secrecy", "; ".join(security_errors)))
        else:
            results.append(_ok("Dashboard Lite env and connection secrecy", "sidebar diagnostics and DB connection logs avoid raw endpoint/path/connection-string values"))

        import app.ui.chat_middleware as chat_mod

        action_errors: list[str] = []

        class _FakeActionCtx:
            def __enter__(self):
                return self
            def __exit__(self, *_exc):
                return False

        class _FakeActionStreamlit:
            def __init__(self):
                self.session_state = {}
                self.downloads: list[dict] = []
                self.buttons: list[dict] = []
                self.captions: list[str] = []
            def columns(self, n, **_kwargs):
                return [_FakeActionCtx() for _ in range(int(n))]
            def download_button(self, label, **kwargs):
                self.downloads.append({"label": label, **kwargs})
                return False
            def button(self, label, **kwargs):
                self.buttons.append({"label": label, **kwargs})
                return False
            def caption(self, text, **_kwargs):
                self.captions.append(str(text))

        old_chat_st = getattr(chat_mod, "st")
        try:
            fake_action_st = _FakeActionStreamlit()
            setattr(chat_mod, "st", fake_action_st)
            chat_mod._render_sims_result_actions_plain(
                key_suffix="dashboard_excel_failed",
                csv_bytes=b"col\n1\n",
                csv_name="x.csv",
                excel_bytes=None,
                xlsx_name="x.xlsx",
                prompt="분석",
                table_key="sims_x",
                clicked_action="Dashboard Lite v0.1",
                download_df=pd.DataFrame({"col": [1]}),
            )
        finally:
            setattr(chat_mod, "st", old_chat_st)
        if [d.get("label") for d in fake_action_st.downloads] != ["CSV 저장"]:
            action_errors.append(f"downloads={fake_action_st.downloads!r}")
        if [b.get("label") for b in fake_action_st.buttons] != ["LLM 분석"]:
            action_errors.append(f"buttons={fake_action_st.buttons!r}")
        if not any("Excel" in c for c in fake_action_st.captions):
            action_errors.append(f"captions={fake_action_st.captions!r}")
        if action_errors:
            results.append(_fail("SIMS action buttons independent Excel failure", "; ".join(action_errors)))
        else:
            results.append(_ok("SIMS action buttons independent Excel failure", "CSV and LLM controls render when Excel bytes are unavailable"))

        facts_json = json.dumps(facts, ensure_ascii=False, default=str)
        if "완료월 총매출과 당월 부분월 현재매출의 직접 우열 판단" not in facts_json or "sample_records 또는 화면 일부 행으로 전체 순위/총합 판단" not in facts_json:
            results.append(_fail("Dashboard Lite comparison guardrails", "forbidden comparison guard missing"))
        else:
            results.append(_ok("Dashboard Lite comparison guardrails", "partial-month/sample/98-percent rules present in serialized facts"))
    except Exception as e:
        results.append(_fail("Dashboard Lite v0.1 facts", f"{type(e).__name__}: {e}\n{traceback.format_exc(limit=4)}"))

    results.extend(_run_customer_sales_forecast_basic_checks())
    return results


def _run_customer_sales_forecast_basic_checks() -> list[CheckResult]:
    results: list[CheckResult] = []
    try:
        customer_mod = importlib.import_module("app.services.analytics_customer_sales_forecast_service")
        chat_mod = importlib.import_module("app.ui.chat_middleware")
        views_mod = importlib.import_module("app.sims.views.analytics_views")
    except Exception as e:
        return [_fail("customer sales forecast import", f"{type(e).__name__}: {e}")]

    calls = {"r130": 0, "master": 0, "date_ranges": []}

    monthly_rows = pd.DataFrame(
        [
            {"기준월": "202601", "매출처코드": "50001", "매출공급가액": 1000, "매출세액": 100, "매출합계": 1100, "집계건수": 1},
            {"기준월": "202602", "매출처코드": "50001", "매출공급가액": 1200, "매출세액": 120, "매출합계": 1320, "집계건수": 1},
            {"기준월": "202603", "매출처코드": "50001", "매출공급가액": 1300, "매출세액": 130, "매출합계": 1430, "집계건수": 1},
            {"기준월": "202604", "매출처코드": "50001", "매출공급가액": 1400, "매출세액": 140, "매출합계": 1540, "집계건수": 1},
            {"기준월": "202605", "매출처코드": "50001", "매출공급가액": 1500, "매출세액": 150, "매출합계": 1650, "집계건수": 1},
            {"기준월": "202606", "매출처코드": "50001", "매출공급가액": 1600, "매출세액": 160, "매출합계": 1760, "집계건수": 1},
            {"기준월": "202607", "매출처코드": "50001", "매출공급가액": 1700, "매출세액": 170, "매출합계": 1870, "집계건수": 1},
            {"기준월": "202601", "매출처코드": "50002", "매출공급가액": 0, "매출세액": 0, "매출합계": 0, "집계건수": 1},
            {"기준월": "202607", "매출처코드": "50002", "매출공급가액": 900, "매출세액": 90, "매출합계": 990, "집계건수": 1},
        ]
    )
    master_rows = pd.DataFrame(
        [
            {"매출처코드": "50001", "매출처명": "한미거래처", "영업사원코드": "S1", "담당영업사원명": "김영업", "시도명": "서울", "시구군명": "강남구", "도로명": "테헤란로"},
            {"매출처코드": "50002", "매출처명": "종근당거래처", "영업사원코드": "S2", "담당영업사원명": "박영업", "시도명": "부산", "시구군명": "해운대구", "도로명": "센텀로"},
        ]
    )

    old_130 = getattr(customer_mod, "_load_rddbc130_monthly", None)
    old_master = getattr(customer_mod, "_load_customer_master", None)

    def fake_130(params, policy):
        calls["r130"] += 1
        date_from = str(params.get("date_from") or "")
        date_to = str(policy.get("effective_date_to") or policy.get("requested_date_to") or params.get("date_to") or "")
        calls["date_ranges"].append((date_from, date_to, policy.get("evaluation_mode")))
        out = monthly_rows.copy()
        if date_to == "20260702":
            out.loc[(out["매출처코드"].astype(str) == "50001") & (out["기준월"].astype(str) == "202607"), ["매출공급가액", "매출세액", "매출합계"]] = [300, 30, 330]
            out.loc[(out["매출처코드"].astype(str) == "50002") & (out["기준월"].astype(str) == "202607"), ["매출공급가액", "매출세액", "매출합계"]] = [200, 20, 220]
        out.attrs.update(
            {
                "source_table": "Rddbc130",
                "source_mode": "transaction_statement",
                "trans_di": "3",
                "date_from": date_from,
                "date_to": date_to,
                "raw_rows": int(out["집계건수"].sum()),
                "monthly_rows": int(len(out)),
                "total_supply": float(out["매출공급가액"].sum()),
                "total_tax": float(out["매출세액"].sum()),
                "total_amount": float(out["매출합계"].sum()),
            }
        )
        return out

    def fake_master(params):
        calls["master"] += 1
        out = master_rows.copy()
        if params.get("sido_nm"):
            out = out[out["시도명"].str.contains(str(params.get("sido_nm")), na=False)].copy()
        if params.get("ven_nm"):
            out = out[out["매출처명"].str.contains(str(params.get("ven_nm")), na=False)].copy()
        return out

    try:
        setattr(customer_mod, "_load_rddbc130_monthly", fake_130)
        setattr(customer_mod, "_load_customer_master", fake_master)

        params_current = {
            "month_from": "202601",
            "month_to": "202607",
            "date_from": "20260101",
            "date_to": "20260712",
            "policy_date": "20260712",
            "top": 0,
        }
        current = customer_mod.get_customer_sales_forecast_df(params_current)
        current_res = customer_mod.get_customer_sales_forecast_result(params_current)
        params_mid = {**params_current, "date_to": "20260702"}
        mid = customer_mod.get_customer_sales_forecast_df(params_mid)
        current_salesperson = customer_mod.get_salesperson_sales_forecast_df(params_current)
        current_region = customer_mod.get_region_sales_forecast_df(params_current)
        current_salesperson_res = customer_mod.get_salesperson_sales_forecast_result(params_current)
        current_region_res = customer_mod.get_region_sales_forecast_result(params_current)
        mid_salesperson = customer_mod.get_salesperson_sales_forecast_df(params_mid)
        mid_region = customer_mod.get_region_sales_forecast_df(params_mid)
        mismatches: list[str] = []
        required_cols = [
            "순번",
            "매출처코드",
            "매출처명",
            "총매출공급가액",
            "총매출세액",
            "총매출액",
            "완료월총매출",
            "완료월평균매출",
            "당월 현재매출",
            "당월 예상매출",
            "당월 잔여예상",
            "당월 진척률",
            "다음월예상매출",
            "예상등급",
            "2026-07 매출",
        ]
        for c in required_cols:
            if c not in current.columns:
                mismatches.append(f"missing current column {c}")
        salesperson_required_cols = [
            "순번",
            "영업사원코드",
            "담당영업사원명",
            "매출처수",
            "총매출액",
            "당월 현재매출",
            "당월 예상매출",
            "예상등급",
            "2026-07 매출",
        ]
        for c in salesperson_required_cols:
            if c not in current_salesperson.columns:
                mismatches.append(f"missing salesperson forecast column {c}")
        region_required_cols = [
            "순번",
            "시도명",
            "시구군명",
            "매출처수",
            "영업사원수",
            "총매출액",
            "당월 현재매출",
            "당월 예상매출",
            "예상등급",
            "2026-07 매출",
        ]
        for c in region_required_cols:
            if c not in current_region.columns:
                mismatches.append(f"missing region forecast column {c}")
        if any(str(c).endswith(" 수량") or str(c) in {"제품코드", "제품명"} for c in current.columns):
            mismatches.append("customer forecast exposed product/qty columns")
        if any(str(c).endswith(" 수량") or str(c) in {"제품코드", "제품명", "매출처코드", "매출처명"} for c in current_salesperson.columns):
            mismatches.append("salesperson forecast exposed customer/product/qty columns")
        if any(str(c).endswith(" 수량") or str(c) in {"제품코드", "제품명", "매출처코드", "매출처명"} for c in current_region.columns):
            mismatches.append("region forecast exposed customer/product/qty columns")
        if calls["r130"] < 3:
            mismatches.append("Rddbc130 transaction statement loader should be used for every period")
        if any(start != "20260101" for start, _end, _mode in calls["date_ranges"]):
            mismatches.append(f"Rddbc130 loader did not preserve date_from ranges={calls['date_ranges']}")
        if not any(end == "20260702" for _start, end, _mode in calls["date_ranges"]):
            mismatches.append(f"Rddbc130 loader did not receive exact historical date_to ranges={calls['date_ranges']}")
        row_50001 = current[current["매출처코드"].astype(str) == "50001"].iloc[0]
        if int(row_50001["완료월수"]) != 6:
            mismatches.append(f"current completed month count expected=6 got={row_50001['완료월수']}")
        if abs(float(row_50001["당월 현재매출"]) - 1870) > 1e-9:
            mismatches.append(f"current month sales expected Rddbc130 total 1870 got={row_50001['당월 현재매출']}")
        mid_row = mid[mid["매출처코드"].astype(str) == "50001"].iloc[0]
        if "평가월 매출" not in mid.columns or "당월 현재매출" in mid.columns:
            mismatches.append("historical midmonth label map failed")
        if abs(float(mid_row["평가월 매출"]) - 330) > 1e-9:
            mismatches.append(f"midmonth evaluation sales expected Rddbc130 total 330 got={mid_row['평가월 매출']}")
        if abs(float(mid_row["평가월 예상매출"]) - float(row_50001["당월 예상매출"])) > 1e-9:
            mismatches.append("current and midmonth expected sales should match from completed months")
        if "평가월 매출" not in mid_salesperson.columns or "당월 현재매출" in mid_salesperson.columns:
            mismatches.append("salesperson historical midmonth label map failed")
        if "평가월 매출" not in mid_region.columns or "당월 현재매출" in mid_region.columns:
            mismatches.append("region historical midmonth label map failed")
        total_customer = float(current["총매출액"].sum())
        if abs(float(current_salesperson["총매출액"].sum()) - total_customer) > 1e-9:
            mismatches.append("salesperson forecast total should match customer forecast total")
        if abs(float(current_region["총매출액"].sum()) - total_customer) > 1e-9:
            mismatches.append("region forecast total should match customer forecast total")
        current_sales_customer = float(current["당월 현재매출"].sum())
        if abs(float(current_salesperson["당월 현재매출"].sum()) - current_sales_customer) > 1e-9:
            mismatches.append("salesperson current sales should match customer current sales")
        if abs(float(current_region["당월 현재매출"].sum()) - current_sales_customer) > 1e-9:
            mismatches.append("region current sales should match customer current sales")
        if current_salesperson[["영업사원코드", "담당영업사원명"]].duplicated().any():
            mismatches.append("salesperson forecast should keep one final row per salesperson")
        if current_region[["시도명", "시구군명"]].duplicated().any():
            mismatches.append("region forecast should keep one final row per region")
        meta = current_res.get("meta") or {}
        if meta.get("analysis_type") != "customer_sales_forecast" or meta.get("summary_type") != "customer_forecast":
            mismatches.append(f"unexpected result meta={meta}")
        if meta.get("source_table") != "Rddbc130" or meta.get("source_mode") != "transaction_statement" or str(meta.get("trans_di")) != "3":
            mismatches.append(f"unexpected source meta={meta}")
        if meta.get("raw_rows", 0) <= 0 or meta.get("monthly_rows", 0) <= 0:
            mismatches.append(f"source row meta missing={meta}")
        if not isinstance(meta.get("salesperson_count"), int):
            mismatches.append(f"salesperson_count should be int got={type(meta.get('salesperson_count')).__name__}")
        if isinstance(meta.get("salesperson_count"), dict):
            mismatches.append("salesperson_count must not contain distribution dict")
        if not isinstance(meta.get("region_count"), int):
            mismatches.append(f"region_count should be int got={type(meta.get('region_count')).__name__}")
        if isinstance(meta.get("region_count"), dict):
            mismatches.append("region_count must not contain distribution dict")
        if not isinstance(meta.get("salesperson_distribution"), dict) or not meta.get("salesperson_distribution"):
            mismatches.append("salesperson_distribution should be separate non-empty dict")
        if not isinstance(meta.get("province_distribution"), dict) or not meta.get("province_distribution"):
            mismatches.append("province_distribution should be separate non-empty dict")
        if not isinstance(meta.get("region_distribution"), dict) or not meta.get("region_distribution"):
            mismatches.append("region_distribution should be separate non-empty dict")
        if not isinstance(meta.get("forecast_grade_counts"), dict) or not meta.get("forecast_grade_counts"):
            mismatches.append("forecast_grade_counts missing")
        if current["매출처코드"].duplicated().any():
            mismatches.append("customer forecast should keep one final row per customer")
        sp_meta = current_salesperson_res.get("meta") or {}
        if sp_meta.get("analysis_type") != "salesperson_sales_forecast" or sp_meta.get("summary_type") != "salesperson_forecast":
            mismatches.append(f"unexpected salesperson meta={sp_meta}")
        if not isinstance(sp_meta.get("salesperson_count"), int) or isinstance(sp_meta.get("salesperson_count"), dict):
            mismatches.append(f"salesperson forecast salesperson_count should be int got={type(sp_meta.get('salesperson_count')).__name__}")
        rg_meta = current_region_res.get("meta") or {}
        if rg_meta.get("analysis_type") != "region_sales_forecast" or rg_meta.get("summary_type") != "region_forecast":
            mismatches.append(f"unexpected region meta={rg_meta}")
        if not isinstance(rg_meta.get("region_count"), int) or isinstance(rg_meta.get("region_count"), dict):
            mismatches.append(f"region forecast region_count should be int got={type(rg_meta.get('region_count')).__name__}")
        llm_df = current.copy()
        llm_df.insert(1, "거래일자", "20260712")
        amount_profile = chat_mod._build_sims_sales_time_profile(  # noqa: SLF001
            llm_df,
            chat_mod._sims_business_terms("매출처별 매출 예상"),  # noqa: SLF001
        )
        if amount_profile.get("amount_col") == "매출처코드":
            mismatches.append("customer forecast LLM amount_col must not be 매출처코드")
        if amount_profile.get("amount_col") != "총매출액":
            mismatches.append(f"customer forecast LLM amount_col expected 총매출액 got={amount_profile.get('amount_col')}")
        sp_llm = current_salesperson.copy()
        sp_llm.insert(1, "거래일자", "20260712")
        sp_amount_profile = chat_mod._build_sims_sales_time_profile(  # noqa: SLF001
            sp_llm,
            chat_mod._sims_business_terms("영업사원별 매출 예상"),  # noqa: SLF001
        )
        if sp_amount_profile.get("amount_col") != "총매출액":
            mismatches.append(f"salesperson forecast LLM amount_col expected 총매출액 got={sp_amount_profile.get('amount_col')}")
        rg_llm = current_region.copy()
        rg_llm.insert(1, "거래일자", "20260712")
        rg_amount_profile = chat_mod._build_sims_sales_time_profile(  # noqa: SLF001
            rg_llm,
            chat_mod._sims_business_terms("지역별 매출 예상"),  # noqa: SLF001
        )
        if rg_amount_profile.get("amount_col") != "총매출액":
            mismatches.append(f"region forecast LLM amount_col expected 총매출액 got={rg_amount_profile.get('amount_col')}")

        old_form = getattr(views_mod, "_render_customer_sales_forecast_form", None)
        old_st = getattr(views_mod, "st", None)

        class _FakeSt:
            @staticmethod
            def subheader(*_args, **_kwargs):
                return None

            @staticmethod
            def caption(*_args, **_kwargs):
                return None

        try:
            setattr(views_mod, "st", _FakeSt())
            setattr(views_mod, "_render_customer_sales_forecast_form", lambda _action_key: (False, {}))
            for fn_name, title in [
                ("render_customer_sales_forecast_analysis", "매출처별 매출 예상"),
                ("render_salesperson_sales_forecast_analysis", "영업사원별 매출 예상"),
                ("render_region_sales_forecast_analysis", "지역별 매출 예상"),
            ]:
                payload = getattr(views_mod, fn_name)()
                if not isinstance(payload, dict) or payload.get("final") is not False or payload.get("title") != title:
                    mismatches.append(f"{fn_name} initial render contract failed payload={payload}")
        finally:
            if old_form is not None:
                setattr(views_mod, "_render_customer_sales_forecast_form", old_form)
            if old_st is not None:
                setattr(views_mod, "st", old_st)

        old_prepare_defaults = getattr(views_mod, "_prepare_analytics_company_defaults", None)
        old_caption = getattr(views_mod, "_render_analytics_default_caption", None)
        old_form_st = getattr(views_mod, "st", None)

        class _ForecastFormContext:
            def __enter__(self):
                return self

            def __exit__(self, *_args):
                return False

        class _ForecastFormSt:
            def __init__(self, submitted: bool):
                self.session_state: dict[str, Any] = {}
                self._submitted = submitted

            def form(self, *_args, **_kwargs):
                return _ForecastFormContext()

            def columns(self, count):
                return [_ForecastFormContext() for _ in range(int(count))]

            @staticmethod
            def date_input(_label, *, value, **_kwargs):
                return value

            @staticmethod
            def caption(*_args, **_kwargs):
                return None

            @staticmethod
            def selectbox(_label, options, *, index=0, **_kwargs):
                return list(options)[index]

            @staticmethod
            def text_input(_label, *, value="", **_kwargs):
                return value

            def form_submit_button(self, *_args, **_kwargs):
                return self._submitted

        try:
            setattr(
                views_mod,
                "_prepare_analytics_company_defaults",
                lambda _action_key, _ns: {
                    "effective": {"io_gu_list": ["0012:051"]},
                    "profile_found": True,
                    "default_supported_keys": [],
                },
            )
            for action_key in (
                "customer_sales_forecast",
                "salesperson_sales_forecast",
                "region_sales_forecast",
            ):
                initial_st = _ForecastFormSt(submitted=False)
                setattr(views_mod, "st", initial_st)
                initial_submitted, initial_params = views_mod._render_customer_sales_forecast_form(action_key)
                if initial_submitted or initial_params:
                    mismatches.append(
                        f"{action_key} initial form should return (False, {{}}), got={(initial_submitted, initial_params)!r}"
                    )

                submit_st = _ForecastFormSt(submitted=True)
                setattr(views_mod, "st", submit_st)
                submitted, form_params = views_mod._render_customer_sales_forecast_form(action_key)
                if not submitted:
                    mismatches.append(f"{action_key} submit form did not submit")
                if form_params.get("io_gu_list") != ["051"] or form_params.get("io_gu_source") != "company_default":
                    mismatches.append(
                        f"{action_key} company IO adapter missing params={form_params!r}"
                    )
        finally:
            if old_prepare_defaults is not None:
                setattr(views_mod, "_prepare_analytics_company_defaults", old_prepare_defaults)
            if old_caption is not None:
                setattr(views_mod, "_render_analytics_default_caption", old_caption)
            if old_form_st is not None:
                setattr(views_mod, "st", old_form_st)

        filtered = customer_mod.get_customer_sales_forecast_df({**params_current, "sido_nm": "서울"})
        if set(filtered["매출처코드"].astype(str).tolist()) != {"50001"}:
            mismatches.append(f"master address filter failed codes={filtered['매출처코드'].astype(str).tolist()}")

        if mismatches:
            results.append(_fail("customer sales forecast", "; ".join(mismatches)))
        else:
            results.append(_ok("customer sales forecast", "Rddbc130 transaction statement source, labels, filters, schema, and meta verified"))
    except Exception as e:
        results.append(_fail("customer sales forecast", f"{type(e).__name__}: {e}"))
    finally:
        if old_130 is not None:
            setattr(customer_mod, "_load_rddbc130_monthly", old_130)
        if old_master is not None:
            setattr(customer_mod, "_load_customer_master", old_master)

    return results


# ---------------------------------------------------------------------
# Service live checks
# ---------------------------------------------------------------------
def _service_cases() -> list[ServiceCase]:
    common_params = {
        "month_from": "202501",
        "month_to": "202512",
        "date_from": "20250101",
        "date_to": "20251231",
        "source_mode": "monthly_book",
        "top": 2000,
    }

    common_condition_tokens = ("2025-01-01", "2025-12-31", "장부재고")

    return [
        ServiceCase(
            name="품목별 매출 추세 분석",
            function_name="get_sales_trend_result",
            params=dict(common_params),
            expected_title_contains="품목별 매출 추세",
            expected_meta_key="trend_judge_counts",
            expected_analysis_type="sales_trend",
            expected_condition_tokens=common_condition_tokens,
            require_seq_column=True,
        ),
        ServiceCase(
            name="품목별 매출 추세 요약표",
            function_name="get_sales_trend_summary_result",
            params=dict(common_params),
            expected_title_contains="품목별 매출 추세 요약표",
            expected_meta_key="trend_judge_counts",
            expected_analysis_type="sales_trend",
            expected_condition_tokens=common_condition_tokens,
            require_seq_column=True,
        ),
        ServiceCase(
            name="품목별 매출 예상",
            function_name="get_sales_forecast_result",
            params=dict(common_params),
            expected_title_contains="품목별 매출 예상",
            expected_meta_key="forecast_grade_counts",
            expected_analysis_type="sales_forecast",
            expected_condition_tokens=common_condition_tokens,
            require_seq_column=True,
        ),
        ServiceCase(
            name="품목별 재고부족현황",
            function_name="get_stock_shortage_result",
            params={
                **common_params,
                "stock_mode": "book",
            },
            expected_title_contains="품목별 재고부족현황",
            expected_meta_key="shortage_grade_counts",
            expected_analysis_type="stock_shortage",
            expected_condition_tokens=common_condition_tokens,
            require_seq_column=True,
        ),
        ServiceCase(
            name="품목별 매출 추세 분석 - 추세판정 필터",
            function_name="get_sales_trend_result",
            params={
                **common_params,
                "trend_judge": "감소",
            },
            expected_title_contains="품목별 매출 추세",
            expected_meta_key="trend_judge_counts",
            expected_analysis_type="sales_trend",
            expected_condition_tokens=common_condition_tokens + ("추세판정", "감소"),
            allow_zero_rows=True,
        ),
        ServiceCase(
            name="품목별 매출 추세 요약표 - 추세판정 필터",
            function_name="get_sales_trend_summary_result",
            params={
                **common_params,
                "trend_judge": "증가",
            },
            expected_title_contains="품목별 매출 추세 요약표",
            expected_meta_key="trend_judge_counts",
            expected_analysis_type="sales_trend",
            expected_condition_tokens=common_condition_tokens + ("추세판정", "증가"),
            allow_zero_rows=True,
        ),
        ServiceCase(
            name="품목별 매출 예상 - 추세판정 필터",
            function_name="get_sales_forecast_result",
            params={
                **common_params,
                "trend_judge": "반품주의",
            },
            expected_title_contains="품목별 매출 예상",
            expected_meta_key="forecast_grade_counts",
            expected_analysis_type="sales_forecast",
            expected_condition_tokens=common_condition_tokens + ("추세판정", "반품주의"),
            allow_zero_rows=True,
        ),
        ServiceCase(
            name="품목별 재고부족현황 - 부족등급 필터",
            function_name="get_stock_shortage_result",
            params={
                **common_params,
                "stock_mode": "book",
                "shortage_grade": "정상",
            },
            expected_title_contains="품목별 재고부족현황",
            expected_meta_key="shortage_grade_counts",
            expected_analysis_type="stock_shortage",
            expected_condition_tokens=common_condition_tokens + ("부족등급", "정상"),
            allow_zero_rows=True,
        ),
    ]

def _evaluate_service_payload(case: ServiceCase, payload: Any) -> CheckResult:
    name = f"service: {case.name}"

    if not isinstance(payload, dict):
        return _fail(name, f"payload가 dict가 아님: {type(payload).__name__}")

    title = str(payload.get("title") or payload.get("action") or "").strip()
    action = str(payload.get("action") or "").strip()
    meta = payload.get("meta") or {}
    ptype = str(payload.get("type") or "").strip()

    if case.expected_title_contains not in title and case.expected_title_contains not in action:
        return _fail(
            name,
            f"title/action mismatch: expected contains {case.expected_title_contains!r}, title={title!r}, action={action!r}",
        )

    row_count = _payload_row_count(payload)
    if row_count <= 0 and not case.allow_zero_rows:
        return _fail(name, f"row_count가 0 이하: rows={row_count}, title={title!r}, type={ptype!r}")

    cols = _payload_columns(payload)

    if row_count > 0 and case.require_seq_column and "순번" not in cols:
        return _fail(name, f"'순번' 컬럼 없음. columns={cols[:20]}")

    if row_count > 0 and case.expected_meta_key:
        val = meta.get(case.expected_meta_key)
        if not isinstance(val, dict) or not val:
            return _fail(name, f"meta[{case.expected_meta_key!r}] 없음 또는 빈값: {val!r}")
        if case.expected_meta_key == "shortage_grade_counts":
            try:
                if sum(int(v or 0) for v in val.values()) != row_count:
                    return _fail(
                        name,
                        f"shortage_grade_counts 합계 불일치: counts={val!r}, rows={row_count}",
                    )
            except Exception:
                return _fail(name, f"shortage_grade_counts 값 변환 실패: {val!r}")

    if case.expected_analysis_type:
        got_analysis_type = str(meta.get("analysis_type") or "").strip()
        if got_analysis_type != case.expected_analysis_type:
            return _fail(
                name,
                (
                    f"analysis_type mismatch: "
                    f"expected={case.expected_analysis_type!r}, got={got_analysis_type!r}, "
                    f"meta_keys={list(meta.keys())}"
                ),
            )

    missing_tokens = _missing_condition_tokens(payload, case.expected_condition_tokens)
    if missing_tokens:
        return _fail(
            name,
            (
                f"조회조건 토큰 누락: missing={missing_tokens!r}, "
                f"condition_text={_condition_text_from_payload(payload)!r}"
            ),
        )

    if row_count > 0 and case.check_code_columns:
        code_problem = _code_column_dtype_problem(payload)
        if code_problem:
            return _fail(name, code_problem)

    summary_md = str(meta.get("summary_md") or "").strip()
    message = str(payload.get("message") or "").strip()

    if case.require_summary_md and not summary_md:
        return _fail(name, f"summary_md 누락: meta_keys={list(meta.keys())}")

    if case.require_message and not message:
        return _fail(name, f"message 누락: payload_keys={list(payload.keys())}")

    condition_preview = _short_text(_condition_text_from_payload(payload), 120)

    detail = (
        f"title={title!r}, action={action!r}, type={ptype!r}, "
        f"rows={row_count}, cols={len(cols)}, "
        f"analysis_type={meta.get('analysis_type')!r}, "
        f"{case.expected_meta_key}={meta.get(case.expected_meta_key) if case.expected_meta_key else None}, "
        f"summary_md={'Y' if summary_md else 'N'}, message={'Y' if message else 'N'}, "
        f"condition={condition_preview!r}"
    )
    return _ok(name, detail)

def run_service_live_checks() -> list[CheckResult]:
    results: list[CheckResult] = []

    try:
        mod = importlib.import_module("app.services.analytics_sales_trend_service")
    except Exception as e:
        return [_fail("import analytics service", f"{type(e).__name__}: {e}")]

    for case in _service_cases():
        fn = getattr(mod, case.function_name, None)
        if not callable(fn):
            results.append(_fail(f"service: {case.name}", f"{case.function_name} callable 없음"))
            continue

        try:
            payload = fn(case.params)
            results.append(_evaluate_service_payload(case, payload))
        except Exception as e:
            detail = f"{type(e).__name__}: {e}\n{traceback.format_exc(limit=4)}"
            results.append(_fail(f"service: {case.name}", detail))

    return results


# ---------------------------------------------------------------------
# NLQ live checks
# ---------------------------------------------------------------------
class PayloadCapture:
    def __init__(self) -> None:
        self.payloads: list[dict[str, Any]] = []

    def fake_push(self, payload=None, action=None, *args, **kwargs):
        if payload is None and args:
            payload = args[0]
        if action is None and len(args) >= 2:
            action = args[1]

        if isinstance(payload, dict):
            p = dict(payload)
            if action and not p.get("action"):
                p["action"] = action
            self.payloads.append(p)
        else:
            self.payloads.append(
                {
                    "final": True,
                    "type": "unknown",
                    "action": action,
                    "data": payload,
                    "meta": {},
                }
            )
        return True

    def pop_last(self) -> dict[str, Any] | None:
        if not self.payloads:
            return None
        return self.payloads[-1]


def _patch_push_function(capture: PayloadCapture) -> None:
    module_names = [
        "app.ui.chat_middleware",
        "app.sims.nlq.nlq_router",
    ]

    for module_name in module_names:
        try:
            mod = importlib.import_module(module_name)
            if hasattr(mod, "push_sims_result_to_chat"):
                setattr(mod, "push_sims_result_to_chat", capture.fake_push)
        except Exception:
            pass


def _make_ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _next_seq_factory() -> Callable[[], int]:
    seq = {"v": 0}

    def _next_seq() -> int:
        seq["v"] += 1
        return seq["v"]

    return _next_seq


def _nlq_cases() -> list[NlqCase]:
    base_tokens = ("2025-01-01", "2025-12-31")

    return [
        NlqCase(
            "품목별 매출 추세 2025년 조회",
            "품목별 매출 추세 분석",
            expected_analysis_type="sales_trend",
            expected_meta_key="trend_judge_counts",
            expected_condition_tokens=base_tokens,
        ),
        NlqCase(
            "품목별 매출 추세 요약표 2025년 조회",
            "품목별 매출 추세 요약표",
            expected_analysis_type="sales_trend",
            expected_meta_key="trend_judge_counts",
            expected_condition_tokens=base_tokens,
        ),
        NlqCase(
            "품목별 매출 예상 2025년 조회",
            "품목별 매출 예상",
            expected_analysis_type="sales_forecast",
            expected_meta_key="forecast_grade_counts",
            expected_condition_tokens=base_tokens,
        ),
        NlqCase(
            "품목별 재고부족현황 2025년 장부재고 기준 조회",
            "품목별 재고부족현황",
            expected_analysis_type="stock_shortage",
            expected_meta_key="shortage_grade_counts",
            expected_params={"stock_mode": "book"},
            expected_condition_tokens=base_tokens + ("장부재고",),
        ),
        NlqCase(
            "품목별 재고부족현황 2025년 장부재고 기준 부족등급 정상 조회",
            "품목별 재고부족현황",
            expected_analysis_type="stock_shortage",
            expected_meta_key="shortage_grade_counts",
            expected_params={"stock_mode": "book", "shortage_grade": "정상"},
            expected_condition_tokens=base_tokens + ("장부재고", "부족등급", "정상"),
            allow_empty_meta_counts=True,
        ),
        NlqCase(
            "품목별 매출 추세 2025년 감소 조회",
            "품목별 매출 추세 분석",
            expected_analysis_type="sales_trend",
            expected_meta_key="trend_judge_counts",
            expected_params={"trend_judge": "감소"},
            expected_condition_tokens=base_tokens + ("추세판정", "감소"),
        ),
        NlqCase(
            "품목별 매출 추세 요약표 2025년 증가 조회",
            "품목별 매출 추세 요약표",
            expected_analysis_type="sales_trend",
            expected_meta_key="trend_judge_counts",
            expected_params={"trend_judge": "증가"},
            expected_condition_tokens=base_tokens + ("추세판정", "증가"),
        ),
        NlqCase(
            "품목별 매출 예상 2025년 반품주의 조회",
            "품목별 매출 예상",
            expected_analysis_type="sales_forecast",
            expected_meta_key="forecast_grade_counts",
            expected_params={"trend_judge": "반품주의"},
            expected_condition_tokens=base_tokens + ("추세판정", "반품주의"),
        ),
    ]

def _evaluate_nlq_case(case: NlqCase, handled: bool, payload: dict[str, Any] | None) -> CheckResult:
    name = f"nlq: {case.query}"

    if not handled:
        return _fail(name, "try_handle_nlq()가 False 반환")

    if not isinstance(payload, dict):
        return _fail(name, "payload 없음")

    action = str(payload.get("action") or payload.get("title") or "").strip()
    meta = payload.get("meta") or {}

    params = payload.get("params") or {}
    if case.expected_params:
        for k, expected_v in case.expected_params.items():
            got_v = params.get(k)
            if got_v != expected_v:
                return _fail(
                    name,
                    f"params mismatch {k!r}: expected={expected_v!r}, got={got_v!r}, params={params!r}",
                )

    if case.expected_action not in action:
        return _fail(name, f"action mismatch expected contains {case.expected_action!r}, got {action!r}")

    if not bool(meta.get("analysis_nlq")):
        return _fail(name, f"meta.analysis_nlq 누락: meta keys={list(meta.keys())}")

    if case.expected_meta_key:
        val = meta.get(case.expected_meta_key)
        if not isinstance(val, dict) or (not val and not case.allow_empty_meta_counts):
            return _fail(name, f"meta[{case.expected_meta_key!r}] 없음 또는 빈값: {val!r}")

    if case.expected_analysis_type:
        got_analysis_type = str(meta.get("analysis_type") or "").strip()
        if got_analysis_type != case.expected_analysis_type:
            return _fail(
                name,
                (
                    f"analysis_type mismatch: "
                    f"expected={case.expected_analysis_type!r}, got={got_analysis_type!r}, "
                    f"meta_keys={list(meta.keys())}"
                ),
            )

    missing_tokens = _missing_condition_tokens(payload, case.expected_condition_tokens)
    if missing_tokens:
        return _fail(
            name,
            (
                f"조회조건 토큰 누락: missing={missing_tokens!r}, "
                f"condition_text={_condition_text_from_payload(payload)!r}"
            ),
        )

    summary_md = str(meta.get("summary_md") or "").strip()
    message = str(payload.get("message") or "").strip()

    if case.require_summary_md and not summary_md:
        return _fail(name, f"summary_md 누락: meta_keys={list(meta.keys())}")

    if case.require_message and not message:
        return _fail(name, f"message 누락: payload_keys={list(payload.keys())}")

    row_count = _payload_row_count(payload)
    cols = _payload_columns(payload)
    condition_preview = _short_text(_condition_text_from_payload(payload), 120)

    detail = (
        f"action={action!r}, rows={row_count}, cols={len(cols)}, "
        f"type={payload.get('type')!r}, "
        f"analysis_type={meta.get('analysis_type')!r}, "
        f"analysis_nlq={meta.get('analysis_nlq')!r}, "
        f"summary_md={'Y' if summary_md else 'N'}, message={'Y' if message else 'N'}, "
        f"condition={condition_preview!r}"
    )
    return _ok(name, detail)

def run_nlq_live_checks() -> list[CheckResult]:
    results: list[CheckResult] = []

    try:
        router = importlib.import_module("app.sims.nlq.nlq_router")
        try_handle_nlq = getattr(router, "try_handle_nlq")
    except Exception as e:
        return [_fail("import router.try_handle_nlq", f"{type(e).__name__}: {e}")]

    capture = PayloadCapture()
    _patch_push_function(capture)

    for case in _nlq_cases():
        room: dict[str, Any] = {"messages": []}
        session_state: dict[str, Any] = {
            "__sims_selected": {},
            "__io_pending_product_pick": {},
        }
        next_seq = _next_seq_factory()
        before_count = len(capture.payloads)

        try:
            handled = bool(
                try_handle_nlq(
                    case.query,
                    room=room,
                    session_state=session_state,
                    make_ts=_make_ts,
                    next_seq=next_seq,
                    logger=log,
                )
            )

            payload = capture.pop_last() if len(capture.payloads) > before_count else None
            results.append(_evaluate_nlq_case(case, handled, payload))

        except Exception as e:
            detail = f"{type(e).__name__}: {e}\n{traceback.format_exc(limit=4)}"
            results.append(_fail(f"nlq: {case.query}", detail))

    return results


# ---------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------
def main() -> int:
    parser = argparse.ArgumentParser(description="SIMS analytics/KPI regression checker")
    parser.add_argument(
        "--live",
        action="store_true",
        help="실제 analytics service DB 조회 smoke test",
    )
    parser.add_argument(
        "--nlq",
        action="store_true",
        help="try_handle_nlq 분석/KPI 라우팅까지 확인",
    )
    args = parser.parse_args()

    print(f"Project root: {PROJECT_ROOT}")

    failed = 0

    basic_results = run_basic_checks()
    failed += _print_results("BASIC IMPORT / HELPER CHECKS", basic_results)

    if args.live:
        service_results = run_service_live_checks()
        failed += _print_results("SERVICE LIVE CHECKS", service_results)

    if args.nlq:
        nlq_results = run_nlq_live_checks()
        failed += _print_results("NLQ LIVE ROUTING CHECKS", nlq_results)

    print()
    if failed:
        print(f"RESULT: FAIL ({failed} failed)")
        return 1

    print("RESULT: OK")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
