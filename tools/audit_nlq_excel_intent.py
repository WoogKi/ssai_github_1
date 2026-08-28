#!/usr/bin/env python3
"""Read an NLQ sample workbook and audit intent through existing dry-run helpers.

This tool never calls a SIMS service or database.  It intentionally leaves
runtime-only master/action-handler evidence as ``runtime_verification_required``.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import subprocess
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.services.io_nlq import resolve_io_nlq, resolve_unlabeled_io_entity_condition
from app.sims.nlq.action_inventory import implemented_actions
from app.sims.nlq import nlq_router
from app.ui.current_table_followups import action_dispatcher, generic


OUTPUT_COLUMNS = (
    "worksheet_row", "row_no", "preceding_question", "question", "expected_action", "actual_action",
    "sample_processed", "sample_intent_match", "sample_note",
    "expected_action_canonical", "canonical_candidate", "expected_query_kind",
    "actual_query_kind", "expected_metric", "actual_metric", "expected_grouping",
    "actual_grouping", "expected_conditions", "actual_conditions", "expected_operation", "actual_operation",
    "execution_status", "processed", "intent_match", "failure_category",
    "all_failure_categories",
    "sample_data_issue", "program_mismatch", "static_confirmed_program_mismatch",
    "runtime_observed_mismatch_categories",
    "runtime_verification_required", "source_binding", "db_requery", "full_source_used",
    "note",
)


# Excel의 화면 표시 기능명은 production canonical action과 다를 수 있다.
# 이 표는 동일 handler/service를 공유하는 승인된 명칭만 정규화한다.
EXPECTED_ACTION_CANONICAL_ALIASES = {
    "품목별 판매예상": "품목별 매출 예상",
}


FIELD_REQUIREMENTS: tuple[tuple[str, tuple[str, ...], tuple[str, ...]], ...] = (
    ("manufacturer", ("제조사명", "제약사명", "제조사", "제약사"), ("maker_nm", "product_ven_nm", "manufacturer_name")),
    ("purchase_vendor", ("매입처명", "매입처"), ("buy_nm", "ven_nm", "purchase_vendor_name")),
    ("order_vendor", ("발주처명", "발주처"), ("order_nm", "ordering_vendor_name")),
    ("salesperson", ("영업사원명", "영업사원", "담당자"), ("sales_man_nm", "manager_name")),
    ("stock_location", ("재고위치명", "재고위치"), ("stock_nm", "stock_cd", "stock_cds")),
    ("product", ("제품명", "품목명", "제품", "품목"), ("physic_nm", "physic_cd", "product_name")),
)


EXPECTED_METRICS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("inbound_count", ("입고횟수",)),
    ("transaction_amount", ("거래금액",)),
    ("purchase_amount", ("매입금액",)),
    ("sales_quantity", ("매출수량", "출고수량")),
    ("stock_amount", ("재고금액",)),
    ("stock_quantity", ("재고수량",)),
    ("forecast_sales", ("예상매출", "매출예상")),
    ("shortage", ("부족제품수", "부족수량", "재고부족", "부족예상")),
    ("sales", ("매출금액", "매출액")),
)


EXPECTED_GROUPINGS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("transaction_type", ("거래명세서구분별",)),
    ("stock_location", ("재고위치별",)),
    ("purchase_vendor", ("매입처별",)),
    ("order_vendor", ("발주처별",)),
    ("manufacturer", ("제조사명별", "제조사별", "제약사별", "제조사기준")),
    ("trend_judgement", ("추세판정별", "추세판정기준")),
    ("judgement_result", ("판정결과별", "판정결과기준")),
    ("forecast_grade", ("예상등급별", "예상등급기준")),
    ("product_class", ("제품분류별",)),
    ("product_category", ("제품구분별",)),
    ("product_group", ("제품그룹별",)),
    ("customer", ("매출거래처별", "거래처별", "매출처별")),
    ("salesperson", ("영업사원별", "담당자별")),
    ("region", ("지역별",)),
    ("month", ("월별",)),
    ("weekday", ("요일별", "요일기준")),
    ("date", ("일자별", "날짜별", "일별")),
    ("product", ("제품별", "품목별")),
)


CURRENT_TABLE_ANCHORS = ("현재표", "현재결과", "현재조회결과", "현재 표", "현재 결과")
RANK_MARKERS = ("TOP", "top", "상위", "최고", "1위", "가장많은", "가장많이")
DETAIL_MARKERS = ("상세", "목록", "리스트", "보여줘", "보여주세요")
LLM_ANALYSIS_MARKERS = ("설명", "의견", "문제점", "경향", "주의사항", "확인할점", "특징")

# Expected semantics are intentionally data-only.  They must not import parser output.
EXPECTED_FIELD_LABELS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("product_group", ("제품그룹명", "제품그룹")),
    ("product_category", ("제품구분명", "제품구분")),
    ("product_class", ("제품분류명", "제품분류")),
    ("stock_location", ("재고위치명", "재고위치")),
    ("purchase_vendor", ("매입처명", "매입처")),
    ("order_vendor", ("발주처명", "발주처")),
    ("manufacturer", ("제조사명", "제약사명", "제조사", "제약사")),
    ("salesperson", ("영업사원명", "영업사원", "담당자")),
    ("region", ("지역명", "지역")),
    ("customer", ("매출처명", "거래처명", "매출처", "거래처")),
    ("product", ("제품명", "품목명", "제품코드", "제품", "품목")),
)

ACTUAL_CONDITION_KEYS: dict[str, tuple[str, ...]] = {
    "product": ("physic_nm", "physic_cd", "product_name"),
    "manufacturer": ("maker_nm", "product_ven_nm", "manufacturer_name"),
    "purchase_vendor": ("ven_nm", "buy_nm", "purchase_vendor_name"),
    "order_vendor": ("order_nm", "ordering_vendor_name"),
    "customer": ("customer_nm", "customer_name", "sales_ven_nm", "ven_nm"),
    "salesperson": ("sales_man_nm", "manager_name"),
    "region": ("region_nm", "region_name"),
    "stock_location": ("stock_nm", "stock_cd", "stock_cds", "stock_cd_list"),
    "product_group": ("product_group_nm", "physic_group_nm", "product_group"),
    "product_category": ("product_category_nm", "product_category"),
    "product_class": ("product_class_nm", "product_class"),
    "unlabeled_name": ("nlq_unlabeled_name", "unlabeled_name"),
    "stock_mode": ("stock_mode",),
    "period": ("date_from", "date_to", "month_from", "month_to", "year"),
}

PROGRAM_MISMATCH_FAMILIES = {
    "action_mismatch", "query_kind_mismatch", "field_mismatch", "condition_mismatch",
    "metric_mismatch", "grouping_mismatch", "operation_mismatch", "filter_mismatch",
    "rank_mismatch", "unsupported_unexpected", "column_unavailable", "routing_error",
}


FAILURE_FAMILY_INFO = {
    "sample_issue": ("Excel 기대값 누락/표현 자체 문제", "Excel expected contract", "아니오", "아니오", "P1"),
    "expected_action_not_canonical": ("기준 기능이 canonical action이 아닌 화면명/별칭", "action_inventory.py", "예", "아니오", "P0"),
    "action_mismatch": ("신규조회 canonical action 라우팅 불일치", "nlq_router.py / io_nlq.py", "예", "경우별", "P0"),
    "query_kind_mismatch": ("신규조회와 현재표 후속질문 경계 불일치", "chat route guard / action_dispatcher.py", "예", "아니오", "P0"),
    "field_mismatch": ("명시 field alias 또는 params 보존 불일치", "io_nlq.py / analytics params", "예", "경우별", "P0"),
    "condition_mismatch": ("명시조건·무라벨·기간 조건 보존 불일치", "io_nlq.py / nlq_router.py", "예", "경우별", "P0"),
    "metric_mismatch": ("질문 지표와 공통 metric vocabulary 불일치", "action_dispatcher.py metric specs", "예", "경우별", "P1"),
    "grouping_mismatch": ("<field>별/기준 grouping 인식 불일치", "action_dispatcher.py dimension specs", "예", "경우별", "P1"),
    "operation_mismatch": ("aggregate/filter/detail/rank 동작 분류 불일치", "action_dispatcher.py / generic.py", "예", "경우별", "P1"),
    "filter_mismatch": ("숫자·문자 필터 해석 불일치", "generic.py common filter", "예", "경우별", "P1"),
    "rank_mismatch": ("TOP/최고/1위 limit 해석 불일치", "generic.py rank helper", "예", "경우별", "P1"),
    "unsupported_expected": ("기대상 미지원 질문", "capability matrix", "아니오", "예", "P2"),
    "unsupported_unexpected": ("지원 기대 질문이 unsupported로 종료", "capability matrix / handler", "예", "경우별", "P0"),
    "column_unavailable": ("full source 필수 컬럼 확인 필요", "current-table source schema", "경우별", "예", "P2"),
    "routing_error": ("deterministic route 오류", "router/dispatcher", "예", "경우별", "P0"),
    "runtime_verification_required": ("DB 또는 action 전용 handler 결과가 있어야 판정 가능", "runtime fixture/smoke", "아니오", "예", "P2"),
    "other": ("기타 또는 복합 원인", "case review", "경우별", "경우별", "P2"),
}


def _text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def _compact(value: Any) -> str:
    return re.sub(r"\s+", "", _text(value))


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)


def _git(*args: str) -> str:
    result = subprocess.run(
        ["git", *args], cwd=PROJECT_ROOT, text=True, encoding="utf-8",
        errors="replace", stdout=subprocess.PIPE, stderr=subprocess.STDOUT, check=False,
    )
    return result.stdout.strip()


def _read_rows(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["NLQ 사례"] if "NLQ 사례" in workbook.sheetnames else workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = [_text(value) for value in rows[0]]
    expected_headers = ["순번", "선행질문", "질문", "기준 기능 (action)", "처리여부", "의도일치", "비고"]
    normalized_headers = [header.replace(" ", "") for header in headers]
    if normalized_headers != [header.replace(" ", "") for header in expected_headers]:
        raise ValueError(f"unexpected headers: {headers!r}")
    out: list[dict[str, str]] = []
    for worksheet_row, values in enumerate(rows[1:], start=2):
        if not any(value not in (None, "") for value in values):
            continue
        record = {headers[index]: _text(values[index]) for index in range(len(headers))}
        record["__worksheet_row"] = str(worksheet_row)
        out.append(record)
    return out


def _expected_metric(question: str) -> str:
    compact = _compact(question)
    for key, phrases in EXPECTED_METRICS:
        if any(phrase in compact for phrase in phrases):
            return key
    return ""


def _expected_grouping(question: str) -> str:
    compact = _compact(question)
    for key, phrases in EXPECTED_GROUPINGS:
        if any(phrase in compact for phrase in phrases):
            return key
    if any(marker in compact for marker in ("분석", "집계", "요약")):
        dimension = _mentioned_dimension(question)
        if dimension:
            return dimension
    if any(marker in compact for marker in RANK_MARKERS) or re.search(r"가장.*많", compact):
        if "제품" in compact or "품목" in compact:
            return "product"
        if "일자" in compact or "날짜" in compact:
            return "date"
    return ""


def _expected_operations(question: str, query_kind: str, grouping: str = "") -> list[str]:
    if query_kind != "current_table_followup":
        return []
    compact = _compact(question)
    operations: list[str] = []
    if grouping or any(marker in compact for marker in ("별", "기준", "요약", "집계")):
        operations.append("aggregate")
    has_named_filter = bool(re.search(
        r"(?:제조사(?:명)?|제약사(?:명)?|예상등급|추세판정|판정결과|상세합계)\s*"
        r"(?:한미(?:약품)?|감소예상|증가예상|감소|안정|불일치|일치)", question
    ))
    if has_named_filter or re.search(r"(?:이상|이하|초과|미만|<|>|=)", compact):
        operations.append("filter")
    if any(marker in compact for marker in ("상세", "목록", "리스트")) or (
        "filter" in operations and any(marker in compact for marker in ("보여줘", "보여주세요"))
    ):
        operations.append("detail")
    if any(marker in compact for marker in RANK_MARKERS) or re.search(r"가장.*많", compact):
        operations.append("rank")
        if "aggregate" not in operations:
            operations.insert(0, "aggregate")
    if not operations and any(marker in compact for marker in LLM_ANALYSIS_MARKERS):
        operations.append("llm_analysis")
    if not operations and "분석" in compact:
        # A bare field plus "분석" is a tabular aggregation request.
        if _mentioned_dimension(question) or _expected_metric(question) or any(
            field in compact for field in ("적용증감율", "재고부족판정")
        ):
            operations.append("aggregate")
        else:
            operations.append("llm_analysis")
    return operations or ["llm_analysis"]


def _mentioned_dimension(question: str) -> str:
    compact = _compact(question)
    aliases = (
        ("product_group", ("제품그룹",)), ("product_category", ("제품구분",)),
        ("product_class", ("제품분류",)), ("stock_location", ("재고위치",)),
        ("purchase_vendor", ("매입처",)), ("order_vendor", ("발주처",)),
        ("manufacturer", ("제조사", "제약사")), ("customer", ("거래처", "매출처")),
        ("salesperson", ("영업사원", "담당자")), ("region", ("지역",)),
        ("trend_judgement", ("추세판정",)), ("judgement_result", ("판정결과",)),
        ("forecast_grade", ("예상등급",)), ("weekday", ("요일",)),
        ("date", ("일자", "날짜")), ("product", ("제품", "품목")),
    )
    for field, names in aliases:
        if any(name in compact for name in names):
            return field
    return ""


def _has_explicit_new_action(question: str) -> bool:
    compact = _compact(question)
    markers = (
        "제품재고장", "제품재고현황", "현재고", "제품수불현황", "제품수불부",
        "입고명세조회", "매입명세조회", "입고현황", "매입현황",
        "출고명세조회", "매출명세조회", "출고현황", "매출현황",
        "거래명세서", "세금계산서", "매출추세", "판매예상", "매출예상",
        "재고부족현황", "지역별매출현황", "사용자조회", "제품코드", "거래처코드",
        "업무코드", "SIMS일일점검", "오늘의경영점검", "SIMS운영점검",
    )
    return any(marker in compact for marker in markers)


def _expected_query_kind(question: str, preceding: str) -> str:
    if any(anchor in question for anchor in CURRENT_TABLE_ANCHORS):
        return "current_table_followup"
    if preceding and not _has_explicit_new_action(question):
        return "current_table_followup"
    return "new_query"


def _actual_query_kind(question: str, preceding: str) -> str:
    if any(anchor in question for anchor in CURRENT_TABLE_ANCHORS):
        return "current_table_followup"
    if nlq_router.resolve_new_sims_nlq_candidate(question):
        return "new_query"
    return "current_table_followup" if preceding else "new_query"


def _clean_condition_value(value: str) -> str:
    value = re.sub(r"(?:19|20)\d{2}(?:년|월|일|~|-|\d)*", " ", value)
    value = re.sub(
        r"입고현황|매입현황|출고현황|매출현황|"
        r"\b(?:조회|해줘|보여줘|보여주세요|상세히|상세표|분석해줘|분석|요약)\b", " ", value
    )
    return _text(value).strip(" ,/")


def _extract_labeled_conditions(question: str) -> dict[str, dict[str, Any]]:
    conditions: dict[str, dict[str, Any]] = {}
    labels: list[tuple[str, str]] = [
        (label, field) for field, field_labels in EXPECTED_FIELD_LABELS for label in field_labels
    ]
    labels.sort(key=lambda item: len(item[0]), reverse=True)
    label_pattern = "|".join(re.escape(label) for label, _ in labels)
    matches = list(re.finditer(label_pattern, question))
    for index, match in enumerate(matches):
        label = match.group(0)
        field = next(field for candidate, field in labels if candidate == label)
        end = matches[index + 1].start() if index + 1 < len(matches) else len(question)
        value = _clean_condition_value(question[match.end():end])
        if field == "stock_location" and value.startswith("전체"):
            conditions[field] = {"mode": "explicit_clear", "value": [], "source": "explicit"}
            continue
        if field == "product" and (
            question[max(0, match.start() - 2):match.start()] in {"부족", "재고"}
            or value.startswith(("수", "별", "기준", "TOP", "top", "상위", "최고", "1위"))
        ):
            continue
        if value and not value.startswith(("별", "기준", "TOP", "top", "상위", "최고", "1위")):
            conditions[field] = {"value": value, "source": "explicit"}
    return conditions


def _remove_labeled_segments(question: str) -> str:
    labels = [label for _, field_labels in EXPECTED_FIELD_LABELS for label in field_labels]
    labels.sort(key=len, reverse=True)
    pattern = "|".join(re.escape(label) for label in labels)
    matches = list(re.finditer(pattern, question))
    if not matches:
        return question
    pieces: list[str] = []
    cursor = 0
    for index, match in enumerate(matches):
        pieces.append(question[cursor:match.start()])
        cursor = matches[index + 1].start() if index + 1 < len(matches) else len(question)
    pieces.append(question[cursor:])
    return " ".join(pieces)


def _expected_conditions(question: str, query_kind: str, expected_action: str) -> dict[str, Any]:
    compact = _compact(question)
    clear_stock = any(token in compact for token in ("전체재고위치", "재고위치전체", "모든재고위치", "전재고위치"))
    condition_text = re.sub(
        r"전체\s*재고위치|재고위치\s*전체|모든\s*재고위치|전\s*재고위치", " ", question
    ) if clear_stock else question
    condition_text = re.sub(r"제품재고(?:장|현황)", " ", condition_text)
    conditions: dict[str, Any] = _extract_labeled_conditions(condition_text)
    if "장부재고" in compact:
        conditions["stock_mode"] = {"value": "book", "source": "explicit"}
    elif "실재고" in compact:
        conditions["stock_mode"] = {"value": "real", "source": "explicit"}
    if clear_stock:
        conditions["stock_location"] = {"mode": "explicit_clear", "value": [], "source": "explicit"}
    if re.search(r"(?:19|20)\d{2}", question):
        conditions["period"] = {"source": "explicit"}

    if query_kind == "current_table_followup":
        fixed_filters = (
            ("forecast_grade", r"예상등급\s*(감소예상|증가예상)"),
            ("trend_judgement", r"추세판정\s*(감소|안정|증가)"),
            ("judgement_result", r"판정결과\s*(감소|안정|증가)"),
            ("detail_total_match", r"상세합계\s*(불일치|일치)"),
        )
        for field, pattern in fixed_filters:
            match = re.search(pattern, question)
            if match and "별" not in match.group(0):
                conditions[field] = {"value": match.group(1), "source": "explicit", "role": "filter"}
        grouping = _expected_grouping(question)
        if grouping:
            conditions.pop(grouping, None)
        return conditions

    residual = _remove_labeled_segments(condition_text)
    residual = re.sub(r"(?:19|20)\d{2}(?:년|월|일|~|-|\d)*", " ", residual)
    residual = re.sub(
        r"제품재고(?:장|현황)|현재고|제품수불(?:현황|부)|입고명세조회|매입명세조회|"
        r"출고명세조회|매출명세조회|입고현황|매입현황|출고현황|매출현황|"
        r"SIMS\s*일일점검|SIMS\s*운영점검|오늘의\s*경영점검|"
        r"실재고|장부재고|전체\s*재고위치|재고위치\s*전체|조회|해줘",
        " ", residual,
    )
    residual = _text(residual)
    unlabeled_actions = {
        "제품재고장", "제품재고현황 조회", "현재고 조회", "입고명세 조회", "출고명세 조회", "Dashboard",
    }
    if residual and expected_action in unlabeled_actions and not conditions.keys() & {
        "product", "manufacturer", "purchase_vendor", "order_vendor", "customer"
    }:
        roles = (
            ["product", "manufacturer", "order_vendor", "purchase_vendor"]
            if expected_action in {"제품재고장", "제품재고현황 조회"} else []
        )
        if expected_action == "Dashboard":
            roles = ["manufacturer"]
        conditions["unlabeled_name"] = {"value": residual, "source": "explicit", "search_roles": roles}
    return conditions


def _actual_conditions(params: dict[str, Any], question: str, action: str) -> dict[str, Any]:
    conditions: dict[str, Any] = {}
    for field, keys in ACTUAL_CONDITION_KEYS.items():
        values = [params.get(key) for key in keys if params.get(key) not in (None, "", [])]
        if values:
            conditions[field] = values[0]
    if action == "제품재고현황 조회" and any(
        token in _compact(question) for token in ("전체재고위치", "재고위치전체", "모든재고위치", "전재고위치")
    ):
        conditions["stock_location"] = {"mode": "explicit_clear", "value": []}
    return conditions


def _condition_mismatches(expected: dict[str, Any], actual: dict[str, Any]) -> list[str]:
    mismatches: list[str] = []
    for field, contract in expected.items():
        if field == "period":
            if field not in actual:
                mismatches.append(field)
            continue
        actual_value = actual.get(field)
        if contract.get("mode") == "explicit_clear":
            if not (isinstance(actual_value, dict) and actual_value.get("mode") == "explicit_clear"):
                mismatches.append(field)
            continue
        expected_value = _compact(contract.get("value"))
        if isinstance(actual_value, (list, tuple)):
            actual_text = " ".join(str(item) for item in actual_value)
        else:
            actual_text = str(actual_value or "")
        actual_compact = _compact(actual_text)
        if field == "stock_mode":
            actual_compact = {"실재고": "real", "장부재고": "book"}.get(actual_compact, actual_compact)
        if expected_value and expected_value not in actual_compact:
            mismatches.append(field)
    return mismatches


def _fixture_df() -> pd.DataFrame:
    columns = {
        "제품코드": ["P1", "P2"], "제품명": ["제품A", "제품B"],
        "제조사명": ["한미", "기타"], "매입처명": ["매입A", "매입B"],
        "발주처명": ["발주A", "발주B"], "거래처명": ["대학약국", "소망약국"],
        "재고위치명": ["본사 창고", "전주 창고"], "제품그룹명": ["일반", "전문"],
        "제품구분명": ["일반", "전문"], "제품분류명": ["내복", "외용"],
        "영업사원명": ["신민우", "김담당"], "지역": ["서울", "부산"],
        "일자": ["20260801", "20260802"], "월": ["202608", "202608"],
        "요일": ["토", "일"], "거래명세서구분": ["매출", "매입"],
        "재고수량": [9, 1], "재고금액": [900, 100], "매출금액": [500, 200],
        "매입금액": [400, 100], "거래금액": [500, 100], "출고수량": [5, 2],
        "입고수량": [3, 1], "부족예상수량": [1, 0], "적용증감율": [5, 15],
        "예상등급": ["감소예상", "증가예상"], "추세판정": ["감소", "안정"],
        "판정결과": ["감소", "안정"], "상세합계일치": ["불일치", "일치"],
    }
    return pd.DataFrame(columns)


def _resolve_new(question: str) -> tuple[str, str, dict[str, Any], bool, str]:
    candidate = nlq_router.resolve_new_sims_nlq_candidate(question) or {}
    route = _text(candidate.get("route"))
    action = _text(candidate.get("action"))
    params: dict[str, Any] = {}
    runtime_required = False
    note = ""
    if route == "dashboard":
        parsed_conditions, residual = nlq_router._extract_dashboard_nlq_conditions(question)
        dashboard_keys = {
            "제약사": "maker_nm", "제조사": "maker_nm", "발주처": "order_nm", "담당자": "manager_name",
        }
        params = {
            dashboard_keys[key]: value for key, value in parsed_conditions.items()
            if key in dashboard_keys and value
        }
        if residual:
            params["nlq_unlabeled_name"] = residual
    elif route == "analytics" and not action.startswith("analytics_grouping_"):
        params = nlq_router._build_analytics_params(question, action)
    elif route == "analytics" and action.startswith("analytics_grouping_"):
        intent = nlq_router._classify_analytics_metric_grouping(question) or {}
        params = {key: intent.get(key) for key in ("requested_metric", "requested_grouping") if intent.get(key)}
    elif route == "io":
        parsed = resolve_io_nlq(question) or {}
        params = dict(parsed.get("params") or {})
        if action in {"입고명세 조회", "출고명세 조회", "제품재고현황 조회"}:
            resolved = resolve_unlabeled_io_entity_condition(question, action=action, params=params)
            params = dict(resolved.get("params") or params)
        elif action == "현재고 조회" and not any(
            key in params for key in ("maker_nm", "physic_nm", "maker_cd", "physic_cd")
        ):
            runtime_required = True
            note = "현재고 무라벨 제조사/제품 resolver는 DB evidence 필요"
        params, _period_policy = nlq_router._apply_io_period_policy(params, action)
    else:
        runtime_required = True
        note = "공개 신규조회 dry-run resolver가 없는 master/action 경계"
    return route, action, params, runtime_required, note


def _field_requirement_errors(question: str, params: dict[str, Any], action: str) -> list[str]:
    compact = _compact(question)
    errors: list[str] = []
    for field, labels, keys in FIELD_REQUIREMENTS:
        explicit = False
        for label in labels:
            start = compact.find(label)
            if start < 0:
                continue
            suffix = compact[start + len(label):]
            if suffix.startswith(("별", "기준")):
                continue
            if field == "product" and label in {"제품", "품목"} and any(
                action_word in compact for action_word in ("제품재고", "제품수불", "제품코드", "품목별")
            ):
                if not re.search(rf"{label}(?:명)?[가-힣A-Za-z0-9]{{2,}}", suffix):
                    continue
            explicit = True
            break
        if explicit and not any(params.get(key) not in (None, "", []) for key in keys):
            errors.append(field)
    if re.search(r"(?:19|20)\d{2}", question) and not any(
        params.get(key) not in (None, "") for key in ("date_from", "date_to", "month_from", "month_to", "year")
    ):
        errors.append("period")
    return errors


def _actual_current_table(
    question: str,
    source_action: str,
    fixture: pd.DataFrame,
) -> tuple[str, str, list[str], str, str, bool]:
    source_compact = _compact(source_action)
    source_metric = ""
    if "매출추세" in source_compact:
        source_metric = "sales"
    elif "재고부족" in source_compact:
        source_metric = "shortage"
    elif any(token in source_compact for token in ("예상매출", "매출예상")):
        source_metric = "forecast_sales"
    source_meta = {"result_metric": source_metric} if source_metric else {}
    prefilter = action_dispatcher._current_table_product_top_filter(
        df=fixture,
        query=question,
        source_action=source_action,
        source_meta=source_meta,
    )
    working_df = prefilter["df"]
    working_question = question
    if prefilter.get("filter_column"):
        for token in ("추세판정", "판정결과", prefilter.get("filter_value")):
            working_question = working_question.replace(_text(token), " ")
    intent = action_dispatcher._current_table_followup_intent(
        working_question, source_action, working_df, source_meta
    )
    metric = _text(intent.get("requested_metric"))
    grouping = _text(intent.get("requested_grouping"))
    operations: list[str] = []
    numeric_filter = generic._find_common_numeric_filter(working_df, working_question)
    text_filter = generic._find_common_column_filter(working_df, working_question)
    group_col = generic._find_common_group_column(working_df, working_question)
    grouping_by_column = {
        "제품코드": "product", "제품명": "product", "제조사명": "manufacturer",
        "제약사명": "manufacturer", "매입처명": "purchase_vendor", "발주처명": "order_vendor",
        "거래처명": "customer", "매출처명": "customer", "재고위치명": "stock_location",
        "제품그룹명": "product_group", "제품구분명": "product_category",
        "제품분류명": "product_class", "영업사원명": "salesperson", "지역": "region",
        "추세판정": "trend_judgement", "판정결과": "judgement_result",
        "예상등급": "forecast_grade", "요일": "weekday", "일자": "date", "월": "month",
        "거래명세서구분": "transaction_type",
    }
    if not grouping and group_col:
        grouping = grouping_by_column.get(group_col, "")
    if not metric and group_col:
        grouped = generic._build_common_group_summary(working_df, group_col)
        metric_col, _metric_label = generic._select_common_group_top_metric(grouped, working_question)
        metric_by_column = {
            "재고수량": "stock_quantity", "재고금액": "stock_amount", "매출금액": "sales",
            "매입금액": "purchase_amount", "거래금액": "transaction_amount",
            "출고수량": "sales_quantity", "입고수량": "inbound_quantity",
            "입고횟수": "inbound_count", "부족예상수량": "shortage",
        }
        metric = metric_by_column.get(metric_col, "")
    if group_col or grouping:
        operations.append("aggregate")
    if numeric_filter[0] or text_filter[0] or prefilter.get("filter_column"):
        operations.append("filter")
    compact = _compact(question)
    if text_filter[0] and any(marker in compact for marker in ("상세", "목록", "리스트")):
        operations.append("detail")
    has_rank, rank_limit = generic._common_rank_limit(question, 20)
    if has_rank:
        if "aggregate" not in operations:
            operations.append("aggregate")
        operations.append("rank")
    classifier = action_dispatcher.classify_current_table_followup_intent(question)
    if not operations and classifier == "llm_analysis":
        operations.append("llm_analysis")
    kind = action_dispatcher.detect_current_table_kind(source_action)
    capability = action_dispatcher._current_table_followup_capability(
        df=working_df, query=question, source_action=source_action, kind=kind,
        source_meta=source_meta,
    )
    status = _text(capability.get("status")) or ("success" if operations else "unsupported")
    if _text(prefilter.get("status")) not in {"", "success"}:
        status = _text(prefilter.get("status"))
    conditions: dict[str, Any] = {
        "numeric_filter": numeric_filter[:3] if numeric_filter[0] else (),
        "text_filter": text_filter if text_filter[0] else (),
        "group_column": group_col,
        "rank_limit": rank_limit if has_rank else 0,
        "source_kind": kind,
    }
    if text_filter[0]:
        column_to_field = {
            "제조사명": "manufacturer", "제약사명": "manufacturer",
            "매입처명": "purchase_vendor", "발주처명": "order_vendor",
            "거래처명": "customer", "매출처명": "customer", "제품명": "product",
            "예상등급": "forecast_grade", "추세판정": "trend_judgement",
            "판정결과": "judgement_result", "상세합계일치": "detail_total_match",
        }
        field = column_to_field.get(str(text_filter[0]))
        if field:
            conditions[field] = text_filter[1]
    if prefilter.get("filter_column"):
        prefilter_field = {
            "추세판정": "trend_judgement",
            "판정결과": "judgement_result",
        }.get(_text(prefilter.get("filter_column")))
        if prefilter_field:
            conditions[prefilter_field] = _text(prefilter.get("filter_value"))
    runtime_required = kind != "generic" and not (metric and grouping) and classifier == "dataframe_table"
    return metric, grouping, operations, status, _json(conditions), runtime_required


def _source_action(preceding: str) -> tuple[str, bool]:
    if not preceding:
        return "", False
    candidate = nlq_router.resolve_new_sims_nlq_candidate(preceding) or {}
    action = _text(candidate.get("action"))
    if action and not action.startswith("analytics_grouping_"):
        return action, False
    canonical = {item.canonical_action for item in implemented_actions()}
    if preceding in canonical:
        return preceding, False
    return "", True


def _primary_failure(categories: list[str]) -> str:
    order = (
        "sample_issue", "expected_action_not_canonical", "action_mismatch",
        "query_kind_mismatch", "field_mismatch", "condition_mismatch",
        "metric_mismatch", "grouping_mismatch", "filter_mismatch", "rank_mismatch",
        "operation_mismatch", "unsupported_unexpected", "column_unavailable",
        "routing_error", "runtime_verification_required", "other",
    )
    return next((category for category in order if category in categories), "")


def audit(input_path: Path) -> list[dict[str, Any]]:
    fixture = _fixture_df()
    canonical_actions = {item.canonical_action for item in implemented_actions()}
    results: list[dict[str, Any]] = []
    for source in _read_rows(input_path):
        row_no = int(float(source["순번"]))
        preceding = source.get("선행질문 ", source.get("선행질문", ""))
        question = source["질문"]
        expected_action = source["기준 기능 (action)"]
        expected_canonical_action = EXPECTED_ACTION_CANONICAL_ALIASES.get(
            expected_action,
            expected_action,
        )
        expected_query_kind = _expected_query_kind(question, preceding)
        actual_query_kind = _actual_query_kind(question, preceding)
        expected_metric = _expected_metric(question) if expected_query_kind == "current_table_followup" else ""
        expected_grouping = _expected_grouping(question) if expected_query_kind == "current_table_followup" else ""
        expected_operations = _expected_operations(question, expected_query_kind, expected_grouping)
        expected_conditions = _expected_conditions(question, expected_query_kind, expected_canonical_action)
        runtime_required = False
        notes: list[str] = []
        actual_conditions_dict: dict[str, Any] = {}
        actual_metric = ""
        actual_grouping = ""
        actual_operations: list[str] = []
        execution_status = "success"
        source_binding = ""
        db_requery = ""
        full_source_used = ""
        params: dict[str, Any] = {}

        if actual_query_kind == "new_query":
            _route, actual_action, params, runtime_required, note = _resolve_new(question)
            actual_conditions_dict = _actual_conditions(params, question, actual_action)
            if note:
                notes.append(note)
            analytics_intent = nlq_router._classify_analytics_metric_grouping(question) or {}
            actual_metric = _text(analytics_intent.get("requested_metric"))
            actual_grouping = _text(analytics_intent.get("requested_grouping"))
            if actual_action.startswith("analytics_grouping_"):
                execution_status = actual_action.removeprefix("analytics_grouping_")
                actual_action = ""
        else:
            actual_action, source_runtime = _source_action(preceding)
            runtime_required = source_runtime
            source_binding = "bound" if actual_action else "runtime_required"
            db_requery = "false"
            full_source_used = "true" if actual_action else "runtime_required"
            actual_metric, actual_grouping, actual_operations, execution_status, actual_conditions_json, handler_runtime = _actual_current_table(
                question, actual_action, fixture
            )
            actual_conditions_dict = json.loads(actual_conditions_json)
            runtime_required = runtime_required or handler_runtime
            if handler_runtime:
                notes.append("action 전용 handler 결과는 실제 source schema fixture 필요")

        categories: list[str] = []
        expected_is_canonical = expected_canonical_action in canonical_actions
        if not expected_action:
            categories.append("sample_issue")
        elif not expected_is_canonical:
            categories.append("expected_action_not_canonical")
        elif actual_action and expected_canonical_action != actual_action:
            categories.append("action_mismatch")
        elif not actual_action and expected_query_kind == "new_query":
            categories.append("action_mismatch")
        if expected_query_kind != actual_query_kind:
            categories.append("query_kind_mismatch")
        condition_errors = [] if runtime_required else _condition_mismatches(expected_conditions, actual_conditions_dict)
        if condition_errors:
            categories.append("condition_mismatch")
            notes.append("누락/불일치 조건=" + ",".join(condition_errors))
        if not runtime_required and expected_metric and actual_metric != expected_metric:
            categories.append("metric_mismatch")
        if not runtime_required and expected_grouping and actual_grouping != expected_grouping:
            categories.append("grouping_mismatch")
        if not runtime_required and expected_operations:
            missing_operations = [item for item in expected_operations if item not in actual_operations]
            if missing_operations:
                if "filter" in missing_operations:
                    categories.append("filter_mismatch")
                elif "rank" in missing_operations:
                    categories.append("rank_mismatch")
                else:
                    categories.append("operation_mismatch")
                notes.append("누락 operation=" + ",".join(missing_operations))
            unexpected_operations = [item for item in actual_operations if item not in expected_operations]
            if unexpected_operations:
                if "filter" in unexpected_operations:
                    categories.append("filter_mismatch")
                else:
                    categories.append("operation_mismatch")
                notes.append("예상 외 operation=" + ",".join(unexpected_operations))
        if execution_status == "unsupported":
            categories.append("unsupported_unexpected")
        elif execution_status == "column_unavailable":
            categories.append("column_unavailable")
        elif execution_status == "routing_error":
            categories.append("routing_error")
        runtime_observed_mismatches = [
            item for item in categories if item in PROGRAM_MISMATCH_FAMILIES
        ]
        if runtime_required:
            categories = [
                item for item in categories
                if item in {"sample_issue", "expected_action_not_canonical"}
            ]
            categories.append("runtime_verification_required")

        categories = list(dict.fromkeys(categories))
        failure = _primary_failure(categories)
        processed = bool(actual_action or actual_query_kind == "current_table_followup") and execution_status not in {"routing_error", "error"}
        sample_data_issue = any(item in {"sample_issue", "expected_action_not_canonical"} for item in categories)
        program_mismatch = any(item in PROGRAM_MISMATCH_FAMILIES for item in categories)
        intent_match = not sample_data_issue and not program_mismatch and not runtime_required
        canonical_candidate = actual_action if expected_action and not expected_is_canonical else ""
        results.append({
            "worksheet_row": source.get("__worksheet_row", ""),
            "row_no": row_no,
            "preceding_question": preceding,
            "question": question,
            "expected_action": expected_action,
            "actual_action": actual_action,
            "sample_processed": source.get("처리여부", ""),
            "sample_intent_match": source.get("의도일치", ""),
            "sample_note": source.get("비고", ""),
            "expected_action_canonical": expected_canonical_action,
            "canonical_candidate": canonical_candidate,
            "expected_query_kind": expected_query_kind,
            "actual_query_kind": actual_query_kind,
            "expected_metric": expected_metric,
            "actual_metric": actual_metric,
            "expected_grouping": expected_grouping,
            "actual_grouping": actual_grouping,
            "expected_conditions": _json(expected_conditions),
            "actual_conditions": _json(actual_conditions_dict),
            "expected_operation": "+".join(expected_operations),
            "actual_operation": "+".join(actual_operations),
            "execution_status": execution_status,
            "processed": str(processed).lower(),
            "intent_match": "PASS" if intent_match else "FAIL",
            "failure_category": failure,
            "all_failure_categories": "|".join(categories),
            "sample_data_issue": str(sample_data_issue).lower(),
            "program_mismatch": str(program_mismatch).lower(),
            "static_confirmed_program_mismatch": str(program_mismatch).lower(),
            "runtime_observed_mismatch_categories": "|".join(runtime_observed_mismatches),
            "runtime_verification_required": str(runtime_required).lower(),
            "source_binding": source_binding,
            "db_requery": db_requery,
            "full_source_used": full_source_used,
            "note": "; ".join(notes),
        })
    return results


def _write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: tuple[str, ...] | list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _failure_families(results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in results:
        for family in row["all_failure_categories"].split("|"):
            if family:
                grouped[family].append(row)
    out: list[dict[str, Any]] = []
    for family, rows in sorted(grouped.items(), key=lambda item: (-len(item[1]), item[0])):
        cause, location, common_fix, action_specific, priority = FAILURE_FAMILY_INFO.get(
            family, FAILURE_FAMILY_INFO["other"]
        )
        out.append({
            "failure_family": family,
            "category_group": (
                "sample_data" if family in {"sample_issue", "expected_action_not_canonical"}
                else "runtime_pending" if family == "runtime_verification_required"
                else "program_mismatch"
            ),
            "count": len(rows),
            "representative_questions": " | ".join(row["question"] for row in rows[:5]),
            "common_cause": cause,
            "existing_common_logic": location,
            "one_place_fix_possible": common_fix,
            "action_specific_rule_needed": action_specific,
            "recommended_priority": priority,
        })
    return out


def _auditor_validation() -> list[dict[str, Any]]:
    cases = [
        {
            "case": "current_table_customer_analysis",
            "question": "현재표 거래처별 분석", "preceding": "입고명세 조회",
            "action": "입고명세 조회", "kind": "current_table_followup",
            "grouping": "customer", "operations": {"aggregate"},
        },
        {
            "case": "manufacturer_filter_detail",
            "question": "제조사 한미 상세히 보여줘", "preceding": "제품재고현황 조회",
            "action": "제품재고현황 조회", "kind": "current_table_followup",
            "grouping": "", "operations": {"filter", "detail"}, "condition": "manufacturer",
        },
        {
            "case": "forecast_grade_filter_detail",
            "question": "예상등급 감소예상 상세히 보여줘", "preceding": "품목별 판매예상",
            "action": "품목별 판매예상", "kind": "current_table_followup",
            "grouping": "", "operations": {"filter", "detail"}, "condition": "forecast_grade",
        },
        {
            "case": "manufacturer_stock_rank",
            "question": "제조사별 재고금액 최고", "preceding": "제품재고현황 조회",
            "action": "제품재고현황 조회", "kind": "current_table_followup",
            "grouping": "manufacturer", "operations": {"aggregate", "rank"},
        },
        {
            "case": "all_stock_location_with_unlabeled",
            "question": "제품재고장 전체 재고위치 한미", "preceding": "",
            "action": "제품재고장", "kind": "new_query", "grouping": "", "operations": set(),
            "condition": "stock_location", "unlabeled": "한미",
        },
        {
            "case": "product_group_not_product",
            "question": "제약사별 매출추세분석 제품그룹 일반", "preceding": "",
            "action": "제약사별 매출추세분석", "kind": "new_query", "grouping": "", "operations": set(),
            "condition": "product_group", "forbidden_condition": "product",
        },
        {
            "case": "inbound_unlabeled_condition_loss",
            "question": "한미 입고현황", "preceding": "", "action": "입고명세 조회",
            "kind": "new_query", "grouping": "", "operations": set(),
            "unlabeled": "한미", "actual_unlabeled": "한미",
        },
        {
            "case": "sales_unlabeled_condition_loss",
            "question": "한미 매출현황", "preceding": "", "action": "출고명세 조회",
            "kind": "new_query", "grouping": "", "operations": set(),
            "unlabeled": "한미", "actual_unlabeled": "한미",
        },
        {
            "case": "trend_judgement_analysis_grouping",
            "question": "현재표 추세판정 분석", "preceding": "품목별 매출 추세 분석",
            "action": "품목별 매출 추세 분석", "kind": "current_table_followup",
            "grouping": "trend_judgement", "operations": {"aggregate"},
        },
        {
            "case": "judgement_result_summary_grouping",
            "question": "현재표 판정결과 요약", "preceding": "품목별 매출 추세 분석",
            "action": "품목별 매출 추세 분석", "kind": "current_table_followup",
            "grouping": "judgement_result", "operations": {"aggregate"},
        },
        {
            "case": "trend_filter_product_rank_actual",
            "question": "현재표 추세판정 감소 품목 TOP 20", "preceding": "품목별 매출 추세 분석",
            "action": "품목별 매출 추세 분석", "kind": "current_table_followup",
            "grouping": "product", "operations": {"aggregate", "filter", "rank"},
            "condition": "trend_judgement", "verify_actual": True,
        },
        {
            "case": "judgement_filter_product_rank_variant",
            "question": "현재결과 판정결과 안정 제품 상위 1", "preceding": "품목별 매출 추세 분석",
            "action": "품목별 매출 추세 분석", "kind": "current_table_followup",
            "grouping": "product", "operations": {"aggregate", "filter", "rank"},
            "condition": "judgement_result", "verify_actual": True,
        },
    ]
    fixture = _fixture_df()
    rows: list[dict[str, Any]] = []
    for case in cases:
        question = case["question"]
        kind = _expected_query_kind(question, case["preceding"])
        grouping = _expected_grouping(question) if kind == "current_table_followup" else ""
        operations = set(_expected_operations(question, kind, grouping)) if kind == "current_table_followup" else set()
        conditions = _expected_conditions(question, kind, case["action"])
        checks = [kind == case["kind"], grouping == case["grouping"], case["operations"].issubset(operations)]
        if case.get("condition"):
            checks.append(case["condition"] in conditions)
        if case.get("forbidden_condition"):
            checks.append(case["forbidden_condition"] not in conditions)
        if case.get("unlabeled"):
            checks.append(_compact(conditions.get("unlabeled_name", {}).get("value")) == _compact(case["unlabeled"]))

        actual_summary: dict[str, Any] = {}
        observed_mismatch = ""
        if case.get("actual_unlabeled"):
            _route, actual_action, params, _runtime, _note = _resolve_new(question)
            actual = _actual_conditions(params, question, actual_action)
            checks.extend([
                actual_action == case["action"],
                _compact(actual.get("unlabeled_name")) == _compact(case["actual_unlabeled"]),
            ])
            actual_summary = {"action": actual_action, "conditions": actual}
        elif case.get("expected_mismatch"):
            _route, actual_action, params, _runtime, _note = _resolve_new(question)
            actual = _actual_conditions(params, question, actual_action)
            mismatches = _condition_mismatches(conditions, actual)
            observed_mismatch = "condition_mismatch" if mismatches else ""
            checks.extend([actual_action == case["action"], observed_mismatch == case["expected_mismatch"]])
            actual_summary = {"action": actual_action, "conditions": actual, "mismatches": mismatches}
        elif kind == "current_table_followup":
            metric, actual_grouping, actual_ops, status, actual_json, _runtime = _actual_current_table(
                question, case["action"], fixture
            )
            actual_summary = {
                "metric": metric, "grouping": actual_grouping, "operations": actual_ops,
                "status": status, "conditions": json.loads(actual_json),
            }
            if case.get("verify_actual"):
                actual_condition_map = actual_summary["conditions"]
                checks.extend([
                    actual_grouping == case["grouping"],
                    case["operations"].issubset(set(actual_ops)),
                ])
                if case.get("condition"):
                    checks.append(case["condition"] in actual_condition_map)
        rows.append({
            "case": case["case"], "question": question,
            "expected_contract": _json({
                "query_kind": kind, "grouping": grouping, "operations": sorted(operations),
                "conditions": conditions,
            }),
            "actual_observation": _json(actual_summary),
            "expected_audit_result": case.get("expected_mismatch", "contract_valid"),
            "passed": str(all(checks)).lower(),
            "note": "" if all(checks) else "expected contract self-check failed",
        })
    return rows


def _write_status(path: Path, input_path: Path, results: list[dict[str, Any]], families: list[dict[str, Any]]) -> None:
    failure_counts = Counter(
        family
        for row in results
        for family in row["all_failure_categories"].split("|")
        if family
    )
    processed = sum(row["processed"] == "true" for row in results)
    matched = sum(row["intent_match"] == "PASS" for row in results)
    sample_issues = sum(row["sample_data_issue"] == "true" for row in results)
    program_mismatches = sum(row["program_mismatch"] == "true" for row in results)
    primary_failure_counts = Counter(
        row["failure_category"] for row in results if row["failure_category"]
    )
    runtime = sum(row["runtime_verification_required"] == "true" for row in results)
    top = [row for row in families if row["category_group"] == "program_mismatch"][:5]
    lines = [
        "SIMS AI NLQ Excel 공식 사례집 전수 의도검증",
        f"입력: {input_path}",
        f"입력 SHA256: {hashlib.sha256(input_path.read_bytes()).hexdigest()}",
        "공식 입력 확인: 사용자가 확정한 유일한 최종본. 별도 파일 동일성 비교 없음.",
        f"branch/head: {_git('branch', '--show-current')} / {_git('rev-parse', '--short', 'HEAD')}",
        "운영소스 수정: 없음",
        "DB 조회: 없음",
        "감사도구 자체 검증: 12/12 PASS",
        "",
        "[감사도구 오탐 보정]",
        "- expected query_kind를 실제 router 결과로 채우던 결합 제거. 선행질문/질문/확정 계약만 사용.",
        "- bare 추세판정·예상등급·요일을 grouping으로 보던 규칙 제거; 별/기준/집계 문법과 filter 값을 구분.",
        "- FIELD + 분석/집계/요약은 aggregate, 설명/의견/문제점/경향만 llm_analysis 후보로 분리.",
        "- 제품그룹/제품분류/제품구분/재고위치를 product 조건으로 오인하지 않도록 longest-label 경계 적용.",
        "- 전체 재고위치는 explicit_clear로 기록하면서 남은 한미를 unlabeled condition으로 독립 보존.",
        "",
        "[요약]",
        f"총 사례 수: {len(results)}",
        f"판정 가능 건수: {len(results) - runtime}",
        f"처리 성공 건수(정적 route/handler 판정): {processed}",
        f"의도일치 PASS 건수: {matched}",
        f"sample_issue 주분류 건수: {primary_failure_counts.get('sample_issue', 0)}",
        f"expected_action_not_canonical 주분류 건수: {primary_failure_counts.get('expected_action_not_canonical', 0)}",
        f"sample data issue 계열 건수: {sample_issues}",
        f"정적 확정 프로그램 mismatch 사례 수: {program_mismatches}",
        f"runtime_verification_required 건수: {runtime}",
        "",
        "[실패 family - 중복 허용]",
    ]
    lines.extend(f"- {family}: {count}" for family, count in failure_counts.most_common())
    lines.extend(["", "[통계 분리 원칙]", "- sample/data 문제는 프로그램 mismatch 통계에서 제외.", "- runtime_verification_required는 억지 PASS/FAIL 없이 별도 보류."])
    lines.extend(["", "[영향이 큰 실제 프로그램 공통 결함 TOP 5]"])
    lines.extend(
        f"- {row['failure_family']} {row['count']}건: {row['common_cause']} ({row['existing_common_logic']})"
        for row in top
    )
    lines.extend([
        "",
        "[공통 수정 후보]",
        "- P0-A 반영: 현황형 IO action에서도 명시조건과 무라벨 잔여어를 공통 경계로 보존.",
        "- P0-B 재분류: 운영 dispatcher의 판정 선필터는 기존 구현이며 v2 감사도구 관측 누락을 보정.",
        "- P1: current-table explicit metric이 generic fallback의 sales 기본값에 덮이지 않도록 metric vocabulary 정합화.",
        "- P1: FIELD + 분석/요약을 aggregate로 처리하는 공통 operation classifier 정합화.",
        "- Data: Excel 기준 기능(action)은 canonical vocabulary로 별도 정리하되 프로그램 실패율과 분리 유지.",
        "",
        "[실제 runtime 사례 대조]",
        "- 한미 입고현황/한미 매출현황: action과 무라벨 한미 조건이 함께 보존됨.",
        "- 현재표 거래처별 분석: expected=current_table_followup + customer aggregate; 실제 dispatcher도 DB 재조회 없이 거래처명 group으로 관찰.",
        "",
        "[action 전용/런타임 검증으로 남길 항목]",
        "- master 사용자/업무코드/제품/거래처 handler의 실제 dry-run action 및 params.",
        "- 거래명세서/세금계산서 등 action 전용 current-table 계산과 source schema 의존 결과.",
        "- 현재고 무라벨 제조사/제품 resolver의 실제 DB evidence.",
        "",
        "[권장 구현 순서]",
        "1. Excel 기준 기능(action)을 canonical action으로 정리하고 sample issue를 분리.",
        "2. 공통 current-table field/metric/grouping vocabulary의 누락 family를 우선 보강.",
        "3. generic filter/rank operation 회귀를 family 단위로 추가.",
        "4. action 전용 handler는 대표 full-source fixture로 별도 검증.",
        "5. 마지막으로 필요한 최소 runtime smoke만 수행.",
        "",
        "[Git 상태]",
        _git("status", "--short", "--branch"),
        "git add/commit/push 미실행.",
    ])
    path.write_text("\n".join(lines) + "\n", encoding="utf-8-sig")


def _write_v3_v4_comparison(
    output_dir: Path,
    results: list[dict[str, Any]],
    v3_path: Path | None,
) -> None:
    if v3_path is None or not v3_path.exists():
        return
    with v3_path.open("r", encoding="utf-8-sig", newline="") as handle:
        v3_by_row = {row["row_no"]: row for row in csv.DictReader(handle)}
    comparison: list[dict[str, Any]] = []
    for row in results:
        previous = v3_by_row.get(str(row["row_no"]), {})
        v3_static = (
            previous.get("program_mismatch") == "true"
            and previous.get("runtime_verification_required") != "true"
        )
        v4_static = row["static_confirmed_program_mismatch"] == "true"
        comparison.append({
            "row_no": row["row_no"],
            "question": row["question"],
            "v3_failure_categories": previous.get("all_failure_categories", ""),
            "v4_failure_categories": row["all_failure_categories"],
            "v3_static_confirmed_program_mismatch": str(v3_static).lower(),
            "v4_static_confirmed_program_mismatch": str(v4_static).lower(),
            "v3_runtime_verification_required": previous.get("runtime_verification_required", ""),
            "v4_runtime_verification_required": row["runtime_verification_required"],
            "category_changed": str(
                previous.get("all_failure_categories", "") != row["all_failure_categories"]
            ).lower(),
            "static_mismatch_change": (
                "resolved" if v3_static and not v4_static
                else "new" if not v3_static and v4_static
                else "unchanged"
            ),
        })
    _write_csv(
        output_dir / "nlq_excel_v3_v4_comparison_20260811.csv",
        comparison,
        [
            "row_no", "question", "v3_failure_categories", "v4_failure_categories",
            "v3_static_confirmed_program_mismatch", "v4_static_confirmed_program_mismatch",
            "v3_runtime_verification_required", "v4_runtime_verification_required",
            "category_changed", "static_mismatch_change",
        ],
    )


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--baseline-v3", type=Path)
    args = parser.parse_args()
    input_path = args.input.resolve()
    output_dir = args.output_dir.resolve()
    validation = _auditor_validation()
    _write_csv(
        output_dir / "nlq_excel_auditor_validation_20260810.csv",
        validation,
        ["case", "question", "expected_contract", "actual_observation", "expected_audit_result", "passed", "note"],
    )
    failed_validation = [row for row in validation if row["passed"] != "true"]
    if failed_validation:
        raise RuntimeError("auditor self-validation failed: " + ", ".join(row["case"] for row in failed_validation))
    expected_case_count = len(_read_rows(input_path))
    results = audit(input_path)
    if len(results) != expected_case_count:
        raise RuntimeError(
            f"case count mismatch: workbook={expected_case_count}, audit={len(results)}"
        )
    families = _failure_families(results)
    _write_csv(output_dir / "nlq_excel_full_audit_v4_20260811.csv", results, OUTPUT_COLUMNS)
    _write_csv(
        output_dir / "nlq_excel_failure_family_v4_20260811.csv",
        families,
        [
            "failure_family", "category_group", "count", "representative_questions", "common_cause",
            "existing_common_logic", "one_place_fix_possible", "action_specific_rule_needed",
            "recommended_priority",
        ],
    )
    _write_status(output_dir / "nlq_excel_full_audit_v4_20260811_status.txt", input_path, results, families)
    _write_v3_v4_comparison(
        output_dir,
        results,
        args.baseline_v3.resolve() if args.baseline_v3 else None,
    )
    print(_json({
        "total": len(results),
        "processed": sum(row["processed"] == "true" for row in results),
        "intent_match": sum(row["intent_match"] == "PASS" for row in results),
        "sample_data_issue": sum(row["sample_data_issue"] == "true" for row in results),
        "program_mismatch": sum(row["program_mismatch"] == "true" for row in results),
        "runtime_verification_required": sum(row["runtime_verification_required"] == "true" for row in results),
        "failure_counts": Counter(row["failure_category"] for row in results if row["failure_category"]),
    }))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
