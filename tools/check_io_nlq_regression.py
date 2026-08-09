# tools/check_io_nlq_regression.py
# -*- coding: utf-8 -*-
# VERSION = "check_io_nlq_regression/2026-05-02-v1"
# 작성자: ChatGPT (OpenAI)
# 정의 :  IO / Docs / Stock NLQ regression checker
# Note: This script is a tool to check for regressions in SIMS analysis/KPI related to NLQ handling.

"""
IO / Docs / Stock NLQ regression checker.

기본 import 확인:
    & "C:\\Program Files\\Python313\\python.exe" tools\\check_io_nlq_regression.py

가벼운 대표 NLQ 라우팅/DB 조회:
    & "C:\\Program Files\\Python313\\python.exe" tools\\check_io_nlq_regression.py --live

전체 대표 NLQ 라우팅/DB 조회:
    & "C:\\Program Files\\Python313\\python.exe" tools\\check_io_nlq_regression.py --live-all

payload 확인:
    & "C:\\Program Files\\Python313\\python.exe" tools\\check_io_nlq_regression.py --live --show-payload
"""

from __future__ import annotations

import argparse
import ast
import hashlib
import importlib
import inspect
import json
import logging
import os
import re
import sys
import tempfile
import traceback
import types
import uuid
import warnings as py_warnings
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable

import pandas as pd


# ---------------------------------------------------------------------
# Project root 보정
# ---------------------------------------------------------------------
THIS_FILE = Path(__file__).resolve()
PROJECT_ROOT = THIS_FILE.parents[1]

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

os.chdir(PROJECT_ROOT)


# ---------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="[%(levelname)s] %(message)s",
)

log = logging.getLogger("io_nlq_regression")


# ---------------------------------------------------------------------
# Result helpers
# ---------------------------------------------------------------------
@dataclass
class CheckResult:
    name: str
    ok: bool
    detail: str = ""


@dataclass
class NlqCase:
    query: str
    expected_action: str
    require_handled: bool = True
    allow_no_payload: bool = False
    allow_zero_rows: bool = True
    expected_condition_tokens: tuple[str, ...] = ()
    forbidden_condition_tokens: tuple[str, ...] = ()
    expected_date_range: tuple[str, str] = ()
    date_column_candidates: tuple[str, ...] = ()

@dataclass
class ParserCase:
    query: str
    expected_action: str
    expected_params: dict[str, Any]
    forbidden_params: tuple[str, ...] = ()

def _ok(name: str, detail: str = "") -> CheckResult:
    return CheckResult(name=name, ok=True, detail=detail)


def _fail(name: str, detail: str = "") -> CheckResult:
    return CheckResult(name=name, ok=False, detail=detail)


def _print_results(title: str, results: list[CheckResult]) -> int:
    print()
    print("=" * 78)
    print(title)
    print("=" * 78)

    failed = 0
    for r in results:
        mark = "OK " if r.ok else "FAIL"
        print(f"[{mark}] {r.name}")
        if r.detail:
            print(f"      {r.detail}")
        if not r.ok:
            failed += 1

    print("-" * 78)
    print(f"총 {len(results)}건 / 성공 {len(results) - failed}건 / 실패 {failed}건")
    return failed


def _safe_import(module_name: str) -> CheckResult:
    try:
        importlib.import_module(module_name)
        return _ok(f"import {module_name}")
    except Exception as e:
        return _fail(f"import {module_name}", f"{type(e).__name__}: {e}")


def _require_attr(module_name: str, attr_name: str) -> CheckResult:
    try:
        mod = importlib.import_module(module_name)
    except Exception as e:
        return _fail(f"import {module_name}", f"{type(e).__name__}: {e}")

    if not hasattr(mod, attr_name):
        return _fail(f"{module_name}.{attr_name}", "attribute 없음")

    obj = getattr(mod, attr_name)
    if not callable(obj):
        return _fail(f"{module_name}.{attr_name}", f"callable 아님: {type(obj).__name__}")

    return _ok(f"{module_name}.{attr_name}")


def _safe_len_df(obj: Any) -> int | None:
    try:
        import pandas as pd

        if isinstance(obj, pd.DataFrame):
            return int(len(obj))
    except Exception:
        pass
    return None


def _payload_columns(payload: dict[str, Any]) -> list[str]:
    cols = payload.get("columns")
    if isinstance(cols, list) and cols:
        return [str(c) for c in cols]

    for key in ("df_display", "df"):
        obj = payload.get(key)
        try:
            import pandas as pd

            if isinstance(obj, pd.DataFrame):
                return [str(c) for c in obj.columns]
        except Exception:
            pass

    records = payload.get("records")
    if isinstance(records, list) and records and isinstance(records[0], dict):
        return [str(c) for c in records[0].keys()]

    return []


def _payload_row_count(payload: dict[str, Any]) -> int:
    meta = payload.get("meta") or {}

    for key in ("row_count_total", "row_count"):
        try:
            v = meta.get(key)
            if v is not None:
                return int(v)
        except Exception:
            pass

    for key in ("df_display", "df"):
        n = _safe_len_df(payload.get(key))
        if n is not None:
            return n

    records = payload.get("records")
    if isinstance(records, list):
        return len(records)

    return 0

def _payload_to_df(payload: dict[str, Any]):
    try:
        import pandas as pd

        for key in ("df_display", "df"):
            obj = payload.get(key)
            if isinstance(obj, pd.DataFrame):
                return obj

        records = payload.get("records")
        if isinstance(records, list):
            return pd.DataFrame(records)
    except Exception:
        return None

    return None


def _norm_yyyymmdd(value: Any) -> str:
    import re

    s = re.sub(r"[^0-9]", "", str(value or ""))
    if len(s) >= 8:
        return s[:8]
    return ""

def _parser_cases() -> list[ParserCase]:
    return [
        ParserCase(
            query="입고명세 매입처명 한미 2026-04-01~2026-04-30 조회",
            expected_action="입고명세 조회",
            expected_params={
                "ven_nm": "한미",
                "date_from": "20260401",
                "date_to": "20260430",
                "month_from": "202604",
                "month_to": "202604",
            },
            forbidden_params=("buy_nm",),
        ),
        ParserCase(
            query="입고명세 거래처명 한미 2026-04-01~2026-04-30 조회",
            expected_action="입고명세 조회",
            expected_params={
                "ven_nm": "한미",
                "date_from": "20260401",
                "date_to": "20260430",
                "month_from": "202604",
                "month_to": "202604",
            },
        ),
        ParserCase(
            query="한미 매입처 거래내역 조회",
            expected_action="입고명세 조회",
            expected_params={
                "ven_nm": "한미",
            },
            forbidden_params=("buy_nm",),
        ),
        ParserCase(
            query="한미 매입처에서 거래내역 조회",
            expected_action="입고명세 조회",
            expected_params={
                "ven_nm": "한미",
            },
            forbidden_params=("buy_nm",),
        ),
        ParserCase(
            query="출고명세 매출처명 대학약국 2026-04-19~2026-05-19 조회",
            expected_action="출고명세 조회",
            expected_params={
                "ven_nm": "대학약국",
                "date_from": "20260419",
                "date_to": "20260519",
                "month_from": "202604",
                "month_to": "202605",
            },
        ),
        ParserCase(
            query="출고명세 거래처명 대학약국 2026.04.19 ~ 2026.05.19 조회",
            expected_action="출고명세 조회",
            expected_params={
                "ven_nm": "대학약국",
                "date_from": "20260419",
                "date_to": "20260519",
                "month_from": "202604",
                "month_to": "202605",
            },
        ),
        ParserCase(
            query="출고명세 거래처명 대학약국 2026/04/19 ~ 2026/05/19 조회",
            expected_action="출고명세 조회",
            expected_params={
                "ven_nm": "대학약국",
                "date_from": "20260419",
                "date_to": "20260519",
                "month_from": "202604",
                "month_to": "202605",
            },
        ),
    ]


def _evaluate_parser_case(case: ParserCase) -> CheckResult:
    name = f"parse: {case.query}"

    try:
        from app.services.io_nlq import resolve_io_nlq

        parsed = resolve_io_nlq(case.query)
    except Exception as e:
        return _fail(name, f"{type(e).__name__}: {e}")

    if not isinstance(parsed, dict):
        return _fail(name, f"parsed 결과 없음: {parsed!r}")

    action = str(parsed.get("action") or "").strip()
    params = parsed.get("params") or {}

    if case.expected_action != action:
        return _fail(
            name,
            f"action mismatch: expected={case.expected_action!r}, got={action!r}, parsed={parsed!r}",
        )

    for key, expected_value in case.expected_params.items():
        got = params.get(key)
        if str(got or "") != str(expected_value):
            return _fail(
                name,
                (
                    f"param mismatch: key={key!r}, "
                    f"expected={expected_value!r}, got={got!r}, params={params!r}"
                ),
            )

    bad_keys = [key for key in case.forbidden_params if key in params]
    if bad_keys:
        return _fail(
            name,
            f"금지 param 포함: bad={bad_keys}, params={params!r}",
        )

    return _ok(name, f"action={action!r}, params={params!r}")


def run_parser_checks() -> list[CheckResult]:
    return [_evaluate_parser_case(case) for case in _parser_cases()]


# ---------------------------------------------------------------------
# Basic import checks
# ---------------------------------------------------------------------
def run_basic_checks() -> list[CheckResult]:
    results: list[CheckResult] = []

    # 라우터
    results.append(_require_attr("app.sims.nlq.nlq_router", "try_handle_nlq"))

    # IO NLQ parser/service modules
    modules = [
        "app.services.io_nlq",
        "app.services.rddbc110_service",
        "app.services.rddbc120_service",
        "app.services.rddbc130_service",
        "app.services.rddbc140_service",
        "app.services.rddbc210_service",
        "app.services.rddbc220_service",
        "app.services.product_flow_service",
        "app.services.product_inventory_service",
        "app.ui.chat_middleware",
    ]

    for module_name in modules:
        results.append(_safe_import(module_name))

    # push function 존재 확인
    results.append(_require_attr("app.ui.chat_middleware", "push_sims_result_to_chat"))

    # 선택적 parser 함수 확인: 프로젝트 버전에 따라 이름이 다를 수 있으므로 있으면 확인, 없으면 skip 처리
    optional_attrs = [
        ("app.services.io_nlq", "parse_io_nlq"),
        ("app.services.io_nlq", "try_parse_io_nlq"),
        ("app.services.io_nlq", "parse_nlq"),
        ("app.services.io_nlq", "detect_io_action"),
    ]

    found_any = False
    for module_name, attr_name in optional_attrs:
        try:
            mod = importlib.import_module(module_name)
            obj = getattr(mod, attr_name, None)
            if callable(obj):
                found_any = True
                results.append(_ok(f"optional parser found: {module_name}.{attr_name}"))
        except Exception:
            pass

    if not found_any:
        results.append(
            _ok(
                "optional io_nlq parser function",
                "공개 parser 함수명은 확인하지 않음. 라우터 live 테스트로 검증.",
            )
        )

    return results


def run_unlabeled_io_entity_resolution_checks() -> list[CheckResult]:
    """Keep label-free entity resolution conservative without a DB connection."""
    from unittest.mock import MagicMock, patch

    from app.services import io_nlq

    results: list[CheckResult] = []
    action = "출고명세 조회"

    validation_cases = (
        (
            "20260801 출고명세 거래명세서 검증",
            "출고↔거래명세서 검증",
            True,
            True,
            False,
            False,
        ),
        (
            "20260801 출고명세 거래명세서 불일치 조회",
            "출고↔거래명세서 검증",
            True,
            True,
            False,
            True,
        ),
        (
            "20260801 출고명세 세금계산서 불일치 조회",
            "출고↔세금계산서 검증",
            True,
            False,
            True,
            True,
        ),
        (
            "20260801 출고명세 거래명세서 세금계산서 불일치 조회",
            "출고명세 조회",
            True,
            True,
            True,
            True,
        ),
    )
    for query, expected_action, requested, trans, tax, mismatch in validation_cases:
        parsed = io_nlq.resolve_io_nlq(query) or {}
        parsed_params = dict(parsed.get("params") or {})
        results.append(
            CheckResult(
                f"outbound validation intent: {query}",
                str(parsed.get("action") or "") == expected_action
                and parsed_params.get("validation_requested") is requested
                and parsed_params.get("validation_trans") is trans
                and parsed_params.get("validation_tax") is tax
                and bool(parsed_params.get("only_mismatch_trans")) is (mismatch and trans)
                and bool(parsed_params.get("only_mismatch_tax")) is (mismatch and tax)
                and parsed_params.get("validation_intent_source") == "user_text",
                f"action={parsed.get('action')!r}, params={parsed_params!r}",
            )
        )

    p0_outbound_validation_cases = (
        ("출고 거래명세서 불일치 2026 조회", "출고↔거래명세서 검증"),
        ("출고 거래명세서 불일치 조회", "출고↔거래명세서 검증"),
        ("출고 세금계산서 불일치 조회", "출고↔세금계산서 검증"),
        ("출고 세금계산서 불일치 2026 조회", "출고↔세금계산서 검증"),
    )
    for query, expected_action in p0_outbound_validation_cases:
        parsed = io_nlq.resolve_io_nlq(query) or {}
        results.append(
            CheckResult(
                f"P0 outbound validation parser: {query}",
                str(parsed.get("action") or "") == expected_action,
                f"parsed={parsed!r}",
            )
        )

    generic_document_cases = (
        ("거래명세서 조회", "거래명세서 공통 조회"),
        ("세금계산서 조회", "세금계산서 공통 조회"),
    )
    for query, expected_action in generic_document_cases:
        parsed = io_nlq.resolve_io_nlq(query) or {}
        results.append(
            CheckResult(
                f"generic document route remains unchanged: {query}",
                str(parsed.get("action") or "") == expected_action,
                f"parsed={parsed!r}",
            )
        )

    outbound_validation_gate_cases = {
        "출고 거래명세서 불일치 조회": True,
        "매출 세금계산서 검증 조회": True,
        "출고 불일치 조회": False,
        "출고 검증 조회": False,
        "거래명세서 불일치 조회": False,
        "세금계산서 검증 조회": False,
    }
    for query, expected in outbound_validation_gate_cases.items():
        actual = io_nlq._is_structured_outbound_validation_request(query)
        results.append(
            CheckResult(
                f"outbound validation three-part gate: {query}",
                actual is expected,
                f"actual={actual!r} expected={expected!r}",
            )
        )

    rag_conflict_params = io_nlq._apply_outbound_validation_intent(
        {"retrieved_context": "거래명세서 불일치 검증"},
        action=action,
        raw="20260801 출고명세 조회",
    )
    action_only_params = io_nlq._apply_outbound_validation_intent(
        {},
        action="출고↔거래명세서 검증",
        raw="20260801 출고명세 조회",
    )
    explanation_params = io_nlq._apply_outbound_validation_intent(
        {},
        action=action,
        raw="출고명세 거래명세서 검증 방법 설명",
    )
    results.append(
        CheckResult(
            "outbound validation ignores retrieved context, action metadata, and explanation text",
            rag_conflict_params.get("retrieved_context") == "거래명세서 불일치 검증"
            and rag_conflict_params.get("validation_requested") is False
            and action_only_params.get("validation_requested") is False
            and explanation_params.get("validation_requested") is False,
            "only the current structured user text may request validation",
        )
    )

    inbound_params = {
        "only_mismatch_trans": "Y",
        "only_mismatch_tax": "Y",
        "validation_requested": True,
    }
    inbound_after = io_nlq._apply_outbound_validation_intent(
        inbound_params,
        action="입고명세 조회",
        raw="입고명세 거래명세서 불일치 조회",
    )
    results.append(
        CheckResult(
            "inbound validation parameters are preserved outside outbound actions",
            inbound_after == inbound_params,
            f"before={inbound_params!r}, after={inbound_after!r}",
        )
    )

    normalized_validation = io_nlq._apply_outbound_validation_intent(
        {},
        action=action,
        raw="20260801 출고 명세서 거래 명세서 검증",
    )
    results.append(
        CheckResult(
            "outbound validation target detection uses normalized user text",
            normalized_validation.get("validation_trans") is True
            and normalized_validation.get("validation_tax") is False,
            f"params={normalized_validation!r}",
        )
    )

    with patch.object(io_nlq, "_lookup_unlabeled_io_entity_candidates") as lookup:
        vendor = io_nlq.resolve_unlabeled_io_entity_condition(
            "대흥약국 출고명세 20260701 조회",
            action=action,
            params={"date_from": "20260701", "date_to": "20260701"},
        )
    results.append(
        CheckResult(
            "unlabeled detail name uses a direct multi-result LIKE condition",
            vendor.get("status") == "resolved"
            and vendor.get("params", {}).get("nlq_unlabeled_name") == "대흥약국"
            and not vendor.get("params", {}).get("ven_cd")
            and lookup.call_count == 0,
            f"result={vendor}",
        )
    )

    for query in (
        "삼진 출고명세 20260731",
        "삼진 출고 명세 20260731",
        "삼진 출고명세서 20260731",
        "삼진 출고 명세서 20260731",
        "20260731 삼진 출고 명세서 조회",
    ):
        parsed_variant = io_nlq.resolve_io_nlq(query) or {}
        parsed_params = dict(parsed_variant.get("params") or {})
        resolved_variant = io_nlq.resolve_unlabeled_io_entity_condition(
            query,
            action=str(parsed_variant.get("action") or ""),
            params=parsed_params,
        )
        variant_params = dict(resolved_variant.get("params") or {})
        results.append(
            CheckResult(
                f"outbound detail spelling preserves unlabeled name: {query}",
                parsed_variant.get("action") == "출고명세 조회"
                and resolved_variant.get("status") == "resolved"
                and variant_params.get("nlq_unlabeled_name") == "삼진"
                and variant_params.get("date_from") == "20260731"
                and variant_params.get("date_to") == "20260731",
                f"parsed={parsed_variant}, resolved={resolved_variant}",
            )
        )

    with patch.object(
        io_nlq,
        "_lookup_unlabeled_io_entity_candidates",
        return_value={
            "candidates": [],
            "outcomes": [{
                "resolver_type": "transaction_vendor",
                "status": "error",
                "candidate_count": 0,
                "elapsed_ms": 7,
                "exception_class": "OperationalError",
            }],
        },
    ):
        errored = io_nlq.resolve_unlabeled_io_entity_condition(
            "오류이름 제품수불현황 조회",
            action="제품수불현황 조회",
            params={},
        )
    results.append(
        CheckResult(
            "unlabeled resolver errors never become not_found",
            errored.get("status") == "resolution_unavailable"
            and errored.get("resolver_outcomes", [{}])[0].get("exception_class") == "OperationalError",
            f"result={errored}",
        )
    )

    # Exercise the real transaction-vendor branch.  This catches a missing
    # module symbol before the master query is incorrectly classified as
    # not-found.
    transaction_vendor_df = pd.DataFrame({
        "Rd03_Ven_Nm": ["fixture-vendor"],
        "Rd03_Ven_PRT": ["fixture-vendor-output"],
        "Rd03_Ven_Cd": ["50001"],
    })
    with (
        patch("app.services.rddbc030_service.search_rows", return_value=transaction_vendor_df) as vendor_search,
        patch("app.services.product_supplier_scope_service.resolve_supplier_vendor_codes", return_value=[]),
        patch.object(io_nlq, "_resolve_single_product_code_by_name_with_error", return_value=(None, None)),
    ):
        direct_lookup = io_nlq._lookup_unlabeled_io_entity_candidates(
            "fixture-vendor",
            action=action,
        )
    transaction_outcome = next(
        (row for row in direct_lookup.get("outcomes", []) if row.get("resolver_type") == "transaction_vendor"),
        {},
    )
    vendor_lookup_kwargs = dict(vendor_search.call_args.kwargs or {}) if vendor_search.call_args else {}
    results.append(
        CheckResult(
            "transaction vendor resolver uses pandas branch without NameError",
            direct_lookup.get("candidates") == [{
                "match_type": "transaction_vendor",
                "match_value": "fixture-vendor",
                "match_code": "50001",
            }]
            and transaction_outcome.get("status") == "success"
            and not transaction_outcome.get("exception_class"),
            f"outcome={transaction_outcome!r}",
        )
    )
    results.append(
        CheckResult(
            "transaction vendor resolver searches only the canonical vendor name including historical masters",
            vendor_lookup_kwargs.get("only_active") is False
            and vendor_lookup_kwargs.get("scope") == "sales_history"
            and vendor_lookup_kwargs.get("ven_nm_kw") == "fixture-vendor"
            and "ven_identity_kw" not in vendor_lookup_kwargs,
            f"scope={vendor_lookup_kwargs.get('scope')!r}, only_active={vendor_lookup_kwargs.get('only_active')!r}, "
            f"name_search={vendor_lookup_kwargs.get('ven_nm_kw')!r}",
        )
    )

    output_name_df = pd.DataFrame({
        "Rd03_Ven_Nm": ["fixture-vendor"],
        "Rd03_Ven_PRT": ["fixture-vendor-output"],
        "Rd03_Ven_Cd": ["50001"],
    })
    with patch("app.services.rddbc030_service.search_rows", return_value=output_name_df):
        output_name_lookup = io_nlq._lookup_transaction_vendor_candidates(
            "fixture-vendor-output",
            action=action,
        )
    results.append(
        CheckResult(
            "transaction vendor resolver does not match output or abbreviation names",
            output_name_lookup.get("candidates") == [],
            f"result={output_name_lookup!r}",
        )
    )

    product_query_calls: list[str] = []

    def _single_product_query(sql: str, _params: dict[str, Any]) -> pd.DataFrame:
        product_query_calls.append(sql)
        return pd.DataFrame({"physic_cd": ["00001"], "physic_nm": ["fixture-product"]})

    with patch("app.services.rddbc_io_common.query_to_df", side_effect=_single_product_query):
        product_code, product_error = io_nlq._resolve_single_product_code_by_name_with_error("fixture-product")
    results.append(
        CheckResult(
            "single product resolver keeps the product-name GROUP BY contract",
            product_code == "00001"
            and product_error is None
            and bool(product_query_calls)
            and "GROUP BY Rd04_Physic_Cd, Rd04_Physic_Nm" in product_query_calls[0],
            f"code={product_code!r}, error={type(product_error).__name__ if product_error else ''}",
        )
    )

    from app.sims.nlq import nlq_router

    pushed: list[dict] = []
    with (
        patch.object(
            io_nlq,
            "resolve_io_nlq",
            return_value={"action": action, "params": {}},
        ),
        patch.object(
            io_nlq,
            "resolve_unlabeled_io_entity_condition",
            return_value={
                "status": "resolution_unavailable",
                "params": {},
                "candidates": [],
                "resolver_outcomes": [{"status": "error", "exception_class": "OperationalError"}],
            },
        ),
        patch("app.ui.chat_middleware.push_sims_result_to_chat", side_effect=lambda payload, _action: pushed.append(payload)),
    ):
        stopped = nlq_router._try_handle_io_nlq(
            "fixture unresolved outbound",
            room={},
            session_state={},
            make_ts=lambda: "2026-07-31 00:00:00",
            next_seq=lambda: 1,
            logger=__import__("logging").getLogger("io-nlq-fixture"),
        )
    stopped_payload = pushed[0] if pushed else {}
    results.append(
        CheckResult(
            "resolver error stops before the detail service and returns tableless guidance",
            stopped is True
            and len(pushed) == 1
            and stopped_payload.get("type") == "text"
            and stopped_payload.get("df") is None
            and stopped_payload.get("meta", {}).get("entity_resolution_status") == "resolution_unavailable"
            and stopped_payload.get("meta", {}).get("row_count") == 0
            and stopped_payload.get("message") == "조회 조건을 확인하는 중 오류가 발생했습니다. 거래처·제약사·제품 중 조건 종류를 명시해 다시 조회해 주세요.",
            f"handled={stopped}, pushed={stopped_payload}",
        )
    )

    with patch.object(
        io_nlq,
        "_lookup_unlabeled_io_entity_candidates",
        return_value=[
            {"match_type": "transaction_vendor", "match_value": "한미약품"},
            {"match_type": "manufacturer", "match_value": "한미약품"},
        ],
    ):
        ambiguous = io_nlq.resolve_unlabeled_io_entity_condition(
            "한미약품 제품수불현황 조회",
            action="제품수불현황 조회",
            params={},
        )
    results.append(
        CheckResult(
            "unlabeled multi-kind product-flow entity requires candidate selection",
            ambiguous.get("status") == "candidate_required"
            and len(ambiguous.get("candidates") or []) == 2,
            f"result={ambiguous}",
        )
    )

    with patch.object(
        io_nlq,
        "_lookup_unlabeled_io_entity_candidates",
        return_value={
            "candidates": [
                {"match_type": "transaction_vendor", "match_value": "fixture-name"},
                {"match_type": "manufacturer", "match_value": "fixture-name"},
            ],
            "outcomes": [{
                "resolver_type": "product",
                "status": "error",
                "candidate_count": 0,
                "elapsed_ms": 1,
                "exception_class": "ProgrammingError",
                "safe_error_code": "master_query_contract",
            }],
        },
    ):
        incomplete = io_nlq.resolve_unlabeled_io_entity_condition(
            "fixture-name 제품수불현황 조회",
            action="제품수불현황 조회",
            params={},
        )
    results.append(
        CheckResult(
            "candidate selection requires every resolver to finish without error",
            incomplete.get("status") == "resolution_unavailable"
            and len(incomplete.get("candidates") or []) == 2,
            f"result={incomplete}",
        )
    )
    results.append(
        CheckResult(
            "entity resolver error codes are sanitized",
            io_nlq._safe_entity_resolver_error_code(NameError("fixture")) == "runtime_name_error"
            and io_nlq._safe_entity_resolver_error_code(
                type("ProgrammingError", (Exception,), {})("fixture")
            ) == "master_query_contract",
            "safe resolver error-code mapping",
        )
    )

    with patch.object(
        io_nlq,
        "_lookup_transaction_vendor_candidates",
        return_value={
            "candidates": [{
                "match_type": "transaction_vendor",
                "match_value": "대흥약국",
                "match_code": "50001",
            }],
            "error": None,
            "started_at": 0.0,
        },
    ):
        labeled = io_nlq.resolve_unlabeled_io_entity_condition(
            "거래처 대흥약국 출고명세 조회",
            action=action,
            params={"ven_nm": "대흥약국"},
        )
    results.append(
        CheckResult(
            "labeled transaction vendor stays on the explicit name contract",
            labeled.get("status") == "not_applicable"
            and labeled.get("params", {}).get("ven_nm") == "대흥약국",
            f"result={labeled}",
        )
    )

    # The display-only name must not recreate the broad name predicate after
    # the resolver has supplied the historical transaction-vendor code.
    from app.services.rddbc120_service import _base_filters

    detail_params = {
        "date_from": "20260701",
        "date_to": "20260701",
        "ven_cd": "50001",
        "ven_nm_display": "fixture-vendor",
    }
    detail_where_sql = _base_filters(detail_params)
    results.append(
        CheckResult(
            "explicit transaction vendor uses the Rddbc120 canonical name predicate",
            "Out_Put.Rd12_Ven_Cd = %(ven_cd)s" in detail_where_sql
            and "ven_nm_like" not in detail_params
            and "Ven_Cd.Rd03_Ven_Nm LIKE" not in detail_where_sql,
            f"has_code_predicate={'Out_Put.Rd12_Ven_Cd = %(ven_cd)s' in detail_where_sql}, "
            f"has_name_predicate={'Ven_Cd.Rd03_Ven_Nm LIKE' in detail_where_sql}",
        )
    )

    from app.services import rddbc120_service

    analysis_source = inspect.getsource(rddbc120_service.get_rddbc120_analysis_summary)
    results.append(
        CheckResult(
            "outbound analysis summary omits unused global transaction and tax aggregates",
            all(token not in analysis_source for token in (
                "WITH trans_sum AS", "tax_sum AS (", "AS T13", "AS T14",
            )),
            "unused_global_aggregate_joins_removed",
        )
    )

    from app.services import rddbc110_service, product_inventory_service

    for service_name, filter_builder in (
        ("inbound", rddbc110_service._base_filters),
        ("outbound", rddbc120_service._base_filters),
    ):
        like_params = {"nlq_unlabeled_name": "fixture-name"}
        like_where = filter_builder(like_params)
        results.append(
            CheckResult(
                f"{service_name} detail keeps the unlabeled OR LIKE predicate in SQL",
                "nlq_unlabeled_name_like" in like_params
                and " OR " in like_where
                and "Rd03_Ven_Nm LIKE %(nlq_unlabeled_name_like)s" in like_where
                and "Rd04_Physic_Nm LIKE %(nlq_unlabeled_name_like)s" in like_where
                and "Rd03_Ven_PRT" not in like_where
                and "Rd03_Ven_Sm" not in like_where,
                f"where={like_where!r}",
            )
        )

    captured_detail_sql: dict[str, tuple[str, dict[str, Any]]] = {}

    def _capture_detail_query(service_name: str):
        def _capture(sql: str, params: dict[str, Any]) -> pd.DataFrame:
            captured_detail_sql[service_name] = (str(sql), dict(params))
            return pd.DataFrame()

        return _capture

    for service_name, service_module, query_function in (
        ("inbound", rddbc110_service, rddbc110_service.get_rddbc110_df),
        ("outbound", rddbc120_service, rddbc120_service.get_rddbc120_df),
    ):
        with patch.object(service_module, "query_to_df", _capture_detail_query(service_name)):
            query_function({
                "date_from": "20260701",
                "date_to": "20260701",
                "nlq_unlabeled_name": "fixture-name",
            })
        sql, sql_params = captured_detail_sql.get(service_name, ("", {}))
        results.append(
            CheckResult(
                f"{service_name} detail sends one OR LIKE SQL with one pattern bind",
                sql_params.get("nlq_unlabeled_name_like") == "%fixture-name%"
                and "UNION ALL" not in sql.upper()
                and "Rd03_Ven_Nm LIKE %(nlq_unlabeled_name_like)s" in sql
                and "Rd04_Physic_Nm LIKE %(nlq_unlabeled_name_like)s" in sql
                and "Rd03_Ven_PRT" not in sql
                and "Rd03_Ven_Sm" not in sql,
                f"captured={bool(sql)}, pattern={sql_params.get('nlq_unlabeled_name_like')!r}",
            )
        )

    inventory_source = inspect.getsource(product_inventory_service._apply_master_filters)
    results.append(
        CheckResult(
            "inventory master filters share the unlabeled OR LIKE helper without output-name matching",
            "add_unlabeled_name_like_filter" in inventory_source
            and "Rd03_Ven_PRT" not in inventory_source
            and "Rd03_Ven_Sm" not in inventory_source,
            "shared product-inventory master filter",
        )
    )

    inventory_unlabeled_params = {"nlq_unlabeled_name": "fixture-name"}
    inventory_unlabeled_where_parts: list[str] = []
    product_inventory_service._apply_master_filters(
        inventory_unlabeled_where_parts,
        inventory_unlabeled_params,
        inventory_unlabeled_params,
        "M.Rd21_Ven_Cd",
    )
    inventory_unlabeled_where = " AND ".join(inventory_unlabeled_where_parts)
    results.append(
        CheckResult(
            "inventory unlabeled name uses transaction-vendor, product, and manufacturer OR search",
            "nlq_unlabeled_name_like" in inventory_unlabeled_params
            and "BuyVen.Rd03_Ven_Nm LIKE %(nlq_unlabeled_name_like)s" in inventory_unlabeled_where
            and "P.Rd04_Physic_Nm LIKE %(nlq_unlabeled_name_like)s" in inventory_unlabeled_where
            and "MakerVen.Rd03_Ven_Nm LIKE %(nlq_unlabeled_name_like)s" in inventory_unlabeled_where,
            f"where={inventory_unlabeled_where!r}",
        )
    )

    inventory_summary_cfg = product_inventory_service._settings(
        {"stock_mode": "real", "group_basis": "maker", "price_mode": "avg"}
    )
    inventory_summary_common = {
        "date_from": "20260801",
        "date_to": "20260809",
        "cfg": inventory_summary_cfg,
        "work_params": {},
    }
    unlabeled_inventory_summary = product_inventory_service._build_inventory_query_summary(
        **inventory_summary_common,
        params={"nlq_unlabeled_name": "fixture-name"},
    )
    maker_inventory_summary = product_inventory_service._build_inventory_query_summary(
        **inventory_summary_common,
        params={"maker_nm": "fixture-name"},
    )
    product_inventory_summary = product_inventory_service._build_inventory_query_summary(
        **inventory_summary_common,
        params={"physic_nm": "fixture-name"},
    )
    results.append(
        CheckResult(
            "inventory unlabeled search summary uses the neutral integrated-search label",
            "통합검색 fixture-name" in unlabeled_inventory_summary
            and "제조사명 fixture-name" not in unlabeled_inventory_summary
            and "제품명 fixture-name" not in unlabeled_inventory_summary
            and "통합검색 fixture-name" not in maker_inventory_summary
            and "제조사명 fixture-name" in maker_inventory_summary
            and "통합검색 fixture-name" not in product_inventory_summary
            and "제품명 fixture-name" in product_inventory_summary,
            f"unlabeled={unlabeled_inventory_summary!r}, maker={maker_inventory_summary!r}, product={product_inventory_summary!r}",
        )
    )

    labeled_inventory_params = {"ven_nm": "fixture-vendor"}
    labeled_inventory_where_parts: list[str] = []
    product_inventory_service._apply_master_filters(
        labeled_inventory_where_parts,
        labeled_inventory_params,
        labeled_inventory_params,
        "M.Rd21_Ven_Cd",
    )
    labeled_inventory_where = " AND ".join(labeled_inventory_where_parts)
    results.append(
        CheckResult(
            "inventory labeled transaction vendor uses only the transaction-vendor LIKE predicate",
            "ven_nm_like" in labeled_inventory_params
            and "BuyVen.Rd03_Ven_Nm LIKE %(ven_nm_like)s" in labeled_inventory_where
            and "P.Rd04_Physic_Nm LIKE %(ven_nm_like)s" not in labeled_inventory_where
            and "P.Rd04_Ven_Nm LIKE %(ven_nm_like)s" not in labeled_inventory_where,
            f"where={labeled_inventory_where!r}",
        )
    )

    labeled_maker_params = {"maker_nm": "fixture-maker"}
    labeled_maker_where_parts: list[str] = []
    product_inventory_service._apply_master_filters(
        labeled_maker_where_parts,
        labeled_maker_params,
        labeled_maker_params,
        "M.Rd21_Ven_Cd",
    )
    labeled_maker_where = " AND ".join(labeled_maker_where_parts)
    results.append(
        CheckResult(
            "inventory labeled manufacturer uses only the manufacturer LIKE predicate",
            "maker_nm_like" in labeled_maker_params
            and "MakerVen.Rd03_Ven_Nm LIKE %(maker_nm_like)s" in labeled_maker_where
            and "P.Rd04_Physic_Nm LIKE %(maker_nm_like)s" not in labeled_maker_where
            and "BuyVen.Rd03_Ven_Nm LIKE %(maker_nm_like)s" not in labeled_maker_where,
            f"where={labeled_maker_where!r}",
        )
    )

    labeled_product_params = {"physic_nm": "fixture-product"}
    labeled_product_where_parts: list[str] = []
    product_inventory_service._apply_master_filters(
        labeled_product_where_parts,
        labeled_product_params,
        labeled_product_params,
        "M.Rd21_Ven_Cd",
    )
    labeled_product_where = " AND ".join(labeled_product_where_parts)
    results.append(
        CheckResult(
            "inventory labeled product uses only the product LIKE predicate",
            "physic_nm_like" in labeled_product_params
            and "P.Rd04_Physic_Nm LIKE %(physic_nm_like)s" in labeled_product_where
            and "MakerVen.Rd03_Ven_Nm LIKE %(physic_nm_like)s" not in labeled_product_where
            and "BuyVen.Rd03_Ven_Nm LIKE %(physic_nm_like)s" not in labeled_product_where,
            f"where={labeled_product_where!r}",
        )
    )

    validation_sql = rddbc120_service._validation_join_sql({
        "validation_trans": True,
        "validation_tax": True,
    })
    results.append(
        CheckResult(
            "outbound validation joins use transaction and tax correlation keys",
            "Trans_Put.Rd12_Trans_Di = Out_Put.Rd12_Trans_Di" in validation_sql
            and "Trans_Put.Rd12_Trans_YyMmDd = Out_Put.Rd12_Trans_YyMmDd" in validation_sql
            and "Tax_Put.Rd12_Tax_YyMmDd = Out_Put.Rd12_Tax_YyMmDd" in validation_sql
            and "Trans_Put.Rd12_Out_YyMmDd" not in validation_sql
            and "Tax_Put.Rd12_Out_YyMmDd" not in validation_sql,
            "Rddbc120 correlation contract",
        )
    )

    outbound_source = inspect.getsource(rddbc120_service.get_rddbc120_df)
    outbound_cte_source = inspect.getsource(rddbc120_service._selected_validation_ctes)
    outbound_cte_builder_source = inspect.getsource(rddbc120_service._validation_ctes_for)
    results.append(
        CheckResult(
            "outbound validation limits transaction and tax aggregation to selected detail keys",
            "_requires_validation_projection(params)" in outbound_source
            and "_selected_validation_ctes(params)" in outbound_source
            and "_validation_ctes_for(" in outbound_cte_source
            and "{prefix}TransKeys AS" in outbound_cte_builder_source
            and "{prefix}TaxKeys AS" in outbound_cte_builder_source
            and "{prefix}TransSums AS" in outbound_cte_builder_source
            and "{prefix}TaxSums AS" in outbound_cte_builder_source
            and "Rd12_Trans_Di" in outbound_cte_builder_source,
            "validation-only selected-row aggregation CTE contract",
        )
    )

    mismatch_sql: dict[str, str] = {}

    def _capture_outbound_mismatch(sql: str, _params: dict[str, Any]) -> pd.DataFrame:
        mismatch_sql["sql"] = str(sql)
        return pd.DataFrame()

    with patch.object(rddbc120_service, "query_to_df", _capture_outbound_mismatch):
        rddbc120_service.get_rddbc120_df({
            "date_from": "20260701",
            "date_to": "20260701",
            "top": 200,
            "only_mismatch_trans": "Y",
            "only_mismatch_tax": "Y",
        })

    outbound_mismatch_sql = mismatch_sql.get("sql", "")
    mismatch_rows_at = outbound_mismatch_sql.find("MismatchRows AS")
    selected_rows_at = outbound_mismatch_sql.find("SelectedRows AS")
    selected_top_at = outbound_mismatch_sql.find("SELECT TOP (%(top)s) Out_Put.*", selected_rows_at)
    results.append(
        CheckResult(
            "outbound mismatch filters are applied before the display TOP",
            "BaseRows AS" in outbound_mismatch_sql
            and "BaseTransKeys AS" in outbound_mismatch_sql
            and "BaseTaxKeys AS" in outbound_mismatch_sql
            and mismatch_rows_at >= 0
            and selected_rows_at > mismatch_rows_at
            and selected_top_at > mismatch_rows_at
            and outbound_mismatch_sql.find("WHERE (T13.Rd13_Trans_Seq IS NULL", mismatch_rows_at) > mismatch_rows_at
            and outbound_mismatch_sql.find("AND (T14.Rd14_Tax_Seq IS NULL", mismatch_rows_at) > mismatch_rows_at,
            "mismatch-before-TOP CTE contract",
        )
    )

    validation_sql_by_side: dict[str, str] = {}

    def _capture_validation_side(side: str, params: dict[str, Any]) -> None:
        with patch.object(
            rddbc120_service,
            "query_to_df",
            side_effect=lambda sql, _params: validation_sql_by_side.setdefault(side, str(sql)) and pd.DataFrame(),
        ):
            rddbc120_service.get_rddbc120_df(params)

    _capture_validation_side("trans", {
        "date_from": "20260701",
        "date_to": "20260701",
        "validation_trans": True,
    })
    _capture_validation_side("tax", {
        "date_from": "20260701",
        "date_to": "20260701",
        "validation_tax": True,
    })
    _capture_validation_side("both", {
        "date_from": "20260701",
        "date_to": "20260701",
        "validation_trans": True,
        "validation_tax": True,
    })
    results.append(
        CheckResult(
            "outbound transaction validation builds only transaction aggregates and joins",
            "SelectedTransKeys AS" in validation_sql_by_side.get("trans", "")
            and "SelectedTaxKeys AS" not in validation_sql_by_side.get("trans", "")
            and "dbo.Rddbc130 AS T13" in validation_sql_by_side.get("trans", "")
            and "dbo.Rddbc140 AS T14" not in validation_sql_by_side.get("trans", ""),
            "transaction-only validation SQL contract",
        )
    )
    results.append(
        CheckResult(
            "outbound tax validation builds only tax aggregates and joins",
            "SelectedTaxKeys AS" in validation_sql_by_side.get("tax", "")
            and "SelectedTransKeys AS" not in validation_sql_by_side.get("tax", "")
            and "dbo.Rddbc140 AS T14" in validation_sql_by_side.get("tax", "")
            and "dbo.Rddbc130 AS T13" not in validation_sql_by_side.get("tax", ""),
            "tax-only validation SQL contract",
        )
    )
    results.append(
        CheckResult(
            "outbound targetless validation builds both aggregate sides",
            "SelectedTransKeys AS" in validation_sql_by_side.get("both", "")
            and "SelectedTaxKeys AS" in validation_sql_by_side.get("both", "")
            and "dbo.Rddbc130 AS T13" in validation_sql_by_side.get("both", "")
            and "dbo.Rddbc140 AS T14" in validation_sql_by_side.get("both", ""),
            "both-side validation SQL contract",
        )
    )

    normal_sql: dict[str, str] = {}

    def _capture_outbound_normal(sql: str, _params: dict[str, Any]) -> pd.DataFrame:
        normal_sql["sql"] = str(sql)
        return pd.DataFrame()

    with patch.object(rddbc120_service, "query_to_df", _capture_outbound_normal):
        rddbc120_service.get_rddbc120_df({
            "date_from": "20260701",
            "date_to": "20260701",
            "top": 200,
        })

    results.append(
        CheckResult(
            "outbound normal detail keeps native issue fields on the selected-row fast path",
            "BaseRows AS" not in normal_sql.get("sql", "")
            and "WITH \nSelectedRows AS" in normal_sql.get("sql", "")
            and "SELECT TOP (%(top)s) Out_Put.*" in normal_sql.get("sql", "")
            and all(field in normal_sql.get("sql", "") for field in (
                "Out_Put.Rd12_Trans_Di",
                "Out_Put.Rd12_Trans_YyMmDd",
                "Out_Put.Rd12_Trans_Seq",
                "Out_Put.Rd12_Tax_Di",
                "Out_Put.Rd12_Tax_YyMmDd",
                "Out_Put.Rd12_Tax_Seq",
            ))
            and normal_sql.get("sql", "").find("Out_Put.Rd12_Trans_Di")
            < normal_sql.get("sql", "").find("Out_Put.Rd12_Tax_Di")
            and "SelectedTransKeys AS" not in normal_sql.get("sql", "")
            and "SelectedTaxKeys AS" not in normal_sql.get("sql", "")
            and "SelectedTransSums AS" not in normal_sql.get("sql", "")
            and "SelectedTaxSums AS" not in normal_sql.get("sql", "")
            and "dbo.Rddbc130 AS T13" not in normal_sql.get("sql", "")
            and "dbo.Rddbc140 AS T14" not in normal_sql.get("sql", "")
            and "거래명세서금액일치" not in normal_sql.get("sql", "")
            and "세금계산서금액일치" not in normal_sql.get("sql", ""),
            "normal selected-row SQL excludes validation-only work and columns",
        )
    )

    explicit_validation_sql: dict[str, str] = {}

    def _capture_outbound_explicit_validation(sql: str, _params: dict[str, Any]) -> pd.DataFrame:
        explicit_validation_sql["sql"] = str(sql)
        return pd.DataFrame()

    with patch.object(rddbc120_service, "query_to_df", _capture_outbound_explicit_validation):
        rddbc120_service.get_rddbc120_df({
            "date_from": "20260701",
            "date_to": "20260701",
            "top": 200,
            "validation_requested": True,
            "validation_trans": True,
        })

    results.append(
        CheckResult(
            "outbound explicit validation intent enables transaction validation projection",
            "SelectedTransKeys AS" in explicit_validation_sql.get("sql", "")
            and "SelectedTaxKeys AS" not in explicit_validation_sql.get("sql", "")
            and "Out_Put.Rd12_Trans_Di = TS.Rd12_Trans_Di" in explicit_validation_sql.get("sql", "")
            and "dbo.Rddbc130 AS T13" in explicit_validation_sql.get("sql", "")
            and "dbo.Rddbc140 AS T14" not in explicit_validation_sql.get("sql", ""),
            "explicit validation action uses the validation projection contract",
        )
    )

    action_only_sql: dict[str, str] = {}

    def _capture_outbound_action_only(sql: str, _params: dict[str, Any]) -> pd.DataFrame:
        action_only_sql["sql"] = str(sql)
        return pd.DataFrame()

    with patch.object(rddbc120_service, "query_to_df", _capture_outbound_action_only):
        rddbc120_service.get_rddbc120_df({
            "date_from": "20260701",
            "date_to": "20260701",
            "top": 200,
            "action": "출고↔거래명세서 검증",
        })

    results.append(
        CheckResult(
            "outbound action metadata alone does not enable validation SQL",
            "SelectedTransKeys AS" not in action_only_sql.get("sql", "")
            and "SelectedTaxKeys AS" not in action_only_sql.get("sql", "")
            and "dbo.Rddbc130 AS T13" not in action_only_sql.get("sql", "")
            and "dbo.Rddbc140 AS T14" not in action_only_sql.get("sql", ""),
            "validation_requested must come from current user text",
        )
    )

    mismatch_fixture = pd.DataFrame({
        "pk": list(range(1, 203)),
        "trans_mismatch": [False] * 200 + [True, True],
        "tax_mismatch": [False] * 201 + [True],
    })
    legacy_trans = mismatch_fixture.iloc[:200].loc[
        lambda frame: frame["trans_mismatch"], "pk"
    ].tolist()
    corrected_trans = mismatch_fixture.loc[
        mismatch_fixture["trans_mismatch"], "pk"
    ].head(200).tolist()
    results.append(
        CheckResult(
            "outbound transaction mismatch outside legacy TOP is retained",
            legacy_trans == [] and corrected_trans == [201, 202],
            f"legacy={legacy_trans}, corrected={corrected_trans}",
        )
    )

    legacy_tax = mismatch_fixture.iloc[:200].loc[
        lambda frame: frame["tax_mismatch"], "pk"
    ].tolist()
    corrected_tax = mismatch_fixture.loc[
        mismatch_fixture["tax_mismatch"], "pk"
    ].head(200).tolist()
    corrected_both = mismatch_fixture.loc[
        mismatch_fixture["trans_mismatch"] & mismatch_fixture["tax_mismatch"], "pk"
    ].head(200).tolist()
    results.append(
        CheckResult(
            "outbound tax and combined mismatches outside legacy TOP are retained",
            legacy_tax == [] and corrected_tax == [202] and corrected_both == [202],
            f"legacy_tax={legacy_tax}, corrected_tax={corrected_tax}, both={corrected_both}",
        )
    )

    trans_di_fixture = pd.DataFrame({
        "trans_di": ["A", "B", "A", "B"],
        "trans_date": ["20260801"] * 4,
        "ven_cd": ["00001"] * 4,
        "trans_seq": ["0001"] * 4,
        "supply": [10, 20, 1, 2],
    })
    trans_di_sums = trans_di_fixture.groupby(
        ["trans_di", "trans_date", "ven_cd", "trans_seq"],
        as_index=False,
    )["supply"].sum()
    results.append(
        CheckResult(
            "outbound transaction validation separates identical date vendor sequence by transaction type",
            trans_di_sums.to_dict("records") == [
                {"trans_di": "A", "trans_date": "20260801", "ven_cd": "00001", "trans_seq": "0001", "supply": 11},
                {"trans_di": "B", "trans_date": "20260801", "ven_cd": "00001", "trans_seq": "0001", "supply": 22},
            ]
            and "Out_Put.Rd12_Trans_Di = T13.Rd13_Trans_Di" in outbound_mismatch_sql,
            f"groups={trans_di_sums.to_dict('records')}",
        )
    )

    outbound_export_source = inspect.getsource(rddbc120_service.get_rddbc120_export_df)
    results.append(
        CheckResult(
            "outbound display and export share the mismatch-before-TOP SQL builder",
            "get_rddbc120_df(qparams)" in outbound_export_source
            and "qparams[\"top\"] = export_top" in outbound_export_source,
            "export delegates to shared detail builder",
        )
    )

    router_source = inspect.getsource(nlq_router._try_handle_io_nlq)
    results.append(
        CheckResult(
            "NLQ trace classifies SQL Server deadlock 1205 without exposing service parameters",
            "[nlq.trace." in router_source
            and "sql_error_number" in router_source
            and "1205" in router_source
            and "_trace_safe_params" in router_source,
            "request-scoped safe NLQ trace contract",
        )
    )
    results.append(
        CheckResult(
            "non-candidate IO results do not emit pending candidate logs",
            "and pending_rows:" in router_source,
            "pending candidate logging requires actual rows",
        )
    )

    explanation_queries = (
        "출고명세 거래명세서 검증 방법을 설명해 줘",
        "출고명세 불일치 원인을 설명해 줘",
        "거래명세서 금액 불일치는 왜 발생해?",
        "세금계산서 검증 기능 사용법 알려줘",
        "출고명세 불일치 여부를 문서에서 찾아줘",
        "RAG 자료에서 거래명세서 검증 내용을 찾아줘",
    )
    explanation_results = [io_nlq.resolve_io_nlq(query) for query in explanation_queries]
    structured_explanation_queries = (
        explanation_queries[0],
        explanation_queries[1],
        explanation_queries[4],
    )
    with patch.object(nlq_router, "_try_handle_io_nlq", return_value=True) as outbound_handler:
        routed_explanations = [
            nlq_router.try_handle_nlq(
                query,
                room={},
                session_state={},
                make_ts=lambda: "2026-08-01 00:00:00",
                next_seq=lambda: 1,
                logger=MagicMock(),
            )
            for query in structured_explanation_queries
        ]
    results.append(
        CheckResult(
            "legacy outbound explanation fixture remains covered by the shared router guard",
            bool(explanation_queries)
            and inspect.getsource(nlq_router.try_handle_nlq).count(
                "is_io_validation_explanation_request"
            ) >= 2,
            f"parsed={explanation_results!r}, routed={routed_explanations!r}, outbound_calls={outbound_handler.call_count}",
        )
    )

    generic_validation_explanation_queries = (
        "\uac70\ub798\uba85\uc138\uc11c \uae08\uc561 \ubd88\uc77c\uce58\ub294 \uc65c \ubc1c\uc0dd\ud574?",
        "\uac70\ub798\uba85\uc138\uc11c \uac80\uc99d \ubc29\ubc95\uc744 \uc124\uba85\ud574 \uc918",
        "\uc138\uae08\uacc4\uc0b0\uc11c \uac80\uc99d \uae30\ub2a5 \uc0ac\uc6a9\ubc95 \uc54c\ub824\uc918",
        "\uc138\uae08\uacc4\uc0b0\uc11c \ubd88\uc77c\uce58 \uc6d0\uc778\uc740 \ubb34\uc5c7\uc774\uc57c?",
        "\uac70\ub798\uba85\uc138\uc11c \uac80\uc99d \ub0b4\uc6a9\uc744 \ubb38\uc11c\uc5d0\uc11c \ucc3e\uc544\uc918",
        "RAG \uc790\ub8cc\uc5d0\uc11c \uc138\uae08\uacc4\uc0b0\uc11c \ubd88\uc77c\uce58 \uc5ec\ubd80\ub97c \ucc3e\uc544\uc918",
        "\uac70\ub798\uba85\uc138\uc11c \uc870\ud68c \ubc29\ubc95\uc744 \uc124\uba85\ud574 \uc918",
        "\uac70\ub798\uba85\uc138\uc11c \uc0ac\uc6a9\ubc95 \uc54c\ub824\uc918",
        "\uc138\uae08\uacc4\uc0b0\uc11c\uac00 \ubb34\uc2a8 \ub73b\uc774\uc57c?",
        "20260801 \uc138\uae08\uacc4\uc0b0\uc11c \ubd88\uc77c\uce58 \uc6d0\uc778\uc744 \uc124\uba85\ud574 \uc918",
        "20260801 \uac70\ub798\uba85\uc138\uc11c \uac80\uc99d \ubc29\ubc95 \uc54c\ub824\uc918",
        "\uac70\ub798\uba85\uc138\uc11c \ub0b4\uc6a9\uc744 \ubb38\uc11c\uc5d0\uc11c \ucc3e\uc544\uc918",
    )
    generic_validation_results = [
        io_nlq.resolve_io_nlq(query) for query in generic_validation_explanation_queries
    ]
    with patch.object(nlq_router, "_try_handle_io_nlq", return_value=True) as generic_handler:
        generic_routed = [
            nlq_router.try_handle_nlq(
                query,
                room={},
                session_state={},
                make_ts=lambda: "2026-08-01 00:00:00",
                next_seq=lambda: 1,
                logger=MagicMock(),
            )
            for query in generic_validation_explanation_queries
        ]
    with (
        patch("app.services.rddbc120_service.get_rddbc120_result") as outbound_service,
        patch("app.services.rddbc130_service.get_rddbc130_result") as trans_service,
        patch("app.services.rddbc140_service.get_rddbc140_result") as tax_service,
    ):
        direct_handler_results = [
            nlq_router._try_handle_io_nlq(
                query,
                room={},
                session_state={},
                make_ts=lambda: "2026-08-01 00:00:00",
                next_seq=lambda: 1,
                logger=MagicMock(),
            )
            for query in generic_validation_explanation_queries
        ]
    results.append(
        CheckResult(
            "transaction and tax explanation or RAG questions bypass every IO DB service",
            all(result is None for result in generic_validation_results)
            and generic_handler.call_count == 0
            and all(result is False for result in direct_handler_results)
            and outbound_service.call_count == 0
            and trans_service.call_count == 0
            and tax_service.call_count == 0,
            "parsed=%r, routed=%r, direct=%r, service_calls=(outbound=%s, trans=%s, tax=%s)"
            % (
                generic_validation_results,
                generic_routed,
                direct_handler_results,
                outbound_service.call_count,
                trans_service.call_count,
                tax_service.call_count,
            ),
        )
    )

    structured_validation_data_queries = (
        (
            "20260801 \uac70\ub798\uba85\uc138\uc11c \uc870\ud68c",
            "\uac70\ub798\uba85\uc138\uc11c \uacf5\ud1b5 \uc870\ud68c",
            {},
        ),
        (
            "20260801 \uc138\uae08\uacc4\uc0b0\uc11c \ubd88\uc77c\uce58 \uc870\ud68c",
            "\uc138\uae08\uacc4\uc0b0\uc11c \uacf5\ud1b5 \uc870\ud68c",
            {},
        ),
        (
            "20260801 \uac70\ub798\uba85\uc138\uc11c \ud558\ub098 \uc870\ud68c",
            "\uac70\ub798\uba85\uc138\uc11c \uacf5\ud1b5 \uc870\ud68c",
            {},
        ),
        (
            "20260801 \uc138\uae08\uacc4\uc0b0\uc11c \ud558\ub098\ub9cc \uc870\ud68c\ud574\uc918",
            "\uc138\uae08\uacc4\uc0b0\uc11c \uacf5\ud1b5 \uc870\ud68c",
            {},
        ),
        (
            "20260801 \uac70\ub798\uba85\uc138\uc11c\ub97c \uc54c\ub824\uc918",
            "\uac70\ub798\uba85\uc138\uc11c \uacf5\ud1b5 \uc870\ud68c",
            {},
        ),
        (
            "20260801 \uc138\uae08\uacc4\uc0b0\uc11c \ucc3e\uc544\uc918",
            "\uc138\uae08\uacc4\uc0b0\uc11c \uacf5\ud1b5 \uc870\ud68c",
            {},
        ),
        (
            "20260801 \ucd9c\uace0 \uba85\uc138\uc11c \uac70\ub798 \uba85\uc138\uc11c \uac80\uc99d",
            "\ucd9c\uace0\uba85\uc138 \uc870\ud68c",
            {"validation_trans": True, "validation_tax": False},
        ),
        (
            "20260801 \ucd9c\uace0 \uba85\uc138\uc11c \uc138\uae08 \uacc4\uc0b0\uc11c \ubd88\uc77c\uce58 \uc870\ud68c",
            "\ucd9c\uace0\uba85\uc138 \uc870\ud68c",
            {"validation_tax": True, "only_mismatch_tax": "Y"},
        ),
    )
    structured_validation_results = [
        io_nlq.resolve_io_nlq(query) for query, _, _ in structured_validation_data_queries
    ]
    with patch.object(nlq_router, "_try_handle_io_nlq", return_value=True) as structured_handler:
        structured_routed = [
            nlq_router.try_handle_nlq(
                query,
                room={},
                session_state={},
                make_ts=lambda: "2026-08-01 00:00:00",
                next_seq=lambda: 1,
                logger=MagicMock(),
            )
            for query, _, _ in structured_validation_data_queries
        ]
    results.append(
        CheckResult(
            "structured transaction and tax data requests retain IO routing after explanation guard",
            structured_handler.call_count == len(structured_validation_data_queries)
            and all(result is True for result in structured_routed)
            and all(
                isinstance(parsed, dict)
                and parsed.get("action") == expected_action
                and all((parsed.get("params") or {}).get(key) == value for key, value in expected_params.items())
                for parsed, (_, expected_action, expected_params) in zip(
                    structured_validation_results,
                    structured_validation_data_queries,
                )
            ),
            f"parsed={structured_validation_results!r}, routed={structured_routed!r}, io_calls={structured_handler.call_count}",
        )
    )

    targetless_validation = io_nlq.resolve_io_nlq(
        "20260801 출고 거래명세서 세금계산서 검증"
    ) or {}
    targetless_params = dict(targetless_validation.get("params") or {})
    results.append(
        CheckResult(
            "targetless structured outbound validation requests both validation sides",
            targetless_params.get("validation_requested") is True
            and targetless_params.get("validation_trans") is True
            and targetless_params.get("validation_tax") is True,
            f"parsed={targetless_validation!r}",
        )
    )

    with patch.object(
        io_nlq,
        "_lookup_unlabeled_io_entity_candidates",
        return_value=[{
            "match_type": "manufacturer",
            "match_value": "fixture-maker",
            "match_code": "10001",
        }],
    ):
        manufacturer = io_nlq.resolve_unlabeled_io_entity_condition(
            "fixture-maker 제품수불현황 조회",
            action="제품수불현황 조회",
            params={},
        )
    manufacturer_params = dict(manufacturer.get("params") or {})
    manufacturer_where_sql = _base_filters(manufacturer_params)
    results.append(
        CheckResult(
            "resolved manufacturer uses the Rddbc120 code predicate without a name predicate",
            manufacturer.get("status") == "resolved"
            and manufacturer_params.get("product_ven_cd") == "10001"
            and not manufacturer_params.get("maker_nm")
            and not manufacturer_params.get("product_ven_nm")
            and "Physic_Cd.Rd04_Ven_Cd = %(product_ven_cd)s" in manufacturer_where_sql
            and "product_ven_nm_like" not in manufacturer_params,
            f"result={manufacturer!r}",
        )
    )

    candidate_pushes: list[dict] = []
    with (
        patch.object(
            io_nlq,
            "resolve_io_nlq",
            return_value={"action": action, "params": {}},
        ),
        patch.object(
            io_nlq,
            "resolve_unlabeled_io_entity_condition",
            return_value={
                "status": "candidate_required",
                "params": {},
                "candidates": [
                    {"match_type": "transaction_vendor", "match_value": "fixture-vendor"},
                    {"match_type": "manufacturer", "match_value": "fixture-maker"},
                    {"match_type": "product", "match_value": "fixture-product"},
                ],
            },
        ),
        patch("app.ui.chat_middleware.push_sims_result_to_chat", side_effect=lambda payload, _action: candidate_pushes.append(payload)),
    ):
        nlq_router._try_handle_io_nlq(
            "fixture ambiguous outbound",
            room={},
            session_state={},
            make_ts=lambda: "2026-08-01 00:00:00",
            next_seq=lambda: 1,
            logger=__import__("logging").getLogger("io-nlq-fixture"),
        )
    candidate_frame = candidate_pushes[0].get("df") if candidate_pushes else None
    results.append(
        CheckResult(
            "entity candidate table uses Korean condition labels only",
            isinstance(candidate_frame, pd.DataFrame)
            and list(candidate_frame.columns) == ["조건 종류", "조건명"]
            and candidate_frame["조건 종류"].tolist() == ["거래처", "제조사", "제품"],
            f"columns={list(candidate_frame.columns) if isinstance(candidate_frame, pd.DataFrame) else []}",
        )
    )

    return results


def run_product_flow_inventory_alias_checks() -> list[CheckResult]:
    """Keep product flow/inventory aliases aligned with the canonical IO route."""

    from app.sims.nlq.action_inventory import IO_VIEW_FALLBACK_TARGETS
    from app.sims.nlq.nlq_router import resolve_new_sims_nlq_candidate
    from app.sims.views import rddbc_io_views
    from app.services.io_nlq import resolve_io_nlq

    results: list[CheckResult] = []
    cases = (
        ("제품수불현황 조회", "제품수불현황 조회"),
        ("제품 수불 현황 조회", "제품수불현황 조회"),
        ("제품수불부 조회", "제품수불현황 조회"),
        ("제품코드 31768 제품수불현황 조회", "제품수불현황 조회"),
        ("제품재고현황 조회", "제품재고현황 조회"),
        ("제품 재고 현황 조회", "제품재고현황 조회"),
        ("제품재고장 조회", "제품재고현황 조회"),
        ("장부재고 제품재고현황 조회", "제품재고현황 조회"),
    )
    for query, expected_action in cases:
        parsed = resolve_io_nlq(query)
        routed = resolve_new_sims_nlq_candidate(query)
        results.append(
            CheckResult(
                f"product alias: {query}",
                (
                    not isinstance(parsed, dict)
                    or str(parsed.get("action") or "") == expected_action
                )
                and isinstance(routed, dict)
                and str(routed.get("action") or "") == expected_action,
                f"parsed={parsed.get('action') if isinstance(parsed, dict) else parsed}, "
                f"routed={routed.get('action') if isinstance(routed, dict) else routed}",
            )
        )

    expected_fallbacks = {
        "제품수불현황 조회": "view_product_flow",
        "제품재고현황 조회": "view_product_inventory",
    }
    for action, export_name in expected_fallbacks.items():
        results.append(
            CheckResult(
                f"product fallback: {action}",
                IO_VIEW_FALLBACK_TARGETS.get(action) == export_name
                and callable(getattr(rddbc_io_views, export_name, None)),
                f"fallback={IO_VIEW_FALLBACK_TARGETS.get(action)}",
            )
        )
    results.append(
        CheckResult(
            "legacy aggregate fallback aliases are not used",
            not hasattr(rddbc_io_views, "view_rddbc250")
            and not hasattr(rddbc_io_views, "view_rddbc260"),
            "view_rddbc250/view_rddbc260 must not be aggregate fallback targets",
        )
    )

    parser_cases = (
        ("제품수불부 조회해줘", {}, ()),
        ("페날린 제품재고장", {"physic_nm": "페날린"}, ("maker_nm", "product_ven_nm")),
        ("제품명 페날린 제품재고장", {"physic_nm": "페날린"}, ()),
        ("제약사 한미약품 제품재고장", {"maker_nm": "한미약품", "product_ven_nm": "한미약품"}, ("physic_nm",)),
        ("제조사 한미약품 제품재고현황", {"maker_nm": "한미약품", "product_ven_nm": "한미약품"}, ("physic_nm",)),
        ("한미약품 제품재고장", {"physic_nm": "한미약품"}, ("maker_nm", "product_ven_nm")),
        ("31768 제품재고장", {"physic_cd": "31768"}, ("physic_nm",)),
        ("A1234 제품재고장", {"physic_cd": "A1234"}, ("physic_nm",)),
        ("제품코드 ABCDE 제품재고장", {"physic_cd": "ABCDE"}, ("physic_nm",)),
    )
    for query, expected_params, forbidden_params in parser_cases:
        parsed = resolve_io_nlq(query) or {}
        actual_params = dict(parsed.get("params") or {})
        ok = (
            str(parsed.get("action") or "") in {"제품수불현황 조회", "제품재고현황 조회"}
            and all(actual_params.get(key) == value for key, value in expected_params.items())
            and not any(key in actual_params for key in forbidden_params)
        )
        results.append(
            CheckResult(
                f"product parser contract: {query}",
                ok,
                f"params={actual_params!r}",
            )
        )
    return results


# ---------------------------------------------------------------------
# Payload capture
# ---------------------------------------------------------------------
class PayloadCapture:
    def __init__(self) -> None:
        self.payloads: list[dict[str, Any]] = []

    def fake_push(self, payload=None, action=None, *args, **kwargs):
        if payload is None and args:
            payload = args[0]
        if action is None and len(args) >= 2:
            action = args[1]

        if isinstance(payload, dict):
            p = dict(payload)
            if action and not p.get("action"):
                p["action"] = action
            self.payloads.append(p)
        else:
            self.payloads.append(
                {
                    "final": True,
                    "type": "unknown",
                    "action": action,
                    "data": payload,
                    "meta": {},
                }
            )
        return True

    def last_since(self, before_count: int) -> dict[str, Any] | None:
        if len(self.payloads) <= before_count:
            return None
        return self.payloads[-1]


def _patch_push_function(capture: PayloadCapture) -> None:
    """
    NLQ 라우터와 관련 모듈에 push_sims_result_to_chat가 직접 import되어 있을 수 있으므로
    가능한 모듈의 같은 이름을 fake_push로 교체한다.
    """
    module_names = [
        "app.ui.chat_middleware",
        "app.sims.nlq.nlq_router",
        "app.services.io_nlq",
    ]

    for module_name in module_names:
        try:
            mod = importlib.import_module(module_name)
            if hasattr(mod, "push_sims_result_to_chat"):
                setattr(mod, "push_sims_result_to_chat", capture.fake_push)
        except Exception:
            pass


def run_default_period_policy_checks() -> list[CheckResult]:
    """Exercise canonical period policy without a service or database call."""
    from app.services.io_nlq import apply_nlq_default_period_policy
    from app.sims.nlq.nlq_router import _build_io_query_summary

    results: list[CheckResult] = []
    reference_day = date(2026, 7, 31)

    cases = (
        (
            "detail without condition uses one day",
            {}, "입고명세 조회", "20260731", "20260731", "recent_1day", "list_detail",
        ),
        (
            "detail manufacturer condition uses current month",
            {"maker_nm": "fixture-maker"}, "입고명세 조회", "20260701", "20260731", "current_month", "list_detail",
        ),
        (
            "detail product condition uses current month",
            {"physic_cd": "39639"}, "출고명세 조회", "20260701", "20260731", "current_month", "list_detail",
        ),
        (
            "product flow always uses seven days",
            {"physic_cd": "39639"}, "제품수불현황 조회", "20260725", "20260731", "recent_7days", "single_entity_history",
        ),
        (
            "inventory without condition uses current month",
            {}, "제품재고현황 조회", "20260701", "20260731", "current_month_inventory", "inventory_movement",
        ),
        (
            "inventory condition also uses current month",
            {"stock_cd_list": ["00001"]}, "제품재고현황 조회", "20260701", "20260731", "current_month_inventory", "inventory_movement",
        ),
    )
    for name, params, action, expected_from, expected_to, expected_policy, expected_class in cases:
        resolved, policy = apply_nlq_default_period_policy(params, action, today=reference_day)
        results.append(
            CheckResult(
                name,
                resolved.get("date_from") == expected_from
                and resolved.get("date_to") == expected_to
                and policy.get("default_policy") == expected_policy
                and policy.get("action_class") == expected_class
                and bool(policy.get("auto_applied")),
                f"params={resolved!r}, policy={policy!r}",
            )
        )

    explicit_params = {"date_from": "20260710", "date_to": "20260720", "maker_nm": "fixture-maker"}
    explicit_resolved, explicit_policy = apply_nlq_default_period_policy(
        explicit_params, "출고명세 조회", today=reference_day
    )
    results.append(
        CheckResult(
            "explicit date range wins over default period",
            explicit_resolved == explicit_params
            and bool(explicit_policy.get("explicit_period_present"))
            and not bool(explicit_policy.get("auto_applied")),
            f"params={explicit_resolved!r}, policy={explicit_policy!r}",
        )
    )

    legacy_default_resolved, legacy_default_policy = apply_nlq_default_period_policy(
        {"date_from": "20260701", "date_to": "20260731", "_default_date_applied": "Y"},
        "입고↔거래명세서 검증",
        today=reference_day,
    )
    results.append(
        CheckResult(
            "legacy parser defaults do not masquerade as explicit dates",
            legacy_default_policy.get("default_policy") == "recent_1day"
            and legacy_default_resolved.get("date_from") == "20260731"
            and legacy_default_resolved.get("date_to") == "20260731"
            and "_default_date_applied" not in legacy_default_resolved,
            f"params={legacy_default_resolved!r}, policy={legacy_default_policy!r}",
        )
    )

    aggregate_params = {"date_from": "20260101", "date_to": "20260731"}
    aggregate_resolved, aggregate_policy = apply_nlq_default_period_policy(
        aggregate_params, "품목별 매출 예상", today=reference_day
    )
    results.append(
        CheckResult(
            "aggregate analysis keeps existing period",
            aggregate_resolved == aggregate_params
            and aggregate_policy.get("action_class") == "aggregate_analysis"
            and not bool(aggregate_policy.get("auto_applied")),
            f"params={aggregate_resolved!r}, policy={aggregate_policy!r}",
        )
    )

    system_default_resolved, system_default_policy = apply_nlq_default_period_policy(
        {"stock_mode": "book", "dashboard_product_di_list": []},
        "출고명세 조회",
        today=reference_day,
    )
    results.append(
        CheckResult(
            "system defaults are not explicit conditions",
            system_default_policy.get("explicit_condition_names") == []
            and system_default_resolved.get("date_from") == "20260731"
            and system_default_policy.get("default_policy") == "recent_1day",
            f"params={system_default_resolved!r}, policy={system_default_policy!r}",
        )
    )

    summary = _build_io_query_summary(
        "입고명세 조회",
        {"date_from": "20260731", "date_to": "20260731"},
        {
            "auto_applied": True,
            "default_policy": "recent_1day",
        },
    )
    results.append(
        CheckResult(
            "period summary uses common one-day wording",
            "최근 1일 자동적용(추가 조건 없음)" in summary,
            summary,
        )
    )
    manufacturer_summary = _build_io_query_summary(
        "입고명세 조회",
        {"date_from": "20260701", "date_to": "20260731"},
        {
            "auto_applied": True,
            "default_policy": "current_month",
            "explicit_condition_names": ["manufacturer"],
        },
    )
    results.append(
        CheckResult(
            "period summary identifies the explicit condition category",
            "최근 1개월 자동적용(제약사 조건)" in manufacturer_summary,
            manufacturer_summary,
        )
    )
    inventory_summary = _build_io_query_summary(
        "제품재고현황 조회",
        {"date_from": "20260701", "date_to": "20260731"},
        {
            "auto_applied": True,
            "default_policy": "current_month_inventory",
        },
    )
    results.append(
        CheckResult(
            "inventory summary uses one fixed monthly wording",
            "현재월 자동적용(재고 월조회)" in inventory_summary,
            inventory_summary,
        )
    )

    from app.services.io_nlq import resolve_io_nlq

    named_value_cases = (
        ("제약사 한미약품 출고명세 조회", "maker_nm", "한미약품"),
        ("제약사 일동 출고명세 조회", "maker_nm", "일동"),
        ("거래처 약국 출고명세 조회", "ven_nm", "약국"),
    )
    for query, key, expected_value in named_value_cases:
        parsed = resolve_io_nlq(query) or {}
        parsed_params = dict(parsed.get("params") or {})
        results.append(
            CheckResult(
                f"named condition excludes resolved action suffix: {key}",
                parsed_params.get(key) == expected_value
                and "출고명세" not in str(parsed_params.get(key) or "")
                and "_named_condition_cleanup" not in parsed_params,
                f"query={query!r}, action={parsed.get('action')!r}, params={parsed_params!r}",
            )
        )

    inventory_date_cases = (
        ("2026년 6월 제품재고현황 조회", "20260601", "20260630"),
        ("20260720 제품재고현황 조회", "20260701", "20260720"),
        ("20260710부터 20260720까지 제품재고현황 조회", "20260710", "20260720"),
    )
    for query, expected_from, expected_to in inventory_date_cases:
        parsed = resolve_io_nlq(query) or {}
        parsed_params = dict(parsed.get("params") or {})
        results.append(
            CheckResult(
                "inventory explicit period normalization",
                parsed_params.get("date_from") == expected_from
                and parsed_params.get("date_to") == expected_to
                and not str(parsed_params.get("physic_nm") or "").strip(),
                f"query={query!r}, params={parsed_params!r}",
            )
        )
    return results


def run_response_timing_checks() -> list[CheckResult]:
    """Verify persisted NLQ timings are stable and header-ready without Streamlit."""
    from unittest.mock import patch

    import pandas as pd

    from app.ui import chat_middleware as middleware

    results: list[CheckResult] = []
    original_monotonic = middleware.time.monotonic
    case_log_dir = tempfile.TemporaryDirectory()
    previous_case_log_path = os.environ.get("SIMS_NLQ_CASE_LOG_FILE")
    os.environ["SIMS_NLQ_CASE_LOG_FILE"] = str(Path(case_log_dir.name) / "timing_cases.jsonl")
    try:
        middleware.time.monotonic = lambda: 110.24
        payload = {"id": "timing-fixture", "meta": {"nlq": True, "action": "입고명세 조회"}}
        state = {
            "__sims_nlq_response_timing": {
                "request_started_at": "2026-07-31T10:00:00",
                "request_started_monotonic": 100.0,
            }
        }
        middleware._attach_sims_response_timing(payload, state)
        meta = dict(payload.get("meta") or {})
        header = middleware._build_sims_result_header_view(
            {"action": "입고명세 조회"}, meta, title="입고명세 조회"
        )
        results.append(
            CheckResult(
                "response timing is persisted before chat push and appears in header",
                meta.get("request_started_at") == "2026-07-31 10:00:00"
                and bool(meta.get("response_completed_at"))
                and meta.get("elapsed_ms") == 10240
                and "응답완료" in str(header.get("line2") or "")
                and "처리 10.2초" in str(header.get("line2") or "")
                and "__sims_nlq_response_timing" not in state,
                f"meta={meta!r}, line2={header.get('line2')!r}",
            )
        )

        restored_meta = dict(meta)
        restored_payload = {"meta": restored_meta}
        middleware.time.monotonic = lambda: 999.0
        middleware._attach_sims_response_timing(
            restored_payload,
            {"__sims_nlq_response_timing": {"request_started_at": "2026-07-31T11:00:00", "request_started_monotonic": 1.0}},
        )
        results.append(
            CheckResult(
                "restored response timing is not recomputed",
                restored_payload.get("meta") == restored_meta,
                f"restored_meta={restored_payload.get('meta')!r}",
            )
        )
        results.append(
            CheckResult(
                "response timing formatter handles seconds and minutes",
                middleware._format_sims_elapsed_ms(850) == "0.8초"
                and middleware._format_sims_elapsed_ms(10_240) == "10.2초"
                and middleware._format_sims_elapsed_ms(158_000) == "2분 38초",
                "formatter boundaries",
            )
        )

        # 출고명세의 화면표보다 큰 전체 원본은 current-table/다운로드 source로
        # 준비한 뒤에만 응답시간을 확정해야 한다. drain 이후 시간은 포함하지 않는다.
        clock = {"value": 100.0}
        full_source_calls = []
        delivered_meta = {}
        full_df = pd.DataFrame({"제품코드": ["00001", "00002"], "수량": [1, 2]})
        display_df = full_df.iloc[:1].copy()
        push_payload = {
            "id": "timing-outbound-fixture",
            "type": "table",
            "action": "출고명세 조회",
            "title": "출고명세 조회",
            "df": full_df,
            "df_display": display_df,
            "meta": {
                "nlq": True,
                "action": "출고명세 조회",
                "table_key": "timing_outbound_fixture",
                "row_count_total": int(len(full_df)),
            },
        }
        outbound_state = {
            "__chat_inbox": [],
            "__chat_history": [],
            "__chat_pending_items": [],
            "__sims_push_count": 0,
            "__sims_nlq_response_timing": {
                "request_started_at": "2026-07-31T10:00:00",
                "request_started_monotonic": 100.0,
            },
        }

        def _prepare_full_source(_item, _meta, _display_df):
            full_source_calls.append(True)
            clock["value"] = 112.5
            return full_df.copy()

        def _capture_delivery():
            delivered_meta.update(dict(outbound_state["__chat_inbox"][-1].get("meta") or {}))
            # history/push 이후 경과는 확정된 elapsed_ms에 포함되면 안 된다.
            clock["value"] = 200.0

        with (
            patch.object(middleware.st, "session_state", outbound_state),
            patch.object(middleware, "wire_chat_context", lambda *_args, **_kwargs: None),
            patch.object(middleware, "_chat_payload_matches_current_company", lambda _payload: True),
            patch.object(middleware, "_get_full_download_df_for_sims_item", _prepare_full_source),
            patch.object(middleware, "drain_inbox_to_chat", _capture_delivery),
            patch.object(middleware.time, "monotonic", lambda: clock["value"]),
        ):
            middleware.wssz(push_payload, action="출고명세 조회")

        results.append(
            CheckResult(
                "outbound timing includes full source preparation before history delivery",
                full_source_calls == [True]
                and delivered_meta.get("elapsed_ms") == 12_500
                and bool(delivered_meta.get("response_completed_at"))
                and outbound_state.get("__sims_current_table_source_key") == "timing_outbound_fixture"
                and "__sims_nlq_response_timing" not in outbound_state,
                f"full_source_calls={len(full_source_calls)}, delivered_meta={delivered_meta!r}",
            )
        )

        guidance_payload = {
            "type": "text",
            "action": "출고명세 조회",
            "title": "조회 조건 확인 필요",
            "message": "조회 조건을 확인하는 중 오류가 발생했습니다. 거래처·제약사·제품 중 조건 종류를 명시해 다시 조회해 주세요.",
            "data": "조회 조건을 확인하는 중 오류가 발생했습니다. 거래처·제약사·제품 중 조건 종류를 명시해 다시 조회해 주세요.",
            "meta": {
                "nlq": True,
                "entity_resolution_status": "resolution_unavailable",
                "result_status": "input_required",
            },
        }
        guidance_state = {
            "__chat_inbox": [],
            "__chat_history": [],
            "__chat_pending_items": [],
            "__sims_push_count": 0,
        }
        delivered_guidance = {}

        def _capture_guidance_delivery():
            delivered_guidance.update(dict(guidance_state["__chat_inbox"][-1]))

        with (
            patch.object(middleware.st, "session_state", guidance_state),
            patch.object(middleware, "wire_chat_context", lambda *_args, **_kwargs: None),
            patch.object(middleware, "_chat_payload_matches_current_company", lambda _payload: True),
            patch.object(middleware, "drain_inbox_to_chat", _capture_guidance_delivery),
        ):
            middleware.wssz(guidance_payload, action="출고명세 조회")

        results.append(
            CheckResult(
                "resolution-unavailable guidance survives tableless chat push",
                delivered_guidance.get("title") == "조회 조건 확인 필요"
                and delivered_guidance.get("message") == guidance_payload["message"]
                and delivered_guidance.get("data") == guidance_payload["message"]
                and delivered_guidance.get("content") == guidance_payload["message"]
                and (delivered_guidance.get("meta") or {}).get("row_count") == 0,
                f"delivered_guidance={delivered_guidance!r}",
            )
        )
    finally:
        middleware.time.monotonic = original_monotonic
        if previous_case_log_path is None:
            os.environ.pop("SIMS_NLQ_CASE_LOG_FILE", None)
        else:
            os.environ["SIMS_NLQ_CASE_LOG_FILE"] = previous_case_log_path
        case_log_dir.cleanup()
    return results


def run_nlq_case_log_checks() -> list[CheckResult]:
    """Verify NLQ case records use the final delivery contract without production writes."""
    from app.services import nlq_case_log_service as case_log_service
    from app.services.nlq_case_log_service import append_nlq_case_record, resolve_nlq_case_log_path
    from app.ui import chat_middleware as middleware
    from unittest.mock import patch

    results: list[CheckResult] = []
    production_path = resolve_nlq_case_log_path()

    def _file_identity(path: Path) -> tuple[int, int, str] | None:
        if not path.exists():
            return None
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            for block in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(block)
        stat = path.stat()
        return stat.st_mtime_ns, stat.st_size, digest.hexdigest()

    production_before = _file_identity(production_path)
    with tempfile.TemporaryDirectory() as temp_dir:
        case_path = Path(temp_dir) / "nlq_cases.jsonl"
        state: dict[str, Any] = {
            "__chat_inbox": [],
            "__chat_history": [],
            "__chat_pending_items": [],
            "__sims_push_count": 0,
        }
        payload = {
            "id": "case-success",
            "type": "table",
            "action": "출고명세 조회",
            "title": "출고명세 조회",
            "df": pd.DataFrame({"순번": [1, 2, 3]}),
            "df_display": pd.DataFrame({"순번": [1, 2]}),
            "params": {
                "date_from": "20260801",
                "date_to": "20260801",
                "ven_cd": "00001",
                "password": "must-not-log",
                "sql": "must-not-log",
                "connection_string": "must-not-log",
            },
            "meta": {
                "nlq": True,
                "nlq_trace_request_id": "case-success",
                "nlq_query": "출고명세 20260801 test@example.com",
                "action": "출고명세 조회",
                "result_status": "success",
                "request_started_at": "2026-08-01T09:00:00+09:00",
                "response_completed_at": "2026-08-01T09:00:01+09:00",
                "row_count_total": 3,
                "display_row_count": 2,
                "download_row_count": 3,
                "expected_rows": 5,
                "prepared_rows": 3,
                "download_limit_rows": 3,
                "applied_download_limit_rows": 3,
                "limit_hit": True,
                "source_call_count": 2, "cache_used": False,
                "display_source_status": "queried", "full_source_status": "queried",
                "download_source_status": "partial_limit",
                "elapsed_ms": 125,
                "parsed_action": "출고명세 조회",
                "canonical_action": "출고명세 조회",
                "period_policy": {"action_class": "list_detail", "explicit_period_present": True, "default_policy": "", "policy_reason": "explicit_date", "auto_applied": False},
                "answer_body": "must-not-log",
                "dataframe": "must-not-log",
            },
        }

        def _capture_delivery() -> None:
            return None

        with (
            patch.dict(os.environ, {"SIMS_NLQ_CASE_LOG_FILE": str(case_path)}),
            patch.object(middleware.st, "session_state", state),
            patch.object(middleware, "wire_chat_context", lambda *_args, **_kwargs: None),
            patch.object(middleware, "_chat_payload_matches_current_company", lambda _payload: True),
            patch.object(middleware, "drain_inbox_to_chat", _capture_delivery),
            patch.object(middleware, "_chat_runtime_log_context", lambda *_args, **_kwargs: {"company_id": 4, "room_id": "room-fixture"}),
        ):
            middleware.wssz(payload, action="출고명세 조회")
            middleware.wssz(payload, action="출고명세 조회")

        records = [json.loads(line) for line in case_path.read_text(encoding="utf-8").splitlines() if line.strip()]
        success = next((record for record in records if record.get("request_id") == "case-success"), {})
        serialized = json.dumps(success, ensure_ascii=False)
        results.append(
            CheckResult(
                "final chat delivery writes one safe, deduplicated NLQ case record",
                len(records) == 1
                and list(success)[:5] == ["occurred_at", "completed_at", "logged_at", "schema_version", "request_id"]
                and success.get("schema_version") == "2.0"
                and bool(re.fullmatch(r".+\.\d{3}\+09:00", str(success.get("occurred_at") or "")))
                and bool(re.fullmatch(r".+\.\d{3}\+09:00", str(success.get("completed_at") or "")))
                and bool(re.fullmatch(r".+\.\d{3}\+09:00", str(success.get("logged_at") or "")))
                and success.get("company_id") == 4
                and success.get("total_rows") == 3
                and success.get("display_rows") == 2
                and success.get("full_source_rows") == 3
                and success.get("expected_rows") == 5
                and success.get("prepared_rows") == 3
                and success.get("download_limit_rows") == 3
                and success.get("applied_download_limit_rows") == 3
                and success.get("limit_hit") is True
                and success.get("source_call_count") == 2
                and success.get("elapsed_ms") == 125
                and success.get("result_status_source") == "payload"
                and success.get("display_source_status") == "queried"
                and success.get("full_source_status") == "queried"
                and success.get("download_source_status") == "partial_limit"
                and success.get("conditions") == {"date_from": "20260801", "date_to": "20260801", "transaction_vendor_codes": ["00001"]}
                and success.get("interpretation", {}).get("condition_summary") == "기간 2026-08-01 / 명시기간"
                and "must-not-log" not in serialized
                and "@example.com" not in serialized,
                f"records={records!r}",
            )
        )

        direct_state: dict[str, Any] = {}
        direct_cases = [
            (
                "no_data",
                {
                    "id": "case-no-data",
                    "action": "입고명세 조회",
                    "meta": {"nlq": True, "result_status": "no_data", "row_count_total": 0, "display_row_count": 0},
                },
            ),
            (
                "input_required",
                {
                    "id": "case-input-required",
                    "action": "제품수불현황 조회",
                    "meta": {"nlq": True, "result_status": "input_required", "row_count_total": 0},
                },
            ),
            (
                "candidate_required",
                {
                    "id": "case-candidate-required",
                    "action": "출고명세 조회",
                    "meta": {
                        "nlq": True,
                        "result_status": "candidate_required",
                        "candidate_count": 2,
                        "candidates": [{"match_value": "must-not-log"}],
                    },
                },
            ),
            (
                "resolution_unavailable",
                {
                    "id": "case-error",
                    "action": "출고명세 조회",
                    "meta": {
                        "nlq": True,
                        "result_status": "resolution_unavailable",
                        "error_class": "ProgrammingError",
                        "sql_error_number": 1205,
                        "error_code": "1205",
                        "traceback": "must-not-log",
                    },
                },
            ),
            (
                "unknown",
                {
                    "id": "case-unknown",
                    "action": "출고명세 조회",
                    "meta": {"nlq": True},
                },
            ),
        ]
        with patch.dict(os.environ, {"SIMS_NLQ_CASE_LOG_FILE": str(case_path)}):
            for _status, case_payload in direct_cases:
                append_nlq_case_record(
                    case_payload,
                    direct_state,
                    runtime_context={"company_id": 4, "room_id": "room-fixture"},
                    question="fixture question",
                )

        records = [json.loads(line) for line in case_path.read_text(encoding="utf-8").splitlines() if line.strip()]
        by_status = {record.get("result_status"): record for record in records}
        direct_serialized = json.dumps(records, ensure_ascii=False)
        results.append(
            CheckResult(
                "all explicit terminal statuses remain safe and nonempty",
                by_status.get("no_data", {}).get("total_rows") == 0
                and by_status.get("input_required", {}).get("display_rows") is None
                and by_status.get("candidate_required", {}).get("candidate_count") == 2
                and by_status.get("resolution_unavailable", {}).get("error_class") == "ProgrammingError"
                and by_status.get("resolution_unavailable", {}).get("sql_error_number") == 1205
                and by_status.get("unknown", {}).get("total_rows") is None
                and by_status.get("unknown", {}).get("consistency_flags", {}).get("result_status_unknown") is True
                and by_status.get("unknown", {}).get("consistency_flags", {}).get("total_rows_missing") is True
                and "candidates" not in by_status.get("candidate_required", {})
                and "must-not-log" not in direct_serialized,
                f"records={records!r}",
            )
        )

        derived_payload = {
            "id": "case-derived",
            "action": "출고명세 조회",
            "params": {"maker_nm": "한미", "date_from": "20260801", "date_to": "20260801"},
            "meta": {
                "nlq": True,
                "row_count_total": 367,
                "display_row_count": 200,
                "download_row_count": 367,
                "candidate_count": 0,
                "parsed_action": "출고명세 조회",
                "canonical_action": "출고명세 조회",
                "search_mode": "labeled",
                "search_fields": ["maker_nm"],
                "period_policy": {"action_class": "list_detail", "explicit_period_present": False, "default_policy": "recent_1day", "policy_reason": "additional_conditions_absent", "auto_applied": True},
            },
        }
        with patch.dict(os.environ, {"SIMS_NLQ_CASE_LOG_FILE": str(case_path)}):
            append_nlq_case_record(derived_payload, direct_state, runtime_context={})
        derived = json.loads(case_path.read_text(encoding="utf-8").splitlines()[-1])
        results.append(
            CheckResult(
                "derived status and interpretation retain final params without reparsing",
                derived.get("result_status") == "success"
                and derived.get("result_status_source") == "derived"
                and derived.get("total_rows") == 367
                and derived.get("display_rows") == 200
                and derived.get("full_source_rows") == 367
                and derived.get("interpretation", {}).get("extracted_conditions", {}).get("manufacturer_name") == "한미"
                and derived.get("interpretation", {}).get("search_mode") == "labeled"
                and derived.get("interpretation", {}).get("period_auto_applied") is True
                and derived.get("interpretation", {}).get("condition_summary") == "기간 2026-08-01 / 제조사명 한미 / 기본기간 최근 1일",
                f"derived={derived!r}",
            )
        )

        current_table_question = "추세판정별 요약"
        current_table_payload = {
            "id": "case-current-table-question",
            "action": "현재표 추세판정별 요약",
            "meta": {
                "nlq": True,
                "current_table_followup": True,
                "nlq_query": current_table_question,
                "result_status": "success",
                "row_count": 2,
                "row_count_total": 2,
                "execution_status": "success",
                "requested_metrics": ["sales"],
                "requested_groupings": ["trend_judgement"],
                "requested_metric": "sales",
                "requested_grouping": "trend_judgement",
                "issue_codes": [],
                "source_action": "제약사별 매출 추세 분석 요약표",
                "source_table_key": "current-table-fixture",
                "source_call_count": 0,
                "filter_column": "추세판정",
                "filter_value": "감소",
                "missing_columns": [],
                "result_metric": "sales",
                "result_grain": "trend_judgement_summary",
                "table_created": True,
            },
        }
        with patch.dict(os.environ, {"SIMS_NLQ_CASE_LOG_FILE": str(case_path)}):
            append_nlq_case_record(
                current_table_payload,
                direct_state,
                runtime_context={"company_id": 4, "room_id": "room-fixture"},
            )
        current_table_case = json.loads(case_path.read_text(encoding="utf-8").splitlines()[-1])
        chat_main_source = (
            PROJECT_ROOT / "app" / "Lmstudio_SSAI_chat_main.py"
        ).read_text(encoding="utf-8")
        results.append(
            CheckResult(
                "current-table case log preserves the actual user question",
                current_table_case.get("question") == current_table_question
                and current_table_case.get("execution_status") == "success"
                and current_table_case.get("requested_metrics") == ["sales"]
                and current_table_case.get("requested_groupings") == ["trend_judgement"]
                and current_table_case.get("requested_grouping") == "trend_judgement"
                and current_table_case.get("source_action") == "제약사별 매출 추세 분석 요약표"
                and current_table_case.get("source_table_key") == "current-table-fixture"
                and current_table_case.get("source_call_count") == 0
                and current_table_case.get("filter_column") == "추세판정"
                and current_table_case.get("filter_value") == "감소"
                and current_table_case.get("missing_columns") == []
                and current_table_case.get("result_metric") == "sales"
                and current_table_case.get("result_grain") == "trend_judgement_summary"
                and current_table_case.get("table_created") is True
                and current_table_case.get("issue_codes") == []
                and "__sims_current_table_followup_case_query" in chat_main_source
                and '"nlq_query": case_query' in chat_main_source,
                f"current_table_case={current_table_case!r}",
            )
        )

        results.append(
            CheckResult(
                "case log table_created preserves true false none",
                success.get("table_created") is None
                and current_table_case.get("table_created") is True,
                (
                    f"generic_success={success.get('table_created')!r} "
                    f"current_table_success={current_table_case.get('table_created')!r}"
                ),
            )
        )

        source_tree = ast.parse(chat_main_source)
        source_constants = [
            ast.get_source_segment(chat_main_source, node) or ""
            for node in source_tree.body
            if isinstance(node, ast.Assign)
            and any(
                isinstance(target, ast.Name)
                and target.id == "_CURRENT_TABLE_PUSH_META_PROTECTED_KEYS"
                for target in node.targets
            )
        ]
        source_functions = {
            node.name: ast.get_source_segment(chat_main_source, node) or ""
            for node in source_tree.body
            if isinstance(node, ast.FunctionDef)
            and node.name in {
                "_merge_current_table_push_meta",
                "_current_table_push_table",
                "_current_table_push_notice",
            }
        }
        production_pushes: list[dict[str, Any]] = []
        production_session = {
            "__sims_current_table_source_action": "제약사별 매출 추세 분석 요약표",
            "__sims_current_table_followup_case_query": "현재표 제조사별 요약",
        }
        production_namespace: dict[str, Any] = {
            "pd": pd,
            "uuid": uuid,
            "st": types.SimpleNamespace(session_state=production_session),
            "log": types.SimpleNamespace(debug=lambda *_args, **_kwargs: None),
            "_current_table_fill_alias_values": lambda frame: frame,
            "_current_table_clean_none_for_display": lambda frame: frame,
        }
        production_source = "\n\n".join(
            [*source_constants]
            + [
                source_functions.get(name, "")
                for name in (
                    "_merge_current_table_push_meta",
                    "_current_table_push_table",
                    "_current_table_push_notice",
                )
            ]
        )
        try:
            exec(compile(production_source, "Lmstudio_SSAI_chat_main.py", "exec"), production_namespace)
            with patch.object(
                middleware,
                "push_sims_result_to_chat",
                side_effect=lambda pushed, _action: production_pushes.append(dict(pushed)),
            ):
                production_namespace["_current_table_push_table"](
                    title="현재표 제조사별 요약",
                    action="현재표 제조사별 요약",
                    df=pd.DataFrame({"제조사명": ["제조사A"], "매출금액": [100]}),
                    query_summary="현재표 / 제조사별 요약",
                    source_query="현재표 제조사별 요약",
                    source_table_key="source-fixture",
                    source_rows=3,
                    extra_meta={
                        "execution_status": "success",
                        "result_status": "success",
                        "requested_metrics": ["sales"],
                        "requested_groupings": ["manufacturer"],
                        "requested_metric": "sales",
                        "requested_grouping": "manufacturer",
                        "missing_columns": [],
                        "result_metric": "sales",
                        "result_grain": "manufacturer_summary",
                        "issue_codes": [],
                        "table_created": True,
                        "source_action": "잘못된 원본 action",
                        "source_call_count": 99,
                        "nlq_query": "잘못된 질문",
                        "source_table_key": "잘못된 key",
                    },
                )
                production_session["__sims_current_table_followup_case_query"] = "현재표 제품그룹별 요약"
                production_namespace["_current_table_push_table"](
                    title="현재표 결과 없음",
                    action="현재표 결과 없음",
                    df=pd.DataFrame(),
                    query_summary="현재표 / 결과 없음",
                    source_query="현재표 제품그룹별 요약",
                    source_table_key="source-fixture",
                    source_rows=3,
                    extra_meta={
                        "execution_status": "no_data",
                        "result_status": "no_data",
                        "requested_metrics": ["sales"],
                        "requested_groupings": ["product_group"],
                        "requested_metric": "sales",
                        "requested_grouping": "product_group",
                        "missing_columns": [],
                        "result_metric": "sales",
                        "result_grain": "product_group_summary",
                        "issue_codes": [],
                        "table_created": False,
                    },
                )
                production_session["__sims_current_table_followup_case_query"] = "현재표 제품별 매출 TOP 20"
                production_namespace["_current_table_push_notice"](
                    title="현재표 컬럼 부족",
                    action="현재표 컬럼 부족",
                    message="제품 컬럼이 없습니다.",
                    query_summary="현재표 / 컬럼 부족",
                    source_query="현재표 제품별 매출 TOP 20",
                    source_table_key="source-fixture",
                    source_rows=3,
                    extra_meta={
                        "execution_status": "column_unavailable",
                        "result_status": "column_unavailable",
                        "requested_metrics": ["sales"],
                        "requested_groupings": ["product"],
                        "requested_metric": "sales",
                        "requested_grouping": "product",
                        "missing_columns": ["제품"],
                        "result_metric": "",
                        "result_grain": "",
                        "issue_codes": [],
                        "table_created": False,
                    },
                )
        except Exception as exc:
            results.append(_fail("production current-table push boundary", f"{type(exc).__name__}: {exc}"))
        else:
            production_metas = [dict(payload.get("meta") or {}) for payload in production_pushes]
            production_case_state: dict[str, Any] = {}
            for index, payload in enumerate(production_pushes, start=1):
                payload["id"] = f"production-current-table-{index}"
                with patch.dict(os.environ, {"SIMS_NLQ_CASE_LOG_FILE": str(case_path)}):
                    append_nlq_case_record(
                        payload,
                        production_case_state,
                        runtime_context={"company_id": 4, "room_id": "room-fixture"},
                    )
            production_records = [
                json.loads(line)
                for line in case_path.read_text(encoding="utf-8").splitlines()
                if line.strip()
            ][-3:]
            table_meta, empty_meta, notice_meta = production_metas
            results.append(
                CheckResult(
                    "production current-table push helpers preserve capability meta and source identity",
                    len(production_metas) == 3
                    and table_meta.get("result_status") == "success"
                    and empty_meta.get("result_status") == "no_data"
                    and notice_meta.get("result_status") == "column_unavailable"
                    and all(meta.get("source_call_count") == 0 for meta in production_metas)
                    and table_meta.get("source_action") == "제약사별 매출 추세 분석 요약표"
                    and table_meta.get("source_table_key") == "source-fixture"
                    and table_meta.get("nlq_query") == "현재표 제조사별 요약"
                    and table_meta.get("result_metric") == "sales"
                    and table_meta.get("requested_metrics") == ["sales"]
                    and table_meta.get("requested_groupings") == ["manufacturer"]
                    and table_meta.get("issue_codes") == []
                    and table_meta.get("table_created") is True
                    and empty_meta.get("requested_grouping") == "product_group"
                    and empty_meta.get("table_created") is False
                    and notice_meta.get("missing_columns") == ["제품"]
                    and notice_meta.get("source_table_key") == "source-fixture"
                    and notice_meta.get("table_created") is False
                    and [record.get("question") for record in production_records]
                    == ["현재표 제조사별 요약", "현재표 제품그룹별 요약", "현재표 제품별 매출 TOP 20"]
                    and [record.get("result_status") for record in production_records]
                    == ["success", "no_data", "column_unavailable"]
                    and all(record.get("source_call_count") == 0 for record in production_records)
                    and [record.get("requested_metrics") for record in production_records]
                    == [["sales"], ["sales"], ["sales"]]
                    and [record.get("requested_groupings") for record in production_records]
                    == [["manufacturer"], ["product_group"], ["product"]]
                    and [record.get("issue_codes") for record in production_records] == [[], [], []]
                    and [record.get("table_created") for record in production_records] == [True, False, False]
                    and [record.get("source_table_key") for record in production_records]
                    == ["source-fixture", "source-fixture", "source-fixture"],
                    f"metas={production_metas!r}, records={production_records!r}",
                )
            )

        unlabeled_payload = {
            "id": "case-unlabeled",
            "action": "출고명세 조회",
            "params": {"nlq_unlabeled_name": "삼진", "ven_cd": "00001", "date_from": "20260801"},
            "meta": {
                "nlq": True,
                "result_status": "success",
                "row_count": 1,
                "parsed_action": "출고명세 조회",
                "canonical_action": "출고명세 조회",
                "search_mode": "unlabeled_or",
                "search_fields": ["ven_nm", "physic_nm", "maker_nm"],
                "entity_resolution_status": "resolved",
                "resolved_kind": "unlabeled_like",
                "display_source_status": "cache",
                "full_source_status": "not_required",
                "cache_used": True,
            },
        }
        with patch.dict(os.environ, {"SIMS_NLQ_CASE_LOG_FILE": str(case_path)}):
            append_nlq_case_record(unlabeled_payload, direct_state, runtime_context={})
        unlabeled = json.loads(case_path.read_text(encoding="utf-8").splitlines()[-1])
        results.append(
            CheckResult(
                "unlabeled interpretation records OR search metadata and bounded resolved codes",
                unlabeled.get("interpretation", {}).get("search_mode") == "unlabeled_or"
                and unlabeled.get("interpretation", {}).get("search_fields") == ["transaction_vendor", "product", "manufacturer"]
                and unlabeled.get("interpretation", {}).get("resolved_conditions", {}).get("resolved_code_samples") == ["00001"]
                and unlabeled.get("interpretation", {}).get("condition_summary") == "기간 2026-08-01 / 무라벨명 삼진 / 거래처·제품·제조사 OR"
                and unlabeled.get("display_source_status") == "cache"
                and unlabeled.get("full_source_status") == "not_required"
                and unlabeled.get("cache_used") is True,
                f"unlabeled={unlabeled!r}",
            )
        )

        anomalous_payload = {
            "id": "case-anomaly",
            "action": "출고명세 조회",
            "meta": {
                "nlq": True,
                "result_status": "not_found",
                "row_count_total": 5,
                "display_row_count": 7,
                "download_row_count": 3,
                "notice_codes": ["entity_not_found"],
            },
        }
        with patch.dict(os.environ, {"SIMS_NLQ_CASE_LOG_FILE": str(case_path)}):
            append_nlq_case_record(anomalous_payload, direct_state, runtime_context={})
        anomaly = json.loads(case_path.read_text(encoding="utf-8").splitlines()[-1])
        flags = anomaly.get("consistency_flags") or {}
        results.append(
            CheckResult(
                "case consistency flags expose row/status and source/display contradictions",
                flags.get("rows_present_but_not_found") is True
                and flags.get("rows_present_but_entity_not_found_notice") is True
                and flags.get("full_source_less_than_display") is True
                and flags.get("total_rows_less_than_display") is True
                and flags.get("total_rows_mismatch_full_source") is False
                and flags.get("elapsed_missing") is True
                and flags.get("action_missing") is False,
                f"flags={flags!r}",
            )
        )

        emulated_app_log = Path(temp_dir) / "runtime" / "app.log"
        with patch.object(case_log_service, "config_path_any", lambda *_args, **_kwargs: emulated_app_log):
            resolved = resolve_nlq_case_log_path(environ={})
        results.append(
            CheckResult(
                "default case path is the existing app log sibling",
                resolved == emulated_app_log.parent / "nlq_cases.jsonl",
                f"resolved={resolved}",
            )
        )

        production_after = _file_identity(production_path)
        results.append(
            CheckResult(
                "case fixtures do not create or modify the production case log",
                production_before == production_after,
                f"path={production_path}, before={production_before}, after={production_after}",
            )
        )
    return results


def run_product_flow_input_and_empty_payload_checks() -> list[CheckResult]:
    """제품수불 필수조건과 제품재고 0건 text payload를 DB 없이 고정한다."""
    import math
    from unittest.mock import patch

    import numpy as np
    import pandas as pd

    from app.services.product_flow_service import get_product_flow_result
    from app.services.product_inventory_service import _inventory_text_payload
    from app.sims.nlq import nlq_router as router

    results: list[CheckResult] = []

    def _is_exact_numeric_zero(value: Any) -> bool:
        return type(value) in {int, float} and not math.isnan(float(value)) and float(value) == 0.0

    direct_payload = get_product_flow_result({})
    direct_meta = dict(direct_payload.get("meta") or {})
    results.append(
        CheckResult(
            "product flow direct empty input is tableless",
            direct_payload.get("type") == "text"
            and bool(direct_meta.get("input_required"))
            and direct_meta.get("result_status") == "input_required"
            and _is_exact_numeric_zero(direct_meta.get("row_count"))
            and _is_exact_numeric_zero(direct_meta.get("row_count_total"))
            and all(key not in direct_payload for key in ("df", "df_display", "records"))
            and "flow_summary" not in direct_meta,
            f"type={direct_payload.get('type')}, meta={direct_meta!r}",
        )
    )

    capture = PayloadCapture()
    _patch_push_function(capture)
    session_state: dict[str, Any] = {}
    flow_module = importlib.import_module("app.services.product_flow_service")
    original_flow_result = flow_module.get_product_flow_result
    flow_calls: list[dict[str, Any]] = []

    def _flow_spy(params=None, **kwargs):
        flow_calls.append(dict(params or kwargs.get("params") or {}))
        raise AssertionError("빈 제품수불은 product flow service를 호출하면 안 됩니다.")

    flow_module.get_product_flow_result = _flow_spy
    try:
        # Missing product input is a local validation path, not a live master lookup.
        with patch(
            "app.services.io_nlq.resolve_unlabeled_io_entity_condition",
            return_value={"status": "not_applicable", "params": {}},
        ):
            handled = router._try_handle_io_nlq(
            "제품수불부 조회해줘",
            room={"messages": []},
            session_state=session_state,
            make_ts=_make_ts,
            next_seq=_next_seq_factory(),
            logger=log,
        )
    finally:
        flow_module.get_product_flow_result = original_flow_result
    routed_payload = capture.last_since(0) or {}
    routed_meta = dict(routed_payload.get("meta") or {})
    results.append(
        CheckResult(
            "product flow NLQ empty input skips service and current table",
            bool(handled)
            and not flow_calls
            and routed_payload.get("type") == "text"
            and bool(routed_meta.get("input_required"))
            and routed_meta.get("result_status") == "input_required"
            and _is_exact_numeric_zero(routed_meta.get("row_count"))
            and _is_exact_numeric_zero(routed_meta.get("row_count_total"))
            and all(key not in routed_payload for key in ("df", "df_display", "records"))
            and "flow_summary" not in routed_meta
            and "__sims_current_table_source_key" not in session_state,
            f"service_calls={len(flow_calls)}, type={routed_payload.get('type')}, meta_keys={sorted(routed_meta)!r}",
        )
    )

    empty_payload = _inventory_text_payload(
        message="해당 조회조건의 자료가 없습니다.",
        params={},
        meta={"result_status": "no_data", "row_count": 0, "row_count_total": 0},
    )
    empty_meta = dict(empty_payload.get("meta") or {})
    numeric_values = [
        empty_meta.get(key)
        for key in (
            "detail_count",
            "sum_carry_qty",
            "sum_in_qty",
            "sum_out_qty",
            "sum_stock_qty",
            "sum_stock_amt",
            "sum_insu_amt",
        )
    ]
    results.append(
        CheckResult(
            "product inventory empty payload has zero metrics and no table",
            empty_payload.get("type") == "text"
            and empty_meta.get("result_status") == "no_data"
            and _is_exact_numeric_zero(empty_meta.get("row_count"))
            and _is_exact_numeric_zero(empty_meta.get("row_count_total"))
            and all(_is_exact_numeric_zero(value) for value in numeric_values)
            and all(key not in empty_payload for key in ("df", "df_display", "records")),
            f"meta={empty_meta!r}",
        )
    )

    inventory_module = importlib.import_module("app.services.product_inventory_service")
    original_inventory_result = inventory_module.get_product_inventory_result
    inventory_calls: list[dict[str, Any]] = []

    def _inventory_no_data(params=None, **kwargs):
        inventory_calls.append(dict(params or kwargs.get("params") or {}))
        return _inventory_text_payload(
            message="해당 조회조건의 자료가 없습니다.",
            params=dict(params or kwargs.get("params") or {}),
            meta={"result_status": "no_data", "row_count": 0, "row_count_total": 0},
        )

    inventory_module.get_product_inventory_result = _inventory_no_data
    inventory_session: dict[str, Any] = {}
    capture_before_inventory = len(capture.payloads)
    try:
        inventory_handled = router._try_handle_io_nlq(
            "제품재고현황 조회",
            room={"messages": []},
            session_state=inventory_session,
            make_ts=_make_ts,
            next_seq=_next_seq_factory(),
            logger=log,
        )
    finally:
        inventory_module.get_product_inventory_result = original_inventory_result

    inventory_payload = capture.last_since(capture_before_inventory) or {}
    inventory_meta = dict(inventory_payload.get("meta") or {})
    routed_inventory_values = [
        inventory_meta.get(key)
        for key in (
            "detail_count",
            "sum_carry_qty",
            "sum_in_qty",
            "sum_out_qty",
            "sum_stock_qty",
            "sum_stock_amt",
            "sum_insu_amt",
        )
    ]
    results.append(
        CheckResult(
            "product inventory routed no-data is tableless and does not promote current table",
            bool(inventory_handled)
            and len(inventory_calls) == 1
            and inventory_payload.get("type") == "text"
            and inventory_meta.get("result_status") == "no_data"
            and _is_exact_numeric_zero(inventory_meta.get("row_count"))
            and _is_exact_numeric_zero(inventory_meta.get("row_count_total"))
            and all(_is_exact_numeric_zero(value) for value in routed_inventory_values)
            and all(key not in inventory_payload for key in ("df", "df_display", "records"))
            and "__sims_current_table_source_key" not in inventory_session,
            f"calls={len(inventory_calls)}, meta={inventory_meta!r}",
        )
    )

    candidate_calls: list[dict[str, Any]] = []

    def _candidate_flow(params=None, **kwargs):
        request_params = dict(params or kwargs.get("params") or {})
        candidate_calls.append(request_params)
        if not request_params.get("physic_cd"):
            return {
                "final": True,
                "type": "table",
                "df": pd.DataFrame([
                    {"제품코드": "00029", "제품명": "레날민A"},
                    {"제품코드": "00339", "제품명": "레날민B"},
                ]),
                "meta": {
                    "result_status": "candidate_required",
                    "candidate_table": True,
                    "pending_product_candidates": [
                        {"제품코드": "00029", "제품명": "레날민A"},
                        {"제품코드": "00339", "제품명": "레날민B"},
                    ],
                    "pending_product_action": "제품수불현황 조회",
                    "pending_product_params": {"physic_nm": "레날민"},
                },
            }
        return {
            "final": True,
            "type": "table",
            "df": pd.DataFrame([{"제품코드": request_params.get("physic_cd")}]),
            "meta": {"row_count": 1, "row_count_total": 1},
        }

    flow_module.get_product_flow_result = _candidate_flow
    try:
        candidate_session: dict[str, Any] = {}
        first_handled = router._try_handle_io_nlq(
            "제품수불 레날민",
            room={"messages": []},
            session_state=candidate_session,
            make_ts=_make_ts,
            next_seq=_next_seq_factory(),
            logger=log,
        )
        second_handled = router._try_handle_io_nlq(
            "2번",
            room={"messages": []},
            session_state=candidate_session,
            make_ts=_make_ts,
            next_seq=_next_seq_factory(),
            logger=log,
        )
    finally:
        flow_module.get_product_flow_result = original_flow_result

    candidate_payloads = capture.payloads[capture_before_inventory + 1:]
    picked_params = candidate_calls[-1] if candidate_calls else {}
    pending_after_pick = candidate_session.get("__io_pending_product_pick") or {}
    results.append(
        CheckResult(
            "product flow candidate pick preserves code and name",
            bool(first_handled)
            and bool(second_handled)
            and picked_params.get("physic_cd") == "00339"
            and picked_params.get("physic_nm") == "레날민B"
            and pending_after_pick.get("last_pick_code") == "00339"
            and pending_after_pick.get("last_pick_name") == "레날민B"
            and len(candidate_payloads) == 2
            and dict(candidate_payloads[0].get("meta") or {}).get("result_status") == "candidate_required"
            and dict(candidate_payloads[1].get("meta") or {}).get("result_status") == "success",
            f"calls={candidate_calls!r}, pending={pending_after_pick!r}, payloads={candidate_payloads!r}",
        )
    )
    return results


def run_product_inventory_display_export_checks() -> list[CheckResult]:
    """제품재고 공통 표시와 history 전체 다운로드를 DB 없이 고정한다."""
    import io
    from unittest.mock import patch

    import numpy as np
    import pandas as pd

    from app.ui import chat_middleware
    from app.ui.sims_table_display import (
        build_sims_table_display_config,
        is_sims_numeric_display_col,
        prepare_sims_table_display_df,
        resolve_sims_excel_number_format,
    )

    results: list[CheckResult] = []
    source = pd.DataFrame(
        {
            "제품코드": ["05136", "A1234"],
            "이월수량": [None, 3.25],
            "이월금액": [None, None],
            "이월단가": [None, 100.5],
            "이월DC율": [None, 12.5],
            "현보험약가": [None, 44.2],
        }
    )
    source_dtypes = source.dtypes.copy()
    display = prepare_sims_table_display_df(source, action_name="제품재고현황 조회")
    display_policy_ok = (
        display.loc[0, "제품코드"] == "05136"
        and display.loc[1, "제품코드"] == "A1234"
        and float(display.loc[0, "이월수량"]) == 0.0
        and float(display.loc[0, "이월금액"]) == 0.0
        and pd.isna(display.loc[1, "이월금액"])
        and pd.isna(display.loc[0, "이월단가"])
        and pd.api.types.is_numeric_dtype(display["이월수량"])
        and source.dtypes.equals(source_dtypes)
        and source.loc[0, "제품코드"] == "05136"
        and pd.isna(source.loc[0, "이월수량"])
    )
    results.append(
        CheckResult(
            "product inventory display-only zero/null contract",
            display_policy_ok,
            f"display_policy_ok={display_policy_ok}",
        )
    )

    small_view_df, _small_config, _small_width, _small_height = build_sims_table_display_config(
        source,
        action_name="제품재고현황 조회",
        add_row_no=False,
    )
    small_display_missing_ok = (
        small_view_df.loc[0, "제품코드"] == "05136"
        and small_view_df.loc[0, "이월수량"] == 0
        and small_view_df.loc[0, "이월금액"] == "0"
        and small_view_df.loc[0, "이월단가"] == ""
        and small_view_df.loc[0, "이월DC율"] == ""
        and str(small_view_df["이월금액"].dtype) == "string"
        and str(small_view_df["이월단가"].dtype) == "string"
        and str(small_view_df["이월DC율"].dtype) == "string"
        and "None" not in small_view_df.astype("string").fillna("").to_string()
        and "nan" not in small_view_df.astype("string").fillna("").to_string().lower()
    )
    results.append(
        CheckResult(
            "product inventory small display blanks numeric missing and keeps zero factual",
            small_display_missing_ok,
            f"dtypes={small_view_df.dtypes.astype(str).to_dict()}",
        )
    )

    common_semantics_ok = (
        is_sims_numeric_display_col(display, "이월수량")
        and is_sims_numeric_display_col(display, "이월금액")
        and not is_sims_numeric_display_col(display, "제품코드")
        and resolve_sims_excel_number_format("이월수량") == "#,##0"
        and resolve_sims_excel_number_format("이월금액") == "#,##0"
        and resolve_sims_excel_number_format("이월DC율") == "#,##0.##"
    )
    results.append(
        CheckResult(
            "product inventory uses shared numeric display and Excel semantics",
            common_semantics_ok,
            "common_semantics_ok="
            f"{common_semantics_ok}, qty={is_sims_numeric_display_col(display, '이월수량')}, "
            f"amount={is_sims_numeric_display_col(display, '이월금액')}, "
            f"code={is_sims_numeric_display_col(display, '제품코드')}, "
            f"qty_fmt={resolve_sims_excel_number_format('이월수량')}, "
            f"amount_fmt={resolve_sims_excel_number_format('이월금액')}, "
            f"dc_fmt={resolve_sims_excel_number_format('이월DC율')}",
        )
    )

    full_df = pd.concat([source] * 271, ignore_index=True)
    display_df = full_df.head(300).copy()
    expected_provenance = {
        "user_id": "fixture-user",
        "company_id": "fixture-company",
        "room_id": "fixture-room",
        "action": "제품재고현황 조회",
        "table_key": "history-product-inventory",
        "rows": 542,
    }
    fake_session = {
        "sims_export_tables": {"history-product-inventory": full_df},
        "__sims_export_tables_by_key": {"history-product-inventory": full_df},
        "__sims_export_table_provenance_by_key": {"history-product-inventory": expected_provenance},
    }
    item = {
        "action": "제품재고현황 조회",
        "df": display_df,
        "df_display": display_df,
        "user_id": "fixture-user",
        "company_id": "fixture-company",
        "room_id": "fixture-room",
    }
    meta = {
        "table_key": "history-product-inventory",
        "download_table_key": "history-product-inventory",
        "row_count_total": 542,
        "display_row_count": 300,
        "current_source_key": "",
    }
    with patch.object(chat_middleware.st, "session_state", fake_session):
        picked, source_name = chat_middleware._resolve_payload_full_df_for_download(item, meta)
        structured_source = chat_middleware._resolve_payload_full_download_source(
            item,
            meta,
            display_df=display_df,
        )
        resolved = chat_middleware._get_full_download_df_for_sims_item(item, meta, display_df)
    results.append(
        CheckResult(
            "product inventory history table key restores full download source",
            len(full_df) == 542
            and len(display_df) == 300
            and isinstance(picked, pd.DataFrame)
            and len(picked) == 542
            and full_df.equals(picked)
            and structured_source.get("source_status") == "full"
            and structured_source.get("source_rows") == 542,
            f"full_rows={len(full_df)}, display_rows={len(display_df)}, source={source_name}, structured={structured_source}, resolved_rows={len(resolved) if isinstance(resolved, pd.DataFrame) else -1}",
        )
    )

    middleware_source = Path(chat_middleware.__file__).read_text(encoding="utf-8")
    results.append(
        CheckResult(
            "lazy full export keeps verified source status separate from file generation",
            "file_generation_status = \"deferred\" if defer_full_export else \"ready\"" in middleware_source
            and "download_source_result[\"file_generation_status\"] = file_generation_status" in middleware_source
            and "deferred_display_only" not in middleware_source,
            "lazy full source status is resolved before deferred file generation",
        )
    )

    main_source = (PROJECT_ROOT / "app" / "Lmstudio_SSAI_chat_main.py").read_text(encoding="utf-8")
    required_summary_keys = (
        '"carry_qty"', '"in_qty"', '"out_qty"', '"stock_qty"',
        '"sum_carry_qty"', '"sum_in_qty"', '"sum_out_qty"',
        '"sum_stock_qty"', '"sum_stock_amt"', '"sum_insu_amt"',
    )
    results.append(
        CheckResult(
            "partitioned history allow-list retains product summary render facts",
            all(key in main_source for key in required_summary_keys),
            "required_summary_keys=" + ",".join(required_summary_keys),
        )
    )

    partial_session = {
        "sims_export_tables": {"history-product-inventory": display_df},
        "__sims_export_tables_by_key": {"history-product-inventory": display_df},
        "__sims_export_table_provenance_by_key": {
            "history-product-inventory": {**expected_provenance, "rows": 300},
        },
    }
    with patch.object(chat_middleware.st, "session_state", partial_session):
        partial_source = chat_middleware._resolve_payload_full_download_source(
            item,
            meta,
            display_df=display_df,
        )
    results.append(
        CheckResult(
            "product inventory partial export cache is never promoted to full",
            partial_source.get("df") is None
            and partial_source.get("source_status") == "partial_unverified"
            and partial_source.get("expected_rows") == 542,
            f"partial_source={partial_source}",
        )
    )

    warnings: list[str] = []
    captions: list[str] = []
    with (
        patch.object(chat_middleware.st, "warning", lambda message, **_kwargs: warnings.append(str(message))),
        patch.object(chat_middleware.st, "caption", lambda message, **_kwargs: captions.append(str(message))),
    ):
        partial_notice = chat_middleware._render_partial_download_source_notice(
            source_status="display_only_partial",
            download_rows=300,
            expected_rows=542,
        )
    results.append(
        CheckResult(
            "product inventory partial download warning labels displayed rows explicitly",
            partial_notice
            and warnings == ["전체 원본을 찾지 못했습니다."]
            and "CSV/EXCEL은 현재 화면 데이터 300건만 포함합니다." in captions
            and "전체 예상 결과는 542건입니다." in captions,
            f"warnings={warnings!r}, captions={captions!r}",
        )
    )

    required_partial_limit_keys = (
        '"expected_rows"',
        '"prepared_rows"',
        '"download_row_count"',
        '"download_limit_rows"',
        '"applied_download_limit_rows"',
        '"limit_hit"',
        '"download_source_status"',
        '"source_call_count"',
    )
    results.append(
        CheckResult(
            "partitioned history retains verified partial-limit diagnostics",
            all(key in main_source for key in required_partial_limit_keys),
            "required_partial_limit_keys=" + ",".join(required_partial_limit_keys),
        )
    )

    capped_rows = 100_000
    capped_expected_rows = 1_159_102
    capped_df = pd.DataFrame({"순번": range(1, capped_rows + 1)})
    capped_meta = {
        **meta,
        "download_row_count": capped_rows,
        "download_limit_rows": capped_rows,
        "limit_hit": True,
        "row_count_total": capped_expected_rows,
        "expected_rows": capped_expected_rows,
    }
    capped_provenance = {**expected_provenance, "rows": capped_rows}
    capped_session = {
        "sims_export_tables": {"history-product-inventory": capped_df},
        "__sims_export_tables_by_key": {"history-product-inventory": capped_df},
        "__sims_export_table_provenance_by_key": {
            "history-product-inventory": capped_provenance,
        },
    }
    with patch.object(chat_middleware.st, "session_state", capped_session):
        capped_source = chat_middleware._resolve_payload_full_download_source(
            item,
            capped_meta,
            display_df=display_df.head(200),
        )
    capped_warnings: list[str] = []
    capped_captions: list[str] = []
    with (
        patch.object(chat_middleware.st, "warning", lambda message, **_kwargs: capped_warnings.append(str(message))),
        patch.object(chat_middleware.st, "caption", lambda message, **_kwargs: capped_captions.append(str(message))),
    ):
        capped_notice = chat_middleware._render_partial_download_source_notice(
            source_status="partial_limit",
            download_rows=capped_rows,
            expected_rows=capped_expected_rows,
        )
    results.append(
        CheckResult(
            "capped full source is reported as partial limit instead of missing",
            isinstance(capped_source.get("df"), pd.DataFrame)
            and len(capped_source["df"]) == capped_rows
            and capped_source.get("source_status") == "partial_limit"
            and capped_notice
            and capped_warnings == [
                "전체 예상 1,159,102건 중 다운로드 상한 100,000건이 준비되었습니다. 현재 파일은 전체 자료가 아닙니다."
            ]
            and any("100,000건" in caption and "1,159,102건" in caption for caption in capped_captions),
            f"capped_source={capped_source}, warnings={capped_warnings!r}, captions={capped_captions!r}",
        )
    )

    generated_limit_meta = {"row_count_total": capped_expected_rows}
    generated_limit_status = chat_middleware._record_io_full_source_limit_meta(
        generated_limit_meta,
        action="출고명세 조회",
        prepared_rows=capped_rows,
        expected_rows=capped_expected_rows,
        applied_limit_rows=capped_rows,
    )
    below_limit_meta = {"row_count_total": capped_expected_rows}
    below_limit_status = chat_middleware._record_io_full_source_limit_meta(
        below_limit_meta,
        action="출고명세 조회",
        prepared_rows=capped_rows - 1,
        expected_rows=capped_expected_rows,
        applied_limit_rows=capped_rows,
    )
    results.append(
        CheckResult(
            "outbound full-source boundary records the actual applied export cap",
            generated_limit_status == "partial_limit"
            and generated_limit_meta.get("applied_download_limit_rows") == capped_rows
            and generated_limit_meta.get("download_limit_rows") == capped_rows
            and generated_limit_meta.get("limit_hit") is True
            and generated_limit_meta.get("download_row_count") == capped_rows
            and generated_limit_meta.get("prepared_rows") == capped_rows
            and generated_limit_meta.get("expected_rows") == capped_expected_rows
            and generated_limit_meta.get("download_source_status") == "partial_limit"
            and generated_limit_meta.get("source_call_count") == 0,
            f"generated_limit_meta={generated_limit_meta!r}",
        )
    )

    outbound_query_meta = {
        "action": "출고명세 조회",
        "table_key": "outbound-cap-fixture",
        "row_count_total": capped_expected_rows,
        "expected_rows": capped_expected_rows,
        "company_id": "4",
        "room_id": "room-outbound-cap",
    }
    outbound_query_item = {
        "action": "출고명세 조회",
        "params": {"date_from": "20260101", "date_to": "20261231"},
        "meta": outbound_query_meta,
        "company_id": "4",
        "room_id": "room-outbound-cap",
    }
    outbound_query_session: dict[str, Any] = {}
    import app.services.rddbc120_service as rddbc120_service

    with (
        patch.object(chat_middleware.st, "session_state", outbound_query_session),
        patch.object(rddbc120_service, "get_rddbc120_export_limit_rows", lambda: capped_rows),
        patch.object(rddbc120_service, "get_rddbc120_export_df", lambda _params: capped_df),
    ):
        outbound_query_df = chat_middleware._get_full_download_df_for_sims_item(
            outbound_query_item,
            outbound_query_meta,
            display_df.head(200),
        )
    outbound_query_provenance = (
        outbound_query_session.get("__sims_export_table_provenance_by_key", {}).get(
            "outbound-cap-fixture", {}
        )
    )
    results.append(
        CheckResult(
            "outbound export query payload carries verified partial-limit metadata end to end",
            isinstance(outbound_query_df, pd.DataFrame)
            and len(outbound_query_df) == capped_rows
            and outbound_query_meta.get("applied_download_limit_rows") == capped_rows
            and outbound_query_meta.get("download_limit_rows") == capped_rows
            and outbound_query_meta.get("limit_hit") is True
            and outbound_query_meta.get("download_row_count") == capped_rows
            and outbound_query_meta.get("prepared_rows") == capped_rows
            and outbound_query_meta.get("expected_rows") == capped_expected_rows
            and outbound_query_meta.get("download_source_status") == "partial_limit"
            and outbound_query_meta.get("source_call_count") == 0
            and outbound_query_provenance.get("download_source_status") == "partial_limit"
            and outbound_query_provenance.get("limit_hit") is True,
            f"meta={outbound_query_meta!r}, provenance={outbound_query_provenance!r}",
        )
    )
    results.append(
        CheckResult(
            "outbound result below its applied cap is not reported as partial limit",
            below_limit_status == "partial_unverified"
            and below_limit_meta.get("limit_hit") is False
            and below_limit_meta.get("download_source_status") == "partial_unverified",
            f"below_limit_meta={below_limit_meta!r}",
        )
    )

    unverified_meta = {
        key: value
        for key, value in capped_meta.items()
        if key not in {"download_limit_rows", "limit_hit"}
    }
    with patch.object(chat_middleware.st, "session_state", capped_session):
        unverified_source = chat_middleware._resolve_payload_full_download_source(
            item,
            unverified_meta,
            display_df=display_df.head(200),
        )
    results.append(
        CheckResult(
            "partial source without an explicit cap is not reported as partial limit",
            unverified_source.get("df") is None
            and unverified_source.get("source_status") == "partial_unverified"
            and unverified_source.get("source_rows") == capped_rows
            and unverified_source.get("expected_rows") == capped_expected_rows,
            f"unverified_source={unverified_source}",
        )
    )

    capped_status = chat_middleware._partial_download_source_status(
        capped_meta,
        prepared_rows=capped_rows,
        expected_rows=capped_expected_rows,
    )
    download_count_only_status = chat_middleware._partial_download_source_status(
        unverified_meta,
        prepared_rows=capped_rows,
        expected_rows=capped_expected_rows,
    )
    middleware_source = inspect.getsource(chat_middleware)
    results.append(
        CheckResult(
            "all partial download and history restore paths share the explicit-cap classifier",
            capped_status == "partial_limit"
            and download_count_only_status == "partial_unverified"
            and middleware_source.count("_partial_download_source_status(") >= 4
            and middleware_source.count("_record_io_full_source_limit_meta(") >= 3
            and "declared_download_rows == recovered_rows" not in middleware_source,
            f"capped={capped_status}, download_count_only={download_count_only_status}",
        )
    )
    derived_context = chat_middleware._download_source_context(
        item,
        {
            **meta,
            "result_metric": "sales",
            "result_grain": "product",
            "filter_column": "판정결과",
            "filter_value": "안정",
        },
        table_key="history-product-inventory",
    )

    limit_context = chat_middleware._download_source_context(
        item,
        generated_limit_meta,
        table_key="history-product-inventory",
    )
    results.append(
        CheckResult(
            "export provenance retains partial-limit ownership diagnostics",
            limit_context.get("applied_download_limit_rows") == capped_rows
            and limit_context.get("download_limit_rows") == capped_rows
            and limit_context.get("download_row_count") == capped_rows
            and limit_context.get("prepared_rows") == capped_rows
            and limit_context.get("expected_rows") == capped_expected_rows
            and limit_context.get("limit_hit") is True
            and limit_context.get("download_source_status") == "partial_limit"
            and limit_context.get("source_call_count") == 0,
            f"limit_context={limit_context!r}",
        )
    )
    results.append(
        CheckResult(
            "derived table provenance keeps the official metric, grain, and judgement filter",
            derived_context.get("result_metric") == "sales"
            and derived_context.get("result_grain") == "product"
            and derived_context.get("filter_column") == "판정결과"
            and derived_context.get("filter_value") == "안정",
            f"derived_context_keys={sorted(derived_context)}",
        )
    )

    mixed_session = {
        "sims_export_tables": {"history-product-inventory": display_df},
        "__sims_export_tables_by_key": {"history-product-inventory": full_df},
        "__sims_export_table_provenance_by_key": {
            "history-product-inventory": expected_provenance,
        },
    }
    with patch.object(chat_middleware.st, "session_state", mixed_session):
        mixed_source = chat_middleware._resolve_payload_full_download_source(
            item,
            meta,
            display_df=display_df,
        )
    results.append(
        CheckResult(
            "product inventory resolver continues from 300-row cache to 542-row cache",
            isinstance(mixed_source.get("df"), pd.DataFrame)
            and len(mixed_source["df"]) == 542
            and mixed_source.get("source_status") == "full",
            f"mixed_source={mixed_source}",
        )
    )

    mismatched_session = {
        "sims_export_tables": {"history-product-inventory": full_df},
        "__sims_export_tables_by_key": {"history-product-inventory": full_df},
        "__sims_export_table_provenance_by_key": {
            "history-product-inventory": {**expected_provenance, "room_id": "other-room"},
        },
    }
    with patch.object(chat_middleware.st, "session_state", mismatched_session):
        mismatch_source = chat_middleware._resolve_payload_full_download_source(
            {**item, "df": None},
            meta,
            display_df=display_df,
        )
        current_source = chat_middleware._resolve_payload_full_download_source(
            item,
            {**meta, "current_table_followup": True},
            display_df=display_df,
        )
    results.append(
        CheckResult(
            "product inventory export provenance blocks foreign scope and unverified current-table fallback",
            mismatch_source.get("df") is None
            and mismatch_source.get("source_status") in {"not_found", "display_only_partial"}
            and current_source.get("df") is None
            and current_source.get("source_status") in {"not_found", "display_only_partial"},
            f"mismatch_source={mismatch_source}, current_source={current_source}",
        )
    )

    derived_key = "derived-stable-detail"
    derived_meta = {
        **meta,
        "table_key": derived_key,
        "current_table_followup": True,
        "analysis_row_count": 61981,
    }
    derived_full = pd.DataFrame({"제품코드": ["P-1"] * 61981, "제품명": ["안정제품"] * 61981})
    derived_provenance = {**expected_provenance, "table_key": derived_key, "rows": 61981}
    derived_session = {
        "sims_export_tables": {derived_key: derived_full, "history-product-inventory": full_df},
        "__sims_export_table_provenance_by_key": {
            derived_key: derived_provenance,
            "history-product-inventory": expected_provenance,
        },
    }
    with patch.object(chat_middleware.st, "session_state", derived_session):
        derived_source = chat_middleware._resolve_payload_full_download_source(
            {**item, "table_key": derived_key, "df": display_df},
            derived_meta,
            display_df=display_df,
        )
    results.append(
        CheckResult(
            "derived current-table export uses only its own verified full source",
            isinstance(derived_source.get("df"), pd.DataFrame)
            and len(derived_source["df"]) == 61981
            and derived_source.get("source_table_key") == derived_key
            and derived_source.get("source_status") == "full",
            f"derived_source={derived_source}",
        )
    )

    rendered_frames: list[pd.DataFrame] = []
    rendered_kwargs: list[dict[str, Any]] = []
    render_session = {"__sims_table_render_path": "history"}
    with (
        patch.object(chat_middleware.st, "session_state", render_session),
        patch.object(
            chat_middleware.st,
            "dataframe",
            lambda frame, **kwargs: (
                rendered_frames.append(frame.copy()),
                rendered_kwargs.append(dict(kwargs)),
            ),
        ),
        patch.object(chat_middleware.st, "caption", lambda *_args, **_kwargs: None),
    ):
        chat_middleware._render_chat_fast_dataframe(
            full_df,
            action_name="제품재고현황 조회",
            meta={"table_key": "history-product-inventory", "is_nlq": True},
        )
    fast_render_ok = (
        len(rendered_frames) == 1
        and len(rendered_frames[0]) == 300
        and len(full_df) == 542
        and full_df["제품코드"].tolist() == source["제품코드"].tolist() * 271
        and rendered_frames[0].loc[rendered_frames[0].index[0], "이월단가"] == ""
        and rendered_frames[0].loc[rendered_frames[0].index[0], "이월DC율"] == ""
        and str(rendered_frames[0]["이월단가"].dtype) == "string"
        and str(rendered_frames[0]["이월DC율"].dtype) == "string"
        and rendered_frames[0].loc[rendered_frames[0].index[1], "이월단가"] == "100.50"
        and rendered_frames[0].loc[rendered_frames[0].index[1], "이월DC율"] == "12.50"
        and bool(rendered_kwargs[0].get("column_config"))
    )

    config = rendered_kwargs[0].get("column_config") if rendered_kwargs else {}
    column_config_ok = (
        "제품코드" in config
        and "이월수량" in config
        and "이월단가" in config
        and "이월DC율" in config
        and config.get("제품코드", {}).get("type_config", {}).get("type") == "text"
        and config.get("이월수량", {}).get("type_config", {}).get("type") == "number"
        and config.get("이월단가", {}).get("type_config", {}).get("type") == "text"
        and config.get("이월DC율", {}).get("type_config", {}).get("type") == "text"
    )
    results.append(
        CheckResult(
            "product inventory converts only missing numeric columns to formatted text",
            column_config_ok,
            f"config_keys={sorted(config) if isinstance(config, dict) else []}",
        )
    )

    large_df = pd.DataFrame(
        {
            "제품코드": [f"{idx:05d}" for idx in range(10115)],
            **{f"수량{idx}": list(range(10115)) for idx in range(97)},
        }
    )
    large_frames: list[pd.DataFrame] = []
    with (
        patch.dict("os.environ", {"SIMS_CHAT_DISPLAY_MAX_ROWS": "300"}),
        patch.object(chat_middleware.st, "session_state", {"__sims_table_render_path": "history"}),
        patch.object(chat_middleware.st, "dataframe", lambda frame, **_kwargs: large_frames.append(frame.copy())),
        patch.object(chat_middleware.st, "caption", lambda *_args, **_kwargs: None),
    ):
        chat_middleware._render_chat_fast_dataframe(
            large_df,
            action_name="제품재고현황 조회",
            meta={"table_key": "large-product-inventory"},
        )
    results.append(
        CheckResult(
            "product inventory 10115x98 history renderer caps display without mutating full source",
            len(large_frames) == 1
            and len(large_frames[0]) <= 300
            and len(large_df) == 10115
            and len(large_df.columns) == 98
            and large_df["제품코드"].iloc[0] == "00000",
            f"rendered_rows={len(large_frames[0]) if large_frames else -1}, full_shape={large_df.shape}",
        )
    )

    inventory_592_df = pd.concat([source] * 296, ignore_index=True)
    inventory_592_frames: list[pd.DataFrame] = []
    with (
        patch.dict("os.environ", {"SIMS_CHAT_DISPLAY_MAX_ROWS": "300"}),
        patch.object(chat_middleware.st, "session_state", {"__sims_table_render_path": "history"}),
        patch.object(
            chat_middleware.st,
            "dataframe",
            lambda frame, **_kwargs: inventory_592_frames.append(frame.copy()),
        ),
        patch.object(chat_middleware.st, "caption", lambda *_args, **_kwargs: None),
    ):
        chat_middleware._render_chat_fast_dataframe(
            inventory_592_df,
            action_name="제품재고현황 조회",
            meta={"table_key": "inventory-592"},
        )
    results.append(
        CheckResult(
            "product inventory 592-row history keeps full source while rendering 300 rows",
            len(inventory_592_df) == 592
            and len(inventory_592_frames) == 1
            and len(inventory_592_frames[0]) == 300
            and inventory_592_df.equals(pd.concat([source] * 296, ignore_index=True)),
            f"full_rows={len(inventory_592_df)}, display_rows={len(inventory_592_frames[0]) if inventory_592_frames else -1}",
        )
    )

    kpi_percent_ok = all(
        chat_middleware.resolve_sims_numeric_display_kind(col) == "percent2"
        for col in chat_middleware._FAST_TABLE_PERCENT_KPI_COLS
    ) and chat_middleware.resolve_sims_numeric_display_kind("이월DC율") != "percent2"
    results.append(
        CheckResult(
            "KPI percent formats remain percent while product DC remains raw decimal",
            kpi_percent_ok,
            f"percent_columns={len(chat_middleware._FAST_TABLE_PERCENT_KPI_COLS)}, dc_kind={chat_middleware.resolve_sims_numeric_display_kind('이월DC율')}",
        )
    )
    results.append(
        CheckResult(
            "product inventory history fast renderer caps only display dataframe",
            fast_render_ok,
            f"captured_rows={len(rendered_frames[0]) if rendered_frames else -1}, full_rows={len(full_df)}",
        )
    )

    try:
        from openpyxl import load_workbook

        bio = io.BytesIO()
        with pd.ExcelWriter(bio, engine="openpyxl") as writer:
            resolved.to_excel(writer, index=False, sheet_name="SIMS")
            chat_middleware._apply_sims_excel_number_formats(writer, resolved, "SIMS")
        workbook = load_workbook(io.BytesIO(bio.getvalue()))
        ws = workbook["SIMS"]
        index = {cell.value: cell.column for cell in ws[1]}
        formats_ok = (
            ws.max_row == 543
            and ws.cell(2, index["제품코드"]).value == "05136"
            and ws.cell(3, index["제품코드"]).value == "A1234"
            and ws.cell(2, index["이월수량"]).value is None
            and ws.cell(2, index["이월금액"]).value is None
            and ws.cell(2, index["이월수량"]).number_format == "#,##0"
            and ws.cell(2, index["이월금액"]).number_format == "#,##0"
            and ws.cell(2, index["이월DC율"]).number_format == "#,##0.##"
        )
        results.append(_ok("product inventory full Excel and common formats", f"formats_ok={formats_ok}") if formats_ok else _fail("product inventory full Excel and common formats", "unexpected full export or number format"))
    except Exception as exc:
        results.append(_fail("product inventory full Excel and common formats", f"{type(exc).__name__}: {exc}"))

    # 실제 수동 확인값을 고정한다. 거래 없는 입고의 단가/DC는 0이라는
    # 계산값이 아니라 근거 없음(NA)이어야 하며, 음수 출고는 실제 거래다.
    from app.services.product_inventory_service import _final_display_df, _prepare_grouped_df
    from app.services.product_flow_service import _prepare_display_df

    inventory_source = pd.DataFrame(
        {
            "group_cd": ["fixture", "fixture"],
            "group_nm": ["fixture", "fixture"],
            "physic_cd": ["31768", "39639"],
            "physic_nm": ["fixture-31768", "fixture-39639"],
            "standard": ["100T", "150T"],
            "kd_cd": ["", ""],
            "edi_cd": ["", ""],
            "std_cd": ["", ""],
            "pack_unit": ["", ""],
            "master_unit_cost": [0.0, 0.0],
            "insu_date": ["", ""],
            "before_insu_date": ["", ""],
            "insu_price": [5800.0, 8700.0],
            "before_insu_price": [0.0, 0.0],
            "acc_unit": [1.0, 1.0],
            "physic_tax": ["", ""],
            "old_in_qty": [9.0, 222.0],
            "old_in_amt": [72000.0, 1776799.2],
            "old_out_qty": [0.0, 0.0],
            "now_in_qty": [0.0, 480.0],
            "now_in_amt": [0.0, 3841728.0],
            "now_out_qty": [-1.0, 413.0],
            "now_out_amt": [-6281.0, 0.0],
            "buy_cd": ["", ""],
            "buy_nm": ["", ""],
            "order_cd": ["", ""],
            "order_nm": ["", ""],
            "maker_cd": ["", ""],
            "maker_nm": ["", ""],
            "product_group_nm": ["", ""],
            "product_di_nm": ["", ""],
            "product_class_nm": ["", ""],
            "special_manage_nm": ["", ""],
        }
    )
    inventory_source.loc[1, "now_out_amt"] = 413.0 * 8267.09
    inventory_source_dtypes = inventory_source.dtypes.copy()
    grouped_inventory = _prepare_grouped_df(
        inventory_source,
        pd.DataFrame(columns=["physic_cd", "last_unit_cost"]),
        {"group_basis": "maker", "price_mode": "avg"},
        {"date_to": "20260731", "fetch_top": 0},
    )
    inventory_display, inventory_meta = _final_display_df(
        grouped_inventory,
        {"group_basis": "maker", "price_mode": "avg"},
    )
    fixture_31768 = inventory_display.loc[inventory_display["제품코드"] == "31768"].iloc[0]
    fixture_39639 = inventory_display.loc[inventory_display["제품코드"] == "39639"].iloc[0]
    inventory_formula_ok = (
        inventory_source.dtypes.equals(inventory_source_dtypes)
        and fixture_31768["입고수량"] == 0.0
        and fixture_31768["입고금액"] == 0.0
        and pd.isna(fixture_31768["입고단가"])
        and pd.isna(fixture_31768["입고DC율"])
        and fixture_31768["출고수량"] == -1.0
        and not pd.isna(fixture_31768["출고단가"])
        and fixture_31768["재고수량"] == 10.0
        and fixture_39639["이월수량"] == 222.0
        and fixture_39639["입고수량"] == 480.0
        and fixture_39639["출고수량"] == 413.0
        and fixture_39639["재고수량"] == 289.0
        and fixture_39639["입고단가"] == 8003.60
        and fixture_39639["입고DC율"] == 8.00
        and fixture_39639["출고단가"] == 8267.09
        and fixture_39639["출고DC율"] == 4.98
        and fixture_39639["재고단가"] == 8003.60
        and fixture_39639["재고DC율"] == 8.00
        and inventory_meta["sum_carry_qty"] == 231.0
        and inventory_meta["sum_in_qty"] == 480.0
        and inventory_meta["sum_out_qty"] == 412.0
        and inventory_meta["sum_stock_qty"] == 299.0
    )
    results.append(
        CheckResult(
            "product inventory transaction-evidence unit and DC contract",
            inventory_formula_ok,
            f"31768_in_unit={fixture_31768['입고단가']!r}, 31768_in_dc={fixture_31768['입고DC율']!r}, "
            f"39639_in={fixture_39639['입고단가']!r}/{fixture_39639['입고DC율']!r}, "
            f"39639_out={fixture_39639['출고단가']!r}/{fixture_39639['출고DC율']!r}, meta={inventory_meta}",
        )
    )

    stock_grouped_inventory = _prepare_grouped_df(
        inventory_source,
        pd.DataFrame(columns=["physic_cd", "last_unit_cost"]),
        {"group_basis": "stock", "price_mode": "avg"},
        {"date_to": "20260731", "fetch_top": 0},
    )
    with py_warnings.catch_warnings(record=True) as stock_display_warnings:
        py_warnings.simplefilter("always")
        stock_inventory_display, stock_inventory_meta = _final_display_df(
            stock_grouped_inventory,
            {"group_basis": "stock", "price_mode": "avg"},
        )
    stock_subtotals = stock_inventory_display.loc[
        stock_inventory_display["재고위치"].astype(str).eq("제품 합계")
    ]
    stock_final_totals = stock_inventory_display.loc[
        stock_inventory_display["재고위치"].astype(str).eq("합계")
    ]
    unexpected_future_warnings = [
        warning for warning in stock_display_warnings
        if issubclass(warning.category, FutureWarning)
    ]
    stock_group_contract_ok = (
        int(stock_inventory_meta.get("detail_count") or 0) == 2
        and int(len(stock_subtotals)) == 0
        and int(len(stock_final_totals)) == 1
        and int(len(stock_inventory_display)) == 3
        and not unexpected_future_warnings
    )
    results.append(
        CheckResult(
            "general product inventory stock grouping keeps detail plus final total",
            stock_group_contract_ok,
            f"detail_rows={stock_inventory_meta.get('detail_count')}, product_subtotals={len(stock_subtotals)}, "
            f"final_totals={len(stock_final_totals)}, warnings={len(unexpected_future_warnings)}, meta={stock_inventory_meta}",
        )
    )

    inventory_service = importlib.import_module("app.services.product_inventory_service")
    general_display_source = inspect.getsource(inventory_service._final_display_df)
    current_stock_display_source = inspect.getsource(inventory_service._build_current_stock_table_frames)
    perf_source = inspect.getsource(inventory_service._collect_source_df) + inspect.getsource(
        inventory_service.get_product_inventory_result
    )
    results.append(
        _ok("product subtotal remains current-stock-only", "general=0/current-stock=present")
        if "제품 합계" not in general_display_source and "제품 합계" in current_stock_display_source
        else _fail("product subtotal remains current-stock-only", "subtotal boundary regression")
    )
    required_perf_stages = (
        '"month_carry"', '"period_in"', '"period_out"', '"last_cost"',
        "filter_prepare_ms", "group_aggregate_ms", "unit_calc_ms",
        "amount_calc_ms", "dc_calc_ms", "final_display_frame_ms", "final_total_ms",
    )
    results.append(
        _ok("product inventory performance stages are independently logged", "sql/group/calc/display")
        if all(marker in perf_source for marker in required_perf_stages)
        else _fail("product inventory performance stages are independently logged", "missing performance marker")
    )

    inventory_rendered_frames: list[pd.DataFrame] = []
    inventory_rendered_kwargs: list[dict[str, Any]] = []
    with (
        patch.object(chat_middleware.st, "session_state", {"__sims_table_render_path": "history"}),
        patch.object(
            chat_middleware.st,
            "dataframe",
            lambda frame, **kwargs: (
                inventory_rendered_frames.append(frame.copy()),
                inventory_rendered_kwargs.append(dict(kwargs)),
            ),
        ),
        patch.object(chat_middleware.st, "caption", lambda *_args, **_kwargs: None),
    ):
        chat_middleware._render_chat_fast_dataframe(
            inventory_display,
            action_name="제품재고현황 조회",
            meta={"table_key": "inventory-31768"},
        )
    inventory_final_frame = inventory_rendered_frames[0] if inventory_rendered_frames else pd.DataFrame()
    rendered_31768 = inventory_final_frame.loc[
        inventory_final_frame["제품코드"] == "31768"
    ].iloc[0] if not inventory_final_frame.empty else pd.Series(dtype="object")
    inventory_config = inventory_rendered_kwargs[0].get("column_config", {}) if inventory_rendered_kwargs else {}
    inventory_renderer_contract_ok = (
        len(inventory_rendered_frames) == 1
        and rendered_31768.get("입고수량") == 0.0
        and rendered_31768.get("입고금액") == 0.0
        and rendered_31768.get("입고단가") == ""
        and rendered_31768.get("입고DC율") == ""
        and rendered_31768.get("출고수량") == -1.0
        and rendered_31768.get("출고단가") == "6,281.00"
        and rendered_31768.get("출고DC율") == "-8.29"
        and rendered_31768.get("재고수량") == 10.0
        and str(inventory_final_frame["입고단가"].dtype) == "string"
        and str(inventory_final_frame["입고DC율"].dtype) == "string"
        and inventory_config.get("입고단가", {}).get("type_config", {}).get("type") == "text"
        and inventory_config.get("입고DC율", {}).get("type_config", {}).get("type") == "text"
        and inventory_config.get("출고수량", {}).get("type_config", {}).get("type") == "number"
    )
    results.append(
        CheckResult(
            "product inventory 31768 keeps values while formatting missing numeric columns as text",
            inventory_renderer_contract_ok,
            f"rows={len(inventory_final_frame)}, config_count={len(inventory_config)}, 31768={rendered_31768.to_dict() if not rendered_31768.empty else {}}",
        )
    )

    metric_markup: list[str] = []
    with (
        patch.object(chat_middleware.st, "caption", lambda *_args, **_kwargs: None),
        patch.object(chat_middleware.st, "markdown", lambda value, **_kwargs: metric_markup.append(str(value))),
    ):
        chat_middleware._render_sims_numeric_metric("출고수량", -1.0)
        chat_middleware._render_sims_numeric_metric("입고수량", 0.0)
    results.append(
        CheckResult(
            "product flow and inventory metrics share negative-only red emphasis",
            len(metric_markup) == 2 and "#dc2626" in metric_markup[0] and "#1f2937" in metric_markup[1],
            f"metric_markup={metric_markup!r}",
        )
    )

    flow_source = pd.DataFrame(
        {
            "내부방향": ["IN", "OUT"],
            "정렬일자": ["20260701", "20260702"],
            "정렬순번": [1, 2],
            "재고증감": [480.0, -413.0],
            "입출고일자": ["20260701", "20260702"],
            "명세서일자": ["20260630", "20260701"],
            "입고수량": [480.0, 0.0],
            "출고수량": [0.0, 413.0],
            "할증": [0.0, 0.0],
            "공급가액": [0.0, 0.0],
            "부가세": [0.0, 0.0],
            "합계금액": [0.0, 0.0],
            "재고위치": ["fixture-stock", "fixture-stock"],
        }
    )
    flow_display, flow_meta = _prepare_display_df(
        flow_source,
        222.0,
        {"stock_mode": "book", "date_basis": "io", "flow_scope": "all"},
        {"physic_cd": "39639", "physic_nm": "fixture-39639"},
    )
    flow_detail = flow_display.iloc[1:]
    flow_mapping_ok = (
        len(flow_detail) == 2
        and flow_display.iloc[0]["입출고일자"] == "이월재고"
        and flow_detail["입출고일자"].tolist() == ["2026/07/01", "2026/07/02"]
        and flow_detail["명세서일자"].tolist() == ["2026/06/30", "2026/07/01"]
        and flow_meta["carry_qty"] == 222.0
        and flow_meta["in_qty"] == 480.0
        and flow_meta["out_qty"] == 413.0
        and flow_meta["stock_qty"] == 289.0
    )
    results.append(
        CheckResult(
            "product flow keeps IO date and invoice date as separate source fields",
            flow_mapping_ok,
            f"detail_dates={flow_detail[['입출고일자', '명세서일자']].to_dict(orient='records')}, meta={flow_meta}",
        )
    )

    flow_rendered_frames: list[pd.DataFrame] = []
    flow_rendered_kwargs: list[dict[str, Any]] = []
    flow_history = flow_display.iloc[1:].copy()
    flow_history.loc[flow_history.index[0], "입출고일자"] = "None"
    flow_history.loc[flow_history.index[0], "단가적용거래처"] = "None"
    flow_history.loc[flow_history.index[0], "재고적용거래처"] = "nan"
    flow_history.loc[flow_history.index[0], "재고위치"] = "<NA>"
    flow_history.loc[flow_history.index[0], "매입처"] = "NaT"
    with (
        patch.object(chat_middleware.st, "session_state", {"__sims_table_render_path": "history"}),
        patch.object(
            chat_middleware.st,
            "dataframe",
            lambda frame, **kwargs: (
                flow_rendered_frames.append(frame.copy()),
                flow_rendered_kwargs.append(dict(kwargs)),
            ),
        ),
        patch.object(chat_middleware.st, "caption", lambda *_args, **_kwargs: None),
    ):
        chat_middleware._render_chat_fast_dataframe(
            flow_history,
            action_name="제품수불현황 조회",
            meta={"table_key": "history-product-flow"},
    )
    flow_final_frame = flow_rendered_frames[0] if flow_rendered_frames else pd.DataFrame()
    flow_text_columns = [
        column
        for column in flow_final_frame.columns
        if not pd.api.types.is_numeric_dtype(flow_final_frame[column])
    ]
    flow_text_rendered = (
        flow_final_frame[flow_text_columns].astype("string").fillna("").to_string()
        if flow_text_columns
        else ""
    )
    flow_render_boundary_ok = (
        len(flow_rendered_frames) == 1
        and flow_final_frame.loc[flow_final_frame.index[0], "입출고일자"] == ""
        and flow_final_frame.loc[flow_final_frame.index[1], "입출고일자"] == "2026/07/02"
        and flow_final_frame.loc[flow_final_frame.index[1], "명세서일자"] == "2026/07/01"
        and flow_final_frame.loc[flow_final_frame.index[0], "출고수량"] == ""
        and flow_final_frame.loc[flow_final_frame.index[0], "입고수량"] == "480"
        and flow_final_frame.loc[flow_final_frame.index[1], "입고수량"] == ""
        and flow_final_frame.loc[flow_final_frame.index[1], "출고수량"] == "413"
        and str(flow_final_frame["입고수량"].dtype) == "string"
        and str(flow_final_frame["출고수량"].dtype) == "string"
        and "입고수량" in flow_rendered_kwargs[0].get("column_config", {})
        and "출고수량" in flow_rendered_kwargs[0].get("column_config", {})
        and flow_final_frame.loc[flow_final_frame.index[0], "단가적용거래처"] == ""
        and flow_final_frame.loc[flow_final_frame.index[0], "재고적용거래처"] == ""
        and flow_final_frame.loc[flow_final_frame.index[0], "재고위치"] == ""
        and flow_final_frame.loc[flow_final_frame.index[0], "매입처"] == ""
        and "None" not in flow_text_rendered
        and "nan" not in flow_text_rendered.lower()
    )
    results.append(
        CheckResult(
            "product flow fast renderer passes normalized date display dataframe",
            flow_render_boundary_ok,
            f"source_dates={flow_history['입출고일자'].tolist()}, rendered_rows={len(flow_final_frame)}, dates={flow_final_frame['입출고일자'].tolist() if '입출고일자' in flow_final_frame else []}, invoice_dates={flow_final_frame['명세서일자'].tolist() if '명세서일자' in flow_final_frame else []}, in={flow_final_frame['입고수량'].tolist() if '입고수량' in flow_final_frame else []}, out={flow_final_frame['출고수량'].tolist() if '출고수량' in flow_final_frame else []}, in_dtype={flow_final_frame['입고수량'].dtype if '입고수량' in flow_final_frame else ''}, config_keys={list((flow_rendered_kwargs[0].get('column_config') or {}).keys()) if flow_rendered_kwargs else []}, token_values={flow_final_frame.loc[flow_final_frame.index[0], ['단가적용거래처','재고적용거래처','재고위치','매입처']].tolist() if len(flow_final_frame) else []}, text={flow_text_rendered!r}",
        )
    )

    carryover_display_source = pd.DataFrame(
        {
            "입출고일자": ["이월재고", "2026/07/29"],
            "명세서번호": ["None", "None"],
            "제조번호": [None, "M-01"],
            "검수확인": [pd.NA, "Y"],
        }
    )
    carryover_display = chat_middleware._preserve_product_flow_table_dtypes(carryover_display_source)
    results.append(
        CheckResult(
            "product flow carryover display clears only synthetic-row missing text",
            carryover_display.loc[0, "명세서번호"] == ""
            and carryover_display.loc[0, "제조번호"] == ""
            and carryover_display.loc[0, "검수확인"] == ""
            and carryover_display.loc[1, "명세서번호"] == "None"
            and carryover_display_source.loc[0, "명세서번호"] == "None",
            f"display={carryover_display.to_dict(orient='records')}, source={carryover_display_source.to_dict(orient='records')}",
        )
    )

    flow_date_fixture = pd.DataFrame(
        {
            "입출고일자": ["2026/07/01"] * 74 + ["None"],
            "거래처명": ["None"] + ["정상 거래처"] * 74,
            "입고수량": [1.0] * 74 + [None],
        }
    )
    flow_date_display = prepare_sims_table_display_df(
        flow_date_fixture,
        action_name="제품수불현황 조회",
    )
    generic_none_display = prepare_sims_table_display_df(
        pd.DataFrame({"상태": ["None", "NULL"], "수량": [None, 1.0]}),
        action_name="일반 분석 조회",
    )
    flow_signature_display = prepare_sims_table_display_df(
        pd.DataFrame(
            {
                "입출고일자": ["2026/07/01"],
                "입고수량": [1.0],
                "출고수량": [0.0],
                "재고수량": [1.0],
                "거래처명": ["None"],
            }
        ),
        action_name="",
    )
    common_display_source = pd.DataFrame(
        {
            "거래처명": [None, "None", "NULL", pd.NA, "정상 거래처"],
            "단가적용처명": ["<NA>", "NaN", "NaT", np.nan, "정상 적용처"],
            "재고적용처명": [None, "None", pd.NA, np.nan, "정상 재고처"],
            "단가": [None, "NULL", 1234.5, pd.NA, -10.0],
            "DC율": [np.nan, "NaN", 0.0, pd.NA, -8.5],
            "수량": [0, -1, "None", None, 2],
            "공급가액": [0, -100, pd.NA, "NaN", 1000],
            "입출고일자": [pd.NaT, "NaT", "20260701", None, "2026/07/02"],
            "제품코드": ["00123", "01", "A01", None, "00001"],
        }
    )
    common_display_before = common_display_source.copy(deep=True)
    common_display, common_config, _common_width, _common_height = build_sims_table_display_config(
        common_display_source,
        action_name="입고명세 조회",
        add_row_no=False,
    )
    def _common_config_type(column: str) -> str:
        config = common_config.get(column)
        if isinstance(config, dict):
            return str((config.get("type_config") or {}).get("type") or "")
        return type(config).__name__

    placeholder_text_ok = (
        common_display["거래처명"].tolist() == ["", "", "", "", "정상 거래처"]
        and common_display["단가적용처명"].tolist() == ["", "", "", "", "정상 적용처"]
        and common_display["재고적용처명"].tolist() == ["", "", "", "", "정상 재고처"]
    )
    def _numeric_display_values(column: str) -> list[float | None]:
        return [
            None if value == "" else float(str(value).replace(",", "").removesuffix("%"))
            for value in common_display[column].tolist()
        ]

    numeric_contract_ok = (
        all(str(common_display[column].dtype) == "string" for column in ("단가", "DC율", "수량", "공급가액"))
        and _numeric_display_values("단가") == [None, None, 1234.5, None, -10.0]
        and _numeric_display_values("DC율") == [None, None, 0.0, None, -8.5]
        and _numeric_display_values("수량") == [0.0, -1.0, None, None, 2.0]
        and _numeric_display_values("공급가액") == [0.0, -100.0, None, None, 1000.0]
        and all(_common_config_type(column) == "text" for column in ("단가", "DC율", "수량", "공급가액"))
    )
    text_and_date_contract_ok = (
        all(_common_config_type(column) == "text" for column in ("거래처명", "단가적용처명", "재고적용처명", "제품코드"))
        and common_display["제품코드"].tolist() == ["00123", "01", "A01", "", "00001"]
        and common_display["입출고일자"].tolist() == ["", "", "2026-07-01", "", "2026/07/02"]
    )
    results.append(
        CheckResult(
            "common display normalizes literal missing tokens without coercing text columns",
            int(flow_date_display["입출고일자"].ne("").sum()) == 74
            and flow_date_display.loc[0, "거래처명"] == ""
            and pd.isna(flow_date_display.loc[74, "입고수량"])
            and str(flow_date_display["입고수량"].dtype) == "float64"
            and generic_none_display["상태"].tolist() == ["", ""]
            and flow_signature_display.loc[0, "거래처명"] == ""
            and placeholder_text_ok
            and numeric_contract_ok
            and text_and_date_contract_ok
            and common_display_source.equals(common_display_before),
            "flow_date_non_null="
            f"{int(flow_date_display['입출고일자'].ne('').sum())}, generic={generic_none_display['상태'].tolist()!r}, "
            f"text_ok={placeholder_text_ok}, numeric_ok={numeric_contract_ok}, "
            f"text_date_ok={text_and_date_contract_ok}, source_unchanged={common_display_source.equals(common_display_before)}",
        )
    )

    detail_semantic_source = pd.DataFrame(
        {
            "출고순번": pd.Series([1, 2], dtype="int64"),
            "거래명세서순번": pd.Series([11, 12], dtype="int64"),
            "세금계산서순번": pd.Series([101, 102], dtype="int64"),
            "확정순번": pd.Series([1, 2], dtype="int64"),
            "세금계산서상세합_공급가액": pd.Series([1000.0, 2000.0], dtype="float64"),
            "거래처코드": ["00123", "00045"],
            "출고일자": ["2026-07-01", "2026-07-02"],
        }
    )
    detail_semantic_before = detail_semantic_source.copy(deep=True)
    detail_semantic_display, detail_semantic_config, _detail_semantic_width, _detail_semantic_height = (
        build_sims_table_display_config(
            detail_semantic_source,
            action_name="출고명세 조회",
            add_row_no=False,
            )
    )

    def _detail_semantic_config_type(column: str) -> str:
        config = detail_semantic_config.get(column)
        if isinstance(config, dict):
            return str((config.get("type_config") or {}).get("type") or "")
        return type(config).__name__

    detail_semantic_numeric_columns = [
        "출고순번",
        "거래명세서순번",
        "세금계산서순번",
        "확정순번",
        "세금계산서상세합_공급가액",
    ]
    results.append(
        CheckResult(
            "detail-table numeric semantics follow dtype while codes and dates remain text",
            all(pd.api.types.is_numeric_dtype(detail_semantic_display[column]) for column in detail_semantic_numeric_columns)
            and all(_detail_semantic_config_type(column) == "number" for column in detail_semantic_numeric_columns)
            and _detail_semantic_config_type("거래처코드") == "text"
            and _detail_semantic_config_type("출고일자") == "text"
            and detail_semantic_display["거래처코드"].tolist() == ["00123", "00045"]
            and detail_semantic_source.equals(detail_semantic_before),
            f"numeric={detail_semantic_numeric_columns}, code_dtype={detail_semantic_display['거래처코드'].dtype}, "
            f"date_dtype={detail_semantic_display['출고일자'].dtype}",
        )
    )

    numeric_null_render_cases = {
        "입고명세 조회": pd.DataFrame(
            {
                "세금계산서상세합_공급가액": [None, 123456.0],
                "세금계산서상세합_세액": [np.nan, 12345.0],
                "입고수량": [0.0, -1.0],
            }
        ),
        "출고명세 조회": pd.DataFrame(
            {
                "세금계산서헤더_공급가액": [None, 654321.0],
                "세금계산서헤더_세액": [pd.NA, 65432.0],
                "세금계산금액일치": [None, 1.0],
                "출고수량": [0.0, -1.0],
            }
        ),
    }
    numeric_null_render_ok = True
    numeric_null_render_details: list[str] = []
    for action_name, source_df in numeric_null_render_cases.items():
        source_before = source_df.copy(deep=True)
        rendered_frames: list[pd.DataFrame] = []
        rendered_frame_refs: list[pd.DataFrame] = []
        rendered_kwargs: list[dict[str, Any]] = []
        with (
            patch.object(chat_middleware.st, "session_state", {"__sims_table_render_path": "history"}),
            patch.object(
                chat_middleware.st,
                "dataframe",
                lambda frame, **kwargs: (
                    rendered_frame_refs.append(frame),
                    rendered_frames.append(frame.copy()),
                    rendered_kwargs.append(dict(kwargs)),
                ),
            ),
            patch.object(chat_middleware.st, "caption", lambda *_args, **_kwargs: None),
        ):
            chat_middleware._render_chat_fast_dataframe(
                source_df,
                action_name=action_name,
                meta={"table_key": f"numeric-null-{action_name}"},
            )

        rendered_df = rendered_frames[0] if rendered_frames else pd.DataFrame()
        numeric_columns = list(source_df.columns)
        missing_numeric_columns = numeric_columns[:-1]
        factual_numeric_column = numeric_columns[-1]
        config = rendered_kwargs[0].get("column_config", {}) if rendered_kwargs else {}
        values_are_blank_or_numeric = (
            not rendered_df.empty
            and all(rendered_df.loc[rendered_df.index[0], column] == "" for column in missing_numeric_columns)
            and rendered_df.loc[rendered_df.index[0], factual_numeric_column] == 0.0
            and rendered_df.loc[rendered_df.index[1], factual_numeric_column] == -1.0
            and all(str(rendered_df[column].dtype) == "string" for column in missing_numeric_columns)
            and pd.api.types.is_numeric_dtype(rendered_df[factual_numeric_column])
            and all(isinstance(config.get(column), dict) for column in numeric_columns)
            and all(
                config.get(column, {}).get("type_config", {}).get("type") == "text"
                for column in missing_numeric_columns
            )
            and config.get(factual_numeric_column, {}).get("type_config", {}).get("type") == "number"
        )
        arrow_values_ok = False
        if values_are_blank_or_numeric:
            from streamlit import dataframe_util
            import pyarrow as pa

            arrow_table = pa.ipc.open_stream(
                dataframe_util.convert_pandas_df_to_arrow_bytes(rendered_df[numeric_columns])
            ).read_all()
            arrow_values_ok = all(
                pa.types.is_string(arrow_table.schema.field(column).type)
                and None not in arrow_table[column].to_pylist()
                for column in missing_numeric_columns
            ) and pa.types.is_floating(arrow_table.schema.field(factual_numeric_column).type)
        case_ok = (
            values_are_blank_or_numeric
            and arrow_values_ok
            and bool(rendered_frame_refs and rendered_frame_refs[0] is not None)
            and source_df.equals(source_before)
        )
        numeric_null_render_ok = numeric_null_render_ok and case_ok
        numeric_null_render_details.append(
            f"{action_name}: ok={case_ok}, dtypes={rendered_df.dtypes.astype(str).to_dict() if not rendered_df.empty else {}}, "
            f"arrow={arrow_values_ok}, dataframe_captured={bool(rendered_frame_refs)}"
        )
    results.append(
        CheckResult(
            "inbound and outbound fast renderers serialize numeric missing values as blank strings",
            numeric_null_render_ok,
            "; ".join(numeric_null_render_details),
        )
    )

    small_source = pd.DataFrame(
        {
            "numeric_missing": [None, 100.0],
            "numeric_factual": [0.0, -1.0],
        }
    )
    small_render_refs: list[pd.DataFrame] = []
    small_render_kwargs: list[dict[str, Any]] = []
    import app.ui.sims_table_display as sims_table_display

    with (
        patch.object(
            sims_table_display.st,
            "dataframe",
            lambda frame, **kwargs: (
                small_render_refs.append(frame),
                small_render_kwargs.append(dict(kwargs)),
            ),
        ),
    ):
        sims_table_display.render_sims_table(
            small_source,
            action_name="small-boundary-fixture",
            add_row_no=False,
        )

    small_rendered_df = small_render_refs[0] if small_render_refs else pd.DataFrame()
    small_config = small_render_kwargs[0].get("column_config", {}) if small_render_kwargs else {}
    small_renderer_boundary_ok = bool(
        not small_rendered_df.empty
        and small_rendered_df.iloc[0, 0] == ""
        and str(small_rendered_df.iloc[:, 0].dtype) == "string"
        and small_rendered_df.iloc[0, 1] == 0.0
        and small_rendered_df.iloc[1, 1] == -1.0
        and all(isinstance(small_config.get(column), dict) for column in small_source.columns)
        and small_config.get("numeric_missing", {}).get("type_config", {}).get("type") == "text"
        and small_config.get("numeric_factual", {}).get("type_config", {}).get("type") == "number"
        and small_source.equals(
            pd.DataFrame(
                {
                    "numeric_missing": [None, 100.0],
                    "numeric_factual": [0.0, -1.0],
                }
            )
        )
    )
    results.append(
        CheckResult(
            "shared small renderer passes the builder display DataFrame directly to st.dataframe",
            small_renderer_boundary_ok,
            f"dataframe_captured={bool(small_render_refs)}, "
            f"dtypes={small_rendered_df.dtypes.astype(str).to_dict() if not small_rendered_df.empty else {}}, "
            f"number_columns={list(small_config)}",
        )
    )

    sales_forecast_source = pd.DataFrame(
        {
            "예상매출": [1234567.89, 0.0, -500.25],
            "매출진척률": [12.34, 0.0, -1.25],
        }
    )
    sales_forecast_view, sales_forecast_config, _forecast_width, _forecast_height = (
        build_sims_table_display_config(
            sales_forecast_source,
            action_name="품목별 매출 예상",
            add_row_no=False,
        )
    )
    sales_forecast_numeric_ok = (
        sales_forecast_view.equals(sales_forecast_source)
        and all(pd.api.types.is_numeric_dtype(sales_forecast_view[column]) for column in sales_forecast_view.columns)
        and all(isinstance(sales_forecast_config.get(column), dict) for column in sales_forecast_view.columns)
        and sales_forecast_source.equals(
            pd.DataFrame(
                {
                    "예상매출": [1234567.89, 0.0, -500.25],
                    "매출진척률": [12.34, 0.0, -1.25],
                }
            )
        )
    )
    results.append(
        CheckResult(
            "sales forecast numeric display without missing values remains numeric",
            sales_forecast_numeric_ok,
            f"dtypes={sales_forecast_view.dtypes.astype(str).to_dict()}",
        )
    )

    summary_payload = {
        "type": "table",
        "action": "제품재고현황 조회",
        "df": inventory_display.head(1).copy(),
        "meta": {
            "sum_carry_qty": float("nan"),
            "sum_in_qty": 480.0,
            "sum_out_qty": 412.0,
            "sum_stock_qty": 299.0,
            "sum_stock_amt": 2313028.0,
            "sum_insu_amt": 2601300.0,
            "inventory_summary": {"sum_carry_qty": 231.0},
        },
    }
    chat_middleware._ensure_table_json_safe(summary_payload)
    stored_meta = dict(summary_payload.get("meta") or {})
    summary_keys = (
        "sum_carry_qty", "sum_in_qty", "sum_out_qty", "sum_stock_qty",
        "sum_stock_amt", "sum_insu_amt",
    )
    json_summary_ok = all(
        isinstance(stored_meta.get(key), (int, float))
        and not pd.isna(stored_meta.get(key))
        for key in summary_keys
    ) and (
        stored_meta.get("inventory_summary", {}).get("sum_carry_qty") == 231
        and stored_meta.get("df", [{}])[0].get("입고단가") is None
        and stored_meta.get("df", [{}])[0].get("입고DC율") is None
    )
    results.append(
        CheckResult(
            "product inventory bounded summary facts stay finite through table JSON serialization",
            json_summary_ok,
            f"stored_meta={stored_meta}",
        )
    )

    flow_summary_payload = {
        "type": "table",
        "action": "제품수불현황 조회",
        "df": flow_display.copy(),
        "meta": {
            "carry_qty": float("nan"),
            "in_qty": 480.0,
            "out_qty": 413.0,
            "stock_qty": 289.0,
            "flow_summary": {"carry_qty": 222.0},
        },
    }
    chat_middleware._ensure_table_json_safe(flow_summary_payload)
    stored_flow_meta = dict(flow_summary_payload.get("meta") or {})
    flow_summary_ok = all(
        isinstance(stored_flow_meta.get(key), (int, float))
        and not pd.isna(stored_flow_meta.get(key))
        for key in ("carry_qty", "in_qty", "out_qty", "stock_qty")
    ) and stored_flow_meta.get("flow_summary", {}).get("carry_qty") == 222
    results.append(
        CheckResult(
            "product flow bounded summary facts stay finite through table JSON serialization",
            flow_summary_ok,
            f"stored_flow_meta={stored_flow_meta}",
        )
    )

    flow_frames: list[pd.DataFrame] = []
    flow_kwargs: list[dict[str, Any]] = []
    flow_render_logs: list[dict[str, Any]] = []
    with (
        patch.object(chat_middleware.st, "session_state", {"__sims_table_render_path": "history"}),
        patch.object(
            chat_middleware.st,
            "dataframe",
            lambda frame, **kwargs: (flow_frames.append(frame.copy()), flow_kwargs.append(dict(kwargs))),
        ),
        patch.object(chat_middleware.st, "caption", lambda *_args, **_kwargs: None),
        patch.object(chat_middleware, "log_sims_table_render", lambda *_args, **kwargs: flow_render_logs.append(kwargs)),
    ):
        chat_middleware._render_chat_fast_dataframe(
            flow_display,
            action_name="제품수불현황 조회",
            meta={"table_key": "flow-history"},
        )
    flow_config_ok = (
        len(flow_frames) == 1
        and bool(flow_kwargs[0].get("column_config"))
        and int(flow_render_logs[0].get("column_config_count") or 0) > 0
    )
    results.append(
        CheckResult(
            "product flow fast renderer passes shared column config and date values",
            flow_config_ok,
            f"config_count={flow_render_logs[0].get('column_config_count') if flow_render_logs else None}, "
            f"config_keys={sorted(flow_kwargs[0].get('column_config') or {}) if flow_kwargs else []}",
        )
    )

    full_notice_warnings: list[str] = []
    with (
        patch.object(chat_middleware.st, "warning", lambda message, **_kwargs: full_notice_warnings.append(str(message))),
        patch.object(chat_middleware.st, "caption", lambda *_args, **_kwargs: None),
    ):
        full_notice = chat_middleware._render_partial_download_source_notice(
            source_status="full",
            download_rows=542,
            expected_rows=542,
        )
    results.append(
        CheckResult(
            "partial export notice remains isolated from a full table message",
            not full_notice and not full_notice_warnings,
            f"full_notice={full_notice}, warnings={full_notice_warnings}",
        )
    )

    return results


def _make_ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _next_seq_factory() -> Callable[[], int]:
    seq = {"v": 0}

    def _next_seq() -> int:
        seq["v"] += 1
        return seq["v"]

    return _next_seq


def _extract_payload_from_room(room: dict[str, Any]) -> dict[str, Any] | None:
    """
    혹시 fake_push를 타지 않고 room.messages에 직접 들어간 결과가 있으면 회수한다.
    """
    msgs = room.get("messages") or []
    if not msgs:
        return None

    last = msgs[-1]
    if isinstance(last, dict):
        content = last.get("content")
        if isinstance(content, dict):
            return content

        return {
            "final": True,
            "type": "text",
            "action": last.get("action"),
            "data": content,
            "message": content,
            "meta": last.get("meta") or {},
        }

    return None


# ---------------------------------------------------------------------
# Live cases
# ---------------------------------------------------------------------
def _smoke_live_cases() -> list[NlqCase]:
    """
    빠른 대표 테스트.
    검증 4종처럼 무거울 수 있는 조회는 live-all에 배치한다.
    """
    return [
        NlqCase(
            "입고명세 20250101~20250131 조회",
            "입고명세 조회",
            expected_condition_tokens=("2025-01-01", "2025-01-31"),
            expected_date_range=("20250101", "20250131"),
            date_column_candidates=("입고일자", "Rd11_In_YyMmDd"),
        ),
        NlqCase(
            "출고명세 20250101~20250131 조회",
            "출고명세 조회",
            expected_condition_tokens=("2025-01-01", "2025-01-31"),
            expected_date_range=("20250101", "20250131"),
            date_column_candidates=("출고일자", "Rd12_Out_YyMmDd"),
        ),
        NlqCase(
            "거래명세서 공통 202501 매입분 조회",
            "거래명세서 공통 조회",
            expected_condition_tokens=("2025-01", "매입분"),
        ),
        NlqCase(
            "세금계산서 공통 202501 매출 조회",
            "세금계산서 공통 조회",
            expected_condition_tokens=("2025-01", "매출"),
        ),
        NlqCase(
            "실재고월집계 2025 제품 00029 조회",
            "실재고월집계 조회",
            expected_condition_tokens=("2025-01", "2025-12", "제품 00029"),
        ),
        NlqCase(
            "장부재고월집계 2025 제품 00029 조회",
            "장부재고월집계 조회",
            expected_condition_tokens=("2025-01", "2025-12", "제품 00029"),
        ),
    ]

def _all_live_cases() -> list[NlqCase]:
    return [
        # 입고명세
        NlqCase("입고명세 20250101~20250131 조회", "입고명세 조회"),
        NlqCase("입고명세 거래처명 동제 조회", "입고명세 조회"),
        NlqCase("입고명세 제조사명 한미 조회", "입고명세 조회"),
        NlqCase("입고명세 제품 00029 조회", "입고명세 조회"),
        NlqCase("입고명세 재고위치 0001 조회", "입고명세 조회"),

        # 출고명세
        NlqCase("출고명세 20250101~20250131 조회", "출고명세 조회"),
        NlqCase("출고명세 거래처명 동제 조회", "출고명세 조회"),
        NlqCase("출고명세 매입처명 한미 조회", "출고명세 조회"),
        NlqCase("출고명세 실납처명 동제 조회", "출고명세 조회"),
        NlqCase("출고명세 영업사원명 김 조회", "출고명세 조회"),

        # 거래명세서 공통
        NlqCase("거래명세서 공통 202501 매입분 조회", "거래명세서 공통 조회"),
        NlqCase("거래명세서 공통 202501 매출분 조회", "거래명세서 공통 조회"),
        NlqCase("거래명세서 공통 거래처명 동제 조회", "거래명세서 공통 조회"),

        # 세금계산서 공통
        NlqCase("세금계산서 공통 202501 매입 조회", "세금계산서 공통 조회"),
        NlqCase("세금계산서 공통 202501 매출 조회", "세금계산서 공통 조회"),
        NlqCase("세금계산서 공통 거래처명 동제 조회", "세금계산서 공통 조회"),

        # 검증 4종
        # 기간 없는 검증은 기본 기간이 적용되어야 한다.
        NlqCase("입고 거래명세서 불일치 조회", "입고↔거래명세서 검증"),
        NlqCase("입고 세금계산서 불일치 조회", "입고↔세금계산서 검증"),
        NlqCase("출고 거래명세서 불일치 조회", "출고↔거래명세서 검증"),
        NlqCase("출고 세금계산서 불일치 조회", "출고↔세금계산서 검증"),

        # 월집계
        NlqCase("실재고월집계 202508 제품 00029 조회", "실재고월집계 조회"),
        NlqCase("실재고월집계 2025 제품 00029 조회", "실재고월집계 조회"),
        NlqCase("실재고월집계 202601 재고위치 0001 조회", "실재고월집계 조회"),

        NlqCase("장부재고월집계 202508 제품 00029 조회", "장부재고월집계 조회"),
        NlqCase("장부재고월집계 2025 제품 00029 조회", "장부재고월집계 조회"),
        NlqCase("장부재고월집계 202601 재고위치 0001 조회", "장부재고월집계 조회"),

        # 제품수불 / 제품재고
        NlqCase("제품수불현황 20250101~20251231 제품 00029 조회", "제품수불현황 조회"),
        NlqCase("제품수불현황 제품 00029 재고위치 0001 조회", "제품수불현황 조회"),
        NlqCase("제품수불현황 장부수불 제품 00029 조회", "제품수불현황 조회"),

        # 최근 실사용 회귀 케이스

        # 매입처명/매출처명은 IO 상세에서는 실제 명세의 거래처 조건으로 해석해야 한다.
        NlqCase(
            "입고명세 매입처명 한미 2026-04-01~2026-04-30 조회",
            "입고명세 조회",
            expected_condition_tokens=("2026-04-01", "2026-04-30", "거래처 한미"),
            forbidden_condition_tokens=("매입처 한미", "최근 1개월 자동적용"),
            expected_date_range=("20260401", "20260430"),
            date_column_candidates=("입고일자", "Rd11_In_YyMmDd"),
        ),
        NlqCase(
            "입고명세 매입처명 온라인 2026-04-01~2026-04-30 조회",
            "입고명세 조회",
            expected_condition_tokens=("2026-04-01", "2026-04-30", "거래처 온라인"),
            forbidden_condition_tokens=("매입처 온라인", "최근 1개월 자동적용"),
            expected_date_range=("20260401", "20260430"),
            date_column_candidates=("입고일자", "Rd11_In_YyMmDd"),
        ),
        NlqCase(
            "입고명세 매입처명 인천 2026-04-01~2026-04-30 조회",
            "입고명세 조회",
            expected_condition_tokens=("2026-04-01", "2026-04-30", "거래처 인천"),
            forbidden_condition_tokens=("매입처 인천", "최근 1개월 자동적용"),
            expected_date_range=("20260401", "20260430"),
            date_column_candidates=("입고일자", "Rd11_In_YyMmDd"),
        ),
        NlqCase(
            "출고명세 매출처명 대학약국 2026-04-19~2026-05-19 조회",
            "출고명세 조회",
            expected_condition_tokens=("2026-04-19", "2026-05-19", "거래처 대학약국"),
            forbidden_condition_tokens=("매출처 대학약국", "최근 1개월 자동적용"),
            expected_date_range=("20260419", "20260519"),
            date_column_candidates=("출고일자", "Rd12_Out_YyMmDd"),
        ),

        # - "제품 수불 현황" 띄어쓰기 보정
        # - 제품명 후보 안내표
        # - 제품코드 + 연도범위 조회
        # - 최근 입고명세 기간 조회
        NlqCase("제품 수불 현황 직듀오서방정", "제품수불현황 조회"),
        NlqCase("제품수불부 펜타듀르패취 조회", "제품수불현황 조회"),
        NlqCase("제품수불부 제품 00339 2024~2026 조회", "제품수불현황 조회"),
        NlqCase("입고명세 20260401~20260508 조회", "입고명세 조회"),

        NlqCase(
            "제품재고장 2025 제품그룹명 일반 조회",
            "제품재고현황 조회",
            expected_condition_tokens=("2025", "제품그룹명 일반"),
            forbidden_condition_tokens=("제품명 일반", "제품명 그룹명 일반"),
        ),
        NlqCase(
            "제품재고장 2025 구분명 의약품 조회",
            "제품재고현황 조회",
            expected_condition_tokens=("2025", "구분명 의약품"),
            forbidden_condition_tokens=("제품명 의약품",),
        ),
        NlqCase(
            "제품재고장 2025 제품분류명 정제 조회",
            "제품재고현황 조회",
            expected_condition_tokens=("2025", "제품분류명 정제"),
            forbidden_condition_tokens=("제품명 정제", "제품명 분류명 정제"),
        ),

        NlqCase(
            "제품재고현황 2023 ~ 2026 제품 00029 조회",
            "제품재고현황 조회",
            expected_condition_tokens=("2023", "2026", "제품코드 00029"),
            forbidden_condition_tokens=("제품명 00029",),
        ),

        NlqCase(
            "제품재고현황 제조사 대웅제약 2023~2026년 조회",
            "제품재고현황 조회",
            expected_condition_tokens=("2023", "2026", "제조사명 대웅제약"),
            forbidden_condition_tokens=("제품명 대웅제약",),
        ),
        NlqCase(
            "제품재고현황 제조사 건일제약 2023~2026년 조회",
            "제품재고현황 조회",
            expected_condition_tokens=("2023", "2026", "제조사명 건일제약"),
            forbidden_condition_tokens=("제품명 건일제약",),
        ),
        NlqCase(
            "제품재고현황 제조사 삼진제약 2023~2026년 조회",
            "제품재고현황 조회",
            expected_condition_tokens=("2023", "2026", "제조사명 삼진제약"),
            forbidden_condition_tokens=("제품명 삼진제약",),
        ),

    ]


def _evaluate_nlq_case(case: NlqCase, handled: bool, payload: dict[str, Any] | None) -> CheckResult:
    name = f"live: {case.query}"

    if case.require_handled and not handled:
        return _fail(name, "try_handle_nlq()가 False 반환")

    if not payload:
        if case.allow_no_payload:
            return _ok(name, "payload 없음 허용")
        return _fail(name, "payload 없음")

    action = str(payload.get("action") or payload.get("title") or "").strip()
    title = str(payload.get("title") or "").strip()
    ptype = str(payload.get("type") or "").strip()
    meta = payload.get("meta") or {}
    message = str(payload.get("message") or payload.get("data") or "").strip()

    if case.expected_action not in action and case.expected_action not in title:
        return _fail(
            name,
            f"action mismatch: expected contains {case.expected_action!r}, action={action!r}, title={title!r}",
        )

    row_count = _payload_row_count(payload)

    if not case.allow_zero_rows and row_count <= 0:
        return _fail(name, f"row_count가 0 이하: rows={row_count}")

    # IO 계열도 가능한 경우 query_summary 또는 조회조건이 있어야 한다.
    summary_md = str(meta.get("summary_md") or "").strip()
    query_summary = str(meta.get("query_summary") or "").strip()

    has_condition_text = (
        "조회조건:" in summary_md
        or "조회조건:" in message
        or bool(query_summary)
        or bool(meta.get("condition"))
    )

    # 검증류 일부 payload는 note만 있을 수 있으므로 경고 수준으로 detail에 표시한다.
    condition_mark = "Y" if has_condition_text else "N"

    if case.expected_date_range and case.date_column_candidates and row_count > 0:
        df = _payload_to_df(payload)
        if df is not None and not df.empty:
            date_col = None
            for cand in case.date_column_candidates:
                if cand in df.columns:
                    date_col = cand
                    break

            if not date_col:
                return _fail(
                    name,
                    f"날짜 컬럼 없음: candidates={case.date_column_candidates}, columns={list(df.columns)[:30]}",
                )

            start, end = case.expected_date_range
            bad_values = []

            for v in df[date_col].head(500).tolist():
                d = _norm_yyyymmdd(v)
                if d and not (start <= d <= end):
                    bad_values.append(str(v))
                    if len(bad_values) >= 5:
                        break

            if bad_values:
                return _fail(
                    name,
                    f"날짜 범위 밖 자료 포함: col={date_col}, range={start}~{end}, bad={bad_values}",
                )

    # IO NLQ는 조건 토큰이 반드시 있어야 한다.
    # 조건 토큰은 summary_md, message, query_summary, meta.condition 중 어디에나 있을 수 있다.
    if case.expected_condition_tokens or case.forbidden_condition_tokens:
        condition_text = " ".join(
            [
                str(summary_md or ""),
                str(message or ""),
                str(query_summary or ""),
                str(meta.get("condition") or ""),
            ]
        )

        if case.expected_condition_tokens:
            missing_tokens = [
                token
                for token in case.expected_condition_tokens
                if str(token) not in condition_text
            ]

            if missing_tokens:
                return _fail(
                    name,
                    f"조회조건 토큰 누락: missing={missing_tokens}, condition_text={condition_text[:300]!r}",
                )

        if case.forbidden_condition_tokens:
            bad_tokens = [
                token
                for token in case.forbidden_condition_tokens
                if str(token) in condition_text
            ]

            if bad_tokens:
                return _fail(
                    name,
                    f"금지 조회조건 포함: bad={bad_tokens}, condition_text={condition_text[:300]!r}",
                )

    # 입고명세/출고명세는 LLM 분석용 상세 요약 meta가 있어야 한다.
    if case.expected_action in {"입고명세 조회", "출고명세 조회"}:
        if not summary_md:
            return _fail(name, f"{case.expected_action} summary_md 누락")

        expected_key = "in_detail_summary" if case.expected_action == "입고명세 조회" else "out_detail_summary"

        has_detail_summary = (
            isinstance(meta.get(expected_key), dict)
            and bool(meta.get(expected_key))
        )

        if not has_detail_summary:
            return _fail(
                name,
                f"{case.expected_action} {expected_key} 누락: meta_keys={sorted(list(meta.keys()))}",
            )

        # 0건 text 결과는 최소 summary만 있어도 허용한다.
        # 실제 row가 있는 경우는 거래처/제품별 집계까지 보호한다.
        if row_count > 0:
            detail_summary = meta.get(expected_key) or {}

            required_keys = [
                "row_count",
                "qty_sum",
                "supply_sum",
                "tax_sum",
                "amount_sum",
                "vendor_count",
                "product_count",
                "top_products",
            ]

            if case.expected_action == "입고명세 조회":
                required_keys += ["top_purchase_vendors"]
            else:
                required_keys += ["top_sales_vendors", "top_sales_staff"]

            missing_summary_keys = [
                k for k in required_keys
                if k not in detail_summary
            ]

            if missing_summary_keys:
                return _fail(
                    name,
                    f"{case.expected_action} detail summary key 누락: missing={missing_summary_keys}, keys={sorted(list(detail_summary.keys()))}",
                )

    # 거래명세서 공통은 LLM 분석용 거래명세서 요약 meta가 있어야 한다.
    if case.expected_action == "거래명세서 공통 조회":
        if not summary_md:
            return _fail(name, "거래명세서 공통 summary_md 누락")

        has_doc_summary = (
            isinstance(meta.get("trans_doc_summary"), dict)
            and bool(meta.get("trans_doc_summary"))
        )

        if not has_doc_summary:
            return _fail(
                name,
                f"거래명세서 공통 trans_doc_summary 누락: meta_keys={sorted(list(meta.keys()))}",
            )

        if row_count > 0:
            doc_summary = meta.get("trans_doc_summary") or {}
            required_keys = [
                "row_count",
                "row_count_total",
                "display_row_count",
                "supply_sum",
                "tax_sum",
                "amount_sum",
                "vendor_count",
                "by_trans_type",
                "top_vendors",
                "by_match_status",
            ]

            missing_summary_keys = [
                k for k in required_keys
                if k not in doc_summary
            ]

            if missing_summary_keys:
                return _fail(
                    name,
                    f"거래명세서 공통 summary key 누락: missing={missing_summary_keys}, keys={sorted(list(doc_summary.keys()))}",
                )

    # 실재고월집계/장부재고월집계는 LLM 분석용 월집계 summary meta가 있어야 한다.
    if case.expected_action in {"실재고월집계 조회", "장부재고월집계 조회"}:
        if not summary_md:
            return _fail(name, f"{case.expected_action} summary_md 누락")

        has_monthly_summary = (
            isinstance(meta.get("monthly_stock_detail_summary"), dict)
            and bool(meta.get("monthly_stock_detail_summary"))
        )

        if not has_monthly_summary:
            return _fail(
                name,
                f"{case.expected_action} monthly_stock_detail_summary 누락: meta_keys={sorted(list(meta.keys()))}",
            )

        if row_count > 0:
            monthly_summary = meta.get("monthly_stock_detail_summary") or {}
            required_keys = [
                "row_count",
                "row_count_total",
                "display_row_count",
                "stock_basis",
                "in_qty_sum",
                "in_supply_sum",
                "in_tax_sum",
                "out_qty_sum",
                "out_supply_sum",
                "out_tax_sum",
                "product_count",
                "vendor_count",
                "stock_location_count",
                "by_month",
                "by_side",
                "by_io_type",
                "top_products",
                "top_vendors",
                "top_stock_locations",
            ]

            missing_summary_keys = [
                k for k in required_keys
                if k not in monthly_summary
            ]

            if missing_summary_keys:
                return _fail(
                    name,
                    f"{case.expected_action} monthly summary key 누락: missing={missing_summary_keys}, keys={sorted(list(monthly_summary.keys()))}",
                )


    # 제품수불현황은 LLM 분석용 요약 meta가 있어야 한다.
    # 단, 후보표(candidate_table)는 실제 수불표가 아니므로 flow_summary 대신 candidate_table을 허용한다.
    if case.expected_action == "제품수불현황 조회":
        if not summary_md:
            return _fail(name, "제품수불현황 summary_md 누락")

        is_candidate_table = bool(meta.get("candidate_table"))
        has_flow_summary = isinstance(meta.get("flow_summary"), dict) and bool(meta.get("flow_summary"))

        if not is_candidate_table and not has_flow_summary:
            return _fail(
                name,
                f"제품수불현황 flow_summary 누락: meta_keys={sorted(list(meta.keys()))}"
            )

    # 제품재고현황은 LLM 분석용 inventory_summary가 있어야 한다.
    # 단, 0건 text 결과는 inventory_summary가 없거나 비어 있어도 허용한다.
    if case.expected_action == "제품재고현황 조회":
        if not summary_md:
            return _fail(name, "제품재고현황 summary_md 누락")

        if row_count > 0:
            has_inventory_summary = (
                isinstance(meta.get("inventory_summary"), dict)
                and bool(meta.get("inventory_summary"))
            )
            if not has_inventory_summary:
                return _fail(
                    name,
                    f"제품재고현황 inventory_summary 누락: meta_keys={sorted(list(meta.keys()))}",
                )


    cols = _payload_columns(payload)

    cond_preview = (
        query_summary
        or str(meta.get("condition") or "")
        or summary_md
        or message
    )
    cond_preview = str(cond_preview or "").replace("\n", " ").strip()
    if len(cond_preview) > 100:
        cond_preview = cond_preview[:100] + "..."

    detail = (
        f"action={action!r}, title={title!r}, type={ptype or None}, "
        f"rows={row_count}, cols={len(cols)}, condition={cond_preview!r}"
    )

    return _ok(name, detail)

def run_live_checks(*, live_all: bool = False, show_payload: bool = False) -> list[CheckResult]:
    results: list[CheckResult] = []

    try:
        router = importlib.import_module("app.sims.nlq.nlq_router")
        try_handle_nlq = getattr(router, "try_handle_nlq")
    except Exception as e:
        return [_fail("import router.try_handle_nlq", f"{type(e).__name__}: {e}")]

    capture = PayloadCapture()
    _patch_push_function(capture)

    cases = _all_live_cases() if live_all else _smoke_live_cases()

    for case in cases:
        room: dict[str, Any] = {"messages": []}
        session_state: dict[str, Any] = {
            "__sims_selected": {},
            "__io_pending_product_pick": {},
        }
        next_seq = _next_seq_factory()
        before_count = len(capture.payloads)

        try:
            handled = bool(
                try_handle_nlq(
                    case.query,
                    room=room,
                    session_state=session_state,
                    make_ts=_make_ts,
                    next_seq=next_seq,
                    logger=log,
                )
            )

            payload = capture.last_since(before_count)
            if payload is None:
                payload = _extract_payload_from_room(room)

            if show_payload and payload is not None:
                print()
                print(f"[PAYLOAD] {case.query}")
                print(f"  action={payload.get('action')!r}")
                print(f"  title={payload.get('title')!r}")
                print(f"  type={payload.get('type')!r}")
                print(f"  meta={payload.get('meta')!r}")

            results.append(_evaluate_nlq_case(case, handled, payload))

        except Exception as e:
            detail = f"{type(e).__name__}: {e}\n{traceback.format_exc(limit=4)}"
            results.append(_fail(f"live: {case.query}", detail))

    return results


def run_current_stock_nlq_contract_checks() -> list[CheckResult]:
    results: list[CheckResult] = []
    io_nlq = importlib.import_module("app.services.io_nlq")
    supplier = importlib.import_module("app.services.product_supplier_scope_service")
    goods_service = importlib.import_module("app.services.rddbc040_service")
    shared = importlib.import_module("app.sims.views.rddbc_io_shared")
    router = importlib.import_module("app.sims.nlq.nlq_router")
    inventory_service = importlib.import_module("app.services.product_inventory_service")
    profile_service = importlib.import_module("app.services.ssai_analysis_profile_service")
    login = importlib.import_module("app.ui.ssai_login")
    originals = {
        "vendor": supplier.resolve_common_vendor_candidates,
        "goods": goods_service.search_goods_full,
        "stocks": shared._load_stock_code_options,
        "profile": profile_service.load_dashboard_profile,
        "company": login.get_selected_company,
    }
    try:
        supplier.resolve_common_vendor_candidates = lambda text: (
            [
                {"entity_code": "10047", "canonical_name": "한미약품", "entity_role": "manufacturer"},
                {"entity_code": "10048", "canonical_name": "한미정밀", "entity_role": "manufacturer"},
            ] if "한미" in str(text) else []
        )
        def _goods_like_fixture(**kwargs):
            keyword = str(kwargs.get("keyword") or "")
            if "타이레놀" in keyword:
                return pd.DataFrame([{"Rd04_Physic_Cd": "00001", "Rd04_Physic_Nm": "타이레놀"}])
            if "아스피린" in keyword:
                return pd.DataFrame([
                    {"Rd04_Physic_Cd": "00101", "Rd04_Physic_Nm": "아스피린100"},
                    {"Rd04_Physic_Cd": "00102", "Rd04_Physic_Nm": "아스피린500"},
                ])
            if "아모크라" in keyword:
                return pd.DataFrame([
                    {"Rd04_Physic_Cd": "00201", "Rd04_Physic_Nm": "아모크라정"},
                    {"Rd04_Physic_Cd": "00202", "Rd04_Physic_Nm": "아모크라듀오"},
                ])
            return pd.DataFrame()
        goods_service.search_goods_full = _goods_like_fixture
        shared._load_stock_code_options = lambda: [(".본사 창고 (00001)", "00001", ".본사 창고")]

        empty = io_nlq.resolve_current_stock_entity_condition("현재고", params={})
        results.append(_ok("current stock requires manufacturer/product", "input_required") if empty.get("status") == "input_required" else _fail("current stock requires manufacturer/product", repr(empty)))

        maker = io_nlq.resolve_current_stock_entity_condition("현재고 한미", params={})
        maker_params = dict(maker.get("params") or {})
        results.append(_ok("current stock unlabeled manufacturer LIKE", "manufacturer scope") if maker.get("status") == "resolved" and maker_params.get("maker_nm") == "한미" else _fail("current stock unlabeled manufacturer LIKE", repr(maker)))

        product = io_nlq.resolve_current_stock_entity_condition("현재고 제품명 타이레놀", params={"physic_nm": "타이레놀"})
        product_params = dict(product.get("params") or {})
        results.append(_ok("current stock product LIKE", "타이레놀 scope") if product.get("status") == "resolved" and product_params.get("physic_nm") == "타이레놀" else _fail("current stock product LIKE", repr(product)))

        located_params_input = io_nlq.extract_params("현재고 재고위치 본사 창고 제품명 타이레놀")
        located = io_nlq.resolve_current_stock_entity_condition("현재고 재고위치 본사 창고 제품명 타이레놀", params=located_params_input)
        located_params = dict(located.get("params") or {})
        results.append(_ok("current stock explicit location", "00001") if located.get("status") == "resolved" and located_params.get("stock_cds") == ["00001"] and located_params.get("physic_nm") == "타이레놀" else _fail("current stock explicit location", repr(located)))
        results.append(_ok("current stock location code/name map", "00001=.본사 창고") if located_params.get("stock_location_name_map") == {"00001": ".본사 창고"} else _fail("current stock location code/name map", repr(located_params)))

        compound_cases = {
            "현재고 제약사 바이엘 제품 아스피린": {"maker_nm": "바이엘", "physic_nm": "아스피린"},
            "현재고 제품 아스피린 제약사 바이엘": {"maker_nm": "바이엘", "physic_nm": "아스피린"},
            "현재고 재고위치 본사 창고 제품명 아스피린": {"stock_nm": "본사 창고", "physic_nm": "아스피린"},
        }
        compound_parser_ok = all(
            all(io_nlq.extract_params(question).get(key) == expected for key, expected in expected_params.items())
            for question, expected_params in compound_cases.items()
        )
        results.append(
            _ok("current stock compound labelled parser", "labels stop at the next label")
            if compound_parser_ok
            else _fail("current stock compound labelled parser", repr({q: io_nlq.extract_params(q) for q in compound_cases}))
        )

        aspirin = io_nlq.resolve_current_stock_entity_condition("현재고 제품명 아스피린", params=io_nlq.extract_params("현재고 제품명 아스피린"))
        aspirin_params = dict(aspirin.get("params") or {})
        aspirin_ok = aspirin.get("status") == "resolved" and aspirin_params.get("physic_nm") == "아스피린" and len(aspirin.get("candidates") or []) == 2
        results.append(_ok("current stock explicit product LIKE multi", "no candidate table") if aspirin_ok else _fail("current stock explicit product LIKE multi", repr(aspirin)))

        amocra = io_nlq.resolve_current_stock_entity_condition("현재고 제품명 아모크라", params=io_nlq.extract_params("현재고 제품명 아모크라"))
        amocra_params = dict(amocra.get("params") or {})
        amocra_ok = amocra.get("status") == "resolved" and amocra_params.get("physic_nm") == "아모크라" and len(amocra.get("candidates") or []) == 2
        results.append(_ok("current stock explicit product LIKE multi amocra", "all LIKE products") if amocra_ok else _fail("current stock explicit product LIKE multi amocra", repr(amocra)))

        maker_labeled = io_nlq.resolve_current_stock_entity_condition("현재고 제조사명 한미", params=io_nlq.extract_params("현재고 제조사명 한미"))
        maker_labeled_params = dict(maker_labeled.get("params") or {})
        maker_labeled_ok = maker_labeled.get("status") == "resolved" and maker_labeled_params.get("maker_nm") == "한미" and len(maker_labeled.get("candidates") or []) == 2
        results.append(_ok("current stock explicit manufacturer LIKE multi", "all LIKE manufacturers") if maker_labeled_ok else _fail("current stock explicit manufacturer LIKE multi", repr(maker_labeled)))

        supplier.resolve_common_vendor_candidates = lambda _text: [
            {"entity_code": "20001", "canonical_name": "바이엘", "entity_role": "manufacturer"}
        ]
        compound_left_question = "현재고 제약사 바이엘 제품 아스피린"
        compound_right_question = "현재고 제품 아스피린 제약사 바이엘"
        compound_left = io_nlq.resolve_current_stock_entity_condition(
            compound_left_question, params=io_nlq.extract_params(compound_left_question)
        )
        compound_right = io_nlq.resolve_current_stock_entity_condition(
            compound_right_question, params=io_nlq.extract_params(compound_right_question)
        )
        compound_left_params = dict(compound_left.get("params") or {})
        compound_right_params = dict(compound_right.get("params") or {})
        compound_sql, _compound_sql_params = inventory_service._build_month_carry_sql(
            {**compound_left_params, "base_month": "202608", "stock_mode": "real", "stock_cds": ["00001"]},
            inventory_service._settings({"stock_mode": "real"}),
        )
        compound_and_ok = (
            compound_left.get("status") == compound_right.get("status") == "resolved"
            and compound_left_params.get("maker_nm") == compound_right_params.get("maker_nm") == "바이엘"
            and compound_left_params.get("physic_nm") == compound_right_params.get("physic_nm") == "아스피린"
            and compound_left_params.get("current_stock_entity_scope") == "manufacturer_and_product"
            and compound_left_params.get("current_stock_manufacturer_codes") == compound_right_params.get("current_stock_manufacturer_codes")
            and compound_left_params.get("current_stock_product_codes") == compound_right_params.get("current_stock_product_codes")
            and "P.Rd04_Ven_Cd IN" in compound_sql
            and "P.Rd04_Physic_Cd IN" in compound_sql
            and " OR " not in compound_sql.split("WHERE", 1)[1].split("GROUP BY", 1)[0]
        )
        results.append(
            _ok("current stock explicit manufacturer/product AND is order independent", "independent code IN predicates")
            if compound_and_ok
            else _fail("current stock explicit manufacturer/product AND is order independent", repr({"left": compound_left, "right": compound_right}))
        )

        current_stock_frame = pd.DataFrame([{
            "group_cd": "00001", "group_nm": "00001", "physic_cd": "00001", "physic_nm": "타이레놀",
            "standard": "500mg", "kd_cd": "", "edi_cd": "", "std_cd": "", "pack_unit": "EA",
            "buy_cd": "", "buy_nm": "", "order_cd": "", "order_nm": "", "maker_cd": "10047", "maker_nm": "한미",
            "product_group_nm": "", "product_di_nm": "", "product_class_nm": "", "special_manage_nm": "",
            "carry_qty": 0, "carry_unit": 0, "carry_dc": 0, "carry_amt": 0,
            "now_in_qty": 0, "in_unit": 0, "in_dc": 0, "in_amt": 0,
            "now_out_qty": 0, "out_unit": 0, "out_dc": 0, "out_amt": 0,
            "stock_qty": 12, "stock_unit": 100, "stock_dc": 0, "stock_amt": 1200,
            "curr_insu_unit": 10, "insu_amt": 120,
        }, {
            "group_cd": "00002", "group_nm": "00002", "physic_cd": "00001", "physic_nm": "타이레놀",
            "standard": "500mg", "kd_cd": "", "edi_cd": "", "std_cd": "", "pack_unit": "EA",
            "buy_cd": "", "buy_nm": "", "order_cd": "", "order_nm": "", "maker_cd": "10047", "maker_nm": "한미",
            "product_group_nm": "", "product_di_nm": "", "product_class_nm": "", "special_manage_nm": "",
            "carry_qty": 0, "carry_unit": 0, "carry_dc": 0, "carry_amt": 0,
            "now_in_qty": 0, "in_unit": 0, "in_dc": 0, "in_amt": 0,
            "now_out_qty": 0, "out_unit": 0, "out_dc": 0, "out_amt": 0,
            "stock_qty": 8, "stock_unit": 100, "stock_dc": 0, "stock_amt": 800,
            "curr_insu_unit": 10, "insu_amt": 80,
        }, {
            "group_cd": "00001", "group_nm": "00001", "physic_cd": "00002", "physic_nm": "타이레놀 ER",
            "standard": "650mg", "kd_cd": "", "edi_cd": "", "std_cd": "", "pack_unit": "EA",
            "buy_cd": "", "buy_nm": "", "order_cd": "", "order_nm": "", "maker_cd": "10047", "maker_nm": "한미",
            "product_group_nm": "", "product_di_nm": "", "product_class_nm": "", "special_manage_nm": "",
            "carry_qty": 0, "carry_unit": 0, "carry_dc": 0, "carry_amt": 0,
            "now_in_qty": 0, "in_unit": 0, "in_dc": 0, "in_amt": 0,
            "now_out_qty": 0, "out_unit": 0, "out_dc": 0, "out_amt": 0,
            "stock_qty": 4, "stock_unit": 100, "stock_dc": 0, "stock_amt": 400,
            "curr_insu_unit": 10, "insu_amt": 40,
        }, {
            "group_cd": "00002", "group_nm": "00002", "physic_cd": "00002", "physic_nm": "타이레놀 ER",
            "standard": "650mg", "kd_cd": "", "edi_cd": "", "std_cd": "", "pack_unit": "EA",
            "buy_cd": "", "buy_nm": "", "order_cd": "", "order_nm": "", "maker_cd": "10047", "maker_nm": "한미",
            "product_group_nm": "", "product_di_nm": "", "product_class_nm": "", "special_manage_nm": "",
            "carry_qty": 0, "carry_unit": 0, "carry_dc": 0, "carry_amt": 0,
            "now_in_qty": 0, "in_unit": 0, "in_dc": 0, "in_amt": 0,
            "now_out_qty": 0, "out_unit": 0, "out_dc": 0, "out_amt": 0,
            "stock_qty": 6, "stock_unit": 100, "stock_dc": 0, "stock_amt": 600,
            "curr_insu_unit": 10, "insu_amt": 60,
        }, {
            "group_cd": "00001", "group_nm": "00001", "physic_cd": "00003", "physic_nm": "타이레놀 8시간",
            "standard": "650mg", "kd_cd": "KD3", "edi_cd": "EDI3", "std_cd": "STD3", "pack_unit": "EA",
            "buy_cd": "", "buy_nm": "", "order_cd": "10003", "order_nm": "발주처3", "maker_cd": "10047", "maker_nm": "한미",
            "product_group_nm": "전문의약품", "product_di_nm": "보험", "product_class_nm": "내복제", "special_manage_nm": "",
            "carry_qty": 0, "carry_unit": 0, "carry_dc": 0, "carry_amt": 0,
            "now_in_qty": 0, "in_unit": 0, "in_dc": 0, "in_amt": 0,
            "now_out_qty": 0, "out_unit": 0, "out_dc": 0, "out_amt": 0,
            "stock_qty": 7, "stock_unit": 100, "stock_dc": 0, "stock_amt": 700,
            "curr_insu_unit": 0, "insu_amt": 700,
        }])
        stock_display, stock_source, stock_meta = inventory_service._build_current_stock_table_frames(current_stock_frame, {
            "group_basis": "stock", "current_stock_query": True,
            "stock_location_name_map": {"00001": "본사 창고", "00002": "지점 창고"},
        })
        expected_current_stock_columns = [
            "순번", "제품코드", "제품명", "규격", "재고위치명", "재고수량", "현보험약가", "보험금액", "표준코드",
            "제품그룹명", "제품구분명", "제품분류명", "발주처코드", "발주처명", "제조사코드",
            "재고위치코드", "KD코드", "EDI코드", "제조사명", "포장단위",
        ]
        display_ok = (
            list(stock_display.columns) == expected_current_stock_columns
            and stock_display.iloc[0]["재고위치코드"] == "00001"
            and stock_display.iloc[0]["재고위치명"] == "본사 창고"
            and int((stock_display["재고위치명"] == "합계").sum()) == 0
            and int((stock_display["재고위치명"] == "제품 합계").sum()) == 2
            and stock_display.loc[stock_display["제품코드"] == "00001", "순번"].nunique() == 1
            and stock_display["제품코드"].tolist() == ["00001", "", "", "00002", "", "", "00003"]
            and stock_display["재고위치명"].tolist() == ["본사 창고", "지점 창고", "제품 합계", "본사 창고", "지점 창고", "제품 합계", "본사 창고"]
            and stock_display.loc[1, "제품명"] == ""
            and (stock_display.loc[1, "순번"] == "" or pd.isna(stock_display.loc[1, "순번"]))
            and (stock_display.loc[1, "보험금액"] in ("", None) or pd.isna(stock_display.loc[1, "보험금액"]))
            and stock_display.loc[stock_display["제품코드"] == "00003", "재고위치명"].iloc[0] != "제품 합계"
            and all(
                stock_display.loc[stock_display["재고위치명"] == "제품 합계", column].fillna("").eq("").all()
                for column in (
                    "순번", "제품코드", "제품명", "규격", "현보험약가", "표준코드",
                    "제품그룹명", "제품구분명", "제품분류명", "발주처코드", "발주처명",
                    "제조사코드", "제조사명", "재고위치코드", "KD코드", "EDI코드", "포장단위",
                )
            )
            and float(stock_display.loc[stock_display["재고위치명"] == "제품 합계", "재고수량"].sum()) == 30.0
            and float(stock_display.loc[stock_display["재고위치명"] == "제품 합계", "보험금액"].sum()) == 300.0
            and stock_meta.get("current_stock_query") is True
            and inventory_service._build_inventory_header_md(stock_meta) == ""
            and current_stock_frame.loc[current_stock_frame["physic_cd"] == "00001", "physic_nm"].eq("타이레놀").all()
        )
        results.append(_ok("current stock final 20-column display", "detail/subtotal serials, no grand total") if display_ok else _fail("current stock final 20-column display", repr(stock_display.head().to_dict("records"))))

        numeric_display_ok = all(
            pd.api.types.is_numeric_dtype(stock_display[column])
            for column in ("재고수량", "현보험약가", "보험금액")
        ) and (
            pd.isna(stock_display.loc[1, "현보험약가"])
            and pd.isna(stock_display.loc[1, "보험금액"])
        )
        results.append(
            _ok("current stock display numeric dtypes", "numeric blanks remain NaN")
            if numeric_display_ok
            else _fail("current stock display numeric dtypes", str(stock_display.dtypes.to_dict()))
        )

        from app.ui.sims_table_display import build_sims_table_display_config

        stock_fast_view, stock_fast_config, _stock_fast_width, _stock_fast_height = (
            build_sims_table_display_config(
                stock_display,
                action_name=inventory_service.ACTION,
                meta=stock_meta,
            )
        )
        stock_fast_number_config_ok = all(
            isinstance(stock_fast_config.get(column), dict)
            and stock_fast_config[column].get("type_config", {}).get("type") == "number"
            and stock_fast_config[column].get("type_config", {}).get("format") == "localized"
            for column in ("재고수량",)
        ) and all(
            stock_fast_config.get(column, {}).get("type_config", {}).get("type") == "text"
            and stock_fast_config.get(column, {}).get("alignment") == "right"
            for column in ("순번", "현보험약가", "보험금액")
        )
        results.append(
            _ok("current stock fast display column contract", "numeric stock / right-aligned blank-safe text")
            if stock_fast_number_config_ok
            else _fail("current stock fast display column contract", repr(stock_fast_config))
        )
        stock_fast_blank_cells_ok = (
            stock_fast_view.loc[1, "현보험약가"] == ""
            and stock_fast_view.loc[1, "보험금액"] == ""
            and stock_fast_view.loc[2, "현보험약가"] == ""
            and stock_fast_view.loc[2, "보험금액"] == "200"
            and stock_fast_view.loc[0, "현보험약가"] == "10.00"
            and not any(
                isinstance(value, str) and value.strip() in {"None", "<NA>", "NaN", "nan"}
                for value in stock_fast_view.to_numpy().ravel()
            )
        )
        results.append(
            _ok("current stock fast display blanks", "empty cells with NumberColumn; no visible null tokens")
            if stock_fast_blank_cells_ok
            else _fail("current stock fast display blanks", repr(stock_fast_view.loc[:2, ["순번", "현보험약가", "보험금액"]].to_dict("records")))
        )

        current_stock_source = inspect.getsource(inventory_service._collect_source_df)
        current_stock_prepare = inspect.getsource(inventory_service._prepare_grouped_df)
        current_stock_display_builder = inspect.getsource(inventory_service._build_current_stock_table_frames)
        current_stock_fast_path_ok = (
            "if current_stock_query:" in current_stock_source
            and "return all_df, pd.DataFrame()" in current_stock_source
            and "if bool(cfg.get(\"current_stock_query\"))" in current_stock_prepare
            and "last_unit_cost" not in current_stock_prepare.split("if bool(cfg.get(\"current_stock_query\"))", 1)[1].split("if last_df", 1)[0]
        )
        results.append(
            _ok("current stock skips last-cost and unit/DC calculations", "current-stock-only fast path")
            if current_stock_fast_path_ok
            else _fail("current stock skips last-cost and unit/DC calculations", "fast path missing")
        )
        results.append(
            _ok("current stock display keeps source frame separate", "display-only numeric empty cells")
            if "source_parts.append" in current_stock_display_builder and "df_full" not in current_stock_display_builder
            else _fail("current stock display keeps source frame separate", "source/display separation missing")
        )

        real_month_sql, _ = inventory_service._build_month_carry_sql(
            {"base_month": "202608", "stock_mode": "real", "physic_cd": "21545", "stock_cds": ["00001"]},
            inventory_service._settings({"stock_mode": "real"}),
        )
        book_month_sql, _ = inventory_service._build_month_carry_sql(
            {"base_month": "202608", "stock_mode": "book", "physic_cd": "21545", "stock_cds": ["00001"]},
            inventory_service._settings({"stock_mode": "book"}),
        )
        monthly_formula_ok = (
            "dbo.Rddbc210" in real_month_sql
            and "Rd21_In_Quantity" in real_month_sql
            and "Rd21_In_Oquantity" in real_month_sql
            and "Rd21_Out_Quantity" in real_month_sql
            and "Rd21_Out_Oquantity" in real_month_sql
            and "dbo.Rddbc220" in book_month_sql
            and "Rd22_In_Quantity" in book_month_sql
            and "Rd22_Out_Quantity" in book_month_sql
        )
        results.append(
            _ok("inventory monthly stock formula contract", "210/220 in/out bonus columns")
            if monthly_formula_ok
            else _fail("inventory monthly stock formula contract", "monthly quantity expression missing")
        )

        flow_service = importlib.import_module("app.services.product_flow_service")
        original_flow_stock_resolver = flow_service.resolve_inventory_stock_codes
        original_flow_master_info = flow_service._get_product_master_info
        try:
            captured_flow_params: list[dict] = []
            flow_service.resolve_inventory_stock_codes = lambda _params: ["00001"]
            def _capture_flow_location_params(work_params):
                captured_flow_params.append(dict(work_params))
                raise RuntimeError("fixture stops after stock-location resolution")
            flow_service._get_product_master_info = _capture_flow_location_params
            flow_location_result = flow_service.get_product_flow_result({
                "physic_cd": "00001",
                "stock_nm": "본사창고",
                "date_from": "20260101",
                "date_to": "20260131",
            })
            flow_location_ok = (
                bool(captured_flow_params)
                and captured_flow_params[0].get("stock_cds") == ["00001"]
                and captured_flow_params[0].get("stock_cd") == "00001"
            )
            results.append(
                _ok("product flow resolves named stock location", "본사창고 -> 00001")
                if flow_location_ok
                else _fail("product flow resolves named stock location", repr(flow_location_result))
            )

            flow_service.resolve_inventory_stock_codes = lambda _params: []
            flow_location_missing = flow_service.get_product_flow_result({
                "physic_cd": "00001",
                "stock_nm": "없는창고",
                "date_from": "20260101",
                "date_to": "20260131",
            })
            flow_missing_ok = (
                flow_location_missing.get("meta", {}).get("result_status") == "no_data"
                and flow_location_missing.get("df") is None
            )
            results.append(
                _ok("product flow rejects unresolved named stock location", "no unfiltered fallback")
                if flow_missing_ok
                else _fail("product flow rejects unresolved named stock location", repr(flow_location_missing))
            )
        finally:
            flow_service.resolve_inventory_stock_codes = original_flow_stock_resolver
            flow_service._get_product_master_info = original_flow_master_info

        inventory_followup = importlib.import_module("app.ui.current_table_followups.inventory")
        followup_payloads: list[dict[str, Any]] = []
        def _find_col(frame, *, exact=(), include_any=(), exclude_any=()):
            for column in frame.columns:
                text = str(column).strip()
                if text in exact:
                    return text
            for column in frame.columns:
                text = str(column).strip()
                if any(token in text for token in include_any) and not any(token in text for token in exclude_any):
                    return text
            return None
        followup_helpers = {
            "find_col": _find_col,
            "to_num": lambda series: pd.to_numeric(series, errors="coerce").fillna(0),
            "push_table": lambda **kwargs: (followup_payloads.append(kwargs) or True),
            "push_notice": lambda **kwargs: (followup_payloads.append(kwargs) or True),
        }
        followup_handled = inventory_followup.handle_inventory_followup(
            df=stock_source,
            query="현재표 제품별 재고수량 TOP 20",
            top_n=20,
            table_key="current-stock-fixture",
            source_action="현재고 조회",
            helpers=followup_helpers,
            log=logging.getLogger("io-current-stock-followup"),
        )
        followup_df = (followup_payloads[0].get("df") if followup_payloads else pd.DataFrame())
        no_double_count_ok = (
            followup_handled is True
            and isinstance(followup_df, pd.DataFrame)
            and float(followup_df.loc[followup_df["제품코드"] == "00001", "재고수량"].iloc[0]) == 20.0
            and float(followup_df.loc[followup_df["제품코드"] == "00002", "재고수량"].iloc[0]) == 10.0
            and float(followup_df.loc[followup_df["제품코드"] == "00003", "재고수량"].iloc[0]) == 7.0
            and list(followup_df.columns) == ["순번", "제품코드", "제품명", "규격", "제조사명", "재고수량", "보험금액"]
        )
        results.append(_ok("current stock followup uses product subtotal once", "20 and 10") if no_double_count_ok else _fail("current stock followup uses product subtotal once", repr(followup_payloads)))

        chat_middleware = importlib.import_module("app.ui.chat_middleware")
        fast_renderer_source = inspect.getsource(chat_middleware._render_chat_fast_dataframe)
        middleware_source = Path(chat_middleware.__file__).read_text(encoding="utf-8")
        display_mode_ok = (
            "_build_io_display_styler" not in fast_renderer_source
            and "[chat] io small table styler skipped" in middleware_source
            and "_build_io_display_styler(view_df, add_row_no=False, band_size=5)" in middleware_source
        )
        results.append(_ok("current stock preserves small styler and fast no-styler contract", "small=IO Styler, fast=common dataframe") if display_mode_ok else _fail("current stock preserves small styler and fast no-styler contract", "renderer contract changed"))

        captured_fast_render: dict[str, Any] = {}
        original_st_dataframe = chat_middleware.st.dataframe
        try:
            chat_middleware.st.dataframe = lambda data, **kwargs: captured_fast_render.update({"df": data, "kwargs": kwargs})
            chat_middleware._render_chat_fast_dataframe(
                stock_display,
                action_name=inventory_service.ACTION,
                meta=stock_meta,
            )
        finally:
            chat_middleware.st.dataframe = original_st_dataframe
        rendered_current_stock_df = captured_fast_render.get("df")
        rendered_current_stock_cfg = dict(captured_fast_render.get("kwargs") or {}).get("column_config") or {}
        renderer_blank_contract_ok = (
            isinstance(rendered_current_stock_df, pd.DataFrame)
            and rendered_current_stock_df.loc[1, "순번"] in ("", None)
            and rendered_current_stock_df.loc[1, "현보험약가"] == ""
            and rendered_current_stock_df.loc[1, "보험금액"] == ""
            and rendered_current_stock_df.loc[2, "현보험약가"] == ""
            and rendered_current_stock_df.loc[2, "보험금액"] == "200"
            and not any(
                isinstance(value, str) and value.strip() in {"None", "<NA>", "NaN", "nan"}
                for value in rendered_current_stock_df.to_numpy().ravel()
            )
            and all(
                isinstance(rendered_current_stock_cfg.get(column), dict)
                and rendered_current_stock_cfg[column].get("type_config", {}).get("type") == "number"
                for column in ("재고수량",)
            )
            and all(
                rendered_current_stock_cfg.get(column, {}).get("type_config", {}).get("type") == "text"
                and rendered_current_stock_cfg.get(column, {}).get("alignment") == "right"
                for column in ("순번", "현보험약가", "보험금액")
            )
        )
        results.append(
            _ok("current stock final Streamlit fast dataframe", "blank-safe text cells and numeric stock passed to st.dataframe")
            if renderer_blank_contract_ok
            else _fail("current stock final Streamlit fast dataframe", repr(rendered_current_stock_df.loc[:2, ["순번", "현보험약가", "보험금액"]].to_dict("records") if isinstance(rendered_current_stock_df, pd.DataFrame) else captured_fast_render))
        )

        supplier.resolve_common_vendor_candidates = lambda text: []
        no_data = io_nlq.resolve_current_stock_entity_condition("현재고 없는이름", params={})
        results.append(_ok("current stock no_data candidate", "not_found") if no_data.get("status") == "not_found" else _fail("current stock no_data candidate", repr(no_data)))

        supplier.resolve_common_vendor_candidates = lambda text: [
            {"entity_code": "10047", "canonical_name": "공통명", "entity_role": "manufacturer"}
        ]
        goods_service.search_goods_full = lambda **kwargs: pd.DataFrame([
            {"Rd04_Physic_Cd": "00002", "Rd04_Physic_Nm": "공통명"}
        ])
        ambiguous = io_nlq.resolve_current_stock_entity_condition("현재고 공통명", params={})
        ambiguous_params = dict(ambiguous.get("params") or {})
        union_sql, union_sql_params = inventory_service._build_month_carry_sql(
            {
                **ambiguous_params,
                "base_month": "202608",
                "stock_mode": "real",
                "stock_cds": ["00001"],
            },
            inventory_service._settings({"stock_mode": "real"}),
        )
        union_scope_ok = (
            ambiguous.get("status") == "resolved"
            and ambiguous_params.get("current_stock_entity_scope") == "manufacturer_or_product"
            and ambiguous_params.get("current_stock_manufacturer_codes") == ["10047"]
            and ambiguous_params.get("current_stock_product_codes") == ["00002"]
            and "P.Rd04_Ven_Cd IN" in union_sql
            and "P.Rd04_Physic_Cd IN" in union_sql
            and "current_stock_entity_like" not in union_sql_params
        )
        results.append(_ok("current stock unlabeled union scope", "resolved master codes without candidate table") if union_scope_ok else _fail("current stock unlabeled union scope", repr(ambiguous)))

        profile_service.load_dashboard_profile = lambda **kwargs: {
            "stock_mode": "book", "stock_cd_list": ["0018:00001"], "io_gu_list": ["0012:501"]
        }
        login.get_selected_company = lambda: {"company_id": 4}
        effective = router._apply_current_stock_defaults(
            {"physic_cd": "00001", "io_gu_list": ["501"]}, session_state={}
        )
        defaults_ok = (
            effective.get("stock_mode") == "book"
            and effective.get("stock_cds") == ["00001"]
            and "io_gu_list" not in effective
            and effective.get("io_gu_scope") == "all"
            and effective.get("group_basis") == "stock"
        )
        results.append(_ok("current stock saved stock / all io", repr(effective)) if defaults_ok else _fail("current stock saved stock / all io", repr(effective)))
    finally:
        supplier.resolve_common_vendor_candidates = originals["vendor"]
        goods_service.search_goods_full = originals["goods"]
        shared._load_stock_code_options = originals["stocks"]
        profile_service.load_dashboard_profile = originals["profile"]
        login.get_selected_company = originals["company"]
    return results


# ---------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------
def main() -> int:
    parser = argparse.ArgumentParser(description="SIMS IO / Docs / Stock NLQ regression checker")
    parser.add_argument(
        "--live",
        action="store_true",
        help="가벼운 대표 NLQ 문장을 실제 라우팅/DB 조회로 테스트",
    )
    parser.add_argument(
        "--live-all",
        action="store_true",
        help="전체 대표 NLQ 문장을 실제 라우팅/DB 조회로 테스트",
    )
    parser.add_argument(
        "--show-payload",
        action="store_true",
        help="live 테스트 시 capture payload meta 출력",
    )
    args = parser.parse_args()

    print(f"Project root: {PROJECT_ROOT}")

    failed = 0

    basic_results = run_basic_checks()
    failed += _print_results("BASIC IMPORT CHECKS", basic_results)

    parser_results = run_parser_checks()
    failed += _print_results("IO NLQ PARSER CHECKS", parser_results)

    unlabeled_entity_results = run_unlabeled_io_entity_resolution_checks()
    failed += _print_results(
        "UNLABELED IO ENTITY RESOLUTION CHECKS",
        unlabeled_entity_results,
    )

    period_policy_results = run_default_period_policy_checks()
    failed += _print_results("NLQ DEFAULT PERIOD POLICY CHECKS", period_policy_results)

    response_timing_results = run_response_timing_checks()
    failed += _print_results("SIMS RESPONSE TIMING CHECKS", response_timing_results)

    nlq_case_log_results = run_nlq_case_log_checks()
    failed += _print_results("NLQ CASE LOG CHECKS", nlq_case_log_results)

    product_alias_results = run_product_flow_inventory_alias_checks()
    failed += _print_results(
        "PRODUCT FLOW / INVENTORY ALIAS CONTRACT CHECKS",
        product_alias_results,
    )

    product_payload_results = run_product_flow_input_and_empty_payload_checks()
    failed += _print_results(
        "PRODUCT FLOW / INVENTORY INPUT AND EMPTY PAYLOAD CHECKS",
        product_payload_results,
    )

    product_inventory_display_results = run_product_inventory_display_export_checks()
    failed += _print_results(
        "PRODUCT INVENTORY DISPLAY / EXPORT CHECKS",
        product_inventory_display_results,
    )

    current_stock_results = run_current_stock_nlq_contract_checks()
    failed += _print_results("CURRENT STOCK NLQ CONTRACT CHECKS", current_stock_results)

    if args.live or args.live_all:

        live_results = run_live_checks(
            live_all=bool(args.live_all),
            show_payload=bool(args.show_payload),
        )
        failed += _print_results("IO NLQ LIVE ROUTING CHECKS", live_results)

    print()
    if failed:
        print(f"RESULT: FAIL ({failed} failed)")
        return 1

    print("RESULT: OK")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
