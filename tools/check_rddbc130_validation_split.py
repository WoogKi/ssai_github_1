"""In-memory contract guard for Rddbc130 ordinary versus explicit validation queries."""

from __future__ import annotations

import sys
import io
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.services import io_nlq
from app.services import rddbc130_service as service
from app.sims.nlq import nlq_router
from app.sims.views import rddbc_io_shared
from app.ui import chat_middleware
from app.ui.sims_table_display import resolve_sims_excel_number_format
from app.ui.current_table_followups.action_dispatcher import (
    handle_current_table_followup_by_action,
    is_bound_current_trans_doc_validation_request,
    is_explicit_current_trans_doc_validation_request,
)


def _assert(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def _base_frame() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Rd13_Trans_Di": "1   ",
                "거래명세서구분명": "입고",
                "Rd13_Trans_YyMmDd": "20260831",
                "Rd13_Ven_Cd": "A   ",
                "거래처명": "거래처A",
                "Rd13_Trans_Seq": 1.0,
                "Rd13_Supply_Price": 100,
                "Rd13_Tax_Price": 10,
                "Rd13_Tot_Amt": 110,
                "Rd13_Dc_Amt": 1,
            },
            {
                "Rd13_Trans_Di": "3",
                "거래명세서구분명": "출고",
                "Rd13_Trans_YyMmDd": "20260831",
                "Rd13_Ven_Cd": "B",
                "거래처명": "거래처B",
                "Rd13_Trans_Seq": 2,
                "Rd13_Supply_Price": 200,
                "Rd13_Tax_Price": 20,
                "Rd13_Tot_Amt": 220,
                "Rd13_Dc_Amt": 0,
            },
            {
                "Rd13_Trans_Di": "9",
                "거래명세서구분명": "기타",
                "Rd13_Trans_YyMmDd": "20260830",
                "Rd13_Ven_Cd": "C",
                "거래처명": "거래처C",
                "Rd13_Trans_Seq": 3,
                "Rd13_Supply_Price": 30,
                "Rd13_Tax_Price": 3,
                "Rd13_Tot_Amt": 33,
                "Rd13_Dc_Amt": 2,
            },
            {
                "Rd13_Trans_Di": "3",
                "거래명세서구분명": "출고",
                "Rd13_Trans_YyMmDd": "20260831",
                "Rd13_Ven_Cd": "D",
                "거래처명": "거래처D",
                "Rd13_Trans_Seq": 4,
                "Rd13_Supply_Price": 40,
                "Rd13_Tax_Price": 4,
                "Rd13_Tot_Amt": 44,
                "Rd13_Dc_Amt": 0,
            },
        ]
    )


def _inbound_frame() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"Trans_Di": "1", "Trans_YyMmDd": "20260831", "Ven_Cd": "A", "Trans_Seq": 1, "Sum_Supply": 100, "Sum_Tax": 10},
            {"Trans_Di": "9", "Trans_YyMmDd": "20260830", "Ven_Cd": "C", "Trans_Seq": 3, "Sum_Supply": 30, "Sum_Tax": 3},
        ]
    )


def _outbound_frame() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"Trans_Di": "3", "Trans_YyMmDd": "20260831", "Ven_Cd": "B", "Trans_Seq": 2, "Sum_Supply": 199, "Sum_Tax": 20},
            {"Trans_Di": "9", "Trans_YyMmDd": "20260830", "Ven_Cd": "C", "Trans_Seq": 3, "Sum_Supply": 999, "Sum_Tax": 99},
        ]
    )


def _run_service_contract() -> None:
    original_query = service.query_to_df
    calls: list[str] = []

    def fake_query(sql: str, _params: dict) -> pd.DataFrame:
        calls.append(sql)
        if "dbo.Rddbc110 AS In_Put" in sql:
            return _inbound_frame()
        if "dbo.Rddbc120 AS Out_Put" in sql:
            return _outbound_frame()
        return _base_frame()

    service.query_to_df = fake_query
    try:
        ordinary = service.get_rddbc130_df({"date_from": "20260831", "date_to": "20260831"})
        _assert(len(calls) == 1, "ordinary transaction document query must issue one header query")
        _assert("dbo.Rddbc110" not in calls[0] and "dbo.Rddbc120" not in calls[0], "ordinary query must not include validation detail tables")
        _assert("상세합계일치" not in ordinary.columns, "ordinary result must not fabricate validation columns")

        calls.clear()
        validation = service.get_rddbc130_df(
            {"date_from": "20260831", "date_to": "20260831", "validation_requested": True}
        )
        _assert(len(calls) == 3, "validation must issue header, inbound scoped, and outbound scoped queries")
        _assert("dbo.Rddbc110" not in calls[0] and "dbo.Rddbc120" not in calls[0], "validation header query must remain header-only")
        for detail_sql in calls[1:]:
            _assert("CurrentResultBooks AS" in detail_sql, "current-result validation must scope the displayed header keys")
            _assert("FilteredTransactionKeys AS" in detail_sql, "validation detail query must scope filtered document keys")
            _assert("INNER JOIN FilteredTransactionKeys AS Keys" in detail_sql, "validation detail query must use the scoped native key join")
            _assert("LTRIM(RTRIM" not in detail_sql or "Trans_Seq" not in detail_sql, "validation detail query must not restore Trans_Seq conversion")

        by_vendor = {str(row["Rd13_Ven_Cd"]).strip(): row for _, row in validation.iterrows()}
        _assert(by_vendor["A"]["상세합계일치"] == "Y", "inbound-only matching document changed")
        _assert(by_vendor["A"]["상세합_공급가액"] == 100, "fixed-char pandas merge lost inbound detail supply")
        _assert(by_vendor["A"]["상세합_세액"] == 10, "fixed-char pandas merge lost inbound detail tax")
        _assert(str(by_vendor["A"]["Rd13_Ven_Cd"]) == "A   ", "merge normalization must not mutate header key values")
        _assert(by_vendor["B"]["상세합계일치"] == "N", "outbound-only mismatch document changed")
        _assert(by_vendor["C"]["거래명세서구분명"] == "입고", "both-detail document must retain inbound precedence")
        _assert(by_vendor["C"]["상세합계일치"] == "Y", "both-detail document must compare inbound totals first")
        _assert(by_vendor["D"]["상세합계일치"] == "상세없음", "missing detail must be explicit")
        _assert(pd.isna(by_vendor["D"]["상세합_공급가액"]), "missing detail supply must remain null")
        _assert(pd.isna(by_vendor["D"]["상세합_세액"]), "missing detail tax must remain null")

        calls.clear()
        mismatch = service.get_rddbc130_df(
            {"date_from": "20260831", "date_to": "20260831", "validation_requested": True, "only_mismatch": "Y"}
        )
        _assert(len(calls) == 3, "mismatch request must use the same key-scoped validation path")
        _assert(list(mismatch["Rd13_Ven_Cd"]) == ["B"], "only_mismatch must exclude no-detail rows")

        calls.clear()
        full_range = service.get_rddbc130_df(
            {
                "date_from": "20260831",
                "date_to": "20260831",
                "validation_requested": True,
                "validation_scope": "full_range",
            }
        )
        _assert(len(calls) == 3, "full-range validation must issue header, inbound scoped, and outbound scoped queries")
        _assert("SELECT TOP" not in calls[0], "full-range validation must not lose headers to display TOP")
        for detail_sql in calls[1:]:
            _assert("FilteredBooks AS" in detail_sql, "full-range validation must be driven by filtered Rddbc130 headers")
            _assert("CurrentResultBooks AS" not in detail_sql, "full-range validation must not use display TOP keys")
        _assert(len(full_range) == len(_base_frame()), "full-range fixture result changed")
    finally:
        service.query_to_df = original_query


def _run_current_source_validation_contract() -> None:
    original_query = service.query_to_df
    calls: list[str] = []

    def fake_query(sql: str, _params: dict) -> pd.DataFrame:
        calls.append(sql)
        if "dbo.Rddbc110 AS In_Put" in sql:
            return _inbound_frame()
        if "dbo.Rddbc120 AS Out_Put" in sql:
            return _outbound_frame()
        raise AssertionError("current-result validation must not query Rddbc130")

    service.query_to_df = fake_query
    try:
        current_display = _base_frame().rename(
            columns={
                "Rd13_Trans_Di": "거래명세서구분",
                "Rd13_Trans_YyMmDd": "거래명세서일자",
                "Rd13_Ven_Cd": "거래처코드",
                "Rd13_Trans_Seq": "거래명세서순번",
                "Rd13_Supply_Price": "공급가액",
                "Rd13_Tax_Price": "세액",
            }
        )
        validated = service.validate_rddbc130_current_result_df(current_display)
        _assert(len(calls) == 2, "current-result validation must issue inbound/outbound detail queries once")
        _assert(all("dbo.Rddbc130" not in sql for sql in calls), "current-result validation must not reread Rddbc130")
        _assert(all("CurrentResultKeys" in sql for sql in calls), "current-result validation must drive detail with stashed keys")
        _assert(all("INNER JOIN CurrentResultKeys AS Keys" in sql for sql in calls), "current-result validation must use native key joins")
        _assert("Rd13_Trans_Seq" not in validated.columns, "current-table validation must not leak temporary raw aliases")
        by_vendor = {str(row["거래처코드"]).strip(): row for _, row in validated.iterrows()}
        _assert(by_vendor["B"]["상세합계일치"] == "N", "current-result mismatch comparison changed")
        _assert(by_vendor["D"]["상세합계일치"] == "상세없음", "current-result no-detail state changed")
        _assert(len(validated[validated["상세합계일치"].eq("N")]) == 1, "only actual N is a mismatch")
    finally:
        service.query_to_df = original_query


def _run_empty_validation_summary_reuse_contract() -> None:
    original = nlq_router._get_trans_doc_full_summary
    calls: list[dict] = []

    def forbidden_summary(params: dict) -> dict:
        calls.append(dict(params))
        raise AssertionError("empty validation payload must not requery full summary")

    nlq_router._get_trans_doc_full_summary = forbidden_summary
    try:
        payload = {"df_display": pd.DataFrame(), "meta": {"row_count": 0}}
        result = nlq_router._ensure_trans_doc_llm_summary(
            payload,
            "거래명세서 공통 조회",
            {"validation_requested": True, "only_mismatch": "Y"},
            "기간 2026-08-31",
        )
        _assert(calls == [], "empty validation summary must reuse payload without DB work")
        _assert(result["meta"].get("llm_summary_kind") == "trans_doc_empty", "empty validation summary contract changed")
    finally:
        nlq_router._get_trans_doc_full_summary = original


def _run_parser_contract() -> None:
    ordinary = io_nlq.resolve_io_nlq("오늘 거래명세서 공통 조회")
    _assert(ordinary and ordinary["action"] == "거래명세서 공통 조회", "ordinary transaction document route changed")
    _assert(not ordinary["params"].get("validation_requested"), "ordinary query must not request validation")

    validation = io_nlq.resolve_io_nlq("오늘 거래명세서 공통 상세합계 검증")
    _assert(validation and validation["action"] == "거래명세서 공통 조회", "explicit validation route changed")
    _assert(validation["params"].get("validation_requested") is True, "explicit validation flag missing")
    _assert(validation["params"].get("validation_scope") == "full_range", "new NLQ validation must use full-range scope")

    mismatch = io_nlq.resolve_io_nlq("오늘 거래명세서 공통 부적합자료 조회")
    _assert(mismatch and mismatch["params"].get("validation_requested") is True, "mismatch validation flag missing")
    _assert(mismatch["params"].get("only_mismatch") == "Y", "mismatch-only contract missing")


def _run_analysis_contract() -> None:
    original_query = service.query_to_df
    calls: list[str] = []

    def fake_query(sql: str, _params: dict) -> pd.DataFrame:
        calls.append(sql)
        return pd.DataFrame(
            [
                {"section": "overall", "name": "전체", "row_count": 2, "supply_sum": 300, "tax_sum": 30, "amount_sum": 330, "dc_sum": 1, "mismatch_count": None, "vendor_count": 2},
                {"section": "by_trans_type", "name": "매입분", "row_count": 1, "supply_sum": 100, "tax_sum": 10, "amount_sum": 110, "dc_sum": 1, "mismatch_count": None, "vendor_count": None},
            ]
        )

    service.query_to_df = fake_query
    try:
        summary = service.get_rddbc130_analysis_summary({"date_from": "20260831", "date_to": "20260831"})
        _assert(len(calls) == 1, "ordinary analysis must issue one header-only query")
        _assert("dbo.Rddbc110" not in calls[0] and "dbo.Rddbc120" not in calls[0], "ordinary analysis must not include validation detail tables")
        _assert(summary.get("validation_performed") is False, "ordinary analysis must remain explicitly unvalidated")
        _assert(summary.get("mismatch_count") is None and summary.get("by_match_status") == [], "ordinary analysis must not fabricate validation results")
    finally:
        service.query_to_df = original_query


def _run_current_result_and_formatter_contract() -> None:
    source = _merge_fixture_for_current_result()
    pushed: list[dict] = []
    six_queries = (
        "현재 거래명세서 검증",
        "현재거래명세서 검증",
        "현재 거래명세서 불일치",
        "현재거래명세서 불일치",
        "현재 거래명세서 불일치 확인",
        "현재 거래명세서 상세합계 확인",
    )

    def push_table(**kwargs):
        pushed.append({"kind": "table", **kwargs})
        return True

    def push_notice(**kwargs):
        pushed.append({"kind": "notice", **kwargs})
        return True

    helpers = {
        "find_col": lambda df, exact=(), include_any=(), exclude_any=(): next((c for c in exact if c in df.columns), None),
        "to_num": lambda series: pd.to_numeric(series, errors="coerce").fillna(0),
        "push_table": push_table,
        "push_notice": push_notice,
    }
    for query in six_queries:
        pushed.clear()
        handled = handle_current_table_followup_by_action(
            df=source, query=query, top_n=20, table_key="trans-doc-source", source_action="거래명세서 공통 조회", helpers=helpers, log=type("L", (), {"info": lambda *_args, **_kwargs: None})(),
        )
        _assert(handled and pushed, f"current result validation did not handle {query}")
        _assert(pushed[0]["kind"] == "table", f"current result validation did not return in-memory table for {query}")
        if "불일치" in query:
            _assert(len(pushed[0]["df"]) == 1, "current mismatch must filter existing source only")

    values = pd.Series([12.0, 0.0, pd.NA], dtype="Float64")
    for column in ("거래명세서순번", "전표순번", "배송순번", "피킹출력순번", "Rd13_Delivery_PrinterSeq"):
        normalized = rddbc_io_shared._maybe_to_numeric(values, column)
        _assert(str(normalized.dtype) == "Int64", f"{column} must use nullable integer formatter")
        _assert(normalized.iloc[0] == 12, f"{column} integer value changed")
        _assert(resolve_sims_excel_number_format(column) == "#,##0", f"{column} Excel number format must be integer")

    for query in six_queries:
        _assert(
            is_explicit_current_trans_doc_validation_request(query),
            f"explicit current transaction validation intent missed: {query}",
        )
        _assert(
            is_bound_current_trans_doc_validation_request(
                query,
                source_action="거래명세서 공통 조회",
                current_table_present=True,
            ),
            f"bound current transaction validation route missed: {query}",
        )
        _assert(
            not is_bound_current_trans_doc_validation_request(
                query,
                source_action="출고명세 조회",
                current_table_present=True,
            ),
            f"non-transaction current source misrouted: {query}",
        )
        _assert(
            not is_bound_current_trans_doc_validation_request(
                query,
                source_action="거래명세서 공통 조회",
                current_table_present=False,
            ),
            f"missing current source misrouted: {query}",
        )

    main_source = (ROOT / "app" / "Lmstudio_SSAI_chat_main.py").read_text(encoding="utf-8-sig")
    decision_pos = main_source.index("explicit_current_trans_doc_validation = is_explicit_current_trans_doc_validation_request(user_input)")
    web_pos = main_source.index("web_search_route = parse_web_search_request(user_input)")
    nlq_pos = main_source.index("new_sims_candidate = resolve_new_sims_nlq_candidate(user_input)")
    _assert(decision_pos < web_pos < nlq_pos, "bound current validation must preempt web and new SIMS NLQ routing")

    export_df = pd.DataFrame(
        {
            "거래명세서구분": [1, 3],
            "거래명세서순번": [12, 11],
            "전표순번": [0, 4],
            "배송순번": [0, 7],
            "피킹출력순번": [0, 9],
            "상세합_공급가액": [100, 199],
            "상세합_세액": [10, 20],
            "상세합계일치": ["Y", "N"],
            "단가": [1.25, 2.5],
        }
    )
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="SIMS")
        chat_middleware._apply_sims_excel_number_formats(writer, export_df, "SIMS")
    output.seek(0)
    sheet = load_workbook(output, data_only=True)["SIMS"]
    header_index = {cell.value: cell.column for cell in sheet[1]}
    for column in ("거래명세서순번", "전표순번", "배송순번", "피킹출력순번"):
        cell = sheet.cell(2, header_index[column])
        _assert(isinstance(cell.value, int), f"{column} Excel value must remain numeric integer")
        _assert(cell.number_format == "#,##0", f"{column} Excel format changed: {cell.number_format}")
    _assert(sheet.cell(2, header_index["거래명세서구분"]).number_format == "General", "transaction type code must not use decimal format")
    _assert(sheet.cell(2, header_index["상세합계일치"]).number_format == "General", "validation status must remain textual")
    _assert(sheet.cell(2, header_index["단가"]).number_format == "#,##0.##", "unit-price formatting changed")

    empty_status_df = export_df.copy()
    empty_status_df["상세합계일치"] = pd.Series([None, None], dtype="float64")
    empty_output = io.BytesIO()
    with pd.ExcelWriter(empty_output, engine="openpyxl") as writer:
        empty_status_df.to_excel(writer, index=False, sheet_name="SIMS")
        chat_middleware._apply_sims_excel_number_formats(writer, empty_status_df, "SIMS")
    empty_output.seek(0)
    empty_sheet = load_workbook(empty_output, data_only=True)["SIMS"]
    empty_header_index = {cell.value: cell.column for cell in empty_sheet[1]}
    _assert(
        empty_sheet.cell(2, empty_header_index["상세합계일치"]).number_format == "General",
        "all-empty validation status must not inherit a numeric format",
    )


def _merge_fixture_for_current_result() -> pd.DataFrame:
    return pd.DataFrame([
        {"거래명세서일자": "20260831", "거래처명": "A", "공급가액": 100, "세액": 10, "합계금액": 110, "상세합_공급가액": 100, "상세합_세액": 10, "상세합계일치": "Y"},
        {"거래명세서일자": "20260831", "거래처명": "B", "공급가액": 200, "세액": 20, "합계금액": 220, "상세합_공급가액": 199, "상세합_세액": 20, "상세합계일치": "N"},
    ])


def main() -> None:
    _run_service_contract()
    _run_current_source_validation_contract()
    _run_empty_validation_summary_reuse_contract()
    _run_parser_contract()
    _run_analysis_contract()
    _run_current_result_and_formatter_contract()
    print("RESULT: OK")
    print("ordinary header-only; explicit validation native-key scoped; only_mismatch and inbound precedence preserved")


if __name__ == "__main__":
    main()
