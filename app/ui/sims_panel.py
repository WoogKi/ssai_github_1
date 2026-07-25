# app/ui/sims_panel.py
# VERSION = "chat_middleware/2025-11-11T-stable"
#VERSION = "vendors/2025-12-14-001"

from __future__ import annotations

import logging
from typing import Dict, Any, Optional, Mapping
import io
import re
import os
import uuid
import hashlib
import json

import datetime as dt

import streamlit as st
import pandas as pd
from app.ui.sims_table_display import log_sims_display_fields, log_sims_table_mode, log_sims_table_render

log = logging.getLogger("ssai")

# ==========================================================
# 🧩 Submit 상태 헬퍼
# ==========================================================
def _mark_submitted() -> None:
    ss = st.session_state
    ss["__sims_form_submitted"] = True
    ss["__sims_submitted_form_id"] = ss.get("__sims_form_id")
    ss["__sims_query_submit_seq"] = int(ss.get("__sims_query_submit_seq") or 0) + 1
    selected = ss.get("__sims_selected") or {}
    action = str(selected.get("action") or "").strip()
    if action:
        ss.pop(f"__sims_panel_query_fingerprint::{action}", None)
    log.debug("[panel.form] submitted=True (form_id=%s)", ss.get("__sims_form_id"))


def _is_current_submit() -> bool:
    """현재 렌더가 '방금 제출한 동일 form_id'에 해당하면 True"""
    ss = st.session_state

    if ss.get("__sims_form_submitted") and ss.get("__sims_submitted_form_id") == ss.get("__sims_form_id"):
        # ✅ 1회만 소비
        ss["__sims_form_submitted"] = False
        return True
    return False

def _safe_filename(name: str) -> str:
    return "".join(ch if ch.isalnum() else "_" for ch in str(name)).strip("_") or "result"

def _export_unavailable_message() -> str:
    return "내보내기 기능은 현재 계정에서 사용할 수 없습니다. 필요하시면 관리자에게 요청해 주세요."


def _export_unavailable_help() -> str:
    return "관리자 승인 후 CSV/Excel 저장 기능을 사용할 수 있습니다."

def _clean_ui_text(value: Any) -> str:
    return str(value or "").strip()


def _selected_sims_category() -> str:
    sel = st.session_state.get("__sims_selected") or {}
    return str(sel.get("category") or "").strip()


def _is_io_category() -> bool:
    return _selected_sims_category() == "입출고/명세서/재고"


def _trim_object_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for c in out.columns:
        if not pd.api.types.is_numeric_dtype(out[c]):
            out[c] = (
                out[c]
                .fillna("")
                .astype(str)
                .replace({"None": "", "nan": "", "<NA>": ""})
                .str.strip()
            )
    return out


def _is_code_col_name(col: str) -> bool:
    s = str(col or "").strip()
    if s == "순번":
        return False
    return ("코드" in s) or (s in {"KD코드", "EDI코드", "표준코드"})


def _is_numeric_col(df: pd.DataFrame, col: str) -> bool:
    s = str(col or "").strip()
    s_lower = s.lower()

    if s == "순번":
        return True

    # 1) 날짜/시간 컬럼은 무조건 숫자 취급 금지
    if any(k in s for k in ["일자", "날짜", "일시", "시간"]):
        return False

    # 2) 이름/명칭/코드 계열도 숫자 취급 금지
    if any(k in s for k in ["명", "이름", "코드", "ID", "번호", "적용처"]):
        return False

    if s in {
        "완료월총매출",
        "월평균매출",
        "완료월평균매출",
        "당월 현재매출",
        "당월 예상매출",
        "당월 잔여예상",
        "전월대비매출",
        "총매출공급가액",
        "매출공급가액",
        "매출세액",
        "매출합계",
        "전월대비매출",
        "총매출공급가액",
        "매출공급가액",
        "매출세액",
        "매출합계",
        "최근3개월평균매출",
        "최근6개월평균매출",
        "당월 진척률",
        "최근3개월증감률",
        "적용증감률",
        "월시점 증감률",
        "월시점 적용증감률",
        "월시점 달성률",
    }:
        return True

    # 3) 실제 dtype 이 숫자면 숫자
    if pd.api.types.is_numeric_dtype(df[col]):
        return True

    # 4) 컬럼명 기준 숫자 후보
    return any(k in s for k in ["수량", "금액", "단가", "DC율", "할인율", "보험약가", "보험가", "가격"])

def _is_decimal_col_name(col: str) -> bool:
    s = str(col or "").strip()
    if s in {
        "당월 진척률",
        "최근3개월증감률",
        "적용증감률",
        "월시점 증감률",
        "월시점 적용증감률",
        "월시점 달성률",
        "최근3개월수량증감률",
        "수요증감률",
        "수요적용증감률",
        "평가월 수요진척률",
        "당월 출고진척률",
        "당월 재고충족률",
    }:
        return True
    if s in {
        "완료월총매출",
        "월평균매출",
        "완료월평균매출",
        "당월 현재매출",
        "당월 예상매출",
        "당월 잔여예상",
        "다음월예상매출",
        "3개월예상매출",
        "6개월예상매출",
    }:
        return False
    if s in {"최근3개월평균매출", "최근6개월평균매출", "평균공급단가"}:
        return True
    return any(k in s for k in ["단가", "DC율", "할인율", "보험약가", "보험가"])


def _ensure_seq_col(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "순번" in out.columns:
        return out

    seq = []
    n = 0
    for _ in range(len(out)):
        n += 1
        seq.append(n)

    out.insert(0, "순번", pd.Series(seq, dtype="Int64"))
    return out

def _make_unique_columns(columns) -> list[str]:
    counts: dict[str, int] = {}
    result: list[str] = []

    for col in columns:
        base = str(col)
        n = counts.get(base, 0)
        if n == 0:
            result.append(base)
        else:
            result.append(f"{base}_{n}")
        counts[base] = n + 1

    return result


def _make_styler_safe_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out = out.reset_index(drop=True)
    out.columns = _make_unique_columns(out.columns)
    return out

def _normalize_zero_like_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    for c in out.columns:
        if pd.api.types.is_numeric_dtype(out[c]):
            out[c] = out[c].apply(
                lambda v: 0 if (pd.notna(v) and abs(float(v)) < 1e-12) else v
            )
        else:
            out[c] = (
                out[c]
                .fillna("")
                .astype(str)
                .str.strip()
                .replace({
                    "None": "",
                    "none": "",
                    "nan": "",
                    "NaN": "",
                    "<NA>": "",
                    "NaT": "",
                    "nat": "",
                    "NULL": "",
                    "null": "",
                })
            )
    return out

def _is_large_table_for_fast_render(df: pd.DataFrame) -> bool:
    """Styler 대신 빠른 st.dataframe 렌더를 쓸지 판단."""
    if df is None or df.empty:
        return False

    try:
        cells = int(len(df)) * int(len(df.columns))
    except Exception:
        return False

    threshold = int(os.getenv("SIMS_FAST_TABLE_CELL_THRESHOLD", "6000"))
    return cells >= threshold


def _is_fast_numeric_column(df: pd.DataFrame, col: str) -> bool:
    """
    빠른 표 모드에서 숫자 처리할 컬럼 판정.

    중요:
    - 제품코드/거래처코드/보험코드처럼 숫자처럼 보이는 코드는 숫자로 바꾸면 안 됨.
    - 이름/명칭/일자/등급도 숫자 처리 제외.
    """
    s = str(col or "").strip()
    s_lower = s.lower()

    if s == "순번":
        return True

    exclude_words = [
        "코드",
        "stock_cd",
        "buy_cd",
        "physic_cd",
        "ven_cd",
        "_cd",
        "ID",
        "번호",
        "일자",
        "날짜",
        "일시",
        "시간",
        "기준월",
        "명",
        "이름",
        "재고기준",
        "자료원",
        "추세판정",
        "예상등급",
        "부족등급",
        "예상기준",
    ]
    if any(w in s or w in s_lower for w in exclude_words):
        return False

    include_words = [
        "수량",
        "금액",
        "단가",
        "커버월수",
        "평균",
        "매출",
        "공급가액",
        "세액",
        "율",
        "건수",
        "품목수",
        "거래처수",
        "매입처수",
        "재고적용처수",
        "분석월수",
        "매출발생월수",
        "부족",
        "필요",
    ]

    if any(w in s for w in include_words):
        return True

    try:
        return pd.api.types.is_numeric_dtype(df[col])
    except Exception:
        return False


def _fast_display_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    화면 표시용 빠른 DataFrame.
    - 코드/명칭/일자는 문자 유지
    - 숫자 컬럼만 numeric 유지
    - 금액성 컬럼은 0자리 반올림
    - 평균/단가/율/커버월수는 2자리 반올림
    - Styler는 사용하지 않음
    """
    out = df.copy()

    for col in out.columns:
        s = str(col or "").strip()

        if not _is_fast_numeric_column(out, col):
            continue

        num = pd.to_numeric(
            out[col].astype(str).str.replace(",", "", regex=False),
            errors="coerce",
        )

        int_cols = {
            "완료월총매출",
            "월평균매출",
            "완료월평균매출",
            "당월 현재매출",
            "당월 예상매출",
            "당월 잔여예상",
            "전월대비매출",
            "총매출공급가액",
            "매출공급가액",
            "매출세액",
            "매출합계",
            "평가월 현재매출",
            "평가월 예상매출",
            "평가월 잔여예상",
            "다음월예상매출",
            "3개월예상매출",
            "6개월예상매출",
            "부족예상금액",
            "완료월수",
            "매출발생월수",
            "매입처수",
            "총집계건수",
        }
        percent_cols = {
            "당월 진척률",
            "평가월 진척률",
            "전월대비매출증감률",
            "최근3개월증감률",
            "월시점 최근3개월증감률",
            "적용증감률",
            "월시점 증감률",
            "월시점 적용증감률",
            "월시점 달성률",
            "최근3개월수량증감률",
            "수요증감률",
            "수요적용증감률",
            "평가월 수요진척률",
            "당월 출고진척률",
            "당월 재고충족률",
        }
        decimal_cols = {
            "최근3개월평균매출",
            "최근6개월평균매출",
            "평균공급단가",
            "완료월평균출고수량",
            "최근3개월평균출고수량",
            "최근6개월평균출고수량",
            "당월 예상출고수량",
            "당월 잔여예상출고수량",
            "예상월말재고수량",
            "부족예상수량",
        }

        if s in percent_cols:
            out[col] = num
        elif s in decimal_cols:
            out[col] = num.round(2)
        elif (
            s in int_cols
            or s in {"순번", "조회순번"}
            or s.endswith("건수")
            or "품목수" in s
            or "거래처수" in s
            or "매입처수" in s
            or "재고적용처수" in s
            or "금액" in s
            or "매출" in s
            or "공급가액" in s
            or "세액" in s
            or "가격" in s
        ):
            out[col] = num.round(0)
        else:
            out[col] = num.round(2)

    return out


def _fast_column_config(df: pd.DataFrame) -> dict:
    """
    빠른 st.dataframe용 column_config.
    개별 셀 CSS 없이 NumberColumn만 지정한다.
    """
    cfg: dict = {}

    for col in df.columns:
        s = str(col or "").strip()

        if not _is_fast_numeric_column(df, col):
            continue

        if s in {
            "당월 진척률",
            "평가월 진척률",
            "전월대비매출증감률",
            "최근3개월증감률",
            "월시점 최근3개월증감률",
            "적용증감률",
            "월시점 증감률",
            "월시점 적용증감률",
            "월시점 달성률",
            "최근3개월수량증감률",
            "수요증감률",
            "수요적용증감률",
            "평가월 수요진척률",
            "당월 출고진척률",
            "당월 재고충족률",
        }:
            cfg[col] = st.column_config.NumberColumn(
                s,
                format="%.2f%%",
                step=0.01,
            )
        elif (
            s in {"순번", "조회순번"}
            or s.endswith("건수")
            or "품목수" in s
            or "거래처수" in s
            or "매입처수" in s
            or "재고적용처수" in s
            or "금액" in s
            or "매출" in s
            or "공급가액" in s
            or "세액" in s
            or "가격" in s
        ):
            cfg[col] = st.column_config.NumberColumn(
                s,
                format="localized",
                step=1,
            )
        else:
            cfg[col] = st.column_config.NumberColumn(
                s,
                format="localized",
                step=0.01,
            )

    return cfg


def _render_fast_dataframe(
    df: pd.DataFrame,
    *,
    height: int = 520,
    action_name: str = "",
    meta: Dict[str, Any] | None = None,
) -> None:
    """빠른 표 렌더링."""
    view_df = normalize_display_df_for_streamlit(_fast_display_df(df))
    try:
        view_df, column_config, _table_width, table_height = build_sims_table_display_config(
            view_df,
            action_name=action_name,
            meta=meta or {},
            add_row_no=False,
            row_no_name="순번",
            enable_pinning=True,
            max_pinned_cols=5,
            min_width=720,
            max_width=2600,
            min_height=170,
            max_height=height,
            row_height=32,
        )
        log_sims_display_fields(df, view_df, action=action_name, render_path="panel", mode="fast")
        st.dataframe(
            view_df,
            width="stretch",
            hide_index=True,
            height=table_height,
            column_config=column_config if column_config else None,
        )
    except Exception:
        log.exception("[panel] fast common table render failed")
        cfg = _fast_column_config(view_df)
        st.dataframe(
            view_df,
            width="stretch",
            hide_index=True,
            height=height,
            column_config=cfg if cfg else None,
        )

def _prepare_io_table_df(df: pd.DataFrame) -> pd.DataFrame:
    out = _trim_object_columns(df)
    out = _make_styler_safe_df(out)
    out = _ensure_seq_col(out)
    return out

def _style_io_table(df: pd.DataFrame):
    safe_df = _make_styler_safe_df(df)
    safe_df = _normalize_zero_like_df(safe_df)
    return _build_io_display_styler(safe_df, add_row_no=False, band_size=5)


def _style_inventory_df(df: pd.DataFrame, payload: Dict[str, Any]):
    meta = payload.get("meta") or {}
    group_label = str(meta.get("group_label") or "").strip()

    safe_df = _make_styler_safe_df(df)
    safe_df = _normalize_zero_like_df(safe_df)

    styler = _build_io_display_styler(safe_df, add_row_no=False, band_size=5)

    total_style_df = pd.DataFrame("", index=safe_df.index, columns=safe_df.columns)

    if group_label and group_label in safe_df.columns:
        for idx in safe_df.index:
            try:
                if str(safe_df.loc[idx, group_label]).strip() == "합계":
                    total_style_df.loc[idx, :] = "background-color: #fff2cc; font-weight: 700;"
            except Exception:
                pass

    styler = styler.apply(lambda _: total_style_df, axis=None)

    styler = styler.set_table_styles([
        {
            "selector": "table",
            "props": [
                ("border-collapse", "collapse"),
                ("width", "100%"),
                ("font-size", "13px"),
            ],
        },
        {
            "selector": "th",
            "props": [
                ("text-align", "center"),
                ("font-weight", "700"),
                ("background-color", "#f5f6f7"),
                ("border", "1px solid #eceff3"),
                ("padding", "6px 8px"),
            ],
        },
        {
            "selector": "td",
            "props": [
                ("border", "1px solid #f1f3f5"),
                ("padding", "5px 8px"),
                ("white-space", "nowrap"),
            ],
        },
    ])

    try:
        styler = styler.hide(axis="index")
    except Exception:
        pass

    return styler

def _is_sales_trend_payload(payload: Dict[str, Any], action: str, title: str) -> bool:
    meta = payload.get("meta") or {}
    action_text = str(action or "").strip()
    title_text = str(title or "").strip()
    analysis_type = str(meta.get("analysis_type") or "").strip()
    summary_type = str(meta.get("summary_type") or "").strip()

    return (
        action_text in {
            "품목별 매출 추세 분석",
            "품목별 매출 추세 요약표",
            "품목별 매출 예상",
            "매출처별 매출 예상",
            "영업사원별 매출 예상",
            "지역별 매출 예상",
            "제약사별 매출 추세 분석",
            "제약사별 매출 추세 분석 요약표",
            "품목별 재고부족현황",
            "매입처별 재고부족 현황"
        }
        or title_text in {
            "품목별 매출 추세 분석",
            "품목별 매출 추세 요약표",
            "품목별 매출 예상",
            "매출처별 매출 예상",
            "영업사원별 매출 예상",
            "지역별 매출 예상",
            "제약사별 매출 추세 분석",
            "제약사별 매출 추세 분석 요약표",
            "품목별 재고부족현황",
            "매입처별 재고부족 현황"
        }
        or analysis_type in {"sales_trend", "sales_forecast", "customer_sales_forecast", "salesperson_sales_forecast", "region_sales_forecast", "stock_shortage", "supplier_stock_shortage", "manufacturer_sales_trend", "manufacturer_sales_trend_summary"}
        or summary_type in {"product_summary", "product_forecast", "customer_forecast", "salesperson_forecast", "region_forecast", "product_stock_shortage", "supplier_stock_shortage", "manufacturer_trend_detail", "manufacturer_trend_summary"}
    )

def _pinned_text_column_config(label: str, width: int):
    """
    Streamlit 버전에 따라 pinned 옵션이 없을 수 있으므로 안전하게 처리.
    pinned 지원 버전이면 좌측 고정, 미지원이면 일반 TextColumn으로 fallback.
    """
    try:
        return st.column_config.TextColumn(
            str(label or ""),
            width=width,
            pinned=True,
        )
    except TypeError:
        return st.column_config.TextColumn(
            str(label or ""),
            width=width,
        )

def _fast_width_px_for_large_table(df_src: pd.DataFrame, col_name: str) -> int:
    """
    큰 표 전용 빠른 컬럼 폭 계산.

    기존 _infer_width_px()는 컬럼마다 최대 200개 샘플을 읽어
    문자열 길이를 계산한다. 큰 표에서는 이 계산도 누적되므로,
    큰 표 모드에서는 컬럼명 기준 고정/간이 폭만 사용한다.
    """
    s = str(col_name or "").strip()

    fixed_width_map = {
        "순번": 60,
        "조회순번": 60,
        "기준월": 90,

        "제품코드": 95,
        "제품명": 260,
        "규격": 120,
        "제조사코드": 95,
        "제조사명": 160,
        "제품그룹명": 130,
        "제품구분명": 130,
        "제품분류명": 130,

        "거래처코드": 95,
        "거래처명": 220,
        "매입처코드": 95,
        "매입처명": 220,
        "재고적용처코드": 110,
        "재고적용처명": 220,

        "재고기준": 100,
        "추세판정": 110,
        "예상등급": 110,
        "부족등급": 130,
        "예상기준": 130,

        "현재재고수량": 130,
        "현재재고금액": 150,
        "장부재고평가단가": 150,
        "실재고평가단가": 150,
        "재고커버월수": 120,
    }

    if s in fixed_width_map:
        return fixed_width_map[s]

    if any(k in s for k in ("일자", "날짜")):
        return 105

    if any(k in s for k in ("일시", "시간")):
        return 150

    if _is_code_col_name(s):
        return 95

    try:
        if _is_numeric_col(df_src, col_name):
            if any(k in s for k in ("금액", "매출", "공급가액", "세액")):
                return 145
            if any(k in s for k in ("수량", "부족", "필요")):
                return 125
            if any(k in s for k in ("단가", "평균", "율", "커버월수")):
                return 120
            return 110
    except Exception:
        pass

    if any(k in s for k in ("명", "이름", "거래처", "제품", "제조사", "매입처")):
        return 180

    if any(k in s for k in ("주소", "비고", "설명", "메모", "기타")):
        return 260

    return 110

def _analytics_number_decimals(col: str) -> int | None:
    """
    분석/KPI 표 숫자 표시 자릿수 결정.
    None이면 숫자 포맷 대상 아님.
    """
    c = str(col or "").strip()
    c_lower = c.lower()
    if not c:
        return None

    int_cols = {
        "완료월총매출",
        "월평균매출",
        "완료월평균매출",
        "당월 현재매출",
        "당월 예상매출",
        "당월 잔여예상",
        "다음월예상매출",
        "3개월예상매출",
        "6개월예상매출",
        "부족예상금액",
        "완료월수",
        "매출발생월수",
        "매입처수",
        "총집계건수",
    }
    if c in int_cols:
        return 0

    decimal_cols = {
        "최근3개월평균매출",
        "최근6개월평균매출",
        "평균공급단가",
        "완료월평균출고수량",
        "최근3개월평균출고수량",
        "최근6개월평균출고수량",
        "당월 예상출고수량",
        "당월 잔여예상출고수량",
        "예상월말재고수량",
        "부족예상수량",
    }
    if c in decimal_cols:
        return 2

    percent_cols = {
        "당월 진척률",
        "평가월 진척률",
        "전월대비매출증감률",
        "최근3개월증감률",
        "월시점 최근3개월증감률",
        "적용증감률",
        "월시점 증감률",
        "월시점 적용증감률",
        "월시점 달성률",
        "최근3개월수량증감률",
        "수요증감률",
        "수요적용증감률",
        "평가월 수요진척률",
        "당월 출고진척률",
        "당월 재고충족률",
    }
    if c in percent_cols:
        return 2

    # 문자/코드성 컬럼 제외
    exclude_words = [
        "코드",
        "stock_cd",
        "buy_cd",
        "physic_cd",
        "ven_cd",
        "_cd",
        "명",
        "기준월",
        "재고기준",
        "자료원",
        "추세판정",
        "예상등급",
        "부족등급",
        "예상기준",
    ]
    if any(w in c or w in c_lower for w in exclude_words):
        return None

    # 소수 2자리 계열
    if any(w in c for w in ["커버월수", "증감률", "적용증감률", "율", "평균", "단가"]):
        return 2

    # 정수 표시 계열
    if any(
        w in c
        for w in [
            "수량",
            "금액",
            "매출",
            "공급가액",
            "세액",
            "부족",
            "필요",
            "건수",
            "품목수",
            "거래처수",
            "매입처수",
            "재고적용처수",
            "분석월수",
            "매출발생월수",
        ]
    ):
        return 0

    return None

# 분석/KPI 등급 컬럼 셀 스타일 결정 함수
# - 셀 값에 따라 글자색과 굵기를 반환한다.
# - 분석/KPI 표는 최종 표시 직전에 숫자 포맷을 문자열로 확정한다 (스타일링 경로에서 포맷이 안 먹는 경우가 있어서 우회).
def _grade_cell_style(value: Any) -> str:
    v = str(value or "").strip()

    if v in {"증가", "신규/증가", "상승예상"}:
        return "color:#166534; font-weight:800;"
    if v in {"감소", "감소예상"}:
        return "color:#9a3412; font-weight:800;"
    if v in {"안정", "안정예상"}:
        return "color:#1d4ed8; font-weight:800;"
    if v == "반품주의":
        return "color:#be123c; font-weight:800;"
    if v == "자료부족":
        return "color:#475569; font-weight:800;"
    if v == "신규확인":
        return "color:#6d28d9; font-weight:800;"
    if v in {"재고없음", "1개월내 부족"}:
        return "color:#be123c; font-weight:800;"
    if v in {"2개월내 부족주의", "3개월내 부족주의", "3개월내 부족"}:
        return "color:#9a3412; font-weight:800;"
    if v == "정상":
        return "color:#166534; font-weight:800;"
    if v in {"수요관찰", "재고없음/수요없음"}:
        return "color:#475569; font-weight:800;"
    
    return ""

# 분석/KPI 표 숫자 포맷팅 함수
# - 숫자 값에서 단위(원/개/건/개월/%)를 제거하고, 천 단위 구분 쉼표와 소수점 2자리까지 포맷팅한다.
# - 분석/KPI 표는 최종 표시 직전에 숫자 포맷을 문자열로 확정한다 (스타일링 경로에서 포맷이 안 먹는 경우가 있어서 우회).
def _parse_number_for_style(value: Any) -> float | None:
    s = str(value or "").strip()
    if not s:
        return None

    s = (
        s.replace(",", "")
        .replace("원", "")
        .replace("개", "")
        .replace("건", "")
        .replace("개월", "")
        .replace("%", "")
        .strip()
    )

    if not s:
        return None

    try:
        return float(s)
    except Exception:
        return None


# 분석/KPI 패널 표 스타일링 함수
# - 기본 숫자 포맷/우측 정렬/음수 붉은 글자는 공통 IO 스타일러 사용
# - 품목별 매출 추세 분석 원자료형은 제품코드 묶음별 배경색
# - 요약표/예상/재고부족현황은 5행 단위 배경색
# - 추세판정/예상등급/부족등급은 글자색만 추가 강조
# - 스타일 적용이 안 되는 경우를 대비해 최대한 안전하게 처리하되, 실패하면 스타일 없이 표만 보여준다.
# - 빈 데이터프레임인 경우 인덱스 숨기기 시도 후 스타일 없이 표 보여준다.
# - 이 스타일러는 분석/KPI 패널의 품목별 매출 추세 분석/요약표/예상/재고부족현황 표에 적용하는 것을 권장한다.
def _style_sales_trend_rows(df: pd.DataFrame):
    """
    분석/KPI 패널 표 스타일.

    - 기본 숫자 포맷/우측 정렬/음수 붉은 글자는 공통 IO 스타일러 사용
    - 품목별 매출 추세 분석 원자료형은 제품코드 묶음별 배경색
    - 요약표/예상/재고부족현황은 5행 단위 배경색
    - 추세판정/예상등급/부족등급은 글자색만 추가 강조
    """
    safe_df = _make_styler_safe_df(df)
    safe_df = _normalize_zero_like_df(safe_df)

    try:
        styler = _build_io_display_styler(safe_df, add_row_no=False, band_size=5)
    except Exception:
        styler = safe_df.style

    if safe_df.empty:
        try:
            styler = styler.hide(axis="index")
        except Exception:
            pass
        return styler

    style_df = pd.DataFrame("", index=safe_df.index, columns=safe_df.columns)

    has_product = "제품코드" in safe_df.columns

    if has_product:
        product_keys = (
            safe_df["제품코드"]
            .fillna("")
            .astype(str)
            .str.strip()
            .replace("", pd.NA)
            .ffill()
            .fillna("")
        )
    else:
        product_keys = pd.Series([""] * len(safe_df), index=safe_df.index)

    is_summary_or_forecast = (
        ("예상등급" in safe_df.columns)
        or ("부족등급" in safe_df.columns)
        or ("추세판정" in safe_df.columns and "기준월" not in safe_df.columns)
    )

    last_product = None
    group_no = -1

    for row_pos, idx in enumerate(safe_df.index):
        border_style = ""

        product_cd = str(product_keys.loc[idx] or "").strip()

        if has_product and product_cd != last_product:
            group_no += 1
            last_product = product_cd

            # 원자료형 추세 분석은 제품이 바뀌는 첫 줄에 경계선
            if not is_summary_or_forecast:
                border_style = "border-top: 2px solid #d7e3f2;"

        if is_summary_or_forecast:
            # 요약표/예상/재고부족현황은 5행 단위로 배경색 변경
            bg = "#ffffff" if (row_pos // 5) % 2 == 0 else "#f8fbff"

            if row_pos > 0 and row_pos % 5 == 0:
                border_style = f"{border_style} border-top: 2px solid #cbd5e1;"
        else:
            # 품목별 매출 추세 분석 원자료형은 제품 묶음 기준 배경색
            bg = "#f7fbff" if group_no % 2 == 0 else "#ffffff"

        style_df.loc[idx, :] = f"background-color: {bg}; {border_style}"

    # 음수 금액/수량은 기존 조회표와 맞춰 글자만 붉게 표시
    for c in safe_df.columns:
        if _analytics_number_decimals(str(c)) is None:
            continue

        try:
            nums = safe_df[c].map(_parse_number_for_style)
            nums = pd.to_numeric(nums, errors="coerce")
            neg_mask = nums < 0

            style_df.loc[neg_mask, c] = (
                style_df.loc[neg_mask, c].astype(str)
                + " color:#be123c; font-weight:800;"
            )
        except Exception:
            pass

    # 추세판정 / 예상등급 / 부족등급은 글자색만 강조
    for grade_col in ["추세판정", "예상등급", "부족등급"]:
        if grade_col in safe_df.columns:
            for idx in safe_df.index:
                extra = _grade_cell_style(safe_df.loc[idx, grade_col])
                if extra:
                    style_df.loc[idx, grade_col] = (
                        str(style_df.loc[idx, grade_col] or "")
                        + " "
                        + extra
                    )

    try:
        styler = styler.apply(lambda _: style_df, axis=None)
    except Exception:
        styler = safe_df.style.apply(lambda _: style_df, axis=None)

    try:
        styler = styler.hide(axis="index")
    except Exception:
        pass

    return styler

def _is_sales_trend_summary_payload(payload: Dict[str, Any], action: str, title: str) -> bool:
    meta = payload.get("meta") or {}
    action_text = str(action or "").strip()
    title_text = str(title or "").strip()
    summary_type = str(meta.get("summary_type") or "").strip()
    analysis_type = str(meta.get("analysis_type") or "").strip()

    return (
        action_text in {
            "품목별 매출 추세 요약표",
            "품목별 매출 예상",
            "매출처별 매출 예상",
            "영업사원별 매출 예상",
            "지역별 매출 예상",
            "제약사별 매출 추세 분석",
            "제약사별 매출 추세 분석 요약표",
            "품목별 재고부족현황",
        }
        or title_text in {"품목별 매출 추세 요약표", "품목별 매출 예상", "매출처별 매출 예상", "영업사원별 매출 예상", "지역별 매출 예상", "제약사별 매출 추세 분석", "제약사별 매출 추세 분석 요약표", "품목별 재고부족현황", "매입처별 재고부족 현황"}
        or summary_type in {"product_summary", "product_forecast", "customer_forecast", "salesperson_forecast", "region_forecast", "product_stock_shortage", "supplier_stock_shortage", "manufacturer_trend_detail", "manufacturer_trend_summary"}
        or analysis_type in {"sales_forecast", "customer_sales_forecast", "salesperson_sales_forecast", "region_sales_forecast", "stock_shortage", "supplier_stock_shortage", "manufacturer_sales_trend", "manufacturer_sales_trend_summary"}
    )

def _blank_same_as_previous(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    """
    화면 표시용 반복값 공백 처리.
    원본 데이터는 건드리지 않고 화면용 df에서만 처리한다.
    """
    out = df.copy()

    for col in cols:
        if col not in out.columns:
            continue

        prev = None
        for idx in out.index:
            cur = out.loc[idx, col]

            cur_text = str(cur or "").strip()
            prev_text = str(prev or "").strip()

            if idx != out.index[0] and cur_text == prev_text:
                out.loc[idx, col] = ""
            else:
                prev = cur

    return out


# 날짜/월 표시 보정: 숫자만 추출해서 YYYY-MM 또는 YYYY-MM-DD 형태로 보이도록 한다. 
def _fmt_yyyymm_display(value: Any) -> str:
    s = "".join(ch for ch in str(value or "") if ch.isdigit())
    if len(s) >= 6:
        return f"{s[:4]}-{s[4:6]}"
    return str(value or "").strip()

# 날짜 표시 보정: 숫자만 추출해서 YYYY-MM-DD 형태로 보이도록 한다.
def _prepare_sales_trend_display_df(
    df: pd.DataFrame,
    payload: Dict[str, Any],
    action: str,
    title: str,
) -> pd.DataFrame:
    """
    품목별 매출 추세 분석/요약표 표시 보정.

    - 상세 추세 분석: 제품별 순번, 반복 제품정보 공백 처리
    - 요약표: 행 단위 순번
    """
    out = df.copy()

    if out.empty:
        return out

    if "기준월" in out.columns:
        out["기준월"] = out["기준월"].map(_fmt_yyyymm_display)

    is_summary = _is_sales_trend_summary_payload(payload, action, title)

    # 기존 순번이 있으면 제거 후 다시 만든다.
    if "순번" in out.columns:
        out = out.drop(columns=["순번"])

    if is_summary:
        # 요약표는 1줄 단위 순번
        out.insert(0, "순번", range(1, len(out) + 1))
        return out

    # 상세 추세 분석은 제품별 순번
    if "제품코드" in out.columns:
        product_series = (
            out["제품코드"]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        group_no = []
        last_product = object()
        n = 0

        for product_cd in product_series.tolist():
            if product_cd != last_product:
                n += 1
                last_product = product_cd
            group_no.append(n)

        out.insert(0, "순번", group_no)

        # 같은 제품 정보는 첫 줄만 표시
        blank_cols = [
            "순번",
            "제품코드",
            "제품명",
            "규격",
            "제조사코드",
            "제조사명",
            "제품그룹명",
            "제품구분명",
            "제품분류명",
        ]

        out = _blank_same_as_previous(out, blank_cols)

        # 같은 제품 안에서 같은 기준월이 반복되면 기준월도 공백 처리
        if "기준월" in out.columns:
            product_key = product_series.replace("", pd.NA).ffill().fillna("")
            prev_key = None
            prev_month = None

            for idx in out.index:
                key = str(product_key.loc[idx] or "").strip()
                month = str(out.loc[idx, "기준월"] or "").strip()

                if key == prev_key and month == prev_month:
                    out.loc[idx, "기준월"] = ""
                else:
                    prev_key = key
                    prev_month = month

        return out

    # 제품코드가 없으면 일반 행 순번
    out.insert(0, "순번", range(1, len(out) + 1))
    return out

# ==========================================================
# 🧩 외부 모듈
#   - 뷰는 '표시 + payload 반환'만 담당 (컨텍스트 푸시는 여기서)
# ==========================================================
from app.sims.views import users, codes, vendors, goods, road_address, analytics_views, dashboard_lite, rddbc_io_views
from app.sims.views.rddbc_io_shared import (
    _build_io_display_styler,
    _prepare_io_display_df,
)

from app.ui.chat_middleware import (
    render_sims_context_controls,
    push_sims_result_to_chat,
    get_current_chat_room_id,
    _build_sims_detail_analysis_prompt,
    _expected_analysis_row_count,
    _get_full_download_df_for_sims_item,
    _sims_clicked_llm_context_mismatch,
)
from app.ui.sims_table_display import (
    build_sims_table_display_config,
    normalize_display_df_for_streamlit,
)

from app.ui.ssai_login import require_permission


# 카테고리/액션 레지스트리
_CATEGORIES: Dict[str, Dict[str, Any]] = {
    "사용자": {
        "actions": {
            "사용자목록 + 부서명": users.render_user_list_with_dept,
            "부서별 사용자 수": users.render_user_count_by_dept,
            "부서별사용자수": users.render_user_count_by_dept,  # 별칭
            "최근 입사자": users.render_recent_hires,
        }
    },
    "코드마스터": {
        "actions": {
            "그룹코드조회": codes.render_codes_by_group,
            "그룹별 코드 조회": codes.render_codes_by_group,  # 별칭
            "코드명 검색": codes.render_search_codes,
        }
    },
    "거래처": {
        "actions": {
            "거래처 목록": vendors.render_vendor_list,
            "거래처 상세": vendors.render_vendor_detail,
        }
    },
    "도로명주소": {
        "actions": {
            "도로명주소 조회": road_address.render_road_address_list,
        }
    },
    "제품": {
        "actions": {
            "제품코드 목록": goods.view_goods_list,
            "제품코드 상세": goods.view_goods_detail,
#            "제품코드목록": rddbc_io_views.view_rddbc040,   # 별칭(붙여쓰기)
            "제품코드목록": goods.view_goods_list,   # 별칭(붙여쓰기)
            "제품코드상세": goods.view_goods_detail, # 별칭(붙여쓰기)
        }
    },
    "분석/KPI": {
        "actions": {
            "Dashboard Lite v0.1": dashboard_lite.render_dashboard_lite,
            "제약사별 매출 추세 분석": analytics_views.render_manufacturer_sales_trend_analysis,
            "제약사별 매출 추세 분석 요약표": analytics_views.render_manufacturer_sales_trend_summary_analysis,
            "품목별 매출 추세 분석": analytics_views.render_sales_trend_analysis,
            "품목별 매출 추세 요약표": analytics_views.render_sales_trend_summary_analysis,
            "품목별 매출 예상": analytics_views.render_sales_forecast_analysis,
            "매출처별 매출 예상": analytics_views.render_customer_sales_forecast_analysis,
            "영업사원별 매출 예상": analytics_views.render_salesperson_sales_forecast_analysis,
            "지역별 매출 예상": analytics_views.render_region_sales_forecast_analysis,
            "품목별 재고부족현황": analytics_views.render_stock_shortage_analysis,
            "매입처별 재고부족 현황": analytics_views.render_supplier_stock_shortage_analysis,
        }
    },
    "입출고/명세서/재고": {
        "actions": {
            "입고명세 조회": rddbc_io_views.view_rddbc110,
            "출고명세 조회": rddbc_io_views.view_rddbc120,
            "거래명세서 공통 조회": rddbc_io_views.view_rddbc130,
            "세금계산서 공통 조회": rddbc_io_views.view_rddbc140,
            "실재고월집계 조회": rddbc_io_views.view_rddbc210,
            "장부재고월집계 조회": rddbc_io_views.view_rddbc220,
            "입고↔거래명세서 검증": rddbc_io_views.view_rddbc110_trans_check,
            "입고↔세금계산서 검증": rddbc_io_views.view_rddbc110_tax_check,
            "출고↔거래명세서 검증": rddbc_io_views.view_rddbc120_trans_check,
            "출고↔세금계산서 검증": rddbc_io_views.view_rddbc120_tax_check,
            "제품수불현황 조회": rddbc_io_views.view_product_flow,
            "제품수불현황": rddbc_io_views.view_product_flow,
            "제품재고현황 조회": rddbc_io_views.view_product_inventory,
            "제품재고현황": rddbc_io_views.view_product_inventory,
        }
    },
}

# ==========================================================
# 🧩 상태/초기화 유틸
# ==========================================================
def _ensure_sims_state() -> None:
    ss = st.session_state
    # __sims_open 기본값은 메인에서와 동일하게 False로 맞춘다.
    # (실제 토글 상태는 메인 UI에서만 관리)
    ss.setdefault("__sims_open", False)
    ss.setdefault("__sims_panel_active", False)
    ss.setdefault("__sims_run_flag", False)
    ss.setdefault("__sims_was_final", False)
    ss.setdefault("__sims_form_id", 0)
    ss.setdefault("__sims_run_seq", 0)
    ss.setdefault("__sims_last_action", "")
    ss.setdefault("__sims_selected", None)
    ss.setdefault("__sims_widget_ns", None)
    ss.setdefault("__sims_form_submitted", False)
    ss.setdefault("__sims_submitted_form_id", None)
    ss.setdefault("__sims_prev_widget_keys", set())


def _log_state(tag: str) -> None:
    try:
        ss = st.session_state
        log.info(
            "[panel.state:%s] open=%s, panel_active=%s, run_flag=%s, was_final=%s, "
            "form_id=%s, run_seq=%s, last_action=%r, selected=%r, widget_ns=%r",
            tag,
            ss.get("__sims_open"),
            ss.get("__sims_panel_active"),
            ss.get("__sims_run_flag"),
            ss.get("__sims_was_final"),
            ss.get("__sims_form_id"),
            ss.get("__sims_run_seq"),
            ss.get("__sims_last_action"),
            ss.get("__sims_selected"),
            ss.get("__sims_widget_ns"),
        )
    except Exception:
        pass

def _clear_widget_keys_for(action: str, form_id: int) -> None:
    """액션 변경 시, 이전 액션에서 사용한 위젯/결과 키 정리"""
    try:
        keys = list(st.session_state.keys())

        for k in keys:
            if (
                k.startswith("__users_")
                or k.startswith("__codes_")
                or k.startswith("__ven_")

                or k.startswith("__vendors_")
                or k.startswith("__road_addr_")
                or k.startswith("__analytics_")
                or k.startswith("__goods_")

                or k.startswith("__io")      # "__io110_last_payload"까지 지우기 위해 "__io_" -> "__io" 로 수정
                or k.startswith("__product_flow")
                or k.startswith("__rddbc_")
                or k.startswith("__m040_")
            ):
                del st.session_state[k]

        for extra in [
            "__sims_ctx_reset_btn",
            "__sims_last_push",
            "__sims_last_render_run_seq",
            "__sims_was_final",
            "__sims_run_flag",
            "__sims_inner_submit",
            "__sims_context",
            "__sims_context_text",
            "__sims_ctx_dirty",
        ]:
            st.session_state.pop(extra, None)

        log.info("[panel.main] cleared widget/render keys on action-change; form_id=%s", form_id)
    except Exception:
        log.exception("[panel.main] clear widget keys failed")     

def _current_action_payload_key(action: str) -> Optional[str]:
    return {
        "입고명세 조회": "__io110_last_payload",
        "출고명세 조회": "__io120_last_payload",
        "거래명세서 공통 조회": "__io130_last_payload",
        "세금계산서 공통 조회": "__io140_last_payload",
        "실재고월집계 조회": "__io210_last_payload",
        "장부재고월집계 조회": "__io220_last_payload",
        "입고↔거래명세서 검증": "__io110_tc_last_payload",
        "입고↔세금계산서 검증": "__io110_tx_last_payload",
        "출고↔거래명세서 검증": "__io120_tc_last_payload",
        "출고↔세금계산서 검증": "__io120_tx_last_payload",
        "제품수불현황 조회": "__product_flow_last_payload",
        "제품수불현황": "__product_flow_last_payload",
        "제품재고현황 조회": "__io260_last_payload",
        "제품재고현황": "__io260_last_payload",
    }.get(str(action or "").strip())


def _clear_current_action_payload(action: str) -> None:
    key = _current_action_payload_key(action)
    if key:
        st.session_state.pop(key, None)


def _panel_current_company_stamp() -> Dict[str, str]:
    """현재 선택 회사 식별자를 payload/meta에 기록하기 위한 stamp."""
    try:
        from app.ui.ssai_login import get_selected_company
        company = get_selected_company()
    except Exception:
        company = None

    if not isinstance(company, dict):
        return {"company_id": "", "db_name": "", "company_name": ""}

    return {
        "company_id": str(company.get("company_id") or "").strip(),
        "db_name": str(company.get("db_name") or "").strip(),
        "company_name": str(company.get("company_name") or "").strip(),
    }


def _panel_payload_company_stamp(payload: Dict[str, Any] | None) -> Dict[str, str]:
    """payload/meta에 기록된 회사 stamp 추출."""
    if not isinstance(payload, dict):
        return {"company_id": "", "db_name": "", "company_name": ""}

    try:
        meta = payload.get("meta") if isinstance(payload.get("meta"), dict) else {}
        return {
            "company_id": str(
                meta.get("_ssai_company_id")
                or payload.get("_ssai_company_id")
                or ""
            ).strip(),
            "db_name": str(
                meta.get("_ssai_db_name")
                or payload.get("_ssai_db_name")
                or ""
            ).strip(),
            "company_name": str(
                meta.get("_ssai_company_name")
                or payload.get("_ssai_company_name")
                or ""
            ).strip(),
        }
    except Exception:
        return {"company_id": "", "db_name": "", "company_name": ""}


def _panel_stamp_payload_company(payload: Dict[str, Any]) -> None:
    """패널 최종 payload에 현재 회사 식별자를 기록한다."""
    if not isinstance(payload, dict):
        return

    stamp = _panel_current_company_stamp()
    meta = payload.setdefault("meta", {})

    if not isinstance(meta, dict):
        meta = {}
        payload["meta"] = meta

    if stamp.get("company_id"):
        meta["_ssai_company_id"] = stamp["company_id"]
        payload["_ssai_company_id"] = stamp["company_id"]

    if stamp.get("db_name"):
        meta["_ssai_db_name"] = stamp["db_name"]
        payload["_ssai_db_name"] = stamp["db_name"]

    if stamp.get("company_name"):
        meta["_ssai_company_name"] = stamp["company_name"]
        payload["_ssai_company_name"] = stamp["company_name"]


def _panel_payload_matches_current_company(payload: Dict[str, Any] | None) -> bool:
    """
    payload가 현재 선택 회사와 일치하는지 검사한다.

    - stamp가 없는 legacy payload는 여기서 막지 않는다.
      회사 변경 clear에서 legacy payload/widget cache를 제거한다.
    - stamp가 있는 payload가 현재 회사와 다르면 stash/push/current 승격을 차단한다.
    """
    stamp = _panel_payload_company_stamp(payload)
    payload_company_id = stamp.get("company_id") or ""
    payload_db_name = stamp.get("db_name") or ""

    if not payload_company_id and not payload_db_name:
        return True

    current = _panel_current_company_stamp()
    current_company_id = current.get("company_id") or ""
    current_db_name = current.get("db_name") or ""

    if payload_company_id and current_company_id and payload_company_id != current_company_id:
        return False

    if payload_db_name and current_db_name and payload_db_name != current_db_name:
        return False

    return True


def _panel_stale_payload_log_state(
    payload_stamp: Mapping[str, Any] | None,
    current_stamp: Mapping[str, Any] | None,
) -> dict[str, Any]:
    """Return only non-identifying fields for stale-payload log records."""
    payload = dict(payload_stamp or {})
    current = dict(current_stamp or {})
    payload_db_name = payload.get("db_name") or ""
    current_db_name = current.get("db_name") or ""
    return {
        "payload_company_id": payload.get("company_id") or "",
        "current_company_id": current.get("company_id") or "",
        "db_mismatch": bool(payload_db_name and current_db_name and payload_db_name != current_db_name),
    }


def _dashboard_chat_push_scope(company_id: str, room_id: str) -> str:
    """Build the session-local dedupe scope for one company and chat room."""
    return f"{str(company_id or '').strip()}::{str(room_id or '').strip()}"


def _dashboard_chat_push_is_duplicate(
    session_state: Mapping[str, Any],
    *,
    company_id: str,
    room_id: str,
    signature: str,
) -> bool:
    if not signature or not company_id or not room_id:
        return False
    seen = session_state.get("__dashboard_lite_chat_push_sigs") or {}
    if not isinstance(seen, Mapping):
        return False
    return str(seen.get(_dashboard_chat_push_scope(company_id, room_id)) or "") == signature


def _remember_dashboard_chat_push_signature(
    session_state: Dict[str, Any],
    *,
    company_id: str,
    room_id: str,
    signature: str,
) -> None:
    if not signature or not company_id or not room_id:
        return
    seen = session_state.get("__dashboard_lite_chat_push_sigs")
    if not isinstance(seen, dict):
        seen = {}
        session_state["__dashboard_lite_chat_push_sigs"] = seen
    seen[_dashboard_chat_push_scope(company_id, room_id)] = signature


def _dashboard_event_id_for_push(
    payload: dict[str, Any],
    meta: dict[str, Any],
    cache: dict[str, Any],
) -> str:
    """Reuse the pre-push Dashboard event; generate only for malformed legacy input."""
    existing = str(
        payload.get("id")
        or meta.get("dashboard_event_id")
        or cache.get("dashboard_event_id")
        or ""
    ).strip()
    return existing or str(uuid.uuid4())


def _panel_action_key(category: str, action: str) -> str:
    return f"{str(category or '').strip()}::{str(action or '').strip()}"


def _remember_panel_final_payload(payload: Dict[str, Any], category: str, action: str) -> None:
    """
    패널 최종 조회 결과를 세션에 보관한다.

    이유:
    - [LLM 분석] 버튼 클릭은 rerun을 만든다.
    - rerun 때 view 함수가 final=False payload를 반환하면
      기존 결과/버튼이 다시 렌더되지 않아 버튼 클릭 이벤트가 소비되지 않는다.
    """
    try:
        if not isinstance(payload, dict):
            return
        if not bool(payload.get("final")):
            return

        _panel_stamp_payload_company(payload)

        key = _panel_action_key(category, action)
        st.session_state["__sims_panel_last_final_action"] = key
        st.session_state["__sims_panel_last_final_payload"] = payload
    except Exception:
        log.exception("[panel] remember last final payload failed")


def _get_panel_last_final_payload(category: str, action: str) -> Optional[Dict[str, Any]]:
    """
    현재 action과 같은 마지막 최종 결과를 가져온다.
    """
    try:
        key = _panel_action_key(category, action)
        saved_key = st.session_state.get("__sims_panel_last_final_action")
        saved_payload = st.session_state.get("__sims_panel_last_final_payload")

        if saved_key == key and isinstance(saved_payload, dict):
            return saved_payload
    except Exception:
        log.exception("[panel] get last final payload failed")

    return None


def _clear_panel_last_final_payload() -> None:
    st.session_state.pop("__sims_panel_last_final_action", None)
    st.session_state.pop("__sims_panel_last_final_payload", None)



def _panel_bool_env(name: str, default: bool = True) -> bool:
    """패널 성능 옵션용 bool 환경변수 판정."""
    try:
        raw = str(os.getenv(name, "")).strip().lower()
        if raw in {"1", "true", "yes", "y", "on"}:
            return True
        if raw in {"0", "false", "no", "n", "off"}:
            return False
    except Exception:
        pass
    return bool(default)

def _panel_int_env(name: str, default: int = 0) -> int:
    """패널 성능 옵션용 int 환경변수 판정."""
    try:
        raw = str(os.getenv(name, "")).strip()
        if raw:
            return int(raw)
    except Exception:
        pass
    return int(default)


def _panel_result_target_chat_enabled() -> bool:
    """
    SIMS 패널 조회 결과 표시 정책.

    SIMS_PANEL_RESULT_TARGET=chat 이면:
    - 패널에는 표를 그리지 않는다.
    - 현재표 컨텍스트만 갱신한다.
    - 메인에서 채팅방으로 1회 push한다.
    """
    target = str(os.getenv("SIMS_PANEL_RESULT_TARGET", "chat") or "chat").strip().lower()
    return target in {"chat", "chat_only", "chat-only", "1", "true", "yes", "y", "on"}


def _store_panel_final_payload_for_chat(payload: Dict[str, Any], action: str) -> None:
    """
    메인에서 채팅방으로 push할 최종 패널 payload를 보관한다.

    중요:
    - df는 전체 다운로드/현재표 후속분석 기준
    - df_display/records는 화면 표시 제한 기준
    - meta.summary_md/query_summary/condition은 그대로 유지
    """
    try:
        if not isinstance(payload, dict):
            return

        _panel_stamp_payload_company(payload)
        try:
            meta = payload.setdefault("meta", {})
            if isinstance(meta, dict):
                meta["_panel_source_sig"] = _make_panel_source_sig(action, payload)
        except Exception:
            pass

        st.session_state["__sims_last_final_payload_for_chat"] = payload
        st.session_state["__sims_last_final_payload_for_chat_action"] = str(action or payload.get("action") or "")
    except Exception:
        log.exception("[panel] store final payload for chat failed")


def _panel_chat_push_already_consumed(panel_source_sig: str) -> bool:
    try:
        sig = str(panel_source_sig or "").strip()
        return bool(sig and st.session_state.get("__sims_panel_chat_pushed_source_sig") == sig)
    except Exception:
        return False


def _render_panel_chat_only_done(payload: Dict[str, Any], action: str) -> None:
    """
    채팅창 표시 정책일 때 패널에 남길 최소 안내.
    표/요약/다운로드는 패널에서 렌더링하지 않는다.

    주의:
    - row_count_total 이 항상 DB 전체 건수를 뜻하지는 않는다.
    - 제품코드 목록처럼 서비스에서 MAX_ROWS 만큼만 가져온 경우,
      row_count_total/download_row_count는 "이번 조회로 적재된 건수"일 수 있다.
    """
    try:
        meta = payload.get("meta") or {}

        # DB 조건 전체 건수. 제품 서비스처럼 별도 total을 넘겨줄 때만 사용한다.
        db_total_rows = int(
            meta.get("db_total_count")
            or meta.get("total_count")
            or meta.get("matched_row_count")
            or meta.get("matched_total_count")
            or 0
        )

        df_full = payload.get("df")
        if isinstance(df_full, pd.DataFrame):
            loaded_rows = int(len(df_full))
        else:
            loaded_rows = int(
                meta.get("download_row_count")
                or meta.get("row_count_loaded")
                or meta.get("row_count_total")
                or meta.get("row_count")
                or 0
            )

        display_rows_raw = meta.get("display_row_count")
        if display_rows_raw is None and isinstance(payload.get("df_display"), pd.DataFrame):
            display_rows_raw = len(payload.get("df_display"))
        display_rows = int(display_rows_raw or 0)

    except Exception:
        db_total_rows = 0
        loaded_rows = 0
        display_rows = 0

    if db_total_rows and loaded_rows and db_total_rows > loaded_rows:
        st.success(
            f"조회 완료: 조건 전체 {db_total_rows:,}건 중 {loaded_rows:,}건을 조회했습니다. "
            f"채팅창에는 {display_rows:,}건 표시했습니다. "
            f"다운로드와 현재표 후속분석은 이번 조회 {loaded_rows:,}건 기준입니다."
        )
    elif loaded_rows and display_rows and loaded_rows > display_rows:
        st.success(
            f"조회 완료: {loaded_rows:,}건을 조회했습니다. "
            f"채팅창에는 {display_rows:,}건 표시했습니다. "
            f"다운로드와 현재표 후속분석은 조회된 {loaded_rows:,}건 기준입니다."
        )
    elif loaded_rows:
        st.success(f"조회 완료: {loaded_rows:,}건. 결과는 위 채팅창에 저장되었습니다.")
    else:
        st.info("해당 조회조건의 자료가 없습니다.")

    st.caption("다른 조건으로 다시 조회하려면 아래 조회조건을 수정한 뒤 조회 버튼을 다시 누르세요.")

def _panel_compact_on_chat_rerun_enabled() -> bool:
    """
    채팅 입력/일반 rerun 때 패널 결과표 full render를 생략할지 여부.

    기본값 True:
    - 패널 조회 직후 1회는 full render
    - 이후 채팅의 '현재표 ...' 후속분석 rerun에서는 패널 표/다운로드 재렌더를 접는다.

    필요 시 .env:
      SIMS_PANEL_COMPACT_ON_CHAT_RERUN=0
    """
    return _panel_bool_env("SIMS_PANEL_COMPACT_ON_CHAT_RERUN", True)


def _panel_query_fingerprint(payload: Mapping[str, Any] | None) -> str:
    """Return a deterministic fingerprint for actual submitted query conditions only."""
    if not isinstance(payload, Mapping):
        return ""
    meta = payload.get("meta") if isinstance(payload.get("meta"), Mapping) else {}
    material = {
        "condition": payload.get("condition") or meta.get("condition"),
        "query_summary": meta.get("query_summary"),
        "source_query": meta.get("source_query"),
        "params": meta.get("params") or payload.get("params"),
        "filters": meta.get("filters"),
    }
    try:
        encoded = json.dumps(material, ensure_ascii=False, sort_keys=True, default=str)
    except Exception:
        encoded = str(material)
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()[:16]


def _make_panel_source_sig(action: str, payload: Mapping[str, Any] | None = None) -> str:
    """Identify one explicit panel query execution without treating reruns as new results."""
    try:
        ss = st.session_state
        sel = ss.get("__sims_selected") or {}
        action_name = str(sel.get("action") or action)
        fingerprint = _panel_query_fingerprint(payload)
        if fingerprint:
            ss[f"__sims_panel_query_fingerprint::{action_name}"] = fingerprint
        else:
            fingerprint = str(ss.get(f"__sims_panel_query_fingerprint::{action_name}") or "")
        return "{}::{}::{}::{}::{}".format(
            ss.get("__sims_run_seq"),
            ss.get("__sims_query_submit_seq", 0),
            str(sel.get("category") or ""),
            action_name,
            fingerprint or "no-condition",
        )
    except Exception:
        return str(action or "")


def _panel_force_render_key(panel_source_sig: str) -> str:
    safe = _safe_filename(panel_source_sig or "panel")
    return f"__sims_panel_force_render_once::{safe}"


def _should_compact_panel_result_on_rerun(payload: Dict[str, Any], action: str) -> bool:
    """
    패널 결과표 compact 렌더 여부.

    조건:
    - final payload
    - 같은 run_seq/action의 패널 결과가 이미 current source로 승격됨
    - 사용자가 [패널 표 다시 표시]를 누른 1회가 아님

    목적:
    채팅 입력 때마다 패널 view 결과표와 다운로드 bytes가 반복 렌더되는 것을 막는다.
    """
    if not _panel_compact_on_chat_rerun_enabled():
        return False

    if not isinstance(payload, dict) or not bool(payload.get("final")):
        return False

    panel_source_sig = _make_panel_source_sig(action)
    if not panel_source_sig:
        return False

    try:
        if st.session_state.get("__sims_panel_source_promoted_sig") != panel_source_sig:
            # 아직 이 패널 조회 결과를 full render/source 승격한 적이 없다.
            return False

        force_key = _panel_force_render_key(panel_source_sig)
        if bool(st.session_state.pop(force_key, False)):
            # 사용자가 명시적으로 다시 표시를 눌렀으므로 이번 1회는 full render 허용
            return False

        return True
    except Exception:
        return False


def _render_compact_panel_result_placeholder(
    *,
    payload: Dict[str, Any],
    action: str,
    title: str,
    df_disp: Optional[pd.DataFrame],
    df_full: Optional[pd.DataFrame],
) -> None:
    """
    채팅 rerun 때 패널 결과표 대신 가벼운 접힘 표시만 렌더한다.

    현재표 후속분석 source와 다운로드용 원본 df는 session_state에 이미 유지되어 있으므로,
    화면에서 패널 표를 다시 그리지 않아도 후속분석은 계속 정상 동작한다.
    """
    st.subheader(title)

    meta = payload.get("meta") or {}
    query_summary = str(meta.get("query_summary") or "").strip()
    if not query_summary:
        params = payload.get("params") or {}
        parts = []
        for k, v in params.items():
            text = str(v or "").strip()
            if text:
                parts.append(f"{k} {text}")
        query_summary = " / ".join(parts)

    if query_summary:
        st.caption(f"조회조건: {query_summary}")

    display_rows = int(len(df_disp)) if isinstance(df_disp, pd.DataFrame) else 0
    full_rows = int(len(df_full)) if isinstance(df_full, pd.DataFrame) else display_rows
    display_cols = int(len(df_disp.columns)) if isinstance(df_disp, pd.DataFrame) else 0
    full_cols = int(len(df_full.columns)) if isinstance(df_full, pd.DataFrame) else display_cols

    if full_rows > display_rows:
        count_text = f"전체 {full_rows:,}건 중 화면 {display_rows:,}건"
    else:
        count_text = f"{display_rows:,}건"

    st.caption(
        f"패널 결과 접힘: {count_text} / {max(full_cols, display_cols):,}열 · "
        "현재표 후속분석 기준 데이터는 유지됩니다."
    )

    c1, c2, c3 = st.columns(3)
    panel_source_sig = _make_panel_source_sig(action)
    force_key = _panel_force_render_key(panel_source_sig)
    button_key_base = _safe_filename(panel_source_sig or str(action or title or "panel"))

    with c1:
        if st.button(
            "📊 패널 표 다시 표시",
            key=f"__panel_show_full_again_{button_key_base}",
            width="stretch",
        ):
            st.session_state[force_key] = True
            st.rerun()

    with c2:
        run_llm = st.button(
            "🤖 LLM 분석",
            key=f"__panel_compact_llm_{button_key_base}",
            width="stretch",
        )

    with c3:
        st.caption("다운로드는 표 다시 표시 후 사용")

    if run_llm:
        try:
            prompt = _build_sims_detail_analysis_prompt(
                action_name=str(action or title or ""),
                display_rows=display_rows,
                download_rows=full_rows,
                expected_rows=full_rows,
            )
        except Exception:
            prompt = (
                f"현재 SIMS 패널 조회 결과({title})를 전체 조회조건 기준으로 분석해줘. "
                "핵심 요약, 주요 수치, 주의/확인할 점, 다음 조회 제안 순서로 정리해줘. "
                "내부 영문 key 이름은 답변에 노출하지 마라."
            )

        _run_panel_llm_analysis_from_button(prompt, button_key_base)

    log.info(
        "[panel] compact result render skipped action=%s display_rows=%s full_rows=%s sig=%s",
        action,
        display_rows,
        full_rows,
        panel_source_sig,
    )


def _panel_skip_view_on_chat_rerun_enabled() -> bool:
    """
    채팅 입력으로 인한 rerun 때 패널 view_func() 자체를 건너뛸지 여부.

    기본값 True:
    - 패널 최종 결과가 이미 있고 current source 승격이 끝난 경우
    - 채팅 입력/후속분석 rerun에서는 view 함수 호출을 생략하고 compact placeholder만 표시

    필요 시 .env:
      SIMS_PANEL_SKIP_VIEW_ON_CHAT_RERUN=0
    """
    return _panel_bool_env("SIMS_PANEL_SKIP_VIEW_ON_CHAT_RERUN", True)


def _pop_panel_skip_view_payload_for_chat_rerun(
    *,
    category: str,
    action: str,
    opened_by_run: bool,
    inner_submit: bool,
) -> Optional[Dict[str, Any]]:
    """
    채팅 입력 rerun 전용 view_func() skip 판단.

    메인에서 채팅 입력을 감지하면 __sims_panel_skip_view_once=True 를 세운다.
    이 함수는 그 1회성 신호를 소비하고, 같은 action의 마지막 final payload가
    재사용 가능할 때만 반환한다.

    반환된 payload는 _render_payload()에서 기존 compact placeholder 경로를 탄다.
    """
    ss = st.session_state

    if not _panel_skip_view_on_chat_rerun_enabled():
        return None

    skip_requested = bool(ss.get("__sims_panel_skip_view_once"))
    if not skip_requested:
        return None

    # 이 신호는 채팅 입력 1회용이다.
    # 조회 버튼/패널 열기 렌더에서 남아 있으면 오동작하므로 여기서 소비한다.
    ss.pop("__sims_panel_skip_view_once", None)

    if opened_by_run or inner_submit:
        return None

    payload = _get_panel_last_final_payload(category, action)
    if not isinstance(payload, dict) or not bool(payload.get("final")):
        return None

    # 이미 current source로 승격된 동일 run_seq/action 결과만 view skip 허용.
    # 사용자가 [패널 표 다시 표시]를 누른 경우에는 _should_compact...에서 False가 된다.
    if not _should_compact_panel_result_on_rerun(payload, action):
        return None

    try:
        log.info(
            "[panel.main] skip view on chat rerun: %s.%s sig=%s reason=%s",
            category,
            action,
            _make_panel_source_sig(action),
            ss.pop("__sims_panel_skip_view_reason", "chat_input"),
        )
    except Exception:
        pass

    return payload


# 패널 화면 표시 제한 건수 계산/적용 함수
# - 패널 화면 표시 건수는 payload의 params.top으로 제어한다.
def _panel_display_limit_from_payload(payload: Dict[str, Any]) -> int:
    """
    패널 화면 표시 제한 건수.

    우선순위:
    1. display_top : 화면 표시 전용 건수
    2. view_top    : 향후 별칭
    3. top         : 기존 호환
    """
    try:
        params = payload.get("params") or {}

        for key in ("display_top", "view_top", "top"):
            raw = params.get(key)
            if raw is None or str(raw).strip() == "":
                continue

            top = int(raw)
            if top > 0:
                return top

        return 0
    except Exception:
        return 0
    

def _limit_panel_display_df(
    df: pd.DataFrame,
    *,
    limit: int,
    payload: Dict[str, Any],
    title: str = "",
) -> pd.DataFrame:
    """
    패널 표시용 DataFrame만 제한한다.
    원본 df는 다운로드/LLM/현재표 후속분석 기준으로 유지한다.

    제품재고현황처럼 마지막 합계 행이 있는 경우:
    - 상세 N건 + 합계 1행을 표시한다.
    - 조회 완료 건수 계산에서는 기존 로직이 합계 행을 빼므로 정상적으로 N건이 된다.
    """
    if not isinstance(df, pd.DataFrame) or df.empty:
        return df

    if limit <= 0 or len(df) <= limit:
        return df

    meta = payload.get("meta") or {}
    group_label = str(meta.get("group_label") or "").strip()
    is_inventory = "제품재고현황" in str(title or payload.get("title") or payload.get("action") or "")

    if (
        is_inventory
        and group_label
        and group_label in df.columns
        and len(df) > 0
        and str(df.iloc[-1][group_label]).strip() == "합계"
    ):
        # 상세 limit건 + 합계 1행
        return pd.concat([df.head(limit).copy(), df.tail(1).copy()], ignore_index=True)

    return df.head(limit).copy()


def _apply_panel_display_limit_to_payload(payload: Dict[str, Any], title: str = "") -> None:
    """
    payload 내부의 df_display만 화면 표시 건수로 제한한다.

    중요:
    - payload["df"]는 원본/다운로드/LLM/현재표 후속분석 기준으로 유지
    - payload["df_display"]만 화면 표시용으로 제한
    - records/columns는 df_display 기준으로 다시 맞춘다
    """
    try:
        if not isinstance(payload, dict):
            return

        if not _panel_payload_matches_current_company(payload):
            payload_stamp = _panel_payload_company_stamp(payload)
            current_stamp = _panel_current_company_stamp()
            log_state = _panel_stale_payload_log_state(payload_stamp, current_stamp)
            log.info(
                "[panel] skip stale table stash after company change action=%s payload_company_id=%s current_company_id=%s db_mismatch=%s",
                action,
                log_state["payload_company_id"],
                log_state["current_company_id"],
                log_state["db_mismatch"],
            )
            return

        _panel_stamp_payload_company(payload)

        df_full = payload.get("df")
        df_disp = payload.get("df_display")

        # 기존 build_result_payload가 df만 넣고 df_display를 안 넣는 경우가 있다.
        # 이 경우 화면 표시용 df_display를 df에서 만들어 제한을 적용한다.
        if not isinstance(df_disp, pd.DataFrame):
            if isinstance(df_full, pd.DataFrame):
                df_disp = df_full
                payload["df_display"] = df_disp
            else:
                return
        df_disp = normalize_display_df_for_streamlit(df_disp)
        payload["df_display"] = df_disp

        if not isinstance(df_full, pd.DataFrame) or df_full.empty:
            df_full = df_disp.copy()
            payload["df"] = df_full

        param_limit = _panel_display_limit_from_payload(payload)
        env_limit = _panel_int_env("SIMS_PANEL_DISPLAY_MAX_ROWS", 2000)

        limits = [x for x in [param_limit, env_limit] if isinstance(x, int) and x > 0]
        limit = min(limits) if limits else 0

        if limit <= 0 or len(df_disp) <= limit:
            meta = payload.setdefault("meta", {})
            meta["display_row_count"] = int(len(df_disp))
            meta["row_count_total"] = int(len(df_full))
            meta.setdefault("download_row_count", int(len(df_full)))
            payload["columns"] = list(df_disp.columns)
            payload["records"] = df_disp.to_dict(orient="records")
            return

        limited = _limit_panel_display_df(
            df_disp,
            limit=limit,
            payload=payload,
            title=title,
        )
        limited = normalize_display_df_for_streamlit(limited)

        payload["df_display"] = limited
        payload["columns"] = list(limited.columns)
        payload["records"] = limited.to_dict(orient="records")

        meta = payload.setdefault("meta", {})
        meta["display_row_count"] = int(len(limited))
        meta["row_count_total"] = int(len(df_full))
        meta.setdefault("download_row_count", int(len(df_full)))

        log.info(
            "[panel] display limited action=%s display_rows=%s full_rows=%s limit=%s",
            payload.get("action") or title,
            len(limited),
            len(df_full),
            limit,
        )
    except Exception:
        log.exception("[panel] apply display limit failed")


def _panel_detail_count_for_display(df: pd.DataFrame, payload: Dict[str, Any]) -> int:
    """
    합계행이 있는 표는 합계행을 조회건수에서 제외한다.
    """
    if not isinstance(df, pd.DataFrame) or df.empty:
        return 0

    try:
        meta = payload.get("meta") or {}
        group_label = str(meta.get("group_label") or "").strip()

        if (
            group_label
            and group_label in df.columns
            and str(df.iloc[-1][group_label]).strip() == "합계"
        ):
            return max(0, int(len(df)) - 1)
    except Exception:
        pass

    return int(len(df))


def _render_panel_result_count_caption(payload: Dict[str, Any], df_disp: pd.DataFrame) -> None:
    """
    패널 조회 결과의 전체/화면 표시 건수를 명확히 표시한다.
    """
    if not isinstance(df_disp, pd.DataFrame):
        return

    meta = payload.get("meta") or {}
    df_full = payload.get("df")

    display_count = _panel_detail_count_for_display(df_disp, payload)

    full_count = 0
    try:
        if meta.get("detail_count") is not None:
            full_count = int(meta.get("detail_count") or 0)
    except Exception:
        full_count = 0

    if full_count <= 0 and isinstance(df_full, pd.DataFrame):
        full_count = _panel_detail_count_for_display(df_full, payload)

    if full_count <= 0:
        full_count = display_count

    try:
        limit = _panel_display_limit_from_payload(payload)
    except Exception:
        limit = 0

    if full_count > display_count:
        st.caption(
            f"조회결과: 전체 {full_count:,}건 중 화면 {display_count:,}건 표시"
            + (f" / 화면 조회건수 {limit:,}건" if limit > 0 else "")
        )
    else:
        st.caption(f"조회결과: {display_count:,}건")


def _panel_compact_text(value: Any, *, limit: int = 180) -> str:
    text = re.sub(r"\s+", " ", str(value or "").strip())
    if len(text) > limit:
        return text[: max(0, limit - 1)].rstrip() + "…"
    return text


def _panel_detail_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _panel_result_header_rows(payload: Dict[str, Any], df_disp: pd.DataFrame) -> tuple[int, int, int]:
    meta = payload.get("meta") or {}
    df_full = payload.get("df")
    display_count = _panel_detail_count_for_display(df_disp, payload) if isinstance(df_disp, pd.DataFrame) else 0

    full_count = 0
    for key in ("download_row_count", "detail_count", "row_count_total", "row_count"):
        try:
            raw = meta.get(key)
            if raw is not None:
                full_count = int(raw or 0)
                if full_count > 0:
                    break
        except Exception:
            full_count = 0

    if full_count <= 0 and isinstance(df_full, pd.DataFrame):
        full_count = _panel_detail_count_for_display(df_full, payload)
    if full_count <= 0:
        full_count = display_count

    expected_count = 0
    for key in ("db_total_count", "total_count", "matched_row_count", "matched_total_count"):
        try:
            raw = meta.get(key)
            if raw is not None:
                expected_count = int(raw or 0)
                if expected_count > 0:
                    break
        except Exception:
            expected_count = 0
    if expected_count <= 0:
        expected_count = full_count

    return full_count, display_count, expected_count


def _render_panel_result_compact_header(payload: Dict[str, Any], action: str, title: str, df_disp: pd.DataFrame) -> None:
    meta = payload.get("meta") or {}
    query_summary = str(meta.get("query_summary") or "").strip()

    if not query_summary:
        params = payload.get("params") or {}
        parts = []
        for k, v in params.items():
            text = str(v or "").strip()
            if text:
                parts.append(f"{k} {text}")
        query_summary = " / ".join(parts)

    full_rows, display_rows, expected_rows = _panel_result_header_rows(payload, df_disp)
    if expected_rows and full_rows and expected_rows > full_rows:
        line1 = f"결과: 조건 전체 {expected_rows:,}건 · 조회 {full_rows:,}건"
        if display_rows and display_rows < full_rows:
            line1 += f" · 표 데이터 {display_rows:,}건"
    elif full_rows and display_rows and full_rows > display_rows:
        line1 = f"결과: 전체 {full_rows:,}건 · 표 데이터 {display_rows:,}건"
    elif full_rows:
        line1 = f"결과: {full_rows:,}건"
    else:
        line1 = "결과 정보가 저장되어 있습니다."
    st.caption(line1)

    line2_parts: list[str] = []
    if query_summary:
        line2_parts.append(f"조회조건: {_panel_compact_text(query_summary, limit=160)}")
    if str(meta.get("table_key") or "").strip():
        line2_parts.append("현재표 후속질문 가능")
    if line2_parts:
        st.caption(" · ".join(line2_parts))

    details: list[str] = []
    for label, value in (
        ("조회명", title or action),
        ("조회시각", meta.get("created_at") or meta.get("timestamp") or meta.get("ts") or payload.get("time")),
        ("전체 조회조건", query_summary),
        ("전체 결과 행수", f"{full_rows:,}건" if full_rows else ""),
        ("표시 행수", f"{display_rows:,}건" if display_rows else ""),
        ("다운로드 행수", f"{int(meta.get('download_row_count') or full_rows):,}건" if (meta.get("download_row_count") or full_rows) else ""),
        ("현재표 후속질문", "가능" if meta.get("table_key") else ""),
        ("action", action),
        ("source", meta.get("source")),
        ("source_action", meta.get("source_action") or meta.get("source_table_action")),
        ("table_key", meta.get("table_key")),
        ("source_key", meta.get("source_table_key") or meta.get("source_key")),
    ):
        text = _panel_detail_text(value)
        if text:
            details.append(f"- {label}: {text}")

    if details:
        with st.expander("상세 조회정보", expanded=False):
            st.markdown("\n".join(details))


def _stash_panel_table_for_current_followup(
    payload: Dict[str, Any],
    action: str,
    *,
    record_previous_source_for_prune: bool = False,
) -> None:
    """
    패널 결과를 채팅에 표로 중복 표시하지 않더라도,
    '현재표 ...' 후속분석이 가능하도록 session_state에 테이블만 저장한다.
    """
    try:
        if not isinstance(payload, dict):
            return

        df_full = payload.get("df")
        df_disp = payload.get("df_display")

        if not isinstance(df_full, pd.DataFrame) or df_full.empty:
            if isinstance(df_disp, pd.DataFrame) and not df_disp.empty:
                df_full = df_disp
            else:
                return

        if not isinstance(df_disp, pd.DataFrame) or df_disp.empty:
            df_disp = df_full

        ss = st.session_state
        ss.setdefault("sims_tables", {})
        ss.setdefault("sims_export_tables", {})
        ss.setdefault("__sims_export_tables_by_key", {})

        meta = payload.setdefault("meta", {})
        table_key = str(meta.get("table_key") or f"sims_{uuid.uuid4().hex[:8]}")
        meta["table_key"] = table_key

        # ------------------------------------------------------------
        # 패널 현재표 후속분석 전체 기준 보정
        # ------------------------------------------------------------
        # 패널 조회에서는 df_display뿐 아니라 payload["df"]도
        # 화면 제한 건수(예: 5,000건)만 들어오는 경우가 있다.
        #
        # 원칙:
        # - 화면 표시는 일부 행 가능
        # - 현재표 후속분석은 전체 조회조건 기준
        # - 다운로드도 전체 조회조건 기준
        try:
            display_rows = int(len(df_disp)) if isinstance(df_disp, pd.DataFrame) else 0
            full_rows = int(len(df_full)) if isinstance(df_full, pd.DataFrame) else 0
            expected_rows = _expected_analysis_row_count(meta, max(display_rows, full_rows))

            action_for_export = str(
                action
                or payload.get("action")
                or meta.get("action")
                or payload.get("title")
                or ""
            ).strip()

            is_validation_action = "검증" in action_for_export

            query_limit = int(
                meta.get("fetch_limit")
                or meta.get("display_top")
                or meta.get("top")
                or (payload.get("params") or {}).get("_max_top")
                or (payload.get("params") or {}).get("display_top")
                or (payload.get("params") or {}).get("top")
                or 0
            )

            force_validation_export = bool(
                is_validation_action
                and query_limit > 0
                and full_rows >= query_limit
            )

            if (
                (expected_rows > full_rows or force_validation_export)
                and isinstance(df_disp, pd.DataFrame)
                and not df_disp.empty
            ):
                action_for_export = str(
                    action
                    or payload.get("action")
                    or meta.get("action")
                    or payload.get("title")
                    or ""
                ).strip()

                item_for_export = {
                    "action": action_for_export,
                    "title": payload.get("title"),
                    "params": payload.get("params") or meta.get("params") or {},
                    "meta": meta,
                    "df_display": df_disp,
                    "df": df_full if isinstance(df_full, pd.DataFrame) else None,
                }

                full_df_for_followup = _get_full_download_df_for_sims_item(
                    item_for_export,
                    meta,
                    df_disp,
                )

                if (
                    isinstance(full_df_for_followup, pd.DataFrame)
                    and not full_df_for_followup.empty
                    and len(full_df_for_followup) > full_rows
                ):
                    df_full = full_df_for_followup

                    display_rows_i = int(len(df_disp)) if isinstance(df_disp, pd.DataFrame) else 0
                    full_rows_i = int(len(df_full)) if isinstance(df_full, pd.DataFrame) else 0

                    meta["row_count"] = full_rows_i
                    meta["display_row_count"] = display_rows_i
                    meta["row_count_total"] = full_rows_i
                    meta["download_row_count"] = full_rows_i
                    meta["row_count_total_for_followup"] = full_rows_i

                    if full_rows_i > display_rows_i:
                        payload["message"] = (
                            f"조회결과: 전체 {full_rows_i:,}건 중 화면 {display_rows_i:,}건 표시"
                        )
                    else:
                        payload["message"] = f"조회결과: {display_rows_i:,}건"

                    payload["df"] = df_full

                    meta["download_row_count"] = int(len(df_full))
                    meta["row_count_total_for_followup"] = int(len(df_full))

                    log.info(
                        "[panel] upgraded followup df to full export action=%s table_key=%s display_rows=%s old_full_rows=%s full_rows=%s expected_rows=%s",
                        action_for_export,
                        table_key,
                        display_rows,
                        full_rows,
                        len(df_full),
                        expected_rows,
                    )

        except Exception:
            log.exception("[panel] upgrade followup df to full export failed")

        previous_source_key = str(ss.get("__sims_current_table_source_key") or "").strip()
        previous_source_action = str(ss.get("__sims_current_table_source_action") or "").strip()
        if (
            record_previous_source_for_prune
            and previous_source_key
            and previous_source_key != table_key
            and not bool(meta.get("current_table_followup"))
        ):
            ss["__sims_previous_current_table_source_key_for_prune"] = previous_source_key
            ss["__sims_previous_current_table_source_action_for_prune"] = previous_source_action
            ss["__sims_previous_current_table_source_target_key_for_prune"] = table_key
            ss["__old_table_history_refresh_key_pending"] = table_key
            try:
                log.info(
                    "[current_table.source_transition] previous_key=%s new_key=%s reason=panel_new_result",
                    previous_source_key,
                    table_key,
                )
            except Exception:
                pass

        ss["sims_tables"][table_key] = df_disp
        ss["sims_export_tables"][table_key] = df_full
        ss["__sims_export_tables_by_key"][table_key] = df_full

        ss["__sims_last_table_key"] = table_key
        ss["__sims_last_table_action"] = str(action or payload.get("action") or payload.get("title") or "")
        ss["__sims_current_table_source_key"] = table_key
        ss["__sims_current_table_source_action"] = ss["__sims_last_table_action"]

        meta["table_key"] = table_key
        meta["download_table_key"] = table_key
        meta["display_row_count"] = int(len(df_disp))
        meta["download_row_count"] = int(len(df_full))
        meta.setdefault("row_count_total", int(len(df_full)))

        log.info(
            "[panel] stashed table for current followup table_key=%s display_rows=%s full_rows=%s action=%s",
            table_key,
            len(df_disp),
            len(df_full),
            action,
        )
    except Exception:
        log.exception("[panel] stash table for current followup failed")

def _ensure_panel_llm_context_from_payload(
    *,
    category: str,
    action: str,
) -> bool:
    """
    패널 LLM 분석 실행 직전에 마지막 최종 payload를 SIMS LLM 컨텍스트로 다시 주입한다.

    이유:
    - 패널의 [LLM 분석] 버튼은 fragment 안에서 실행된다.
    - 이때 session_state의 SIMS 컨텍스트가 비어 있으면
      build_messages_with_system이 NO SIMS BLOCK 상태가 된다.
    """
    try:
        payload = _get_panel_last_final_payload(category, action)
        if not isinstance(payload, dict):
            log.warning(
                "[panel.fragment] no remembered payload for LLM context category=%s action=%s",
                category,
                action,
            )
            return False

        meta = payload.get("meta") or {}
        sig = str(
            payload.get("id")
            or meta.get("table_key")
            or meta.get("sig")
            or f"{category}::{action}::{meta.get('row_count') or meta.get('row_count_total') or ''}"
        )

        last_sig = st.session_state.get("__panel_llm_context_sig")
        if last_sig != sig:
            # 패널 화면은 df_display를 쓰지만,
            # LLM 분석 컨텍스트는 가능하면 전체 원본 df를 사용해야 한다.
            payload_for_ctx = dict(payload)

            df_full = payload.get("df")
            df_display = payload.get("df_display")

            if isinstance(df_full, pd.DataFrame) and not df_full.empty:
                payload_for_ctx["df_display"] = df_full
                payload_for_ctx["df"] = df_full
                log.info(
                    "[panel.fragment] use full df for LLM context rows=%s cols=%s",
                    len(df_full),
                    len(df_full.columns),
                )
            elif isinstance(df_display, pd.DataFrame) and not df_display.empty:
                log.info(
                    "[panel.fragment] use display df for LLM context rows=%s cols=%s",
                    len(df_display),
                    len(df_display.columns),
                )

            # push_sims_result_to_chat가 SIMS 분석 컨텍스트를 갱신한다.
            # 단, 패널 LLM 버튼 클릭 시 채팅에 표를 중복 추가하지 않도록 silent 처리한다.
            _stash_panel_table_for_current_followup(payload_for_ctx, action)
            st.session_state["__sims_silent_push"] = True
            push_sims_result_to_chat(payload_for_ctx, action)

            st.session_state["__panel_llm_context_sig"] = sig
            log.info(
                "[panel.fragment] pushed panel payload to SIMS LLM context action=%s sig=%s",
                action,
                sig,
            )
        else:
            log.debug("[panel.fragment] panel LLM context already prepared sig=%s", sig)

        return True

    except Exception:
        log.exception("[panel.fragment] ensure panel LLM context failed")
        return False
    
    
# ==========================================================
# 🧩 메인 진입
#   selected = {"category": "...", "action": "..."}
# ==========================================================
def render_sims_main(selected: Optional[Dict[str, str]]) -> None:
    _ensure_sims_state()
    ss = st.session_state

    # 이번 런이 "SIMS 실행으로 패널을 연 런"인지 먼저 잡아 둔다.
    opened_by_run = bool(ss.get("__sims_run_flag"))
    inner_submit = bool(ss.get("__sims_inner_submit"))

    # 이번 실행에서 prepass가 세운 run_flag를 한 번만 소비하도록 정리
    # - 버튼 클릭 시: prepass에서 True → 여기서 False로 내려줌
    # - 이후 rerun에서는 prepass가 다시 세우지 않는 한 재실행되지 않음
    ss["__sims_run_flag"] = False

    # 최종 여부 플래그는 payload 렌더링 시점에 다시 세팅
    ss["__sims_was_final"] = False

    # 사이드 컨트롤(열림/닫힘 등) — 필요 시 보이기
    render_sims_context_controls()

    if not selected or not selected.get("category") or not selected.get("action"):
        st.info("좌측에서 카테고리와 작업을 선택하세요.")
        return

    category = selected["category"]
    action = selected["action"]
    ss["__sims_selected"] = selected

    # SIMS 실행으로 패널만 다시 연 경우에는,
    # 같은 액션의 이전 조회 payload를 그대로 재사용하지 않도록 현재 액션 payload를 비운다.
    if opened_by_run and not inner_submit:
        _clear_current_action_payload(action)
        _clear_panel_last_final_payload()

    # 액션별 함수 찾기
    view_func = _CATEGORIES.get(category, {}).get("actions", {}).get(action)
    if view_func is None:
        st.warning(f"알 수 없는 작업입니다: {category} / {action}")
        return

    log.info("[panel.main] enter selected=%r", selected)
    _log_state("enter")

    # 액션 변경 시에만 form_id 증가 및 위젯키 정리
    if ss.get("__sims_last_action") != f"{category}::{action}":
        if str(ss.get("__ui_rerun_reason_current") or "") != "current_table_followup":
            ss["__ui_rerun_reason"] = "sims_action_change"
        ss["__sims_form_id"] = int(ss.get("__sims_form_id", 0)) + 1
        ss["__sims_widget_ns"] = str(ss["__sims_form_id"])
        ss["__sims_last_action"] = f"{category}::{action}"
        _clear_widget_keys_for(action, ss["__sims_form_id"])

    # 패널 활성/오픈 상태는 메인(Lmstudio_SSAI_chat_main.py)의 prepass에서만 관리한다.
    # 여기서는 __sims_open / __sims_panel_active 값을 절대 덮어쓰지 않는다.

    # ------------------ 폼 블록 ------------------
    payload: Any = None

    # ------------------------------------------------------------
    # B단계 성능 패치:
    # 채팅 입력/현재표 후속분석으로 생긴 rerun에서는,
    # 이미 final payload가 있는 패널 view 함수 호출 자체를 생략한다.
    # ------------------------------------------------------------
    payload = _pop_panel_skip_view_payload_for_chat_rerun(
        category=category,
        action=action,
        opened_by_run=opened_by_run,
        inner_submit=inner_submit,
    )

    # ─────────────────────────
    # 2) 제출 버튼을 먼저 읽고 → 제출 플래그 세운 다음 → 같은 렌더에서 view_func() 호출
    #    (두 번 눌러야 하는 현상을 제거)
    # ─────────────────────────
    try:
        if payload is None:
            log.info("[panel.main] call view: %s.%s (run_seq=%s, form_id=%s)",
                     category, action, ss.get("__sims_run_seq"), ss.get("__sims_form_id"))
            # 폼 네임스페이스(위젯 key suffix)는 form_id 그대로 사용
            ss["__sims_widget_ns"] = str(ss["__sims_form_id"])

            # ✅ 각 view 함수 내부에서 st.form(...)과 st.form_submit_button(...)을 직접 사용
            _log_state("before_view")
            payload = view_func()
            _log_state("after_view")

            # [LLM 분석] 버튼 클릭 같은 일반 rerun에서는
            # view 함수가 final=False를 반환할 수 있다.
            # 이때 같은 action의 마지막 최종 결과를 복원해야
            # 기존 표/다운로드/LLM 버튼이 다시 렌더되고 버튼 클릭 이벤트가 소비된다.
            if (
                isinstance(payload, dict)
                and not bool(payload.get("final"))
                and not opened_by_run
                and not inner_submit
            ):
                remembered_payload = _get_panel_last_final_payload(category, action)
                if isinstance(remembered_payload, dict):
                    payload = remembered_payload
                    log.debug("[panel] restored last final payload for action=%s.%s", category, action)

    except Exception as e:
        st.error(f"실행 오류: {e}")
        log.exception(
            "sims_panel: view render error (category=%s, action=%s, run_flag=%s, was_final=%s, form_id=%s, widget_ns=%s)",
            category,
            action,
            ss.get("__sims_run_flag"),
            ss.get("__sims_was_final"),
            ss.get("__sims_form_id"),
            ss.get("__sims_widget_ns"),
        )
        return

    # payload 점검
    if not isinstance(payload, dict):
        st.info("결과가 없습니다.")
        log.info("SIMS action=%r -> no dict payload", action)
        return
    # ------------------ 결과 렌더/브리지 ------------------
    _render_payload(payload, action)

    log.info("[panel.main] leave")

def _fmt_metric_value(value: Any) -> str:
    try:
        n = float(value)
    except Exception:
        return str(value or "")

    if abs(n - int(n)) < 1e-9:
        return f"{int(n):,}"
    return f"{n:,.2f}"

def _week_label_52(value: Any) -> str:
    if value is None:
        return "-"

    if isinstance(value, dt.datetime):
        d = value.date()
    elif isinstance(value, dt.date):
        d = value
    else:
        text = str(value).strip()
        if not text:
            return "-"
        digits = "".join(ch for ch in text if ch.isdigit())

        d = None
        try:
            if len(digits) >= 8:
                d = dt.datetime.strptime(digits[:8], "%Y%m%d").date()
            elif len(digits) == 6:
                d = dt.datetime.strptime(digits + "01", "%Y%m%d").date()
        except Exception:
            d = None

        if d is None:
            return text

    weekday_map = ["월", "화", "수", "목", "금", "토", "일"]
    week_no = min(int(d.strftime("%W")) + 1, 52)
    return f"{d:%Y-%m-%d}({weekday_map[d.weekday()]}) / {d.year}년 {week_no:02d}주"

def _render_stock_flow_summary(payload: Dict[str, Any], df_disp: pd.DataFrame) -> None:
    meta = payload.get("meta") or {}
    required = ("carry_qty", "in_qty", "out_qty", "stock_qty")
    if not all(k in meta for k in required):
        return

    params = payload.get("params") or {}
    product_info = meta.get("product_info") or {}

    product_cd = str(params.get("physic_cd") or "").strip()
    product_nm = str(params.get("physic_nm") or "").strip()
    stock_mode = str(params.get("stock_mode") or "").strip()
    flow_scope = str(params.get("flow_scope") or "").strip()
    date_basis = str(params.get("date_basis") or "").strip()
    stock_names = params.get("stock_names") or []
    if not isinstance(stock_names, list):
        stock_names = []

    selected_product = " / ".join([x for x in [product_cd, product_nm] if x]).strip()
    selected_stock = "전체" if not stock_names else ", ".join(map(str, stock_names))

    st.caption(
        f"조건 · 제품: {selected_product or '-'} | "
        f"수불구분: {stock_mode or '-'} | "
        f"조회범위: {flow_scope or '-'} | "
        f"기준일자: {date_basis or '-'} | "
        f"재고위치: {selected_stock}"
    )

    def _fmt_info_value(v: Any) -> str:
        if v is None:
            return ""
        try:
            if pd.isna(v):
                return ""
        except Exception:
            pass

        if isinstance(v, (int, float)):
            return _fmt_metric_value(v)

        s = str(v).strip()
        if s in {"", "None", "nan", "<NA>", "NaT"}:
            return ""
        return s

    if isinstance(product_info, dict) and product_info:
        info_pairs = [
            ("제품코드", product_info.get("제품코드")),
            ("제품명", product_info.get("제품명")),
            ("규격", product_info.get("규격")),
            ("포장단위", product_info.get("포장단위")),
            ("최종보험가", product_info.get("최종보험가")),
            ("보험코드", product_info.get("보험코드")),
            ("표준코드", product_info.get("표준코드")),
            ("제조사명", product_info.get("제조사명")),
            ("발주처명", product_info.get("발주처명")),
            ("제품그룹명", product_info.get("제품그룹명")),
            ("제품분류명", product_info.get("제품분류명")),
            ("특수관리제품명", product_info.get("특수관리제품명")),
        ]

        parts = []
        for label, value in info_pairs:
            text = _fmt_info_value(value)
            if text:
                parts.append(f"{label} {text}")

        if parts:
            st.caption("제품정보: " + " / ".join(parts))

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.metric("이월재고", _fmt_metric_value(meta.get("carry_qty", 0)))
    with c2:
        st.metric("입고수량", _fmt_metric_value(meta.get("in_qty", 0)))
    with c3:
        st.metric("출고수량", _fmt_metric_value(meta.get("out_qty", 0)))
    with c4:
        st.metric("재고수량", _fmt_metric_value(meta.get("stock_qty", 0)))

    detail_count = len(df_disp)
    try:
        if (
            isinstance(df_disp, pd.DataFrame)
            and not df_disp.empty
            and "입출고일자" in df_disp.columns
            and str(df_disp.iloc[0]["입출고일자"]).strip() == "이월재고"
        ):
            detail_count -= 1
    except Exception:
        pass

    st.caption(f"상세건수: {max(detail_count, 0):,}건")

def _render_inventory_summary(payload: Dict[str, Any], df_disp: pd.DataFrame) -> None:
    meta = payload.get("meta") or {}
    required = (
        "sum_carry_qty",
        "sum_in_qty",
        "sum_out_qty",
        "sum_stock_qty",
        "sum_stock_amt",
        "sum_insu_amt",
    )
    if not all(k in meta for k in required):
        return

    params = payload.get("params") or {}

    def _to_stock_mode_label(v: Any) -> str:
        s = str(v or "").strip().lower()
        return {
            "real": "실재고",
            "book": "장부재고",
            "실재고": "실재고",
            "장부재고": "장부재고",
        }.get(s, str(v or "-"))

    def _to_group_basis_label(v: Any) -> str:
        s = str(v or "").strip().lower()
        return {
            "maker": "제조사별 제품 누적",
            "order": "발주처별 제품 누적",
            "purchase": "매입처+제품 누적",
            "제조사": "제조사별 제품 누적",
            "발주처": "발주처별 제품 누적",
            "매입처": "매입처+제품 누적",
        }.get(s, str(v or "-"))

    def _to_price_mode_label(v: Any) -> str:
        s = str(v or "").strip().lower()
        return {
            "avg": "총평균단가",
            "last": "최종매입가",
            "std": "기준가",
            "insu": "현보험약가",
            "총평균단가": "총평균단가",
            "최종매입가": "최종매입가",
            "기준가": "기준가",
            "현보험약가": "현보험약가",
        }.get(s, str(v or "-"))

    stock_mode = _to_stock_mode_label(params.get("stock_mode"))
    group_basis = _to_group_basis_label(params.get("group_basis"))
    price_mode = _to_price_mode_label(params.get("price_mode"))

    product_cd = _clean_ui_text(params.get("physic_cd"))
    product_nm = _clean_ui_text(params.get("physic_nm"))

    maker_cd = _clean_ui_text(params.get("maker_cd"))
    maker_nm = _clean_ui_text(params.get("maker_nm"))
    order_cd = _clean_ui_text(params.get("order_cd"))
    order_nm = _clean_ui_text(params.get("order_nm"))
    buy_cd = _clean_ui_text(params.get("buy_cd"))
    buy_nm = _clean_ui_text(params.get("buy_nm"))

    date_from = _week_label_52(params.get("date_from"))
    date_to = _week_label_52(params.get("date_to"))

    stock_names = params.get("stock_names") or []
    if not isinstance(stock_names, list):
        stock_names = []

    selected_product = " / ".join([x for x in [product_cd, product_nm] if x]).strip()

    maker_text = " / ".join([x for x in [maker_cd, maker_nm] if x]).strip()
    order_text = " / ".join([x for x in [order_cd, order_nm] if x]).strip()
    buy_text = " / ".join([x for x in [buy_cd, buy_nm] if x]).strip()

    selected_stock = "전체" if not stock_names else ", ".join(map(str, stock_names))

    cond_parts = []
    if maker_text:
        cond_parts.append(f"제조사: {maker_text}")
    if order_text:
        cond_parts.append(f"발주처: {order_text}")
    if buy_text:
        cond_parts.append(f"매입처: {buy_text}")
    vendor_cond = " | ".join(cond_parts) if cond_parts else "거래처조건: 없음"

    st.caption(
        f"조건 · "
        f"조회기간: {date_from} ~ {date_to} | "
        f"제품: {selected_product or '전체'} | "
        f"재고구분: {stock_mode} | "
        f"집계기준: {group_basis} | "
        f"단가기준: {price_mode} | "
        f"{vendor_cond} | "
        f"재고위치: {selected_stock}"
    )

    c1, c2, c3 = st.columns(3)
    with c1:
        st.metric("이월수량", _fmt_metric_value(meta.get("sum_carry_qty", 0)))
    with c2:
        st.metric("입고수량", _fmt_metric_value(meta.get("sum_in_qty", 0)))
    with c3:
        st.metric("출고수량", _fmt_metric_value(meta.get("sum_out_qty", 0)))

    c4, c5, c6 = st.columns(3)
    with c4:
        st.metric("재고수량", _fmt_metric_value(meta.get("sum_stock_qty", 0)))
    with c5:
        st.metric("재고금액", _fmt_metric_value(meta.get("sum_stock_amt", 0)))
    with c6:
        st.metric("보험금액", _fmt_metric_value(meta.get("sum_insu_amt", 0)))

    detail_count = len(df_disp)
    try:
        group_label = str(meta.get("group_label") or "").strip()
        if (
            isinstance(df_disp, pd.DataFrame)
            and not df_disp.empty
            and group_label
            and group_label in df_disp.columns
            and str(df_disp.iloc[-1][group_label]).strip() == "합계"
        ):
            detail_count -= 1
    except Exception:
        pass

    st.caption(f"집계건수: {max(detail_count, 0):,}건  ·  마지막 행은 합계입니다.")


def _fmt_header_num(value: Any, unit: str = "", *, decimals: int | None = None) -> str:
    if decimals is None and unit == "원":
        decimals = 0

    n = _parse_number_for_style(value)
    if n is None:
        text = str(value or "").strip()
    else:
        if decimals is not None:
            text = f"{n:,.{decimals}f}"
        elif abs(n - int(n)) < 1e-9:
            text = f"{int(n):,}"
        else:
            text = f"{n:,.2f}".rstrip("0").rstrip(".")

    if not text:
        text = "-"

    return f"{text}{unit}" if unit and text != "-" else text


def _render_simple_analysis_header(payload: Dict[str, Any]) -> None:
    """
    NLQ 분석/KPI 결과용 간단 헤더.
    패널 조회 화면의 헤더까지는 아니어도, 핵심 요약을 표 위에 보여준다.
    """
    meta = payload.get("meta") or {}
    if not isinstance(meta, dict):
        return

    # 패널 조회는 analytics_views에서 이미 헤더를 그리므로,
    # NLQ 결과에만 여기서 헤더를 추가한다.
    if not meta.get("nlq") and not meta.get("analysis_nlq"):
        return

    analysis_type = str(meta.get("analysis_type") or "").strip()
    title = str(payload.get("title") or payload.get("action") or "").strip()

    if analysis_type == "stock_shortage":
        st.markdown("### 재고부족요약")

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.info(f"품목수 : {_fmt_header_num(meta.get('product_count'), '개')}")
        with c2:
            st.info(f"부족품목수 : {_fmt_header_num(meta.get('shortage_item_count'), '개')}")
        with c3:
            st.info(f"현재재고수량 : {_fmt_header_num(meta.get('sum_current_stock_qty'), '개')}")
        with c4:
            st.info(f"현재재고금액 : {_fmt_header_num(meta.get('sum_current_stock_amt'), '원')}")

        c5, c6, c7, c8 = st.columns(4)
        with c5:
            st.info(f"평가월 예상수요 : {_fmt_header_num(meta.get('sum_current_month_expected_out_qty'), '개')}")
        with c6:
            st.info(f"평가월 실제수요 : {_fmt_header_num(meta.get('sum_current_month_out_qty'), '개')}")
        with c7:
            st.warning(f"평가월 잔여예상수요 : {_fmt_header_num(meta.get('sum_current_month_remaining_out_qty'), '개')}")
        with c8:
            st.success(f"평가월 수요진척률 : {_fmt_header_num(meta.get('current_month_demand_progress_pct'), '%', decimals=2)}")

        c9, c10, c11, c12 = st.columns(4)
        with c9:
            st.warning(f"부족예상수량 : {_fmt_header_num(meta.get('sum_expected_shortage_qty'), '개')}")
        with c10:
            st.warning(f"부족예상금액 : {_fmt_header_num(meta.get('sum_expected_shortage_amt'), '원')}")
        with c11:
            st.info(f"재고충족률 : {_fmt_header_num(meta.get('overall_stock_fill_rate'), '%', decimals=2)}")
        with c12:
            st.info(f"재고기준 : {str(meta.get('stock_label') or '-')}")

        st.caption(
            f"자료원: {meta.get('source_label') or '-'} / "
            f"현재고원천: {meta.get('stock_source_label') or meta.get('stock_label') or '-'} / "
            f"조회건수: {meta.get('row_count_total') or meta.get('row_count') or 0}건"
        )

        counts = meta.get("shortage_grade_counts") or {}
        if isinstance(counts, dict) and counts:
            order = [
                "재고없음",
                "1개월내 부족",
                "2개월내 부족주의",
                "3개월내 부족주의",
                "정상",
                "수요관찰",
                "재고없음/수요없음",
            ]
            keys = [k for k in order if k in counts]
            keys += [k for k in counts.keys() if k not in keys]
            line = " / ".join(f"{k} {_fmt_header_num(counts.get(k), '개')}" for k in keys)
            st.caption(f"부족등급별 제품수: {line}")

        return

    if analysis_type in {"manufacturer_sales_trend", "manufacturer_sales_trend_summary"} or meta.get("summary_type") in {"manufacturer_trend_detail", "manufacturer_trend_summary"}:
        show_manufacturer_extended = analysis_type == "manufacturer_sales_trend_summary" or meta.get("summary_type") == "manufacturer_trend_summary"
        st.markdown("### 매출추세요약")
        if meta.get("period_caption"):
            st.caption(str(meta.get("period_caption")))
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.info(f"총매출액 : {_fmt_header_num(meta.get('sum_sales_amt'), '원')}")
        with c2:
            st.info(f"매출공급가액 : {_fmt_header_num(meta.get('sum_supply_amt'), '원')}")
        with c3:
            st.info(f"매출세액 : {_fmt_header_num(meta.get('sum_tax_amt'), '원')}")
        with c4:
            st.info(f"제약사수 : {_fmt_header_num(meta.get('manufacturer_count') or meta.get('group_count'), '개')}")

        c5, c6, c7, c8 = st.columns(4)
        with c5:
            st.info(f"제품수 : {_fmt_header_num(meta.get('product_count'), '개')}")
        with c6:
            st.info(f"매입처수 : {_fmt_header_num(meta.get('buy_vendor_count') or meta.get('purchase_vendor_count'), '개')}")
        with c7:
            st.info(f"분석월수 : {_fmt_header_num(meta.get('month_count'), '개월')}")
        with c8:
            st.info(f"자료원 : {meta.get('source_label') or '-'}")

        if not show_manufacturer_extended:
            st.caption(
                f"평가월: {meta.get('evaluation_month') or '-'} / "
                f"자료원: {meta.get('source_label') or '-'} / "
                f"조회건수: {meta.get('row_count_total') or meta.get('row_count') or 0}건"
            )
            return

        st.markdown(f"### {meta.get('current_progress_title') or ('당월 진행 요약' if meta.get('evaluation_mode') == 'current_monthly' else '평가월 진행 요약')}")
        p1, p2, p3, p4, p5, p6 = st.columns(6)
        with p1:
            st.info(f"완료월수 : {_fmt_header_num(meta.get('completed_month_count'), '개월')}")
        with p2:
            st.info(f"완료월평균매출 : {_fmt_header_num(meta.get('avg_completed_month_sales_amt'), '원')}")
        with p3:
            st.info(f"{meta.get('current_sales_label') or '당월 현재매출'} : {_fmt_header_num(meta.get('sum_current_month_sales_amt'), '원')}")
        with p4:
            st.warning(f"{meta.get('current_expected_label') or '당월 예상매출'} : {_fmt_header_num(meta.get('sum_current_month_expected_amt'), '원')}")
        with p5:
            st.warning(f"{meta.get('current_remaining_label') or '당월 잔여예상'} : {_fmt_header_num(meta.get('sum_current_month_remaining_expected_amt'), '원')}")
        with p6:
            st.success(f"{meta.get('current_progress_label') or '당월 진척률'} : {_fmt_header_num(meta.get('current_month_progress_pct'), '%', decimals=2)}")

        trend_counts = meta.get("trend_judge_counts") or {}
        st.markdown("### 추세판정별 제약사수")
        j1, j2, j3, j4 = st.columns(4)
        with j1:
            st.success(f"증가 : {_fmt_header_num(trend_counts.get('증가', 0), '개')}")
        with j2:
            st.warning(f"감소 : {_fmt_header_num(trend_counts.get('감소', 0), '개')}")
        with j3:
            st.info(f"안정 : {_fmt_header_num(trend_counts.get('안정', 0), '개')}")
        with j4:
            st.info(f"자료부족 : {_fmt_header_num(trend_counts.get('자료부족', 0), '개')}")

        st.caption(
            f"평가월: {meta.get('evaluation_month') or '-'} / "
            f"자료원: {meta.get('source_label') or '-'} / "
            f"조회건수: {meta.get('row_count_total') or meta.get('row_count') or 0}건"
        )
        return

    customer_group_forecast_types = {"customer_sales_forecast", "salesperson_sales_forecast", "region_sales_forecast"}
    if analysis_type in {"sales_forecast", *customer_group_forecast_types}:
        current_expected = float(meta.get("sum_current_month_expected_amt") or 0)
        current_sales = float(meta.get("sum_current_month_sales_amt") or 0)
        current_progress = meta.get("current_month_progress_pct")
        if current_progress is None:
            current_progress = (current_sales / current_expected * 100) if abs(current_expected) >= 1e-12 else 0

        st.markdown(f"### {meta.get('current_progress_title') or '당월 매출예상 요약'}")

        c1, c2, c3, c4, c5 = st.columns(5)
        with c1:
            st.info(f"완료월평균매출 : {_fmt_header_num(meta.get('avg_completed_month_sales_amt'), '원')}")
        with c2:
            st.info(f"{meta.get('current_sales_label') or '당월 현재매출'} : {_fmt_header_num(meta.get('sum_current_month_sales_amt'), '원')}")
        with c3:
            st.warning(f"{meta.get('current_expected_label') or '당월 예상매출'} : {_fmt_header_num(meta.get('sum_current_month_expected_amt'), '원')}")
        with c4:
            st.warning(f"{meta.get('current_remaining_label') or '당월 잔여예상'} : {_fmt_header_num(meta.get('sum_current_month_remaining_expected_amt'), '원')}")
        with c5:
            st.success(f"{meta.get('current_progress_label') or '당월 진척률'} : {_fmt_header_num(current_progress, '%', decimals=2)}")

        st.markdown("### 중장기 예상")

        f1, f2, f3, f4 = st.columns(4)
        with f1:
            st.info(f"총매출액 : {_fmt_header_num(meta.get('sum_sales_amt'), '원')}")
        with f2:
            st.warning(f"다음월예상매출 : {_fmt_header_num(meta.get('sum_next_month_forecast_amt'), '원')}")
        with f3:
            st.warning(f"3개월예상매출 : {_fmt_header_num(meta.get('sum_3month_forecast_amt'), '원')}")
        with f4:
            st.warning(f"6개월예상매출 : {_fmt_header_num(meta.get('sum_6month_forecast_amt'), '원')}")

        if analysis_type in customer_group_forecast_types:
            s1, s2, s3, s4 = st.columns(4)
            with s1:
                st.info(f"매출처수 : {_fmt_header_num(meta.get('customer_count'), '개')}")
            with s2:
                st.info(f"영업사원수 : {_fmt_header_num(meta.get('salesperson_count'), '명')}")
            with s3:
                st.info(f"지역수 : {_fmt_header_num(meta.get('region_count'), '개')}")
            with s4:
                st.info(f"분석월수 : {_fmt_header_num(meta.get('month_count'), '개월')}")
            st.info(f"자료원 : {meta.get('source_label') or '-'}")

        counts = meta.get("forecast_grade_counts") or {}
        if isinstance(counts, dict) and counts:
            order = ["상승예상", "감소예상", "안정예상", "신규확인", "반품주의", "자료부족"]
            keys = [k for k in order if k in counts]
            keys += [k for k in counts.keys() if k not in keys]
            line = " / ".join(f"{k} {_fmt_header_num(counts.get(k), '개')}" for k in keys)
            st.caption(f"예상등급별 {'매출처수' if analysis_type in customer_group_forecast_types else '제품수'}: {line}")

        return

    if analysis_type == "sales_trend" or "매출 추세" in title:
        st.markdown("### 매출추세요약")

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.info(f"총매출액 : {_fmt_header_num(meta.get('sum_sales_amt'), '원')}")
        with c2:
            st.info(f"출고수량 : {_fmt_header_num(meta.get('sum_qty'), '개')}")
        with c3:
            st.info(f"품목수 : {_fmt_header_num(meta.get('product_count'), '개')}")
        with c4:
            st.info(f"분석월수 : {_fmt_header_num(meta.get('month_count'), '개월')}")

        current_expected = float(meta.get("sum_current_month_expected_amt") or 0)
        current_sales = float(meta.get("sum_current_month_sales_amt") or 0)
        current_progress = meta.get("current_month_progress_pct")
        if current_progress is None:
            current_progress = (current_sales / current_expected * 100) if abs(current_expected) >= 1e-12 else 0

        st.markdown("### 당월 진행 요약")
        p1, p2, p3, p4, p5, p6 = st.columns(6)
        with p1:
            st.info(f"완료월수 : {_fmt_header_num(meta.get('completed_month_count'), '개월')}")
        with p2:
            st.info(f"완료월평균매출 : {_fmt_header_num(meta.get('avg_completed_month_sales_amt'), '원')}")
        with p3:
            st.info(f"당월 현재매출 : {_fmt_header_num(meta.get('sum_current_month_sales_amt'), '원')}")
        with p4:
            st.warning(f"당월 예상매출 : {_fmt_header_num(meta.get('sum_current_month_expected_amt'), '원')}")
        with p5:
            st.warning(f"당월 잔여예상 : {_fmt_header_num(meta.get('sum_current_month_remaining_expected_amt'), '원')}")
        with p6:
            st.success(f"당월 진척률 : {_fmt_header_num(current_progress, '%', decimals=2)}")

# ==========================================================
# 🧩 현재표 source 정리
# ==========================================================
def _clear_current_table_source_after_empty_payload(action: str, payload: Dict[str, Any], reason: str = "") -> None:
    """
    최종 조회가 표 없는 결과(0건/안내문)로 끝났을 때 이전 현재표가 계속 잡히는 것을 방지한다.

    예:
    - 검증 화면 조회 결과가 0건인데 이전 세금계산서 표가 현재표 source로 남는 경우
    - 이후 사용자가 "현재표 불일치 목록"을 입력하면 이전 표를 분석하는 문제 차단
    """
    try:
        ss = st.session_state
        prev_key = ss.get("__sims_current_table_source_key")
        prev_action = ss.get("__sims_current_table_source_action")

        for key in (
            "__sims_current_table_source_key",
            "__sims_current_table_source_action",
            "__sims_current_table_source_analysis_ctx",            
            "__sims_last_table_key",
            "__sims_last_table_action",
            "__sims_context",
            "__sims_context_text",
            "__sims_context_obj",
            "__sims_analysis_ctx",
            "__sims_latest_analysis_key",            
        ):
            ss.pop(key, None)

        log.info(
            "[panel] cleared current table source after empty final payload action=%s prev_key=%s prev_action=%s reason=%s",
            action,
            prev_key,
            prev_action,
            reason or str(payload.get("action") or payload.get("title") or ""),
        )
    except Exception:
        log.exception("[panel] clear current table source after empty final payload failed")


def _payload_has_table_rows(payload: Dict[str, Any], df_full: Any, df_disp: Any) -> bool:
    """payload가 현재표 source로 승격할 수 있는 실제 표 행을 갖는지 판정."""
    try:
        if isinstance(df_disp, pd.DataFrame) and not df_disp.empty:
            return True
        if isinstance(df_full, pd.DataFrame) and not df_full.empty:
            return True

        records = payload.get("records")
        if isinstance(records, list) and len(records) > 0:
            return True

    except Exception:
        pass

    return False


# ==========================================================
# 🧩 결과 렌더링
#    - df/df_display 또는 records/columns 자동 처리
#    - 최종 결과이면 채팅 브리지로도 푸시(테이블/텍스트)
# ==========================================================
def _render_payload(payload: Dict[str, Any], action: str) -> None:
    """payload를 화면에 표시하고, 필요 시 채팅 브리지로 전달 (다운로드 버튼은 폼 밖에서)"""

    final = bool(payload.get("final"))

    # 회사 변경 직후 이전 회사 payload가 rerun/compact 경로로 다시 올라오는 것을 차단한다.
    if final and not _panel_payload_matches_current_company(payload):
        payload_stamp = _panel_payload_company_stamp(payload)
        current_stamp = _panel_current_company_stamp()
        log_state = _panel_stale_payload_log_state(payload_stamp, current_stamp)

        _clear_panel_last_final_payload()

        for key in (
            "__sims_last_final_payload_for_chat",
            "__sims_last_final_payload_for_chat_action",
            "__sims_panel_source_promoted_sig",
            "__sims_last_push_sig",
            "__sims_last_push",
        ):
            st.session_state.pop(key, None)

        st.session_state["__sims_was_final"] = False
        st.session_state["__sims_run_flag"] = False
        st.session_state["__sims_inner_submit"] = False

        log.info(
            "[panel] skip stale final payload after company change action=%s payload_company_id=%s current_company_id=%s db_mismatch=%s",
            action,
            log_state["payload_company_id"],
            log_state["current_company_id"],
            log_state["db_mismatch"],
        )
        return

    if final:
        _panel_stamp_payload_company(payload)

    # Dashboard is a chat-history result, not a panel result. Keep the panel
    # focused on its conditions while preserving every completed Dashboard as
    # an independently renderable message in the active room.
    if final and str(payload.get("type") or "").strip().lower() == "dashboard_lite":
        ss = st.session_state
        meta = dict(payload.get("meta") or {})
        cache = meta.get("dashboard_cache") if isinstance(meta.get("dashboard_cache"), dict) else {}
        signature = str(cache.get("cache_key") or cache.get("query_fingerprint") or "")
        company_id = str(_panel_current_company_stamp().get("company_id") or "").strip()
        room_id = get_current_chat_room_id()
        if not room_id:
            log.warning("[dashboard.chat_push] room_resolved=False pushed=False")
            ss["__sims_inner_submit"] = False
            return
        duplicate_skip = _dashboard_chat_push_is_duplicate(
            ss,
            company_id=company_id,
            room_id=room_id,
            signature=signature,
        )
        # A completed Dashboard payload is produced only by its explicit form
        # submit path.  Do not require the panel's generic submit flag here:
        # Dashboard owns its form and the signature prevents rerun duplicates.
        if not duplicate_skip:
            try:
                room_before_push = next(
                    (
                        room
                        for room in (ss.get("chat_rooms") or [])
                        if isinstance(room, dict) and str(room.get("id") or "") == room_id
                    ),
                    {},
                )
                # Dashboard title generation may already have changed
                # auto_created.  Use the pre-mutation marker instead of
                # inferring pending state from the changed room object.
                was_pending_room = (
                    str(ss.get("__dashboard_lite_pending_room_id") or "") == room_id
                    or bool(room_before_push.get("auto_created"))
                )
                ss.pop("__sims_silent_push", None)
                # Reuse the event created by Dashboard before push. The UUID
                # fallback is only for malformed/legacy payloads.
                event_id = _dashboard_event_id_for_push(payload, meta, cache)
                payload["id"] = event_id
                meta["room_id"] = room_id
                meta["dashboard_event_id"] = event_id
                if isinstance(cache, dict):
                    cache["room_id"] = room_id
                    cache["dashboard_event_id"] = event_id
                    meta["dashboard_cache"] = cache
                active_cache = ss.get("__dashboard_lite_result")
                if isinstance(active_cache, dict):
                    active_cache["room_id"] = room_id
                    active_cache["dashboard_event_id"] = event_id
                payload["meta"] = meta
                if was_pending_room:
                    ss["__chat_room_nav_request"] = {
                        "target_room_id": room_id,
                        "reason": "pending_to_persisted",
                        "request_id": event_id,
                    }
                # push_sims_result_to_chat drains into the in-memory history.
                # The main entrypoint consumes this marker immediately after
                # panel rendering and uses its normal partition save path.
                ss["__dashboard_lite_pending_persist"] = {
                    "room_id": room_id,
                    "event_id": event_id,
                    "was_pending_room": was_pending_room,
                }
                ss["__chat_save_perf_event_id"] = event_id
                ss["__chat_save_dirty_reason"] = "content_changed"
                push_sims_result_to_chat(payload, action)
                primary_event_id = str(
                    active_cache.get("dashboard_event_id")
                    if isinstance(active_cache, dict)
                    else ""
                ).strip()
                payload_event_id = str(payload.get("id") or "").strip()
                meta_event_id = str(meta.get("dashboard_event_id") or "").strip()
                snapshot_event_id = str(
                    cache.get("dashboard_event_id")
                    if isinstance(cache, dict)
                    else ""
                ).strip()
                linked_events = (
                    primary_event_id,
                    payload_event_id,
                    meta_event_id,
                    snapshot_event_id,
                )
                log.info(
                    "[dashboard.event_link] stage=push primary_event_present=%s payload_event_present=%s snapshot_event_present=%s partition_event_present=%s all_equal=%s",
                    bool(primary_event_id),
                    bool(payload_event_id and meta_event_id),
                    bool(snapshot_event_id),
                    False,
                    bool(all(linked_events) and len(set(linked_events)) == 1),
                )
                _remember_dashboard_chat_push_signature(
                    ss,
                    company_id=company_id,
                    room_id=room_id,
                    signature=signature,
                )
                log.info(
                    "[dashboard.chat_push] room_id=%s event_id=%s render_target=chat duplicate_skip=False",
                    room_id,
                    event_id,
                )
            except Exception as exc:
                log.warning("[dashboard.chat_push] render_target=chat pushed=False error_type=%s", type(exc).__name__)
        else:
            log.info(
                "[dashboard.chat_push] room_id=%s event_id= render_target=chat duplicate_skip=%s",
                room_id,
                duplicate_skip,
            )
        ss["__sims_inner_submit"] = False
        ss["__sims_last_render_run_seq"] = ss.get("__sims_run_seq")
        return

    # 메인 컨테이너에서 이번 프레임이 '최종 조회'였는지 판단할 수 있도록 플래그 반영
    st.session_state["__sims_was_final"] = final

    # 제목 / 기본 파일명
    title = payload.get("title") or action
    base  = _safe_filename(title)
    is_sales_trend_payload = _is_sales_trend_payload(payload, action, title)

    # ------------------------------------------------------
    # 📌 파일명 자동화용: 카테고리/액션/타이틀 조합
    # ------------------------------------------------------
    ss = st.session_state
    sel = ss.get("__sims_selected") or {}
    cat = _safe_filename(sel.get("category") or "")
    act = _safe_filename(sel.get("action") or action)

    # 파일명 구성 요소
    file_base_parts = [p for p in [cat, act, title] if p]
    file_base = "_".join(_safe_filename(p) for p in file_base_parts) or base

    # 1) DataFrame 스타일
    df_full = payload.get("df")
    df_disp = payload.get("df_display")

    # 2) records/columns 스타일
    records = payload.get("records")
    columns = payload.get("columns")
    df_from_records: Optional[pd.DataFrame] = None

    if df_full is None and df_disp is None and records is not None and columns is not None:
        try:
            df_from_records = pd.DataFrame.from_records(records, columns=columns)
            df_full = df_from_records
            df_disp = df_from_records
        except Exception:
            log.warning("failed to build DataFrame from records/columns")

    # 화면 표시용 df_display는 조회건수(top)만큼만 제한한다.
    # 원본 df는 다운로드/LLM/현재표 후속분석 기준으로 유지한다.
    _apply_panel_display_limit_to_payload(payload, str(title or action or ""))
    df_full = payload.get("df")
    df_disp = payload.get("df_display")

    # 최종 조회가 표 없는 안내/0건으로 끝난 경우, 이전 현재표 source를 제거한다.
    # 그렇지 않으면 다음 "현재표 ..." 질문이 이전 조회표를 잘못 잡는다.
    if final and not _payload_has_table_rows(payload, df_full, df_disp):
        _clear_current_table_source_after_empty_payload(
            action,
            payload,
            reason="no_table_rows",
        )

        # 0건 조회도 채팅창에 안내 메시지를 1회 올린다.
        # 표가 없다고 skip되면 사용자는 조회가 실행됐는지 알 수 없다.
        try:
            meta = dict(payload.get("meta") or {})
            no_data_msg = str(
                payload.get("message")
                or payload.get("data")
                or meta.get("message")
                or meta.get("empty_message")
                or ""
            ).strip()

            empty_default_msg = (
                "검증 결과 이상 자료가 없습니다."
                if "검증" in str(action or payload.get("action") or payload.get("title") or "")
                else "해당 조회조건의 자료가 없습니다."
            )

            if (
                not no_data_msg
                or no_data_msg in {"None", "nan", "NaN"}
                or no_data_msg in {
                    "조회 결과가 없습니다.",
                    "조회 결과가 없습니다",
                    "해당 자료가 없습니다.",
                    "해당 자료가 없습니다",
                    "해당 자료 없습니다.",
                    "해당 자료 없습니다",
                    "해당 조회조건의 자료가 없습니다.",
                    "해당 조회조건의 자료가 없습니다",
                }
            ):
                no_data_msg = empty_default_msg

            empty_payload = dict(payload)
            empty_payload.update(
                {
                    "final": True,
                    "type": "text",
                    "title": f"📋 조회 결과 — {action}",
                    "action": action,
                    "data": no_data_msg,
                    "df": None,
                    "df_display": None,
                    "records": [],
                    "columns": [],
                }
            )

            meta.update(
                {
                    "row_count": 0,
                    "row_count_total": 0,
                    "display_row_count": 0,
                    "download_row_count": 0,
                    "column_count": 0,
                    "empty_result": True,
                    "_force_push": True,
                    "action": action,
                }
            )

            if not str(meta.get("query_summary") or "").strip():
                qs = str(meta.get("condition") or "").strip()
                if not qs:
                    qs = str(action or "").strip()
                meta["query_summary"] = qs

            empty_payload["meta"] = meta

            _store_panel_final_payload_for_chat(empty_payload, action)

            try:
                sel = st.session_state.get("__sims_selected") or {}
                _remember_panel_final_payload(
                    empty_payload,
                    str(sel.get("category") or ""),
                    str(sel.get("action") or action),
                )
            except Exception:
                pass

            if _panel_result_target_chat_enabled():
                st.info(no_data_msg)
                st.caption("다른 조건으로 다시 조회하려면 아래 조회조건을 수정한 뒤 조회 버튼을 다시 누르세요.")
                return

        except Exception:
            log.exception("[panel] store empty payload for chat failed")


    # ------------------------------------------------------------
    # 채팅/일반 rerun에서는 이미 승격된 패널 결과표를 접어서 렌더 비용을 줄인다.
    # - 최초 패널 조회 결과는 기존처럼 full render
    # - 이후 현재표 후속분석 입력으로 인한 rerun에서는 placeholder만 표시
    # ------------------------------------------------------------
    if (
        isinstance(df_disp, pd.DataFrame)
        and final
        and _should_compact_panel_result_on_rerun(payload, action)
    ):
        try:
            sel = st.session_state.get("__sims_selected") or {}

            # compact 분기로 빠지더라도 채팅 push용 최종 payload는 반드시 보관한다.
            # 특히 그룹코드조회처럼 0건 조회 후 같은 action에서 다시 131건이 조회되는 경우,
            # 여기서 return 해버리면 메인에서 chat.panel.push skip: no df 가 발생한다.
            if _panel_result_target_chat_enabled():
                try:
                    panel_source_sig = _make_panel_source_sig(action, payload)
                    if _panel_chat_push_already_consumed(panel_source_sig):
                        log.info(
                            "[panel] skip duplicate panel render: already pushed action=%s sig=%s",
                            action,
                            panel_source_sig,
                        )
                        _render_panel_chat_only_done(payload, action)
                        return
                    else:
                        df_for_meta: Optional[pd.DataFrame] = None
                        if isinstance(df_full, pd.DataFrame) and not df_full.empty:
                            df_for_meta = df_full
                        elif isinstance(df_disp, pd.DataFrame) and not df_disp.empty:
                            df_for_meta = df_disp

                        if df_for_meta is not None:
                            _enrich_payload_meta_with_basic_stats(payload, df_for_meta)

                        _stash_panel_table_for_current_followup(
                            payload,
                            action,
                            record_previous_source_for_prune=True,
                        )
                        _store_panel_final_payload_for_chat(payload, action)

                except Exception:
                    log.exception("[panel] store compact payload for chat failed")

            _remember_panel_final_payload(
                payload,
                str(sel.get("category") or ""),
                str(sel.get("action") or action),
            )
        except Exception:
            log.exception("[panel] remember final payload before compact render failed")

        _render_compact_panel_result_placeholder(
            payload=payload,
            action=action,
            title=str(title or action or ""),
            df_disp=df_disp,
            df_full=df_full if isinstance(df_full, pd.DataFrame) else None,
        )
        return


    # ------------------------------------------------------------
    # SIMS 패널 조회 결과 표시 정책:
    # - 결과 표는 패널에 직접 렌더링하지 않는다.
    # - 현재표 후속분석용 source/context는 여기서 갱신한다.
    # - 메인에서 채팅방으로 1회 push한다.
    # ------------------------------------------------------------
    if (
        final
        and _panel_result_target_chat_enabled()
        and (
            isinstance(df_disp, pd.DataFrame)
            or isinstance(df_full, pd.DataFrame)
            or isinstance(payload.get("records"), list)
        )
    ):
        try:
            ss = st.session_state
            panel_source_sig = _make_panel_source_sig(action, payload)

            df_for_meta: Optional[pd.DataFrame] = None
            if isinstance(df_full, pd.DataFrame) and not df_full.empty:
                df_for_meta = df_full
            elif isinstance(df_disp, pd.DataFrame) and not df_disp.empty:
                df_for_meta = df_disp

            if df_for_meta is not None:
                _enrich_payload_meta_with_basic_stats(payload, df_for_meta)

            # 현재표 후속분석 source는 반드시 갱신
            if ss.get("__sims_panel_source_promoted_sig") != panel_source_sig:
                _stash_panel_table_for_current_followup(
                    payload,
                    action,
                    record_previous_source_for_prune=True,
                )
                ss["__sims_panel_source_promoted_sig"] = panel_source_sig

                log.info(
                    "[panel] promoted panel result as current source action=%s sig=%s",
                    action,
                    panel_source_sig,
                )

            # 메인 push용 payload 보관
            _store_panel_final_payload_for_chat(payload, action)

            # 패널에는 표/다운로드를 그리지 않고 안내만 표시
            _render_panel_chat_only_done(payload, action)

            try:
                sel = st.session_state.get("__sims_selected") or {}
                _remember_panel_final_payload(
                    payload,
                    str(sel.get("category") or ""),
                    str(sel.get("action") or action),
                )
            except Exception:
                log.exception("[panel] remember final payload for chat-only failed")

            return

        except Exception:
            log.exception("[panel] chat-only result handling failed")

    # 화면 표시
    if isinstance(df_disp, pd.DataFrame):
        st.subheader(title)

        meta = payload.get("meta") or {}
        _render_panel_result_compact_header(payload, action, str(title or action or ""), df_disp)

        if df_disp.empty:
            data = payload.get("data")
            if isinstance(data, str) and data.strip():
                st.warning(data)
            else:
                st.info("해당 조회조건의 자료가 없습니다.")

            if final:
                st.info("조회 완료: 0건")
        else:
            is_flow = "제품수불현황" in str(title)
            is_inventory = "제품재고현황" in str(title)

            if is_flow:
                _render_stock_flow_summary(payload, df_disp)

            if is_inventory:
                _render_inventory_summary(payload, df_disp)

            render_df = df_disp.copy()

            if _is_io_category():
                # SIMS 조회 화면의 IO/명세서/재고 표도 공용 표 렌더러를 사용한다.
                # 목적:
                # - 좌측 고정 pinned 적용
                # - 채팅/NLQ 결과표와 컬럼폭/높이/순번 처리 통일
                # - Styler 경로에서는 Streamlit pinned가 적용되지 않으므로 column_config 경로로 전환
                render_df = _prepare_io_display_df(df_disp, add_row_no=True)
                render_df = _normalize_zero_like_df(render_df)
                safe_df = _make_styler_safe_df(render_df)

                try:
                    view_df, column_config, table_width, table_height = build_sims_table_display_config(
                        safe_df,
                        action_name=str(action or title or ""),
                        meta=payload.get("meta") or {},
                        add_row_no=False,      # _prepare_io_display_df에서 이미 순번 보강
                        row_no_name="순번",
                        enable_pinning=True,
                        max_pinned_cols=4,
                        min_width=720,
                        max_width=1650,
                        min_height=170,
                        max_height=520,
                        row_height=32,
                    )

                    render_df = view_df
                    log_sims_display_fields(safe_df, view_df, action=str(action or title or ""), render_path="panel", mode="small")

                    log.debug(
                        "[panel] io common table render action=%s rows=%s cols=%s natural_width=%s height=%s pinned=%s",
                        action,
                        len(view_df),
                        len(view_df.columns),
                        table_width,
                        table_height,
                        True,
                    )

                    st.dataframe(
                        view_df,
                        width="stretch",
                        hide_index=True,
                        height=table_height,
                        column_config=column_config if column_config else None,
                    )

                except Exception:
                    log.exception("[PANEL_IO_COMMON_RENDER_FAIL] action=%s", action)

                    st.dataframe(
                        safe_df,
                        width="stretch",
                        hide_index=True,
                        height=520,
                    )

            else:
                view_df = _trim_object_columns(df_disp.copy())

                if is_sales_trend_payload:
                    view_df = _prepare_sales_trend_display_df(
                        view_df,
                        payload=payload,
                        action=action,
                        title=title,
                    )

                    # 분석/KPI 표는 최종 표시 직전에 숫자 포맷을 문자열로 확정한다.
                    # Styler.format / column_config 경로가 안 먹는 경우를 우회한다.

                def _clean_sample_text(v: Any) -> str:
                    if v is None:
                        return ""
                    try:
                        if pd.isna(v):
                            return ""
                    except Exception:
                        pass
                    s = str(v)
                    s = s.replace("\u00a0", " ")
                    s = re.sub(r"\s+", " ", s).strip()
                    if s in {"None", "nan", "<NA>", "NaT"}:
                        return ""
                    return s

                def _infer_width_px(df_src: pd.DataFrame, col_name: str) -> int:
                    s = str(col_name or "").strip()
                    header_len = len(s)

                    if _is_large_table_for_fast_render(df_src):
                        return _fast_width_px_for_large_table(df_src, col_name)

                    try:
                        sample = (
                            df_src[col_name]
                            .head(200)
                            .map(_clean_sample_text)
                            .tolist()
                        )
                        max_len = max([header_len] + [len(x) for x in sample])
                    except Exception:
                        max_len = header_len

                    # 0) 컬럼명 기준 고정 우선폭
                    fixed_width_map = {
                        # 공통 코드/ID
                        "순번": 60,
                        "그룹코드": 90,
                        "상세코드": 90,
                        "항목코드": 90,
                        "사용자코드": 95,
                        "사용자ID": 150,
                        "사번": 95,
                        "제품코드": 95,
                        "거래처코드": 95,
                        "보험코드": 95,
                        "바코드": 130,

                        # 명칭/이름
                        "사용자명": 180,
                        "부서명": 120,
                        "직책": 95,
                        "영업지역": 120,
                        "재고위치": 110,
                        "거래처명": 220,
                        "제품명": 220,
                        "상품명": 220,
                        "한글명": 220,
                        "코드종류": 170,
                        "영문명": 170,
                        "약칭": 120,

                        # 등록/수정
                        "등록자": 120,
                        "수정자": 120,
                        "등록일자": 95,
                        "수정일자": 95,
                        "등록일시": 150,
                        "수정일시": 150,

                        # 연락처/기타
                        "이메일": 220,
                        "휴대폰": 120,
                        "전화번호": 120,
                        "사무실전화": 120,
                        "기타1": 180,
                        "기타2": 180,
                        "기타3": 180,
                        "비고": 260,
                        "설명": 260,
                        "메모": 260,
                        "주소": 320,

                        # 숫자성
                        "계산단위": 90,
                        "보험가격": 110,
                        "보험단가": 110,
                        "이전보험가격": 120,
                        "이전보험단가": 120,
                        "단가": 100,
                    }

                    if s in fixed_width_map:
                        return fixed_width_map[s]

                    # 1) 날짜/시간
                    if any(k in s for k in ("등록일자", "수정일자", "일자", "날짜")):
                        return 95
                    if any(k in s for k in ("등록일시", "수정일시", "일시", "시간")):
                        return 150

                    # 2) 코드류
                    if _is_code_col_name(s):
                        return 90

                    # 3) 숫자류
                    if _is_numeric_col(df_src, col_name):
                        if max_len <= 6:
                            return 90
                        if max_len <= 12:
                            return 110
                        return 130

                    # 4) 이름/명칭류
                    if any(k in s for k in ("명", "이름", "거래처", "사용자", "제품", "부서", "직책", "영업지역")):
                        if max_len <= 8:
                            return 110
                        if max_len <= 16:
                            return 150
                        return 220

                    # 5) 설명/주소/비고류
                    if any(k in s for k in ("주소", "비고", "설명", "메모", "기타")):
                        if max_len <= 20:
                            return 180
                        if max_len <= 40:
                            return 260
                        return 360

                    # 6) 일반 텍스트
                    px = 28 + max_len * 9
                    if px < 85:
                        px = 85
                    if px > 320:
                        px = 320
                    return int(px)
                
                def _number_col_config(df_src: pd.DataFrame, name: str):
                    s = str(name or "").strip()
                    if not _is_numeric_col(df_src, name):
                        return None

                    width_px = _infer_width_px(df_src, name)

                    if s == "계산단위":
                        return st.column_config.NumberColumn(
                            s,
                            format="localized",
                            step=0.001,
                            width=width_px,
                        )

                    if s in {
                        "당월 진척률",
                        "최근3개월증감률",
                        "적용증감률",
                        "월시점 증감률",
                        "월시점 적용증감률",
                        "월시점 달성률",
                        "최근3개월수량증감률",
                        "수요증감률",
                        "수요적용증감률",
                        "평가월 수요진척률",
                        "당월 출고진척률",
                        "당월 재고충족률",
                    }:
                        return st.column_config.NumberColumn(
                            s,
                            format="%.2f%%",
                            step=0.01,
                            width=width_px,
                        )

                    if _is_decimal_col_name(s):
                        return st.column_config.NumberColumn(
                            s,
                            format="localized",
                            step=0.01,
                            width=width_px,
                        )

                    return st.column_config.NumberColumn(
                        s,
                        format="localized",
                        step=1,
                        width=width_px,
                    )

                def _text_col_config(df_src: pd.DataFrame, name: str):
                    return st.column_config.TextColumn(
                        str(name or ""),
                        width=_infer_width_px(df_src, name),
                    )

                column_config = {}
                for c in view_df.columns:
                    num_cfg = _number_col_config(view_df, c)
                    if num_cfg is not None:
                        column_config[c] = num_cfg
                    else:
                        column_config[c] = _text_col_config(view_df, c)

                # 품목별 매출 추세 분석 전용: 좌측 주요 컬럼 고정
                if is_sales_trend_payload:
                    pinned_widths = {
                        "순번": 60,
                        "기준월": 90,
                        "제품코드": 95,
                        "제품명": 260,
                        "규격": 110,
                        "제조사명": 160,
                    }

                    for c, w in pinned_widths.items():
                        if c in view_df.columns:
                            column_config[c] = _pinned_text_column_config(c, w)

                table_width = int(
                    min(
                        max(sum(_infer_width_px(view_df, c) for c in view_df.columns) + 80, 900),
                        2600,
                    )
                )

                if is_sales_trend_payload:
                    _render_simple_analysis_header(payload)

                    table_mode_info = {"mode": "fast" if _is_large_table_for_fast_render(view_df) else "small"}
                    try:
                        table_mode_info = log_sims_table_mode(view_df, action=action, render_path="panel")
                    except Exception:
                        log.debug("[sims.table_mode] panel log failed", exc_info=True)
                    if str(table_mode_info.get("mode") or "") == "fast":
                        st.caption("빠른 표 모드: 분석/KPI 큰 표는 속도를 위해 셀 색상/굵은 글씨 서식을 생략합니다.")
                        log_sims_table_render(
                            view_df,
                            action=action,
                            render_path="panel",
                            mode="fast",
                            renderer="_render_fast_dataframe",
                            height=520,
                            visible_rows=min(int(len(view_df)), 300),
                            width_mode="fast",
                            column_config_count=0,
                        )
                        _render_fast_dataframe(
                            view_df,
                            height=520,
                            action_name=action,
                            meta=meta,
                        )

                    else:
                        log_sims_display_fields(
                            df_disp if isinstance(df_disp, pd.DataFrame) else view_df,
                            view_df,
                            action=action,
                            render_path="panel",
                            mode="small",
                        )
                        log_sims_table_render(
                            view_df,
                            action=action,
                            render_path="panel",
                            mode="small",
                            renderer="st.dataframe",
                            height=520,
                            visible_rows=min(int(len(view_df)), max(int((520 - 48) / 32), 0)),
                            width_mode="width",
                            column_config_count=len(column_config or {}),
                        )
                        try:
                            st.dataframe(
                                view_df,
                                width=table_width,
                                hide_index=True,
                                height=520,
                                column_config=column_config if column_config else None,
                            )
                        except Exception:
                            log.warning("[PANEL_SALES_TREND_STYLE_FAIL] fallback to plain dataframe")
                            st.dataframe(
                                view_df,
                                width=table_width,
                                hide_index=True,
                                height=520,
                                column_config=column_config if column_config else None,
                            )
                else:
                    if _is_large_table_for_fast_render(view_df):
                        st.caption("빠른 표 모드: 큰 표는 속도를 위해 셀 색상/굵은 글씨 서식을 생략합니다.")
                        _render_fast_dataframe(
                            view_df,
                            height=420,
                            action_name=action,
                            meta=meta,
                        )
                    else:
                        st.dataframe(
                            view_df,
                            width=table_width,
                            hide_index=True,
                            height=420,
                            column_config=column_config if column_config else None,
                        )

            display_count = 0
            try:
                base_df = render_df if isinstance(render_df, pd.DataFrame) else df_disp
                if isinstance(base_df, pd.DataFrame):
                    display_count = len(base_df)
            except Exception:
                display_count = 0

            try:
                if is_flow:
                    if (
                        isinstance(df_disp, pd.DataFrame)
                        and "입출고일자" in df_disp.columns
                        and len(df_disp) > 0
                        and str(df_disp.iloc[0]["입출고일자"]).strip() == "이월재고"
                    ):
                        display_count -= 1

                if is_inventory:
                    meta = payload.get("meta") or {}
                    group_label = str(meta.get("group_label") or "").strip()
                    if (
                        isinstance(render_df, pd.DataFrame)
                        and group_label
                        and group_label in render_df.columns
                        and len(render_df) > 0
                        and str(render_df.iloc[-1][group_label]).strip() == "합계"
                    ):
                        display_count -= 1
            except Exception:
                pass

            # ------------------------------------------------------------
            # 패널 조회 결과를 현재표 후속분석 source로 등록
            # ------------------------------------------------------------
            # Panel(A) 조회 결과는 채팅에 표를 중복 push하지 않기 때문에,
            # push_sims_result_to_chat()을 타지 않는 경우가 있다.
            # 이 경우 화면에는 패널 표가 조회되어도
            # 채팅의 "현재표 ..." 후속분석 source는 이전 NLQ/이전표로 남는다.
            #
            # 같은 run_seq/action에서는 1회만 승격하고,
            # 이후 일반 rerun에서는 위쪽 compact placeholder가 동작한다.
            if final:
                try:
                    ss = st.session_state
                    panel_source_sig = _make_panel_source_sig(action, payload)

                    if ss.get("__sims_panel_source_promoted_sig") != panel_source_sig:
                        df_for_meta: Optional[pd.DataFrame] = None
                        if isinstance(df_full, pd.DataFrame) and not df_full.empty:
                            df_for_meta = df_full
                        elif isinstance(df_disp, pd.DataFrame) and not df_disp.empty:
                            df_for_meta = df_disp

                        if df_for_meta is not None:
                            _enrich_payload_meta_with_basic_stats(payload, df_for_meta)

                        _stash_panel_table_for_current_followup(
                            payload,
                            action,
                            record_previous_source_for_prune=True,
                        )

                        ss["__sims_panel_source_promoted_sig"] = panel_source_sig

                        log.info(
                            "[panel] promoted panel result as current source action=%s sig=%s",
                            action,
                            panel_source_sig,
                        )

                except Exception:
                    log.exception("[panel] promote panel current source failed")

            # 채팅 브리지 푸시
            # - 실제 조회 submit 런에서만 push
            # - 단순 rerun(채팅 Enter 등)에서는 이전 payload 재push 금지
            if final and bool(st.session_state.get("__sims_inner_submit")):
                try:
                    df_for_meta: Optional[pd.DataFrame] = None
                    if isinstance(df_full, pd.DataFrame) and not df_full.empty:
                        df_for_meta = df_full
                    elif isinstance(df_disp, pd.DataFrame) and not df_disp.empty:
                        df_for_meta = df_disp

                    if df_for_meta is not None:
                        _enrich_payload_meta_with_basic_stats(payload, df_for_meta)

                    # 패널 조회 결과는 패널 안에서 이미 표/버튼을 렌더한다.
                    # 채팅에도 같은 표를 push하면 표와 LLM 버튼이 위/아래로 중복된다.
                    # 따라서:
                    # 1) 현재표 후속분석용 table_key는 직접 stash
                    # 2) push_sims_result_to_chat는 silent 모드로 호출해 LLM 컨텍스트만 갱신
                    _stash_panel_table_for_current_followup(payload, action)
                    st.session_state["__sims_silent_push"] = True
                    push_sims_result_to_chat(payload, action)

                    # 이번 submit은 1회만 소비
                    st.session_state["__sims_inner_submit"] = False
                    st.session_state["__sims_last_render_run_seq"] = st.session_state.get("__sims_run_seq")

                except Exception:
                    log.exception("push_sims_result_to_chat failed")

            if final:
                st.success(f"조회 완료: {max(display_count, 0)}건")

                try:
                    sel = st.session_state.get("__sims_selected") or {}
                    _remember_panel_final_payload(
                        payload,
                        str(sel.get("category") or ""),
                        str(sel.get("action") or action),
                    )
                except Exception:
                    log.exception("[panel] remember final payload before downloads failed")

                _render_downloads(
                    df_disp,
                    file_base,
                    df_full=df_full if isinstance(df_full, pd.DataFrame) else None,
                    action_name=title,
                )



def _enrich_payload_meta_with_basic_stats(payload: Dict[str, Any], df: pd.DataFrame) -> None:
    """LLM이 활용할 수 있는 기본 요약(meta.summary_text 등)을 자동 추가"""
    try:
        meta = payload.setdefault("meta", {})
        row_count = int(len(df))
        col_count = int(len(df.columns))

        meta.setdefault("row_count", row_count)
        meta.setdefault("column_count", col_count)

        cols = list(map(str, df.columns))
        dept_col = next((c for c in cols if "부서" in c), None)
        date_col = next((c for c in cols if "일자" in c or "등록" in c), None)

        summary_parts = [f"행 수: {row_count}건, 열 수: {col_count}개"]

        # 부서 컬럼 요약
        if dept_col:
            top5 = df[dept_col].value_counts().head(5).to_dict()
            meta.setdefault("dept_column", dept_col)
            meta.setdefault("dept_top5", top5)
            summary_parts.append(f"부서 컬럼 '{dept_col}' 상위 5개 분포: {top5}")

        # 날짜 컬럼 요약
        if date_col:
            meta.setdefault("date_column", date_col)
            summary_parts.append(f"날짜/등록 관련 컬럼: {date_col}")

        summary_text = " / ".join(summary_parts)
        meta.setdefault("summary_text", summary_text)

    except Exception:
        log.exception("[panel.meta] enrich meta failed")

@st.fragment
def _render_panel_result_actions_fragment(
    *,
    key_suffix: str,
    csv_bytes: bytes,
    csv_name: str,
    excel_bytes: Optional[bytes],
    xlsx_name: str,
    prompt: str,
) -> None:
    """
    SIMS 패널 결과 하단 액션 영역.

    - CSV / EXCEL 다운로드는 EXPORT_EXCEL 권한 필요
    - LLM 분석은 다운로드 권한과 별개로 유지
    """
    can_export_excel = require_permission("EXPORT_EXCEL", show_error=False)

    c1, c2, c3 = st.columns(3)

    with c1:
        if can_export_excel:
            st.download_button(
                "🧾 CSV로 저장",
                data=csv_bytes,
                file_name=csv_name,
                mime="text/csv",
                key=f"__panel_dl_csv_{key_suffix}",
                width="stretch",
            )
        else:
            st.button(
                "🔒 CSV 저장",
                disabled=True,
                key=f"__panel_dl_csv_disabled_{key_suffix}",
                width="stretch",
                help=_export_unavailable_help(),
            )

    with c2:
        if can_export_excel:
            if excel_bytes is not None:
                st.download_button(
                    "⬇ Excel로 저장",
                    data=excel_bytes,
                    file_name=xlsx_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"__panel_dl_xlsx_{key_suffix}",
                    width="stretch",
                )
            else:
                st.caption("Excel 엔진 없음")
        else:
            st.button(
                "🔒 Excel 저장",
                disabled=True,
                key=f"__panel_dl_xlsx_disabled_{key_suffix}",
                width="stretch",
                help=_export_unavailable_help(),
            )

    with c3:
        run_llm = st.button(
            "🤖 LLM 분석",
            key=f"__panel_llm_analysis_fragment_btn_{key_suffix}",
            width="stretch",
        )

    if not can_export_excel:
        st.info(_export_unavailable_message())

    if run_llm:
        _run_panel_llm_analysis_from_button(prompt, key_suffix)

# ==========================================================
# 🧩 결과 렌더링 - 액션 영역 (다운로드/LLM
#    - 패널 결과 하단에 CSV/EXCEL/LLM 버튼 표시
#    - LLM 분석은 버튼 클릭 시 fragment 내부에서 실행하여 전체 앱 rerun 최소화
# ==========================================================
def _get_panel_download_lazy_threshold_rows() -> int:
    """
    패널 대형표 다운로드 lazy 기준 행 수.
    기본 5,000건 이상이면 CSV/XLSX bytes를 즉시 만들지 않는다.
    """
    try:
        return max(0, int(os.getenv("SIMS_DOWNLOAD_LAZY_ROW_THRESHOLD", "5000")))
    except Exception:
        return 5000


def _run_panel_llm_analysis_from_button(prompt: str, key_suffix: str) -> None:
    runner = st.session_state.get("__sims_llm_analysis_runner")
    log.info("[panel.fragment] LLM analysis clicked runner=%s key=%s", callable(runner), key_suffix)

    if not callable(runner):
        st.warning("LLM 분석 실행기가 아직 준비되지 않았습니다. 잠시 후 다시 눌러 주세요.")
        return

    sel = st.session_state.get("__sims_selected") or {}
    category = str(sel.get("category") or "")
    action = str(sel.get("action") or "")

    payload = _get_panel_last_final_payload(category, action)
    if not isinstance(payload, dict):
        st.warning("현재 SIMS 조회 결과를 LLM 분석 컨텍스트로 준비하지 못했습니다. 조회를 다시 실행한 뒤 눌러 주세요.")
        return

    meta = payload.get("meta") or {}
    table_key = str(meta.get("table_key") or meta.get("download_table_key") or payload.get("table_key") or "").strip()
    clicked_action = str(payload.get("action") or meta.get("action") or action or payload.get("title") or "").strip()
    if not table_key:
        st.warning("선택한 표의 식별값을 찾을 수 없어 LLM 분석을 실행하지 않았습니다.")
        return

    ss = st.session_state
    restore_keys = (
        "__sims_current_table_source_key",
        "__sims_current_table_source_action",
        "__sims_current_table_source_analysis_ctx",
        "__sims_last_table_key",
        "__sims_last_table_action",
        "__sims_analysis_ctx",
    )
    before_state = {k: ss.get(k) for k in restore_keys if k in ss}
    missing_before = {k for k in restore_keys if k not in ss}

    ok_ctx = False
    try:
        ok_ctx = _ensure_panel_llm_context_from_payload(
            category=category,
            action=action,
        )
    finally:
        for k in restore_keys:
            if k in missing_before:
                ss.pop(k, None)
            elif k in before_state:
                ss[k] = before_state[k]

    if not ok_ctx:
        st.warning("현재 SIMS 조회 결과를 LLM 분석 컨텍스트로 준비하지 못했습니다. 조회를 다시 실행한 뒤 눌러 주세요.")
        return

    cache = ss.get("__sims_analysis_ctx_by_table_key") or {}
    analysis_ctx = cache.get(table_key)
    mismatch_reason = _sims_clicked_llm_context_mismatch(analysis_ctx, table_key, clicked_action)
    log.info(
        "[sims.analysis.button] source=panel table_key=%s action=%s context_action=%s mismatch=%s",
        table_key,
        clicked_action,
        analysis_ctx.get("action") if isinstance(analysis_ctx, dict) else "",
        mismatch_reason or "",
    )
    if mismatch_reason:
        st.warning("선택한 표의 LLM 분석 컨텍스트가 유효하지 않아 분석을 실행하지 않았습니다.")
        return

    st.markdown("#### 🤖 LLM 분석 결과")

    try:
        runner(
            prompt,
            analysis_ctx_override=analysis_ctx,
            clicked_table_key=table_key,
            clicked_action=clicked_action,
            clicked_message_id="",
        )
    except Exception:
        log.exception("[panel.fragment] LLM analysis failed")
        st.error("LLM 분석 중 오류가 발생했습니다.")


def _render_panel_result_actions_lazy(
    *,
    key_suffix: str,
    download_df: pd.DataFrame,
    csv_name: str,
    xlsx_name: str,
    prompt: str,
) -> None:
    """
    SIMS 패널 결과 하단 액션 영역 lazy 버전.

    - 작은 표: 기존처럼 CSV/EXCEL/LLM 버튼 즉시 표시
    - 큰 표: [Excel 다운로드 준비]를 눌렀을 때만 CSV/XLSX bytes 생성
    """
    if not isinstance(download_df, pd.DataFrame) or download_df.empty:
        c1, c2, c3 = st.columns(3)
        with c1:
            st.caption("다운로드할 표가 없습니다.")
        with c3:
            run_llm = st.button(
                "🤖 LLM 분석",
                key=f"__panel_llm_analysis_lazy_empty_{key_suffix}",
                width="stretch",
            )

        if run_llm:
            _run_panel_llm_analysis_from_button(prompt, key_suffix)

        return

    threshold_rows = _get_panel_download_lazy_threshold_rows()
    row_count = int(len(download_df))
    col_count = int(len(download_df.columns))

    can_export_excel = require_permission("EXPORT_EXCEL", show_error=False)

    if not can_export_excel:
        st.info(_export_unavailable_message())

        c1, c2, c3 = st.columns(3)

        with c1:
            st.button(
                "🔒 CSV 저장",
                disabled=True,
                key=f"__panel_dl_csv_lazy_disabled_{key_suffix}",
                width="stretch",
                help=_export_unavailable_help(),
            )

        with c2:
            st.button(
                "🔒 Excel 저장",
                disabled=True,
                key=f"__panel_dl_xlsx_lazy_disabled_{key_suffix}",
                width="stretch",
                help=_export_unavailable_help(),
            )

        with c3:
            run_llm = st.button(
                "🤖 LLM 분석",
                key=f"__panel_llm_analysis_lazy_no_export_{key_suffix}",
                width="stretch",
            )

        if run_llm:
            _run_panel_llm_analysis_from_button(prompt, key_suffix)

        return


    supplier_detail_key = str(getattr(download_df, "attrs", {}).get("supplier_detail_key") or "").strip()
    force_lazy_supplier_excel = bool(supplier_detail_key)
    is_large_download = force_lazy_supplier_excel or (threshold_rows > 0 and row_count >= threshold_rows)
    cache_key_suffix = f"{key_suffix}::{supplier_detail_key}" if supplier_detail_key else key_suffix

    ready_key = f"__panel_download_ready::{cache_key_suffix}"
    bytes_key = f"__panel_download_bytes::{cache_key_suffix}"

    ss = st.session_state
    is_ready = bool(ss.get(ready_key))

    if is_large_download and not is_ready:
        st.caption(
            f"대형표 다운로드: {row_count:,}건 × {col_count:,}열입니다. "
            "속도를 위해 CSV/EXCEL 파일은 [Excel 다운로드 준비]를 누른 뒤 생성합니다."
        )

        c1, c2, c3 = st.columns(3)

        with c1:
            if st.button(
                "Excel 다운로드 준비",
                key=f"__panel_prepare_download_{key_suffix}",
                width="stretch",
            ):
                ss[ready_key] = True
                st.rerun()

        with c2:
            st.caption("CSV/EXCEL 버튼은 준비 후 표시됩니다.")

        with c3:
            run_llm = st.button(
                "🤖 LLM 분석",
                key=f"__panel_llm_analysis_lazy_{key_suffix}",
                width="stretch",
            )

        if run_llm:
            _run_panel_llm_analysis_from_button(prompt, key_suffix)

        return

    cached = ss.get(bytes_key)
    col_sig = tuple(str(c) for c in download_df.columns)

    if (
        isinstance(cached, dict)
        and cached.get("rows") == row_count
        and cached.get("cols") == col_count
        and tuple(cached.get("col_sig") or []) == col_sig
        and isinstance(cached.get("csv_bytes"), (bytes, bytearray))
    ):
        csv_bytes = cached["csv_bytes"]
        xbytes = cached.get("excel_bytes")
        log.debug(
            "[panel.download] bytes cache hit key=%s rows=%s cols=%s",
            key_suffix,
            row_count,
            col_count,
        )
    else:
        t0 = dt.datetime.now()

        csv_buf = io.StringIO()
        download_df.to_csv(csv_buf, index=False, encoding="utf-8-sig")
        csv_bytes = csv_buf.getvalue().encode("utf-8-sig")

        xbytes = _xlsx_bytes(download_df)

        ss[bytes_key] = {
            "rows": row_count,
            "cols": col_count,
            "col_sig": col_sig,
            "csv_bytes": csv_bytes,
            "excel_bytes": xbytes,
        }

        try:
            elapsed = (dt.datetime.now() - t0).total_seconds()
        except Exception:
            elapsed = 0

        log.info(
            "[panel.download] bytes prepared lazy=%s rows=%s cols=%s %.3fs",
            is_large_download,
            row_count,
            col_count,
            elapsed,
        )

    _render_panel_result_actions_fragment(
        key_suffix=key_suffix,
        csv_bytes=csv_bytes,
        csv_name=csv_name,
        excel_bytes=xbytes,
        xlsx_name=xlsx_name,
        prompt=prompt,
    )

# 📥 다운로드 / LLM 분석 버튼 렌더링 (폼 외부, 패널 최하단)
# - CSV/XLSX 다운로드 버튼과 함께, LLM 분석 버튼도 같이 렌더링한다.
def _render_downloads(
    df: pd.DataFrame,
    base: str,
    *,
    df_full: Optional[pd.DataFrame] = None,
    action_name: str = "",
) -> None:
    """폼 외부에서 CSV/XLSX/LLM 분석 버튼을 한 줄로 렌더.

    원칙:
    - 화면 표시는 df_display 기준
    - Excel/CSV 다운로드는 df_full이 있으면 df_full 우선
    - df_full이 없으면 기존처럼 화면 df 사용
    """
    if df is None:
        return

    display_df = df
    download_df = df

    if (
        isinstance(df_full, pd.DataFrame)
        and not df_full.empty
        and (
            not isinstance(display_df, pd.DataFrame)
            or len(df_full) >= len(display_df)
        )
    ):
        download_df = df_full

    # 패널 stash 단계에서 전체 export DF가 session에 저장된 경우,
    # _render_downloads()에 전달된 df_full보다 session의 전체 DF를 우선 사용한다.
    try:
        ss = st.session_state
        table_key = str(
            ss.get("__sims_current_table_source_key")
            or ss.get("__sims_last_table_key")
            or ""
        ).strip()

        if table_key:
            by_key = ss.get("__sims_export_tables_by_key") or {}
            export_tables = ss.get("sims_export_tables") or {}

            session_full_df = None
            if isinstance(by_key, dict):
                session_full_df = by_key.get(table_key)
            if not isinstance(session_full_df, pd.DataFrame) and isinstance(export_tables, dict):
                session_full_df = export_tables.get(table_key)

            if (
                isinstance(session_full_df, pd.DataFrame)
                and not session_full_df.empty
                and isinstance(display_df, pd.DataFrame)
                and len(session_full_df) > len(display_df)
            ):
                df_full = session_full_df
                download_df = session_full_df

    except Exception:
        log.exception("[panel.download] use session full df failed")

    display_rows = int(len(display_df)) if isinstance(display_df, pd.DataFrame) else 0
    download_rows = int(len(download_df)) if isinstance(download_df, pd.DataFrame) else 0
    display_cols = int(len(display_df.columns)) if isinstance(display_df, pd.DataFrame) else 0
    download_cols = int(len(download_df.columns)) if isinstance(download_df, pd.DataFrame) else 0

    log.info(
        "[panel.download] uses %s df action=%s rows=%s display_rows=%s cols=%s display_cols=%s",
        "full" if download_df is df_full else "display",
        str(action_name or base),
        download_rows,
        display_rows,
        download_cols,
        display_cols,
    )

    if download_rows > display_rows:
        st.caption(
            f"CSV/EXCEL 다운로드 기준: 전체 조회조건 {download_rows:,}건 "
            f"(화면 표시 {display_rows:,}건)"
        )
    elif download_cols > display_cols:
        st.caption(
            f"CSV/EXCEL 다운로드 기준: 전체 원본 컬럼 {download_cols:,}개 "
            f"(화면 표시 컬럼 {display_cols:,}개)"
        )

    ns = f"{st.session_state.get('__sims_form_id')}_{st.session_state.get('__sims_run_seq')}_{_safe_filename(base)}"
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")


    try:
        prompt = _build_sims_detail_analysis_prompt(
            action_name=str(action_name or base or ""),
            display_rows=display_rows,
            download_rows=download_rows,
            expected_rows=download_rows,
        )
    except Exception:
        log.exception("[panel] build detail analysis prompt failed")
        prompt = (
            f"현재 SIMS 패널 조회 결과({base})를 전체 조회조건 기준으로 분석해줘. "
            "핵심 요약, 주요 수치, 주의/확인할 점, 다음 조회 제안 순서로 정리해줘. "
            "내부 영문 key 이름은 답변에 노출하지 마라."
        )

    _render_panel_result_actions_lazy(
        key_suffix=ns,
        download_df=download_df,
        csv_name=f"{base}_{ts}.csv",
        xlsx_name=f"{base}_{ts}.xlsx",
        prompt=prompt,
    )    

def _write_supplier_stock_shortage_excel_if_any(writer: Any, df: pd.DataFrame, engine: str) -> bool:
    attrs = getattr(df, "attrs", {}) if isinstance(df, pd.DataFrame) else {}
    detail_df = attrs.get("supplier_detail_df") if isinstance(attrs, dict) else None
    if not isinstance(detail_df, pd.DataFrame):
        detail_key = str((attrs or {}).get("supplier_detail_key") or "").strip()
        detail_store = st.session_state.get("__sims_supplier_stock_shortage_detail_tables") or {}
        if detail_key and isinstance(detail_store, dict):
            detail_df = detail_store.get(detail_key)
    if not isinstance(detail_df, pd.DataFrame) or detail_df.empty:
        return False
    summary_df = df.copy()
    detail_excel_df = detail_df.copy()
    summary_df.to_excel(writer, index=False, sheet_name="매입처별요약")
    detail_excel_df.to_excel(writer, index=False, sheet_name="제품매입처상세")
    _apply_sims_excel_number_formats(writer, summary_df, "매입처별요약", engine)
    _apply_sims_excel_number_formats(writer, detail_excel_df, "제품매입처상세", engine)
    return True


def _xlsx_bytes(df: pd.DataFrame) -> Optional[bytes]:
    """xlsxwriter 또는 openpyxl이 있으면 XLSX 바이트를 반환, 없으면 None."""
    try:
        import xlsxwriter  # noqa: F401
        engine = "xlsxwriter"
    except Exception:
        try:
            import openpyxl  # noqa: F401
            engine = "openpyxl"
        except Exception:
            return None
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine=engine) as w:
        if not _write_supplier_stock_shortage_excel_if_any(w, df, engine):
            df.to_excel(w, index=False, sheet_name="SIMS")
            _apply_sims_excel_number_formats(w, df, "SIMS", engine)
    return buf.getvalue()


def _apply_sims_excel_number_formats(writer: Any, df: pd.DataFrame, sheet_name: str, engine: str) -> None:
    money_cols = {
        "완료월총매출",
        "월평균매출",
        "완료월평균매출",
        "당월 현재매출",
        "당월 예상매출",
        "당월 잔여예상",
        "전월대비매출",
        "총매출공급가액",
        "매출공급가액",
        "매출세액",
        "매출합계",
        "평가월 현재매출",
        "평가월 예상매출",
        "평가월 잔여예상",
        "다음월예상매출",
        "3개월예상매출",
        "6개월예상매출",
        "부족예상금액",
    }
    decimal_money_cols = {
        "최근3개월평균매출",
        "최근6개월평균매출",
        "평균공급단가",
        "완료월평균출고수량",
        "최근3개월평균출고수량",
        "최근6개월평균출고수량",
        "당월 예상출고수량",
        "당월 잔여예상출고수량",
        "예상월말재고수량",
        "부족예상수량",
    }
    percent_cols = {
        "당월 진척률",
        "평가월 진척률",
        "전월대비매출증감률",
        "최근3개월증감률",
        "월시점 최근3개월증감률",
        "적용증감률",
        "월시점 증감률",
        "월시점 적용증감률",
        "월시점 달성률",
        "최근3개월수량증감률",
        "수요증감률",
        "수요적용증감률",
        "평가월 수요진척률",
        "당월 출고진척률",
        "당월 재고충족률",
    }

    try:
        if engine == "xlsxwriter":
            workbook = writer.book
            worksheet = writer.sheets.get(sheet_name)
            if worksheet is None:
                return
            money_fmt = workbook.add_format({"num_format": "#,##0"})
            pct_fmt = workbook.add_format({"num_format": "0.00\\%"})
            header_fmt = workbook.add_format({"bold": True, "bg_color": "#E5E7EB", "border": 1})
            try:
                worksheet.freeze_panes(1, 0)
                if len(df.columns) > 0:
                    worksheet.autofilter(0, 0, max(len(df), 1), len(df.columns) - 1)
                for idx, col in enumerate(df.columns):
                    worksheet.write(0, idx, col, header_fmt)
            except Exception:
                pass
            for idx, col in enumerate(df.columns):
                name = str(col or "").strip()
                width = 24 if name in {"제약사명", "분석자료원"} else None
                if name in money_cols:
                    worksheet.set_column(idx, idx, width, money_fmt)
                elif name in decimal_money_cols:
                    dec_fmt = workbook.add_format({"num_format": "#,##0.00"})
                    worksheet.set_column(idx, idx, width, dec_fmt)
                elif name in percent_cols:
                    worksheet.set_column(idx, idx, width, pct_fmt)
                elif width:
                    worksheet.set_column(idx, idx, width)
            return

        if engine == "openpyxl":
            worksheet = writer.sheets.get(sheet_name)
            if worksheet is None:
                return
            try:
                from openpyxl.styles import Font, PatternFill, Border, Side
                from openpyxl.utils import get_column_letter
                worksheet.freeze_panes = "A2"
                if len(df.columns) > 0:
                    worksheet.auto_filter.ref = worksheet.dimensions
                header_fill = PatternFill("solid", fgColor="E5E7EB")
                thin = Side(style="thin", color="D1D5DB")
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                for idx, col in enumerate(df.columns, start=1):
                    if str(col or "").strip() in {"제약사명", "분석자료원"}:
                        worksheet.column_dimensions[get_column_letter(idx)].width = 24
            except Exception:
                pass
            for idx, col in enumerate(df.columns, start=1):
                name = str(col or "").strip()
                if name in money_cols:
                    fmt = "#,##0"
                elif name in decimal_money_cols:
                    fmt = "#,##0.00"
                elif name in percent_cols:
                    fmt = "0.00\\%"
                else:
                    continue
                for row in range(2, len(df) + 2):
                    worksheet.cell(row=row, column=idx).number_format = fmt
    except Exception:
        log.exception("[panel] apply excel number formats failed")

# ==========================================================
# (선택) 사이드·상태 제어 보조 함수 — 필요 시 호출
# ==========================================================
def set_run_flag(open_panel: bool = True) -> None:
    """외부(엔트리)에서 '실행' 버튼 클릭 시 호출"""
    ss = st.session_state

    # __sims_open은 사이드바 토글 위젯 전용 키다.
    # 실행 버튼이 이 값을 True로 되돌리면 사용자가 패널을 닫아도 다시 열리는 문제가 생긴다.
    # 실행 직후 보조 노출은 __sims_force_open만 사용한다.
    ss["__sims_force_open"] = bool(open_panel)
    ss["__sims_panel_active"] = True
    ss["__sims_run_flag"] = True
    ss["__sims_run_seq"] = int(ss.get("__sims_run_seq", 0)) + 1
    log.info("[ui] SIMS 실행 클릭됨 — user_open=%s force_open=%s form_id=%s run_seq=%s selected=%r",
             ss.get("__sims_open"), ss.get("__sims_force_open"), ss.get("__sims_form_id"), ss.get("__sims_run_seq"), ss.get("__sims_selected"))


def _force_open_sims_toggles() -> None:
    """
    메인에서 SIMS 패널 토글이 켜질 때 호출하는 보조 함수.
    현재는 별도 상태를 건드릴 필요가 없어서 no-op으로 둔다.
    향후 컨텍스트 컨트롤/허브 패널 expander와 연동할 때 이 함수를 확장하면 된다.
    """
    _ensure_sims_state()
    log.debug("[panel.main] _force_open_sims_toggles() called")
