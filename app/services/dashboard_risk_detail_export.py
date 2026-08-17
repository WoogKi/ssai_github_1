"""Excel export helpers for Dashboard Lite risk-detail facts.

This module deliberately receives only already-calculated facts.  It must not
load DB data, read Streamlit session state, or persist download bytes.
"""

from __future__ import annotations

from io import BytesIO
import logging
import time
from typing import Any, Mapping, Sequence

import pandas as pd

from app.ui.chat_middleware import _apply_sims_excel_number_formats, _sanitize_dataframe_for_excel


log = logging.getLogger("ssai.sims.dashboard_risk_detail_export")

DETAIL_SHEET_NAME = "위험품목상세"
VENDOR_SHEET_NAME = "매입처별위험요약"
SCOPE_SHEET_NAME = "조회조건"
INVENTORY_DETAIL_SHEET_NAME = "재고현황상세"

_AMOUNT_COLUMNS = {"현재재고금액", "재고평가단가", "위험보정부족예상금액", "긴급부족금액", "부족주의금액", "전체위험보정부족금액"}
_QTY_COLUMNS = {"현재재고수량", "당월현재출고수량", "당월기준예상출고수량", "진행속도기준월말예상출고수량", "위험보정예상출고수량", "위험보정잔여예상수요", "위험보정부족예상수량", "위험보정부족수량"}
_PCT_COLUMNS = {"위험보정재고준비율"}

_DETAIL_EXPORT_COLUMNS = (
    "위험상태", "위험사유", "제품코드", "제품명", "규격", "제조사명", "제품그룹명", "제품구분명", "제품분류명",
    "주요매입처코드", "주요매입처명", "주요매입처상태", "주요매입처선정기준", "재고기준",
    "현재재고수량", "현재재고금액", "재고평가단가", "당월현재출고수량", "당월기준예상출고수량",
    "진행속도기준월말예상출고수량", "위험보정예상출고수량", "위험보정잔여예상수요", "위험보정재고준비율",
    "위험보정부족예상수량", "위험보정부족예상금액", "수요급증여부", "수요급증상위분류", "수요급증세부분류",
    "수요급증사유",
)


def _as_frame(rows: Any) -> pd.DataFrame:
    if isinstance(rows, pd.DataFrame):
        return rows.copy()
    return pd.DataFrame(list(rows or []))


def _risk_detail_export_frame(rows: Any) -> pd.DataFrame:
    """Keep business columns only; never export private filter/sort helpers."""
    frame = _as_frame(rows)
    columns = [column for column in _DETAIL_EXPORT_COLUMNS if column in frame.columns]
    return frame.loc[:, columns].copy() if columns else pd.DataFrame()


def _apply_dashboard_risk_formats(writer: Any, df: pd.DataFrame, sheet_name: str) -> None:
    """Use common SIMS styling, then add Dashboard risk-detail formats."""
    _apply_sims_excel_number_formats(writer, df, sheet_name)
    worksheet = writer.sheets.get(sheet_name)
    if worksheet is None:
        return
    for column_index, column_name in enumerate(df.columns, start=1):
        name = str(column_name or "")
        if name in _AMOUNT_COLUMNS:
            number_format = "#,##0"
        elif name in _QTY_COLUMNS:
            number_format = "#,##0.00"
        elif name in _PCT_COLUMNS:
            number_format = "0.00\\%"
        else:
            continue
        for row_index in range(2, len(df) + 2):
            worksheet.cell(row=row_index, column=column_index).number_format = number_format

    try:
        from openpyxl.utils import get_column_letter

        for column_index, column_name in enumerate(df.columns, start=1):
            values = df.iloc[:500, column_index - 1].fillna("").astype(str)
            max_length = int(values.str.len().max()) if not values.empty else 0
            width = min(36, max(10, len(str(column_name or "")) + 2, max_length + 2))
            worksheet.column_dimensions[get_column_letter(column_index)].width = width
    except Exception:
        pass


def build_dashboard_risk_detail_excel_bytes(
    risk_detail_rows: Sequence[Mapping[str, Any]] | pd.DataFrame,
    vendor_rows: Sequence[Mapping[str, Any]] | pd.DataFrame,
    query_conditions: Sequence[Mapping[str, Any]] | pd.DataFrame,
) -> tuple[bytes, dict[str, int]]:
    """Build the three in-memory Dashboard risk-detail sheets with no side effects."""
    started = time.perf_counter()
    detail_df = _sanitize_dataframe_for_excel(_risk_detail_export_frame(risk_detail_rows))
    vendor_df = _sanitize_dataframe_for_excel(_as_frame(vendor_rows))
    scope_df = _sanitize_dataframe_for_excel(_as_frame(query_conditions))

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        detail_df.to_excel(writer, index=False, sheet_name=DETAIL_SHEET_NAME)
        vendor_df.to_excel(writer, index=False, sheet_name=VENDOR_SHEET_NAME)
        scope_df.to_excel(writer, index=False, sheet_name=SCOPE_SHEET_NAME)
        _apply_dashboard_risk_formats(writer, detail_df, DETAIL_SHEET_NAME)
        _apply_dashboard_risk_formats(writer, vendor_df, VENDOR_SHEET_NAME)
        _apply_dashboard_risk_formats(writer, scope_df, SCOPE_SHEET_NAME)

    payload = output.getvalue()
    info = {
        "export_rows": int(len(detail_df)),
        "vendor_summary_rows": int(len(vendor_df)),
        "sheet_count": 3,
        "bytes_size": int(len(payload)),
        "elapsed_ms": int((time.perf_counter() - started) * 1000),
    }
    log.info(
        "[dashboard.risk_detail_export] export_rows=%s vendor_summary_rows=%s sheet_count=%s bytes_size=%s permission_allowed=True elapsed_ms=%s success=True",
        info["export_rows"], info["vendor_summary_rows"], info["sheet_count"], info["bytes_size"], info["elapsed_ms"],
    )
    return payload, info


def build_dashboard_inventory_detail_excel_bytes(
    detail_rows: Sequence[Mapping[str, Any]] | pd.DataFrame,
    query_conditions: Sequence[Mapping[str, Any]] | pd.DataFrame,
) -> tuple[bytes, dict[str, int]]:
    """Export the already-filtered unified inventory detail without any reload."""
    started = time.perf_counter()
    detail_df = _sanitize_dataframe_for_excel(_as_frame(detail_rows))
    private_columns = [column for column in detail_df.columns if str(column).startswith("_")]
    if private_columns:
        detail_df = detail_df.drop(columns=private_columns)
    scope_df = _sanitize_dataframe_for_excel(_as_frame(query_conditions))

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        detail_df.to_excel(writer, index=False, sheet_name=INVENTORY_DETAIL_SHEET_NAME)
        scope_df.to_excel(writer, index=False, sheet_name=SCOPE_SHEET_NAME)
        _apply_dashboard_risk_formats(writer, detail_df, INVENTORY_DETAIL_SHEET_NAME)
        _apply_dashboard_risk_formats(writer, scope_df, SCOPE_SHEET_NAME)

    payload = output.getvalue()
    info = {
        "export_rows": int(len(detail_df)),
        "sheet_count": 2,
        "bytes_size": int(len(payload)),
        "elapsed_ms": int((time.perf_counter() - started) * 1000),
    }
    log.info(
        "[dashboard.inventory_detail_export] export_rows=%s sheet_count=%s bytes_size=%s permission_allowed=True elapsed_ms=%s success=True",
        info["export_rows"], info["sheet_count"], info["bytes_size"], info["elapsed_ms"],
    )
    return payload, info
